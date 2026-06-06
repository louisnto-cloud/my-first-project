#!/usr/bin/env python3
"""Enterprise competitor intelligence workbook for the Organika Sparkling / MUV launch.
Builds a ~25-tab analyst-grade .xlsx from six research streams."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------- palette ----------------
NAVY="14304F"; TEAL="2E7D8A"; STEEL="3E5C76"; LIGHT="EAF1F4"; LIGHT2="F4F8FA"
AMBER="FFF2CC"; AMBERHEAD="B45309"; GREEN="E2EFDA"; GREENHEAD="1E7D32"
RED="F8D7DA"; REDHEAD="9C2A2A"; WHITE="FFFFFF"; GREY="595959"; LINKBLUE="1155CC"

def F(**k): return Font(name="Calibri", **k)
HEAD   = F(bold=True, color=WHITE, size=10)
TITLE  = F(bold=True, color=WHITE, size=18)
SUB    = F(bold=True, color=NAVY, size=12)
BODY   = F(size=10, color="000000")
BODYB  = F(size=10, bold=True, color="000000")
SMALL  = F(size=9, italic=True, color=GREY)
LINK   = F(size=9, color=LINKBLUE, underline="single")

thin=Side(style="thin", color="BFBFBF")
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True, vertical="top")
WRAPC=Alignment(wrap_text=True, vertical="center")
CTR=Alignment(horizontal="center", vertical="center", wrap_text=True)

def fill(c): return PatternFill("solid", fgColor=c)

def hdrow(ws,row,n,color=NAVY):
    for c in range(1,n+1):
        x=ws.cell(row=row,column=c); x.font=HEAD; x.fill=fill(color); x.alignment=CTR; x.border=BORDER

def table(ws,start,headers,rows,widths=None,zebra=True,hcolor=NAVY,rowh=None):
    n=len(headers)
    for j,h in enumerate(headers,1): ws.cell(row=start,column=j,value=h)
    hdrow(ws,start,n,hcolor)
    r=start+1
    for i,row in enumerate(rows):
        for j,val in enumerate(row,1):
            x=ws.cell(row=r,column=j,value=val); x.font=BODY; x.alignment=WRAP; x.border=BORDER
            if zebra and i%2==1: x.fill=fill(LIGHT)
        if rowh: ws.row_dimensions[r].height=rowh
        r+=1
    if widths:
        for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
    return r

def banner(ws,row,text,n,color=TEAL,fontcolor=WHITE):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=n)
    c=ws.cell(row=row,column=1,value=text)
    c.font=F(bold=True,color=fontcolor,size=10); c.fill=fill(color)
    c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1)
    return row+1

def callout(ws,row,text,n,fillc=GREEN,fontc=GREENHEAD,h=44):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=n)
    c=ws.cell(row=row,column=1,value=text); c.font=F(bold=True,color=fontc,size=10)
    c.fill=fill(fillc); c.alignment=WRAP; c.border=BORDER
    ws.row_dimensions[row].height=h
    return row+1

def title(ws,text,n,sub=None):
    ws.sheet_view.showGridLines=False
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    c=ws.cell(row=1,column=1,value=text); c.font=TITLE; c.fill=fill(NAVY); c.alignment=CTR
    ws.row_dimensions[1].height=30
    if sub:
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=n)
        s=ws.cell(row=2,column=1,value=sub); s.font=F(bold=True,color=WHITE,size=10)
        s.fill=fill(TEAL); s.alignment=CTR; ws.row_dimensions[2].height=20
        return 4
    return 3

wb=Workbook()

# ============================================================ 00 COVER
ws=wb.active; ws.title="00 Cover"; ws.sheet_view.showGridLines=False
for col,w in zip("ABCDEFGHI",[3,20,20,20,20,20,20,20,3]): ws.column_dimensions[col].width=w
ws.merge_cells("B2:H2"); t=ws["B2"]
t.value="FUNCTIONAL & WELLNESS SPARKLING BEVERAGES"; t.font=TITLE; t.fill=fill(NAVY); t.alignment=CTR
ws.row_dimensions[2].height=40
ws.merge_cells("B3:H3"); s=ws["B3"]
s.value="Enterprise Competitive Intelligence & Market-Entry Dossier"; s.font=F(bold=True,color=WHITE,size=13); s.fill=fill(TEAL); s.alignment=CTR
ws.row_dimensions[3].height=26
ws.merge_cells("B4:H4"); s2=ws["B4"]
s2.value="Prepared for the Organika Sparkling / “MUV” launch decision  ·  United States + Canada"; s2.font=F(italic=True,color=NAVY,size=11); s2.alignment=CTR
ws.row_dimensions[4].height=20

meta=[("Document type","Competitive intelligence + go-to-market dossier (analyst-grade)"),
 ("Subject","Functional soda, hydration/electrolyte, and wellness/collagen sparkling categories"),
 ("Decision supported","Whether/how to launch a Canadian functional collagen sparkling beverage (“MUV”)"),
 ("Geography","US (primary evidence base) with Canada market + regulatory read-across"),
 ("Brands analyzed","8 core profiles + ~30 extended landscape entries across 6 sub-categories"),
 ("Research method","6 parallel research streams, multi-source web research, per-claim citations, confidence-tagged"),
 ("Date prepared","2026-06-05"),
 ("Status","Draft for internal review — verify amber-flagged items before external circulation")]
r=6
for k,v in meta:
    ws.cell(row=r,column=2,value=k).font=SUB
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=8)
    c=ws.cell(row=r,column=3,value=v); c.font=BODY; c.alignment=WRAP
    ws.row_dimensions[r].height=28; r+=1

r+=1
r=callout(ws,r,("CRITICAL CAVEAT — A discrete “Organika Sparkling” ready-to-drink CAN could not be verified in public sources. "
 "Organika's documented fizzy/collagen products are POWDER (Electrolytes + Enhanced Collagen) and LIQUID collagen water. "
 "If “MUV” is a new RTD can, confirm format, flavours, dose and price with the brand. The entire Canada regulatory analysis "
 "assumes an RTD food-format beverage."),7,fillc=AMBER,fontc=AMBERHEAD,h=72)
r+=1
ws.cell(row=r,column=2,value="Confidence legend used throughout:").font=BODYB; r+=1
for txt,col,fc in [("HARD DATA — published / company-reported / scanner (Circana/SPINS) / regulatory primary",GREEN,GREENHEAD),
                   ("DIRECTIONAL — research-firm projection or single-source; definitions vary",LIGHT,GREY),
                   ("FLAG / VERIFY — estimate, secondary-only, or decision-critical assumption to confirm",AMBER,AMBERHEAD)]:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
    c=ws.cell(row=r,column=2,value=txt); c.font=F(bold=True,size=9,color=fc); c.fill=fill(col); c.alignment=WRAP; c.border=BORDER
    ws.row_dimensions[r].height=18; r+=1

# ============================================================ 01 TOC
ws=wb.create_sheet("01 Contents")
nr=title(ws,"Table of Contents",3)
toc=[
 ("00 Cover","Title, scope, caveats, confidence legend"),
 ("01 Contents","This page"),
 ("02 Executive Summary","The thesis, the 7 findings that matter, the recommendation"),
 ("03 Methodology","Research design, source hierarchy, confidence framework, limitations"),
 ("04 Market Sizing","TAM/SAM/SOM, category sizes & growth, US + Canada"),
 ("05 Trends & Macro","Sugar reduction, GLP-1, gut health, beauty-from-within, white space"),
 ("06 Consumer","Segmentation, collagen buyer, efficacy-evidence risk, buyer criteria"),
 ("07 Comparison Matrix","8 core brands, key metrics side by side"),
 ("08 Extended Landscape","~30 brands across 6 sub-categories"),
 ("09 Channel & Retail","Launch channel & expansion sequence per brand"),
 ("10 Channel Economics","Margin waterfall, retailer/distributor/trade-spend, DTC/Amazon"),
 ("11 Formulation","The winning claim, hero ingredient, claim risk"),
 ("12 Marketing","Founder story, virality engine, what moved the needle"),
 ("13 Launch Playbooks","Phased GTM timelines (Poppi/OLIPOP/LMNT/Liquid I.V./Celsius)"),
 ("14 Patterns & Failures","7 success patterns + 6 failure modes"),
 ("15 Outcomes & M&A","Revenue, funding, valuations, acquisition comps"),
 ("16 SWOT","Per-brand strengths/weaknesses/opportunities/threats"),
 ("17 Scoring Model","Weighted competitive scoring (editable weights)"),
 ("18 Canada Market","Retail banners & share, entry playbooks, pricing"),
 ("19 Canada Regulatory","NHP vs Food vs Supplemented Foods; SFCI; collagen/caffeine; Bill 96"),
 ("20 GTM Recommendation","The MUV go-to-market thesis & decisions"),
 ("21 Launch Plan","Phased 0–18 month plan with workstreams"),
 ("22 Risk Register","Ranked risks, likelihood/impact, mitigations"),
 ("23 KPI Dashboard","Targets & benchmarks by phase"),
 ("24 Sources","Master bibliography, tagged by type"),
 ("25 Glossary","Terms & abbreviations"),
]
nr=table(ws,nr,["Tab","Contents"],toc,widths=[26,84],rowh=20)

# ============================================================ 02 EXEC SUMMARY
ws=wb.create_sheet("02 Executive Summary")
nr=title(ws,"Executive Summary",2,"How competitors launched, what made them win, and what it means for MUV")
ws.column_dimensions["A"].width=26; ws.column_dimensions["B"].width=92
nr=banner(ws,nr,"THE THESIS",2,color=STEEL)
ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=2)
c=ws.cell(row=nr,column=1,value=("The functional-beverage winners of 2018–2025 did NOT win on formula. They won on a repeatable launch mechanic: a "
 "founder-authority insight, ONE obsessively-executed trial tactic, and a data-driven (velocity/repeat-rate) retail sell-in — sequenced up a "
 "channel ladder from natural/DTC to mass. For MUV, the opportunity is real (gut-health + beauty-from-within + sugar-cutting tailwinds and a "
 "domestic-producer cost edge), but it sits squarely at the intersection of the category's two biggest risks: collagen-efficacy/claims liability "
 "and strategic-owned competition (Nestlé already owns Vital Proteins). Win condition: efficacious dose + claim discipline + proven repeat "
 "before chasing doors + a Costco-Canada/Amazon.ca beachhead Organika is uniquely positioned to seize."))
c.font=BODY; c.alignment=WRAP; c.fill=fill(LIGHT2); c.border=BORDER; ws.row_dimensions[nr].height=118; nr+=2

findings=[
 ("1. Channel ladder, never skipped","Winners sequenced natural/independent or DTC → specialty (Whole Foods/Sprouts) → conventional grocery → mass (Target/Walmart) → club/convenience — earning each rung with proven velocity. OLIPOP launched in 40 NorCal indie grocers; Poppi seeded farmers markets → Whole Foods."),
 ("2. One trial tactic, executed obsessively","Poppi = TikTok creator seeding; LMNT = free-sample-pay-shipping + podcast host-reads; Liquid I.V. = weekend/Costco roadshow sampling; OLIPOP = natural-channel velocity; Celsius = Amazon + gym influencers. Pick one, go deep."),
 ("3. Sell-in on DATA, not flavour","Retail buyers buy velocity and repeat rate, not taste. OLIPOP's “astronomical repurchase rate” and Liquid I.V.'s POS data were the sell-in. This is the single most repeated lever."),
 ("4. Category-creation language commands price","“Prebiotic soda”, “hydration multiplier” — escaping commodity comparison supports a ~$2–$3/can premium. Modern soda hit $1.8B in 2024 (+83% YoY, Circana)."),
 ("5. The collagen graveyard is real","Standalone “beauty collagen sparkling water” indie brands repeatedly flopped (Goldn Hour, BOONS, Luster & Lum). A 2025 AJM meta-analysis found NO proven skin-aging benefit and an industry-funding bias — a live claims liability. Collagen-as-feature inside a broader brand survives; collagen-as-hero does not."),
 ("6. Claims discipline is existential","Poppi paid $8.9M settling a gut-health claim its 2g fiber couldn't support. Dose to efficacy and use structure/function-safe language, or budget for litigation."),
 ("7. Canada is a FOOD, not an NHP, problem","An RTD collagen can is regulated as a food/Supplemented Food (CFIA), not an NHP — so Organika's NPN-backed beauty claims do NOT transfer to the can, and a mandatory collagen caution statement applies. This reshapes the entire claim strategy."),
]
nr=banner(ws,nr,"THE SEVEN FINDINGS THAT MATTER",2,color=STEEL)
nr=table(ws,nr,["Finding","So what"],findings,widths=[26,92],rowh=78)
nr+=1
nr=callout(ws,nr,("RECOMMENDATION (detail on Tab 20): GO — conditional. Lead with a functional/benefit-forward proposition (gut + hydration + a credible "
 "collagen dose) under a “by Canada's #1 collagen brand” halo, NOT narrow beauty-from-within. Beachhead = Costco Canada + Amazon.ca + natural "
 "channel + Organika's existing pharmacy/health-food base. Price at/below the ~$3.49 CAD premium-functional shelf norm, protected by domestic "
 "production. Resolve the Supplemented-Foods claim/label pathway with regulatory counsel BEFORE artwork. Prove repeat rate before chasing doors."),2,h=92)

# ============================================================ 03 METHODOLOGY
ws=wb.create_sheet("03 Methodology")
nr=title(ws,"Methodology & Confidence Framework",2)
ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=88
rows=[
 ("Research design","Six parallel research streams: (1) market sizing, (2) channel economics, (3) extended competitor sweep, (4) Canada market + regulatory, (5) launch-playbook teardowns, (6) consumer trends + risk. Each ran independent multi-source web research with per-claim citation."),
 ("Source hierarchy","Decision-grade → scanner data (Circana/SPINS), company/SEC filings (LMNT Reg CF), regulatory primary (Health Canada, CFIA, FDR), transaction/press releases. Directional → syndicated research-firm market sizes (Grand View, FMI, Mordor) which vary widely by category definition. Lowest → single secondary sources / social estimates."),
 ("Confidence tagging","Every load-bearing figure is tagged HARD / DIRECTIONAL / FLAG. Where sources conflict (collagen market size; prebiotic soda; adaptogen market) both figures are shown and the conflict flagged."),
 ("Known limitations","(a) Many primary publisher pages (CNBC, Entrepreneur, Circana, canada.ca) block automated fetch (HTTP 403); figures were captured via search extracts of those same pages plus corroborating sources — re-verify load-bearing numbers from the cited URLs before external use. (b) Private-company margins and Canadian per-SKU listing fees are confidential; ranges are industry rules-of-thumb. (c) Female-skew % and licensed Circana/Mintel splits sit behind paywalls."),
 ("What would strengthen this","Licensed Circana/SPINS pull for Canada; a regulatory-counsel opinion on the Supplemented-Foods claim set; primary interviews with 2–3 Canadian distributors (UNFI Canada, Tree of Life, Purity Life); a co-packer capacity/quote for a Canadian RTD line."),
]
nr=table(ws,nr,["Element","Detail"],rows,widths=[30,88],rowh=92)

# ============================================================ 04 MARKET SIZING
ws=wb.create_sheet("04 Market Sizing")
nr=title(ws,"Market Sizing — Category Sizes & Growth",7,"Anchor TAM on Circana scanner data; treat research-firm category sizes as directional floors")
us=[
 ["Modern / better-for-you soda (scanner)","US","$1.8B","2024","+83% YoY (from $983M 2023)","Circana / Beverage Industry","HARD"],
 ["Prebiotic/probiotic soda (research-firm)","US","$171.6M","2024","7.8% (25–30) → $268.8M","Grand View","DIRECTIONAL – conflicts w/ scanner"],
 ["Prebiotic soda (sales ramp)","US","~$777M","2025","from ~$33M in 2022","just-drinks / SPINS","DIRECTIONAL"],
 ["Electrolyte powder","US","$2.8B","2024","8.9% (25–30) → $4.6B","Grand View / Spherical","HARD-ish (2 sources)"],
 ["Powdered hydration mixes (scanner)","US","$1.5B","2024","+20% YoY","Circana / CNN","HARD"],
 ["Collagen beverage","US","$47.9M","2024","10.9% to 2034","Fact.MR","FLAG – see conflict"],
 ["Collagen beverage (global)","Global","$275–470M","2024","9–10%","Fact.MR / market.us","FLAG – definitions vary"],
 ["Beauty-from-within drinks","Global","$3.24B","2024","8.5% to 2035 → $4.77B","Future Market Insights","DIRECTIONAL"],
 ["Adaptogenic beverages","Global","$1.39B","2024","7.1% to 2034 → $2.73B","GMInsights (FMI conflicts)","FLAG – dueling"],
 ["Sparkling water (total)","US","$6.9B","2024","—","Market Research Future","DIRECTIONAL"],
]
nr=table(ws,nr,["Category","Geo","Size","Year","Growth","Source","Confidence"],us,
   widths=[34,8,12,8,28,24,24],rowh=30)
nr+=1
nr=banner(ws,nr,"CANADA (category-specific scanner data scarce — US-as-proxy convention: Canada ≈ 10% of US)",7,color=TEAL)
ca=[
 ["Functional beverages (broad)","Canada","US$17.35B","2025","4.5% → $27.0B (2035)","Market Research Future","FLAG – very broad def, 1 source"],
 ["Modern soda (US-proxy)","Canada","~$180M","2024","—","Derived (10% of US $1.8B)","ESTIMATE"],
 ["Electrolyte powder (US-proxy)","Canada","~$280M","2024","—","Derived (10% of US $2.8B)","ESTIMATE"],
]
nr=table(ws,nr,["Category","Geo","Size","Year","Growth","Source","Confidence"],ca,
   widths=[34,8,12,8,28,24,24],rowh=26)
nr+=1
# TAM/SAM/SOM
nr=banner(ws,nr,"ILLUSTRATIVE TAM / SAM / SOM FOR MUV (Canada) — framework, plug real numbers",7,color=STEEL)
tss=[
 ["TAM","Canadian functional beverage category","~US$17B (broad) / ~$180M modern-soda-equiv","Total demand pool","DIRECTIONAL"],
 ["SAM","Functional/collagen/prebiotic sparkling sub-segment served by MUV in target channels","Est. low-hundreds of $M CAD (build bottom-up from banner ACV)","Reachable with planned distribution","ESTIMATE – build bottom-up"],
 ["SOM (Yr 1–3)","Realistic share given doors, velocity, trade spend","Model from doors × velocity × price; benchmark new-brand share <1–3%","Capturable near-term","BUILD"],
]
nr=table(ws,nr,["Layer","Definition for MUV","Sizing approach","Meaning","Confidence"],tss,
   widths=[10,40,34,18,18],rowh=44)
nr=callout(ws,nr,("FLAG: Three figures carry material source conflicts — (a) collagen beverage ($47.9M US vs $275–470M global), (b) prebiotic soda "
 "($171.6M research-firm vs $1.8B Circana “modern soda”, different boundaries), (c) adaptogenic beverages (GMInsights vs FMI). Anchor on Circana "
 "scanner reads; treat the rest as directional floors."),7,fillc=AMBER,fontc=AMBERHEAD,h=46)

# ============================================================ 05 TRENDS
ws=wb.create_sheet("05 Trends & Macro")
nr=title(ws,"Trends & Macro Drivers (2025–2026)",4)
rows=[
 ["Sugar reduction","Table stakes","60%+ of US consumers actively cutting sugar; >50% of new bev launches carry sugar-free claim; low-sugar est. +$36B/3yrs (Mintel)","Zero/low sugar is baseline, not a differentiator"],
 ["GLP-1 (Ozempic/Wegovy)","Reshaping demand","~23% of US households have a user; projected 35% of F&B units by 2030; GLP-1 users over-index zero-sugar (+20.5pp); grocery spend -5.3% in 6 mo, partial rebound after ~12 mo","Favor low-cal, protein, fiber, hydration, smaller portions"],
 ["Gut health","Mainstreamed","Prebiotic soda ~$33M (2022) → ~$777M (2025); digestive-claim soda +301.5% $ growth 2023 (SPINS MULO)","The core MUV-adjacent tailwind — ride prebiotic/fiber"],
 ["Beauty-from-within","Accelerating, contested","BFW drinks $3.24B (2024), women 68.4% of revenue, core buyer F 25–45; BUT 2025 AJM meta-analysis: no proven skin-aging benefit + funding bias","Demand is real; efficacy claims are legally risky"],
 ["Protein & fiber","Protein = table stakes; fiber = frontier","Protein market $12.1B→$27.4B by 2034; “fibermaxxing” rising on prebiotic wave","Fiber is a credible, defensible functional hook"],
 ["Adaptogens/nootropics","Rising w/ Gen Z","Ashwagandha, L-theanine, lion's mane, magnesium threonate going mainstream","Optional stack-on; watch claims"],
 ["Dry / non-alc","Structural","Declining alcohol + better-for-you adult drinks framed as durable","Tailwind for premium functional cans"],
 ["Function stacking","Consumer expectation","One drink, multiple need-states (gut+energy, hydration+skin)","Multi-benefit hedges single-fad risk"],
]
nr=table(ws,nr,["Trend","Status","Evidence","Implication for MUV"],rows,widths=[20,20,52,30],rowh=70)
nr=banner(ws,nr,"WHITE SPACE (where analysts say growth goes): prebiotic/fiber soda · collagen in new formats (sparkling) · protein-forward · GLP-1-friendly hydration · low/no sugar default",4,color=GREENHEAD)

# ============================================================ 06 CONSUMER
ws=wb.create_sheet("06 Consumer")
nr=title(ws,"Consumer Segmentation & The Collagen Buyer",3)
rows=[
 ["Demand engine","Gen Z + Millennials","~80% of Gen Z, ~75% of Millennials consume functional beverages regularly; 36% of adults but 41%+ of wellness spend (McKinsey). US wellness market $480B (2024)."],
 ["Gender skew (category)","Female-led growth","Women = fastest-growing segment (~5.3% growth); over-index on probiotic/fortified. Hard female-% split is paywalled — DATA GAP."],
 ["Collagen / BFW buyer","Women 25–45","Women = 68.4% of beauty-from-within drink revenue (2024); buy anti-aging, hydration, hair/nail outcomes; clinically-backed claims drive repeat/subscription."],
 ["Efficacy & skepticism","RISK","2025 AJM systematic review/meta-analysis: NO clinical evidence collagen prevents/treats skin aging; industry-funded trials showed benefit, independent ones did not. Positive studies used 2.5–10g/day, effect at 4–8 wks."],
 ["Willingness to pay","Attribute-gated","Younger consumers are price-conscious AND premium-willing for a named function / clean label / brand-community. Premiumization is real but not blanket."],
]
nr=table(ws,nr,["Dimension","Read","Detail & evidence"],rows,widths=[22,22,74],rowh=60)
nr+=1
nr=banner(ws,nr,"RETAILER / CLUB BUYER CRITERIA (what gets a new functional bev listed)",3,color=STEEL)
buy=[["Velocity","Gating metric — will the SKU turn fast enough to earn its facing"],
 ["Repeat rate","Buyers discount discount-driven velocity; they want durable repeat & customer-level margin"],
 ["Margin / trade support","Plan 35–45% trade spend in early scale; slotting, promo, free-fill"],
 ["Differentiation + demand-gen","Category-builder story + brand-funded marketing (the Poppi/OLIPOP social-first benchmark)"]]
nr=table(ws,nr,["Criterion","What buyers actually want"],buy,widths=[22,96],rowh=30)
nr=callout(ws,nr,"IMPLICATION: MUV should dose collagen to a defensible level (2.5g+), lead with gut/hydration function, and prove repeat rate in a beachhead before pitching national buyers. The collagen efficacy debate makes substantiation files and structure/function-safe language mandatory.",3,h=52)

# ============================================================ 07 COMPARISON MATRIX
ws=wb.create_sheet("07 Comparison Matrix")
nr=title(ws,"Core Competitor Comparison Matrix",12)
H=["Brand","Category","Founded","Founder(s)","First channel","Core claim","Format","Key spec","Unit price","Doors / reach","Peak revenue","Outcome"]
rows=[
 ["Poppi","Modern soda","2016","A. & S. Ellsworth","Farmers market → Whole Foods","Prebiotic soda + ACV, low sugar","12oz can","~2g fiber, 2–5g sugar, ~25 cal","~$2.25/can","~36,000+ US; 120+ retailers","~$500M (2024 est.)","PepsiCo $1.95B, May 2025"],
 ["OLIPOP","Modern soda","2018","B. Goodwin, D. Lester","40 NorCal indie grocers","9g fiber, gut health","12oz can","up to 9g fiber, 2–5g sugar","~$3.00/can","~35,000+ US (7,000+ chains)","$400M+ (2024)","Independent; $1.85B val (Feb 2025)"],
 ["LMNT","Electrolyte","2018","R. Wolf + Ketogains","DTC + podcast","Zero-sugar, high-sodium, keto","Powder sticks","1,000mg sodium, 0g sugar","~$1.00–1.50/stick","DTC-led + Amazon","$206M sales (2023, SEC)","Independent (Reg CF)"],
 ["Liquid I.V.","Electrolyte","2012","B. Cohen (+2)","Whole Foods sampling → Costco","Hydration Multiplier (CTT) 2–3x","Powder sticks","500mg sodium, 11g carbs","~$0.67/stick (Costco)","80,000+ US stores","~$200M at exit; $1B+ now","Unilever, Sept 2020 (~$500M est.)"],
 ["Nuun","Electrolyte","2004","T. Moxey","Cycling/endurance specialty","Effervescent low-sugar tablet","Tablets (+powder)","300mg sodium, 1g sugar","~$0.60–0.70/serv","Broad; #1 run/bike/outdoor","$10M+ (2013)","Nestlé HS, July 2021"],
 ["Vital Proteins Sparkling","Collagen sparkling","2025 line","(Nestlé-owned)","Mass retail","Beauty-from-within: VERISOL + Vit C","Can","0g sugar, 15 cal, collagen","$2.50/can","Mass (Nestlé)","n/a (new)","Nestlé-owned"],
 ["Recess","Functional sparkling","2018","B. Witte","DTC → retail","Relaxation: magnesium + adaptogens","Can","Mag L-threonate, adaptogens","Premium","15,000+ stores","n/a","Indep.; $30M Series B (2025)"],
 ["Aura Bora","Sparkling water","2020","P. & M. Voge","Shark Tank → DTC/retail","Herbal/floral, 0 cal/sugar","Can","0 cal/sugar/sodium","~$1.99–2.29/can","11,000 stores","$12M (2024)","Majority stake to Next In Natural"],
]
nr=table(ws,nr,H,rows,widths=[16,14,9,16,22,22,12,18,14,22,16,24],rowh=56)
ws.freeze_panes="A4"
nr=callout(ws,nr,"COLLAGEN-SPARKLING GRAVEYARD: Goldn Hour (CA, closed 2025), BOONS (acquired then discontinued), Luster & Lum (defunct). Standalone “beauty collagen sparkling water” indie brands repeatedly flopped — narrow positioning, premium price, no clinical moat, undercapitalized.",12,fillc=AMBER,fontc=AMBERHEAD,h=44)

# ============================================================ 08 EXTENDED LANDSCAPE
ws=wb.create_sheet("08 Extended Landscape")
nr=title(ws,"Extended Competitor Landscape (~30 brands)",6)
def seg(ws,nr,name):
    return banner(ws,nr,name,6,color=STEEL)
H6=["Brand","Founded / Founder","Positioning / hero","Price / footprint","Funding / exit","Status"]
nr=seg(ws,nr,"FUNCTIONAL SODA / PREBIOTIC")
nr=table(ws,nr,H6,[
 ["Culture Pop","2020; Tom First (Nantucket Nectars)","Probiotic soda, live cultures, real juice","Multi-region retail","$21M raised (~2024)","Active"],
 ["Health-Ade (Booch Pop)","Kombucha 2012; Pop 2020","Prebiotic soda; flagship kombucha","~$250M/yr retail","Acq. Generous Brands $500M","Active"],
 ["Mighty Pop","2023; by Beliv","Pre+pro+postbiotics, 3g fiber/can","US","Beliv portfolio","Active; sold out initial run"],
 ["De La Calle","2020; Matthews & Martin del Campo","Tepache (fermented pineapple), low sugar","Retail","$7M (2022, KarpReilly)","Active"],
 ["United Sodas","2019; Brooklyn","Premium low-cal, DTC flavor packs","~2,500 stores","Undisclosed (FLAG)","Active"],
 ["Wildwonder","~2018; Rosa Li","Sparkling pre+probiotic gut drink","Sprouts, Target chilled","Shark Tank $500K/6% (2023)","Active"],
 ["Simply Pop","2025; Coca-Cola","Prebiotic soda (incumbent entry)","Mass","Corporate","Active"],
],widths=[16,26,28,18,22,16],rowh=30)
nr=seg(ws,nr,"SPARKLING / FLAVORED WATER")
nr=table(ws,nr,H6,[
 ["Spindrift","2010; Bill Creelman","Real squeezed fruit, only ingredient","National","$75.7M; recap by Gryphon","Active"],
 ["Waterloo","2017; Cusack & Christopher (CAVU)","Bold zero-cal sparkling water","National","Majority sale 2020 (Flexis-led, NOT KDP)","Active"],
 ["Sound","~2015; Kelly & Najjar","Organic sparkling tea/botanicals, unsweetened","Retail","Acq. Next In Natural","Active"],
 ["Hint","2005; Kara Goldin","Unsweetened fruit-essence water","National","$51.5M; ~$150M val","Active, indep."],
 ["AHA","2020; Coca-Cola","Zero-cal flavor-pair sparkling","Reduced","Corporate","Distribution cut"],
 ["Bubly","2018; PepsiCo","Playful zero-cal sparkling","National","Corporate","Active"],
],widths=[16,26,28,18,22,16],rowh=28)
nr=seg(ws,nr,"ELECTROLYTE / HYDRATION (expanded)")
nr=table(ws,nr,H6,[
 ["Gatorlyte","2022; PepsiCo/Gatorade","Rapid-rehydration 5-electrolyte","Full distribution","Corporate","Active"],
 ["Prime Hydration","2022; Logan Paul & KSI","Influencer-led; BCAAs, coconut water","Mass","~$1.2–1.3B peak 2023, then down ~75% (FLAG)","Contracting"],
 ["Cure Hydration","~2018; L. Picasso","Plant-based, no added sugar","Retail","$9.88M total; $5.6M A (2023)","Active"],
 ["Skratch Labs","2012; Dr. Allen Lim","“Real food” sports hydration","Specialty","Bootstrapped","Active"],
 ["DripDrop","2008; Dr. Dolhun","Medical-grade ORS","Retail","$3M+$5.6M early","Active (PBC)"],
 ["Ultima Replenisher","25+ yrs","Zero-sugar electrolyte powder","Retail","Undisclosed (FLAG)","Active"],
],widths=[16,26,28,18,22,16],rowh=28)
nr=seg(ws,nr,"ENERGY / FUNCTIONAL (playbook reference)")
nr=table(ws,nr,H6,[
 ["Celsius","2004; NASDAQ 2017","Thermogenic energy; fitness/social + PepsiCo","100,000+ doors","PepsiCo $550M/8.5% (2022); bought Alani Nu $1.8B","Public"],
 ["Alani Nu","2018; Katy Hearn","Influencer/DTC → mass","Mass","~$595M rev 2024; sold to Celsius $1.8B","Acquired"],
 ["Ghost","2016","Community/influencer, licensed flavors","Mass","KDP 60% (2024)","Acquired (KDP)"],
],widths=[16,26,28,18,22,16],rowh=28)
nr=seg(ws,nr,"COLLAGEN / BEAUTY / ADAPTOGEN")
nr=table(ws,nr,H6,[
 ["Vital Proteins","K. Seidensticker","Leading US collagen (supp/bev/food)","National","Nestlé (full 2022)","Active"],
 ["Moon Juice","2011; A.C. Bacon","Adaptogenic dusts/supplements","DTC + retail","Undisclosed (FLAG)","Active"],
 ["Wylde One","S. Park","Adaptogenic elixirs/lattes (powder)","Anthropologie, UO","Undisclosed (FLAG)","Active"],
 ["Olly","—","Gummy vitamins incl. beauty/collagen","Mass","Unilever (2019)","Active"],
 ["Halo","Ambiguous","Ingestible beauty supplement (NOT a beverage)","—","—","Clarify"],
 ["Mud\\Wtr","2018; S. Heath","Coffee-alt w/ mushrooms","DTC-led","~$1M seed; later FLAG","Active"],
],widths=[16,26,28,18,22,16],rowh=28)
nr=seg(ws,nr,"CANADIAN FUNCTIONAL BEVERAGE BRANDS")
nr=table(ws,nr,H6,[
 ["Collective Arts","2013; Hamilton ON","Craft + non-alc/wellness/cocktails","National","Independent","Active"],
 ["Flow","2015; N. Reichenbach","Alkaline spring water, B-Corp","~46k stores NA","TSX-listed; ~C$30M","Active (pressured)"],
 ["Wize Tea","Montreal","Sparkling tea upcycled from coffee leaves","Regional","Undisclosed (FLAG)","Active"],
 ["Partake","2017; T. Fleming","Non-alc beer; Hop Twist hop water (2025)","National","Kickstarter origin","Active"],
 ["Greenhouse Juice","Toronto","Cold-pressed juice + functional","Regional","Independent","Active"],
 ["Rise Kombucha","2008; Montreal","Leading CA kombucha; “Better Soda” prebiotic (2023)","National","Buddha Brands","Active"],
 ["Loop / Crazy D's / Cove Soda","Various (CA)","Real CA prebiotic/better-soda players (verify)","Regional","—","Active (suggested swap-ins)"],
],widths=[16,26,28,18,22,16],rowh=28)
nr=callout(ws,nr,"VERIFICATION NOTES: “GoodSip”, “Rad”, “Local”, “Botaniste” could not be verified as real Canadian functional-beverage brands and were dropped — replace with Loop Mission / Crazy D's / Cove Soda / Buddha Brands. Waterloo was acquired by a Flexis-led group (2020), NOT Keurig Dr Pepper. “Halo” is a supplement, not an RTD. Re-verify all FLAG funding figures before circulation.",6,fillc=AMBER,fontc=AMBERHEAD,h=60)

# ============================================================ 09 CHANNEL & RETAIL
ws=wb.create_sheet("09 Channel & Retail")
nr=title(ws,"Channel & Retail Strategy",5,"Where they launched, why, and how they sequenced expansion")
H=["Brand","Launch channel (why)","Expansion sequence","Reach today","Canada presence"]
rows=[
 ["Poppi","Dallas farmers market; a Whole Foods buyer found the booth in ~3 weeks — natural channel validated product first","Farmers mkt → Whole Foods national → DTC/Amazon (#1 soda 2023) → grocery → mass/club","~36,000+ US, 120+ retailers","Launched Canada Aug 13, 2024 (first intl): Loblaws, Metro, Save-On, Costco, Sobeys, Well.ca, Natura"],
 ["OLIPOP","40 independent NorCal natural grocers (Oct 2019) — deliberately seeded natural channel","Indie CA → Sprouts/WF → Kroger/Safeway → Target/Walmart/Costco + Starbucks; 4,500 stores by Nov 2020","~35,000+ US","Thin official; via NevrAsk distributor; Amazon.ca, Natura"],
 ["LMNT","DTC + podcast sponsorship — owned the customer from day one","DTC/sub → Amazon → select retail (Vitamin Shoppe, Target) + Sparkling can","DTC-led + Amazon","Ships to Canada from site (free >~$100)"],
 ["Liquid I.V.","Whole Foods “Supplier 101” + founder sampling every weekend → 12 WF by 2015; then Costco","Whole Foods → Costco national (2019, ~516 whs) → broad retail; 30k→80k stores","80,000+ US stores","Costco Canada FIRST (2023 via Unilever); then Amazon.ca, Walmart, Loblaws, Jean Coutu"],
 ["Nuun","Cycling/endurance/tri specialty (2004) — grassroots in athlete community","Run/bike specialty → REI/Whole Foods → broad grocery/mass","Broad; #1 run/bike/outdoor","Dedicated Canadian retailers page"],
 ["Celsius","Amazon dominance + gym/fitness influencers","Amazon → Target DSD (2020) → 100k doors → PepsiCo distribution (2022)","100,000+ doors","Jan 2024 via PepsiCo distribution"],
 ["Vital Proteins Sparkling","Mass retail from the start (Nestlé muscle) — skipped indie-seeding","Big-brand mass launch","Mass","Not yet verified at CA retail"],
 ["Recess","DTC-first, built brand world online, then retail","DTC → Target/Kroger/Sprouts/WF","15,000+ stores","Not verified"],
]
nr=table(ws,nr,H,rows,widths=[14,34,40,20,38],rowh=88)
ws.freeze_panes="A5"
nr=callout(ws,nr,"READ-ACROSS: Costco Canada + Amazon.ca + natural channel is the proven Canadian beachhead (how Liquid I.V. and Poppi entered). Organika already holds Canadian pharmacy/health-food/Costco/Well.ca relationships imported brands must buy into — a genuine structural head-start.",5,h=46)

# ============================================================ 10 CHANNEL ECONOMICS
ws=wb.create_sheet("10 Channel Economics")
nr=title(ws,"Channel Economics & Margin Waterfall",4,"Only Amazon fees, Costco margin & DTC CAC are HARD; the rest are operator rules-of-thumb — model as ranges")
nr=banner(ws,nr,"MARGIN WATERFALL — canned functional beverage",4,color=STEEL)
mw=[["Unit COGS target","< $1.00/unit (liquid+can+co-pack at scale)","RULE OF THUMB"],
 ["12oz decorated can (component)","~$0.15/can at intro minimums","RULE OF THUMB"],
 ["Co-pack run (10,000 cases)","≈ $200,000; MOQ 2,000–5,000 cases/SKU","RULE OF THUMB"],
 ["Blended gross margin (early)","40–50%; aspiration 50–65% at scale","RULE OF THUMB"],
 ["GM by channel","Retail ~35–45%; DTC ~40–50% (before CAC/freight)","RULE OF THUMB"],
 ["COGS warning","Excluding freight/duty/co-pack overstates GM by 3–8 pts — the #1 error","RULE OF THUMB"]]
nr=table(ws,nr,["Line","Typical value","Tag"],mw,widths=[28,62,22],rowh=26)
nr+=1
nr=banner(ws,nr,"CHANNEL TAKE-RATES (who takes margin off the top)",4,color=STEEL)
ch=[["Grocery (conventional)","~30–50% retailer margin; bev often 30–40%","RULE OF THUMB"],
 ["Natural/premium (Whole Foods)","45%+ (slower turns)","RULE OF THUMB"],
 ["Club (Costco)","Max 14% markup (15% Kirkland); FY24 actual GM 10.92% — volume not margin play","HARD"],
 ["Distributor (general)","15–25% markup; UNFI avg ~13–14% (range 6–30%)","HARD-ish"],
 ["DSD vs warehouse","DSD can save ~5–10% in fees vs UNFI/KeHE routing","RULE OF THUMB"],
 ["Broker","3–5% of net sales + often $5–10k/mo retainer","RULE OF THUMB"],
 ["Amazon","8–45% referral (most 15%) + FBA $3.73–$6.97/unit ($25+ heavy bev)","HARD (published)"],
 ["DTC CAC (food & bev)","~$45–$53 (lowest vertical); supplements ~$89; shipping $8–15/order","HARD (benchmark)"]]
nr=table(ws,nr,["Channel / cost","Take / value","Tag"],ch,widths=[26,64,22],rowh=28)
nr+=1
nr=banner(ws,nr,"TRADE SPEND & FEES",4,color=STEEL)
ts=[["Slotting fee","$250 → $250,000 by retailer/scope; ~$25,000/item regional cluster","RULE OF THUMB"],
 ["Trade spend % of revenue","10–20% (emerging brands high end); US trade spend >$200B, ~20% of gross, #2 P&L line","HARD (aggregate)"],
 ["Free fill","1 free case/SKU/store on first order — cheaper substitute for slotting","RULE OF THUMB"],
 ["LMNT subscription model","$200M+ rev 2023, ~20% net margin, 250k+ customers; free-8-pack (pay shipping) funnel","HARD (company/SEC)"]]
nr=table(ws,nr,["Item","Value","Tag"],ts,widths=[24,66,22],rowh=30)
nr+=1
nr=banner(ws,nr,"CANADA COST STACK (additive on US economics)",4,color=TEAL)
cac=[["Grocery concentration","Loblaw+Sobeys+Metro ≈ 80% w/ Walmart+Costco; all signed Grocery Code of Conduct (in force Jan 1, 2025)","HARD"],
 ["Listing/slotting fees","Charged but NOT publicly disclosed; Code now lets suppliers challenge arbitrary fees — model at ≥ US slotting","No public $"],
 ["Bilingual packaging","EN+FR mandatory; French ~2× longer; non-compliance fines up to $20,000; Bill 96 adds rules","HARD (regulatory)"],
 ["Distributors (CA)","UNFI Canada, Tree of Life Canada, Purity Life, Ecotrend","HARD (existence)"],
 ["Freight/FX/duty","FX (US$1≈C$1.15+) + cross-border freight + duty; model as premium; local co-pack is the margin fix","Qualitative"]]
nr=table(ws,nr,["Item","Value","Tag"],cac,widths=[22,68,22],rowh=30)
nr=callout(ws,nr,"DTC-for-beverage reality check: heavy to ship, low online AOV, CAC can exceed contribution. Historically ~90% of beverage startups fail; the survivable DTC model (LMNT) is subscription + free-sample acquisition + asset-light profitability — not paid-media scale.",4,fillc=AMBER,fontc=AMBERHEAD,h=46)

# ============================================================ 11 FORMULATION
ws=wb.create_sheet("11 Formulation")
nr=title(ws,"Formulation & Positioning — the claim that won",5)
H=["Brand","Winning claim","Hero ingredient & spec","Positioning evolution","Claim risk"]
rows=[
 ["Poppi","Prebiotic soda you can drink daily","Prebiotic fiber + ACV; ~2g fiber, 2–5g sugar","ACV “tonic” (Mother Beverage) → 2020 bright “prebiotic soda”","$8.9M settlement — 2g fiber too little for claim"],
 ["OLIPOP","“New kind of soda”, serious fiber","Up to 9g fiber (cassava/chicory/artichoke) + botanicals","Stable gut-health since launch; ~4× Poppi fiber","No suit found; higher fiber = stronger substantiation"],
 ["LMNT","Zero-sugar, high-sodium for keto/athletes","1,000mg sodium, 200mg K, 60mg Mg, 0g sugar","Salt-first, science-forward","Sodium is the polarizing hook"],
 ["Liquid I.V.","Hydration Multiplier, 2–3× faster","Cellular Transport Tech; 500mg sodium, 11g carbs","Added sugar-free line later","CTT efficacy is marketing-led; sugar a vulnerability"],
 ["Nuun","Carb-free effervescent electrolyte","300mg sodium, 1g sugar, 15 cal, stevia","Endurance carb-free → added powders","Clean low-sugar story"],
 ["Vital Proteins Sparkling","Beauty-from-within + clinical collagen at mass price","VERISOL collagen + 100% DV Vit C; 0g sugar","Extends #1 collagen powder into a can","Strongest collagen claim, but see 2025 efficacy meta-analysis"],
 ["Recess","Relaxation — sell the benefit not the molecule","Magnesium L-threonate + adaptogens","Pivoted CBD → relaxation (CBD <8% of sales)","KEY LESSON: lead with felt benefit"],
 ["Aura Bora","Herbal/floral, zero everything","0 cal/sugar/sodium; herb+fruit+flower","Flavor-forward, no health-claim baggage","Low regulatory risk"],
]
nr=table(ws,nr,H,rows,widths=[16,28,30,30,30],rowh=64)
nr=callout(ws,nr,"READ-ACROSS for MUV: “collagen-as-hero” has a graveyard AND a 2025 no-benefit meta-analysis. Winners wrap collagen in a big brand + clinical claim at mass price (Vital Proteins) OR sell a broader felt benefit (Recess). Lead MUV with gut+hydration function + a credible collagen dose under the “Canada's #1 collagen brand” halo — not narrow beauty-from-within.",5,h=58)

# ============================================================ 12 MARKETING
ws=wb.create_sheet("12 Marketing")
nr=title(ws,"Marketing & Brand Build — the virality engine",4)
H=["Brand","Founder story","Growth engine","What moved the needle"]
rows=[
 ["Poppi","Allison & Stephen Ellsworth; Shark Tank 2018 ($400k/25%, Rohan Oza)","TikTok creator seeding + 2 Super Bowls (2024 first ad, 2025 return)","Pink aesthetic, Alix Earle/Emily Mariko seeding, vending-machine stunt (2025, mixed)"],
 ["OLIPOP","Ben Goodwin (ex-Obi) & David Lester; $100k from Obi sale","DTC repurchase rates + reactive social","“Soda Stories”; needled Poppi's vending spend during 2025 SB — earned media without paying for the game"],
 ["LMNT","Robb Wolf (biochemist) + Ketogains (Villaseñor)","Podcast-sponsorship moat (Huberman, Attia)","Host-read trust + free-sample funnel; $54.6M mktg + $7.1M sampling in 2023 (~26.5% of rev)"],
 ["Liquid I.V.","Brandin Cohen, college brainchild (+2)","In-store/Costco roadshow sampling + give-back","Founder sampled every weekend; POS-data sell-in; Emerson Group for national muscle"],
 ["Nuun","Tim Moxey, British Ironman","Grassroots endurance-community events","Authentic credibility in run/bike/tri before mainstream"],
 ["Celsius","Turnaround 2012; Li Ka-shing $15M (2015)","Amazon + fitness influencers/gym sampling","Amazon velocity → Target DSD → PepsiCo $550M distribution"],
 ["Vital Proteins Sparkling","K. Seidensticker; Nestlé-owned","Big-brand media + retail muscle + collagen loyalty","Distribution & #1-collagen halo, not scrappy virality"],
 ["Recess","Benjamin Witte","Distinctive pastel brand world + culture-tuned social","Owned a vibe + the “relaxation” moment"],
]
nr=table(ws,nr,H,rows,widths=[16,32,30,40],rowh=66)
nr=callout(ws,nr,"READ-ACROSS: The two cheapest high-ROI challenger playbooks are LMNT's trusted-host podcast model and OLIPOP's reactive earned-media social — both winnable without Super Bowl budgets. Pair with Costco-roadshow sampling (Liquid I.V.). Brand tone must evolve with stage (Poppi's vending stunt misfired once mainstream).",4,h=52)

# ============================================================ 13 LAUNCH PLAYBOOKS
ws=wb.create_sheet("13 Launch Playbooks")
nr=title(ws,"Launch Playbook Teardowns — phased GTM timelines",5)
H=["Brand","Phase 0 (pre-launch)","Phase 1 (launch, yr 0–1)","Phase 2 (scale, yr 1–3)","Phase 3 (maturity/exit)"]
rows=[
 ["Poppi","2015–16: gut-health insight; ~$90k self-funded; “Mother Beverage” glass bottles","Farmers markets → Whole Foods regional; ~$500k in 18 mo; Shark Tank 2018 ($400k/25% Oza)","2020 rebrand to cans + national launch (Mar 2020); TikTok creator seeding; #1 Amazon soda 2023","2024 first Super Bowl ad; >$500M rev; PepsiCo $1.95B closed May 2025"],
 ["OLIPOP","Founders' prior brand Obi; $100k seed; nostalgic-soda + fiber insight","Late 2018: 40 NorCal indie grocers; ~$852k yr 1; ~$2.49/can; gut-health science story","Natural → conventional → mass; “astronomical repurchase” sell-in; $73.4M (2022); $39.7M Series B","$200M (2023)→$400M+ (2024); 7,000+ grocers; $50M @ $1.85B (Feb 2025); INDEPENDENT"],
 ["LMNT","2018–19: keto-sodium insight; Robb Wolf + Ketogains; high-sodium no-sugar formula","Late 2019 DTC-only; free Sample Pack (pay ~$5 shipping); profitable yr 1 (claim)","Podcast sponsorships (Huberman et al., ~10k episodes); ~$50M run-rate by Q1 2022","$206M sales (2023, SEC); private, founder-controlled; selective retail; NO exit"],
 ["Liquid I.V.","2012: adults using Pedialyte insight; WHO-ORS-based CTT formula","Whole Foods “Supplier 101”; founder samples weekly → 12 WF by 2015","Costco roadshow sampling; Emerson Group (2015); ~$100M within 5 yrs","Unilever Sept 2020 (~$500M est.); 30k→80k stores; $1B+ brand"],
 ["Celsius","2004 (Elite FX); near-bankruptcy/delisting ~2010 — over-marketing before PMF","Turnaround 2012; Li Ka-shing $15M (2015); NASDAQ uplisting 2017 (~$36M rev)","Amazon ~15–20% of sales (2020); Target DSD; 100k+ doors by 2021","PepsiCo $550M/8.5% (2022) → ~#3 energy; bought Alani Nu $1.8B (2025)"],
]
nr=table(ws,nr,H,rows,widths=[12,28,28,28,28],rowh=92)
ws.freeze_panes="A4"
nr=callout(ws,nr,"DATE TRAP: Poppi's FIRST Super Bowl ad = 2024 (LVIII); the vending-machine controversy = the 2025 (LIX) return. Many secondary sources conflate them.",5,fillc=AMBER,fontc=AMBERHEAD,h=30)

# ============================================================ 14 PATTERNS & FAILURES
ws=wb.create_sheet("14 Patterns & Failures")
nr=title(ws,"Success Patterns & Failure Modes",2)
ws.column_dimensions["A"].width=4; ws.column_dimensions["B"].width=114
nr=banner(ws,nr,"SEVEN SUCCESS PATTERNS",2,color=GREENHEAD)
pats=["Founder-authority insight tied to a personal health problem — cheap, durable brand moat (Ellsworth gut health, Wolf sodium, Cohen Pedialyte).",
 "ONE concentrated trial-generation tactic, executed obsessively, before broad spend.",
 "Sell-in on DATA (velocity/repurchase), not flavour — the single most repeated lever.",
 "Sequenced channel ladder — natural/DTC → specialty → grocery → mass → club; never skip rungs.",
 "Strategic-investor signal early (Oza, Li Ka-shing, celebrity rounds) — capital + credibility + distribution.",
 "Category-creation language to escape commodity comparison and hold a $2–$3/can premium.",
 "Endgame split: strategic exit (Poppi/Liquid I.V./Celsius) vs independence (OLIPOP/LMNT) — both valid; scale eventually needs a partner or heavy capital."]
r=nr
for i,p in enumerate(pats,1):
    ws.cell(row=r,column=1,value=i).font=BODYB
    c=ws.cell(row=r,column=2,value=p); c.font=BODY; c.alignment=WRAP; c.border=BORDER
    c.fill=fill(LIGHT2 if i%2 else WHITE); ws.row_dimensions[r].height=30; r+=1
r+=1
r=banner(ws,r,"SIX FAILURE MODES",2,color=REDHEAD)
fails=["Over-distribution before product-market fit — Celsius's own ~2010 near-bankruptcy is the canonical case. Velocity must precede footprint.",
 "Undercapitalized / capital-intensive DTC — beverages are heavy physical goods; ~90% of beverage startups failed historically.",
 "Premium price with no defensible moat / thin claims — ~32% of consumers cite “too expensive”; premium needs credible function + repeat.",
 "Trial without retention — sampling/virality wasted if repurchase is weak (inverse of the OLIPOP/Liquid I.V. data story).",
 "Brand/tone missteps at scale — Poppi's 2025 vending stunt; tactics must evolve with brand stage.",
 "Wrong timing / flavour failure — the baseline reasons ~85% of new CPG products fail within 2 years."]
for i,p in enumerate(fails,1):
    ws.cell(row=r,column=1,value=i).font=BODYB
    c=ws.cell(row=r,column=2,value=p); c.font=BODY; c.alignment=WRAP; c.border=BORDER
    c.fill=fill(RED if i%2 else WHITE); ws.row_dimensions[r].height=30; r+=1

# ============================================================ 15 OUTCOMES & M&A
ws=wb.create_sheet("15 Outcomes & M&A")
nr=title(ws,"Financial Outcomes & M&A Comps",6)
H=["Brand","Revenue trajectory","Funding / valuation","Exit / status","Category share","Profitability"]
rows=[
 ["Poppi","$13M(20)→$65M(22)→$100M+(23)→~$500M(24)","VC-backed pre-exit","PepsiCo $1.95B (net ~$1.65B), May 2025","~38% modern soda (24)","n/a"],
 ["OLIPOP","$852k(18)→$200M+(23)→$400M+(24)","$50M Series C @ $1.85B (Feb 2025, JPM)","Independent","~32.7% modern soda (24)","Cash-profitable early 2024"],
 ["LMNT","$206M sales (2023, SEC)","Reg CF / Republic","Independent","n/a","~20% net margin (2023)"],
 ["Liquid I.V.","~$100M in 5 yrs → ~$200M exit → $1B+ now","—","Unilever Sept 2020 (~$500M est.)","n/a","Unilever +$80M MO plant"],
 ["Nuun","$10M+ (2013)","TSG Consumer (2017)","Nestlé HS July 2021 (undisclosed)","#1 run/bike/outdoor","n/a"],
 ["Celsius","multi-$B (public)","Li Ka-shing $15M (2015)","Public; PepsiCo $550M (2022); bought Alani Nu $1.8B","~#3 US energy (~11.5%)","Profitable"],
 ["Health-Ade","~$250M/yr retail","~$54M raised","Generous Brands $500M","—","—"],
 ["Alani Nu","~$595M (2024)","Congo Brands","Celsius $1.8B (2025)","—","—"],
 ["Vital Proteins","(in Nestlé HS)","—","Nestlé (full 2022)","—","—"],
 ["Aura Bora","$12M (2024)","~$22M raised","Majority to Next In Natural (2025)","—","—"],
]
nr=table(ws,nr,H,rows,widths=[14,34,26,30,18,18],rowh=34)
nr=callout(ws,nr,"M&A READ: Strategics now own most winners — PepsiCo (Poppi, Celsius stake), Unilever (Liquid I.V., Olly), Nestlé (Vital Proteins, Nuun), KDP (Ghost). For a challenger the realistic endgame is acquisition, not durable independence — but Coke's organic Simply Pop shows incumbents may CLONE rather than buy, compressing the premium. OLIPOP (independent, $1.85B) is the counter-example.",6,h=58)

# ============================================================ 16 SWOT
ws=wb.create_sheet("16 SWOT")
nr=title(ws,"Per-Brand SWOT",5)
H=["Brand","Strengths","Weaknesses","Opportunities","Threats"]
rows=[
 ["Poppi","Mass brand love, TikTok engine, PepsiCo distribution","Thin fiber dose, claims litigation history","PepsiCo scale, international","Commoditization, Simply Pop, claim scrutiny"],
 ["OLIPOP","9g fiber moat, repeat rate, independence, $1.85B val","Premium price, single category","Mass expansion, intl, format extensions","PepsiCo-backed Poppi, Coke Simply Pop, private label"],
 ["LMNT","Profitable, capital-light, podcast moat, 20% net margin","Polarizing high sodium, niche, DTC-dependent","Retail expansion, sparkling line, intl","Electrolyte crowding (Prime, LMNT clones)"],
 ["Liquid I.V.","Unilever scale, Costco engine, $1B+ brand","Sugar content, undifferentiated vs new entrants","Global expansion, sugar-free line","Crowded hydration, Prime, private label"],
 ["Vital Proteins Sparkling","Nestlé muscle, #1 collagen halo, clinical claim, mass price","New format unproven, collagen efficacy debate","Owns beauty-from-within mainstreaming","2025 no-benefit meta-analysis, claims risk"],
 ["MUV (Organika)","Domestic CA producer, #1 CA collagen halo, existing CA shelf, cost edge","Unproven RTD, food-format claim limits, small vs strategics","Costco CA/Amazon beachhead, gut+beauty stack, Supplemented-Food pathway","Nestlé/Vital Proteins, collagen efficacy/claims liability, shelf saturation"],
]
nr=table(ws,nr,H,rows,widths=[16,26,26,26,26],rowh=58)
ws.freeze_panes="A4"

# ============================================================ 17 SCORING MODEL
ws=wb.create_sheet("17 Scoring Model")
nr=title(ws,"Weighted Competitive Scoring Model",10,"Scores 1–5 (5=strongest). Edit weights in row 4; totals auto-calc. Illustrative scores — calibrate with your data")
crit=[("Brand strength",0.15),("Distribution reach",0.15),("Claim defensibility",0.15),("Margin/price power",0.10),
 ("Innovation/format",0.10),("Marketing engine",0.10),("Capital/backing",0.15),("Canada readiness",0.10)]
# header
hdr=["Brand"]+[c for c,_ in crit]+["Weighted score"]
ws.cell(row=nr,column=1,value="WEIGHTS →").font=BODYB
for j,(c,w) in enumerate(crit,2):
    x=ws.cell(row=nr,column=j,value=w); x.font=BODYB; x.alignment=CTR; x.number_format="0%"; x.fill=fill(AMBER); x.border=BORDER
ws.cell(row=nr,column=2+len(crit),value="").border=BORDER
weight_row=nr; nr+=1
hdrr=nr
for j,h in enumerate(hdr,1): ws.cell(row=hdrr,column=j,value=h)
hdrow(ws,hdrr,len(hdr),NAVY)
scores=[
 ("Poppi",[5,5,2,3,3,5,5,2]),
 ("OLIPOP",[5,5,4,3,4,4,4,2]),
 ("LMNT",[4,3,3,4,3,5,4,2]),
 ("Liquid I.V.",[4,5,3,3,3,4,5,3]),
 ("Vital Proteins Sparkling",[4,5,4,3,3,3,5,2]),
 ("Recess",[3,3,3,3,4,4,3,1]),
 ("MUV (Organika)",[3,3,3,4,3,2,2,5]),
]
r=nr+1
firstdata=r
for i,(b,sc) in enumerate(scores):
    ws.cell(row=r,column=1,value=b).font=BODY; ws.cell(row=r,column=1).border=BORDER
    if i%2: ws.cell(row=r,column=1).fill=fill(LIGHT)
    for j,s in enumerate(sc,2):
        x=ws.cell(row=r,column=j,value=s); x.font=BODY; x.alignment=CTR; x.border=BORDER
        if i%2: x.fill=fill(LIGHT)
    # weighted = sumproduct(scores, weights)
    cols=[get_column_letter(j) for j in range(2,2+len(crit))]
    wcol=get_column_letter(2+len(crit))
    formula="=SUMPRODUCT("+get_column_letter(2)+str(r)+":"+get_column_letter(1+len(crit))+str(r)+","+\
            get_column_letter(2)+"$"+str(weight_row)+":"+get_column_letter(1+len(crit))+"$"+str(weight_row)+")"
    x=ws.cell(row=r,column=2+len(crit),value=formula); x.font=BODYB; x.alignment=CTR; x.border=BORDER
    x.number_format="0.00"
    if i%2: x.fill=fill(LIGHT)
    r+=1
ws.column_dimensions["A"].width=22
for j in range(2,2+len(crit)): ws.column_dimensions[get_column_letter(j)].width=12
ws.column_dimensions[get_column_letter(2+len(crit))].width=14
nr=r+1
nr=callout(ws,nr,"HOW TO USE: weights (amber row) are editable and should sum to 100%; weighted score = SUMPRODUCT(scores × weights). Scores are illustrative analyst judgment — recalibrate against your own velocity/claims/financial data. MUV scores high on Canada readiness and margin/price power, low on marketing engine and capital — which is exactly where the launch plan must invest.",10,h=58)

# ============================================================ 18 CANADA MARKET
ws=wb.create_sheet("18 Canada Market")
nr=title(ws,"Canada Market — Retail Landscape & Entry Playbooks",5)
nr=banner(ws,nr,"GROCERY BANNERS & APPROXIMATE SHARE (varies by methodology — treat as ranges)",5,color=STEEL)
ban=[["Loblaw","Loblaws, No Frills, Shoppers Drug Mart, RC Superstore, Maxi","~26–30%","Largest; pharmacy via Shoppers"],
 ["Empire/Sobeys","Sobeys, Safeway, IGA, FreshCo, Foodland","~20–21%","#2"],
 ["Costco Canada","Costco","~14%","Proven beachhead for new functional bev"],
 ["Walmart Canada","Walmart","~12%","Mass"],
 ["Metro","Metro, Food Basics, Jean Coutu, Super C","~11–12%","Strong QC + ON; pharmacy via Jean Coutu"],
 ["Pattison/Save-On","Save-On-Foods, Buy-Low, Choices, Quality Foods","Regional (West)","187 Save-On stores"]]
nr=table(ws,nr,["Retailer","Banners","Est. share","Note"],ban,widths=[16,40,14,30],rowh=28)
nr+=1
nr=banner(ws,nr,"NATURAL & E-COMMERCE CHANNEL",5,color=TEAL)
nat=[["Whole Foods Canada","Natural anchor (Amazon-owned)","Premium natural shelf"],
 ["Healthy Planet","“Canada's largest online health store”, ~42 ON stores","Omnichannel natural"],
 ["Nature's Fare / Goodness Me","Regional natural (BC / ON)","Regional"],
 ["Amazon.ca","Primary online channel","Liquid I.V./OLIPOP/Poppi all sell here"],
 ["Well.ca","Health/beauty/natural (McKesson-owned)","Organika already sells here"],
 ["Natura Market / Vitasave / SPUD","Online natural marketplaces","Front-run official distribution"]]
nr=table(ws,nr,["Channel","What it is","Note"],nat,widths=[24,42,34],rowh=24)
nr+=1
nr=banner(ws,nr,"HOW US BRANDS ENTERED CANADA (the playbook)",5,color=STEEL)
ent=[["Liquid I.V.","2023 (first market outside US)","Via Unilever; Costco FIRST → Amazon.ca, Walmart, Loblaws, Jean Coutu"],
 ["Poppi","Aug 13, 2024 (first international)","Brand-led; Loblaws, Metro, Save-On, Maxi, Safeway, Costco, Sobeys + Well.ca, Natura"],
 ["OLIPOP","Limited","Distributor NevrAsk; Amazon.ca, Natura (specialty e-comm)"],
 ["Celsius","Jan 2024","PepsiCo exclusive distribution (extends 2022 US deal)"],
 ["Prime","2023","Hydration sells; ENERGY blocked (200mg caffeine > 180mg cap)"]]
nr=table(ws,nr,["Brand","When","Entry model"],ent,widths=[14,24,62],rowh=28)
nr+=1
nr=banner(ws,nr,"PRICING",5,color=TEAL)
pr=[["Premium functional can","~CAD $3.49 MSRP (Poppi); promo ~$2.50","Shelf norm to anchor against"],
 ["Cross-border premium","FX (US$1≈C$1.15+) + freight + duty","US$2.50 can lands well above C$3.49 imported"],
 ["GST/HST","5–15% by province; import GST on duty-incl value","Plus tariff status to verify (relief extended to Dec 16, 2025)"],
 ["Margin fix","Local/contract co-packing in Canada","Organika's domestic production is the structural edge"]]
nr=table(ws,nr,["Item","Value","Note"],pr,widths=[20,42,38],rowh=26)
nr=callout(ws,nr,"WHAT WORKED: (1) enter via a big-box anchor (Costco) or a strategic distributor/parent DSD (Unilever, PepsiCo); (2) reformulate to Canadian limits BEFORE launch (Prime Energy is the cautionary tale); (3) seed Amazon.ca + Natura/Well.ca ahead of mainstream grocery. Organika already has #2 and the CA shelf relationships.",5,h=50)

# ============================================================ 19 CANADA REGULATORY
ws=wb.create_sheet("19 Canada Regulatory")
nr=title(ws,"Canada Regulatory Pathway — the decision-critical tab",3,"An RTD collagen can is a FOOD / Supplemented Food (CFIA), NOT an NHP")
nr=banner(ws,nr,"CLASSIFICATION — NHP vs FOOD vs SUPPLEMENTED FOOD",3,color=STEEL)
cls=[["Core finding","An RTD sparkling beverage looks/packages/consumes like food → regulated as a FOOD by Health Canada Food Directorate + CFIA, OUTSIDE the NHP framework","HARD (Health Canada guidance)"],
 ["Consequence","The can generally CANNOT carry an NPN or make NPN-licensed therapeutic/beauty claims. NPNs attach to capsules/tablets/controlled-dose NHPs","HARD"],
 ["Likely pathway","A can with added vitamins/minerals/amino acids/electrolytes/caffeine beyond normal food levels = a SUPPLEMENTED FOOD under the Supplemented Foods Regulations (in force Jul 21, 2022)","HARD"],
 ["What SFR sets","(a) List of Permitted Supplemental Ingredients w/ max amounts; (b) Supplemented Food Facts table; (c) cautionary statements; (d) Supplemented Food Caution Identifier (SFCI)","HARD"],
 ["SFCI box","Black ‘!’ “Supplemented/Supplémenté” box — REQUIRED only when the product needs cautionary statements; cannot be used decoratively. A modest formula may avoid it — confirm per ingredient","DIRECTIONAL – confirm"],
 ["Compliance dates","Legacy transition ends Dec 31, 2025; ALL supplemented foods need compliant labels by Jan 1, 2026","HARD-ish"]]
nr=table(ws,nr,["Item","Detail","Confidence"],cls,widths=[20,74,24],rowh=46)
nr+=1
nr=banner(ws,nr,"INGREDIENT-SPECIFIC RULES",3,color=STEEL)
ing=[["Collagen caution (DO NOT MISS)","Foods with hydrolyzed collagen/gelatin must show on the principal display panel, in same type size as the common name: “CAUTION, DO NOT USE AS SOLE SOURCE OF NUTRITION” (FDR C.R.C. c.870)","HARD (regulation)"],
 ["Caffeine cap","Caffeinated supplemented foods: max 180mg/serving & /container — this BLOCKED Prime Energy (200mg). Recommend MUV be caffeine-free to avoid the energy-drink caution regime","HARD"],
 ["Claims","“Beauty/skin/anti-aging” structure-function claims are an NHP construct — high-risk on a food/supplemented-food label; only permitted food/function/nutrient-content claims allowed","DIRECTIONAL – counsel needed"]]
nr=table(ws,nr,["Rule","Detail","Confidence"],ing,widths=[20,74,24],rowh=50)
nr+=1
nr=banner(ws,nr,"LABELLING",3,color=TEAL)
lab=[["Bilingual EN/FR","Mandatory on all required info incl. Nutrition/Supplemented Food Facts","HARD"],
 ["Quebec Bill 96 / OQLF","From Jun 1, 2025 French ≥ as prominent as any other language; trademark descriptors in French; grace to Jun 1, 2027; fines $3,000–$30,000","DIRECTIONAL"],
 ["Facts table","Supplemented Food Facts table (bilingual) replaces standard NFt where applicable","HARD"]]
nr=table(ws,nr,["Item","Detail","Confidence"],lab,widths=[20,74,24],rowh=34)
nr+=1
# route comparison
nr=banner(ws,nr,"ROUTE COMPARISON",3,color=STEEL)
rt=[["Format","Capsule/tablet/controlled-dose","RTD can / beverage ← MUV"],
 ["Pre-market approval","NPN licensing required","No NPN; self-comply with SFR"],
 ["Claims","Monograph/clinical (“beauty from within” possible)","Limited to permitted food/function claims"],
 ["Label","NHP facts box + NPN","Supplemented Food Facts + possible SFCI box + collagen caution"],
 ["Speed to market","Slower (licensing queue)","Faster (self-compliance)"]]
nr=table(ws,nr,["Dimension","NHP route","Food / Supplemented Food route (MUV)"],rt,widths=[20,40,58],rowh=30)
nr=callout(ws,nr,"VERIFY BEFORE ARTWORK (mandatory): (1) whether MUV's exact collagen+electrolyte formula triggers the SFCI caution box; (2) which skin/beauty claim language survives on a food label; (3) tariff status on US-origin inputs post-Dec 16, 2025. Engage Canadian regulatory counsel on the claim set FIRST — it reshapes packaging, marketing, and the entire value proposition.",3,fillc=AMBER,fontc=AMBERHEAD,h=56)

# ============================================================ 20 GTM RECOMMENDATION
ws=wb.create_sheet("20 GTM Recommendation")
nr=title(ws,"MUV Go-to-Market Recommendation",2)
ws.column_dimensions["A"].width=26; ws.column_dimensions["B"].width=92
nr=callout(ws,nr,"VERDICT: GO — conditional on resolving the Supplemented-Foods claim/label pathway and dosing collagen to a defensible level. The tailwinds are real and Organika has a genuine structural edge in Canada; the risks (collagen efficacy/claims, strategic competition) are manageable with discipline.",2,h=52)
recs=[
 ("1. Positioning","Lead with a FUNCTIONAL benefit stack (gut/prebiotic + hydration + a credible collagen dose) under a “by Canada's #1 collagen brand” halo. Avoid narrow beauty-from-within hero claims — the graveyard + 2025 efficacy meta-analysis make it the riskiest lane. Function-stacking also hedges single-fad risk."),
 ("2. Formulation","Dose collagen to a substantiable level (2.5g+; positive studies used 2.5–10g/day). Zero/low sugar is table stakes. Add prebiotic fiber for a defensible gut hook (OLIPOP's 9g beat Poppi's 2g in court). Caffeine-free to avoid the 180mg energy-drink caution regime."),
 ("3. Channel beachhead","Costco Canada + Amazon.ca + natural channel (Whole Foods CA, Healthy Planet) + Organika's existing pharmacy/health-food/Well.ca base. This mirrors Liquid I.V./Poppi's Canada entry — and Organika already holds the relationships. Do NOT chase national grocery before proving velocity."),
 ("4. Pricing","Anchor at/below the ~$3.49 CAD premium-functional shelf norm. Domestic production lets MUV protect margin where imported US brands cannot — a durable advantage. Plan 35–45% trade spend in early scale; use free-fill over cash slotting where possible."),
 ("5. Marketing","Run the cheap challenger playbooks: trusted-host podcast sponsorship (LMNT model) + reactive earned-media social (OLIPOP model) + Costco roadshow sampling (Liquid I.V. model). Skip Super Bowl-tier spend. Sell-in to buyers on velocity/repeat data, not taste."),
 ("6. Claims & legal","Engage Canadian regulatory counsel on the Supplemented-Foods claim set BEFORE artwork. Build substantiation files. Include the mandatory collagen “do not use as sole source of nutrition” caution. Dose-to-claim discipline is existential (Poppi paid $8.9M)."),
 ("7. Moat & exit","Differentiate on a real, defensible spec + Canadian-made authenticity + Organika's collagen credibility. Plan for an acquisition-oriented endgame (strategics own the category) while building OLIPOP-style independence optionality through profitability and repeat rate."),
]
nr=table(ws,nr,["Decision","Recommendation"],recs,widths=[26,92],rowh=78)

# ============================================================ 21 LAUNCH PLAN
ws=wb.create_sheet("21 Launch Plan")
nr=title(ws,"Phased Launch Plan (0–18 months)",5)
H=["Phase","Timing","Objectives","Key workstreams","Success gate"]
rows=[
 ["Phase 0 — Foundation","Months 0–3","Lock product, claims, regulatory pathway, co-pack","Finalize formula & collagen dose; regulatory-counsel claim opinion (Supplemented Foods); co-packer quote & MOQ; bilingual + Bill 96 artwork; brand/positioning","Regulatory sign-off on claim set + artwork approved"],
 ["Phase 1 — Beachhead","Months 3–9","Prove velocity & repeat in a contained beachhead","Costco CA roadshows + Amazon.ca + Well.ca/Natura + Organika pharmacy/health-food base; podcast + reactive social; sampling","Repeat rate + velocity above category benchmark"],
 ["Phase 2 — Scale","Months 9–18","Earn conventional grocery on proven data","Sell-in to Loblaw/Sobeys/Metro on velocity data; distributor (UNFI CA / Tree of Life); managed trade spend; add SKUs on repeat data","National listings won on data, not discounts"],
 ["Phase 3 — Defend & extend","18 mo+","Category leadership & optionality","Format/flavour extensions; deepen moat; evaluate strategic-partner vs independent scale","Profitable contribution + exit optionality"],
]
nr=table(ws,nr,H,rows,widths=[18,12,30,46,30],rowh=80)
ws.freeze_panes="A4"

# ============================================================ 22 RISK REGISTER
ws=wb.create_sheet("22 Risk Register")
nr=title(ws,"Risk Register",6,"Likelihood × Impact (H/M/L). Ranked by exposure")
H=["#","Risk","Category","Likelihood","Impact","Mitigation"]
rows=[
 [1,"Collagen efficacy / claims liability (2025 meta-analysis; Poppi-style dose-insufficiency suit)","Regulatory/Legal","High","High","Dose to efficacy (2.5g+); structure/function-safe language; substantiation files; lead with gut/hydration not beauty"],
 [2,"Strategic-owned competition (Nestlé owns Vital Proteins; PepsiCo/Unilever firepower)","Competitive","High","High","Compete on Canadian-made + domestic cost edge + Organika collagen credibility; niche beachhead before they react"],
 [3,"Canada claim/label pathway misread (food vs NHP; SFCI; collagen caution)","Regulatory","Med","High","Regulatory-counsel opinion before artwork; design to Supplemented Foods framework; caffeine-free"],
 [4,"Over-distribution before product-market fit","Execution","Med","High","Prove velocity/repeat in beachhead before national; gate Phase 2 on data"],
 [5,"Slotting / trade-spend cash burn","Financial","Med","High","Use free-fill over cash slotting; 35–45% trade-spend budget; Code of Conduct leverage on arbitrary fees"],
 [6,"Thin differentiation / shelf saturation (prebiotic shelf filling fast)","Category","Med","Med","Defensible spec (fiber + collagen dose) + function stacking; Canadian-made story"],
 [7,"Private-label encroachment (~$277B, ~21% share, growing 2×)","Category","Med","Med","Brand equity + clinical-grade ingredients retailers can't easily copy"],
 [8,"DTC CAC / heavy-ship economics if DTC-led","Financial","Med","Med","Lead retail/beachhead not paid DTC; subscription + sampling only"],
 [9,"Co-packer capacity / collagen sourcing / cost inflation","Operations","Med","Med","Lock co-pack capacity early; qualify 2nd collagen supplier; hedge raws"],
 [10,"FX / tariff / freight premium on cross-border inputs","Financial","Med","Med","Domestic CA co-pack; monitor tariff status post-Dec 2025"],
 [11,"GLP-1 demand shift away from sweet/snack","Macro","Med","Low","Low-cal, fiber, protein, hydration positioning aligns with GLP-1 users"],
 [12,"Brand/tone misstep at scale","Brand","Low","Med","Evolve tactics with stage; avoid Poppi-style stunts once mainstream"],
]
nr=table(ws,nr,H,rows,widths=[4,38,16,12,10,44],rowh=46)
ws.freeze_panes="A4"
# color likelihood/impact
for rr in range(nr-len(rows),nr):
    for col in (4,5):
        v=ws.cell(row=rr,column=col).value
        if v=="High": ws.cell(row=rr,column=col).fill=fill(RED)
        elif v=="Med": ws.cell(row=rr,column=col).fill=fill(AMBER)
        elif v=="Low": ws.cell(row=rr,column=col).fill=fill(GREEN)
        ws.cell(row=rr,column=col).alignment=CTR

# ============================================================ 23 KPI DASHBOARD
ws=wb.create_sheet("23 KPI Dashboard")
nr=title(ws,"KPI Dashboard & Targets",5,"Benchmarks from this research; set MUV targets before launch")
H=["KPI","Why it matters","Category benchmark","MUV target (set)","Phase"]
rows=[
 ["Repeat purchase rate","The metric buyers actually buy on","OLIPOP cited “astronomical”; aim well above category avg","[set]","Phase 1+"],
 ["Velocity (units/store/week)","Gates shelf retention & expansion","Buyers reward durable, non-discount velocity","[set]","Phase 1+"],
 ["Gross margin","Survival","Aim 50–65% at scale (40–50% early); watch COGS definition","[set]","All"],
 ["Trade spend % of revenue","#2 P&L line; cash burn risk","10–20%; emerging brands high end","≤45% early","Phase 1–2"],
 ["CAC (if DTC/sub)","DTC viability","F&B DTC ~$45–$53","≤ contribution margin","Phase 1+"],
 ["Doors / distribution","Reach — but only on proven data","Poppi ~36k / OLIPOP ~35k US (mature)","Beachhead first","Phase 1–3"],
 ["Net revenue","Scale","$852k (OLIPOP yr1) → $400M+ (yr6)","[set]","All"],
 ["Brand awareness / social","Demand-gen efficiency","TikTok/podcast-led; earned-media ratio","[set]","Phase 1+"],
]
nr=table(ws,nr,H,rows,widths=[24,32,38,20,14],rowh=34)

# ============================================================ 24 SOURCES
ws=wb.create_sheet("24 Sources")
nr=title(ws,"Master Bibliography",3,"Tagged by type. Several primary pages block automated fetch (403) — re-verify load-bearing numbers from the cited URLs before external use")
srcs=[
 ["PepsiCo — Poppi acquisition ($1.95B, closed May 2025)","pepsico.com/newsroom","Primary"],
 ["Poppi $8.9M gut-health class-action settlement (final approval Apr 2026)","classaction.org; bevnet.com; axios.com","Primary/Trade"],
 ["OLIPOP $50M Series C @ $1.85B; revenue; profitability","cnbc.com; fooddive.com; bevindustry.com","Primary/Trade"],
 ["Modern soda $1.8B 2024 (+83%); brand shares (Poppi 38% / OLIPOP 32.7%)","Circana; bevindustry.com; csnews.com","Scanner/Trade"],
 ["LMNT FY2023 ($206M sales, ~20% net, $54.6M mktg, $7.1M sampling, 250k+ customers)","SEC Reg CF CIK 1871551","Primary (strongest)"],
 ["Powdered hydration $1.5B (+20%); electrolyte powder $2.8B","Circana/CNN; Grand View","Scanner/Research-firm"],
 ["Liquid I.V. Costco national launch (2019, ~516 whs); Unilever acq. Sept 2020","liquid-iv.com; unilever.com; bevnet.com","Primary/Trade"],
 ["Nuun — Nestlé Health Science acquisition (July 2021)","nestle.com; businesswire.com","Primary"],
 ["Vital Proteins sparkling ($2.50/can, VERISOL); Nestlé ownership","fooddive.com; wwd.com; foodbusinessnews.net","Trade/Primary"],
 ["Celsius — PepsiCo $550M/8.5% (2022); Alani Nu $1.8B (2025)","bevnet.com; bloomberg.com; businesswire.com","Primary/Trade"],
 ["Health-Ade — Generous Brands $500M; $4M kombucha settlement","fooddive.com; bursor.com","Trade"],
 ["Collagen efficacy — 2025 systematic review/meta-analysis (no proven skin benefit; funding bias)","Am. Journal of Medicine; nutraingredients.com","Primary (peer-reviewed)"],
 ["GLP-1 impact (23% households; 35% of units by 2030; spend -5.3%)","Morning Consult; Circana; Food Dive","Research/Trade"],
 ["Beauty-from-within drinks $3.24B, women 68.4%","Future Market Insights; Grand View","Research-firm (directional)"],
 ["Channel economics (Amazon fees, Costco margin, DTC CAC, slotting, distributor)","sellingpartners.aboutamazon.com; massivemoats; eightx.co; balancedbusinessgroup.com","Mixed (Amazon/Costco HARD)"],
 ["Canada — NHP vs Food classification (food-format guidance)","canada.ca Health Canada guidance","Primary (regulatory)"],
 ["Canada — Supplemented Foods Regulations (in force Jul 2022; SFCI; Jan 2026)","canada.ca; inspection.canada.ca; fasken.com","Primary/Legal"],
 ["Canada — collagen caution statement (FDR C.R.C. c.870)","laws.justice.gc.ca","Primary (regulation)"],
 ["Canada — caffeine 180mg cap (Prime Energy blocked)","canada.ca; stack3d.com","Primary/Trade"],
 ["Canada — Bill 96 / OQLF packaging (Jun 2025)","mltaikins.com; stikeman.com","Legal"],
 ["Canada — grocery share & Code of Conduct (Jan 2025)","USDA FAS; statista.com; cbc.ca; canadacode.org","Govt/Trade"],
 ["US brand Canada entry (Liquid I.V. 2023, Poppi Aug 2024, Celsius Jan 2024)","foodincanada.com; newswire.ca; prnewswire.com","Trade/Primary"],
 ["Launch playbooks (Poppi/OLIPOP/LMNT/Liquid I.V./Celsius timelines)","fortune.com; entrepreneur.com; cnbc.com; inc.com; hubermanlab.com","Trade/Primary"],
 ["Extended landscape (Culture Pop, Spindrift, Prime, Cure, etc.)","bevnet.com; techcrunch.com; prnewswire.com; wikipedia","Trade (verify FLAGs)"],
 ["Private label ~$277B, ~21% share","grocerydive.com; nielseniq.com","Trade"],
 ["CPG failure rate ~85% in 2 years","foodnavigator-usa.com; eightx.co","Trade (directional)"],
]
nr=table(ws,nr,["Claim / topic","Source(s)","Type"],srcs,widths=[58,40,20],rowh=26)
nr+=1
nr=callout(ws,nr,"PROCESS NOTE: Numbers were captured via multi-source web research with per-claim citation. Primary publisher and canada.ca pages frequently block automated fetch (HTTP 403); load-bearing figures should be re-pulled directly from the cited URLs (or licensed Circana/SPINS/Mintel portals) before external publication. The two highest-stakes regulatory claims — (a) RTD beverages are foods not NHPs, (b) the Supplemented-Foods/SFCI pathway with Jan 1 2026 compliance — are each corroborated by an official source + ≥1 independent secondary.",3,fillc=AMBER,fontc=AMBERHEAD,h=58)

# ============================================================ 25 GLOSSARY
ws=wb.create_sheet("25 Glossary")
nr=title(ws,"Glossary & Abbreviations",2)
ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=96
gl=[
 ["ACV","Apple cider vinegar (Poppi's original hero ingredient)"],
 ["ACV (retail)","Also: All-Commodity Volume — a retail distribution/weighting metric (context-dependent)"],
 ["CFIA","Canadian Food Inspection Agency — regulates food-format products in Canada"],
 ["CTT","Cellular Transport Technology — Liquid I.V.'s hydration claim"],
 ["DSD","Direct Store Delivery — brand/bottler delivers & merchandises in-store (Coca-Cola model)"],
 ["DTC","Direct-to-consumer (e-commerce / subscription)"],
 ["FBA","Fulfillment by Amazon"],
 ["FDR","Food and Drug Regulations (Canada)"],
 ["GLP-1","Glucagon-like peptide-1 agonists (Ozempic/Wegovy) — weight-loss drugs reshaping food demand"],
 ["GM / COGS","Gross margin / Cost of goods sold"],
 ["MOQ","Minimum order quantity (co-packer runs)"],
 ["MULO","Multi-Outlet (SPINS/Circana retail measurement universe)"],
 ["NHP / NPN","Natural Health Product / Natural Product Number (Canada licensing for supplements)"],
 ["PMF","Product-market fit"],
 ["RTD","Ready-to-drink (a finished beverage, vs powder/tablet)"],
 ["SAM/SOM/TAM","Serviceable available / obtainable / total addressable market"],
 ["SFCI","Supplemented Food Caution Identifier — Canada's ‘!’ caution box for supplemented foods"],
 ["SFR","Supplemented Foods Regulations (Canada, in force Jul 21 2022)"],
 ["SPINS / Circana","Syndicated retail scanner-data providers (point-of-sale sales & velocity)"],
 ["Slotting fee","Payment to a retailer to list a new SKU / secure shelf space"],
 ["Trade spend","Promotional/discount spend to retailers — typically the #2 P&L line after COGS"],
 ["UNFI / KeHE","Major North American natural/specialty food distributors"],
 ["VERISOL","A clinically-studied collagen peptide (Vital Proteins sparkling)"],
]
nr=table(ws,nr,["Term","Definition"],gl,widths=[22,96],rowh=22)

out="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb.save(out)
print("saved:",out)
print("sheets:",len(wb.sheetnames))
for s in wb.sheetnames: print(" -",s)
