#!/usr/bin/env python3
"""Loop batch I19: navigation legend on the Contents tab — explains the color coding and how to move
around, so the interface is self-explaining. Safe append (no row inserts, no merged-cell edits)."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; STEEL="3E5C76"; LIGHT="EAF1F4"; GOLD="C9A227"; GREY="595959"; WHITE="FFFFFF"
def F(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

toc=wb["01 Contents"]
r=toc.max_row+2
toc.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
h=toc.cell(row=r,column=1,value="HOW TO NAVIGATE"); h.font=F(bold=True,color=WHITE,size=11); h.fill=fill(STEEL); h.alignment=Alignment(horizontal="left",vertical="center",indent=1)
toc.row_dimensions[r].height=20
r+=1
legend=[
 ("Click any tab name above","jumps straight to that tab"),
 ("🏠 Start Here","the front door — grouped cards to every section (workbook opens here)"),
 ("🏠 Home (top-right of every tab)","returns to Start Here"),
 ("Gold tab","interactive: Financial Model, Scenario Comparison, Go-NoGo, Peer Set, Data Confidence, etc."),
 ("Executive Brief 1-page","the whole recommendation on one printable page"),
 ("Confidence tags ✅ / ◐ / ⚠","hard evidence / directional / estimate-or-flag (see Data Confidence scoreboard)"),
]
for i,(a,b) in enumerate(legend):
    x=toc.cell(row=r,column=1,value=a); x.font=F(bold=True,size=10); x.border=BORDER; x.alignment=Alignment(vertical="center",wrap_text=True)
    y=toc.cell(row=r,column=2,value=b); y.font=F(size=10); y.border=BORDER; y.alignment=Alignment(wrap_text=True,vertical="center")
    if i%2: x.fill=fill(LIGHT); y.fill=fill(LIGHT)
    toc.row_dimensions[r].height=22
    r+=1
wb.save(PATH)
print("I19 navigation legend added to Contents.")
