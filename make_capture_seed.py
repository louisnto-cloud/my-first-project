#!/usr/bin/env python3
"""
Generate capture_seed.csv — a head-start worksheet for the live-capture pass.

Reads the newest Amazon_ca_*.xlsx and emits ONE row per product LINE in the exact
capture schema (capture_template.csv), pre-filled with everything already known
(brand, category, size, flavours, nutrition, sweetener, claims). The COMMERCIAL
columns (price, S&S, coupon, promo, rating, #ratings, badge, BSR) and the
asin/url are left BLANK for you to fill from the live Amazon.ca listing.

Why this helps: you don't re-type stable data or risk transcribing nutrition
wrong — you only capture what genuinely needs the live page. `ingest_captures.py`
skips any row whose asin is blank, so a half-filled seed never pollutes the
workbook; rows light up only once you add a real ASIN + price.

Usage:  python3 make_capture_seed.py [workbook.xlsx] [out.csv]
"""
import sys, csv, glob
from openpyxl import load_workbook

HEADER=["asin","url","brand","product_title","category","format_size","unit_ml",
 "pack_count","price_cad","ss_offered","ss_pct","coupon_value_cad","promo_text",
 "flavours","caffeine_mg","sodium_mg","potassium_mg","sugar_g","calories",
 "sweetener","claims","star_rating","num_ratings","badge","bsr"]
MASTERS_PREFIX="M"
def pre(h): return str(h).split(".")[0].strip() if h else ""

def main():
    args=sys.argv[1:]
    wbs=[a for a in args if a.endswith(".xlsx")]; csvs=[a for a in args if a.endswith(".csv")]
    wbpath=wbs[0] if wbs else sorted(glob.glob("Amazon_ca_*.xlsx"))[-1]
    out=csvs[0] if csvs else "capture_seed.csv"
    wb=load_workbook(wbpath, read_only=True)
    lines={}  # base sku -> aggregated record
    order=[]
    for sn in wb.sheetnames:
        if not (sn.startswith(MASTERS_PREFIX) and " - " in sn): continue
        ws=wb[sn]; rows=ws.iter_rows(values_only=True)
        hdr=next(rows); idx={pre(h):i for i,h in enumerate(hdr)}
        def g(row,p):
            i=idx.get(p); return row[i] if i is not None and i<len(row) else None
        for row in rows:
            brand=g(row,"1")
            if brand in (None,""): continue
            full=str(g(row,"2") or "")           # "Line | Flavour | pack"
            base=full.split(" | ")[0].strip()
            flav=full.split(" | ")[1].strip() if " | " in full else ""
            key=(brand,base)
            cat=str(g(row,"3") or ""); catnum=cat.split(")")[0].strip() if ")" in cat else ""
            if key not in lines:
                lines[key]={"asin":"","url":"","brand":brand,"product_title":base,
                 "category":catnum,"format_size":g(row,"4") or "","unit_ml":g(row,"H1") or "",
                 "pack_count":"","price_cad":"","ss_offered":"","ss_pct":"",
                 "coupon_value_cad":"","promo_text":"","flavours":[],
                 "caffeine_mg":g(row,"16"),"sodium_mg":g(row,"17"),"potassium_mg":g(row,"18"),
                 "sugar_g":g(row,"19"),"calories":g(row,"20"),"sweetener":g(row,"21") or "",
                 "claims":g(row,"22") or "","star_rating":"","num_ratings":"","badge":"","bsr":""}
                order.append(key)
            if flav and flav not in lines[key]["flavours"]:
                lines[key]["flavours"].append(flav)
    with open(out,"w",newline="",encoding="utf-8") as fh:
        w=csv.DictWriter(fh,fieldnames=HEADER); w.writeheader()
        for key in order:
            rec=dict(lines[key]); rec["flavours"]=", ".join(rec["flavours"])
            for k in ("caffeine_mg","sodium_mg","potassium_mg","sugar_g","calories","unit_ml"):
                if rec[k] is None: rec[k]=""
            w.writerow(rec)
    print(f"Wrote {len(order)} line-level seed rows to {out} (from {wbpath})")
    print("Fill asin, url, price_cad and the other commercial columns from the live")
    print("Amazon.ca listing, then run: python3 ingest_captures.py")

if __name__=="__main__":
    main()
