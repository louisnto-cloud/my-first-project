#!/usr/bin/env python3
"""Deep-research integration: adds 'Electrolyte BFY Deep-Dive' and 'Canada Regulatory v2'
tabs (Canada-first sparkling/electrolyte/BFY research), wired into Contents."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; TEAL="2E7D8A"; STEEL="3E5C76"; LIGHT="EAF1F4"; LIGHT2="F4F8FA"
AMBER="FFF2CC"; AMBERHEAD="B45309"; GREEN="E2EFDA"; GREENHEAD="1E7D32"; GOLD="C9A227"; WHITE="FFFFFF"; GREY="595959"; LINKBLUE="1155CC"
def F(**k): return Font(name="Calibri",**k)
HEAD=F(bold=True,color=WHITE,size=10); TITLE=F(bold=True,color=WHITE,size=18); BODY=F(size=10); BODYB=F(size=10,bold=True); SMALL=F(size=9,italic=True,color=GREY)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top"); CTR=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)
def fill(c): return PatternFill("solid",fgColor=c)
def title(ws,text,n,sub=None):
    ws.sheet_view.showGridLines=False
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    c=ws.cell(row=1,column=1,value=text); c.font=TITLE; c.fill=fill(NAVY); c.alignment=CTR; ws.row_dimensions[1].height=30
    if sub:
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=n)
        s=ws.cell(row=2,column=1,value=sub); s.font=F(bold=True,color=WHITE,size=10); s.fill=fill(TEAL); s.alignment=CTR; ws.row_dimensions[2].height=20
        return 4
    return 3
def banner(ws,row,text,n,color=STEEL):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=n)
    c=ws.cell(row=row,column=1,value=text); c.font=F(bold=True,color=WHITE,size=10); c.fill=fill(color); c.alignment=Alignment(horizontal="left",vertical="center",indent=1); return row+1
def table(ws,start,headers,rows,widths=None,hcolor=NAVY,rowh=None):
    n=len(headers)
    for j,h in enumerate(headers,1):
        c=ws.cell(row=start,column=j,value=h); c.font=HEAD; c.fill=fill(hcolor); c.alignment=CTR; c.border=BORDER
    r=start+1
    for i,row in enumerate(rows):
        for j,val in enumerate(row,1):
            c=ws.cell(row=r,column=j,value=val); c.font=BODY; c.alignment=WRAP; c.border=BORDER
            if i%2: c.fill=fill(LIGHT)
        if rowh: ws.row_dimensions[r].height=rowh
        r+=1
    if widths:
        for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
    return r
def callout(ws,row,text,n,fillc=GREEN,fontc=GREENHEAD,h=44):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=n)
    c=ws.cell(row=row,column=1,value=text); c.font=F(bold=True,color=fontc,size=10); c.fill=fill(fillc); c.alignment=WRAP; c.border=BORDER; ws.row_dimensions[row].height=h; return row+1
def navlink(ws):
    c=ws.cell(row=1,column=14,value="↩ Contents"); c.hyperlink="#'01 Contents'!A1"; c.font=F(bold=True,color=LINKBLUE,underline="single",size=8); c.alignment=Alignment(horizontal="right")

# ============================================================ ELECTROLYTE/BFY DEEP-DIVE
dd=wb.create_sheet("Electrolyte BFY Deep-Dive")
nr=title(dd,"Electrolyte / Sparkling / Better-For-You — Canada-First Deep-Dive",6,
   "Deep-research pass (2026-06-26): MÜV's real competitive lane · ✅HARD ◐DIRECTIONAL ⚠FLAG")
nr=callout(dd,nr,("TWO FINDINGS THAT RESHAPE THE READ: (1) MÜV is a ready-to-drink SPARKLING CAN (verified: ingredient list begins “Carbonated water”, "
 "SKU 4338, ~$14.99), not a powder. (2) Canada is moving ALL sport-electrolyte products (RTD, powder, effervescent) from NHP to SUPPLEMENTED FOODS by "
 "Dec 31, 2027. So MÜV competes in the daily-wellness sparkling lane and is food-regulated; win on format + Canadian-made, not on collagen-beauty claims."),6,fillc=AMBER,fontc=AMBERHEAD,h=60)

nr=banner(dd,nr,"1 · MARKET SIZING (Canada-first)",6)
nr=table(dd,nr,["Metric","Value","Geo","Conf.","Source"],[
 ["Powdered hydration mixes","$1.5B (2024), +20% YoY; 4 straight yrs double-digit; units +21.4%","US","✅","Circana via CNN"],
 ["Gatorade powder enhancers","+200% over 4 years","US","✅","Circana via CNN"],
 ["Electrolyte powder market","~$8.7B (2024), ~8.8% CAGR to 2030; powder fastest-growing format","Global","◐","Grand View; Mordor"],
 ["NA electrolyte share","38.2% of global hydration drinks (2024)","N. America","◐","Grand View"],
 ["Canada powdered hydration","~US$404M → $738.5M (10.7% CAGR)","Canada","⚠◐","Grand View (3rd-party est.)"],
 ["Canada US-proxy cross-check","~$150M (10% of US $1.5B) — lower than Grand View","Canada","⚠","Derived"],
 ["Liquid I.V.","#1 US powdered hydration; 2024 intl expansion incl. Canada","US/Canada","✅","CNN; Unilever"],
],widths=[26,40,12,8,24],rowh=30)

nr=banner(dd,nr,"2 · TRENDS",6)
nr=table(dd,nr,["Trend","Read","Conf."],[
 ["Electrolyte/hydration boom","#WaterTok “flavor your water” is the engine behind +20% powder growth; sparkling is a demo-able hero for short-form video","✅/◐"],
 ["Sugar-free + clean label","Now table stakes, not a differentiator (stevia, magnesium bisglycinate)","✅"],
 ["Sodium split","High-sodium “serious” (LMNT 1,000mg) vs moderate daily-wellness (Liquid I.V. 500mg, Nuun 300mg)","✅"],
 ["Functional stacking","Differentiation lives in electrolytes + magnesium + prebiotic fibre (+ collagen on NHP line)","◐"],
],widths=[24,72,10],rowh=34)

nr=banner(dd,nr,"3 · COMPETITOR FORMULATION & PRICE LADDER",6)
nr=table(dd,nr,["Brand","Sodium/serv","Format","Price/serv","Lane"],[
 ["LMNT","1,000 mg","Powder stick (+Sparkling can)","~$1.50 ($1.30 sub)","High-sodium performance/keto"],
 ["Liquid I.V.","500 mg","Powder stick","~$1.56 (~$1.00 Costco)","Daily wellness"],
 ["Nuun","300 mg","Effervescent tablet","~$0.75","Everyday / endurance"],
 ["Prime","low","RTD + powder","hype-priced","Youth / viral"],
 ["BioSteel","moderate","Powder / RTD","mid","Athletic / “clean sport”"],
 ["Organika electrolyte sachets","—","3.5g powder sachets","~$29.99/30 (Costco)","Daily, Canadian-made"],
 ["Organika MÜV (this product)","TBD ⚠","Sparkling RTD CAN (SKU 4338)","$14.99 / pack","Daily-wellness sparkling"],
],widths=[24,12,24,18,28],rowh=26)
nr=callout(dd,nr,"WHITE SPACE ◐: a flavored SPARKLING electrolyte is genuinely rare (LMNT Sparkling can; Nuun tablets fizz). MÜV's carbonation is a real sensory differentiator that aids flavour release → more consumption, no hydration penalty (Beverage Hydration Index).",6,h=40)

nr=banner(dd,nr,"4 · GO-TO-MARKET & THE BIOSTEEL CAUTIONARY TALE",6)
nr=table(dd,nr,["Brand / case","Canada entry","Lesson"],[
 ["Liquid I.V.","Costco Canada national launch (Jul 2023), tight 2-SKU; festival + Costco roadshow sampling","Costco = the proven “one big door” beachhead; sampling drives trial"],
 ["LMNT","Amazon.ca + DTC; podcast sponsorships + “free sample, just pay shipping”","Cheap, high-ROI trial funnel; borrow it (not the 1,000mg sodium)"],
 ["BioSteel (Canadian)","Broad omnichannel on NHL/McDavid/Mahomes halo","⚠ Canopy bought 72% for $50.7M (2019); burned >$3 COGS per $1 rev ($24M rev vs $90M COGS); CCAA bankruptcy Sept 2023; sold ~$30.4M to Crosby/Coachwood; mfg in-housed Windsor 2025; lost NHL to Coke's BodyArmor. MORAL: athlete halo + broad distribution on a heavy cost base = bankruptcy. Velocity & unit economics FIRST."],
 ["Powder vs RTD economics","Powder ships ~60–70% cheaper, shelf-stable","MÜV-the-CAN forgoes that edge → margin case leans on Organika's BC manufacturing"],
],widths=[20,34,54],rowh=52)

nr=banner(dd,nr,"5 · RECOMMENDATIONS FOR MÜV (Canada)",6)
nr=table(dd,nr,["#","Recommendation"],[
 ["1","Beachhead = Costco Canada (hero, roadshow sampling) + Amazon.ca / Well.ca / DTC — channels Organika already holds. Prove velocity before national grocery."],
 ["2","Position “daily-wellness sparkling hydration, made in Canada” — between Liquid I.V. (flat, foreign-owned) and LMNT (high-sodium performance). Avoid athletic combat with the reviving BioSteel."],
 ["3","Design the label to the Supplemented-Food framework NOW (SFFt, front-of-pack symbol, cautions if triggered). Lead with hydration/electrolyte + prebiotic-fibre function."],
 ["4","Keep collagen / beauty messaging on the NHP collagen-powder line — NOT on the MÜV can (food-coded, claims restricted)."],
 ["5","Marketing: borrow LMNT's sample-pack funnel + #WaterTok creators (cheaper than US podcasts); fizz as the demo hero. Lean on BC manufacturing for margin."],
],widths=[6,110],rowh=40)
nr=callout(dd,nr,"SOURCES & CONFIDENCE: Circana via CNN (category); Grand View/Mordor (powder market, ◐); organika.com (MÜV SKU 4338, ingredient list); Globe & Mail/CBC/PRNewswire/Coachwood (BioSteel); strategyonline.ca/Unilever (Liquid I.V. Canada); drinklmnt.com/hubermanlab.com (LMNT). Revenue & market-size figures are estimates/projections (⚠). Several canada.ca/retailer pages block automated fetch (403) — confirm load-bearing items on the live page.",6,fillc=LIGHT,fontc=GREY,h=46)
navlink(dd); dd.sheet_properties.tabColor=TEAL
dd.page_setup.orientation="landscape"; dd.page_setup.fitToWidth=1; dd.page_setup.fitToHeight=0; dd.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)

# ============================================================ CANADA REGULATORY v2
rv=wb.create_sheet("Canada Regulatory v2")
nr=title(rv,"Canada Regulatory v2 — Sport-Electrolyte Reclassification (2026)",3,
   "Supersedes the powder/NHP framing on Tab 19. ✅HARD ◐DIRECTIONAL ⚠FLAG")
nr=callout(rv,nr,("HEADLINE: Health Canada Public Notice (Apr 2026) reclassifies ALL sport-electrolyte products — RTD beverages, powders/concentrates AND effervescent "
 "tablets — from Natural Health Products (NHP/NPN) to FOODS / Supplemented Foods. Existing NPN holders transition by Dec 31, 2027; NEW products must comply "
 "with the Food & Drug Regulations immediately. Oral Rehydration Solutions (clinical) stay NHP."),3,fillc=AMBER,fontc=AMBERHEAD,h=58)

nr=banner(rv,nr,"FORMAT IS NOT THE DECIDER (the key nuance)",3)
nr=table(rv,nr,["Point","Detail","Conf."],[
 ["Format ≠ classification","Per Health Canada's food–NHP interface guidance, even powders & effervescent tablets for reconstitution CAN be foods. Classification weighs representation, composition, format, perception & history — together.","✅ (primary)"],
 ["What pushes toward FOOD","Beverage representation; scoop/serving + a Nutrition/Supplemented-Food Facts table; consumed freely like ordinary food/drink.","✅"],
 ["What pushes toward NHP","Dosage forms (capsules, single-dose sticks, measured drops) + regimented dosing + defined daily limits + conditions of use + therapeutic/structure-function claims.","✅"],
 ["So the real driver is CLAIMS","Electrolyte/hydration framing = food. Collagen skin/joint/hair/beauty claims = NHP-only. The claim, not the powder-vs-RTD format, decides the lane.","✅"],
],widths=[22,72,12],rowh=44)

nr=banner(rv,nr,"CLAIMS FORK",3)
nr=table(rv,nr,["If the product is…","Framework","Allowed claims"],[
 ["Plain electrolyte (can OR powder)","Supplemented Food","hydration, “replenish electrolytes”, source-of-magnesium/potassium, sugar-free, nutrient-content & acceptable function claims"],
 ["Electrolyte + collagen w/ beauty/joint claims","NHP (NPN)","“supports skin/hair/nails/joints”, collagen structure-function (dose-format powder is the defensible vehicle)"],
 ["Oral rehydration (clinical)","NHP (NPN)","therapeutic rehydration"],
],widths=[30,18,60],rowh=40)

nr=banner(rv,nr,"SUPPLEMENTED FOODS FRAMEWORK & LABELLING",3)
nr=table(rv,nr,["Item","Requirement","Conf."],[
 ["Framework","In force July 21, 2022. Added vitamins/minerals/amino acids per List of Permitted Supplemental Ingredients & Categories.","✅"],
 ["Supplemented Food Facts table (SFFt)","Modified Nutrition Facts table with a “Supplemented with” line (name + amount of each supplemental ingredient).","✅"],
 ["Caution Identifier (SFCI)","Black-&-white identifier on the principal display panel WHEN cautionary statements are required (children, pregnancy, etc.).","✅"],
 ["Caffeine","180 mg cap is PER-SERVING for the caffeinated-beverage category (400 mg/day is the reference). Moot for caffeine-free MÜV.","✅ (precision)"],
 ["Collagen caution","If hydrolyzed collagen is present: “CAUTION, DO NOT USE AS SOLE SOURCE OF NUTRITION” on the PDP, same type size as common name (FDR).","✅"],
 ["Bilingual EN/FR","Mandatory on all required info incl. the facts table.","✅"],
 ["Quebec Bill 96 / OQLF","From Jun 1, 2025 generic/descriptive terms in a trademark need French; “Sparkling Electrolytes” likely needs a French descriptor. Sell-through grace to Jun 1, 2027.","◐ (trademark call)"],
],widths=[20,74,12],rowh=38)

nr=banner(rv,nr,"PRACTICAL TAKEAWAY FOR MÜV",3)
nr=table(rv,nr,["Decision","Guidance"],[
 ["Classification","MÜV (RTD electrolyte can, caffeine-free) = a Supplemented Food. No NPN; SFFt + cautions/SFCI as applicable."],
 ["Claims","Hydration/electrolyte/“replenish” + prebiotic-fibre function are safe. Do NOT put collagen beauty/joint claims on the can — keep them on the NHP collagen powder."],
 ["Format choice","For a plain electrolyte line, powder vs RTD now sit in the SAME (food) lane — choose on cost/shelf-life/experience, not regulation. Only a collagen-claims strategy argues for a dose-format NHP powder."],
 ["Timeline","Design to the food/Supplemented-Food framework from day one to avoid a 2027 relabel; new products must comply with FDR now."],
],widths=[20,96],rowh=44)
nr=callout(rv,nr,"OPEN ITEMS ⚠: confirm exact transition date (Dec 31 2027 vs Jan 1 2028) on the canada.ca notice; MÜV's current licensing status (food vs any legacy NPN); verbatim collagen monograph wording; Bill 96 trademark analysis for “MÜV Sparkling Electrolytes”. canada.ca/LNHPD pages block automated fetch (403) — verify on the live primary page before a compliance filing.",3,fillc=LIGHT,fontc=GREY,h=46)
navlink(rv); rv.sheet_properties.tabColor=GREENHEAD
rv.page_setup.orientation="landscape"; rv.page_setup.fitToWidth=1; rv.page_setup.fitToHeight=0; rv.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)

# ============================================================ CONTENTS + ORDER
toc=wb["01 Contents"]
for tab,desc in [("Electrolyte BFY Deep-Dive","Canada-first electrolyte/sparkling/BFY research; MÜV's real lane (deep-research)"),
                 ("Canada Regulatory v2","Sport-electrolyte → Supplemented Foods reclassification; claims fork (supersedes Tab 19)")]:
    r=toc.max_row+1
    a=toc.cell(row=r,column=1,value=tab); a.hyperlink=f"#'{tab}'!A1"; a.font=F(bold=True,color=LINKBLUE,underline="single",size=10); a.fill=fill(AMBER); a.border=BORDER; a.alignment=LEFT
    b=toc.cell(row=r,column=2,value=desc); b.font=BODY; b.border=BORDER; b.alignment=WRAP; toc.row_dimensions[r].height=18

# place both after '19 Canada Regulatory' (capture refs BEFORE removing)
s1=wb["Canada Regulatory v2"]; s2=wb["Electrolyte BFY Deep-Dive"]
wb._sheets.remove(s1); wb._sheets.remove(s2)
idx=wb.sheetnames.index("19 Canada Regulatory")+1
wb._sheets.insert(idx,s1); wb._sheets.insert(idx+1,s2)

wb.properties.description="32-tab dossier: interactive scenario P&L, dashboard, Go/No-Go, battlecards, verification log, and a Canada-first electrolyte/BFY deep-dive. MÜV verified as an RTD sparkling can; Canada electrolyte→Supplemented-Food reclassification reflected."
wb.save(PATH)
print("deep-research tabs added. sheets:",len(wb.sheetnames))
for s in wb.sheetnames: print("  -",s)
