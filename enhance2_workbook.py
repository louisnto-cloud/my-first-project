#!/usr/bin/env python3
"""Second 10x pass: scenario-driven full P&L + break-even + sensitivity heatmap,
Go/No-Go decision tab, competitor battlecards, rebuilt Contents, re-pointed dashboard."""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation

PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)

NAVY="14304F"; TEAL="2E7D8A"; STEEL="3E5C76"; LIGHT="EAF1F4"; LIGHT2="F4F8FA"
AMBER="FFF2CC"; AMBERHEAD="B45309"; GREEN="E2EFDA"; GREENHEAD="1E7D32"
GOLD="C9A227"; WHITE="FFFFFF"; GREY="595959"; RED="F8D7DA"; REDH="9C2A2A"
def F(**k): return Font(name="Calibri", **k)
HEAD=F(bold=True,color=WHITE,size=10); TITLE=F(bold=True,color=WHITE,size=18)
BODY=F(size=10); BODYB=F(size=10,bold=True); SMALL=F(size=9,italic=True,color=GREY)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top"); CTR=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)
def fill(c): return PatternFill("solid",fgColor=c)
YEL=fill("FFF7CC")

def title(ws,text,n,sub=None):
    ws.sheet_view.showGridLines=False
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    c=ws.cell(row=1,column=1,value=text); c.font=TITLE; c.fill=fill(NAVY); c.alignment=CTR
    ws.row_dimensions[1].height=30
    if sub:
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=n)
        s=ws.cell(row=2,column=1,value=sub); s.font=F(bold=True,color=WHITE,size=10); s.fill=fill(TEAL); s.alignment=CTR
        ws.row_dimensions[2].height=20
        return 4
    return 3
def banner(ws,row,text,n,color=STEEL):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=n)
    c=ws.cell(row=row,column=1,value=text); c.font=F(bold=True,color=WHITE,size=10); c.fill=fill(color)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1); return row+1
def callout(ws,row,text,n,fillc=GREEN,fontc=GREENHEAD,h=44):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=n)
    c=ws.cell(row=row,column=1,value=text); c.font=F(bold=True,color=fontc,size=10)
    c.fill=fill(fillc); c.alignment=WRAP; c.border=BORDER; ws.row_dimensions[row].height=h; return row+1

# remove old Dashboard + Financial Model to rebuild coherently
for nm in ["Dashboard","Financial Model"]:
    if nm in wb.sheetnames: del wb[nm]

# =====================================================================
# FINANCIAL MODEL v2 — scenario-driven full P&L
# =====================================================================
fm=wb.create_sheet("Financial Model")
title(fm,"MUV Financial Model — Scenario P&L, Break-even & Sensitivity",8,
   "Pick a scenario in the dropdown (C4). Yellow = editable inputs; white = live formulas. Illustrative, not a forecast.")
fm.column_dimensions["A"].width=3
fm.column_dimensions["B"].width=32
for col,w in zip("CDEFGH",[14,14,14,14,4,30]): fm.column_dimensions[col].width=w

def lab(r,t,b=False,col=2):
    c=fm.cell(row=r,column=col,value=t); c.font=BODYB if b else BODY; c.alignment=LEFT; return c
def inp(r,col,v,fmt=None,b=True):
    c=fm.cell(row=r,column=col,value=v); c.font=BODYB if b else BODY; c.fill=YEL; c.border=BORDER; c.alignment=CTR
    if fmt:c.number_format=fmt
    return c
def cal(r,col,f,fmt=None,b=False,bg=LIGHT2):
    c=fm.cell(row=r,column=col,value=f); c.font=BODYB if b else BODY; c.border=BORDER; c.alignment=CTR; c.fill=fill(bg)
    if fmt:c.number_format=fmt
    return c

# Scenario selector
lab(4,"Scenario (pick):",True)
sel=fm.cell(row=4,column=3,value="Base"); sel.font=F(bold=True,color=NAVY,size=11); sel.fill=YEL; sel.border=BORDER; sel.alignment=CTR
dv=DataValidation(type="list",formula1='"Conservative,Base,Aggressive"',allow_blank=False)
fm.add_data_validation(dv); dv.add(sel)
lab(5,"Scenario index:")
cal(5,3,'=MATCH(C4,{"Conservative","Base","Aggressive"},0)','0')
IDX="$C$5"
lab(4,"← change this dropdown to flip the whole model",col=6); fm.cell(row=4,column=6).font=SMALL

# Assumptions matrix
hr=7
for j,t in enumerate(["Assumption","Conservative","Base","Aggressive","ACTIVE"]):
    c=fm.cell(row=hr,column=2+j,value=t); c.font=HEAD; c.fill=fill(NAVY); c.alignment=CTR; c.border=BORDER
fm.cell(row=hr,column=7).value=None
# rows: name, cons, base, aggr, fmt
amx=[
 ("List price / can (CAD)",3.29,3.49,3.79,'"$"#,##0.00'),
 ("COGS / can (CAD)",1.20,1.05,0.95,'"$"#,##0.00'),
 ("Retailer margin %",0.38,0.35,0.33,'0%'),
 ("Distributor margin %",0.18,0.15,0.12,'0%'),
 ("Trade spend % of net",0.25,0.20,0.15,'0%'),
 ("Cans per case",12,12,12,'0'),
 ("Doors — Year 1",300,500,900,'#,##0'),
 ("Doors — Year 2",1200,2000,3500,'#,##0'),
 ("Doors — Year 3",3500,6000,11000,'#,##0'),
 ("Velocity Y1 (u/store/wk)",2.0,3.0,4.5,'0.0'),
 ("Velocity Y2 (u/store/wk)",3.0,4.0,6.0,'0.0'),
 ("Velocity Y3 (u/store/wk)",3.5,5.0,7.5,'0.0'),
 ("Marketing / year (CAD)",150000,300000,600000,'"$"#,##0'),
 ("Fixed overhead / year (CAD)",250000,350000,500000,'"$"#,##0'),
 ("Slotting (Year 1 one-time)",50000,100000,200000,'"$"#,##0'),
]
A={}
r=hr+1
for name,cv,bv,av,fmt in amx:
    lab(r,name)
    inp(r,3,cv,fmt); inp(r,4,bv,fmt); inp(r,5,av,fmt)
    cal(r,6,f"=CHOOSE({IDX},C{r},D{r},E{r})",fmt,True,GREEN)
    A[name]=r; fm.row_dimensions[r].height=16; r+=1
# active refs
P=f"$F${A['List price / can (CAD)']}"; CO=f"$F${A['COGS / can (CAD)']}"
RM=f"$F${A['Retailer margin %']}"; DM=f"$F${A['Distributor margin %']}"; TR=f"$F${A['Trade spend % of net']}"
CASE=f"$F${A['Cans per case']}"
MKT=f"$F${A['Marketing / year (CAD)']}"; FIX=f"$F${A['Fixed overhead / year (CAD)']}"; SLOT=f"$F${A['Slotting (Year 1 one-time)']}"

r+=1
r=banner(fm,r,"PER-CAN ECONOMICS (active scenario)",8)
e=r
lab(e,"Net revenue / can",True);  NET=cal(e,3,f"={P}*(1-{RM})*(1-{DM})",'"$"#,##0.00',True); NETc=f"$C${e}"
lab(e+1,"Gross profit / can",True); cal(e+1,3,f"=C{e}-{CO}",'"$"#,##0.00',True); GPc=f"$C${e+1}"
lab(e+2,"Gross margin %",True); cal(e+2,3,f"=C{e+1}/C{e}",'0.0%',True)
lab(e+3,"Trade spend / can"); cal(e+3,3,f"=C{e}*{TR}",'"$"#,##0.00'); TRc=f"$C${e+3}"
lab(e+4,"Contribution / can",True); cal(e+4,3,f"=C{e+1}-C{e+3}",'"$"#,##0.00',True); CONc=f"$C${e+4}"
r=e+6

r=banner(fm,r,"3-YEAR P&L  (CAD)",8)
ph=r
for j,t in enumerate(["Line","Year 1","Year 2","Year 3"]):
    c=fm.cell(row=ph,column=2+j,value=t); c.font=HEAD; c.fill=fill(STEEL); c.alignment=CTR; c.border=BORDER
P3={}
def prow(rr,name,builder,fmt='"$"#,##0',bold=False,bg=LIGHT2):
    lab(rr,name,bold)
    for j,col in enumerate([3,4,5]):
        cal(rr,col,builder(get_column_letter(col),j),fmt,bold,bg)
    P3[name]=rr
r=ph+1
doors=[A['Doors — Year 1'],A['Doors — Year 2'],A['Doors — Year 3']]
vel=[A['Velocity Y1 (u/store/wk)'],A['Velocity Y2 (u/store/wk)'],A['Velocity Y3 (u/store/wk)']]
prow(r,"Distribution (doors)",lambda L,j:f"=$F${doors[j]}",'#,##0'); r+=1
prow(r,"Velocity (u/store/wk)",lambda L,j:f"=$F${vel[j]}",'0.0'); r+=1
UNITS=r; prow(r,"Annual units",lambda L,j:f"={L}{P3['Distribution (doors)']}*{L}{P3['Velocity (u/store/wk)']}*52",'#,##0'); r+=1
NETr=r; prow(r,"Net revenue",lambda L,j:f"={L}{UNITS}*{NETc}",bold=True); r+=1
prow(r,"COGS",lambda L,j:f"=-{L}{UNITS}*{CO}"); r+=1
GPr=r; prow(r,"Gross profit",lambda L,j:f"={L}{NETr}+{L}{r-1}",bold=True); r+=1
prow(r,"Trade spend",lambda L,j:f"=-{L}{UNITS}*{TRc}"); r+=1
prow(r,"Marketing",lambda L,j:f"=-{MKT}"); r+=1
prow(r,"Fixed overhead",lambda L,j:f"=-{FIX}"); r+=1
prow(r,"Slotting (one-time)",lambda L,j:(f"=-{SLOT}" if j==0 else "=0")); r+=1
EBr=r; prow(r,"EBITDA",lambda L,j:f"={L}{GPr}+{L}{r-3}+{L}{r-2}+{L}{r-1}+{L}{r-4}",bold=True,bg=GREEN); r+=1
prow(r,"EBITDA margin %",lambda L,j:f"={L}{EBr}/{L}{NETr}",'0.0%'); r+=1
r+=1

r=banner(fm,r,"BREAK-EVEN (Year-1 cost load ÷ contribution per can)",8)
b=r
lab(b,"Year-1 fixed + marketing + slotting"); cal(b,3,f"={MKT}+{FIX}+{SLOT}",'"$"#,##0',True)
lab(b+1,"Break-even units (cans)"); cal(b+1,3,f"=C{b}/{CONc}",'#,##0',True)
lab(b+2,"Break-even cases"); cal(b+2,3,f"=C{b+1}/{CASE}",'#,##0')
lab(b+3,"Year-1 actual units"); cal(b+3,3,f"=C{UNITS}",'#,##0')
lab(b+4,"Year-1 surplus / (gap) to break-even",True); cal(b+4,3,f"=C{b+3}-C{b+1}",'#,##0',True,AMBER)
r=b+6

r=banner(fm,r,"SENSITIVITY — Contribution per can (CAD):  price (down) × COGS (across), active margins",8)
sh=r
prices=[3.09,3.29,3.49,3.69,3.99]; cogs=[0.85,0.95,1.05,1.20,1.35]
# header row: COGS across cols D..H (col 4..8); corner label col C
fm.cell(row=sh,column=3,value="P＼COGS").font=SMALL
for j,cg in enumerate(cogs):
    c=fm.cell(row=sh,column=4+j,value=cg); c.font=BODYB; c.number_format='"$"#,##0.00'; c.fill=fill(STEEL); c.font=F(bold=True,color=WHITE,size=9); c.alignment=CTR; c.border=BORDER
grid_first=sh+1
for i,pr in enumerate(prices):
    rr=sh+1+i
    pc=fm.cell(row=rr,column=3,value=pr); pc.font=F(bold=True,color=WHITE,size=9); pc.fill=fill(STEEL); pc.number_format='"$"#,##0.00'; pc.alignment=CTR; pc.border=BORDER
    for j,cg in enumerate(cogs):
        col=4+j; L=get_column_letter(col)
        # net = price*(1-RM)*(1-DM); contribution = net*(1-TR) - cogs
        f=f"=($C{rr}*(1-{RM})*(1-{DM}))*(1-{TR})-{L}${sh}"
        c=fm.cell(row=rr,column=col,value=f); c.number_format='"$"#,##0.00'; c.alignment=CTR; c.border=BORDER; c.font=BODY
grid_last=sh+len(prices)
fm.conditional_formatting.add(f"D{grid_first}:H{grid_last}",ColorScaleRule(
    start_type='min',start_color='F8696B',mid_type='percentile',mid_value=50,mid_color='FFEB84',end_type='max',end_color='63BE7B'))
r=grid_last+2
callout(fm,r,("Scenario logic: the ACTIVE column = CHOOSE(scenario, Conservative, Base, Aggressive); every downstream cell reads ACTIVE, so the "
 "dropdown re-drives the entire P&L, break-even and (margins) sensitivity. EBITDA excludes freight/FX and financing — add as further lines for a "
 "full forecast. Replace all yellow cells with real co-packer, distributor and buyer numbers before relying on outputs."),8,fillc=AMBER,fontc=AMBERHEAD,h=58)

# conditional data bars on P&L net & ebitda
fm.conditional_formatting.add(f"C{NETr}:E{NETr}",DataBarRule(start_type='min',end_type='max',color="3E5C76"))
fm.conditional_formatting.add(f"C{EBr}:E{EBr}",DataBarRule(start_type='min',end_type='max',color="1E7D32"))
fm.freeze_panes="C5"

# =====================================================================
# DASHBOARD v2 — re-pointed + EBITDA chart
# =====================================================================
dash=wb.create_sheet("Dashboard")
title(dash,"Executive Dashboard",12,"Category sizing · leader trajectories · share · competitive scores · MUV scenario P&L")
for col in "ABCDEFGHIJKL": dash.column_dimensions[col].width=11
kpis=[("Modern soda US '24","$1.8B","+83% YoY"),("Poppi → PepsiCo","$1.95B","2025"),
      ("OLIPOP valuation","$1.85B","2025"),("Collagen-indie survival","Low","graveyard"),
      ("Canada beachhead","Costco+Amazon","proven")]
r0=4; col=1
for t,v,s in kpis:
    dash.merge_cells(start_row=r0,start_column=col,end_row=r0,end_column=col+1)
    dash.cell(row=r0,column=col,value=t).font=F(bold=True,color=WHITE,size=8); dash.cell(row=r0,column=col).fill=fill(STEEL); dash.cell(row=r0,column=col).alignment=CTR
    dash.merge_cells(start_row=r0+1,start_column=col,end_row=r0+1,end_column=col+1)
    dash.cell(row=r0+1,column=col,value=v).font=F(bold=True,color=NAVY,size=14); dash.cell(row=r0+1,column=col).fill=fill(LIGHT); dash.cell(row=r0+1,column=col).alignment=CTR
    dash.merge_cells(start_row=r0+2,start_column=col,end_row=r0+2,end_column=col+1)
    dash.cell(row=r0+2,column=col,value=s).font=F(italic=True,color=GREY,size=8); dash.cell(row=r0+2,column=col).fill=fill(LIGHT); dash.cell(row=r0+2,column=col).alignment=CTR
    for rr in (r0,r0+1,r0+2):
        for cc in (col,col+1): dash.cell(row=rr,column=cc).border=BORDER
    col+=2
for rr in (r0,r0+1,r0+2): dash.row_dimensions[rr].height=18

DR=70
def put(sr,sc,ttl,hdrs,rows):
    dash.cell(row=sr,column=sc,value=ttl).font=BODYB
    for j,h in enumerate(hdrs): dash.cell(row=sr+1,column=sc+j,value=h).font=SMALL
    for i,row in enumerate(rows):
        for j,v in enumerate(row): dash.cell(row=sr+2+i,column=sc+j,value=v)
cat=[("Modern soda",1.8),("Hydration powder",1.5),("Electrolyte powder",2.8),("Sparkling water",6.9),("Collagen bev US",0.05),("BFW global",3.24)]
put(DR,1,"Category $B",["Cat","$B"],cat)
yrs=[2020,2021,2022,2023,2024]; poppi=[13,26,65,100,500]; olipop=[20,40,73,200,400]
put(DR,5,"Revenue $M",["Yr","Poppi","OLIPOP"],list(zip(yrs,poppi,olipop)))
share=[("Poppi",38),("OLIPOP",32.7),("Zevia",12.1),("Jarritos",10.3),("Fever-Tree",4),("Other",2.9)]
put(DR,9,"Share %",["Brand","%"],share)
weights=[0.15,0.15,0.15,0.10,0.10,0.10,0.15,0.10]
sc={"Poppi":[5,5,2,3,3,5,5,2],"OLIPOP":[5,5,4,3,4,4,4,2],"LMNT":[4,3,3,4,3,5,4,2],"Liquid I.V.":[4,5,3,3,3,4,5,3],
    "Vital Proteins":[4,5,4,3,3,3,5,2],"Recess":[3,3,3,3,4,4,3,1],"MUV (Organika)":[3,3,3,4,3,2,2,5]}
scores=[(k,round(sum(s*w for s,w in zip(v,weights)),2)) for k,v in sc.items()]
put(DR,13,"Score",["Brand","Score"],scores)

c1=BarChart(); c1.type="col"; c1.title="Category Sizes (US, $B)"; c1.height=7; c1.width=14
c1.add_data(Reference(dash,min_col=2,min_row=DR+1,max_row=DR+1+len(cat)),titles_from_data=True)
c1.set_categories(Reference(dash,min_col=1,min_row=DR+2,max_row=DR+1+len(cat))); c1.legend=None; dash.add_chart(c1,"A8")
c2=LineChart(); c2.title="Revenue Trajectory ($M)"; c2.height=7; c2.width=14
c2.add_data(Reference(dash,min_col=6,min_row=DR+1,max_col=7,max_row=DR+1+len(yrs)),titles_from_data=True)
c2.set_categories(Reference(dash,min_col=5,min_row=DR+2,max_row=DR+1+len(yrs))); dash.add_chart(c2,"G8")
c3=PieChart(); c3.title="Modern Soda Share '24 (%)"; c3.height=7; c3.width=10
c3.add_data(Reference(dash,min_col=10,min_row=DR+1,max_row=DR+1+len(share)),titles_from_data=True)
c3.set_categories(Reference(dash,min_col=9,min_row=DR+2,max_row=DR+1+len(share)))
c3.dataLabels=DataLabelList(); c3.dataLabels.showVal=True; dash.add_chart(c3,"A23")
c4=BarChart(); c4.type="bar"; c4.title="Weighted Competitive Score"; c4.height=7; c4.width=14
c4.add_data(Reference(dash,min_col=14,min_row=DR+1,max_row=DR+1+len(scores)),titles_from_data=True)
c4.set_categories(Reference(dash,min_col=13,min_row=DR+2,max_row=DR+1+len(scores))); c4.legend=None
c4.dataLabels=DataLabelList(); c4.dataLabels.showVal=True; dash.add_chart(c4,"E23")
c5=BarChart(); c5.type="col"; c5.title="MUV Scenario P&L — Net Rev & EBITDA (CAD)"; c5.height=7; c5.width=14
c5.add_data(Reference(fm,min_col=2,min_row=NETr,max_col=5,max_row=NETr),titles_from_data=True,from_rows=True)
c5.add_data(Reference(fm,min_col=2,min_row=EBr,max_col=5,max_row=EBr),titles_from_data=True,from_rows=True)
c5.set_categories(Reference(fm,min_col=3,min_row=ph,max_col=5,max_row=ph)); dash.add_chart(c5,"J23")
for rr in range(DR,DR+12): dash.row_dimensions[rr].hidden=True
dash.merge_cells(start_row=39,start_column=1,end_row=39,end_column=12)
dash.cell(row=39,column=1,value="Illustrative; built from cited research (see ‘24 Sources’). The P&L chart is driven live by the Financial Model scenario dropdown.").font=SMALL

# =====================================================================
# GO / NO-GO DECISION
# =====================================================================
gng=wb.create_sheet("Go-NoGo Decision")
nr=title(gng,"Go / No-Go Decision Scorecard",6,"Weighted gate criteria. Edit yellow weight & score (1–5) cells; verdict auto-computes")
crit=[("Market opportunity / tailwinds","Category growth & white space support entry",0.15,4),
 ("Right to win (Organika assets)","Domestic mfg, #1 collagen halo, CA shelf relationships",0.20,4),
 ("Claim defensibility / dose","Substantiable collagen+fiber dose; avoids Poppi-style risk",0.15,3),
 ("Regulatory clarity (Canada)","Supplemented-Foods pathway resolved with counsel",0.15,3),
 ("Unit economics / margin","Contribution positive at ≤$3.49 CAD with domestic COGS",0.15,4),
 ("Competitive resilience","Defensible vs Nestlé/Vital Proteins & private label",0.10,2),
 ("Go-to-market feasibility","Costco/Amazon/natural beachhead + cheap marketing playbooks",0.10,4)]
hr=nr
for j,t in enumerate(["Criterion","What 'pass' looks like","Weight","Score 1–5","Weighted"]):
    c=gng.cell(row=hr,column=1+j,value=t); c.font=HEAD; c.fill=fill(NAVY); c.alignment=CTR; c.border=BORDER
r=hr+1; first=r
for name,desc,w,s in crit:
    gng.cell(row=r,column=1,value=name).font=BODYB; gng.cell(row=r,column=1).border=BORDER; gng.cell(row=r,column=1).alignment=LEFT
    gng.cell(row=r,column=2,value=desc).font=BODY; gng.cell(row=r,column=2).border=BORDER; gng.cell(row=r,column=2).alignment=WRAP
    iw=gng.cell(row=r,column=3,value=w); iw.fill=YEL; iw.number_format='0%'; iw.border=BORDER; iw.alignment=CTR; iw.font=BODYB
    isc=gng.cell(row=r,column=4,value=s); isc.fill=YEL; isc.border=BORDER; isc.alignment=CTR; isc.font=BODYB
    cw=gng.cell(row=r,column=5,value=f"=C{r}*D{r}"); cw.number_format='0.00'; cw.border=BORDER; cw.alignment=CTR; cw.fill=fill(LIGHT2)
    gng.row_dimensions[r].height=30; r+=1
last=r-1
gng.cell(row=r,column=1,value="TOTALS").font=BODYB
gng.cell(row=r,column=3,value=f"=SUM(C{first}:C{last})").number_format='0%'; gng.cell(row=r,column=3).font=BODYB; gng.cell(row=r,column=3).alignment=CTR; gng.cell(row=r,column=3).border=BORDER
ws_total=f"=SUMPRODUCT(C{first}:C{last},D{first}:D{last})"
tot=gng.cell(row=r,column=5,value=ws_total); tot.number_format='0.00'; tot.font=F(bold=True,size=12,color=NAVY); tot.alignment=CTR; tot.fill=fill(GREEN); tot.border=BORDER
gng.cell(row=r,column=4,value="Weighted /5 →").font=BODYB; gng.cell(row=r,column=4).alignment=Alignment(horizontal="right")
total_row=r
for col,w in zip("ABCDE",[30,40,10,10,10]): gng.column_dimensions[col].width=w
r+=2
gng.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
gng.cell(row=r,column=1,value="Auto verdict (thresholds):").font=BODYB
v=gng.cell(row=r,column=3,value=f'=IF(E{total_row}>=3.5,"GO",IF(E{total_row}>=2.8,"GO — conditional","NO-GO / rework"))')
gng.merge_cells(start_row=r,start_column=3,end_row=r,end_column=5)
v.font=F(bold=True,size=12,color=WHITE); v.fill=fill(GREENHEAD); v.alignment=CTR; v.border=BORDER
gng.conditional_formatting.add(f"E{first}:E{last}",DataBarRule(start_type='num',start_value=0,end_type='num',end_value=5,color="2E7D8A"))
r+=2
callout(gng,r,("Current illustrative score lands in “GO — conditional”: the upside (Organika's right-to-win + tailwinds + margin) is strong; the gates "
 "to clear are claim/dose defensibility, Canadian regulatory sign-off, and competitive resilience. Recalibrate every yellow cell with your own evidence."),6,h=46)

# =====================================================================
# BATTLECARDS
# =====================================================================
bc=wb.create_sheet("Battlecards")
nr=title(bc,"Competitor Battlecards",6,"One block per key competitor — for sales/exec quick reference")
cards=[
 ("POPPI (PepsiCo)","Prebiotic soda, mass-market","$1.95B PepsiCo exit (2025); ~$500M rev; ~38% modern-soda share",
  "TikTok-creator virality + 2 Super Bowls; farmers-mkt→Whole Foods→mass","2g fiber, $8.9M gut-health settlement — thin claim",
  "Out-claim on dose + Canadian-made; they can't match domestic authenticity"),
 ("OLIPOP","Prebiotic 'new soda', 9g fiber","$1.85B val (2025), independent, cash-profitable; ~$400M rev",
  "Natural-channel seeding + 'astronomical repurchase' sell-in","Premium $3/can; single category; no strategic distribution",
  "Match fiber credibility; beat on collagen+hydration stack & price"),
 ("VITAL PROTEINS Sparkling (Nestlé)","Collagen sparkling, beauty-from-within","Nestlé-owned; #1 collagen halo; $2.50/can mass price",
  "Big-brand distribution + clinical VERISOL claim","2025 meta-analysis: no proven skin benefit; new format unproven",
  "Function-forward (gut+hydration) not pure beauty; domestic cost edge; avoid their efficacy-claim exposure"),
 ("LMNT","Zero-sugar high-sodium electrolyte","$206M sales (2023), ~20% net margin, independent",
  "Podcast host-reads + free-sample funnel; DTC/subscription","Polarizing sodium; powder not RTD; niche",
  "Borrow the cheap podcast + sampling playbook for MUV's beachhead"),
 ("LIQUID I.V. (Unilever)","Hydration multiplier powder","Unilever; $1B+ brand; 80,000+ US doors; Costco engine",
  "Costco roadshow sampling + POS-data sell-in","Sugar content; undifferentiated vs new entrants",
  "Copy the Costco-Canada sampling beachhead they used to enter Canada in 2023"),
]
H=["Competitor","Positioning","Scale / outcome","Winning play","Vulnerability","How MUV beats / borrows"]
hr=nr
for j,t in enumerate(H):
    c=bc.cell(row=hr,column=1+j,value=t); c.font=HEAD; c.fill=fill(NAVY); c.alignment=CTR; c.border=BORDER
r=hr+1
for i,row in enumerate(cards):
    for j,v in enumerate(row):
        c=bc.cell(row=r,column=1+j,value=v); c.font=BODYB if j==0 else BODY; c.alignment=WRAP; c.border=BORDER
        if i%2: c.fill=fill(LIGHT)
        if j==5: c.fill=fill(GREEN)
    bc.row_dimensions[r].height=70; r+=1
for col,w in zip("ABCDEF",[20,22,28,30,28,34]): bc.column_dimensions[col].width=w
bc.freeze_panes="A"+str(hr+1)

# =====================================================================
# REBUILD CONTENTS (01) to reflect full tab set
# =====================================================================
if "01 Contents" in wb.sheetnames: del wb["01 Contents"]
toc=wb.create_sheet("01 Contents")
nr=title(toc,"Table of Contents",2)
entries=[
 ("00 Cover","Title, scope, caveats, confidence legend"),
 ("01 Contents","This page"),
 ("02 Executive Summary","Thesis, 7 findings, recommendation"),
 ("Dashboard","Visual snapshot — 5 charts + KPI strip"),
 ("Financial Model","Scenario-driven P&L, break-even, sensitivity heatmap (interactive)"),
 ("03 Methodology","Research design, source hierarchy, confidence, limitations"),
 ("04 Market Sizing","TAM/SAM/SOM, category sizes & growth (US + Canada)"),
 ("05 Trends & Macro","Sugar, GLP-1, gut health, beauty-from-within, white space"),
 ("06 Consumer","Segmentation, collagen buyer, efficacy risk, buyer criteria"),
 ("07 Comparison Matrix","8 core brands side by side (auto-filter)"),
 ("08 Extended Landscape","~30 brands across 6 sub-categories"),
 ("09 Channel & Retail","Launch channel & expansion sequence"),
 ("10 Channel Economics","Margin waterfall, distributor/trade-spend, DTC/Amazon"),
 ("11 Formulation","The winning claim, hero ingredient, claim risk"),
 ("12 Marketing","Founder story, virality engine"),
 ("13 Launch Playbooks","Phased GTM timelines"),
 ("14 Patterns & Failures","7 success patterns + 6 failure modes"),
 ("15 Outcomes & M&A","Revenue, funding, valuations, acquisition comps"),
 ("16 SWOT","Per-brand SWOT incl. MUV"),
 ("17 Scoring Model","Weighted competitive scoring (heatmap)"),
 ("18 Canada Market","Retail banners & share, entry playbooks, pricing"),
 ("19 Canada Regulatory","NHP vs Food vs Supplemented Foods; SFCI; caffeine; Bill 96"),
 ("20 GTM Recommendation","The MUV go-to-market thesis & decisions"),
 ("Go-NoGo Decision","Weighted gate scorecard with auto verdict (interactive)"),
 ("Battlecards","One-page-per-competitor quick reference"),
 ("21 Launch Plan","Phased 0–18 month plan"),
 ("22 Risk Register","Ranked risks, likelihood/impact, mitigations"),
 ("23 KPI Dashboard","Targets & benchmarks by phase"),
 ("24 Sources","Master bibliography, tagged by type"),
 ("25 Glossary","Terms & abbreviations"),
]
for j,t in enumerate(["Tab","Contents"]):
    c=toc.cell(row=nr,column=1+j,value=t); c.font=HEAD; c.fill=fill(NAVY); c.alignment=CTR; c.border=BORDER
r=nr+1
for i,(tab,desc) in enumerate(entries):
    a=toc.cell(row=r,column=1,value=tab); a.font=BODYB; a.border=BORDER; a.alignment=LEFT
    b=toc.cell(row=r,column=2,value=desc); b.font=BODY; b.border=BORDER; b.alignment=WRAP
    if i%2:
        a.fill=fill(LIGHT); b.fill=fill(LIGHT)
    if tab in ("Dashboard","Financial Model","Go-NoGo Decision","Battlecards"):
        a.fill=fill(AMBER); a.font=F(bold=True,color=AMBERHEAD,size=10)
    toc.row_dimensions[r].height=18; r+=1
toc.column_dimensions["A"].width=26; toc.column_dimensions["B"].width=84

# =====================================================================
# TAB COLORS + ORDER
# =====================================================================
colormap={"Dashboard":GOLD,"Financial Model":GOLD,"Go-NoGo Decision":GOLD,"Battlecards":GOLD,"01 Contents":NAVY}
for nm,c in colormap.items():
    if nm in wb.sheetnames: wb[nm].sheet_properties.tabColor=c

def move_after(name,after):
    s=wb[name]; wb._sheets.remove(s)
    idx=wb.sheetnames.index(after)+1; wb._sheets.insert(idx,s)
# order: Cover, Contents, Exec, Dashboard, Financial Model, ... , then Go-NoGo & Battlecards after GTM Recommendation
move_after("01 Contents","00 Cover")
move_after("Dashboard","02 Executive Summary")
move_after("Financial Model","Dashboard")
move_after("Go-NoGo Decision","20 GTM Recommendation")
move_after("Battlecards","Go-NoGo Decision")

# print setup for new sheets
from openpyxl.worksheet.properties import PageSetupProperties
for ws in (fm,dash,gng,bc,toc):
    ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
    ws.oddFooter.right.text="&P of &N"; ws.oddFooter.left.text="Organika MUV — Competitor Intelligence (confidential draft)"

wb.save(PATH)
print("v2 saved. sheets:",len(wb.sheetnames))
for s in wb.sheetnames: print("  -",s)
