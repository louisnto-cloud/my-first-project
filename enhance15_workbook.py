#!/usr/bin/env python3
"""Loop batch I21+I22+I17:
 I21 'How to Present' tab — one talking point per key tab for a leadership readout.
 I22 'Next Steps & Gates' tab — immediate actions / decision gates with a status column.
 I17 number-format: format the hidden chart-data helper cells (removes 'General' numerics)."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; STEEL="3E5C76"; LIGHT="EAF1F4"; AMBER="FFF2CC"; AMBERHEAD="B45309"; GREEN="E2EFDA"; GREENHEAD="1E7D32"; GREY="595959"; GOLD="C9A227"; WHITE="FFFFFF"; LINKBLUE="1155CC"
def F(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CTR=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

def new_tab(name,title_txt,sub_txt,ncols):
    if name in wb.sheetnames: del wb[name]
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    t=ws.cell(row=1,column=1,value=title_txt); t.font=F(bold=True,color=WHITE,size=15); t.fill=fill(NAVY); t.alignment=CTR; ws.row_dimensions[1].height=26
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)
    s=ws.cell(row=2,column=1,value=sub_txt); s.font=F(italic=True,color=GREY,size=10); s.alignment=CTR
    h=ws.cell(row=1,column=13,value="🏠 Home"); h.hyperlink="#'🏠 Start Here'!A1"; h.font=F(bold=True,color=LINKBLUE,underline="single",size=8); h.alignment=Alignment(horizontal="right")
    ws.sheet_properties.tabColor=GOLD
    ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0; ws.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
    return ws
def hdr(ws,row,cols):
    for j,h in enumerate(cols):
        x=ws.cell(row=row,column=1+j,value=h); x.font=F(bold=True,color=WHITE,size=10); x.fill=fill(STEEL); x.alignment=CTR; x.border=BORDER
def rows_into(ws,start,data,widths,rowh=30,statuscol=None):
    for i,row in enumerate(data):
        for j,v in enumerate(row):
            x=ws.cell(row=start+i,column=1+j,value=v); x.font=F(size=10,bold=(j==0)); x.alignment=LEFT if j!=statuscol else CTR; x.border=BORDER
            if i%2: x.fill=fill(LIGHT)
            if statuscol is not None and j==statuscol: x.fill=fill(AMBER); x.font=F(size=9,bold=True,color=AMBERHEAD)
        ws.row_dimensions[start+i].height=rowh
    for j,w in enumerate(widths): ws.column_dimensions[chr(65+j)].width=w

# ---------- I21 How to Present ----------
hp=new_tab("How to Present","How to Present — leadership readout","One line to say on each key tab. Lead with the decision; let the tabs back it up.",4)
hdr(hp,4,["Tab","Say this","Then show"])
data=[
 ("Open","“One decision: GO — conditional. Launch MÜV as daily-wellness sparkling hydration, made in Canada.”","Executive Brief 1-page"),
 ("Executive Brief","“Here's the whole case on one page — verdict, numbers, five moves, risks.”","Executive Brief 1-page"),
 ("Dashboard","“The category is real: modern soda $1.8B (+83%), hydration powders $1.5B (+20%).”","Dashboard charts"),
 ("MÜV Peer Set","“We compete with LMNT, Liquid I.V., Nuun — not Poppi. Sparkling is our wedge.”","peer ladder + scatter"),
 ("Financial Model","“Flip the scenario. Even Conservative has a defined break-even; Base is healthy.”","scenario dropdown"),
 ("Canada Regulatory v2","“It's a food, not an NHP. We label to Supplemented Foods now; collagen claims stay on the powder.”","claims fork table"),
 ("Go-No-Go","“The weighted gates say GO — conditional. The conditions are dose/claims and regulatory sign-off.”","auto-verdict"),
 ("Data Confidence / Verification Log","“Every load-bearing number is primary-sourced; estimates are flagged, not hidden.”","scoreboard"),
 ("Risk Register","“Top three risks — claims liability, big-CPG rivals, over-distribution — all mitigable.”","heatmap"),
 ("The ask","“Approve Phase 0: confirm MÜV specs, regulatory sign-off, co-packer quote, Costco beachhead.”","Next Steps & Gates"),
]
rows_into(hp,5,data,[24,68,24],rowh=32)

# ---------- I22 Next Steps & Gates ----------
ns=new_tab("Next Steps & Gates","Next Steps & Decision Gates","Immediate actions to move MÜV forward. Status is illustrative — set owners/dates with the team.",6)
hdr(ns,4,["#","Action","Owner","Timing","Why it's a gate","Status"])
gates=[
 ("1","Confirm MÜV on-can specs (sodium mg, pack size, any collagen variant)","Brand / R&D","Now","Unblocks claims + fills the last model input","Open"),
 ("2","Regulatory counsel sign-off on the claim set (Supplemented Food)","Regulatory","Wk 1–3","De-risks the label; must precede artwork","Open"),
 ("3","Canadian co-packer quote + capacity","Ops","Wk 1–4","Real COGS replaces the illustrative model input","Open"),
 ("4","Bilingual + Quebec Bill 96 artwork","Brand","Wk 3–6","Compliance; long-lead for print","Open"),
 ("5","Costco Canada roadshow slot + Amazon.ca / Well.ca listings","Sales","Wk 4–9","The beachhead trial engine","Open"),
 ("6","Define velocity + repeat-rate targets (KPIs)","Sales / Finance","Wk 2","Gates Phase 2 national sell-in","Open"),
 ("7","Board approval of Phase 0 budget","Leadership","Wk 1","Greenlights the above","Open"),
]
rows_into(ns,5,gates,[5,44,16,12,34,10],rowh=32,statuscol=5)
ns.cell(row=13,column=2,value="Sequenced to the Launch Plan (Tab 21): #1–4 = Phase 0 foundation, #5–6 = Phase 1 beachhead.").font=F(italic=True,size=9,color=GREY)

# place both after GTM Recommendation; contents
def place_after(name,after):
    s=wb[name]; wb._sheets.remove(s); wb._sheets.insert(wb.sheetnames.index(after)+1,s)
place_after("How to Present","20 GTM Recommendation")
place_after("Next Steps & Gates","How to Present")
toc=wb["01 Contents"]
for nm,desc in [("How to Present","Leadership readout — one talking point per key tab"),
                ("Next Steps & Gates","Immediate actions & decision gates with status")]:
    r=toc.max_row+1
    a=toc.cell(row=r,column=1,value=nm); a.hyperlink=f"#'{nm}'!A1"; a.font=F(bold=True,color=LINKBLUE,underline="single",size=10); a.fill=fill(GOLD); a.border=BORDER; a.alignment=LEFT
    b=toc.cell(row=r,column=2,value=desc); b.font=F(size=10); b.border=BORDER; b.alignment=LEFT

# ---------- I17 format hidden chart-data helpers ----------
scn=wb["Scenario Comparison"]
for row in scn.iter_rows():
    for c in row:
        if isinstance(c.value,(int,float)) and c.number_format in (None,"General"):
            c.number_format='#,##0'

wb.save(PATH)
print("I21 How to Present + I22 Next Steps & Gates added; I17 helper formats set. sheets:",len(wb.sheetnames))
