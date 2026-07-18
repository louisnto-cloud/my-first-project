#!/usr/bin/env python3
"""Loop batch: Start Here headline strip; I29 sources-by-tier count; I28 KPI peer-price benchmark bar."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; STEEL="3E5C76"; LIGHT="EAF1F4"; AMBER="FFF2CC"; GREEN="E2EFDA"; GREENHEAD="1E7D32"; GREY="595959"; GOLD="C9A227"; WHITE="FFFFFF"; TEAL="2E7D8A"
def F(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

# ---------- Start Here headline strip (row 4, currently empty) ----------
sh=wb["🏠 Start Here"]
sh.merge_cells(start_row=4,start_column=2,end_row=4,end_column=5)
c=sh.cell(row=4,column=2,value="US hydration powders $1.5B (+20%)  ·  sparkling electrolyte = white space  ·  MÜV = food-regulated RTD can  ·  Verdict: GO — conditional")
c.font=F(bold=True,color=GREENHEAD,size=9); c.fill=fill(GREEN); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=BORDER
sh.row_dimensions[4].height=16

# ---------- I29: sources-by-tier count ----------
src=wb["24 Sources"]
# find 'Type' column
tcol=trow=None
for row in src.iter_rows():
    for cell in row:
        if isinstance(cell.value,str) and cell.value.strip()=="Type":
            tcol=cell.column; trow=cell.row; break
    if tcol: break
counts={}
if tcol:
    for r in range(trow+1,src.max_row+1):
        v=src.cell(row=r,column=tcol).value
        if not isinstance(v,str) or not v.strip(): continue
        key = "Primary" if "Primary" in v else ("Regulatory" if "regulator" in v.lower() else ("Scanner" if "Scanner" in v or "scanner" in v else ("Trade" if "Trade" in v else "Other")))
        counts[key]=counts.get(key,0)+1
    total=sum(counts.values()) or 1
    r=src.max_row+2
    src.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
    summ="SOURCE MIX: " + " · ".join(f"{k} {v}" for k,v in sorted(counts.items(), key=lambda x:-x[1])) + f"  (total {total} rows; load-bearing claims lean primary/regulatory)."
    x=src.cell(row=r,column=1,value=summ); x.font=F(bold=True,size=9,color=STEEL); x.fill=fill(LIGHT); x.alignment=Alignment(wrap_text=True,vertical="center"); x.border=BORDER
    src.row_dimensions[r].height=24

# ---------- I28: KPI peer-price benchmark bar ----------
kpi=wb["23 KPI Dashboard"]
base=kpi.max_row+3
kpi.cell(row=base,column=1,value="(peer price/serving benchmark data)").font=F(size=9,italic=True,color=GREY)
data=[("LMNT",1.50),("Liquid I.V.",1.56),("Nuun",0.75),("Organika sachet",1.00)]
kpi.cell(row=base+1,column=1,value="Brand"); kpi.cell(row=base+1,column=2,value="Price/serving")
for i,(b,p) in enumerate(data):
    kpi.cell(row=base+2+i,column=1,value=b); kpi.cell(row=base+2+i,column=2,value=p)
ch=BarChart(); ch.type="bar"; ch.title="Peer price per serving ($) — where to anchor MÜV"; ch.height=6.5; ch.width=13
ch.add_data(Reference(kpi,min_col=2,min_row=base+1,max_row=base+1+len(data)),titles_from_data=True)
ch.set_categories(Reference(kpi,min_col=1,min_row=base+2,max_row=base+1+len(data)))
ch.legend=None; ch.dataLabels=DataLabelList(); ch.dataLabels.showVal=True
kpi.add_chart(ch,"H3")
for rr in range(base,base+2+len(data)): kpi.row_dimensions[rr].hidden=True

wb.save(PATH)
print(f"Start Here strip; I29 source mix {counts}; I28 KPI peer-price bar added.")
