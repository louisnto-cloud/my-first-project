#!/usr/bin/env python3
"""Iteration I2+I3 (easy interface): a '🏠 Start Here' navigator front-door with grouped,
clickable cards; a Home link on every tab; opens on Start Here."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; TEAL="2E7D8A"; STEEL="3E5C76"; LIGHT="EAF1F4"; LIGHT2="F4F8FA"; AMBER="FFF2CC"; AMBERHEAD="B45309"
GREEN="E2EFDA"; GREENHEAD="1E7D32"; GOLD="C9A227"; WHITE="FFFFFF"; GREY="595959"; LINKBLUE="1155CC"
def F(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); med=Side(style="medium",color="FFFFFF")
CARD_BORDER=Border(left=med,right=med,top=med,bottom=med)
CTR=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

if "🏠 Start Here" in wb.sheetnames: del wb["🏠 Start Here"]
sh=wb.create_sheet("🏠 Start Here")
sh.sheet_view.showGridLines=False
for col,w in zip("ABCDEFGH",[3,26,26,26,26,3,3,3]): sh.column_dimensions[col].width=w

sh.merge_cells("B2:E2"); t=sh["B2"]; t.value="MÜV — Canada Launch Dossier"; t.font=F(bold=True,color=WHITE,size=22); t.fill=fill(NAVY); t.alignment=CTR; sh.row_dimensions[2].height=44
sh.merge_cells("B3:E3"); s=sh["B3"]; s.value="Competitive intelligence & go-to-market for Organika's sparkling electrolyte RTD.  Click a card to jump in."; s.font=F(italic=True,color=WHITE,size=11); s.fill=fill(TEAL); s.alignment=CTR; sh.row_dimensions[3].height=24

def card(r,c,label,target,color):
    sh.merge_cells(start_row=r,start_column=c,end_row=r+1,end_column=c)
    cell=sh.cell(row=r,column=c,value=label)
    cell.hyperlink=f"#'{target}'!A1"; cell.font=F(bold=True,color=WHITE,size=11); cell.fill=fill(color)
    cell.alignment=CTR; cell.border=CARD_BORDER
    sh.row_dimensions[r].height=22; sh.row_dimensions[r+1].height=22

def section(r,ttl):
    sh.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
    c=sh.cell(row=r,column=2,value=ttl); c.font=F(bold=True,color=NAVY,size=12); c.alignment=LEFT
    sh.row_dimensions[r].height=22

r=5
section(r,"▸ See the answer"); r+=1
card(r,2,"Executive Summary","02 Executive Summary",STEEL); card(r,3,"GTM Recommendation","20 GTM Recommendation",GREENHEAD); card(r,4,"Go / No-Go Decision","Go-NoGo Decision",GREENHEAD); r+=3
section(r,"▸ Explore the numbers"); r+=1
card(r,2,"Dashboard","Dashboard",TEAL); card(r,3,"Financial Model","Financial Model",TEAL); card(r,4,"Scenario Comparison","Scenario Comparison",TEAL); r+=3
section(r,"▸ Know the competition"); r+=1
card(r,2,"MÜV Peer Set","MÜV Peer Set",GOLD); card(r,3,"Comparison Matrix","07 Comparison Matrix",STEEL); card(r,4,"Extended Landscape","08 Extended Landscape",STEEL); r+=3
section(r,"▸ Canada launch"); r+=1
card(r,2,"Canada Regulatory v2","Canada Regulatory v2",GREENHEAD); card(r,3,"Electrolyte BFY Deep-Dive","Electrolyte BFY Deep-Dive",TEAL); card(r,4,"Canada Market","18 Canada Market",STEEL); r+=3
section(r,"▸ Trust the data"); r+=1
card(r,2,"Verification Log","Verification Log",GOLD); card(r,3,"Sources","24 Sources",GREY); card(r,4,"Full Contents Index","01 Contents",NAVY); r+=3

sh.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
note=sh.cell(row=r,column=2,value=("Verified anchor: MÜV Sparkling Electrolytes is a ready-to-drink sparkling CAN (SKU 4338), food-regulated in Canada. "
 "Gold tabs are interactive. Every tab has a “↩ Home” link (top-right) back to this page."))
note.font=F(italic=True,color=GREY,size=9); note.fill=fill(LIGHT); note.alignment=LEFT; note.border=Border(left=thin,right=thin,top=thin,bottom=thin)
sh.row_dimensions[r].height=34

# Home link on every other tab (col 13), keep Contents link in col 14
for name in wb.sheetnames:
    if name=="🏠 Start Here": continue
    ws=wb[name]
    h=ws.cell(row=1,column=13,value="🏠 Home"); h.hyperlink="#'🏠 Start Here'!A1"; h.font=F(bold=True,color=LINKBLUE,underline="single",size=8); h.alignment=Alignment(horizontal="right")

# make Start Here first + active
s=wb["🏠 Start Here"]; wb._sheets.remove(s); wb._sheets.insert(0,s)
wb.active=0
for ws in wb.worksheets: ws.sheet_view.tabSelected = (ws.title=="🏠 Start Here")
sh.sheet_properties.tabColor=GOLD
sh.page_setup.orientation="landscape"; sh.page_setup.fitToWidth=1; sh.page_setup.fitToHeight=0; sh.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)

# contents entry
toc=wb["01 Contents"]
r=toc.max_row+1
a=toc.cell(row=r,column=1,value="🏠 Start Here"); a.hyperlink="#'🏠 Start Here'!A1"; a.font=F(bold=True,color=LINKBLUE,underline="single",size=10); a.fill=fill(GOLD); a.border=Border(left=thin,right=thin,top=thin,bottom=thin); a.alignment=LEFT
b=toc.cell(row=r,column=2,value="Navigator front-door — click a card to jump to any section"); b.font=F(size=10); b.border=Border(left=thin,right=thin,top=thin,bottom=thin); b.alignment=LEFT

wb.save(PATH)
print("Start Here front-door added. sheets:",len(wb.sheetnames))
