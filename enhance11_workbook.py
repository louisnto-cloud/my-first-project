#!/usr/bin/env python3
"""Loop batch I11+I12:
 I11 break-even-units-by-scenario bar chart on the Scenario Comparison tab.
 I12 glossary: append a 'Regulatory & format terms' section (SFFt, SFCI, ORS, NPN, NHP interface, RTD…)."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; STEEL="3E5C76"; LIGHT="EAF1F4"; GREY="595959"
def F(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

# ---------- I11: break-even by scenario (recompute; same logic as Financial Model) ----------
S={"Conservative":dict(price=3.29,cogs=1.20,ret=0.38,dist=0.15+0.03,trade=0.25,mkt=150000,fixed=250000,slot=50000),
   "Base":dict(price=3.49,cogs=1.05,ret=0.35,dist=0.15,trade=0.20,mkt=300000,fixed=350000,slot=100000),
   "Aggressive":dict(price=3.79,cogs=0.95,ret=0.33,dist=0.12,trade=0.15,mkt=600000,fixed=500000,slot=200000)}
def be(s):
    net=s["price"]*(1-s["ret"])*(1-s["dist"]); cont=(net-s["cogs"])-net*s["trade"]
    load=s["mkt"]+s["fixed"]+s["slot"]
    return round(load/cont) if cont>0 else 0
order=["Conservative","Base","Aggressive"]; vals=[be(S[k]) for k in order]

sc=wb["Scenario Comparison"]
base=sc.max_row+3
sc.cell(row=base,column=1,value="(break-even chart data)").font=F(size=9,italic=True,color=GREY)
sc.cell(row=base+1,column=1,value="Scenario"); sc.cell(row=base+1,column=2,value="Break-even units")
for i,k in enumerate(order):
    sc.cell(row=base+2+i,column=1,value=k); sc.cell(row=base+2+i,column=2,value=vals[i])
ch=BarChart(); ch.type="col"; ch.title="Break-even units by scenario (Yr-1 cost load ÷ contribution/can)"; ch.height=7.5; ch.width=15
ch.add_data(Reference(sc,min_col=2,min_row=base+1,max_row=base+1+len(order)),titles_from_data=True)
ch.set_categories(Reference(sc,min_col=1,min_row=base+2,max_row=base+1+len(order)))
ch.legend=None; ch.dataLabels=DataLabelList(); ch.dataLabels.showVal=True
sc.add_chart(ch,"H4")
for rr in range(base,base+2+len(order)): sc.row_dimensions[rr].hidden=True

# ---------- I12: glossary regulatory terms ----------
gl=wb["25 Glossary"]
r=gl.max_row+2
gl.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
h=gl.cell(row=r,column=1,value="REGULATORY & FORMAT TERMS (Canada)"); h.font=F(bold=True,color="FFFFFF",size=10); h.fill=fill(STEEL); h.alignment=Alignment(horizontal="left",vertical="center",indent=1)
r+=1
terms=[
 ("RTD","Ready-to-drink — a finished beverage (e.g., a can), vs a powder/tablet to reconstitute."),
 ("NHP","Natural Health Product — Canadian licensed supplement category; bears an NPN."),
 ("NPN","Natural Product Number — 8-digit licence number on an NHP."),
 ("Supplemented Food","A food with added vitamins/minerals/amino acids beyond normal fortification, under the Supplemented Foods Regulations (in force 2022-07-21)."),
 ("SFFt","Supplemented Food Facts table — modified Nutrition Facts table with a 'Supplemented with' line."),
 ("SFCI","Supplemented Food Caution Identifier — the black-&-white '!' identifier required when cautionary statements apply."),
 ("ORS","Oral Rehydration Solution — clinical rehydration product; stays regulated as an NHP (carve-out from the electrolyte→food shift)."),
 ("Food–NHP interface","Health Canada guidance that classifies food-format products; format is not decisive — representation + claims are."),
 ("Fibersol","A soluble prebiotic dietary fibre used in MÜV."),
 ("Bill 96 / OQLF","Quebec French-language packaging rules; generic/descriptive trademark terms need French."),
]
for i,(term,defn) in enumerate(terms):
    a=gl.cell(row=r,column=1,value=term); a.font=F(bold=True,size=10); a.border=BORDER; a.alignment=Alignment(vertical="top",wrap_text=True)
    b=gl.cell(row=r,column=2,value=defn); b.font=F(size=10); b.border=BORDER; b.alignment=Alignment(wrap_text=True,vertical="top")
    if i%2: a.fill=fill(LIGHT); b.fill=fill(LIGHT)
    gl.row_dimensions[r].height=26
    r+=1

wb.save(PATH)
print(f"I11 break-even chart (units {vals}); I12 glossary +{len(terms)} terms.")
