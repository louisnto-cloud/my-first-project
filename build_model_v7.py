"""
Costco UK D93 Portfolio ROI Model — v7

Vs v6:
  - Every tab (Forecast, Summary, Portfolio — plus the existing Inputs)
    now has its own Display Currency dropdown at the top, with its own
    FX rate and symbol pulled from the shared FX Rate Table on Inputs.
  - Each tab is independently selectable. Summary's pull from Forecast
    is adjusted by (Summary.FXRate / Forecast.FXRate) so the numbers
    stay consistent even if the two tabs are set to different currencies.
"""

import openpyxl
from datetime import date
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

NAVY="1F4E78"; BLUE="2E75B6"; LIGHT_BLUE="DDEBF7"
YELLOW="FFF2CC"; GREY="F2F2F2"; WHITE="FFFFFF"
GREEN="C6EFCE"; RED="FFC7CE"; DARK_GREY="595959"

thin = Side(border_style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

f_title       = Font(name="Calibri", size=18, bold=True, color=WHITE)
f_section     = Font(name="Calibri", size=12, bold=True, color=WHITE)
f_subhead     = Font(name="Calibri", size=11, bold=True, color=NAVY)
f_label       = Font(name="Calibri", size=11, bold=True)
f_input       = Font(name="Calibri", size=11, color="0070C0", bold=True)
f_calc        = Font(name="Calibri", size=11)
f_note        = Font(name="Calibri", size=10, italic=True, color=DARK_GREY)
f_kpi_val     = Font(name="Calibri", size=14, bold=True, color=NAVY)
f_white_bold  = Font(name="Calibri", size=11, bold=True, color=WHITE)

fill_title   = PatternFill("solid", fgColor=NAVY)
fill_section = PatternFill("solid", fgColor=BLUE)
fill_input   = PatternFill("solid", fgColor=YELLOW)
fill_calc    = PatternFill("solid", fgColor=GREY)
fill_kpi     = PatternFill("solid", fgColor=LIGHT_BLUE)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
right  = Alignment(horizontal="right", vertical="center")
top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

FMT_GBP_HARD   = '_-£* #,##0_-;[Red]-£* #,##0_-;_-£* "-"??_-;_-@_-'
FMT_GBP_HARD_D = '_-£* #,##0.00_-;[Red]-£* #,##0.00_-;_-£* "-"??_-;_-@_-'
FMT_MONEY      = '_-* #,##0_-;[Red]_-* (#,##0)_-;_-* "-"??_-;_-@_-'
FMT_MONEY_D    = '_-* #,##0.00_-;[Red]_-* (#,##0.00)_-;_-* "-"??_-;_-@_-'
FMT_INT  = "#,##0"
FMT_PCT  = "0.0%"
FMT_MULT = '0.0"x"'
FMT_DATE = "dd-mmm"
FMT_DATE_L = "dd-mmm-yyyy"
FMT_RATE = "0.0000"

WEEKS = 26
FIRST_WEEK_NUM = 1
DEFAULT_START_DATE = date(2026, 1, 12)

CURRENCIES = [
    ("GBP", "£",  1.0000),
    ("USD", "$",  1.2700),
    ("CAD", "C$", 1.7100),
    ("EUR", "€",  1.1700),
]

SCENARIOS = [
    {"key":"BEST","name":"BEST – x2 Demo (Wks 5-6) + TPD",
     "promo":{"Demo":[5,6],"End Cap":[],"Fence":[],"TPD":[5,6],"Markdown":[],"Advert":[]},
     "lift_overrides":{5:3.0,6:3.0},
     "rationale":"Two demo weekends with TPD. Demo lifts ~3x in-week. No end cap, no markdown."},
    {"key":"IDEAL","name":"IDEAL – x2 Demo + End Cap (Wks 5-6) + TPD",
     "promo":{"Demo":[5,6],"End Cap":[5,6],"Fence":[],"TPD":[5,6],"Markdown":[],"Advert":[]},
     "lift_overrides":{5:3.6,6:3.6},
     "rationale":"Demo + end cap + TPD stacked for two weeks. Combined velocity ~3.6x base."},
    {"key":"WORST","name":"WORST – Demo+EndCap+TPD then Markdown (Wks 18-20)",
     "promo":{"Demo":[5,6],"End Cap":[5,6],"Fence":[],"TPD":[5,6],"Markdown":[18,19,20],"Advert":[]},
     "lift_overrides":{5:3.0,6:3.0,18:1.0,19:0.3,20:8.7},
     "rationale":"Promo plan as IDEAL but residual inventory clears at 50% off over Wks 18-20."},
]
PROMO_TYPES = ["Demo","End Cap","Fence","TPD","Markdown","Advert"]

PORTFOLIO_SKUS = [
    {"rank":1,"code":"UK-D93-#1","name":"Bee Propolis Spray 30ml x2",
     "pack":"30ml x 2","sp":15.99,"cost":7.29,"units":17000,"promo":25000,
     "rationale":"Lead pitch — strong GP%, supported by demo + TPD."},
    {"rank":2,"code":"UK-D93-#2","name":"Mag 8-in-1 90's/120's",
     "pack":None,"sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":3,"code":"UK-D93-#3","name":"O1 450g",
     "pack":None,"sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":4,"code":"UK-D93-#4","name":"Belli Bliss Raspberry, 450g",
     "pack":None,"sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":5,"code":"UK-D93-#5","name":"Daily Boost",
     "pack":None,"sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":6,"code":"UK-D93-#6","name":"Berberine 90's/120's",
     "pack":None,"sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":7,"code":"UK-D93-#7","name":"L-theanine Capsules 90's/120's",
     "pack":None,"sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
]

wb = openpyxl.Workbook()
wb.remove(wb.active)


def wc(ws, r, c, v, *, font=None, fill=None, align=None, border=None, number_format=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font is not None: cell.font = font
    if fill is not None: cell.fill = fill
    if align is not None: cell.alignment = align
    if border is not None: cell.border = border
    if number_format is not None: cell.number_format = number_format
    return cell


def section_header(ws, r, text, span_start=2, span_end=4):
    ws.merge_cells(start_row=r, start_column=span_start, end_row=r, end_column=span_end)
    wc(ws, r, span_start, text, font=f_section, fill=fill_section, align=left)
    ws.row_dimensions[r].height = 22


def add_currency_widget(ws, row, default="GBP"):
    """Inline Currency picker on a tab. Returns (cur_ref, fx_ref, sym_ref)
    as absolute cell refs for use in this tab's formulas."""
    wc(ws, row, 2, "Display Currency:", font=f_label, align=right)
    wc(ws, row, 3, default, font=f_input, fill=fill_input, align=center, border=box)
    wc(ws, row, 4, "FX Rate (× GBP):", font=f_label, align=right)
    wc(ws, row, 5,
       f"=VLOOKUP(C{row},Inputs!$B$13:$D$16,3,FALSE)",
       font=f_calc, fill=fill_calc, align=center, border=box, number_format=FMT_RATE)
    wc(ws, row, 6, "Symbol:", font=f_label, align=right)
    wc(ws, row, 7,
       f"=VLOOKUP(C{row},Inputs!$B$13:$D$16,2,FALSE)",
       font=f_calc, fill=fill_calc, align=center, border=box)
    dv = DataValidation(type="list", formula1='"GBP,USD,CAD,EUR"', allow_blank=False)
    dv.error = "Pick one of GBP, USD, CAD, EUR."
    dv.errorTitle = "Invalid currency"
    ws.add_data_validation(dv)
    dv.add(f"C{row}")
    ws.row_dimensions[row].height = 22
    return (f"$C${row}", f"$E${row}", f"$G${row}")


# =============================================================================
# README
# =============================================================================
rd = wb.create_sheet("README")
rd.sheet_view.showGridLines = False
rd.column_dimensions["A"].width = 4
rd.column_dimensions["B"].width = 110

wc(rd,2,2,"Costco UK — D93 Portfolio ROI Model",
   font=Font(name="Calibri",size=20,bold=True,color=NAVY),align=left)
wc(rd,3,2,"v7.0 · Currency dropdown on every tab",
   font=Font(name="Calibri",size=12,italic=True,color=DARK_GREY),align=left)

wc(rd,5,2,"HOW TO USE",font=Font(name="Calibri",size=13,bold=True,color=WHITE),
   fill=fill_section,align=left)

lines = [
    "",
    "1.  Display Currency. Every tab has its own dropdown (GBP / USD / CAD / EUR)",
    "    at the top. Switching the dropdown on a tab converts every money cell on",
    "    that tab and updates the column headers with the right symbol.",
    "",
    "2.  FX rates live in one place — the FX Rate Table on the Inputs sheet.",
    "    Edit any rate there and every tab updates.",
    "",
    "3.  Tabs are INDEPENDENT. Set them all to the same currency for a consistent",
    "    pitch deck view, or set them differently if you want to compare. Summary",
    "    automatically adjusts its pull from Forecast if the two tabs disagree.",
    "",
    "4.  Inputs themselves stay in GBP — the source of truth for the brand's",
    "    accounting currency. The Forecast / Summary / Portfolio tabs are the",
    "    display layer.",
    "",
    "5.  YELLOW = editable. Calc cells (GM/u, GM%, Base Units/Week, Demo Total,",
    "    End Cap Total) hold a formula by default; type over to override.",
    "",
    "6.  Test Start Date (W1) on Inputs drives every week column's date.",
    "",
    "",
]
r = 6
for ln in lines:
    wc(rd,r,2,ln,font=Font(name="Calibri",size=11),align=top_wrap); r += 1

wc(rd,r,2,"COLOUR KEY",font=f_subhead); r += 1
for label,color,desc in [
    ("Yellow",     YELLOW,     "Editable input — change me"),
    ("Light Blue", LIGHT_BLUE, "Summary KPI"),
    ("Grey",       GREY,       "Reference / structure"),
]:
    wc(rd,r,2,f"   {label}: {desc}",align=left,
       fill=PatternFill("solid",fgColor=color),font=Font(name="Calibri",size=11))
    r += 1


# =============================================================================
# INPUTS  (unchanged from v6 — keeps master FX table at B13:D16)
# =============================================================================
inp = wb.create_sheet("Inputs")
inp.sheet_view.showGridLines = False
inp.column_dimensions["A"].width = 2
inp.column_dimensions["B"].width = 38
inp.column_dimensions["C"].width = 16
inp.column_dimensions["D"].width = 78

inp.merge_cells("B2:D2")
wc(inp,2,2,"INPUTS & ASSUMPTIONS — change any yellow cell",
   font=f_title,fill=fill_title,align=center)
inp.row_dimensions[2].height = 32

wc(inp,3,2,"Field",font=f_subhead,fill=fill_calc,align=left,border=box)
wc(inp,3,3,"Value",font=f_subhead,fill=fill_calc,align=center,border=box)
wc(inp,3,4,"Notes / justification",font=f_subhead,fill=fill_calc,align=left,border=box)


def input_row(r, label, value, note, fmt=None, is_formula=False):
    wc(inp,r,2,label,font=f_label,align=left)
    cell_value = f"={value}" if is_formula else value
    cell_font = f_calc if is_formula else f_input
    wc(inp,r,3,cell_value,font=cell_font,fill=fill_input,
       align=right,border=box,number_format=fmt)
    wc(inp,r,4,note,font=f_note,align=top_wrap)
    inp.row_dimensions[r].height = 24


section_header(inp, 5, "1.  Display Currency")
wc(inp,6,2,"Display Currency",font=f_label,align=left)
wc(inp,6,3,"GBP",font=f_input,fill=fill_input,align=center,border=box)
wc(inp,6,4,"Dropdown. Inputs stays in GBP — this selection is informational here. "
            "Each display tab has its own currency dropdown that drives that tab.",
   font=f_note,align=top_wrap)
inp.row_dimensions[6].height = 30

wc(inp,7,2,"FX Rate (auto, from table)",font=f_label,align=left)
wc(inp,7,3,"=VLOOKUP(C6,B13:D16,3,FALSE)",
   font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_RATE)
wc(inp,7,4,"Reference only — display tabs each look up their own rate.",
   font=f_note,align=top_wrap)
inp.row_dimensions[7].height = 22

wc(inp,8,2,"Currency Symbol (auto)",font=f_label,align=left)
wc(inp,8,3,"=VLOOKUP(C6,B13:D16,2,FALSE)",
   font=f_calc,fill=fill_calc,align=center,border=box)
wc(inp,8,4,"Reference only.",font=f_note,align=top_wrap)
inp.row_dimensions[8].height = 22

inp.merge_cells("B10:D10")
wc(inp,10,2,"FX Rate Table — edit rates as needed (rate is per 1 GBP)",
   font=f_subhead,fill=fill_calc,align=left,border=box)

wc(inp,12,2,"Code",  font=f_white_bold,fill=fill_section,align=center,border=box)
wc(inp,12,3,"Symbol",font=f_white_bold,fill=fill_section,align=center,border=box)
wc(inp,12,4,"Rate (× GBP)",font=f_white_bold,fill=fill_section,align=center,border=box)

for i,(code,sym,rate) in enumerate(CURRENCIES):
    r = 13 + i
    wc(inp,r,2,code,font=f_label,fill=fill_calc,align=center,border=box)
    wc(inp,r,3,sym, font=f_input,fill=fill_input,align=center,border=box)
    wc(inp,r,4,rate,font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_RATE)
    inp.row_dimensions[r].height = 20

dv = DataValidation(type="list", formula1='"GBP,USD,CAD,EUR"', allow_blank=False)
inp.add_data_validation(dv)
dv.add("C6")

section_header(inp, 18, "2.  Product & Distribution")
input_row(19,"SKU","Bee Propolis Spray 30ml x 2","Costco UK market test SKU.")
input_row(20,"Test Start Date (W1)",DEFAULT_START_DATE,
          "Week-commencing date for W1.",fmt=FMT_DATE_L)
input_row(21,"Test Warehouses",20,"Warehouses in test.",fmt=FMT_INT)
input_row(22,"Test Period (Weeks)",WEEKS,
          f"Forecast horizon ({WEEKS} weeks).",fmt=FMT_INT)

section_header(inp, 24, "3.  Pricing & Cost (per unit, GBP)")
input_row(25,"Selling Price (MSRP, £)",15.99,"Retail per pack.",fmt=FMT_GBP_HARD_D)
input_row(26,"Cost Price (landed, £)",7.29,"Brand cost delivered.",fmt=FMT_GBP_HARD_D)
input_row(27,"Gross Margin per unit (£)","C25-C26",
          "Default = SP − Cost.",fmt=FMT_GBP_HARD_D,is_formula=True)
input_row(28,"Gross Margin %","IFERROR((C25-C26)/C25,0)",
          "Default = GP ÷ SP.",fmt=FMT_PCT,is_formula=True)

section_header(inp, 30, "4.  Base Demand (no promo)")
input_row(31,"Units / Warehouse / Week",13.79,
          "Prior 30ml SKU run-rate (500 × 0.8 / 29).",fmt="#,##0.00")
input_row(32,"Base Units / Week (all warehouses)","C31*C21",
          "Default = per-wh rate × warehouses.",fmt=FMT_INT,is_formula=True)

section_header(inp, 34, "5.  Promo Economics (GBP)")
input_row(35,"Demo £ / warehouse / week",199,"Costco demo rate.",fmt=FMT_GBP_HARD)
input_row(36,"Demo cost factor",1.86,"Overhead multiplier.",fmt=FMT_MULT)
input_row(37,"Demo Total £ / week","C35*C21*C36",
          "Default = rate × wh × factor.",fmt=FMT_GBP_HARD,is_formula=True)
input_row(38,"End Cap £ / warehouse / booking",850,"End cap rental.",fmt=FMT_GBP_HARD)
input_row(39,"End Cap cost factor",1.86,"Same multiplier.",fmt=FMT_MULT)
input_row(40,"End Cap booking cycle (weeks)",2,"Per-week cost = total ÷ this.",fmt=FMT_INT)
input_row(41,"End Cap Total £ / week","C38*C21*C39/C40",
          "Default = rate × wh × factor ÷ cycle.",fmt=FMT_GBP_HARD,is_formula=True)
input_row(42,"Fence £ / week (when active)",0,"Not currently budgeted.",fmt=FMT_GBP_HARD)
input_row(43,"Advertising £ / week (when active)",0,"Not currently budgeted.",fmt=FMT_GBP_HARD)
input_row(44,"TPD discount % (of net)",0.25,"20% off MSRP ≈ 25% on net.",fmt=FMT_PCT)
input_row(45,"Markdown discount %",0.50,"Clearance pricing.",fmt=FMT_PCT)

wb.defined_names["StartDate"]  = DefinedName("StartDate",  attr_text="Inputs!$C$20")
wb.defined_names["WHs"]        = DefinedName("WHs",        attr_text="Inputs!$C$21")
wb.defined_names["Period"]     = DefinedName("Period",     attr_text="Inputs!$C$22")
wb.defined_names["Price"]      = DefinedName("Price",      attr_text="Inputs!$C$25")
wb.defined_names["Cost"]       = DefinedName("Cost",       attr_text="Inputs!$C$26")
wb.defined_names["BaseUnits"]  = DefinedName("BaseUnits",  attr_text="Inputs!$C$32")
wb.defined_names["DemoCost"]   = DefinedName("DemoCost",   attr_text="Inputs!$C$37")
wb.defined_names["EndCapCost"] = DefinedName("EndCapCost", attr_text="Inputs!$C$41")
wb.defined_names["FenceCost"]  = DefinedName("FenceCost",  attr_text="Inputs!$C$42")
wb.defined_names["AdvertCost"] = DefinedName("AdvertCost", attr_text="Inputs!$C$43")
wb.defined_names["TPDPct"]     = DefinedName("TPDPct",     attr_text="Inputs!$C$44")
wb.defined_names["MdPct"]      = DefinedName("MdPct",      attr_text="Inputs!$C$45")

section_header(inp, 47, "6.  Scenario activity & demand lift",span_start=2,span_end=4)
inp.merge_cells("B48:D48")
wc(inp,48,2,
   "Toggle 1 = promo active, 0 = off. Lift is demand multiplier vs base. "
   "Dates auto-fill.",
   font=f_note,align=top_wrap)
inp.row_dimensions[48].height = 30

WEEK_START_COL = 6
WEEK_END_COL = WEEK_START_COL + WEEKS - 1
for col in range(WEEK_START_COL, WEEK_END_COL + 1):
    inp.column_dimensions[get_column_letter(col)].width = 7.5

scenario_grid_start = {}
current = 50
for sc in SCENARIOS:
    inp.merge_cells(start_row=current,start_column=2,end_row=current,end_column=WEEK_END_COL)
    wc(inp,current,2,sc["name"],font=f_white_bold,fill=fill_title,align=left)
    inp.row_dimensions[current].height = 22
    current += 1

    inp.merge_cells(start_row=current,start_column=2,end_row=current,end_column=WEEK_END_COL)
    wc(inp,current,2,f"Rationale: {sc['rationale']}",font=f_note,align=top_wrap)
    inp.row_dimensions[current].height = 30
    current += 1

    wc(inp,current,2,"Week",font=f_label,fill=fill_calc,align=left)
    for i in range(WEEKS):
        wc(inp,current,WEEK_START_COL+i,f"W{FIRST_WEEK_NUM+i}",
           font=f_white_bold,fill=PatternFill("solid",fgColor=BLUE),
           align=center,border=box)
    current += 1

    wc(inp,current,2,"Wk-commencing",font=f_label,fill=fill_calc,align=left)
    for i in range(WEEKS):
        wc(inp,current,WEEK_START_COL+i,f"=StartDate+7*{i}",
           font=Font(name="Calibri",size=9,color=DARK_GREY),
           fill=fill_calc,align=center,border=box,number_format=FMT_DATE)
    current += 1

    scenario_grid_start[sc["key"]] = current

    for promo in PROMO_TYPES:
        wc(inp,current,2,promo,font=f_label,align=left)
        active = set(sc["promo"].get(promo,[]))
        for i in range(WEEKS):
            wknum = FIRST_WEEK_NUM + i
            wc(inp,current,WEEK_START_COL+i,
               1 if wknum in active else 0,
               font=f_input,fill=fill_input,align=center,border=box,
               number_format="0;;;@")
        current += 1

    wc(inp,current,2,"Lift Multiplier (x base)",font=f_label,align=left)
    for i in range(WEEKS):
        wknum = FIRST_WEEK_NUM + i
        wc(inp,current,WEEK_START_COL+i,
           sc["lift_overrides"].get(wknum,1.0),
           font=f_input,fill=fill_input,align=center,border=box,
           number_format=FMT_MULT)
    current += 1
    current += 1

inp.freeze_panes = "F5"


# =============================================================================
# FORECAST  (currency widget at row 3)
# =============================================================================
fc = wb.create_sheet("Forecast")
fc.sheet_view.showGridLines = False
fc.column_dimensions["A"].width = 2
fc.column_dimensions["B"].width = 32
for col in range(WEEK_START_COL,WEEK_END_COL+1):
    fc.column_dimensions[get_column_letter(col)].width = 11
TOTAL_COL = WEEK_END_COL + 1
fc.column_dimensions[get_column_letter(TOTAL_COL)].width = 14

fc.merge_cells(start_row=2,start_column=2,end_row=2,end_column=TOTAL_COL)
wc(fc,2,2,'="FORECAST — Weekly P&L by Scenario  (Display: "&$G$3&" "&$C$3&")"',
   font=f_title,fill=fill_title,align=center)
fc.row_dimensions[2].height = 32

FC_CUR, FC_FX, FC_SYM = add_currency_widget(fc, 3, default="GBP")

PNL_ROWS = [
    ("POS Units",                                "units"),
    ('="Selling Price ("&$G$3&")"',              "price"),
    ('="Revenue ("&$G$3&")"',                    "rev"),
    ('="Cost Price ("&$G$3&")"',                 "cost"),
    ('="Raw COGS ("&$G$3&")"',                   "cogs"),
    ('="— Demo Spend ("&$G$3&")"',               "demo"),
    ('="— End Cap Spend ("&$G$3&")"',            "endcap"),
    ('="— Fence Spend ("&$G$3&")"',              "fence"),
    ('="— TPD Spend ("&$G$3&")"',                "tpd"),
    ('="— Markdown Spend ("&$G$3&")"',           "markdown"),
    ('="— Advertising Spend ("&$G$3&")"',        "advert"),
    ('="Total Promo Spend ("&$G$3&")"',          "promo_total"),
    ('="Gross Profit ("&$G$3&")"',               "gp"),
    ('="Net Profit after promo ("&$G$3&")"',     "net"),
]
PROMO_ROW_OFFSET = {p:i for i,p in enumerate(PROMO_TYPES)}
LIFT_OFFSET = len(PROMO_TYPES)

fc_row = 5  # leave row 4 blank as breathing space below the currency widget
scenario_summary_rows = {}

for sc in SCENARIOS:
    fc.merge_cells(start_row=fc_row,start_column=2,end_row=fc_row,end_column=TOTAL_COL)
    wc(fc,fc_row,2,sc["name"],font=f_white_bold,fill=fill_title,align=left)
    fc.row_dimensions[fc_row].height = 22
    fc_row += 1

    wc(fc,fc_row,2,"Week",font=f_label,fill=fill_calc,align=left)
    for i in range(WEEKS):
        wc(fc,fc_row,WEEK_START_COL+i,f"W{FIRST_WEEK_NUM+i}",
           font=f_white_bold,fill=PatternFill("solid",fgColor=BLUE),
           align=center,border=box)
    wc(fc,fc_row,TOTAL_COL,"TOTAL",
       font=f_white_bold,fill=fill_title,align=center,border=box)
    fc_row += 1

    wc(fc,fc_row,2,"Wk-commencing",font=f_label,fill=fill_calc,align=left)
    for i in range(WEEKS):
        wc(fc,fc_row,WEEK_START_COL+i,f"=StartDate+7*{i}",
           font=Font(name="Calibri",size=9,color=DARK_GREY),
           fill=fill_calc,align=center,border=box,number_format=FMT_DATE)
    wc(fc,fc_row,TOTAL_COL,"",fill=fill_calc,border=box)
    fc_row += 1

    scenario_pnl_rows = {}
    grid_top = scenario_grid_start[sc["key"]]
    promo_rows_on_inp = {p:grid_top+PROMO_ROW_OFFSET[p] for p in PROMO_TYPES}
    lift_row_on_inp = grid_top + LIFT_OFFSET

    for label_formula, key in PNL_ROWS:
        scenario_pnl_rows[key] = fc_row
        bold = key in ("rev","cogs","promo_total","gp","net","units")
        navy_bold = key in ("gp","net")
        lbl_font = (Font(name="Calibri",size=11,bold=True,color=NAVY)
                    if navy_bold else (f_label if bold else f_calc))
        wc(fc,fc_row,2,label_formula,font=lbl_font,align=left)

        for i in range(WEEKS):
            wcol = get_column_letter(WEEK_START_COL+i)
            if key == "units":
                formula = f"=ROUND(BaseUnits*Inputs!{wcol}{lift_row_on_inp},0)"
            elif key == "price":
                formula = f"=Price*{FC_FX}"
            elif key == "rev":
                formula = f"={wcol}{scenario_pnl_rows['units']}*{wcol}{scenario_pnl_rows['price']}"
            elif key == "cost":
                formula = f"=Cost*{FC_FX}"
            elif key == "cogs":
                formula = f"={wcol}{scenario_pnl_rows['units']}*{wcol}{scenario_pnl_rows['cost']}"
            elif key == "demo":
                formula = f"=Inputs!{wcol}{promo_rows_on_inp['Demo']}*DemoCost*{FC_FX}"
            elif key == "endcap":
                formula = f"=Inputs!{wcol}{promo_rows_on_inp['End Cap']}*EndCapCost*{FC_FX}"
            elif key == "fence":
                formula = f"=Inputs!{wcol}{promo_rows_on_inp['Fence']}*FenceCost*{FC_FX}"
            elif key == "tpd":
                formula = (f"=Inputs!{wcol}{promo_rows_on_inp['TPD']}"
                           f"*{wcol}{scenario_pnl_rows['units']}"
                           f"*{wcol}{scenario_pnl_rows['price']}*TPDPct")
            elif key == "markdown":
                formula = (f"=Inputs!{wcol}{promo_rows_on_inp['Markdown']}"
                           f"*{wcol}{scenario_pnl_rows['units']}"
                           f"*{wcol}{scenario_pnl_rows['price']}*MdPct")
            elif key == "advert":
                formula = f"=Inputs!{wcol}{promo_rows_on_inp['Advert']}*AdvertCost*{FC_FX}"
            elif key == "promo_total":
                formula = (f"=SUM({wcol}{scenario_pnl_rows['demo']}:"
                           f"{wcol}{scenario_pnl_rows['advert']})")
            elif key == "gp":
                formula = f"={wcol}{scenario_pnl_rows['rev']}-{wcol}{scenario_pnl_rows['cogs']}"
            elif key == "net":
                formula = f"={wcol}{scenario_pnl_rows['gp']}-{wcol}{scenario_pnl_rows['promo_total']}"
            else:
                formula = ""
            num_fmt = (FMT_INT if key == "units"
                       else FMT_MONEY_D if key in ("price","cost")
                       else FMT_MONEY)
            wc(fc,fc_row,WEEK_START_COL+i,formula,
               font=f_calc,fill=fill_calc,align=right,border=box,number_format=num_fmt)

        first_w = get_column_letter(WEEK_START_COL)
        last_w = get_column_letter(WEEK_END_COL)
        if key in ("price","cost"):
            tot_formula = f"=Price*{FC_FX}" if key == "price" else f"=Cost*{FC_FX}"
        else:
            tot_formula = f"=SUM({first_w}{fc_row}:{last_w}{fc_row})"
        wc(fc,fc_row,TOTAL_COL,tot_formula,
           font=Font(name="Calibri",size=11,bold=True,color=NAVY),
           fill=fill_kpi,align=right,border=box,
           number_format=(FMT_INT if key == "units"
                          else FMT_MONEY_D if key in ("price","cost")
                          else FMT_MONEY))
        fc_row += 1

    scenario_summary_rows[sc["key"]] = scenario_pnl_rows
    fc_row += 2

fc.freeze_panes = "C5"


# =============================================================================
# SUMMARY
# =============================================================================
sm = wb.create_sheet("Summary")
sm.sheet_view.showGridLines = False
sm.column_dimensions["A"].width = 2
sm.column_dimensions["B"].width = 36
for i in range(len(SCENARIOS)):
    sm.column_dimensions[get_column_letter(3+i)].width = 22

end_col = 2 + len(SCENARIOS)
sm.merge_cells(start_row=2,start_column=2,end_row=2,end_column=end_col)
wc(sm,2,2,
   '="ROI SUMMARY — Bee Propolis Scenario Comparison  (Display: "&$G$3&" "&$C$3&")"',
   font=f_title,fill=fill_title,align=center)
sm.row_dimensions[2].height = 32

SM_CUR, SM_FX, SM_SYM = add_currency_widget(sm, 3, default="GBP")
# Reference to Forecast tab's FX rate so Summary can compensate
FC_FX_FROM_SM = f"Forecast!{FC_FX}"

sm.merge_cells(start_row=4,start_column=2,end_row=4,end_column=end_col)
wc(sm,4,2,
   '="Test window: W1 ("&TEXT(StartDate,"dd-mmm-yyyy")&")  →  W"&Period&" ("&TEXT(StartDate+7*(Period-1),"dd-mmm-yyyy")&")"',
   font=Font(name="Calibri",size=11,italic=True,color=DARK_GREY),
   fill=fill_calc,align=center)

hr = 6
wc(sm,hr,2,"Metric",font=f_white_bold,fill=fill_section,align=left)
for i,sc in enumerate(SCENARIOS):
    wc(sm,hr,3+i,sc["key"],font=f_white_bold,fill=fill_section,align=center)
sm.row_dimensions[hr].height = 24

TC = get_column_letter(TOTAL_COL)
KPIS = [
    ('Units Sold',                              "units",      FMT_INT,    False, False),
    ('="Revenue ("&$G$3&")"',                   "rev",        FMT_MONEY,  False, True),
    ('="Raw COGS ("&$G$3&")"',                  "cogs",       FMT_MONEY,  False, True),
    ('="Gross Profit ("&$G$3&")"',              "gp",         FMT_MONEY,  True,  True),
    ('Gross Margin %',                          "gp_pct",     FMT_PCT,    True,  False),
    ('="Demo Spend ("&$G$3&")"',                "demo",       FMT_MONEY,  False, True),
    ('="End Cap Spend ("&$G$3&")"',             "endcap",     FMT_MONEY,  False, True),
    ('="Fence Spend ("&$G$3&")"',               "fence",      FMT_MONEY,  False, True),
    ('="TPD Spend ("&$G$3&")"',                 "tpd",        FMT_MONEY,  False, True),
    ('="Markdown Spend ("&$G$3&")"',            "markdown",   FMT_MONEY,  False, True),
    ('="Advertising Spend ("&$G$3&")"',         "advert",     FMT_MONEY,  False, True),
    ('="Total Promo Spend ("&$G$3&")"',         "promo_total",FMT_MONEY,  True,  True),
    ('="Net Profit ("&$G$3&")"',                "net",        FMT_MONEY,  True,  True),
    ('Net Margin %',                            "net_pct",    FMT_PCT,    True,  False),
    ('ROI on Promo Spend',                      "roi",        FMT_PCT,    True,  False),
    ('Avg Units / WH / Week',                   "uphpw",      FMT_INT,    False, False),
]

sr = hr + 1
kpi_row_idx = {}
for label,key,fmt,bold,is_money in KPIS:
    kpi_row_idx[key] = sr
    wc(sm,sr,2,label,
       font=Font(name="Calibri",size=11,bold=bold,color=NAVY if bold else "000000"),
       fill=fill_kpi if bold else None,align=left,border=box)
    # Cross-tab FX adjustment factor — applied to money values pulled from Forecast.
    # Forecast values are already in Forecast's currency; multiply by (SM_FX / FC_FX)
    # to re-express in Summary's currency.
    fx_adj = f"*({SM_FX}/{FC_FX_FROM_SM})"
    for i,sc in enumerate(SCENARIOS):
        rows = scenario_summary_rows[sc["key"]]
        if key in rows:
            base = f"Forecast!{TC}{rows[key]}"
            formula = f"={base}{fx_adj}" if is_money else f"={base}"
        elif key == "gp_pct":
            formula = f"=IFERROR(Forecast!{TC}{rows['gp']}/Forecast!{TC}{rows['rev']},0)"
        elif key == "net_pct":
            formula = f"=IFERROR(Forecast!{TC}{rows['net']}/Forecast!{TC}{rows['rev']},0)"
        elif key == "roi":
            formula = f"=IFERROR(Forecast!{TC}{rows['net']}/Forecast!{TC}{rows['promo_total']},0)"
        elif key == "uphpw":
            formula = f"=IFERROR(Forecast!{TC}{rows['units']}/WHs/Period,0)"
        else:
            formula = ""
        wc(sm,sr,3+i,formula,
           font=f_kpi_val if bold else f_calc,
           fill=fill_kpi if bold else None,
           align=right,border=box,number_format=fmt)
    sm.row_dimensions[sr].height = 22 if not bold else 26
    sr += 1

last_col_letter = get_column_letter(2+len(SCENARIOS))
for k in ("net","roi"):
    r = kpi_row_idx[k]
    rng = f"C{r}:{last_col_letter}{r}"
    sm.conditional_formatting.add(rng,
        CellIsRule(operator="lessThan",formula=["0"],
                   fill=PatternFill("solid",fgColor=RED),
                   font=Font(bold=True,color="9C0006")))
    sm.conditional_formatting.add(rng,
        CellIsRule(operator="greaterThanOrEqual",formula=["0"],
                   fill=PatternFill("solid",fgColor=GREEN),
                   font=Font(bold=True,color="006100")))

sr += 2
sm.merge_cells(start_row=sr,start_column=2,end_row=sr,end_column=end_col)
wc(sm,sr,2,"Why each scenario?",font=f_section,fill=fill_section,align=left)
sm.row_dimensions[sr].height = 22
sr += 1
for sc in SCENARIOS:
    wc(sm,sr,2,sc["key"],
       font=Font(name="Calibri",size=11,bold=True,color=NAVY),align=top_wrap)
    sm.merge_cells(start_row=sr,start_column=3,end_row=sr,end_column=end_col)
    wc(sm,sr,3,sc["rationale"],font=f_note,align=top_wrap)
    sm.row_dimensions[sr].height = 38
    sr += 1

sm.freeze_panes = "C7"


# =============================================================================
# PORTFOLIO
# =============================================================================
pf = wb.create_sheet("Portfolio")
pf.sheet_view.showGridLines = False
pf.column_dimensions["A"].width = 2
widths = {"B":5,"C":12,"D":32,"E":12,"F":9,"G":9,"H":9,"I":8,"J":11,"K":13,
          "L":13,"M":13,"N":12,"O":13,"P":8,"Q":8,"R":40}
for col,w in widths.items():
    pf.column_dimensions[col].width = w

pf.merge_cells("B2:R2")
wc(pf,2,2,
   '="PORTFOLIO — D93 Pitch SKUs · Annual Y1 Roll-Up  (Display: "&$G$3&" "&$C$3&")"',
   font=f_title,fill=fill_title,align=center)
pf.row_dimensions[2].height = 32

PF_CUR, PF_FX, PF_SYM = add_currency_widget(pf, 3, default="GBP")

pf.merge_cells("B4:R4")
wc(pf,4,2,
   "SP / Cost / Promo are entered in the currently-selected display currency. "
   "Bee Propolis pre-filled; fill SKUs 2-7 from Asana into the yellow cells. "
   "Bottom row blends across SKUs.",
   font=Font(name="Calibri",size=11,italic=True,color=DARK_GREY),align=top_wrap)
pf.row_dimensions[4].height = 30

HDR_ROW = 6
hdrs = [
    ("B","#"),("C","Code"),("D","SKU"),("E","Pack"),
    ('F','="SP ("&$G$3&")"'),
    ('G','="Cost ("&$G$3&")"'),
    ('H','="GM/u ("&$G$3&")"'),
    ("I","GM %"),
    ("J","Y1 Units"),
    ('K','="Revenue ("&$G$3&")"'),
    ('L','="COGS ("&$G$3&")"'),
    ('M','="GP ("&$G$3&")"'),
    ('N','="Promo ("&$G$3&")"'),
    ('O','="Net Profit ("&$G$3&")"'),
    ("P","Net %"),("Q","ROI %"),
    ("R","Rationale / status"),
]
for col_letter,txt in hdrs:
    al = left if col_letter in ("C","D","E","R") else center
    wc(pf,HDR_ROW,column_index_from_string(col_letter),txt,
       font=f_white_bold,fill=fill_section,align=al,border=box)

sku_start_row = HDR_ROW + 1
for idx,sku in enumerate(PORTFOLIO_SKUS):
    r = sku_start_row + idx
    wc(pf,r,2,sku["rank"],font=f_label,fill=fill_calc,align=center,border=box,number_format=FMT_INT)
    wc(pf,r,3,sku["code"],font=f_label,fill=fill_calc,align=left,border=box)
    wc(pf,r,4,sku["name"],font=f_label,fill=fill_calc,align=left,border=box)
    wc(pf,r,5,sku["pack"],font=f_input,fill=fill_input,align=left,border=box)
    wc(pf,r,6,sku["sp"],  font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_MONEY_D)
    wc(pf,r,7,sku["cost"],font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_MONEY_D)
    wc(pf,r,8,f"=IFERROR(F{r}-G{r},0)",      font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_MONEY_D)
    wc(pf,r,9,f"=IFERROR((F{r}-G{r})/F{r},0)",font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_PCT)
    wc(pf,r,10,sku["units"],font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_INT)
    wc(pf,r,11,f"=IFERROR(J{r}*F{r},0)",font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_MONEY)
    wc(pf,r,12,f"=IFERROR(J{r}*G{r},0)",font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_MONEY)
    wc(pf,r,13,f"=IFERROR(K{r}-L{r},0)",
       font=Font(name="Calibri",size=11,bold=True,color=NAVY),
       fill=fill_kpi,align=right,border=box,number_format=FMT_MONEY)
    wc(pf,r,14,sku["promo"],font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_MONEY)
    wc(pf,r,15,f"=IFERROR(M{r}-N{r},0)",
       font=Font(name="Calibri",size=11,bold=True,color=NAVY),
       fill=fill_kpi,align=right,border=box,number_format=FMT_MONEY)
    wc(pf,r,16,f"=IFERROR(O{r}/K{r},0)",font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_PCT)
    wc(pf,r,17,f"=IFERROR(O{r}/N{r},0)",font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_PCT)
    wc(pf,r,18,sku["rationale"],font=f_input,fill=fill_input,align=left,border=box)
    pf.row_dimensions[r].height = 20

last_sku_row = sku_start_row + len(PORTFOLIO_SKUS) - 1
tr = last_sku_row + 1

for c in range(2,8):
    wc(pf,tr,c,"",fill=fill_title,border=box)
wc(pf,tr,4,"PORTFOLIO TOTAL / BLENDED",
   font=f_white_bold,fill=fill_title,align=left,border=box)
wc(pf,tr,9,f"=IFERROR(SUM(M{sku_start_row}:M{last_sku_row})/SUM(K{sku_start_row}:K{last_sku_row}),0)",
   font=f_white_bold,fill=fill_title,align=right,border=box,number_format=FMT_PCT)
for col_letter,fmt in [("J",FMT_INT),("K",FMT_MONEY),("L",FMT_MONEY),
                       ("M",FMT_MONEY),("N",FMT_MONEY),("O",FMT_MONEY)]:
    wc(pf,tr,column_index_from_string(col_letter),
       f"=SUM({col_letter}{sku_start_row}:{col_letter}{last_sku_row})",
       font=f_white_bold,fill=fill_title,align=right,border=box,number_format=fmt)
wc(pf,tr,16,f"=IFERROR(O{tr}/K{tr},0)",
   font=f_white_bold,fill=fill_title,align=right,border=box,number_format=FMT_PCT)
wc(pf,tr,17,f"=IFERROR(O{tr}/N{tr},0)",
   font=f_white_bold,fill=fill_title,align=right,border=box,number_format=FMT_PCT)
wc(pf,tr,18,"Blended across SKUs",font=f_white_bold,fill=fill_title,align=left,border=box)
pf.row_dimensions[tr].height = 26

pf.freeze_panes = "B7"


out = "/home/user/my-first-project/Costco_UK_D93_Portfolio_ROI_v7.xlsx"
wb.save(out)
print(f"Saved: {out}")
