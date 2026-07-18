#!/usr/bin/env python3
"""
Build Commercial_Control_Tower_Template_v2.xlsx — the operational landing zone
for the Claude Cowork Commercial OS v4.

One workbook per project (one Costco workbook, one FDM workbook, one Sales Ops
workbook). Tabs mirror the Control Tower Workbook Guide:

    Dashboard, Action Tracker, Decision Log, Risk Register, Customer Asks,
    Deck Backlog, Meeting Queue, File Index

Run:  python3 build_control_tower.py
Out:  Commercial_Control_Tower_Template_v2.xlsx  (next to this script)

Reproducible: no timestamps or randomness. Re-running overwrites the file
byte-for-byte identically.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

# ---- palette (calm, print-friendly, readable) -----------------------------
INK = "1F2A37"          # headers text
HEADER_FILL = "1F3A5F"  # deep slate blue
HEADER_TEXT = "FFFFFF"
BAND = "F2F5F9"         # zebra band
ACCENT = "C0491F"       # terracotta accent (dashboard)
GOOD = "1E7F4F"
WARN = "B4690E"
BAD = "B42318"
NOTE_FILL = "FFF8E7"

thin = Side(style="thin", color="D0D7DE")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=HEADER_TEXT, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def band_rows(ws, first_row, last_row, ncols):
    for r in range(first_row, last_row + 1):
        if (r - first_row) % 2 == 1:
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BAND)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_list_validation(ws, col_letter, options, first=2, last=400):
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(options), allow_blank=True)
    dv.error = "Pick a value from the list"
    dv.errorTitle = "Invalid entry"
    dv.prompt = "Choose: " + ", ".join(options)
    ws.add_data_validation(dv)
    dv.add("%s%d:%s%d" % (col_letter, first, col_letter, last))


STATUS = ["Open", "In Progress", "Waiting", "Blocked", "Done", "Dropped"]
PRIORITY = ["P0", "P1", "P2", "P3"]
ATYPE = ["Customer-facing", "Internal"]
SEVERITY = ["High", "Medium", "Low"]
LIKELIHOOD = ["High", "Medium", "Low"]
DECISION_STATUS = ["Decided", "Revisit", "Superseded"]
PROCESSED = ["Yes", "No", "Partial", "Duplicate"]
FILE_STATUS = ["Approved (source of truth)", "Working copy", "Review copy", "Archive", "Superseded"]
YESNO = ["Yes", "No"]

wb = Workbook()

# =====================================================================
# DASHBOARD  (built last in layout, first in tab order)
# =====================================================================
dash = wb.active
dash.title = "Dashboard"
dash.sheet_view.showGridLines = False
set_widths(dash, [3, 34, 14, 30, 14])

dash["B2"] = "Commercial Control Tower"
dash["B2"].font = Font(bold=True, size=20, color=INK)
dash["B3"] = "Claude Cowork Commercial OS v4  ·  one workbook per project"
dash["B3"].font = Font(size=10, italic=True, color="5B6570")
dash["B4"] = "Set the project name in cell D4 →"
dash["B4"].font = Font(size=9, color="8A94A0")
dash["D4"] = "[Project]"
dash["D4"].font = Font(bold=True, size=12, color=ACCENT)
dash["D4"].fill = PatternFill("solid", fgColor=NOTE_FILL)
dash["D4"].alignment = Alignment(horizontal="center")
dash["D4"].border = BORDER

# KPI tiles: label / formula
kpis = [
    ("Open actions", "=COUNTIFS('Action Tracker'!F:F,\"<>Done\",'Action Tracker'!F:F,\"<>Dropped\",'Action Tracker'!B:B,\"<>\")-1"),
    ("Overdue actions", "=COUNTIFS('Action Tracker'!G:G,\"<\"&TODAY(),'Action Tracker'!F:F,\"<>Done\",'Action Tracker'!F:F,\"<>Dropped\")"),
    ("P0 / P1 open", "=COUNTIFS('Action Tracker'!E:E,\"P0\",'Action Tracker'!F:F,\"<>Done\")+COUNTIFS('Action Tracker'!E:E,\"P1\",'Action Tracker'!F:F,\"<>Done\")"),
    ("Actions missing owner", "=COUNTIFS('Action Tracker'!C:C,\"\",'Action Tracker'!B:B,\"<>\")-1"),
    ("Actions missing due date", "=COUNTIFS('Action Tracker'!G:G,\"\",'Action Tracker'!B:B,\"<>\")-1"),
    ("Open customer asks", "=COUNTIFS('Customer Asks'!G:G,\"<>Done\",'Customer Asks'!G:G,\"<>Dropped\",'Customer Asks'!B:B,\"<>\")-1"),
    ("Live risks (High)", "=COUNTIFS('Risk Register'!C:C,\"High\",'Risk Register'!G:G,\"<>Done\")"),
    ("Open decisions to revisit", "=COUNTIF('Decision Log'!G:G,\"Revisit\")"),
    ("Deck backlog open", "=COUNTIFS('Deck Backlog'!G:G,\"<>Done\",'Deck Backlog'!B:B,\"<>\")-1"),
    ("Meetings unprocessed", "=COUNTIF('Meeting Queue'!E:E,\"No\")+COUNTIF('Meeting Queue'!E:E,\"Partial\")"),
]
r = 6
dash.cell(row=r, column=2, value="Health at a glance").font = Font(bold=True, size=12, color=ACCENT)
r += 1
for label, formula in kpis:
    lc = dash.cell(row=r, column=2, value=label)
    lc.font = Font(size=11, color=INK)
    lc.border = BORDER
    lc.fill = PatternFill("solid", fgColor=BAND)
    vc = dash.cell(row=r, column=3, value=formula)
    vc.font = Font(bold=True, size=12, color=INK)
    vc.alignment = Alignment(horizontal="center")
    vc.border = BORDER
    r += 1

# highlight the two that should scream if non-zero
dash.conditional_formatting.add(
    "C8:C16",
    CellIsRule(operator="greaterThan", formula=["0"], font=Font(color=BAD, bold=True)),
)
# overdue tile (row 8 is 'Overdue actions' -> C8) already covered

dash.cell(row=r + 1, column=2, value="The rule that matters").font = Font(bold=True, color=ACCENT)
dash.cell(row=r + 2, column=2,
          value="No summary is finished until the parts that matter have landed here.").font = Font(italic=True, color="5B6570")
dash.cell(row=r + 4, column=2, value="Tabs: Action Tracker · Decision Log · Risk Register · Customer Asks · Deck Backlog · Meeting Queue · File Index").font = Font(size=9, color="8A94A0")
dash.cell(row=r + 5, column=2, value="Update in review-copy mode by default. Never mark Done without evidence.").font = Font(size=9, color="8A94A0")

# =====================================================================
def make_sheet(title, headers, widths, examples, validations, band_to=60):
    ws = wb.create_sheet(title)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))
    set_widths(ws, widths)
    for ri, row in enumerate(examples, start=2):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
    band_rows(ws, 2, band_to, len(headers))
    for col_letter, options in validations:
        add_list_validation(ws, col_letter, options)
    return ws


# ---- Action Tracker -------------------------------------------------
at = make_sheet(
    "Action Tracker",
    ["ID", "Action", "Owner", "Priority", "Type", "Status", "Due Date",
     "Created", "Source Meeting", "Notes / Evidence"],
    [8, 40, 16, 10, 16, 13, 13, 13, 24, 34],
    [
        ["A-001", "Send Costco buyer the revised pack architecture options", "Louis", "P1",
         "Customer-facing", "In Progress", "=TODAY()+3", "=TODAY()-1",
         "Costco JBP prep", "Buyer asked for 2 vs 3 pack options with landed cost"],
        ["A-002", "Confirm item setup timing with supply planning", "", "P0",
         "Internal", "Open", "=TODAY()+1", "=TODAY()-1",
         "Costco JBP prep", "MISSING OWNER — assign before it slips"],
        ["A-003", "Update forecast file with committed promo weeks", "Teresa", "P2",
         "Internal", "Waiting", "=TODAY()-2", "=TODAY()-5",
         "Weekly sync", "OVERDUE example — waiting on promo calendar sign-off"],
    ],
    [("D", PRIORITY), ("E", ATYPE), ("F", STATUS)],
)
# overdue due dates in red
at.conditional_formatting.add(
    "G2:G400",
    FormulaRule(formula=["AND($G2<>\"\",$G2<TODAY(),$F2<>\"Done\",$F2<>\"Dropped\")"],
                fill=PatternFill("solid", fgColor="FCE8E6"), font=Font(color=BAD, bold=True)),
)
# P0 rows accent
at.conditional_formatting.add(
    "A2:J400",
    FormulaRule(formula=["$D2=\"P0\""], font=Font(color=BAD, bold=True)),
)
for row in ("2", "3", "4"):
    at["G" + row].number_format = "yyyy-mm-dd"
    at["H" + row].number_format = "yyyy-mm-dd"

# ---- Decision Log ---------------------------------------------------
make_sheet(
    "Decision Log",
    ["ID", "Decision", "Date", "Owner", "Rationale", "Source", "Status", "Notes"],
    [8, 40, 13, 16, 34, 22, 12, 30],
    [
        ["D-001", "Lead with 3-pack architecture for Costco line review", "=TODAY()-1",
         "Aaron", "Better landed cost story and cleaner shelf block", "Costco JBP prep",
         "Decided", "Revisit if buyer pushes on price gap"],
    ],
    [("G", DECISION_STATUS)],
)

# ---- Risk Register --------------------------------------------------
rr = make_sheet(
    "Risk Register",
    ["ID", "Risk", "Severity", "Likelihood", "Mitigation", "Owner", "Status",
     "Review Date", "Notes"],
    [8, 36, 11, 12, 34, 14, 12, 13, 26],
    [
        ["R-001", "Item setup may miss the Costco reset window", "High", "Medium",
         "Lock setup timing with supply planning this week", "Louis", "Open",
         "=TODAY()+7", "Tied to A-002"],
    ],
    [("C", SEVERITY), ("D", LIKELIHOOD), ("G", STATUS)],
)
rr.conditional_formatting.add(
    "C2:C400",
    CellIsRule(operator="equal", formula=['"High"'],
               fill=PatternFill("solid", fgColor="FCE8E6"), font=Font(color=BAD, bold=True)),
)
rr["H2"].number_format = "yyyy-mm-dd"

# ---- Customer Asks --------------------------------------------------
make_sheet(
    "Customer Asks",
    ["ID", "Ask", "Requested By", "Date", "Internal Owner", "Internal Action Needed",
     "Status", "Due", "Notes"],
    [8, 36, 16, 13, 16, 30, 12, 13, 24],
    [
        ["C-001", "Two vs three pack options with landed cost", "Costco buyer",
         "=TODAY()-1", "Louis", "Build costed pack-architecture comparison", "Open",
         "=TODAY()+3", "Feeds A-001"],
    ],
    [("G", STATUS)],
)

# ---- Deck Backlog ---------------------------------------------------
make_sheet(
    "Deck Backlog",
    ["ID", "Slide / Section", "Current State", "Needed Change", "Why It Matters",
     "Priority", "Status", "Notes"],
    [8, 22, 30, 30, 26, 10, 12, 22],
    [
        ["K-001", "Pack architecture", "Shows 2-pack only", "Add 3-pack option + landed cost",
         "Buyer asked; changes the margin story", "P1", "Open", "After A-001 lands"],
    ],
    [("F", PRIORITY), ("G", STATUS)],
)

# ---- Meeting Queue --------------------------------------------------
make_sheet(
    "Meeting Queue",
    ["ID", "Meeting", "Date", "Artifacts", "Processed?", "Summary File", "Notes"],
    [8, 26, 13, 26, 13, 30, 26],
    [
        ["M-001", "Costco JBP prep", "=TODAY()-1", "Transcript + notes + agenda", "Yes",
         "YYYY-MM-DD - Costco - Meeting Summary - JBP prep", "One queue entry per meeting"],
        ["M-002", "Weekly internal sync", "=TODAY()", "Recording", "No",
         "", "Awaiting Meeting Inbox Processor"],
    ],
    [("E", PROCESSED)],
)

# ---- File Index -----------------------------------------------------
make_sheet(
    "File Index",
    ["Type", "File Name", "Status", "Location", "Last Updated", "Notes"],
    [20, 40, 24, 24, 13, 26],
    [
        ["Deck", "Costco Line Review vX.pptx", "Approved (source of truth)", "/04 Decks",
         "=TODAY()-3", "The one Claude should read first"],
        ["Deck", "REVIEW - Costco Line Review.pptx", "Review copy", "/04 Decks",
         "=TODAY()-1", "Working edits only"],
        ["Workbook", "Commercial_Control_Tower_Template_v2.xlsx", "Approved (source of truth)",
         "/02 Control Tower", "=TODAY()", "This file"],
        ["Forecast", "Costco Forecast vX.xlsx", "Approved (source of truth)", "/05 Workbooks",
         "=TODAY()-4", ""],
    ],
    [("C", FILE_STATUS)],
)

# tab colors
colors = {
    "Dashboard": ACCENT, "Action Tracker": HEADER_FILL, "Decision Log": "3B6E8F",
    "Risk Register": BAD, "Customer Asks": GOOD, "Deck Backlog": WARN,
    "Meeting Queue": "5B6570", "File Index": "1F2A37",
}
for name, col in colors.items():
    wb[name].sheet_properties.tabColor = col

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Commercial_Control_Tower_Template_v2.xlsx")
wb.save(out)
print("wrote", out)
