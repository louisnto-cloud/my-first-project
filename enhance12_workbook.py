#!/usr/bin/env python3
"""Loop batch I13 + I15:
 I13 Risk likelihood×impact heatmap appended below the Risk Register table (fresh grid, no row inserts).
 I15 Assumptions Audit tab — every Financial Model input with value(s) + source/confidence."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; STEEL="3E5C76"; LIGHT="EAF1F4"; AMBER="FFF2CC"; AMBERHEAD="B45309"; GREEN="E2EFDA"; GREENHEAD="1E7D32"; RED="F8D7DA"; REDH="9C2A2A"; WHITE="FFFFFF"; GREY="595959"; GOLD="C9A227"; LINKBLUE="1155CC"
def F(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CTR=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

# ---------- I13: risk heatmap on the Risk Register tab ----------
rr=wb["22 Risk Register"]
base=rr.max_row+3
c=rr.cell(row=base,column=1,value="LIKELIHOOD × IMPACT HEATMAP (risk #s placed by rating)")
rr.merge_cells(start_row=base,start_column=1,end_row=base,end_column=6)
c.font=F(bold=True,color=WHITE,size=10); c.fill=fill(STEEL); c.alignment=LEFT
# grid: rows = Impact High/Med/Low (col A label, cols B/C/D = Likelihood Low/Med/High)
cells={("High","High"):"1, 2", ("High","Med"):"3, 4, 5", ("High","Low"):"—",
       ("Med","High"):"—", ("Med","Med"):"6,7,8,9,10", ("Med","Low"):"12",
       ("Low","High"):"—", ("Low","Med"):"11", ("Low","Low"):"—"}
def colr(imp,lik):
    score={"Low":1,"Med":2,"High":3}[imp]+{"Low":1,"Med":2,"High":3}[lik]
    return RED if score>=5 else (AMBER if score>=4 else GREEN)
# header row
hr=base+1
rr.cell(row=hr,column=1,value="Impact ↓ / Likelihood →").font=F(bold=True,size=9)
for j,lik in enumerate(["Low","Med","High"]):
    x=rr.cell(row=hr,column=2+j,value=lik); x.font=F(bold=True,color=WHITE,size=9); x.fill=fill(NAVY); x.alignment=CTR; x.border=BORDER
for i,imp in enumerate(["High","Med","Low"]):
    row=hr+1+i
    lab=rr.cell(row=row,column=1,value=imp); lab.font=F(bold=True,color=WHITE,size=9); lab.fill=fill(NAVY); lab.alignment=CTR; lab.border=BORDER
    for j,lik in enumerate(["Low","Med","High"]):
        cell=rr.cell(row=row,column=2+j,value=cells[(imp,lik)])
        cell.fill=fill(colr(imp,lik)); cell.alignment=CTR; cell.border=BORDER; cell.font=F(bold=True,size=10)
        rr.row_dimensions[row].height=24
leg=rr.cell(row=hr+4,column=1,value="Red = act now · Amber = manage · Green = monitor. Numbers refer to the risk # in the table above.")
rr.merge_cells(start_row=hr+4,start_column=1,end_row=hr+4,end_column=6); leg.font=F(italic=True,size=8,color=GREY)

# ---------- I15: Assumptions Audit tab ----------
if "Assumptions Audit" in wb.sheetnames: del wb["Assumptions Audit"]
aa=wb.create_sheet("Assumptions Audit")
aa.sheet_view.showGridLines=False
aa.merge_cells("A1:E1"); t=aa["A1"]; t.value="Assumptions Audit — every model input, with source & confidence"; t.font=F(bold=True,color=WHITE,size=14); t.fill=fill(NAVY); t.alignment=CTR; aa.row_dimensions[1].height=26
hdr=["Input","Conservative","Base","Aggressive","Source / confidence"]
for j,h in enumerate(hdr):
    x=aa.cell(row=3,column=1+j,value=h); x.font=F(bold=True,color=WHITE,size=10); x.fill=fill(STEEL); x.alignment=CTR; x.border=BORDER
rows=[
 ("List price / can (CAD)","$3.29","$3.49","$3.79","MÜV SRP ~$14.99/pack; per-can set vs ~$3.49 CAD premium shelf norm ◐"),
 ("COGS / can (CAD)","$1.20","$1.05","$0.95","Industry rule-of-thumb <$1 target at scale; excludes freight/duty ⚠"),
 ("Retailer margin %","38%","35%","33%","Grocery bev 30–40% rule-of-thumb ◐"),
 ("Distributor margin %","18%","15%","12%","UNFI/KeHE ~13–25% range ◐"),
 ("Trade spend % of net","25%","20%","15%","Emerging-brand 10–20%+ ◐"),
 ("Doors Y1 / Y2 / Y3","300/1.2k/3.5k","500/2k/6k","900/3.5k/11k","Illustrative beachhead→scale ramp ⚠"),
 ("Velocity (u/store/wk)","2 / 3 / 3.5","3 / 4 / 5","4.5 / 6 / 7.5","Illustrative; buyers reward velocity ⚠"),
 ("Marketing / yr","$150k","$300k","$600k","Illustrative challenger budget ⚠"),
 ("Fixed overhead / yr","$250k","$350k","$500k","Illustrative ⚠"),
 ("Slotting (Y1 one-time)","$50k","$100k","$200k","Slotting $250–$250k/SKU range ◐"),
]
r=4
for i,row in enumerate(rows):
    for j,v in enumerate(row):
        x=aa.cell(row=r,column=1+j,value=v); x.font=F(size=10,bold=(j==0)); x.alignment=LEFT if j in(0,4) else CTR; x.border=BORDER
        if i%2: x.fill=fill(LIGHT)
    aa.row_dimensions[r].height=26; r+=1
aa.column_dimensions["A"].width=24
for col in "BCD": aa.column_dimensions[col].width=14
aa.column_dimensions["E"].width=52
note=aa.cell(row=r+1,column=1,value="All financial inputs are ILLUSTRATIVE planning scaffolding, not a forecast. Replace each with real co-packer, distributor and buyer numbers. ◐ rule-of-thumb · ⚠ estimate/illustrative.")
aa.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=5); note.font=F(italic=True,size=9,color=GREY); note.fill=fill(AMBER); note.alignment=LEFT; note.border=BORDER; aa.row_dimensions[r+1].height=30
h=aa.cell(row=1,column=13,value="🏠 Home"); h.hyperlink="#'🏠 Start Here'!A1"; h.font=F(bold=True,color=LINKBLUE,underline="single",size=8); h.alignment=Alignment(horizontal="right")
aa.sheet_properties.tabColor=GOLD
aa.page_setup.orientation="landscape"; aa.page_setup.fitToWidth=1; aa.page_setup.fitToHeight=0; aa.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
# place after Financial Model; add to contents
s=wb["Assumptions Audit"]; wb._sheets.remove(s); idx=wb.sheetnames.index("Financial Model")+1; wb._sheets.insert(idx,s)
toc=wb["01 Contents"]; row=toc.max_row+1
a=toc.cell(row=row,column=1,value="Assumptions Audit"); a.hyperlink="#'Assumptions Audit'!A1"; a.font=F(bold=True,color=LINKBLUE,underline="single",size=10); a.fill=fill(GOLD); a.border=BORDER; a.alignment=LEFT
b=toc.cell(row=row,column=2,value="Every model input with its source & confidence tag"); b.font=F(size=10); b.border=BORDER; b.alignment=LEFT

wb.save(PATH)
print("I13 heatmap + I15 Assumptions Audit added. sheets:",len(wb.sheetnames))
