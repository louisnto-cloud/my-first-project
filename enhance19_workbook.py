#!/usr/bin/env python3
"""Loop batch I31: amber-tag FLAG / UNVERIFIED / non-existent brand cells in the Extended Landscape
for scanability (fill only; no restructuring)."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
AMBER="FFF2CC"; AMBERHEAD="B45309"
def fill(c): return PatternFill("solid",fgColor=c)
ws=wb["08 Extended Landscape"]
tagged=0
CUES=("FLAG","UNVERIFIED","LIKELY DOES NOT EXIST","DOES NOT EXIST","AMBIGUOUS","could not verify","Undisclosed (FLAG)")
for row in ws.iter_rows():
    for c in row:
        if isinstance(c.value,str) and any(cue.lower() in c.value.lower() for cue in CUES):
            c.fill=fill(AMBER)
            try: c.font=Font(name="Calibri",size=c.font.size or 10,color=AMBERHEAD,bold=True)
            except Exception: pass
            tagged+=1
# legend note at bottom
r=ws.max_row+2
ws.cell(row=r,column=1,value=f"Amber cells = FLAG / unverified / could-not-confirm ({tagged} tagged). Re-verify before external use.").font=Font(name="Calibri",italic=True,size=9,color=AMBERHEAD)
wb.save(PATH)
print(f"I31 tagged {tagged} FLAG/unverified cells in Extended Landscape.")
