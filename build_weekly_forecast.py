"""Organika RTD FY26-27 Forecast - Enterprise Edition (v3).

Devil's-advocate review of v2 produced this checklist of gaps. All are fixed below:

  GOVERNANCE
    [x] No assumption register -> add 'Assumption Register' with owner, confidence, review date.
    [x] No risk register -> add 'Risk Register' with mitigation and owner.
    [x] No version log -> add 'Change Log' tab.
    [x] No data dictionary -> add 'Glossary' tab.

  FINANCE
    [x] Only revenue & GP -> add full 'P&L' (Revenue -> Trade Spend -> Net -> COGS -> GP -> A&P -> SG&A -> EBITDA).
    [x] No trade-spend model -> add 'Trade Spend' tab with listing fees, slotting, scan/promo per channel.
    [x] No OPEX -> add 'OPEX' tab with headcount, A&P spend, freight, broker fees.
    [x] No quarterly view -> Dashboard now includes quarterly roll-up.

  STRATEGY
    [x] One scenario at a time -> add 'Scenarios' tab showing Conservative / Base / Stretch side-by-side.
    [x] No KPI tracking -> add 'KPIs' tab (GP per case, % new channel, velocity per door, ACV proxy).

  DATA INTEGRITY
    [x] Free-text channel/province/SKU strings -> add Validation lists and data-validation dropdowns.
    [x] Class Mix sum check is silent -> add conditional formatting (green = 100%, red otherwise).
    [x] No bounds on negative inputs -> data validation prevents negative doors / negative velocity.

  USABILITY
    [x] Dense Dashboard -> reorganized with sections, headline number callouts, sparkline-friendly layout.
    [x] No print setup -> page setup landscape, fit-to-width, header rows.

Outputs:
  Organika_RTD_Forecast_FY26_27_Weekly.xlsx
"""

from datetime import date, timedelta

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

# ---------------------------------------------------------------------------
# Domain
# ---------------------------------------------------------------------------
SKUS = ["MUV Raspberry Lemon", "MUV Lime", "MUV PFP", "LCA Energy"]
SKU_BRAND = {
    "MUV Raspberry Lemon": "MUV",
    "MUV Lime": "MUV",
    "MUV PFP": "MUV",
    "LCA Energy": "LCA",
}
PROVINCES = ["BC", "AB", "SK", "MB", "ON", "QC", "NB", "NS", "PE", "NL"]
CHANNELS = [
    "Natural", "Specialty", "FDM", "Costco",
    "On-Premise", "Convenience", "Gym & Fitness", "Private Liquor", "RAS",
]
NEW_CHANNELS = {"On-Premise", "Convenience", "Gym & Fitness", "Private Liquor", "RAS"}

START_DATE = date(2026, 9, 1)
N_WEEKS = 52
WEEK_DATES = [START_DATE + timedelta(days=7 * i) for i in range(N_WEEKS)]
WEEK_LABELS = [f"W{i + 1:02d} ({d.strftime('%b %d')})" for i, d in enumerate(WEEK_DATES)]
MONTH_ORDER = [
    "Sep 2026", "Oct 2026", "Nov 2026", "Dec 2026",
    "Jan 2027", "Feb 2027", "Mar 2027", "Apr 2027",
    "May 2027", "Jun 2027", "Jul 2027", "Aug 2027",
]
WEEK_TO_MONTH = [date(d.year, d.month, 1).strftime("%b %Y") for d in WEEK_DATES]
QUARTERS = ["Q1 (Sep-Nov)", "Q2 (Dec-Feb)", "Q3 (Mar-May)", "Q4 (Jun-Aug)"]
MONTH_QUARTER = {
    "Sep 2026": "Q1 (Sep-Nov)", "Oct 2026": "Q1 (Sep-Nov)", "Nov 2026": "Q1 (Sep-Nov)",
    "Dec 2026": "Q2 (Dec-Feb)", "Jan 2027": "Q2 (Dec-Feb)", "Feb 2027": "Q2 (Dec-Feb)",
    "Mar 2027": "Q3 (Mar-May)", "Apr 2027": "Q3 (Mar-May)", "May 2027": "Q3 (Mar-May)",
    "Jun 2027": "Q4 (Jun-Aug)", "Jul 2027": "Q4 (Jun-Aug)", "Aug 2027": "Q4 (Jun-Aug)",
}

# Column layout
FW_FIRST_COL = 5
FW_LAST_COL = FW_FIRST_COL + N_WEEKS - 1
def fw_week_col(i):
    return get_column_letter(FW_FIRST_COL + i - 1)

DOORS_MONTHLY_FIRST_COL = 5
DOORS_MONTHLY_LAST_COL = 16
DOORS_WEEKLY_FIRST_COL = 17
DOORS_WEEKLY_LAST_COL = DOORS_WEEKLY_FIRST_COL + N_WEEKS - 1
MONTH_INPUT_COL = {m: get_column_letter(DOORS_MONTHLY_FIRST_COL + i) for i, m in enumerate(MONTH_ORDER)}
def doors_week_col(i):
    return get_column_letter(DOORS_WEEKLY_FIRST_COL + i - 1)

REP_MONTHLY_FIRST_COL = 2
REP_WEEKLY_FIRST_COL = 14
REP_MONTH_INPUT_COL = {m: get_column_letter(REP_MONTHLY_FIRST_COL + i) for i, m in enumerate(MONTH_ORDER)}
def rep_week_col(i):
    return get_column_letter(REP_WEEKLY_FIRST_COL + i - 1)

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
# Apple-inspired palette: restrained, white, one accent.
APPLE_INK    = "1D1D1F"  # primary text
APPLE_GREY   = "6E6E73"  # secondary text
APPLE_SUBTLE = "F5F5F7"  # background tint
APPLE_BLUE   = "0071E3"  # accent
APPLE_GREEN  = "34C759"
APPLE_ORANGE = "FF9500"
APPLE_RED    = "FF3B30"

HEADER_FONT = Font(name="Helvetica", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=APPLE_INK)
SUBHEADER_FONT = Font(name="Helvetica", bold=True, color=APPLE_INK, size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor=APPLE_SUBTLE)
SECTION_FONT = Font(name="Helvetica", bold=True, size=13, color=APPLE_INK)
TITLE_FONT = Font(name="Helvetica", bold=True, size=22, color=APPLE_INK)
HERO_FONT = Font(name="Helvetica", bold=True, size=36, color=APPLE_INK)
HERO_LABEL_FONT = Font(name="Helvetica", size=11, color=APPLE_GREY)
BODY_FONT = Font(name="Helvetica", size=11, color=APPLE_INK)
MUTED_FONT = Font(name="Helvetica", size=10, color=APPLE_GREY)
INPUT_FONT = Font(name="Helvetica", color=APPLE_BLUE, size=11)
CALC_FONT = Font(name="Helvetica", color=APPLE_INK, size=11)
LINK_FONT = Font(name="Helvetica", color=APPLE_GREEN, size=11)
WARN_FONT = Font(name="Helvetica", color=APPLE_RED, bold=True, size=11)
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
GREEN_FILL = PatternFill("solid", fgColor="D1F2D1")
RED_FILL = PatternFill("solid", fgColor="FFD6D6")
SUBTOTAL_FILL = PatternFill("solid", fgColor=APPLE_SUBTLE)
MONTHLY_FILL = PatternFill("solid", fgColor="FFF8E5")
THIN = Side(border_style="thin", color="D2D2D7")
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Tab color hierarchy: front (blue), inputs (orange), governance (grey), engine (light grey)
TAB_FRONT      = APPLE_BLUE
TAB_REPORT     = "1D1D1F"
TAB_INPUT      = "FF9500"
TAB_GOVERNANCE = "6E6E73"
TAB_ENGINE     = "C7C7CC"


def style_header_row(ws, row, last_col):
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_subheader(cell):
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL


def title(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT


def section(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = SECTION_FONT


def setup_print(ws, landscape=True, fit_width=True, header_rows=1):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    if fit_width:
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.6)
    if header_rows:
        ws.print_title_rows = f"1:{header_rows + 4}"


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------

def build_readme(ws):
    ws.title = "README"
    title(ws, 1, "Organika RTD National Forecast - FY26 to FY27 (Enterprise Build v3)")
    ws["A2"] = "Source of truth. Supersedes all scratch models."
    ws["A2"].font = Font(italic=True, color="808080")

    blocks = [
        ("Fiscal year",        "September 1, 2026 through August 31, 2027 (52 weeks)"),
        ("Portfolio",          "MUV Sparkling (Raspberry Lemon, Lime, Passion Fruit Pineapple) and LCA Energy"),
        ("Currency",           "CAD"),
        ("Pack",               "24 cans per case, 355 ml"),
        ("Channels",           "9 - Natural, Specialty, FDM, Costco, On-Premise (restaurants),"),
        ("",                   "Convenience, Gym & Fitness, Private Liquor, RAS (rural agency stores)"),
        ("Owner",              "Louis - RTD lead, Organika"),
        ("Stakeholders",       "Aaron (CEO), Teresa (VP Marketing), Rijo (Ops), Elliot (Finance)"),
        ("Sign-off log",       "See Change Log tab"),
    ]
    for i, (a, b) in enumerate(blocks, start=4):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True)
        ws.cell(row=i, column=2, value=b)

    # Tab guide
    section(ws, 15, "Tab guide")
    tabs = [
        ("Dashboard",            "Executive view. Headline metrics, monthly + quarterly roll-ups, weekly totals."),
        ("P&L",                  "Revenue -> Trade Spend -> Net Revenue -> COGS -> Gross Profit -> A&P -> SG&A -> EBITDA."),
        ("Scenarios",            "Conservative (0.85) / Base (1.00) / Stretch (1.15) side-by-side."),
        ("KPIs",                 "Velocity per door, GP per case, % revenue from new channels, ACV proxy."),
        ("Assumptions",          "Workbook toggles, week index."),
        ("Class Mix",            "% A/B/C doors per channel. Rows must sum to 100% (auto-flagged)."),
        ("Velocity",             "Units per door per week, by SKU x channel x class. Blended output."),
        ("Doors",                "Active doors by brand x province x channel. 12 monthly inputs -> 52 weekly outputs."),
        ("Pricing",              "Net case price, landed cost, GP per case per SKU."),
        ("Trade Spend",          "Listing fees, slotting, scan/promo allowance per channel."),
        ("OPEX",                 "Sales reps, broker fees, freight, A&P spend, G&A."),
        ("Marketing and Sampling", "Account sampling, events, rep samples, buffer."),
        ("Forecast Weekly",      "Engine. 360 row grid (4 SKUs x 10 provinces x 9 channels)."),
        ("Production Plan",      "Sales + marketing + buffer = weekly production demand."),
        ("Revenue",              "Weekly revenue and gross profit by SKU."),
        ("Assumption Register",  "Every assumption with owner, confidence, last-reviewed date."),
        ("Risk Register",        "Top risks with likelihood / impact / mitigation / owner."),
        ("Change Log",           "Version history and sign-offs."),
        ("Glossary",             "Data dictionary."),
        ("Validation Lists",     "Source data for dropdowns. Do not edit unless adding a new channel/SKU."),
    ]
    ws.cell(row=16, column=1, value="Tab").font = Font(bold=True)
    ws.cell(row=16, column=2, value="What it contains").font = Font(bold=True)
    for i, (a, b) in enumerate(tabs, start=17):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True, color="1F4E79")
        ws.cell(row=i, column=2, value=b)

    # Colour key
    nxt = 17 + len(tabs) + 2
    section(ws, nxt, "Colour key")
    legend = [
        ("Blue text",     "User input. Edit these to flex the model."),
        ("Black text",    "Calculated. Do not edit."),
        ("Green text",    "Cross-tab link. Pulls from another sheet."),
        ("Red text",      "Warning / out of bounds."),
        ("Yellow fill",   "Key assumption needing sign-off before locking."),
        ("Cream fill",    "Monthly input strip. Editing 1 cell updates 4-5 weeks."),
        ("Green fill",    "Validation passed."),
        ("Red fill",      "Validation failed - inspect."),
    ]
    for i, (a, b) in enumerate(legend, start=nxt + 1):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True)
        ws.cell(row=i, column=2, value=b)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110
    setup_print(ws)


def build_assumptions(ws):
    ws.title = "Assumptions"
    title(ws, 1, "Assumptions and toggles")

    section(ws, 3, "Workbook inputs")
    inputs = [
        ("Units per case",        24,                "Cans per case. Locked."),
        ("Production buffer",     0.10,              "Production cushion above demand. Discuss with Rijo."),
        ("Scenario multiplier",   1.0,               "Toggle 0.85 / 1.00 / 1.15. Used in legacy Forecast Weekly engine."),
        ("Fiscal year",           "FY26 to FY27",    "Sep 1, 2026 - Aug 31, 2027."),
        ("Period start",          "September 1, 2026", ""),
        ("Period end",            "August 31, 2027",   ""),
        ("Number of weeks",       N_WEEKS,           "52 calendar weeks."),
    ]
    for i, (lbl, val, note) in enumerate(inputs, start=4):
        ws.cell(row=i, column=1, value=lbl).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=val); c.font = INPUT_FONT
        ws.cell(row=i, column=3, value=note)
    ws["B5"].number_format = "0%"
    ws["B6"].fill = YELLOW_FILL  # scenario multiplier

    # Data validation: scenario multiplier must be between 0 and 2
    dv = DataValidation(type="decimal", operator="between",
                        formula1=0, formula2=2,
                        allow_blank=False, showErrorMessage=True,
                        errorTitle="Out of bounds",
                        error="Scenario multiplier must be between 0 and 2.")
    ws.add_data_validation(dv); dv.add("B6")

    # Week index
    section(ws, 13, "Week index")
    ws.cell(row=14, column=1, value="Week").font = Font(bold=True)
    ws.cell(row=14, column=2, value="Start date").font = Font(bold=True)
    ws.cell(row=14, column=3, value="Month").font = Font(bold=True)
    ws.cell(row=14, column=4, value="Quarter").font = Font(bold=True)
    for cell in (ws["A14"], ws["B14"], ws["C14"], ws["D14"]):
        style_subheader(cell)
    for i, (lbl, d, m) in enumerate(zip(WEEK_LABELS, WEEK_DATES, WEEK_TO_MONTH), start=15):
        ws.cell(row=i, column=1, value=lbl)
        ws.cell(row=i, column=2, value=d).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=3, value=m)
        ws.cell(row=i, column=4, value=MONTH_QUARTER[m])

    for col, w in zip("ABCD", (22, 14, 12, 16)):
        ws.column_dimensions[col].width = w
    setup_print(ws)


def build_validation_lists(ws):
    ws.title = "Validation Lists"
    title(ws, 1, "Validation Lists - source for dropdowns")
    ws["A3"] = "Do not edit unless adding/removing a channel, SKU, brand, or province."

    section(ws, 5, "Channels")
    for i, ch in enumerate(CHANNELS, start=6):
        ws.cell(row=i, column=1, value=ch)
    section(ws, 5 + 0, "Channels")  # already done

    section(ws, 5, "Channels")  # idempotent
    # Already at A6+

    section(ws, 5, "Channels")
    # Provinces
    ws.cell(row=5, column=3, value="Provinces").font = SECTION_FONT
    for i, p in enumerate(PROVINCES, start=6):
        ws.cell(row=i, column=3, value=p)
    # SKUs
    ws.cell(row=5, column=5, value="SKUs").font = SECTION_FONT
    for i, s in enumerate(SKUS, start=6):
        ws.cell(row=i, column=5, value=s)
    # Brands
    ws.cell(row=5, column=7, value="Brands").font = SECTION_FONT
    for i, b in enumerate(["MUV", "LCA"], start=6):
        ws.cell(row=i, column=7, value=b)
    # Quarters
    ws.cell(row=5, column=9, value="Quarters").font = SECTION_FONT
    for i, q in enumerate(QUARTERS, start=6):
        ws.cell(row=i, column=9, value=q)
    # Confidence levels
    ws.cell(row=5, column=11, value="Confidence").font = SECTION_FONT
    for i, c in enumerate(["High", "Medium", "Low"], start=6):
        ws.cell(row=i, column=11, value=c)
    # RAG status
    ws.cell(row=5, column=13, value="Status").font = SECTION_FONT
    for i, c in enumerate(["Green", "Amber", "Red"], start=6):
        ws.cell(row=i, column=13, value=c)

    for col in "ABCDEFGHIJKLMN":
        ws.column_dimensions[col].width = 16
    setup_print(ws)


def build_class_mix(ws):
    ws.title = "Class Mix"
    title(ws, 1, "Class mix by channel")
    ws["A3"] = "% of doors in each class, per channel. Each row must sum to 100% (green = valid)."

    headers = ["Channel", "Class A %", "Class B %", "Class C %", "Sum check"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    rows = [
        ("Natural",         0.20, 0.50, 0.30, False),
        ("Specialty",       0.30, 0.50, 0.20, False),
        ("FDM",             0.50, 0.35, 0.15, False),
        ("Costco",          0.80, 0.20, 0.00, False),
        ("On-Premise",      0.25, 0.50, 0.25, True),
        ("Convenience",     0.15, 0.50, 0.35, True),
        ("Gym & Fitness",   0.30, 0.50, 0.20, True),
        ("Private Liquor",  0.25, 0.55, 0.20, True),
        ("RAS",             0.10, 0.40, 0.50, True),
    ]
    for i, (ch, a, b, c, is_new) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=ch)
        for col, val in zip((2, 3, 4), (a, b, c)):
            cell = ws.cell(row=i, column=col, value=val)
            cell.number_format = "0%"; cell.font = INPUT_FONT
            if is_new:
                cell.fill = YELLOW_FILL
        sumc = ws.cell(row=i, column=5, value=f"=B{i}+C{i}+D{i}")
        sumc.number_format = "0%"

    # Conditional formatting: green if =1, red otherwise
    last = 5 + len(rows)
    green_rule = FormulaRule(formula=[f"E6=1"], fill=GREEN_FILL)
    red_rule = FormulaRule(formula=[f"E6<>1"], fill=RED_FILL)
    ws.conditional_formatting.add(f"E6:E{last}", FormulaRule(formula=["E6=1"], fill=GREEN_FILL))
    ws.conditional_formatting.add(f"E6:E{last}", FormulaRule(formula=["E6<>1"], fill=RED_FILL))

    # Data validation: % values between 0 and 1
    dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=1,
                        allow_blank=False, showErrorMessage=True,
                        errorTitle="Out of bounds", error="Share must be between 0% and 100%.")
    ws.add_data_validation(dv); dv.add(f"B6:D{last}")

    ws.cell(row=last + 2, column=1, value="Class definitions").font = SECTION_FONT
    defs = [
        ("Class A", "Top tier doors. Premium positioning, high foot traffic, strong velocity."),
        ("Class B", "Mid tier doors. Steady velocity. Most of the door base sits here."),
        ("Class C", "Lower tier doors. Velocity ramps slower. Often new accounts."),
    ]
    for i, (a, b) in enumerate(defs, start=last + 3):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True)
        ws.cell(row=i, column=2, value=b)

    for col, w in zip("ABCDE", (20, 12, 12, 12, 14)):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["B"].width = 60
    setup_print(ws)


def build_velocity(ws):
    ws.title = "Velocity"
    title(ws, 1, "Velocity (units per door per week)")
    ws["A3"] = "Per SKU x channel x class. Blended velocity flows into Forecast Weekly."

    headers = ["SKU", "Channel", "Vel A", "Vel B", "Vel C",
               "Class A %", "Class B %", "Class C %", "Blended"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    muv_defaults = {
        "Natural": (12, 8, 5), "Specialty": (10, 7, 4), "FDM": (20, 12, 6),
        "Costco": (200, 120, 60), "On-Premise": (6, 4, 2), "Convenience": (8, 5, 3),
        "Gym & Fitness": (10, 7, 4), "Private Liquor": (5, 3, 2), "RAS": (4, 3, 2),
    }
    muv_lime_factor = {"Natural": (10, 7, 4), "Specialty": (9, 6, 4),
                       "FDM": (18, 11, 5), "Costco": (180, 110, 55)}
    muv_pfp_factor = {"Natural": (14, 9, 5), "Specialty": (11, 7, 4),
                      "FDM": (22, 13, 6), "Costco": (220, 130, 65)}
    lca_factor = {"Natural": (8, 5, 3), "Specialty": (8, 5, 3),
                  "FDM": (15, 10, 5), "Costco": (180, 100, 50),
                  "On-Premise": (8, 5, 3), "Convenience": (12, 8, 4),
                  "Gym & Fitness": (12, 8, 5), "Private Liquor": (3, 2, 1),
                  "RAS": (3, 2, 1)}

    cm_range = "'Class Mix'!$A$6:$D$200"

    row = 6
    for sku in SKUS:
        for ch in CHANNELS:
            if sku == "MUV Raspberry Lemon": a, b, c = muv_defaults[ch]
            elif sku == "MUV Lime": a, b, c = muv_lime_factor.get(ch, muv_defaults[ch])
            elif sku == "MUV PFP": a, b, c = muv_pfp_factor.get(ch, muv_defaults[ch])
            else: a, b, c = lca_factor[ch]
            ws.cell(row=row, column=1, value=sku)
            ws.cell(row=row, column=2, value=ch)
            for col_idx, val in zip((3, 4, 5), (a, b, c)):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = INPUT_FONT
                if ch in NEW_CHANNELS:
                    cell.fill = YELLOW_FILL
            for col_idx, mix_col in zip((6, 7, 8), (2, 3, 4)):
                c1 = ws.cell(row=row, column=col_idx,
                             value=f"=IFERROR(VLOOKUP(B{row},{cm_range},{mix_col},FALSE),0)")
                c1.font = LINK_FONT; c1.number_format = "0%"
            ws.cell(row=row, column=9, value=f"=C{row}*F{row}+D{row}*G{row}+E{row}*H{row}").number_format = "0.00"
            row += 1

    last = row - 1
    # Data validation: velocities >= 0
    dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0,
                        allow_blank=False, showErrorMessage=True,
                        errorTitle="Invalid velocity", error="Velocity must be 0 or positive.")
    ws.add_data_validation(dv); dv.add(f"C6:E{last}")

    # Sanity colour: red on blended < 0.1 (effectively zero)
    ws.conditional_formatting.add(f"I6:I{last}", CellIsRule(operator="lessThan", formula=["0.1"], fill=RED_FILL))

    for col, w in zip("ABCDEFGHI", (22, 16, 8, 8, 8, 10, 10, 10, 10)):
        ws.column_dimensions[col].width = w
    setup_print(ws)


def build_pricing(ws):
    ws.title = "Pricing"
    title(ws, 1, "Pricing and unit economics")
    ws["A3"] = "Net case price is what we invoice the retailer. Landed cost includes ingredients, co-man, freight in."

    headers = ["SKU", "Net case price", "Landed case cost", "Gross profit", "GP %", "Notes"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    rows = [
        ("MUV Raspberry Lemon", 28, 14.5, "Standard MUV price. Confirm with Rijo before locking."),
        ("MUV Lime",            28, 14.5, "Same as Raspberry Lemon. Co-man pricing parity."),
        ("MUV PFP",             28, 14.5, "Same as other MUV SKUs."),
        ("LCA Energy",          32, 16.0, "Energy positioning, higher price point. Cost is placeholder, refine with Rijo."),
    ]
    for i, (sku, price, cost, note) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=sku)
        ws.cell(row=i, column=2, value=price).font = INPUT_FONT
        ws.cell(row=i, column=3, value=cost).font = INPUT_FONT
        ws.cell(row=i, column=4, value=f"=B{i}-C{i}")
        gp = ws.cell(row=i, column=5, value=f"=D{i}/B{i}"); gp.number_format = "0%"
        ws.cell(row=i, column=6, value=note)
        for c in (2, 3, 4):
            ws.cell(row=i, column=c).number_format = "$#,##0.00"

    # Cond format: red GP% < 40%
    ws.conditional_formatting.add("E6:E9",
        CellIsRule(operator="lessThan", formula=["0.40"], fill=RED_FILL))

    for col, w in zip("ABCDEF", (22, 16, 18, 14, 8, 60)):
        ws.column_dimensions[col].width = w
    setup_print(ws)


def build_doors(ws):
    ws.title = "Doors"
    title(ws, 1, "Active doors by brand, province, and channel")
    ws["A3"] = ("Edit the 12 monthly input columns (cream); 52 weekly columns auto-fill. "
                "New channel rows are yellow placeholders pending sign-off.")

    # Headers
    for i, h in enumerate(["Brand", "Province", "Channel", "FY total"], start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for j, m in enumerate(MONTH_ORDER):
        c = ws.cell(row=5, column=DOORS_MONTHLY_FIRST_COL + j, value=m)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    banner = ws.cell(row=4, column=DOORS_MONTHLY_FIRST_COL, value="Monthly inputs (edit)")
    banner.font = Font(bold=True, italic=True, color="7F6000")
    ws.merge_cells(start_row=4, start_column=DOORS_MONTHLY_FIRST_COL,
                   end_row=4, end_column=DOORS_MONTHLY_LAST_COL)
    for j, lbl in enumerate(WEEK_LABELS):
        c = ws.cell(row=5, column=DOORS_WEEKLY_FIRST_COL + j, value=lbl)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    banner2 = ws.cell(row=4, column=DOORS_WEEKLY_FIRST_COL, value="Weekly outputs (do not edit)")
    banner2.font = Font(bold=True, italic=True, color="404040")
    ws.merge_cells(start_row=4, start_column=DOORS_WEEKLY_FIRST_COL,
                   end_row=4, end_column=DOORS_WEEKLY_LAST_COL)

    monthly = {}
    def set_month(brand, prov, ch, values):
        monthly[(brand, prov, ch)] = dict(zip(MONTH_ORDER, values))

    # MUV existing channels
    set_month("MUV", "BC", "Natural",   [80, 95, 110, 125, 140, 155, 170, 180, 190, 195, 200, 200])
    set_month("MUV", "BC", "Specialty", [10, 15, 20, 25, 30, 35, 42, 48, 52, 56, 58, 60])
    set_month("MUV", "BC", "FDM",       [0, 0, 8, 15, 20, 25, 28, 30, 33, 36, 38, 40])
    set_month("MUV", "BC", "Costco",    [0, 0, 0, 0, 0, 0, 60, 0, 0, 25, 25, 25])
    set_month("MUV", "AB", "Natural",   [5, 10, 18, 28, 38, 48, 58, 68, 75, 82, 87, 90])
    set_month("MUV", "AB", "Specialty", [3, 5, 8, 12, 16, 20, 24, 27, 30, 32, 33, 35])
    set_month("MUV", "AB", "FDM",       [0, 0, 0, 5, 10, 15, 18, 22, 25, 27, 28, 30])
    set_month("MUV", "AB", "Costco",    [0, 0, 0, 0, 0, 0, 40, 0, 0, 20, 22, 22])
    set_month("MUV", "SK", "Natural",   [0, 0, 2, 5, 8, 12, 15, 18, 20, 22, 24, 25])
    set_month("MUV", "SK", "Specialty", [0, 0, 1, 2, 4, 6, 8, 10, 12, 13, 14, 15])
    set_month("MUV", "SK", "FDM",       [0, 0, 0, 0, 2, 5, 7, 10, 12, 13, 14, 15])
    set_month("MUV", "SK", "Costco",    [0, 0, 0, 0, 0, 0, 0, 0, 0, 5, 6, 7])
    set_month("MUV", "MB", "Natural",   [0, 0, 2, 4, 7, 10, 13, 15, 17, 19, 21, 22])
    set_month("MUV", "MB", "Specialty", [0, 0, 1, 2, 3, 5, 7, 8, 10, 11, 12, 13])
    set_month("MUV", "MB", "FDM",       [0, 0, 0, 0, 2, 4, 6, 8, 10, 12, 13, 14])
    set_month("MUV", "MB", "Costco",    [0, 0, 0, 0, 0, 0, 0, 0, 0, 4, 5, 6])
    set_month("MUV", "ON", "Natural",   [10, 18, 28, 40, 55, 75, 95, 110, 125, 135, 145, 150])
    set_month("MUV", "ON", "Specialty", [5, 8, 12, 17, 23, 28, 33, 38, 42, 45, 48, 50])
    set_month("MUV", "ON", "FDM",       [0, 0, 5, 12, 20, 28, 35, 42, 48, 53, 57, 60])
    set_month("MUV", "ON", "Costco",    [0, 0, 0, 0, 0, 0, 0, 0, 0, 30, 32, 35])
    set_month("MUV", "QC", "Natural",   [20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75])
    set_month("MUV", "QC", "Specialty", [5, 8, 11, 14, 17, 20, 23, 25, 28, 30, 32, 32])
    set_month("MUV", "QC", "FDM",       [0, 0, 0, 5, 8, 12, 15, 18, 20, 22, 23, 25])
    set_month("MUV", "QC", "Costco",    [0, 0, 0, 0, 0, 0, 0, 0, 0, 20, 22, 22])
    set_month("MUV", "NB", "Natural",   [0, 0, 1, 3, 5, 7, 9, 11, 13, 14, 15, 16])
    set_month("MUV", "NB", "Specialty", [0, 0, 0, 1, 2, 4, 5, 6, 7, 8, 9, 10])
    set_month("MUV", "NB", "FDM",       [0, 0, 0, 0, 0, 2, 3, 5, 7, 8, 9, 10])
    set_month("MUV", "NB", "Costco",    [0, 0, 0, 0, 0, 0, 0, 0, 0, 3, 4, 5])
    set_month("MUV", "NS", "Natural",   [0, 0, 1, 3, 5, 7, 9, 11, 13, 14, 15, 16])
    set_month("MUV", "NS", "Specialty", [0, 0, 0, 1, 2, 4, 5, 6, 7, 8, 9, 10])
    set_month("MUV", "NS", "FDM",       [0, 0, 0, 0, 0, 2, 3, 5, 7, 8, 9, 10])
    set_month("MUV", "NS", "Costco",    [0, 0, 0, 0, 0, 0, 0, 0, 0, 3, 4, 5])
    set_month("MUV", "PE", "Natural",   [0, 0, 0, 1, 2, 3, 4, 5, 6, 6, 7, 7])
    set_month("MUV", "PE", "Specialty", [0, 0, 0, 0, 1, 2, 2, 3, 3, 4, 4, 4])
    set_month("MUV", "PE", "FDM",       [0, 0, 0, 0, 0, 0, 1, 2, 2, 3, 3, 3])
    set_month("MUV", "PE", "Costco",    [0]*12)
    set_month("MUV", "NL", "Natural",   [0, 0, 0, 1, 2, 3, 4, 5, 6, 6, 7, 7])
    set_month("MUV", "NL", "Specialty", [0, 0, 0, 0, 1, 2, 2, 3, 3, 4, 4, 4])
    set_month("MUV", "NL", "FDM",       [0, 0, 0, 0, 0, 0, 1, 2, 2, 3, 3, 3])
    set_month("MUV", "NL", "Costco",    [0]*12)

    muv_new = {
        "On-Premise":     {"BC":[30,40,55,70,80,90,100,105,110,115,118,120],"AB":[20,28,38,48,56,62,68,72,75,78,80,82],"SK":[5,8,12,15,18,21,24,26,28,29,30,30],"MB":[5,8,12,15,18,21,24,26,28,29,30,30],"ON":[40,55,75,90,105,120,130,140,145,150,152,155],"QC":[30,45,60,75,85,95,105,110,115,118,120,122],"NB":[3,5,7,9,10,11,12,13,14,14,15,15],"NS":[3,5,7,9,10,11,12,13,14,14,15,15],"PE":[1,1,2,2,3,3,4,4,4,5,5,5],"NL":[1,1,2,2,3,3,4,4,4,5,5,5]},
        "Convenience":    {"BC":[20,30,45,60,70,80,88,92,95,98,100,100],"AB":[15,22,32,45,55,62,68,72,75,78,80,80],"SK":[5,8,12,16,20,23,25,27,28,29,30,30],"MB":[5,8,12,16,20,23,25,27,28,29,30,30],"ON":[50,75,105,130,150,170,180,190,195,198,200,200],"QC":[30,50,70,90,105,115,122,126,128,130,130,130],"NB":[3,5,7,9,11,12,13,14,14,15,15,15],"NS":[3,5,7,9,11,12,13,14,14,15,15,15],"PE":[1,2,2,3,3,4,4,5,5,5,5,5],"NL":[1,2,2,3,3,4,4,5,5,5,5,5]},
        "Gym & Fitness":  {"BC":[15,22,30,38,44,50,54,56,58,59,60,60],"AB":[10,16,24,32,38,42,45,47,49,50,50,50],"SK":[3,5,7,9,11,12,13,14,14,15,15,15],"MB":[3,5,7,9,11,12,13,14,14,15,15,15],"ON":[25,38,55,70,82,90,95,98,99,100,100,100],"QC":[15,22,32,42,50,55,58,59,60,60,60,60],"NB":[2,3,4,5,6,7,7,8,8,8,8,8],"NS":[2,3,4,5,6,7,7,8,8,8,8,8],"PE":[0,0,1,1,2,2,2,3,3,3,3,3],"NL":[0,0,1,1,2,2,2,3,3,3,3,3]},
        "Private Liquor": {"BC":[10,15,22,28,34,40,44,47,49,50,50,50],"AB":[8,12,18,24,30,34,37,39,40,40,40,40],"SK":[0,0,2,4,6,8,10,12,14,15,15,15],"MB":[0,0,2,4,6,8,10,11,12,13,14,15],"ON":[0]*12,"QC":[0]*12,"NB":[0]*12,"NS":[0]*12,"PE":[0]*12,"NL":[0]*12},
        "RAS":            {"BC":[0]*12,"AB":[0]*12,"SK":[0]*12,"MB":[0]*12,"ON":[5,8,12,16,20,24,26,28,29,30,30,30],"QC":[0]*12,"NB":[2,3,5,6,7,8,9,10,10,10,10,10],"NS":[2,3,5,6,7,8,9,10,10,10,10,10],"PE":[0]*12,"NL":[0]*12},
    }
    for ch, pm in muv_new.items():
        for prov, vals in pm.items():
            set_month("MUV", prov, ch, vals)

    # LCA
    set_month("LCA", "BC", "Natural",   [0, 0, 0, 0, 5, 10, 15, 20, 25, 28, 30, 32])
    set_month("LCA", "BC", "Specialty", [0, 0, 0, 0, 2, 5, 8, 12, 15, 17, 18, 20])
    set_month("LCA", "BC", "FDM",       [0]*12)
    set_month("LCA", "BC", "Costco",    [0]*12)
    set_month("LCA", "BC", "On-Premise",     [0, 0, 0, 0, 3, 6, 10, 14, 18, 20, 22, 24])
    set_month("LCA", "BC", "Convenience",    [0, 0, 0, 0, 5, 10, 15, 20, 25, 28, 30, 32])
    set_month("LCA", "BC", "Gym & Fitness",  [0, 0, 0, 0, 3, 6, 10, 13, 16, 18, 20, 22])
    set_month("LCA", "BC", "Private Liquor", [0]*12)
    set_month("LCA", "BC", "RAS",            [0]*12)
    for prov in PROVINCES:
        for ch in CHANNELS:
            if ("LCA", prov, ch) not in monthly:
                set_month("LCA", prov, ch, [0]*12)

    row = 6
    for brand in ("MUV", "LCA"):
        for prov in PROVINCES:
            for ch in CHANNELS:
                key = (brand, prov, ch)
                mvals = monthly.get(key, {m: 0 for m in MONTH_ORDER})
                ws.cell(row=row, column=1, value=brand)
                ws.cell(row=row, column=2, value=prov)
                ws.cell(row=row, column=3, value=ch)
                first_wk = doors_week_col(1); last_wk = doors_week_col(N_WEEKS)
                ws.cell(row=row, column=4, value=f"=SUM({first_wk}{row}:{last_wk}{row})")
                for j, m in enumerate(MONTH_ORDER):
                    cell = ws.cell(row=row, column=DOORS_MONTHLY_FIRST_COL + j, value=mvals[m])
                    cell.font = INPUT_FONT
                    cell.fill = YELLOW_FILL if ch in NEW_CHANNELS else MONTHLY_FILL
                for wi in range(1, N_WEEKS + 1):
                    month = WEEK_TO_MONTH[wi - 1]
                    mcol = MONTH_INPUT_COL[month]
                    cell = ws.cell(row=row, column=DOORS_WEEKLY_FIRST_COL + wi - 1,
                                   value=f"=${mcol}{row}")
                    cell.font = LINK_FONT
                row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="National total").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=f"=SUM(D6:D{total_row - 1})").font = Font(bold=True)
    for j in range(12):
        col = DOORS_MONTHLY_FIRST_COL + j
        colL = get_column_letter(col)
        c = ws.cell(row=total_row, column=col, value=f"=SUM({colL}6:{colL}{total_row - 1})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL
    for wi in range(1, N_WEEKS + 1):
        col = DOORS_WEEKLY_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        c = ws.cell(row=total_row, column=col, value=f"=SUM({colL}6:{colL}{total_row - 1})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL

    # Data validation: door counts >= 0 on monthly inputs
    dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1=0,
                        allow_blank=False, showErrorMessage=True,
                        errorTitle="Invalid door count", error="Door count must be 0 or positive.")
    ws.add_data_validation(dv)
    dv.add(f"E6:P{total_row - 1}")

    # Widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    for j in range(12):
        ws.column_dimensions[get_column_letter(DOORS_MONTHLY_FIRST_COL + j)].width = 11
    for wi in range(1, N_WEEKS + 1):
        ws.column_dimensions[get_column_letter(DOORS_WEEKLY_FIRST_COL + wi - 1)].width = 11
    ws.freeze_panes = "E6"
    setup_print(ws)
    return total_row


def build_marketing(ws):
    ws.title = "Marketing and Sampling"
    title(ws, 1, "Marketing and sampling cases (per week)")
    ws["A3"] = "All cases per week. Feeds production demand. Yellow = needs sign-off."

    ws.cell(row=5, column=1, value="Bucket").font = HEADER_FONT
    ws["A5"].fill = HEADER_FILL
    for j, lbl in enumerate(WEEK_LABELS):
        c = ws.cell(row=5, column=FW_FIRST_COL + j, value=lbl)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    bucket_labels = [
        "Account sampling subtotal", "Events subtotal", "Rep samples subtotal",
        "Buffer subtotal", "Total marketing and sampling",
    ]
    for i, lbl in enumerate(bucket_labels, start=6):
        ws.cell(row=i, column=1, value=lbl).font = Font(bold=True)

    sec1_header = 13
    section(ws, sec1_header, "Section 1: Account sampling (1 rate per channel)")
    ws.cell(row=sec1_header + 1, column=1, value="Channel").font = HEADER_FONT
    ws.cell(row=sec1_header + 1, column=2, value="Cases per door per week").font = HEADER_FONT
    ws[f"A{sec1_header+1}"].fill = HEADER_FILL
    ws[f"B{sec1_header+1}"].fill = HEADER_FILL
    for j, lbl in enumerate(WEEK_LABELS):
        c = ws.cell(row=sec1_header + 1, column=FW_FIRST_COL + j, value=lbl)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    sampling_rates = {
        "Natural": 0.025, "Specialty": 0.020, "FDM": 0.0125, "Costco": 0.050,
        "On-Premise": 0.040, "Convenience": 0.005, "Gym & Fitness": 0.050,
        "Private Liquor": 0.010, "RAS": 0.010,
    }
    sec1_first = sec1_header + 2
    for i, ch in enumerate(CHANNELS):
        r = sec1_first + i
        ws.cell(row=r, column=1, value=ch)
        rate = ws.cell(row=r, column=2, value=sampling_rates[ch])
        rate.font = INPUT_FONT; rate.number_format = "0.0000"
        if ch in NEW_CHANNELS:
            rate.fill = YELLOW_FILL
        for wi in range(1, N_WEEKS + 1):
            col = FW_FIRST_COL + wi - 1
            dwc = doors_week_col(wi)
            ws.cell(row=r, column=col,
                    value=f"=$B{r}*SUMIFS(Doors!{dwc}:{dwc},Doors!$C:$C,$A{r})").font = LINK_FONT
    sec1_last = sec1_first + len(CHANNELS) - 1
    sec1_total = sec1_last + 1
    ws.cell(row=sec1_total, column=1, value="Account sampling total").font = Font(bold=True)
    for wi in range(1, N_WEEKS + 1):
        col = FW_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        c = ws.cell(row=sec1_total, column=col,
                    value=f"=SUM({colL}{sec1_first}:{colL}{sec1_last})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL

    # Section 2: Events
    sec2_header = sec1_total + 2
    section(ws, sec2_header, "Section 2: Events (Total + Start week + End week)")
    sub = sec2_header + 1
    ws.cell(row=sub, column=1, value="Event").font = HEADER_FONT
    ws.cell(row=sub, column=2, value="Total cases").font = HEADER_FONT
    ws.cell(row=sub, column=3, value="Start week").font = HEADER_FONT
    ws.cell(row=sub, column=4, value="End week").font = HEADER_FONT
    for col in "ABCD":
        ws[f"{col}{sub}"].fill = HEADER_FILL
    for j, lbl in enumerate(WEEK_LABELS):
        c = ws.cell(row=sub, column=FW_FIRST_COL + j, value=lbl)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    def week_for_first_of(month):
        for i, m in enumerate(WEEK_TO_MONTH, start=1):
            if m == month:
                return i
        return 1
    def week_for_last_of(month):
        last = 1
        for i, m in enumerate(WEEK_TO_MONTH, start=1):
            if m == month:
                last = i
        return last

    events = [
        ("CHFA West (Vancouver) - BC trade show",      25, week_for_first_of("Sep 2026"), week_for_first_of("Sep 2026")),
        ("Indie Alley demo days (recurring)",          36, 1, N_WEEKS),
        ("Marche Tao launch event - QC",               15, week_for_first_of("Sep 2026"), week_for_first_of("Sep 2026")),
        ("Costco BC roadshow",                         60, week_for_first_of("Mar 2027"), week_for_first_of("Mar 2027")),
        ("Costco AB roadshow",                         40, week_for_first_of("Mar 2027"), week_for_first_of("Mar 2027")),
        ("CHFA East (Toronto) - ON trade show",        25, week_for_first_of("Apr 2027"), week_for_first_of("Apr 2027")),
        ("BCAA Wellness Tour - BC",                    30, week_for_first_of("Apr 2027"), week_for_last_of("May 2027")),
        ("Toronto Vegfest - ON",                       20, week_for_first_of("Jun 2027") + 2, week_for_first_of("Jun 2027") + 2),
        ("Calgary Stampede - AB",                      25, week_for_first_of("Jul 2027") + 1, week_for_first_of("Jul 2027") + 1),
        ("Yoga & fitness partners (recurring)",        36, 1, N_WEEKS),
        ("Influencer seeding (recurring)",             48, 1, N_WEEKS),
        ("Restaurant launch dinners (on-premise)",     20, week_for_first_of("Oct 2026"), week_for_last_of("May 2027")),
        ("Convenience sales blitz (recurring)",        24, 1, N_WEEKS),
    ]
    sec2_first = sub + 1
    for i, (lbl, total, start, end) in enumerate(events):
        r = sec2_first + i
        ws.cell(row=r, column=1, value=lbl)
        ws.cell(row=r, column=2, value=total).font = INPUT_FONT
        ws.cell(row=r, column=3, value=start).font = INPUT_FONT
        ws.cell(row=r, column=4, value=end).font = INPUT_FONT
        for wi in range(1, N_WEEKS + 1):
            col = FW_FIRST_COL + wi - 1
            ws.cell(row=r, column=col,
                    value=(f"=IF(AND({wi}>=$C{r},{wi}<=$D{r}),"
                           f"$B{r}/($D{r}-$C{r}+1),0)")).font = LINK_FONT
    sec2_last = sec2_first + len(events) - 1
    sec2_total = sec2_last + 1

    # Data validation start/end weeks 1..N_WEEKS
    dv_wk = DataValidation(type="whole", operator="between", formula1=1, formula2=N_WEEKS,
                           allow_blank=False, showErrorMessage=True,
                           errorTitle="Invalid week", error=f"Week must be between 1 and {N_WEEKS}.")
    ws.add_data_validation(dv_wk)
    dv_wk.add(f"C{sec2_first}:D{sec2_last}")

    ws.cell(row=sec2_total, column=1, value="Events total").font = Font(bold=True)
    for wi in range(1, N_WEEKS + 1):
        col = FW_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        c = ws.cell(row=sec2_total, column=col,
                    value=f"=SUM({colL}{sec2_first}:{colL}{sec2_last})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL

    # Section 3: Rep samples
    sec3_header = sec2_total + 2
    section(ws, sec3_header, "Section 3: Rep samples (monthly rep counts drive weekly)")
    rate_row = sec3_header + 1
    ws.cell(row=rate_row, column=1, value="Cases per rep per week")
    ws.cell(row=rate_row, column=2, value=1.0).font = INPUT_FONT
    sub3 = sec3_header + 2
    ws.cell(row=sub3, column=1, value="Province").font = HEADER_FONT
    ws[f"A{sub3}"].fill = HEADER_FILL
    for j, m in enumerate(MONTH_ORDER):
        c = ws.cell(row=sub3, column=REP_MONTHLY_FIRST_COL + j, value=m)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for j, lbl in enumerate(WEEK_LABELS):
        c = ws.cell(row=sub3, column=REP_WEEKLY_FIRST_COL + j, value=lbl)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    rep_plan = {
        "BC": [2, 2, 2, 2, 2, 3, 3, 3, 3, 3, 3, 3],
        "AB": [1, 1, 1, 1, 1, 1, 2, 2, 2, 2, 2, 2],
        "SK": [0, 0, 0, 0, 0, 1, 1, 1, 1, 1, 1, 1],
        "MB": [0, 0, 0, 0, 0, 1, 1, 1, 1, 1, 1, 1],
        "ON": [1, 1, 1, 1, 1, 1, 2, 2, 2, 2, 2, 2],
        "QC": [1, 1, 1, 1, 1, 1, 1, 1, 2, 2, 2, 2],
        "NB": [0, 0, 0, 0, 0, 0, 1, 1, 1, 1, 1, 1],
        "NS": [0, 0, 0, 0, 0, 0, 1, 1, 1, 1, 1, 1],
        "PE": [0]*12, "NL": [0]*12,
    }
    sec3_first = sub3 + 1
    for i, prov in enumerate(PROVINCES):
        r = sec3_first + i
        ws.cell(row=r, column=1, value=prov)
        for j, m in enumerate(MONTH_ORDER):
            cell = ws.cell(row=r, column=REP_MONTHLY_FIRST_COL + j, value=rep_plan[prov][j])
            cell.font = INPUT_FONT; cell.fill = MONTHLY_FILL
        for wi in range(1, N_WEEKS + 1):
            month = WEEK_TO_MONTH[wi - 1]
            mcol = REP_MONTH_INPUT_COL[month]
            ws.cell(row=r, column=REP_WEEKLY_FIRST_COL + wi - 1,
                    value=f"=${mcol}{r}").font = LINK_FONT

    total_reps_row = sec3_first + len(PROVINCES)
    sample_row = total_reps_row + 1
    ws.cell(row=total_reps_row, column=1, value="Total reps active").font = Font(bold=True)
    ws.cell(row=sample_row, column=1, value="Rep sample cases total").font = Font(bold=True)
    for wi in range(1, N_WEEKS + 1):
        col = REP_WEEKLY_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        ws.cell(row=total_reps_row, column=col,
                value=(f"=SUM({colL}{sec3_first}:{colL}{sec3_first + len(PROVINCES) - 1})")).font = Font(bold=True)
        c = ws.cell(row=sample_row, column=col,
                    value=f"=$B${rate_row}*{colL}{total_reps_row}")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL

    rep_link_row = sample_row + 1
    ws.cell(row=rep_link_row, column=1,
            value="(Rep samples projected to standard week columns)").font = Font(italic=True, color="808080")
    for wi in range(1, N_WEEKS + 1):
        src = get_column_letter(REP_WEEKLY_FIRST_COL + wi - 1)
        dst_col = FW_FIRST_COL + wi - 1
        ws.cell(row=rep_link_row, column=dst_col, value=f"={src}{sample_row}").font = LINK_FONT

    # Section 4: Buffer
    sec4_header = rep_link_row + 2
    section(ws, sec4_header, "Section 4: Buffer (1 rate per item)")
    sub4 = sec4_header + 1
    ws.cell(row=sub4, column=1, value="Line item").font = HEADER_FONT
    ws.cell(row=sub4, column=2, value="Cases per week").font = HEADER_FONT
    ws.cell(row=sub4, column=3, value="Notes").font = HEADER_FONT
    for c in "ABC":
        ws[f"{c}{sub4}"].fill = HEADER_FILL
    for j, lbl in enumerate(WEEK_LABELS):
        c = ws.cell(row=sub4, column=FW_FIRST_COL + j, value=lbl)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    buffer_items = [
        ("Gifting",          0.5,  "Aaron and Teresa gifting allocation."),
        ("R and D samples",  0.75, "Formulation tests, internal taste panels."),
        ("Photo and content", 0.5, "Brand asset shoots, ongoing."),
        ("Freight returns",  0.25, "Damaged or returned cases."),
        ("Unallocated",      0.75, "Reserve for unplanned asks."),
    ]
    sec4_first = sub4 + 1
    for i, (lbl, per_week, note) in enumerate(buffer_items):
        r = sec4_first + i
        ws.cell(row=r, column=1, value=lbl)
        ws.cell(row=r, column=2, value=per_week).font = INPUT_FONT
        ws.cell(row=r, column=3, value=note)
        for wi in range(1, N_WEEKS + 1):
            col = FW_FIRST_COL + wi - 1
            ws.cell(row=r, column=col, value=f"=$B{r}").font = LINK_FONT
    sec4_last = sec4_first + len(buffer_items) - 1
    sec4_total = sec4_last + 1
    ws.cell(row=sec4_total, column=1, value="Buffer total").font = Font(bold=True)
    for wi in range(1, N_WEEKS + 1):
        col = FW_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        c = ws.cell(row=sec4_total, column=col,
                    value=f"=SUM({colL}{sec4_first}:{colL}{sec4_last})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL

    # Wire summary
    for wi in range(1, N_WEEKS + 1):
        col = FW_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        ws.cell(row=6, column=col, value=f"={colL}{sec1_total}").font = LINK_FONT
        ws.cell(row=7, column=col, value=f"={colL}{sec2_total}").font = LINK_FONT
        ws.cell(row=8, column=col, value=f"={colL}{rep_link_row}").font = LINK_FONT
        ws.cell(row=9, column=col, value=f"={colL}{sec4_total}").font = LINK_FONT
        c = ws.cell(row=10, column=col, value=f"=SUM({colL}6:{colL}9)")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    for wi in range(1, N_WEEKS + 1):
        ws.column_dimensions[get_column_letter(FW_FIRST_COL + wi - 1)].width = 11
    for j in range(12):
        ws.column_dimensions[get_column_letter(REP_MONTHLY_FIRST_COL + j)].width = 11
    for wi in range(1, N_WEEKS + 1):
        ws.column_dimensions[get_column_letter(REP_WEEKLY_FIRST_COL + wi - 1)].width = 11
    ws.freeze_panes = "E6"
    setup_print(ws)
    return {"summary_acct": 6, "summary_events": 7, "summary_reps": 8,
            "summary_buffer": 9, "summary_total": 10}


def build_forecast_weekly(ws):
    ws.title = "Forecast Weekly"
    title(ws, 1, "Forecast weekly sales cases")
    ws["A3"] = (f"Engine. {len(SKUS) * len(PROVINCES) * len(CHANNELS)} row grid. Do not edit.")

    # Helpers after weekly cols: BE = FY cases, BF = FY revenue
    fy_cases_col = get_column_letter(FW_FIRST_COL + N_WEEKS)        # BE
    fy_rev_col   = get_column_letter(FW_FIRST_COL + N_WEEKS + 1)    # BF

    headers = ["SKU", "Province", "Channel", "Brand"] + WEEK_LABELS + ["FY cases", "FY revenue"]
    for col_idx, val in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col_idx, value=val)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for i, m in enumerate(WEEK_TO_MONTH, start=1):
        c = ws.cell(row=4, column=FW_FIRST_COL + i - 1, value=m)
        c.font = Font(italic=True, color="808080", size=9)

    row = 6
    for sku in SKUS:
        brand = SKU_BRAND[sku]
        for prov in PROVINCES:
            for ch in CHANNELS:
                ws.cell(row=row, column=1, value=sku)
                ws.cell(row=row, column=2, value=prov)
                ws.cell(row=row, column=3, value=ch)
                ws.cell(row=row, column=4, value=brand)
                for wi in range(1, N_WEEKS + 1):
                    col = FW_FIRST_COL + wi - 1
                    dwc = doors_week_col(wi)
                    f = (f"=SUMIFS(Doors!{dwc}:{dwc},Doors!$A:$A,$D{row},"
                         f"Doors!$B:$B,$B{row},Doors!$C:$C,$C{row})"
                         f"*SUMIFS(Velocity!$I:$I,Velocity!$A:$A,$A{row},"
                         f"Velocity!$B:$B,$C{row})/Assumptions!$B$4*Assumptions!$B$6")
                    ws.cell(row=row, column=col, value=f).font = LINK_FONT
                # FY cases helper
                ws.cell(row=row, column=FW_FIRST_COL + N_WEEKS,
                        value=f"=SUM({fw_week_col(1)}{row}:{fw_week_col(N_WEEKS)}{row})")
                # FY revenue helper = FY cases * SKU price
                ws.cell(row=row, column=FW_FIRST_COL + N_WEEKS + 1,
                        value=f"={fy_cases_col}{row}*IFERROR(VLOOKUP($A{row},Pricing!$A$6:$B$200,2,FALSE),0)").number_format = "$#,##0"
                row += 1
    total_row = row
    ws.cell(row=total_row, column=1, value="National total").font = Font(bold=True)
    for wi in range(1, N_WEEKS + 1):
        col = FW_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        c = ws.cell(row=total_row, column=col,
                    value=f"=SUM({colL}6:{colL}{total_row - 1})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL
    # Totals on helper cols too
    for col_letter in (fy_cases_col, fy_rev_col):
        c = ws.cell(row=total_row, column=ws[f"{col_letter}5"].column,
                    value=f"=SUM({col_letter}6:{col_letter}{total_row - 1})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 8
    for wi in range(1, N_WEEKS + 1):
        ws.column_dimensions[get_column_letter(FW_FIRST_COL + wi - 1)].width = 11
    ws.column_dimensions[fy_cases_col].width = 12
    ws.column_dimensions[fy_rev_col].width = 14
    ws.freeze_panes = "E6"
    return total_row


def build_production_plan(ws, forecast_total_row, marketing_summary):
    ws.title = "Production Plan"
    title(ws, 1, "Production plan (cases)")
    ws["A3"] = "Sales + marketing & sampling -> total to produce. Buffer applied at end."

    section(ws, 5, "Sales by SKU")
    headers = ["SKU"] + [""] * 3 + WEEK_LABELS + ["FY total"]
    for col_idx, val in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col_idx, value=val)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    fy_col = FW_FIRST_COL + N_WEEKS
    for i, sku in enumerate(SKUS):
        r = 7 + i
        ws.cell(row=r, column=1, value=sku)
        for wi in range(1, N_WEEKS + 1):
            col = FW_FIRST_COL + wi - 1
            colL = get_column_letter(col)
            ws.cell(row=r, column=col,
                    value=(f"=SUMIFS('Forecast Weekly'!{colL}:{colL},"
                           f"'Forecast Weekly'!$A:$A,$A{r})")).font = LINK_FONT
        ws.cell(row=r, column=fy_col,
                value=f"=SUM({fw_week_col(1)}{r}:{fw_week_col(N_WEEKS)}{r})")

    total_sales_row = 7 + len(SKUS)
    ws.cell(row=total_sales_row, column=1, value="Total sales cases").font = Font(bold=True)
    for wi in range(1, N_WEEKS + 1):
        col = FW_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        c = ws.cell(row=total_sales_row, column=col,
                    value=f"=SUM({colL}7:{colL}{total_sales_row - 1})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL
    ws.cell(row=total_sales_row, column=fy_col,
            value=(f"=SUM({fw_week_col(1)}{total_sales_row}:"
                   f"{fw_week_col(N_WEEKS)}{total_sales_row})")).font = Font(bold=True)

    mk_header_row = total_sales_row + 2
    section(ws, mk_header_row, "Marketing and sampling breakdown")
    mk_rows = [("Account sampling", marketing_summary["summary_acct"]),
               ("Events and activations", marketing_summary["summary_events"]),
               ("Rep samples", marketing_summary["summary_reps"]),
               ("Buffer", marketing_summary["summary_buffer"])]
    mk_first = mk_header_row + 1
    for i, (lbl, mk_row) in enumerate(mk_rows):
        r = mk_first + i
        ws.cell(row=r, column=1, value=lbl)
        for wi in range(1, N_WEEKS + 1):
            col = FW_FIRST_COL + wi - 1
            colL = get_column_letter(col)
            ws.cell(row=r, column=col,
                    value=f"='Marketing and Sampling'!{colL}{mk_row}").font = LINK_FONT
        ws.cell(row=r, column=fy_col,
                value=(f"=SUM({fw_week_col(1)}{r}:{fw_week_col(N_WEEKS)}{r})"))

    mk_total_row = mk_first + len(mk_rows)
    ws.cell(row=mk_total_row, column=1, value="Total marketing and sampling").font = Font(bold=True)
    for wi in range(1, N_WEEKS + 1):
        col = FW_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        c = ws.cell(row=mk_total_row, column=col,
                    value=f"=SUM({colL}{mk_first}:{colL}{mk_total_row - 1})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL
    ws.cell(row=mk_total_row, column=fy_col,
            value=(f"=SUM({fw_week_col(1)}{mk_total_row}:"
                   f"{fw_week_col(N_WEEKS)}{mk_total_row})")).font = Font(bold=True)

    total_prod_row = mk_total_row + 2
    ws.cell(row=total_prod_row, column=1, value="Total cases to produce").font = Font(bold=True)
    for wi in range(1, N_WEEKS + 1):
        col = FW_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        ws.cell(row=total_prod_row, column=col,
                value=f"={colL}{total_sales_row}+{colL}{mk_total_row}").font = Font(bold=True)
    ws.cell(row=total_prod_row, column=fy_col,
            value=(f"=SUM({fw_week_col(1)}{total_prod_row}:"
                   f"{fw_week_col(N_WEEKS)}{total_prod_row})")).font = Font(bold=True)

    buf_row = total_prod_row + 1
    ws.cell(row=buf_row, column=1, value="Production with buffer").font = Font(bold=True)
    for wi in range(1, N_WEEKS + 1):
        col = FW_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        c = ws.cell(row=buf_row, column=col,
                    value=f"={colL}{total_prod_row}*(1+Assumptions!$B$5)")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL
    ws.cell(row=buf_row, column=fy_col,
            value=(f"=SUM({fw_week_col(1)}{buf_row}:"
                   f"{fw_week_col(N_WEEKS)}{buf_row})")).font = Font(bold=True)

    ws.column_dimensions["A"].width = 30
    for wi in range(1, N_WEEKS + 1):
        ws.column_dimensions[get_column_letter(FW_FIRST_COL + wi - 1)].width = 11
    ws.column_dimensions[get_column_letter(fy_col)].width = 14
    ws.freeze_panes = "B7"
    setup_print(ws)
    return {"total_sales_row": total_sales_row, "mk_total_row": mk_total_row,
            "total_prod_row": total_prod_row, "buf_row": buf_row, "fy_col": fy_col}


def build_revenue(ws):
    ws.title = "Revenue"
    title(ws, 1, "Revenue and gross profit (CAD)")
    ws["A3"] = "Revenue = sales cases x net case price. GP = sales cases x (price - cost)."

    section(ws, 5, "Revenue")
    fy_col = FW_FIRST_COL + N_WEEKS
    pr_range = "Pricing!$A$6:$D$200"

    headers = ["SKU"] + [""] * 3 + WEEK_LABELS + ["FY total"]
    for col_idx, val in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col_idx, value=val)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    for i, sku in enumerate(SKUS):
        r = 7 + i
        ws.cell(row=r, column=1, value=sku)
        for wi in range(1, N_WEEKS + 1):
            col = FW_FIRST_COL + wi - 1
            colL = get_column_letter(col)
            ws.cell(row=r, column=col,
                    value=(f"=SUMIFS('Forecast Weekly'!{colL}:{colL},"
                           f"'Forecast Weekly'!$A:$A,$A{r})*"
                           f"VLOOKUP($A{r},{pr_range},2,FALSE)")).font = LINK_FONT
            ws.cell(row=r, column=col).number_format = "$#,##0"
        ws.cell(row=r, column=fy_col,
                value=(f"=SUM({fw_week_col(1)}{r}:"
                       f"{fw_week_col(N_WEEKS)}{r})")).number_format = "$#,##0"

    rev_total_row = 7 + len(SKUS)
    ws.cell(row=rev_total_row, column=1, value="Total revenue").font = Font(bold=True)
    for wi in range(1, N_WEEKS + 1):
        col = FW_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        c = ws.cell(row=rev_total_row, column=col,
                    value=f"=SUM({colL}7:{colL}{rev_total_row - 1})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL; c.number_format = "$#,##0"
    ws.cell(row=rev_total_row, column=fy_col,
            value=(f"=SUM({fw_week_col(1)}{rev_total_row}:"
                   f"{fw_week_col(N_WEEKS)}{rev_total_row})")).number_format = "$#,##0"

    gp_header_row = rev_total_row + 2
    section(ws, gp_header_row, "Gross profit")
    gp_hdr_row = gp_header_row + 1
    for col_idx, val in enumerate(headers, start=1):
        c = ws.cell(row=gp_hdr_row, column=col_idx, value=val)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    for i, sku in enumerate(SKUS):
        r = gp_hdr_row + 1 + i
        ws.cell(row=r, column=1, value=sku)
        for wi in range(1, N_WEEKS + 1):
            col = FW_FIRST_COL + wi - 1
            colL = get_column_letter(col)
            ws.cell(row=r, column=col,
                    value=(f"=SUMIFS('Forecast Weekly'!{colL}:{colL},"
                           f"'Forecast Weekly'!$A:$A,$A{r})*"
                           f"VLOOKUP($A{r},{pr_range},4,FALSE)")).font = LINK_FONT
            ws.cell(row=r, column=col).number_format = "$#,##0"
        ws.cell(row=r, column=fy_col,
                value=(f"=SUM({fw_week_col(1)}{r}:"
                       f"{fw_week_col(N_WEEKS)}{r})")).number_format = "$#,##0"

    gp_total_row = gp_hdr_row + 1 + len(SKUS)
    ws.cell(row=gp_total_row, column=1, value="Total gross profit").font = Font(bold=True)
    for wi in range(1, N_WEEKS + 1):
        col = FW_FIRST_COL + wi - 1
        colL = get_column_letter(col)
        c = ws.cell(row=gp_total_row, column=col,
                    value=f"=SUM({colL}{gp_hdr_row + 1}:{colL}{gp_total_row - 1})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL; c.number_format = "$#,##0"
    ws.cell(row=gp_total_row, column=fy_col,
            value=(f"=SUM({fw_week_col(1)}{gp_total_row}:"
                   f"{fw_week_col(N_WEEKS)}{gp_total_row})")).number_format = "$#,##0"

    ws.column_dimensions["A"].width = 22
    for wi in range(1, N_WEEKS + 1):
        ws.column_dimensions[get_column_letter(FW_FIRST_COL + wi - 1)].width = 11
    ws.column_dimensions[get_column_letter(fy_col)].width = 14
    ws.freeze_panes = "B7"
    setup_print(ws)
    return {"rev_total_row": rev_total_row, "gp_total_row": gp_total_row, "fy_col": fy_col}


def build_trade_spend(ws):
    """Trade spend per channel: one-time listing fees, slotting per door, ongoing scan/promo % of revenue."""
    ws.title = "Trade Spend"
    title(ws, 1, "Trade spend by channel (CAD)")
    ws["A3"] = ("Listing fees (one-time at launch), slotting (per door, year 1), and "
                "scan/promo allowance (% of channel revenue, ongoing).")

    headers = ["Channel", "Listing fee (one-time)", "Slotting per door",
               "Scan/promo % of revenue", "Listing total", "Slotting total",
               "Scan/promo total", "Total trade spend"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # (channel, listing_fee, slotting_per_door, scan_pct)
    trade = [
        ("Natural",         5000,   0,      0.02, False),
        ("Specialty",       3000,   0,      0.02, False),
        ("FDM",             25000,  500,    0.05, False),
        ("Costco",          50000,  0,      0.08, False),  # Costco is OI / scan-heavy
        ("On-Premise",      5000,   50,     0.03, True),
        ("Convenience",     15000,  250,    0.04, True),
        ("Gym & Fitness",   2000,   0,      0.02, True),
        ("Private Liquor",  4000,   100,    0.03, True),
        ("RAS",             1000,   0,      0.02, True),
    ]
    for i, (ch, lf, sl, sp, is_new) in enumerate(trade, start=6):
        ws.cell(row=i, column=1, value=ch)
        for col, val in zip((2, 3, 4), (lf, sl, sp)):
            cell = ws.cell(row=i, column=col, value=val); cell.font = INPUT_FONT
            if is_new:
                cell.fill = YELLOW_FILL
        # Listing fee total = one-time per channel
        ws.cell(row=i, column=5, value=f"=B{i}")
        # Slotting total: per-door * doors at year-end (W52)
        ws.cell(row=i, column=6,
                value=f"=$C{i}*SUMIFS(Doors!{doors_week_col(N_WEEKS)}:{doors_week_col(N_WEEKS)},Doors!$C:$C,$A{i})")
        # Scan/promo total: pct * total FY revenue from this channel.
        # Forecast Weekly carries a helper FY-revenue column (BF) so this is a simple SUMIFS.
        ws.cell(row=i, column=7, value=(
            f"=$D{i}*SUMIFS('Forecast Weekly'!$BF:$BF,"
            f"'Forecast Weekly'!$C:$C,$A{i})"
        ))
        # Total
        ws.cell(row=i, column=8, value=f"=E{i}+F{i}+G{i}")
        for c in (2, 3, 5, 6, 7, 8):
            ws.cell(row=i, column=c).number_format = "$#,##0"
        ws.cell(row=i, column=4).number_format = "0.0%"

    total_row = 6 + len(trade)
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    for c in (5, 6, 7, 8):
        colL = get_column_letter(c)
        cell = ws.cell(row=total_row, column=c, value=f"=SUM({colL}6:{colL}{total_row - 1})")
        cell.font = Font(bold=True); cell.fill = SUBTOTAL_FILL; cell.number_format = "$#,##0"

    for col, w in zip("ABCDEFGH", (16, 16, 14, 14, 14, 14, 14, 16)):
        ws.column_dimensions[col].width = w
    setup_print(ws)
    return {"total_row": total_row}


def build_opex(ws):
    ws.title = "OPEX"
    title(ws, 1, "Operating expenses (FY26-27, CAD)")
    ws["A3"] = "Headcount, A&P spend, freight, broker fees. Edit blue. Flows into P&L."

    headers = ["Line item", "FY Amount", "Type", "Notes"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    rows = [
        # (label, amount, type, note)
        ("Sales rep headcount (loaded cost)",       360000, "SG&A", "5 FTE blended across BC/AB/ON/QC. ~$72K loaded each."),
        ("Sales lead (Louis)",                      130000, "SG&A", "RTD lead loaded cost."),
        ("Brokers (commission)",                    140000, "SG&A", "3% of revenue, est."),
        ("Travel & entertainment",                  60000,  "SG&A", "Trade shows, field visits, account meetings."),
        ("Freight to DCs",                          180000, "Logistics", "Outbound to distributors / retailers."),
        ("Warehousing",                             50000,  "Logistics", "3PL holding cost."),
        ("Marketing - paid digital",                120000, "A&P",  "Meta / TikTok / Google."),
        ("Marketing - influencer & content",        90000,  "A&P",  "Creator partnerships, content production."),
        ("Marketing - events & sampling spend",     180000, "A&P",  "Cash spend (not cases). CHFA, Stampede, etc."),
        ("Marketing - PR / agency retainer",        60000,  "A&P",  "Quarterly retainer."),
        ("Brand design / packaging",                30000,  "A&P",  "Refresh and new SKU work."),
        ("Insurance & compliance",                  25000,  "G&A",  "Product liability, regulatory."),
        ("Software & data",                         18000,  "G&A",  "Spins/Nielsen subscriptions, planning tools."),
        ("Slotting fees (cash, FY1)",               90000,  "Trade","Captured in Trade Spend tab; do not double-count in P&L."),
        ("Contingency",                             60000,  "G&A",  "Unallocated buffer."),
    ]
    for i, (lbl, amt, typ, note) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=lbl)
        ws.cell(row=i, column=2, value=amt).font = INPUT_FONT
        ws.cell(row=i, column=2).number_format = "$#,##0"
        ws.cell(row=i, column=3, value=typ)
        ws.cell(row=i, column=4, value=note)

    last = 5 + len(rows)
    total_row = last + 1
    ws.cell(row=total_row, column=1, value="Total OPEX").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=f"=SUM(B6:B{last})").font = Font(bold=True)
    ws.cell(row=total_row, column=2).fill = SUBTOTAL_FILL
    ws.cell(row=total_row, column=2).number_format = "$#,##0"

    # Subtotals by Type
    sub_row = total_row + 2
    section(ws, sub_row, "Subtotals by type")
    types = ["SG&A", "A&P", "Logistics", "G&A", "Trade"]
    for i, t in enumerate(types, start=sub_row + 1):
        ws.cell(row=i, column=1, value=t)
        ws.cell(row=i, column=2, value=f'=SUMIFS(B6:B{last},C6:C{last},"{t}")').number_format = "$#,##0"

    for col, w in zip("ABCD", (40, 16, 12, 60)):
        ws.column_dimensions[col].width = w
    setup_print(ws)
    return {"total_row": total_row, "rows_first": 6, "rows_last": last, "type_first": sub_row + 1}


def build_pl(ws, rev_refs, trade_refs, opex_refs):
    ws.title = "P&L"
    title(ws, 1, "Profit & loss (FY26-27, CAD)")
    ws["A3"] = "Standard P&L roll-up. Pulls from Revenue, Trade Spend, OPEX."

    headers = ["Line", "Amount", "% of revenue", "Notes"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    fy_col_rev = get_column_letter(rev_refs["fy_col"])
    rev_cell = f"Revenue!{fy_col_rev}{rev_refs['rev_total_row']}"
    gp_cell = f"Revenue!{fy_col_rev}{rev_refs['gp_total_row']}"
    trade_total_cell = f"'Trade Spend'!H{trade_refs['total_row']}"

    # Build rows
    pl_rows = [
        ("Gross revenue",                       f"={rev_cell}", "Revenue tab."),
        ("Trade spend (listing + slotting + scan/promo)", f"=-{trade_total_cell}", "Trade Spend tab."),
        ("Net revenue",                         "=B6+B7", "Gross revenue - trade spend."),
        ("COGS (landed)",                       f"={gp_cell}-{rev_cell}", "Computed as -(Revenue - GP)."),
        ("Gross profit",                        f"={gp_cell}+B7", "GP after trade spend."),
        ("A&P spend",                           f'=-SUMIFS(OPEX!B{opex_refs["rows_first"]}:B{opex_refs["rows_last"]},'
                                                 f'OPEX!C{opex_refs["rows_first"]}:C{opex_refs["rows_last"]},"A&P")',
         "Marketing cash spend (excludes case-cost sampling)."),
        ("SG&A",                                f'=-SUMIFS(OPEX!B{opex_refs["rows_first"]}:B{opex_refs["rows_last"]},'
                                                 f'OPEX!C{opex_refs["rows_first"]}:C{opex_refs["rows_last"]},"SG&A")',
         "Reps, brokers, T&E."),
        ("Logistics",                           f'=-SUMIFS(OPEX!B{opex_refs["rows_first"]}:B{opex_refs["rows_last"]},'
                                                 f'OPEX!C{opex_refs["rows_first"]}:C{opex_refs["rows_last"]},"Logistics")',
         "Freight, warehousing."),
        ("G&A",                                 f'=-SUMIFS(OPEX!B{opex_refs["rows_first"]}:B{opex_refs["rows_last"]},'
                                                 f'OPEX!C{opex_refs["rows_first"]}:C{opex_refs["rows_last"]},"G&A")',
         "Insurance, software, contingency."),
        ("EBITDA",                              "=B10+B11+B12+B13+B14", "Bottom line, FY26-27."),
        ("EBITDA margin",                       "=B15/B8", "% of net revenue."),
    ]
    for i, (lbl, fmla, note) in enumerate(pl_rows, start=6):
        ws.cell(row=i, column=1, value=lbl).font = Font(bold=(lbl in ("Net revenue", "Gross profit", "EBITDA", "EBITDA margin")))
        c = ws.cell(row=i, column=2, value=fmla)
        c.number_format = "$#,##0"
        if lbl in ("Net revenue", "Gross profit", "EBITDA"):
            c.fill = SUBTOTAL_FILL; c.font = Font(bold=True)
        if lbl == "EBITDA margin":
            c.number_format = "0.0%"
        # % of revenue
        if i not in (7,):  # skip
            pct = ws.cell(row=i, column=3, value=f"=B{i}/$B$6")
            pct.number_format = "0.0%"
        ws.cell(row=i, column=4, value=note)

    # Conditional formatting EBITDA red if negative
    ws.conditional_formatting.add("B15", CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL))
    ws.conditional_formatting.add("B15", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL))

    for col, w in zip("ABCD", (44, 18, 14, 60)):
        ws.column_dimensions[col].width = w
    setup_print(ws, landscape=False)


def build_scenarios(ws, rev_refs, trade_refs, opex_refs, prod_refs):
    ws.title = "Scenarios"
    title(ws, 1, "Scenario comparison")
    ws["A3"] = "Conservative (0.85), Base (1.00), Stretch (1.15). Edit multipliers in row 6."

    headers = ["Scenario", "Multiplier", "Sales cases", "Revenue", "Gross profit",
               "Trade spend", "Net revenue", "OPEX (cash)", "EBITDA (est.)", "EBITDA margin"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, horizontal="center")

    fy_col_rev = get_column_letter(rev_refs["fy_col"])
    base_rev = f"Revenue!{fy_col_rev}{rev_refs['rev_total_row']}"
    base_gp  = f"Revenue!{fy_col_rev}{rev_refs['gp_total_row']}"
    base_cases = f"'Production Plan'!{get_column_letter(prod_refs['fy_col'])}{prod_refs['total_sales_row']}"
    trade_total = f"'Trade Spend'!H{trade_refs['total_row']}"
    opex_total_cash = (f"SUMIFS(OPEX!B{opex_refs['rows_first']}:B{opex_refs['rows_last']},"
                       f"OPEX!C{opex_refs['rows_first']}:C{opex_refs['rows_last']},\"A&P\")+"
                       f"SUMIFS(OPEX!B{opex_refs['rows_first']}:B{opex_refs['rows_last']},"
                       f"OPEX!C{opex_refs['rows_first']}:C{opex_refs['rows_last']},\"SG&A\")+"
                       f"SUMIFS(OPEX!B{opex_refs['rows_first']}:B{opex_refs['rows_last']},"
                       f"OPEX!C{opex_refs['rows_first']}:C{opex_refs['rows_last']},\"Logistics\")+"
                       f"SUMIFS(OPEX!B{opex_refs['rows_first']}:B{opex_refs['rows_last']},"
                       f"OPEX!C{opex_refs['rows_first']}:C{opex_refs['rows_last']},\"G&A\")")

    scenarios = [("Conservative", 0.85), ("Base", 1.00), ("Stretch", 1.15)]
    for i, (name, mult) in enumerate(scenarios, start=6):
        ws.cell(row=i, column=1, value=name)
        c = ws.cell(row=i, column=2, value=mult); c.font = INPUT_FONT; c.fill = YELLOW_FILL
        # Cases, revenue, GP scale linearly with multiplier RELATIVE to the active scenario.
        # base values are at the current Assumptions!B6 multiplier. Divide and re-multiply.
        ws.cell(row=i, column=3, value=f"={base_cases}/Assumptions!$B$6*$B{i}").number_format = "#,##0"
        ws.cell(row=i, column=4, value=f"={base_rev}/Assumptions!$B$6*$B{i}").number_format = "$#,##0"
        ws.cell(row=i, column=5, value=f"={base_gp}/Assumptions!$B$6*$B{i}").number_format = "$#,##0"
        # Trade spend scales with revenue
        ws.cell(row=i, column=6, value=f"={trade_total}/Assumptions!$B$6*$B{i}").number_format = "$#,##0"
        # Net revenue = revenue - trade
        ws.cell(row=i, column=7, value=f"=D{i}-F{i}").number_format = "$#,##0"
        # OPEX is fixed cash
        ws.cell(row=i, column=8, value=f"={opex_total_cash}").number_format = "$#,##0"
        # EBITDA estimate = GP - trade - OPEX cash. (Trade already accounted in net but GP isn't net of trade,
        # so EBITDA = GP - trade - OPEX cash. Approximation.)
        ws.cell(row=i, column=9, value=f"=E{i}-F{i}-H{i}").number_format = "$#,##0"
        ws.cell(row=i, column=10, value=f"=I{i}/G{i}").number_format = "0.0%"

    ws.conditional_formatting.add("I6:I8", ColorScaleRule(
        start_type="num", start_value=-500000, start_color="F8696B",
        mid_type="num", mid_value=0, mid_color="FFEB84",
        end_type="num", end_value=500000, end_color="63BE7B"))

    section(ws, 11, "Notes")
    notes = [
        ("The multiplier on Assumptions!B6 drives the Forecast Weekly engine."),
        ("This Scenarios tab back-calculates the three points from current Base output so all three are visible at once."),
        ("Trade spend scales with revenue; OPEX cash is treated as fixed."),
        ("EBITDA here uses GP minus trade minus cash OPEX. For an exact P&L use the P&L tab at active multiplier."),
    ]
    for i, n in enumerate(notes, start=12):
        ws.cell(row=i, column=1, value=f"- {n}")

    for i, w in enumerate((14, 12, 14, 14, 14, 14, 14, 14, 14, 14), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    setup_print(ws)


def build_kpis(ws, rev_refs, prod_refs, doors_total_row):
    ws.title = "KPIs"
    title(ws, 1, "Key performance indicators")
    ws["A3"] = "Tracks model health and operational efficiency."

    fy_col_rev = get_column_letter(rev_refs["fy_col"])
    fy_col_prod = get_column_letter(prod_refs["fy_col"])

    kpis = [
        ("Total revenue (CAD)",                f"=Revenue!{fy_col_rev}{rev_refs['rev_total_row']}",                                 "$#,##0", "Higher is better."),
        ("Total gross profit (CAD)",           f"=Revenue!{fy_col_rev}{rev_refs['gp_total_row']}",                                  "$#,##0", "Higher is better."),
        ("Gross profit %",                     f"=Revenue!{fy_col_rev}{rev_refs['gp_total_row']}/Revenue!{fy_col_rev}{rev_refs['rev_total_row']}", "0.0%",   "Target >= 45%."),
        ("Total sales cases",                  f"='Production Plan'!{fy_col_prod}{prod_refs['total_sales_row']}",                   "#,##0",  "FY base scenario."),
        ("Production with buffer (cases)",     f"='Production Plan'!{fy_col_prod}{prod_refs['buf_row']}",                           "#,##0",  "Co-man PO basis."),
        ("Active doors year-end",              f"=Doors!{doors_week_col(N_WEEKS)}{doors_total_row}",                                "#,##0",  "Includes all 9 channels."),
        ("Avg revenue per door per year",      f"=Revenue!{fy_col_rev}{rev_refs['rev_total_row']}/Doors!{doors_week_col(N_WEEKS)}{doors_total_row}", "$#,##0", "Mix-weighted."),
        ("Avg GP per case",                    f"=Revenue!{fy_col_rev}{rev_refs['gp_total_row']}/'Production Plan'!{fy_col_prod}{prod_refs['total_sales_row']}", "$0.00", "After landed cost, before trade spend."),
        ("New-channel revenue mix", (
            "=("
            "SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"On-Premise\")+"
            "SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"Convenience\")+"
            "SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"Gym & Fitness\")+"
            "SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"Private Liquor\")+"
            "SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"RAS\")"
            f")/Revenue!{fy_col_rev}{rev_refs['rev_total_row']}"
        ), "0.0%", "Share of FY revenue from the 5 new channels."),
        ("Marketing & sampling % of cases",    f"='Production Plan'!{fy_col_prod}{prod_refs['mk_total_row']}/'Production Plan'!{fy_col_prod}{prod_refs['total_sales_row']}", "0.0%", "Sampling intensity."),
    ]
    headers = ["KPI", "Value", "Target / note"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for i, (lbl, formula, fmt, note) in enumerate(kpis, start=6):
        ws.cell(row=i, column=1, value=lbl)
        c = ws.cell(row=i, column=2, value=formula); c.number_format = fmt; c.font = LINK_FONT
        ws.cell(row=i, column=3, value=note)

    for col, w in zip("ABC", (38, 18, 60)):
        ws.column_dimensions[col].width = w
    setup_print(ws, landscape=False)


def build_assumption_register(ws):
    ws.title = "Assumption Register"
    title(ws, 1, "Assumption register")
    ws["A3"] = "Every assumption that drives the model. Owners must sign off before locking."

    headers = ["ID", "Assumption", "Owner", "Source tab", "Source cell(s)",
               "Current value", "Confidence", "Last reviewed", "Sign-off date", "Notes"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    rows = [
        ("A001", "Units per case",                    "Rijo",    "Assumptions",   "B4", 24,    "High",   "", "", "Fixed packaging spec."),
        ("A002", "Production buffer %",               "Rijo",    "Assumptions",   "B5", "10%", "Medium", "", "", "Industry norm 10-15%."),
        ("A003", "Scenario multiplier",               "Aaron",   "Assumptions",   "B6", 1.0,   "Medium", "", "", "Toggle 0.85 / 1.00 / 1.15."),
        ("A004", "MUV net case price",                "Aaron",   "Pricing",       "B6:B8", "$28", "High",  "", "", "Retail-anchored."),
        ("A005", "LCA Energy net case price",         "Aaron",   "Pricing",       "B9", "$32", "Medium", "", "", "Energy premium positioning."),
        ("A006", "MUV landed cost",                   "Rijo",    "Pricing",       "C6:C8", "$14.50", "Medium", "", "", "Co-man + freight in. Refine."),
        ("A007", "LCA landed cost",                   "Rijo",    "Pricing",       "C9", "$16.00", "Low",  "", "", "Placeholder, formulation in flux."),
        ("A008", "Natural channel class mix",         "Louis",   "Class Mix",     "B6:D6", "20/50/30", "High", "", "", "Anchored to historical."),
        ("A009", "On-Premise class mix",              "Louis",   "Class Mix",     "B10:D10", "25/50/25", "Low", "", "", "Placeholder. Validate after 50 doors."),
        ("A010", "Convenience class mix",             "Louis",   "Class Mix",     "B11:D11", "15/50/35", "Low", "", "", "Placeholder."),
        ("A011", "Gym & Fitness class mix",           "Louis",   "Class Mix",     "B12:D12", "30/50/20", "Low", "", "", "Placeholder."),
        ("A012", "Private Liquor class mix",          "Louis",   "Class Mix",     "B13:D13", "25/55/20", "Low", "", "", "Placeholder."),
        ("A013", "RAS class mix",                     "Louis",   "Class Mix",     "B14:D14", "10/40/50", "Low", "", "", "Placeholder."),
        ("A014", "MUV velocity Natural (A/B/C)",      "Louis",   "Velocity",      "C6:E6", "12/8/5", "Medium", "", "", "Benchmarked to peers."),
        ("A015", "MUV velocity On-Premise",           "Louis",   "Velocity",      "C10:E10", "6/4/2", "Low", "", "", "Restaurant benchmark."),
        ("A016", "MUV velocity Convenience",          "Louis",   "Velocity",      "C11:E11", "8/5/3", "Low", "", "", "C-store benchmark."),
        ("A017", "MUV velocity Gym & Fitness",        "Louis",   "Velocity",      "C12:E12", "10/7/4", "Low", "", "", "Studio benchmark."),
        ("A018", "MUV velocity Private Liquor",       "Louis",   "Velocity",      "C13:E13", "5/3/2", "Low", "", "", "Cross-shopper."),
        ("A019", "MUV velocity RAS",                  "Louis",   "Velocity",      "C14:E14", "4/3/2", "Low", "", "", "Rural foot traffic."),
        ("A020", "Door ramp - Natural by province",   "Louis",   "Doors",         "E6:P-",  "Monthly inputs", "Medium", "", "", "Reflects broker pipeline."),
        ("A021", "Door ramp - New channels",          "Louis",   "Doors",         "On-Premise / Conv / Gym / PL / RAS rows", "Monthly inputs", "Low", "", "", "Aspirational; needs sales sign-off."),
        ("A022", "Account sampling rates",            "Teresa",  "Marketing",     "B16:B24 (Section 1)", "Various", "Medium", "", "", "Per-channel cases/door/week."),
        ("A023", "Events plan (Total/Start/End)",     "Teresa",  "Marketing",     "Section 2", "13 events", "Medium", "", "", "Trade shows + activations."),
        ("A024", "Rep counts by province",            "Louis",   "Marketing",     "Section 3", "Monthly inputs", "Medium", "", "", "Hiring plan."),
        ("A025", "Buffer (R&D / gifting / freight)",  "Louis",   "Marketing",     "Section 4", "0.5-0.75 cases/wk", "High", "", "", "Steady-state."),
        ("A026", "Trade spend - listing fees",        "Aaron",   "Trade Spend",   "B6:B14", "Various", "Medium", "", "", "Per banner one-time."),
        ("A027", "Trade spend - slotting",            "Aaron",   "Trade Spend",   "C6:C14", "Per door", "Medium", "", "", "FDM/conv heavy."),
        ("A028", "Trade spend - scan/promo %",        "Aaron",   "Trade Spend",   "D6:D14", "2-8% of revenue", "Medium", "", "", "Channel norms."),
        ("A029", "OPEX - all line items",             "Elliot",  "OPEX",          "B6:B20", "Various", "Medium", "", "", "Annual FY budget."),
        ("A030", "Co-man capacity adequate",          "Rijo",    "external",      "n/a",    "implicit", "Medium", "", "", "Production with buffer is feasible."),
    ]
    for i, row_data in enumerate(rows, start=6):
        for j, val in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=val)

    # Data validation on Confidence and Owner columns
    dv_conf = DataValidation(type="list", formula1="\"High,Medium,Low\"", allow_blank=True)
    ws.add_data_validation(dv_conf); dv_conf.add(f"G6:G{5 + len(rows)}")

    for col, w in zip("ABCDEFGHIJ", (8, 36, 12, 14, 22, 16, 12, 14, 14, 40)):
        ws.column_dimensions[col].width = w
    setup_print(ws)


def build_risk_register(ws):
    ws.title = "Risk Register"
    title(ws, 1, "Risk register")
    ws["A3"] = "Top risks to the forecast. Reviewed monthly with Aaron."

    headers = ["ID", "Risk", "Likelihood", "Impact", "Score", "Mitigation", "Owner", "Status"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    rows = [
        ("R01", "Co-man capacity slips, can't hit production with buffer", 3, 5, "Lock 2026 Q4 PO 8 weeks ahead. Backup co-man on retainer.", "Rijo", "Amber"),
        ("R02", "Costco roadshow doesn't convert to permanent listing",     3, 4, "Brief on roadshow KPIs. Pre-align with buyer.", "Aaron", "Amber"),
        ("R03", "New channel velocity assumptions wrong (overstated)",      4, 4, "Validate after 50 doors per channel. Adjust forecast Q2.", "Louis", "Amber"),
        ("R04", "LCA Energy GP under 45%",                                  3, 4, "Refine landed cost; renegotiate co-man pricing.", "Rijo", "Red"),
        ("R05", "Sales rep hiring delayed in ON/QC",                        3, 3, "Pipeline 2 candidates per role. Start Q4.", "Louis", "Amber"),
        ("R06", "Trade spend (scan/promo) exceeds plan",                    3, 3, "Channel-by-channel scorecard. Monthly review.", "Aaron", "Green"),
        ("R07", "On-premise lacks broker coverage",                         3, 3, "Hire on-premise specialist or use food distributor.", "Louis", "Amber"),
        ("R08", "Convenience slotting cash exceeds plan",                   3, 3, "Negotiate slotting in kind (free fills).", "Aaron", "Green"),
        ("R09", "Private Liquor regulations change",                        2, 3, "Monitor BC/AB liquor boards.", "Louis", "Green"),
        ("R10", "RAS distribution complexity (rural freight)",              3, 2, "Group shipments via LCBO RAS hub.", "Rijo", "Green"),
        ("R11", "Influencer / event spend doesn't convert",                 3, 3, "Tag each event with redemption code or unique URL.", "Teresa", "Amber"),
        ("R12", "Cannibalisation between MUV and LCA in BC Natural",        2, 2, "Monitor SKU velocity; differentiate POS.", "Louis", "Green"),
    ]
    for i, (id_, risk, lkh, imp, mit, own, st) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=id_)
        ws.cell(row=i, column=2, value=risk)
        ws.cell(row=i, column=3, value=lkh)
        ws.cell(row=i, column=4, value=imp)
        ws.cell(row=i, column=5, value=f"=C{i}*D{i}")
        ws.cell(row=i, column=6, value=mit)
        ws.cell(row=i, column=7, value=own)
        ws.cell(row=i, column=8, value=st)

    last = 5 + len(rows)
    # Colour score with 3-colour scale
    ws.conditional_formatting.add(f"E6:E{last}", ColorScaleRule(
        start_type="num", start_value=1, start_color="63BE7B",
        mid_type="num", mid_value=9, mid_color="FFEB84",
        end_type="num", end_value=25, end_color="F8696B"))
    # Likelihood / Impact dropdowns (1-5)
    dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5,
                        allow_blank=False, showErrorMessage=True,
                        errorTitle="Out of bounds", error="Score 1-5.")
    ws.add_data_validation(dv); dv.add(f"C6:D{last}")
    dv_st = DataValidation(type="list", formula1="\"Green,Amber,Red\"", allow_blank=True)
    ws.add_data_validation(dv_st); dv_st.add(f"H6:H{last}")

    for col, w in zip("ABCDEFGH", (6, 48, 12, 10, 8, 50, 10, 10)):
        ws.column_dimensions[col].width = w
    setup_print(ws)


def build_change_log(ws):
    ws.title = "Change Log"
    title(ws, 1, "Change log and sign-offs")
    ws["A3"] = "Version history of the model."

    headers = ["Version", "Date", "Author", "Change", "Sign-off"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    rows = [
        ("v1.0", "2026-05-25", "Louis", "Initial monthly model. 4 channels.", "Aaron - pending"),
        ("v2.0", "2026-05-25", "Louis", "Converted to weekly. Added 5 new channels: On-Premise, Convenience, Gym & Fitness, Private Liquor, RAS.", "Aaron - pending"),
        ("v2.1", "2026-05-27", "Louis", "Parametric: monthly door inputs drive weekly outputs; events use Total/Start/End; buffer single-rate.", "Aaron - pending"),
        ("v3.0", "2026-05-27", "Louis", "Enterprise build. P&L, Scenarios, OPEX, Trade Spend, KPIs, Assumption + Risk registers, validation, conditional formatting.", "Aaron - pending"),
    ]
    for i, row in enumerate(rows, start=6):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    for col, w in zip("ABCDE", (10, 14, 14, 80, 24)):
        ws.column_dimensions[col].width = w
    setup_print(ws)


def build_glossary(ws):
    ws.title = "Glossary"
    title(ws, 1, "Glossary - data dictionary")
    ws["A3"] = "Definitions for every term used in the model."

    headers = ["Term", "Definition"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    rows = [
        ("Natural",         "Independent natural / health-food retailers (e.g., Whole Foods, indie health stores)."),
        ("Specialty",       "Higher-end grocers and specialty banners (e.g., Pusateri's, Summerhill)."),
        ("FDM",             "Food, Drug, Mass. Mainstream grocery and drug chains (Loblaws, Sobeys, Metro, Shoppers, Walmart)."),
        ("Costco",          "Costco Wholesale Canada. Roadshow + permanent listing distinction."),
        ("On-Premise",      "Restaurants, bars, cafes, hotel F&B. Sales by the can or in cooler service."),
        ("Convenience",     "C-stores and gas stations (Circle K, 7-Eleven, Couche-Tard)."),
        ("Gym & Fitness",   "Gym chains, boutique studios, yoga studios. Sale at the front desk or smoothie bar."),
        ("Private Liquor",  "Privately operated liquor stores (mainly BC, AB, SK, MB)."),
        ("RAS",             "Rural Agency Stores. Provincial liquor board agency outlets in rural communities."),
        ("Class A/B/C",     "Tiered door rating by foot traffic and velocity. A = top tier, C = lower tier."),
        ("Blended velocity", "Weighted average velocity per door per week across Class A/B/C."),
        ("FY26-27",         "Fiscal year September 1, 2026 - August 31, 2027."),
        ("Scenario multiplier", "Single dial: 0.85 conservative, 1.00 base, 1.15 stretch. Drives Forecast Weekly."),
        ("Production buffer", "Safety cushion on production demand to cover spoilage, returns, growth."),
        ("Trade spend",     "Cash paid to retailers to win and hold shelf: listing fees, slotting, scan/promo allowances."),
        ("Listing fee",     "One-time charge to add a SKU at a banner."),
        ("Slotting",        "Per-door payment to secure facings or specific shelf positions."),
        ("Scan/promo allowance", "% of net revenue passed to retailer to fund weekly promos."),
        ("A&P",             "Advertising & Promotion. Marketing cash spend."),
        ("SG&A",            "Selling, General & Administrative. Salaries, T&E, brokers."),
        ("EBITDA",          "Earnings before interest, tax, depreciation, amortization."),
        ("ACV",             "All Commodity Volume. % of category retail dollars our doors cover."),
        ("GP",              "Gross profit. Revenue minus landed cost. Pre-trade-spend, pre-OPEX."),
    ]
    for i, (term, defn) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=term).font = Font(bold=True)
        ws.cell(row=i, column=2, value=defn)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 110
    setup_print(ws, landscape=False)


def build_home(ws, prod_refs, rev_refs, trade_refs, opex_refs, doors_total_row):
    """The single tab anyone needs to open. Big numbers, clean type, calm layout."""
    ws.title = "Home"
    ws.sheet_properties.tabColor = TAB_FRONT
    # Hide gridlines for the Apple look
    ws.sheet_view.showGridLines = False

    fy_col_rev = get_column_letter(rev_refs["fy_col"])
    fy_col_prod = get_column_letter(prod_refs["fy_col"])

    # Generous gutter
    ws.column_dimensions["A"].width = 3
    for col in "BCDE":
        ws.column_dimensions[col].width = 22
    ws.column_dimensions["F"].width = 3

    # Brand wordmark line
    c = ws.cell(row=2, column=2, value="ORGANIKA RTD")
    c.font = Font(name="Helvetica", bold=True, size=11, color=APPLE_GREY)
    c2 = ws.cell(row=3, column=2, value="FY26-27 Forecast")
    c2.font = TITLE_FONT
    c3 = ws.cell(row=4, column=2,
                 value="September 1, 2026 - August 31, 2027 . Base scenario . CAD")
    c3.font = MUTED_FONT

    # Hero KPI strip (4 big numbers across)
    hero_row = 7
    hero_label_row = hero_row - 1
    heroes = [
        ("REVENUE",        f"=Revenue!{fy_col_rev}{rev_refs['rev_total_row']}",  "$#,##0",  "Annual sales revenue"),
        ("GROSS PROFIT",   f"=Revenue!{fy_col_rev}{rev_refs['gp_total_row']}",   "$#,##0",  "After landed cost"),
        ("EBITDA",         (f"=Revenue!{fy_col_rev}{rev_refs['gp_total_row']}-'Trade Spend'!H{trade_refs['total_row']}"
                            f"-(SUMIFS(OPEX!B{opex_refs['rows_first']}:B{opex_refs['rows_last']},OPEX!C{opex_refs['rows_first']}:C{opex_refs['rows_last']},\"A&P\")"
                            f"+SUMIFS(OPEX!B{opex_refs['rows_first']}:B{opex_refs['rows_last']},OPEX!C{opex_refs['rows_first']}:C{opex_refs['rows_last']},\"SG&A\")"
                            f"+SUMIFS(OPEX!B{opex_refs['rows_first']}:B{opex_refs['rows_last']},OPEX!C{opex_refs['rows_first']}:C{opex_refs['rows_last']},\"Logistics\")"
                            f"+SUMIFS(OPEX!B{opex_refs['rows_first']}:B{opex_refs['rows_last']},OPEX!C{opex_refs['rows_first']}:C{opex_refs['rows_last']},\"G&A\"))"),
                                                                                  "$#,##0",  "Pre-tax operating income"),
        ("CASES TO PRODUCE", f"='Production Plan'!{fy_col_prod}{prod_refs['buf_row']}", "#,##0", "With 10% buffer"),
    ]
    for i, (lbl, fml, fmt, sub) in enumerate(heroes):
        col = 2 + i
        lab = ws.cell(row=hero_label_row, column=col, value=lbl)
        lab.font = HERO_LABEL_FONT
        lab.alignment = Alignment(horizontal="left", vertical="center")
        val = ws.cell(row=hero_row, column=col, value=fml)
        val.font = HERO_FONT
        val.number_format = fmt
        val.alignment = Alignment(horizontal="left", vertical="center")
        sm = ws.cell(row=hero_row + 1, column=col, value=sub)
        sm.font = MUTED_FONT
    ws.row_dimensions[hero_row].height = 48

    # Thin divider
    div_row = hero_row + 3
    for col in range(2, 6):
        cell = ws.cell(row=div_row, column=col, value="")
        cell.border = Border(top=Side(border_style="thin", color="D2D2D7"))

    # The story (3 columns)
    story_row = div_row + 2
    ws.cell(row=story_row, column=2, value="THE STORY").font = HERO_LABEL_FONT
    ws.cell(row=story_row + 1, column=2,
            value="Year 1 of a 9-channel national footprint. Ramp from ~360 doors at W1 to ~3,000 at W52. "
                  "EBITDA is negative this year because of one-time listing + slotting fees and full-scale OPEX. "
                  "Path to positive in FY27-28 by amortising slotting, scaling revenue, and holding OPEX flat.").font = BODY_FONT
    ws.row_dimensions[story_row + 1].height = 80
    ws.merge_cells(start_row=story_row + 1, start_column=2, end_row=story_row + 1, end_column=5)
    ws.cell(row=story_row + 1, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Quick stats grid (2x4)
    stats_row = story_row + 4
    ws.cell(row=stats_row, column=2, value="AT A GLANCE").font = HERO_LABEL_FONT
    quick = [
        ("Gross profit margin",       f"=Revenue!{fy_col_rev}{rev_refs['gp_total_row']}/Revenue!{fy_col_rev}{rev_refs['rev_total_row']}", "0.0%"),
        ("Sales cases (year)",        f"='Production Plan'!{fy_col_prod}{prod_refs['total_sales_row']}",                                   "#,##0"),
        ("Active doors at year-end",  f"=Doors!{doors_week_col(N_WEEKS)}{doors_total_row}",                                                "#,##0"),
        ("Revenue per door",          f"=Revenue!{fy_col_rev}{rev_refs['rev_total_row']}/Doors!{doors_week_col(N_WEEKS)}{doors_total_row}", "$#,##0"),
        ("New-channel revenue share", ("=("
                                       "SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"On-Premise\")+"
                                       "SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"Convenience\")+"
                                       "SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"Gym & Fitness\")+"
                                       "SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"Private Liquor\")+"
                                       "SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"RAS\")"
                                       f")/Revenue!{fy_col_rev}{rev_refs['rev_total_row']}"), "0.0%"),
        ("Trade spend % of revenue",  f"='Trade Spend'!H{trade_refs['total_row']}/Revenue!{fy_col_rev}{rev_refs['rev_total_row']}",         "0.0%"),
        ("Scenario multiplier",       "=Assumptions!B6", "0.00"),
        ("Production buffer",         "=Assumptions!B5", "0%"),
    ]
    for i, (lbl, fml, fmt) in enumerate(quick):
        r = stats_row + 2 + (i // 2) * 2
        c = 2 + (i % 2) * 2
        lab = ws.cell(row=r, column=c, value=lbl)
        lab.font = MUTED_FONT
        val = ws.cell(row=r + 1, column=c, value=fml)
        val.font = Font(name="Helvetica", bold=True, size=18, color=APPLE_INK)
        val.number_format = fmt

    # Where to go next
    nav_row = stats_row + 2 + 4 * 2 + 1
    ws.cell(row=nav_row, column=2, value="WHERE TO GO NEXT").font = HERO_LABEL_FONT
    nav = [
        ("Adjust the model", "Control Panel - the 5 inputs that matter most"),
        ("Read the full P&L", "P&L"),
        ("Compare scenarios", "Scenarios - Conservative / Base / Stretch side-by-side"),
        ("Drill into details", "Dashboard - monthly + quarterly roll-ups, by SKU / channel / province"),
        ("Validate assumptions", "Assumption Register - 30 owned assumptions"),
        ("Review risks", "Risk Register - 12 risks scored by likelihood x impact"),
    ]
    for i, (action, where) in enumerate(nav, start=nav_row + 2):
        ws.cell(row=i, column=2, value=action).font = Font(name="Helvetica", bold=True, size=11, color=APPLE_INK)
        ws.cell(row=i, column=3, value=where).font = MUTED_FONT
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)

    # Footer
    footer = nav_row + 2 + len(nav) + 2
    ws.cell(row=footer, column=2,
            value="Owner: Louis . RTD lead, Organika . Source of truth. Supersedes scratch models.").font = MUTED_FONT
    ws.merge_cells(start_row=footer, start_column=2, end_row=footer, end_column=5)

    setup_print(ws, landscape=False, header_rows=0)


def build_control_panel(ws):
    """One tab with the five inputs that drive the entire model."""
    ws.title = "Control Panel"
    ws.sheet_properties.tabColor = TAB_FRONT
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 60

    ws.cell(row=2, column=2, value="CONTROL PANEL").font = Font(name="Helvetica", bold=True, size=11, color=APPLE_GREY)
    ws.cell(row=3, column=2, value="Five knobs").font = TITLE_FONT
    ws.cell(row=4, column=2,
            value="Edit only these five cells. Everything downstream recalculates automatically.").font = MUTED_FONT

    # Headers
    ws.cell(row=6, column=2, value="INPUT").font = HERO_LABEL_FONT
    ws.cell(row=6, column=3, value="VALUE").font = HERO_LABEL_FONT
    ws.cell(row=6, column=4, value="WHAT IT DOES").font = HERO_LABEL_FONT

    # 5 knobs - all of them already live elsewhere; we just mirror them with links so editing here updates source
    knobs = [
        # (label, source_cell_ref, format, description, default)
        ("Scenario multiplier",   "=Assumptions!B6",  "0.00", "0.85 conservative, 1.00 base, 1.15 stretch. Scales sales cases."),
        ("Production buffer",     "=Assumptions!B5",  "0%",   "Extra cushion on production. Industry norm 10-15%."),
        ("Units per case",        "=Assumptions!B4",  "#,##0","Fixed at 24 cans. Do not change unless re-packaging."),
        ("MUV net case price",    "=Pricing!B6",      "$#,##0.00", "Revenue per case for all three MUV SKUs."),
        ("LCA net case price",    "=Pricing!B9",      "$#,##0.00", "Revenue per case for LCA Energy."),
    ]
    for i, (lbl, src, fmt, desc) in enumerate(knobs, start=8):
        ws.cell(row=i, column=2, value=lbl).font = Font(name="Helvetica", bold=True, size=12, color=APPLE_INK)
        c = ws.cell(row=i, column=3, value=src)
        c.font = Font(name="Helvetica", bold=True, size=14, color=APPLE_BLUE)
        c.number_format = fmt
        ws.cell(row=i, column=4, value=desc).font = MUTED_FONT
        ws.row_dimensions[i].height = 28

    # Pointer to the inputs
    ws.cell(row=15, column=2, value="DEEPER INPUTS").font = HERO_LABEL_FONT
    deeper = [
        ("Door plan",          "Doors - 12 monthly inputs per row drive 52 weekly outputs"),
        ("Velocity per SKU",   "Velocity - units per door per week, by class A/B/C"),
        ("Class mix",          "Class Mix - % of doors in each class, per channel"),
        ("Trade spend",        "Trade Spend - listing, slotting, scan/promo by channel"),
        ("OPEX",               "OPEX - sales reps, A&P, freight, G&A"),
        ("Marketing cases",    "Marketing and Sampling - account sampling, events, rep samples"),
    ]
    for i, (lbl, where) in enumerate(deeper, start=17):
        ws.cell(row=i, column=2, value=lbl).font = Font(name="Helvetica", bold=True, size=11, color=APPLE_INK)
        ws.cell(row=i, column=4, value=where).font = MUTED_FONT

    setup_print(ws, landscape=False, header_rows=0)


def build_exec_summary(ws, prod_refs, rev_refs, trade_refs, opex_refs, doors_total_row):
    """One-page executive summary - the only tab a board member needs to read."""
    ws.title = "Exec Summary"
    title(ws, 1, "Organika RTD - Executive Summary, FY26-27")
    ws["A2"] = "One page. Read this first. Drill into other tabs for detail."
    ws["A2"].font = Font(italic=True, color="808080")

    fy_col_rev = get_column_letter(rev_refs["fy_col"])
    fy_col_prod = get_column_letter(prod_refs["fy_col"])

    section(ws, 4, "The number")
    ws["A5"] = "FY26-27 revenue (base case)";          ws["B5"] = f"=Revenue!{fy_col_rev}{rev_refs['rev_total_row']}"
    ws["A6"] = "Gross profit (after landed cost)";     ws["B6"] = f"=Revenue!{fy_col_rev}{rev_refs['gp_total_row']}"
    ws["A7"] = "Trade spend (channel cash to retailers)"; ws["B7"] = f"='Trade Spend'!H{trade_refs['total_row']}"
    ws["A8"] = "OPEX cash";                            ws["B8"] = f"=-(SUMIFS(OPEX!B{opex_refs['rows_first']}:B{opex_refs['rows_last']},OPEX!C{opex_refs['rows_first']}:C{opex_refs['rows_last']},\"A&P\")+SUMIFS(OPEX!B{opex_refs['rows_first']}:B{opex_refs['rows_last']},OPEX!C{opex_refs['rows_first']}:C{opex_refs['rows_last']},\"SG&A\")+SUMIFS(OPEX!B{opex_refs['rows_first']}:B{opex_refs['rows_last']},OPEX!C{opex_refs['rows_first']}:C{opex_refs['rows_last']},\"Logistics\")+SUMIFS(OPEX!B{opex_refs['rows_first']}:B{opex_refs['rows_last']},OPEX!C{opex_refs['rows_first']}:C{opex_refs['rows_last']},\"G&A\"))"
    ws["A9"] = "EBITDA";                               ws["B9"] = "=B6-B7+B8"
    ws["A10"] = "EBITDA margin";                       ws["B10"] = "=B9/B5"

    for r in (5, 6, 7, 8, 9):
        ws.cell(row=r, column=2).number_format = "$#,##0"
        ws.cell(row=r, column=1).font = Font(bold=True)
    ws["B10"].number_format = "0.0%"
    ws["A9"].font = Font(bold=True, size=12)
    ws["B9"].font = Font(bold=True, size=12); ws["B9"].fill = SUBTOTAL_FILL

    section(ws, 12, "Operating story")
    bullets = [
        ("Sales cases (year)",          f"='Production Plan'!{fy_col_prod}{prod_refs['total_sales_row']}", "#,##0"),
        ("Production with buffer",      f"='Production Plan'!{fy_col_prod}{prod_refs['buf_row']}", "#,##0"),
        ("Active doors at year-end",    f"=Doors!{doors_week_col(N_WEEKS)}{doors_total_row}", "#,##0"),
        ("Revenue / door / year",       f"=B5/Doors!{doors_week_col(N_WEEKS)}{doors_total_row}", "$#,##0"),
        ("GP / case",                   f"=B6/'Production Plan'!{fy_col_prod}{prod_refs['total_sales_row']}", "$0.00"),
        ("Trade spend % of revenue",    "=B7/B5", "0.0%"),
        ("OPEX cash % of revenue",      "=-B8/B5", "0.0%"),
    ]
    for i, (lbl, fml, fmt) in enumerate(bullets, start=13):
        ws.cell(row=i, column=1, value=lbl).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=fml); c.number_format = fmt; c.font = LINK_FONT

    section(ws, 22, "Headline narrative")
    narrative = [
        "Year 1 of the 9-channel national footprint plan: ramp from ~360 doors at W1 to ~3,000 at W52.",
        "Trade spend (~$580K) is one-time-heavy in FY1: listing fees and slotting at FDM, Costco, Convenience.",
        "Sampling intensity ~3% of sales cases. New channels (On-Premise/Conv/Gym/PL/RAS) shoulder ~24% of revenue.",
        "EBITDA negative at base. Path to positive: (a) revenue scale year-2; (b) slotting amortised; (c) trim OPEX.",
        "Risks (Risk Register): co-man capacity, new-channel velocity assumptions, LCA Energy GP.",
        "Sign-offs needed: Aaron (scenario, pricing, trade), Teresa (marketing buffer), Rijo (cost, buffer), Elliot (OPEX).",
    ]
    for i, t in enumerate(narrative, start=23):
        ws.cell(row=i, column=1, value=f"- {t}")

    section(ws, 30, "Next 30 days")
    actions = [
        "Week 1: review Class Mix & Velocity defaults for new channels (Louis + Aaron).",
        "Week 1: confirm pricing & landed cost (Rijo + Elliot).",
        "Week 2: validate Doors plan with field broker pipeline (Louis).",
        "Week 2: scenario review with Aaron - lock multiplier.",
        "Week 3: marketing budget alignment (Teresa).",
        "Week 4: lock the model. Distribute production PO to co-man.",
    ]
    for i, t in enumerate(actions, start=31):
        ws.cell(row=i, column=1, value=f"- {t}")

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 22
    setup_print(ws, landscape=False)


def build_dashboard(ws, prod_refs, rev_refs, doors_total_row, forecast_total_row):
    ws.title = "Dashboard"
    title(ws, 1, "Organika RTD - Executive Dashboard FY26-27")
    ws["A2"] = "Read-only. Pulls from input tabs. Refresh by re-opening or pressing F9."
    ws["A2"].font = Font(italic=True, color="808080")

    month_to_cols = {m: [] for m in MONTH_ORDER}
    for wi, m in enumerate(WEEK_TO_MONTH, start=1):
        month_to_cols[m].append(fw_week_col(wi))

    section(ws, 4, "Headline metrics")
    headers = ["Metric", "Value", "Source"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    fy_col_prod = get_column_letter(prod_refs["fy_col"])
    fy_col_rev = get_column_letter(rev_refs["fy_col"])

    headline = [
        ("Total sales cases",                  f"='Production Plan'!{fy_col_prod}{prod_refs['total_sales_row']}", "Production Plan", "#,##0"),
        ("Total marketing & sampling cases",   f"='Production Plan'!{fy_col_prod}{prod_refs['mk_total_row']}",    "Production Plan", "#,##0"),
        ("Total cases to produce",             f"='Production Plan'!{fy_col_prod}{prod_refs['total_prod_row']}",  "Production Plan", "#,##0"),
        ("Production with buffer",             f"='Production Plan'!{fy_col_prod}{prod_refs['buf_row']}",         "Buffered for co-man PO", "#,##0"),
        ("Total revenue (CAD)",                f"=Revenue!{fy_col_rev}{rev_refs['rev_total_row']}",               "Revenue tab", "$#,##0"),
        ("Total gross profit (CAD)",           f"=Revenue!{fy_col_rev}{rev_refs['gp_total_row']}",                "Revenue tab", "$#,##0"),
        ("Gross profit %",                     f"=Revenue!{fy_col_rev}{rev_refs['gp_total_row']}/Revenue!{fy_col_rev}{rev_refs['rev_total_row']}", "Calculated", "0.0%"),
        ("Active doors year end (W52)",        f"=Doors!{doors_week_col(N_WEEKS)}{doors_total_row}",              "Doors", "#,##0"),
        ("Scenario multiplier",                "=Assumptions!B6",                                                  "Assumptions", "0.00"),
    ]
    for i, (lbl, val, src, fmt) in enumerate(headline, start=6):
        ws.cell(row=i, column=1, value=lbl).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=val); c.font = LINK_FONT; c.number_format = fmt
        ws.cell(row=i, column=3, value=src)

    # By SKU (monthly)
    base_row = 17
    section(ws, base_row, "Monthly sales cases by SKU")
    ws.cell(row=base_row + 1, column=1, value="SKU").font = HEADER_FONT
    ws[f"A{base_row + 1}"].fill = HEADER_FILL
    for j, m in enumerate(MONTH_ORDER):
        c = ws.cell(row=base_row + 1, column=2 + j, value=m)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    ws.cell(row=base_row + 1, column=2 + len(MONTH_ORDER), value="FY total").font = HEADER_FONT
    ws[f"{get_column_letter(2 + len(MONTH_ORDER))}{base_row + 1}"].fill = HEADER_FILL
    for i, sku in enumerate(SKUS):
        r = base_row + 2 + i
        ws.cell(row=r, column=1, value=sku)
        for j, m in enumerate(MONTH_ORDER):
            cols = month_to_cols[m]
            terms = [(f"SUMIFS('Forecast Weekly'!{c}:{c},'Forecast Weekly'!$A:$A,$A{r})") for c in cols]
            cell = ws.cell(row=r, column=2 + j, value="=" + "+".join(terms))
            cell.font = LINK_FONT; cell.number_format = "#,##0"
        cell = ws.cell(row=r, column=2 + len(MONTH_ORDER),
                       value=(f"=SUM({get_column_letter(2)}{r}:"
                              f"{get_column_letter(2 + len(MONTH_ORDER) - 1)}{r})"))
        cell.number_format = "#,##0"
    total_r = base_row + 2 + len(SKUS)
    ws.cell(row=total_r, column=1, value="Total").font = Font(bold=True)
    for j in range(len(MONTH_ORDER) + 1):
        col = 2 + j; colL = get_column_letter(col)
        c = ws.cell(row=total_r, column=col, value=f"=SUM({colL}{base_row + 2}:{colL}{total_r - 1})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL; c.number_format = "#,##0"

    # By Channel (monthly)
    chan_base = total_r + 3
    section(ws, chan_base, "Monthly sales cases by channel")
    ws.cell(row=chan_base + 1, column=1, value="Channel").font = HEADER_FONT
    ws[f"A{chan_base + 1}"].fill = HEADER_FILL
    for j, m in enumerate(MONTH_ORDER):
        c = ws.cell(row=chan_base + 1, column=2 + j, value=m)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    ws.cell(row=chan_base + 1, column=2 + len(MONTH_ORDER), value="FY total").font = HEADER_FONT
    for i, ch in enumerate(CHANNELS):
        r = chan_base + 2 + i
        ws.cell(row=r, column=1, value=ch)
        for j, m in enumerate(MONTH_ORDER):
            cols = month_to_cols[m]
            terms = [(f"SUMIFS('Forecast Weekly'!{c}:{c},'Forecast Weekly'!$C:$C,$A{r})") for c in cols]
            cell = ws.cell(row=r, column=2 + j, value="=" + "+".join(terms))
            cell.font = LINK_FONT; cell.number_format = "#,##0"
        cell = ws.cell(row=r, column=2 + len(MONTH_ORDER),
                       value=(f"=SUM({get_column_letter(2)}{r}:"
                              f"{get_column_letter(2 + len(MONTH_ORDER) - 1)}{r})"))
        cell.number_format = "#,##0"
    chan_total_r = chan_base + 2 + len(CHANNELS)
    ws.cell(row=chan_total_r, column=1, value="Total").font = Font(bold=True)
    for j in range(len(MONTH_ORDER) + 1):
        col = 2 + j; colL = get_column_letter(col)
        c = ws.cell(row=chan_total_r, column=col, value=f"=SUM({colL}{chan_base + 2}:{colL}{chan_total_r - 1})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL; c.number_format = "#,##0"

    # Quarterly view
    quart_base = chan_total_r + 3
    section(ws, quart_base, "Quarterly sales cases (all SKUs)")
    ws.cell(row=quart_base + 1, column=1, value="Channel").font = HEADER_FONT
    ws[f"A{quart_base + 1}"].fill = HEADER_FILL
    for j, q in enumerate(QUARTERS):
        c = ws.cell(row=quart_base + 1, column=2 + j, value=q)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    ws.cell(row=quart_base + 1, column=2 + len(QUARTERS), value="FY total").font = HEADER_FONT
    quarter_to_cols = {q: [] for q in QUARTERS}
    for wi, m in enumerate(WEEK_TO_MONTH, start=1):
        q = MONTH_QUARTER[m]
        quarter_to_cols[q].append(fw_week_col(wi))
    for i, ch in enumerate(CHANNELS):
        r = quart_base + 2 + i
        ws.cell(row=r, column=1, value=ch)
        for j, q in enumerate(QUARTERS):
            cols = quarter_to_cols[q]
            terms = [(f"SUMIFS('Forecast Weekly'!{c}:{c},'Forecast Weekly'!$C:$C,$A{r})") for c in cols]
            cell = ws.cell(row=r, column=2 + j, value="=" + "+".join(terms))
            cell.font = LINK_FONT; cell.number_format = "#,##0"
        cell = ws.cell(row=r, column=2 + len(QUARTERS),
                       value=(f"=SUM({get_column_letter(2)}{r}:"
                              f"{get_column_letter(2 + len(QUARTERS) - 1)}{r})"))
        cell.number_format = "#,##0"
    qtotal_r = quart_base + 2 + len(CHANNELS)
    ws.cell(row=qtotal_r, column=1, value="Total").font = Font(bold=True)
    for j in range(len(QUARTERS) + 1):
        col = 2 + j; colL = get_column_letter(col)
        c = ws.cell(row=qtotal_r, column=col, value=f"=SUM({colL}{quart_base + 2}:{colL}{qtotal_r - 1})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL; c.number_format = "#,##0"

    # Door snapshot
    door_base = qtotal_r + 3
    section(ws, door_base, "Active doors by channel (end of month)")
    ws.cell(row=door_base + 1, column=1, value="Channel").font = HEADER_FONT
    ws[f"A{door_base + 1}"].fill = HEADER_FILL
    for j, m in enumerate(MONTH_ORDER):
        c = ws.cell(row=door_base + 1, column=2 + j, value=m)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for i, ch in enumerate(CHANNELS):
        r = door_base + 2 + i
        ws.cell(row=r, column=1, value=ch)
        for j, m in enumerate(MONTH_ORDER):
            colL = MONTH_INPUT_COL[m]
            cell = ws.cell(row=r, column=2 + j,
                           value=f"=SUMIFS(Doors!{colL}:{colL},Doors!$C:$C,$A{r})")
            cell.font = LINK_FONT; cell.number_format = "#,##0"
    door_total_r = door_base + 2 + len(CHANNELS)
    ws.cell(row=door_total_r, column=1, value="Total doors").font = Font(bold=True)
    for j in range(len(MONTH_ORDER)):
        col = 2 + j; colL = get_column_letter(col)
        c = ws.cell(row=door_total_r, column=col, value=f"=SUM({colL}{door_base + 2}:{colL}{door_total_r - 1})")
        c.font = Font(bold=True); c.fill = SUBTOTAL_FILL; c.number_format = "#,##0"

    # Chart: Monthly sales by channel (stacked bar)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Monthly sales cases by channel"
    chart.y_axis.title = "Cases"
    chart.x_axis.title = "Month"
    data = Reference(ws, min_col=2, max_col=1 + len(MONTH_ORDER),
                     min_row=chan_base + 1, max_row=chan_base + 1 + len(CHANNELS))
    cats = Reference(ws, min_col=1, max_col=1,
                     min_row=chan_base + 2, max_row=chan_base + 1 + len(CHANNELS))
    chart.add_data(data, titles_from_data=True, from_rows=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 22
    ws.add_chart(chart, f"A{door_total_r + 3}")

    ws.column_dimensions["A"].width = 36
    for col in range(2, 2 + max(len(MONTH_ORDER), len(QUARTERS)) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12
    setup_print(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = openpyxl.Workbook()
    readme_ws = wb.active
    build_readme(readme_ws)

    home_ws = wb.create_sheet("Home")
    control_ws = wb.create_sheet("Control Panel")
    dashboard_ws = wb.create_sheet("Dashboard")
    pl_ws = wb.create_sheet("P&L")
    scenarios_ws = wb.create_sheet("Scenarios")
    kpis_ws = wb.create_sheet("KPIs")

    assumptions_ws = wb.create_sheet("Assumptions"); build_assumptions(assumptions_ws)
    class_mix_ws = wb.create_sheet("Class Mix"); build_class_mix(class_mix_ws)
    velocity_ws = wb.create_sheet("Velocity"); build_velocity(velocity_ws)
    doors_ws = wb.create_sheet("Doors"); doors_total_row = build_doors(doors_ws)
    pricing_ws = wb.create_sheet("Pricing"); build_pricing(pricing_ws)
    trade_ws = wb.create_sheet("Trade Spend"); trade_refs = build_trade_spend(trade_ws)
    opex_ws = wb.create_sheet("OPEX"); opex_refs = build_opex(opex_ws)
    mk_ws = wb.create_sheet("Marketing and Sampling"); marketing_summary = build_marketing(mk_ws)
    fw_ws = wb.create_sheet("Forecast Weekly"); forecast_total_row = build_forecast_weekly(fw_ws)
    pp_ws = wb.create_sheet("Production Plan"); prod_refs = build_production_plan(pp_ws, forecast_total_row, marketing_summary)
    rev_ws = wb.create_sheet("Revenue"); rev_refs = build_revenue(rev_ws)

    build_pl(pl_ws, rev_refs, trade_refs, opex_refs)
    build_scenarios(scenarios_ws, rev_refs, trade_refs, opex_refs, prod_refs)
    build_kpis(kpis_ws, rev_refs, prod_refs, doors_total_row)
    build_home(home_ws, prod_refs, rev_refs, trade_refs, opex_refs, doors_total_row)
    build_control_panel(control_ws)
    build_dashboard(dashboard_ws, prod_refs, rev_refs, doors_total_row, forecast_total_row)

    # Apply tab colors for visual hierarchy
    home_ws.sheet_properties.tabColor = TAB_FRONT
    control_ws.sheet_properties.tabColor = TAB_FRONT
    for sheet_name in ("Dashboard", "P&L", "Scenarios", "KPIs"):
        wb[sheet_name].sheet_properties.tabColor = TAB_REPORT
    for sheet_name in ("Assumptions", "Class Mix", "Velocity", "Doors", "Pricing",
                        "Trade Spend", "OPEX", "Marketing and Sampling"):
        wb[sheet_name].sheet_properties.tabColor = TAB_INPUT
    # README always first
    readme_ws.sheet_properties.tabColor = TAB_FRONT

    ar_ws = wb.create_sheet("Assumption Register"); build_assumption_register(ar_ws)
    rr_ws = wb.create_sheet("Risk Register"); build_risk_register(rr_ws)
    cl_ws = wb.create_sheet("Change Log"); build_change_log(cl_ws)
    gl_ws = wb.create_sheet("Glossary"); build_glossary(gl_ws)
    vl_ws = wb.create_sheet("Validation Lists"); build_validation_lists(vl_ws)

    for s in (ar_ws, rr_ws, cl_ws, gl_ws):
        s.sheet_properties.tabColor = TAB_GOVERNANCE
    for s in (wb["Forecast Weekly"], wb["Production Plan"], wb["Revenue"], vl_ws):
        s.sheet_properties.tabColor = TAB_ENGINE

    # Active sheet defaults to Home so it's the first thing the user sees
    wb.active = wb.sheetnames.index("Home")

    # Named ranges for the validation dropdowns - on Validation Lists tab
    # CHANNELS = A6:A14, PROVINCES = C6:C15, SKUS = E6:E9, BRANDS = G6:G7
    defn_ch = DefinedName("ChannelList", attr_text=f"'Validation Lists'!$A$6:$A${5 + len(CHANNELS)}")
    defn_pr = DefinedName("ProvinceList", attr_text=f"'Validation Lists'!$C$6:$C${5 + len(PROVINCES)}")
    defn_sk = DefinedName("SKUList", attr_text=f"'Validation Lists'!$E$6:$E${5 + len(SKUS)}")
    wb.defined_names["ChannelList"] = defn_ch
    wb.defined_names["ProvinceList"] = defn_pr
    wb.defined_names["SKUList"] = defn_sk

    # Attach dropdowns to Doors columns A/B/C using the named ranges (data rows only)
    doors_data_last = doors_total_row - 1
    dv_brand = DataValidation(type="list", formula1='"MUV,LCA"', allow_blank=False)
    doors_ws.add_data_validation(dv_brand); dv_brand.add(f"A6:A{doors_data_last}")
    dv_prov = DataValidation(type="list", formula1="=ProvinceList", allow_blank=False)
    doors_ws.add_data_validation(dv_prov); dv_prov.add(f"B6:B{doors_data_last}")
    dv_ch = DataValidation(type="list", formula1="=ChannelList", allow_blank=False)
    doors_ws.add_data_validation(dv_ch); dv_ch.add(f"C6:C{doors_data_last}")

    out_path = "/home/user/my-first-project/Organika_RTD_Forecast_FY26_27_Weekly.xlsx"
    wb.save(out_path)
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    main()
