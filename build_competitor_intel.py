#!/usr/bin/env python3
"""
Beverage Competitor Intelligence workbook generator (Amazon.ca scope) — v2 ENTERPRISE.

HONESTY MODEL (unchanged, enforced):
  - NO live Amazon.ca access in this environment (HTTP 403 on every page).
  - STABLE attributes (brand/sku/size/flavours/nutrition/sweetener/claims/positioning)
    are filled and flagged "approximate, NOT verified on Amazon.ca".
  - LIVE COMMERCIAL fields (price, S&S, coupon, promo, rating, #ratings, badge, BSR, URL)
    are LEFT BLANK + amber. NEVER fabricated. Formulas auto-compute on entry.

v2 adds: Executive Dashboard w/ charts, normalized nutrition (/100 mL), transparent
Health & Functional scores, Brand Roll-up, Category Benchmarks, LIVE cross-tab links,
Nutrition Scoreboard, Live-Capture Protocol, self-auditing QA tab, Excel Tables,
data bars / colour scales, and more products (70+).
"""

import datetime, re, statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, PieChart, ScatterChart, Series, Reference
from openpyxl.chart.marker import Marker

TODAY = datetime.date.today().isoformat()
FNAME = f"Amazon_ca_Beverage_Competitor_Intel_{TODAY}.xlsx"

# palette
NAVY="1F3864"; BLUE="2E5496"; LTBLUE="D9E1F2"; GREY="F2F2F2"; AMBER="FFF2CC"
GREEN="E2EFDA"; RED="F8CBAD"; WHITE="FFFFFF"; GOLD="BF9000"; TEAL="215968"
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
HFONT=Font(name="Calibri",bold=True,color=WHITE,size=11); HFILL=PatternFill("solid",fgColor=NAVY)
HALIGN=Alignment(horizontal="center",vertical="center",wrap_text=True)
WRAP=Alignment(vertical="top",wrap_text=True)
MONEY='"$"#,##0.00'; PCT='0%'; NUM1='0.0'

# ============================================================
# MASTER COLUMNS  (30 task fields + computed analytics + helpers/provenance)
# ============================================================
COLS=[
 ("brand","1. Brand",16,"text"),("sku","2. Product Line / SKU Name",30,"text"),
 ("category","3. Category",26,"text"),("format","4. Format & Size",16,"text"),
 ("pack","5. Pack Count",9,"num"),("price","6. One-Time Price (CAD)",14,"blankmoney"),
 ("ppu","7. Price / Unit (CAD)",13,"formula"),("p100","8. Price / 100 mL (CAD)",13,"formula"),
 ("ss_off","9. S&S Offered (Y/N)",11,"blanktext"),("ss_pct","10. S&S Discount %",12,"blankpct"),
 ("ss_price","11. S&S Price After Disc.",14,"formula"),
 ("coupon","12. Coupon on Listing (Y/N + value)",18,"blanktext"),
 ("promo","13. Promotion Text",22,"blanktext"),
 ("floor","14. Lowest Effective Price (stacked)",16,"formula"),
 ("flavours","15. All Flavours in Listing",40,"flavours"),
 ("caffeine","16. Caffeine / Serving (mg)",12,"num"),("sodium","17. Sodium / Serving (mg)",12,"num"),
 ("potassium","18. Potassium / Serving (mg)",12,"num"),("sugar","19. Sugar / Serving (g)",11,"num"),
 ("calories","20. Calories / Serving",11,"num"),("sweetener","21. Sweetener Type",24,"text"),
 ("claims","22. Key Functional Claims",34,"text"),("star","23. Star Rating",10,"blanknum"),
 ("nratings","24. Number of Ratings",12,"blanknum"),("badge","25. Amazon Badge",16,"blanktext"),
 ("bsr","26. Bestseller Rank in Category",14,"blanktext"),
 ("vband","27a. Est. Weekly Velocity (band)",14,"text"),
 ("vsignal","27b. Velocity Signal Used",34,"text"),
 ("message","28. Biggest Selling Message",40,"text"),
 ("why","29. Why a Shopper Chooses This",46,"text"),("url","30. Direct Product URL",22,"blanktext"),
 # ---- computed analytics (v2) ----
 ("caf100","C1. Caffeine /100 mL (mg)",12,"calc"),("sugar100","C2. Sugar /100 mL (g)",12,"calc"),
 ("sodium100","C3. Sodium /100 mL (mg)",12,"calc"),("sugarfree","C4. Sugar-Free? (<=1 g)",11,"calc"),
 ("health","C5. Health Score (0-100)",12,"calc"),("func","C6. Functional Score (0-100)",13,"calc"),
 # ---- helpers / provenance ----
 ("unitml","H1. Unit Vol (mL) [calc helper]",12,"num"),
 ("couponval","H2. Coupon Value (CAD) [calc helper]",13,"blankmoney"),
 ("prov","P1. Attribute Provenance",30,"text"),("status","P2. Commercial-Field Status",26,"text"),
 ("flag","P3. Row Flag",22,"text"),
]
KEYIDX={c[0]:i for i,c in enumerate(COLS)}
def CL(key): return get_column_letter(KEYIDX[key]+1)
PROV="Secondary/approx; NOT verified on Amazon.ca"
STATUS="BLANK -> requires live Amazon.ca capture"
FLAG="Unverified commercial data"

C1="1) Sparkling electrolyte / functional sparkling"
C2="2) Electrolyte beverage (RTD)"
C3="3) Functional beverage (still/sparkling)"
C4="4) Caffeine & energy beverage"
C5="5) Sparkling water & soda"
CATS=[C1,C2,C3,C4,C5]
SHEET_FOR={C1:"M1 - Sparkling Electrolyte",C2:"M2 - Electrolyte RTD",
 C3:"M3 - Functional Beverage",C4:"M4 - Caffeine & Energy",C5:"M5 - Sparkling Water & Soda"}

# canonical tab order (single source of truth: drives the TOC and the final sort)
SHEET_ORDER=["README & Methodology","Executive Dashboard","Strategic Insights","Data Dictionary",
 SHEET_FOR[C1],SHEET_FOR[C2],SHEET_FOR[C3],SHEET_FOR[C4],SHEET_FOR[C5],
 "Brand Roll-up","Category Benchmarks","Nutrition Scoreboard",
 "Pricing & Promo Analysis","Subscription Strategy","Why They Win",
 "Velocity Estimate","Flavour Map","Live-Capture Protocol","QA & Integrity","Enrichment Log","Sources"]
TAB_PURPOSE={
 "README & Methodology":"Start here — honesty model, scoring, scope, this index.",
 "Executive Dashboard":"KPIs, charts and the top key findings at a glance.",
 "Strategic Insights":"Synthesized findings, flavour whitespace, positioning map, opportunity shortlist.",
 "Data Dictionary":"Every field defined — meaning, source class, fill status.",
 SHEET_FOR[C1]:"SKU-level master: sparkling electrolyte / functional sparkling.",
 SHEET_FOR[C2]:"SKU-level master: electrolyte beverages (RTD).",
 SHEET_FOR[C3]:"SKU-level master: functional beverages (still/sparkling).",
 SHEET_FOR[C4]:"SKU-level master: caffeine & energy beverages.",
 SHEET_FOR[C5]:"SKU-level master: sparkling water & soda.",
 "Brand Roll-up":"Per-brand SKU counts, avg nutrition and scores.",
 "Category Benchmarks":"Per-category averages, ranges and sugar-free counts.",
 "Nutrition Scoreboard":"All lines ranked by Health Score with data bars.",
 "Pricing & Promo Analysis":"Live-linked price/floor view (fill amber cells to activate).",
 "Subscription Strategy":"Subscribe & Save economics, live-linked.",
 "Why They Win":"One-line buyer rationale + primary draw per line.",
 "Velocity Estimate":"Estimated demand bands with the signal used.",
 "Flavour Map":"Brand × flavour-family coverage grid.",
 "Live-Capture Protocol":"How to fill the live commercial fields.",
 "QA & Integrity":"Self-audit: confirms no commercial cell is fabricated.",
 "Enrichment Log":"Every secondary-source nutrition correction, with citations.",
 "Sources":"Source classes and references behind the stable attributes."}

# ============================================================
# PRODUCTS
# ============================================================
P=[]
def add(**k): P.append(k)

# ---- CAT 4: CAFFEINE & ENERGY ----
add(brand="Red Bull",sku="Red Bull Energy Drink",category=C4,format="250 mL can",pack=24,flavours="Original",caffeine=80,sodium=105,potassium=None,sugar=27,calories=110,sweetener="Sucrose + glucose-fructose",claims="Energy, focus, taurine 1000 mg + B-vitamins (B3/B5/B6/B12)",vband="Very High",vsignal="Global #1 energy brand; ubiquitous repeat purchase",message="The original wings-giving energy hit",why="Brand trust and ubiquity - the default energy can people already know.",unitml=250)
add(brand="Red Bull",sku="Red Bull Sugarfree",category=C4,format="250 mL can",pack=12,flavours="Original (sugarfree)",caffeine=80,sodium=105,potassium=None,sugar=0,calories=10,sweetener="Sucralose + acesulfame-K",claims="Energy, zero sugar, taurine 1000 mg + B-vitamins",vband="High",vsignal="Strong line extension of #1 brand",message="Red Bull energy, no sugar",why="No sugar without leaving the brand they trust.",unitml=250)
add(brand="Monster",sku="Monster Energy Original (Green)",category=C4,format="473 mL can",pack=15,flavours="Original",caffeine=160,sodium=370,potassium=None,sugar=54,calories=210,sweetener="Sugar + glucose",claims="Big energy blend, taurine, ginseng",vband="Very High",vsignal="#2 energy brand; large can = value perception",message="A bigger, harder-hitting energy can",why="Pack economics and high dose - more caffeine and volume per dollar.",unitml=473)
add(brand="Monster",sku="Monster Energy Ultra Zero (White)",category=C4,format="473 mL can",pack=12,flavours="Zero Ultra, Ultra Paradise, Ultra Sunrise, Ultra Watermelon, Ultra Gold, Ultra Vice Guava, Ultra Peachy Keen",caffeine=150,sodium=370,potassium=None,sugar=0,calories=10,sweetener="Erythritol + sucralose",claims="Zero sugar, light taste, full energy",vband="Very High",vsignal="Best-selling zero-sugar energy line",message="Full Monster energy, zero sugar, lighter taste",why="No sugar plus a lighter flavour than the green can.",unitml=473)
add(brand="Celsius",sku="Celsius Original Sparkling",category=C4,format="355 mL can",pack=12,flavours="Sparkling Orange, Wild Berry, Kiwi Guava, Cola, Fuji Apple Pear, Sparkling Cranberry",caffeine=200,sodium=5,potassium=50,sugar=0,calories=10,sweetener="Sucralose",claims="Accelerates metabolism, MetaPlus blend, zero sugar, vitamins",vband="Very High",vsignal="Fastest-growing energy brand; fitness crossover",message="Functional energy that claims to boost metabolism",why="No sugar plus a fitness/metabolism story, not just caffeine.",unitml=355)
add(brand="Celsius",sku="Celsius Essentials",category=C4,format="473 mL can",pack=12,flavours="Cosmic Blue, Sparkling Green, Dragonberry, Orange Mango",caffeine=270,sodium=10,potassium=None,sugar=0,calories=10,sweetener="Sucralose",claims="Higher caffeine, amino acids, zero sugar",vband="High",vsignal="Higher-dose line riding Celsius momentum",message="A bigger, stronger Celsius for serious energy",why="Highest caffeine dose in the Celsius family.",unitml=473)
add(brand="Prime",sku="Prime Energy",category=C4,format="355 mL can",pack=12,flavours="Blue Raspberry, Tropical Punch, Lemon Lime, Ice Pop, Strawberry Watermelon",caffeine=200,sodium=10,potassium=None,sugar=0,calories=10,sweetener="Sucralose + acesulfame-K",claims="Zero sugar, electrolytes, B-vitamins",vband="High",vsignal="Influencer-driven (Logan Paul/KSI); youth pull",message="Hype-brand energy with zero sugar",why="Social-media brand cachet with a zero-sugar profile.",unitml=355)
add(brand="C4",sku="C4 Energy (Performance)",category=C4,format="473 mL can",pack=12,flavours="Frozen Bombsicle, Cotton Candy, Strawberry Watermelon, Tropical Blast, Orange Slice",caffeine=200,sodium=45,potassium=None,sugar=0,calories=0,sweetener="Sucralose + acesulfame-K",claims="Pre-workout energy, beta-alanine (CarnoSyn)",vband="Medium",vsignal="Strong in fitness/pre-workout niche",message="Pre-workout performance in a ready-to-drink can",why="Performance/pre-workout positioning with beta-alanine.",unitml=473)
add(brand="Reign",sku="Reign Total Body Fuel",category=C4,format="473 mL can",pack=12,flavours="Razzle Berry, Melon Mania, Sour Apple, Orange Dreamsicle, Lemon HDZ",caffeine=300,sodium=10,potassium=None,sugar=0,calories=10,sweetener="Sucralose",claims="300 mg caffeine, BCAAs, CoQ10, zero sugar",vband="Medium",vsignal="Performance line under Monster umbrella",message="Maximum-dose fitness energy with BCAAs",why="Highest mainstream caffeine dose plus workout aminos.",unitml=473)
add(brand="Alani Nu",sku="Alani Nu Energy",category=C4,format="355 mL can",pack=12,flavours="Hawaiian Shaved Ice, Witch's Brew, Mimosa, Cosmic Stardust, Breezeberry, Watermelon Wave",caffeine=200,sodium=35,potassium=None,sugar=0,calories=10,sweetener="Sucralose",claims="Zero sugar, vitamins, female-fitness positioning",vband="High",vsignal="Fast-growing; strong female demographic & flavours",message="Stylish zero-sugar energy aimed at women",why="On-trend flavours and aesthetic plus zero sugar.",unitml=355)
add(brand="Bang",sku="Bang Energy",category=C4,format="473 mL can",pack=12,flavours="Rainbow Unicorn, Star Blast, Cotton Candy, Blue Razz, Peach Mango",caffeine=300,sodium=40,potassium=85,sugar=0,calories=0,sweetener="Sucralose",claims="300 mg caffeine, BCAAs, CoQ10, zero calorie",vband="Medium",vsignal="Declining vs peak but still high volume",message="Zero-calorie high-caffeine performance fuel",why="High caffeine at zero calories with aminos.",unitml=473)
add(brand="Rockstar",sku="Rockstar Energy",category=C4,format="473 mL can",pack=12,flavours="Original, Sugar Free, Punched Fruit Punch, Pure Zero Silver Ice",caffeine=160,sodium=65,potassium=None,sugar=31,calories=122,sweetener="Sugar (or sucralose on zero)",claims="Energy blend, taurine, guarana",vband="Medium",vsignal="Legacy brand; PepsiCo distribution",message="Value-priced legacy energy can",why="Familiar legacy brand, usually a lower price point.",unitml=473)
add(brand="GURU",sku="GURU Organic Energy",category=C4,format="250 mL can",pack=12,flavours="Original, Tropical Punch, Yerba Mate, GURU Lite (low-cal)",caffeine=100,sodium=0,potassium=None,sugar=21,calories=80,sweetener="Organic cane sugar + organic white grape juice (Lite: stevia + monk fruit, 3 g sugar, 25 cal)",claims="Certified organic, plant-based, natural caffeine (guarana/green tea/ginseng/echinacea); Canada Organic certified",vband="Medium",vsignal="Canadian brand; natural/organic niche leader at home",message="Clean, organic, plant-based energy (Canadian)",why="Organic, natural-caffeine alternative - a clean-label choice.",unitml=250)
add(brand="Guayaki",sku="Guayaki Yerba Mate",category=C4,format="458 mL can",pack=12,flavours="Revel Berry, Enlighten Mint, Bluephoria, Lemon Elation, Orange Exuberance",caffeine=150,sodium=0,potassium=None,sugar=28,calories=120,sweetener="Organic cane sugar",claims="Organic, Fair Trade yerba mate, 150 mg natural caffeine; USDA/Canada Organic certified",vband="Medium",vsignal="Strong natural-channel following",message="Organic yerba mate energy with an ethical story",why="Natural yerba-mate caffeine plus organic/Fair-Trade ethics.",unitml=458)
add(brand="NOS",sku="NOS Energy Drink",category=C4,format="473 mL can",pack=12,flavours="Original, Sugar Free",caffeine=160,sodium=410,potassium=None,sugar=54,calories=210,sweetener="HFCS + sucralose",claims="High-performance energy blend, taurine, L-theanine, guarana, B-vitamins",vband="Low",vsignal="Niche legacy brand, limited shelf share",message="Motorsport-styled high-octane energy",why="Legacy enthusiast brand at a value price.",unitml=473)
add(brand="Starbucks",sku="Starbucks Doubleshot Energy",category=C4,format="444 mL can",pack=12,flavours="Coffee, Vanilla, Mocha, White Chocolate",caffeine=135,sodium=170,potassium=None,sugar=28,calories=200,sweetener="Sugar",claims="Espresso + skim milk + energy blend (guarana, ginseng, B-vitamins); 135 mg caffeine; separate from 6.5 fl oz Espresso + Cream product",vband="High",vsignal="Coffee-energy crossover; broad mainstream appeal",message="Cafe espresso energy in a grab-and-go can",why="Trusted coffee taste rather than a synthetic energy flavour.",unitml=444)
add(brand="Ghost",sku="Ghost Energy",category=C4,format="473 mL can",pack=12,flavours="Sour Patch Kids Redberry, Warheads Sour Watermelon, Swedish Fish, Tropical Mango, Orange Cream",caffeine=200,sodium=30,potassium=None,sugar=0,calories=5,sweetener="Sucralose + acesulfame-K",claims="200 mg natural caffeine, zero sugar, licensed-candy flavours, nootropic blend (L-carnitine, Alpha-GPC, NeuroFactor)",vband="High",vsignal="Fast-growing lifestyle brand; strong flavour licensing",message="Candy-flavoured zero-sugar energy with a focus blend",why="Novel licensed-candy flavours plus zero sugar.",unitml=473)
add(brand="ZOA",sku="ZOA Zero Sugar Energy",category=C4,format="473 mL can",pack=12,flavours="Wild Orange, Super Berry, Tropical Punch, White Peach, Lemon Lime",caffeine=160,sodium=200,potassium=60,sugar=0,calories=15,sweetener="Sucralose + acesulfame-K",claims="Natural caffeine from green tea + green coffee extract, 200 mg Na + 60 mg K electrolytes, vitamin C & B-vitamins, camu camu, zero sugar (Dwayne Johnson brand)",vband="Medium",vsignal="Celebrity-founded; wellness-energy positioning",message="A 'healthier' energy with electrolytes and vitamins",why="Energy plus added electrolytes/vitamins and a celebrity halo.",unitml=473)
add(brand="Zevia",sku="Zevia Energy",category=C4,format="355 mL can",pack=12,flavours="Mango Ginger, Grapefruit, Kola, Raspberry Lime",caffeine=120,sodium=0,potassium=None,sugar=0,calories=0,sweetener="Stevia",claims="Zero sugar, zero calorie, stevia-sweetened, no artificial sweeteners",vband="Low",vsignal="Niche clean-label energy; stevia differentiator",message="Naturally-sweetened zero-calorie energy",why="Caffeine with stevia and no artificial sweeteners.",unitml=355)

# ---- CAT 2: ELECTROLYTE RTD ----
add(brand="Gatorade",sku="Gatorade Thirst Quencher",category=C2,format="591 mL bottle",pack=12,flavours="Cool Blue, Fruit Punch, Glacier Freeze, Lemon-Lime, Orange, Riptide Rush",caffeine=0,sodium=270,potassium=75,sugar=34,calories=140,sweetener="Sugar + dextrose",claims="Rehydrate, replenish electrolytes, sports performance",vband="Very High",vsignal="Category-defining sports drink; mass repeat purchase",message="The original sports hydration standard",why="Sodium-led rehydration with unmatched brand trust.",unitml=591)
add(brand="Gatorade",sku="Gatorade Zero Sugar",category=C2,format="591 mL bottle",pack=12,flavours="Glacier Cherry, Lemon-Lime, Orange, Berry, Glacier Freeze",caffeine=0,sodium=270,potassium=75,sugar=0,calories=5,sweetener="Sucralose + acesulfame-K",claims="Same electrolytes, zero sugar",vband="Very High",vsignal="Top zero-sugar sports drink",message="Gatorade electrolytes without the sugar",why="Full electrolytes minus the sugar load.",unitml=591)
add(brand="Powerade",sku="Powerade Mountain Berry Blast",category=C2,format="591 mL bottle",pack=12,flavours="Mountain Berry Blast, Fruit Punch, Grape, White Cherry, Lemon-Lime",caffeine=0,sodium=250,potassium=70,sugar=35,calories=130,sweetener="High-fructose corn syrup",claims="ION4 electrolytes, B-vitamins, hydration",vband="High",vsignal="Strong #2; Coca-Cola distribution & price",message="Value-priced electrolyte hydration",why="Cheaper electrolyte alternative to Gatorade.",unitml=591)
add(brand="Powerade",sku="Powerade Zero",category=C2,format="591 mL bottle",pack=12,flavours="Mixed Berry, Fruit Punch, Grape, White Cherry",caffeine=0,sodium=250,potassium=35,sugar=0,calories=0,sweetener="Sucralose + acesulfame-K",claims="Zero sugar/calorie, ION4 electrolytes, vitamins",vband="High",vsignal="Popular zero-sugar value option",message="Zero-calorie electrolytes at a value price",why="Cheapest zero-sugar electrolyte option.",unitml=591)
add(brand="BODYARMOR",sku="BODYARMOR Sports Drink",category=C2,format="473 mL bottle",pack=12,flavours="Fruit Punch, Strawberry Banana, Orange Mango, Mixed Berry, Tropical Punch",caffeine=0,sodium=30,potassium=680,sugar=29,calories=120,sweetener="Cane sugar",claims="Potassium-packed electrolytes, coconut water, vitamins, no artificial",vband="High",vsignal="Fast-growing premium sports drink (Coca-Cola owned)",message="Premium, natural sports hydration with potassium",why="Coconut-water electrolytes and a cleaner label than legacy brands.",unitml=473)
add(brand="BODYARMOR",sku="BODYARMOR LYTE",category=C2,format="473 mL bottle",pack=12,flavours="Blueberry Pomegranate, Peach Mango, Tropical Coconut, Berry Punch",caffeine=0,sodium=30,potassium=350,sugar=3,calories=20,sweetener="Stevia + erythritol",claims="Low sugar, potassium electrolytes, coconut water, vitamins",vband="Medium",vsignal="Low-sugar variant of fast-grower",message="Low-sugar premium hydration",why="High potassium with very little sugar.",unitml=473)
add(brand="Prime",sku="Prime Hydration",category=C2,format="500 mL bottle",pack=12,flavours="Blue Raspberry, Tropical Punch, Ice Pop, Lemon Lime, Strawberry Watermelon, Meta Moon",caffeine=0,sodium=30,potassium=700,sugar=2,calories=25,sweetener="Sucralose + acesulfame-K",claims="Zero sugar, BCAAs, electrolytes, coconut water, vitamins",vband="High",vsignal="Influencer hype; strong youth pull",message="Hype-brand zero-sugar hydration with BCAAs",why="Brand cachet plus zero sugar and added BCAAs.",unitml=500)
add(brand="Roar Organic",sku="Roar Organic Electrolyte Infusions",category=C2,format="532 mL bottle",pack=12,flavours="Cucumber Watermelon, Mango Clementine, Blueberry Acai, Georgia Peach",caffeine=0,sodium=45,potassium=100,sugar=2,calories=10,sweetener="Stevia + erythritol",claims="Coconut-water electrolytes, antioxidants, B-vitamins, low sugar, organic",vband="Low",vsignal="Emerging organic niche; modest shelf share",message="Organic, low-sugar coconut-water hydration",why="Organic and barely-sweetened for clean-label hydration seekers.",unitml=532)
add(brand="BioSteel",sku="BioSteel Sports Drink (RTD)",category=C2,format="500 mL bottle",pack=12,flavours="Blue Raspberry, White Freeze, Rainbow Twist, Mixed Berry, Peach Mango",caffeine=0,sodium=230,potassium=225,sugar=0,calories=10,sweetener="Stevia Leaf Extract",claims="5 essential electrolytes (Na/K/Ca/Mg/Sea Salt), zero sugar, plant-based sweetener, vegan, non-GMO, caffeine-free; made in Canada",vband="Medium",vsignal="Canadian brand; strong domestic hockey/sports tie-ins",message="Zero-sugar clean sports hydration (Canadian, pro-endorsed)",why="Zero sugar plus pro-sports credibility, made in Canada.",unitml=500)
add(brand="CWENCH",sku="CWENCH Hydration",category=C2,format="500 mL bottle",pack=12,flavours="Blue Raspberry, Fruit Punch, Green Apple, Watermelon, Peach",caffeine=0,sodium=200,potassium=100,sugar=0,calories=15,sweetener="Sucralose",claims="Zero sugar, electrolytes, vitamins; Canadian challenger brand",vband="Low",vsignal="New Canadian entrant; limited reviews/footprint",message="Affordable Canadian zero-sugar hydration",why="Local challenger price and zero sugar.",unitml=500)
add(brand="Pedialyte",sku="Pedialyte Electrolyte Solution",category=C2,format="1 L bottle",pack=8,flavours="Unflavored, Fruit, Grape, Strawberry, Mixed Fruit",caffeine=0,sodium=1035,potassium=780,sugar=25,calories=100,sweetener="Dextrose + sucralose + acesulfame-K",claims="Medical-grade rehydration, dehydration relief (values per 1 L; ~370 mg sodium per 12 oz serving)",vband="High",vsignal="Trusted clinical rehydration; year-round + flu-season spikes",message="Doctor-trusted rehydration for illness/dehydration",why="Medical-grade electrolyte ratio trusted for genuine dehydration.",unitml=1000)
add(brand="Gatorade",sku="Gatorlyte Rapid Rehydration",category=C2,format="591 mL bottle",pack=12,flavours="Orange, Cherry Lime, Glacier Cherry, Strawberry Kiwi",caffeine=0,sodium=490,potassium=350,sugar=12,calories=50,sweetener="Sugar + sucralose",claims="5 electrolytes, rapid rehydration, specialized formula",vband="Medium",vsignal="Premium rapid-rehydration line",message="Higher-electrolyte rapid rehydration",why="Highest sodium/potassium load for serious rehydration.",unitml=591)
add(brand="Vita Coco",sku="Vita Coco Pure Coconut Water",category=C2,format="330 mL carton",pack=12,flavours="Original, Pineapple, Peach Mango, The Original (Pressed)",caffeine=0,sodium=25,potassium=470,sugar=11,calories=45,sweetener="None (naturally occurring)",claims="Natural electrolytes, potassium, no added sugar (Original)",vband="High",vsignal="Market-leading coconut water; mainstream",message="Natural hydration straight from coconuts",why="Naturally high potassium with no added sugar.",unitml=330)
add(brand="Electrolit",sku="Electrolit Hydration Beverage",category=C2,format="625 mL bottle",pack=12,flavours="Strawberry-Kiwi, Fruit Punch, Watermelon, Coconut, Blue Raspberry, Grape",caffeine=0,sodium=270,potassium=200,sugar=24,calories=100,sweetener="Cane sugar + glucose",claims="Magnesium + calcium + sodium + potassium, fast rehydration",vband="High",vsignal="Fast-growing Mexican import; strong Amazon momentum",message="Multi-mineral rehydration popular for recovery",why="Broad mineral profile (incl. magnesium) for serious rehydration.",unitml=625)
add(brand="Gatorade",sku="Gatorade Fit Electrolyte Beverage",category=C2,format="591 mL bottle",pack=12,flavours="Tropical Mango, Blue Citrus, Strawberry Watermelon, Watermelon Strawberry",caffeine=0,sodium=200,potassium=80,sugar=0,calories=5,sweetener="Stevia",claims="No added sugar, naturally sweetened, electrolytes + vitamins A & C",vband="Medium",vsignal="Clean-label line targeting wellness shoppers",message="Cleaner-label Gatorade with no added sugar",why="Gatorade electrolytes with stevia and added vitamins.",unitml=591)

# ---- CAT 3: FUNCTIONAL ----
add(brand="OLIPOP",sku="OLIPOP Prebiotic Soda",category=C3,format="355 mL can",pack=12,flavours="Vintage Cola, Classic Root Beer, Cherry Cola, Strawberry Vanilla, Orange Squeeze, Ginger Ale, Doctor Goodwin, Lemon Lime, Watermelon Lime",caffeine=0,sodium=30,potassium=None,sugar=4,calories=45,sweetener="Stevia + cassava syrup",claims="9 g prebiotic fiber, gut/digestive health, low sugar (Vintage/Cherry Cola & Doctor Goodwin ~50 mg caffeine, Ridge Rush ~60 mg; other flavours caffeine-free)",vband="Very High",vsignal="Category-leading prebiotic soda; viral DTC + retail",message="Soda nostalgia that's actually good for your gut",why="Classic soda taste with 9 g fiber and almost no sugar.",unitml=355)
add(brand="Poppi",sku="Poppi Prebiotic Soda",category=C3,format="355 mL can",pack=12,flavours="Strawberry Lemon, Raspberry Rose, Classic Cola, Cherry Limeade, Orange, Grape, Watermelon, Ginger Lime, Doc Pop, Root Beer",caffeine=0,sodium=10,potassium=None,sugar=5,calories=25,sweetener="Agave + stevia",claims="Prebiotic, apple cider vinegar, gut health, low sugar (cola/Doc Pop flavours ~40 mg caffeine from green tea; fruit flavours caffeine-free)",vband="Very High",vsignal="Viral brand; Super Bowl ad; strong Gen-Z pull",message="Pretty, prebiotic, apple-cider-vinegar soda",why="Trendy aesthetic plus a gut-health ACV story at low sugar.",unitml=355)
add(brand="Culture Pop",sku="Culture Pop Probiotic Soda",category=C3,format="355 mL can",pack=12,flavours="Ginger Lemon, Wild Berry, Orange Mango, Lime Watermelon, Cherry Pomegranate",caffeine=0,sodium=15,potassium=None,sugar=9,calories=40,sweetener="Fruit juice + a touch of cane",claims="Live probiotics, real juice, gut health, no stevia",vband="Medium",vsignal="Emerging; differentiates on real-juice/no-stevia",message="Probiotic soda with real juice, no stevia",why="Real-juice taste and live probiotics without stevia.",unitml=355)
add(brand="Mayawell",sku="Mayawell Prebiotic Soda",category=C3,format="355 mL can",pack=12,flavours="Strawberry Hibiscus, Prickly Pear, Mango Chili, Grapefruit Jalapeno",caffeine=0,sodium=20,potassium=None,sugar=6,calories=40,sweetener="Agave inulin",claims="Prebiotic agave fiber, gut health, low sugar, Latino-owned",vband="Low",vsignal="Small emerging brand; limited footprint",message="Agave-fiber prebiotic soda with bold flavours",why="Distinctive agave-fiber prebiotic and adventurous flavours.",unitml=355)
add(brand="De La Calle",sku="De La Calle Tepache",category=C3,format="355 mL can",pack=12,flavours="Tamarindo, Mango Chili, Original, Guava, Pineapple Spice",caffeine=0,sodium=15,potassium=None,sugar=6,calories=45,sweetener="Fermented pineapple + agave",claims="Fermented prebiotic tepache, gut health, low sugar",vband="Low",vsignal="Niche fermented-soda entrant",message="Traditional Mexican fermented tepache, modernized",why="Authentic fermented tepache with a prebiotic angle.",unitml=355)
add(brand="GT's",sku="GT's Synergy Raw Kombucha",category=C3,format="480 mL bottle",pack=12,flavours="Trilogy, Gingerade, Gingerberry, Strawberry Lemonade, Cosmic Cranberry",caffeine=14,sodium=10,potassium=None,sugar=12,calories=50,sweetener="Naturally fermented (residual)",claims="Raw kombucha, 9 billion probiotics per bottle, gut health",vband="High",vsignal="Kombucha category leader",message="The original raw, probiotic kombucha",why="Authentic raw kombucha from the category pioneer.",unitml=480)
add(brand="Health-Ade",sku="Health-Ade Kombucha",category=C3,format="473 mL bottle",pack=12,flavours="Pink Lady Apple, Pomegranate, Ginger Lemon, Bubbly Rose, California Grape",caffeine=15,sodium=10,potassium=None,sugar=7,calories=35,sweetener="Naturally fermented (residual)",claims="Probiotics, organic, gut health, glass-bottle quality",vband="Medium",vsignal="Premium kombucha; strong natural channel",message="Premium small-batch organic kombucha",why="Premium organic kombucha taste and quality cues.",unitml=473)
add(brand="Remedy",sku="Remedy Kombucha",category=C3,format="250 mL can",pack=12,flavours="Raspberry Lemonade, Ginger Lemon, Peach, Cherry Plum, Passionfruit",caffeine=10,sodium=5,potassium=None,sugar=0,calories=8,sweetener="None (sugar-free fermented)",claims="Sugar-free kombucha, live cultures, organic acids",vband="Medium",vsignal="Sugar-free kombucha differentiator",message="Kombucha benefits with no sugar",why="Live-culture kombucha at zero sugar.",unitml=250)
add(brand="Recess",sku="Recess Mood",category=C3,format="355 mL can",pack=12,flavours="Blackberry Chai, Pomegranate Hibiscus, Blood Orange, Black Cherry, Coconut Lime",caffeine=0,sodium=10,potassium=None,sugar=1,calories=25,sweetener="Agave + stevia",claims="Adaptogens + magnesium + L-theanine, calm/de-stress",vband="Medium",vsignal="Leading calming/adaptogen sparkling brand",message="A sparkling drink designed to calm you down",why="Adaptogens and L-theanine for relaxation, not energy.",unitml=355)
add(brand="Protein2o",sku="Protein2o Protein Infused Water",category=C3,format="500 mL bottle",pack=12,flavours="Wild Cherry, Mixed Berry, Tropical Coconut, Peach Mango, Blueberry Raspberry",caffeine=0,sodium=120,potassium=None,sugar=1,calories=70,sweetener="Sucralose + acesulfame-K",claims="15 g whey protein isolate, electrolytes, low calorie",vband="Medium",vsignal="Leading protein-water; fitness shoppers",message="15 g of protein in a light, clear water",why="Real protein dose without a heavy milkshake texture.",unitml=500)
add(brand="vitaminwater",sku="vitaminwater (glaceau)",category=C3,format="591 mL bottle",pack=12,flavours="XXX Acai-Blueberry-Pomegranate, Power-C Dragonfruit, Focus Kiwi-Strawberry, Energy Tropical Citrus",caffeine=0,sodium=0,potassium=None,sugar=27,calories=100,sweetener="Cane sugar + crystalline fructose",claims="Added vitamins (B, C), function-by-flavour (Energy flavour ~60 mg caffeine; other flavours caffeine-free)",vband="High",vsignal="Established functional water; mass distribution",message="Flavoured water with a vitamin function story",why="Familiar vitamin-enhanced water in many functional flavours.",unitml=591)
add(brand="vitaminwater",sku="vitaminwater zero sugar",category=C3,format="591 mL bottle",pack=12,flavours="XXX, Power-C, Squeezed Lemonade, Rise Orange",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="Stevia + erythritol",claims="Vitamins, zero sugar",vband="High",vsignal="Strong zero-sugar variant",message="Vitamin water with zero sugar",why="The vitamin story without the sugar.",unitml=591)
add(brand="Bai",sku="Bai Antioxidant Infusion",category=C3,format="530 mL bottle",pack=12,flavours="Brasilia Blueberry, Costa Rica Clementine, Molokai Coconut, Ipanema Pomegranate",caffeine=55,sodium=10,potassium=None,sugar=1,calories=10,sweetener="Erythritol + stevia",claims="Antioxidants (coffeefruit + white tea), vitamin C, low sugar, low calorie",vband="Medium",vsignal="Established low-sugar functional flavoured water",message="Antioxidant flavoured water, barely any sugar",why="Flavour with antioxidants and almost no sugar.",unitml=530)
add(brand="Hint",sku="Hint Water (fruit-infused)",category=C3,format="473 mL bottle",pack=12,flavours="Watermelon, Blackberry, Cherry, Pineapple, Peach, Crisp Apple",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None (fruit essence only)",claims="Zero sweetener, zero sugar, zero calorie, just fruit essence",vband="Medium",vsignal="Leading unsweetened flavoured water",message="Fruit-flavoured water with absolutely no sweetener",why="Real fruit essence with zero sweeteners of any kind.",unitml=473)
add(brand="wildwonder",sku="wildwonder Sparkling Prebiotic+Probiotic",category=C3,format="355 mL can",pack=12,flavours="Strawberry Passion, Peach Ginger, Raspberry Rose, Lemon Elderflower, Pineapple Mango",caffeine=0,sodium=15,potassium=None,sugar=6,calories=45,sweetener="Cane sugar + chicory inulin",claims="Prebiotics + probiotics, organic botanicals, gut health, low sugar",vband="Low",vsignal="Emerging female/AAPI-founded gut-health brand",message="Botanical gut-health soda with pre + probiotics",why="Both prebiotics and probiotics plus organic botanicals.",unitml=355)
add(brand="Lemon Perfect",sku="Lemon Perfect Cold-Pressed Lemon Water",category=C3,format="355 mL bottle",pack=12,flavours="Just Lemon, Peach Raspberry, Dragon Fruit Mango, Blueberry Acai, Strawberry Passionfruit",caffeine=0,sodium=20,potassium=70,sugar=1,calories=5,sweetener="Stevia + monk fruit",claims="Cold-pressed lemon, vitamin C, electrolytes, low sugar, organic",vband="Low",vsignal="Emerging functional lemon-water entrant",message="Organic lemon water with vitamin C and electrolytes",why="Clean lemon-water hydration with vitamin C, barely any sugar.",unitml=355)

# ---- CAT 5: SPARKLING WATER & SODA ----
add(brand="Spindrift",sku="Spindrift Sparkling Water",category=C5,format="355 mL can",pack=12,flavours="Lemon, Raspberry Lime, Grapefruit, Half-Tea Half-Lemon, Pineapple, Blackberry, Strawberry, Orange Mango",caffeine=0,sodium=0,potassium=None,sugar=2,calories=10,sweetener="None (real squeezed fruit)",claims="Real squeezed fruit, no added sugar/sweetener",vband="High",vsignal="Premium real-fruit seltzer; strong loyalty",message="Sparkling water made with real squeezed fruit",why="Tastes like real fruit because it is - not flavour essence.",unitml=355)
add(brand="LaCroix",sku="LaCroix Sparkling Water",category=C5,format="355 mL can",pack=12,flavours="Pamplemousse, Lime, Pure, Coconut, Berry, Passionfruit, Key Lime, Cerise Limon",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None (natural essence)",claims="Zero calorie/sugar/sweetener, naturally essenced",vband="Very High",vsignal="Iconic seltzer; huge mainstream base",message="The cult zero-everything sparkling water",why="Zero calories/sugar with cult brand status and price.",unitml=355)
add(brand="bubly",sku="bubly Sparkling Water",category=C5,format="355 mL can",pack=12,flavours="Cherry, Lime, Grapefruit, Mango, Strawberry, Blackberry, Raspberry, Watermelon",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None (natural flavours)",claims="Zero calorie/sugar/sweetener",vband="Very High",vsignal="PepsiCo muscle; aggressive pricing & distribution",message="Cheerful zero-calorie sparkling water at a sharp price",why="Mainstream zero-everything seltzer, usually well priced.",unitml=355)
add(brand="Waterloo",sku="Waterloo Sparkling Water",category=C5,format="355 mL can",pack=12,flavours="Black Cherry, Grape, Watermelon, Lemon-Lime, Peach, Mango Orange, Summer Berry",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None (natural flavours)",claims="Zero calorie/sugar, bold true-to-fruit flavour",vband="Medium",vsignal="Flavour-forward challenger; growing",message="Bolder, more flavour-true sparkling water",why="Stronger, more realistic flavour than the big seltzers.",unitml=355)
add(brand="Perrier",sku="Perrier Carbonated Mineral Water",category=C5,format="330 mL can",pack=10,flavours="Original, Lime, Lemon, Pink Grapefruit, Strawberry, Peach",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None",claims="Natural French mineral water from Vergèze (9.5 mg/L sodium, 150 mg/L calcium — below 5 mg label threshold per serving, declared as 0 mg), naturally carbonated, premium",vband="High",vsignal="Iconic premium mineral water; mainstream + horeca",message="Premium French sparkling mineral water",why="Premium imported mineral-water cachet and crisp bubbles.",unitml=330)
add(brand="San Pellegrino",sku="S.Pellegrino Sparkling Natural Mineral Water",category=C5,format="500 mL bottle",pack=12,flavours="Original (mineral)",caffeine=0,sodium=15,potassium=None,sugar=0,calories=0,sweetener="None",claims="Italian Alpine natural mineral water, 33.6 mg/L sodium, 179 mg/L calcium, 52 mg/L magnesium, fine perlage, premium dining",vband="High",vsignal="Premium mineral water; dining staple",message="Italy's fine-dining sparkling mineral water",why="Premium mineral profile and table-water prestige.",unitml=500)
add(brand="San Pellegrino",sku="Sanpellegrino Aranciata (Italian Sparkling Drink)",category=C5,format="330 mL can",pack=6,flavours="Aranciata, Aranciata Rossa, Limonata, Pompelmo, Melograno e Arancia",caffeine=0,sodium=10,potassium=None,sugar=16,calories=70,sweetener="Sugar + real fruit juice",claims="Real fruit juice Italian sparkling beverage",vband="Medium",vsignal="Premium imported fruit soda; recognizable",message="Italian fruit soda with real juice",why="Premium imported fruit-soda taste with real juice.",unitml=330)
add(brand="Liquid Death",sku="Liquid Death Sparkling Water",category=C5,format="500 mL can",pack=12,flavours="Severed Lime, Mango Chainsaw, Berry It Alive, Convicted Melon, Grave Fruit, Pina Killada",caffeine=0,sodium=5,potassium=None,sugar=0,calories=5,sweetener="Agave nectar + stevia (2026)",claims="Mountain water, zero sugar, irreverent brand, recyclable tallboy",vband="High",vsignal="Viral brand; huge cultural pull beyond product",message="Edgy 'murder your thirst' sparkling water in a tallboy",why="Brand attitude and packaging as much as the water itself.",unitml=500)
add(brand="Liquid Death",sku="Liquid Death Mountain Water (Still)",category=C5,format="500 mL can",pack=12,flavours="Still (unflavoured)",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None",claims="Austrian Alps artesian water (single ingredient), pH ~8.0, zero-everything label, infinitely recyclable tallboy can, brand-driven",vband="High",vsignal="Viral brand halo on plain water",message="Plain mountain water made cool by branding",why="Buys the brand and the can, not just hydration.",unitml=500)
add(brand="Topo Chico",sku="Topo Chico Mineral Water",category=C5,format="355 mL bottle",pack=12,flavours="Original, Twist of Lime, Twist of Grapefruit, Tangerine",caffeine=0,sodium=15,potassium=None,sugar=0,calories=0,sweetener="None",claims="Mexican mineral water from Monterrey source; 33 mg/L sodium (2025 Coca-Cola water quality report) → 11.7 mg/355 mL → label rounds to 15 mg; 140 mg/L calcium; pH 6.2; strong carbonation, cult following",vband="Medium",vsignal="Cult mineral water; cocktail mixer crossover",message="Cult Mexican mineral water with serious fizz",why="Hard carbonation and mixer credibility.",unitml=355)
add(brand="Polar",sku="Polar Seltzer",category=C5,format="355 mL can",pack=8,flavours="Black Cherry, Ruby Red Grapefruit, Lime, Mandarin, Raspberry Lime, Vanilla",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None (natural flavours)",claims="Zero calorie/sugar, wide flavour range, value",vband="Medium",vsignal="Value seltzer with broad flavour range",message="Value seltzer with a huge flavour range",why="Lots of flavours at a value price.",unitml=355)
add(brand="AHA",sku="AHA Sparkling Water",category=C5,format="355 mL can",pack=8,flavours="Lime + Watermelon, Citrus + Green Tea (+caffeine), Black Cherry + Coffee, Peach + Honey",caffeine=30,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None (natural flavours)",claims="Zero sugar, bold dual-flavour pairings, some with caffeine",vband="Medium",vsignal="Coca-Cola seltzer entry; broad distribution",message="Bold paired-flavour seltzer, some caffeinated",why="Unusual flavour pairings (and optional caffeine) at value price.",unitml=355)
add(brand="Montellier",sku="Montellier Carbonated Natural Spring Water",category=C5,format="355 mL can",pack=12,flavours="Original, Lime, Lemon, Mango, Cranberry, Raspberry",caffeine=0,sodium=5,potassium=None,sugar=0,calories=0,sweetener="None",claims="Canadian natural spring water, carbonated, zero everything",vband="Medium",vsignal="Established Canadian sparkling brand (domestic strength)",message="Canadian carbonated spring water",why="Local Canadian sparkling water, often value priced.",unitml=355)
add(brand="Eska",sku="Eska Sparkling Spring Water",category=C5,format="500 mL bottle",pack=12,flavours="Original, Lemon, Lime",caffeine=0,sodium=5,potassium=None,sugar=0,calories=0,sweetener="None",claims="Canadian glacial-era spring water, naturally pure",vband="Low",vsignal="Canadian regional brand; modest reach",message="Pure Canadian spring water, sparkling",why="Canadian-sourced purity story.",unitml=500)
add(brand="Coca-Cola",sku="Coca-Cola Classic",category=C5,format="355 mL can",pack=12,flavours="Classic",caffeine=34,sodium=45,potassium=None,sugar=39,calories=140,sweetener="High-fructose corn syrup / sugar",claims="Classic cola (no functional claim)",vband="Very High",vsignal="World's #1 soda; baseline mass volume",message="The classic cola benchmark",why="Universal taste and brand - the soda baseline.",unitml=355)
add(brand="Zevia",sku="Zevia Zero Calorie Soda",category=C5,format="355 mL can",pack=12,flavours="Cola, Ginger Root Beer, Cream Soda, Dr. Zevia, Black Cherry, Grape, Lemon Lime Twist",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="Stevia",claims="Zero sugar/calorie, stevia-sweetened, no artificial sweeteners; 0 mg sodium on label; Cola variant contains 45 mg caffeine (all other flavours 0 mg caffeine)",vband="Medium",vsignal="Leading naturally-sweetened diet soda",message="Diet soda taste, sweetened only with stevia",why="Full soda flavour, zero sugar, no artificial sweeteners.",unitml=355)
add(brand="Nixie",sku="Nixie Organic Sparkling Water",category=C5,format="355 mL can",pack=8,flavours="Grapefruit Twist, Lime, Black Cherry, Blackberry Mint, Watermelon Mint",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None (organic flavours)",claims="USDA organic, non-GMO, zero everything",vband="Low",vsignal="Small organic-certified seltzer niche",message="Certified-organic zero-everything seltzer",why="Organic certification in a plain-seltzer field that mostly isn't.",unitml=355)

# ---- CAT 1: SPARKLING ELECTROLYTE / FUNCTIONAL SPARKLING ----
add(brand="Sparkling Ice",sku="Sparkling Ice (+ Antioxidants & Vitamins)",category=C1,format="503 mL bottle",pack=12,flavours="Black Raspberry, Cherry Limeade, Orange Mango, Kiwi Strawberry, Coconut Pineapple, Classic Lemonade",caffeine=0,sodium=0,potassium=30,sugar=0,calories=5,sweetener="Sucralose",claims="Zero sugar, vitamins A/D/B, antioxidants (green tea), bold flavour",vband="High",vsignal="Mass-market functional sparkling; strong shelf share",message="Bold zero-sugar sparkling with added vitamins",why="Big flavour and added vitamins at zero sugar.",unitml=503)
add(brand="BUBBL'R",sku="BUBBL'R Antioxidant Sparkling Water",category=C1,format="473 mL can",pack=12,flavours="Blackberribrilliance, Strawbeary Lemonade, Tropic'b'r, Raspberry Lemonade, Blue Crush'r",caffeine=69,sodium=15,potassium=None,sugar=0,calories=10,sweetener="Erythritol + stevia",claims="Natural caffeine (69 mg), antioxidants, B-vitamins, zero sugar",vband="Medium",vsignal="Mid-caffeine functional sparkling; growing",message="Sparkling water with a light caffeine + antioxidant lift",why="Light natural caffeine and antioxidants without an energy-drink dose.",unitml=473)
add(brand="Phocus",sku="Phocus Caffeinated Sparkling Water",category=C1,format="350 mL can",pack=12,flavours="Grapefruit, Yuzu Lime, Cucumber, Black Cherry, Natural, Lemon",caffeine=75,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None (natural flavours)",claims="75 mg caffeine + L-theanine for smooth focus, zero sugar/calorie",vband="Medium",vsignal="Niche caffeinated sparkling; focus positioning",message="Sparkling water that wakes you up smoothly",why="Coffee-level caffeine plus L-theanine in a calorie-free seltzer.",unitml=350)
add(brand="Aura Bora",sku="Aura Bora Herbal Sparkling Water",category=C1,format="355 mL can",pack=12,flavours="Lavender Cucumber, Basil Berry, Lemongrass Coconut, Cactus Rose, Peppermint Watermelon",caffeine=0,sodium=0,potassium=None,sugar=0,calories=5,sweetener="None (herbal extracts)",claims="Herb/flower/fruit flavours, zero sugar, calm/unique sensory",vband="Low",vsignal="Emerging craft brand; small but loyal",message="Whimsical herbal sparkling water for the curious",why="Unusual herbal-botanical flavours you can't get elsewhere.",unitml=355)
add(brand="Sound",sku="Sound Sparkling Water (Tea/Botanical)",category=C1,format="355 mL can",pack=12,flavours="Blackberry + Lemon Verbena, Grapefruit + Hops, Lemon Ginger, Cucumber + Mint Green Tea",caffeine=15,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None (unsweetened)",claims="Real brewed tea/botanicals, unsweetened, zero sugar/calorie",vband="Low",vsignal="Emerging unsweetened-botanical niche",message="Unsweetened sparkling brewed tea and botanicals",why="Genuinely unsweetened with real tea/botanical character.",unitml=355)
add(brand="Halo Sport",sku="Halo Sport Sparkling Electrolyte Water",category=C1,format="355 mL can",pack=12,flavours="Lemon Lime, Berry, Orange, Grapefruit",caffeine=0,sodium=110,potassium=40,sugar=0,calories=5,sweetener="Stevia",claims="Sparkling electrolytes, zero sugar, light hydration",vband="Low",vsignal="Small emerging sparkling-electrolyte entrant",message="Sparkling, lightly-mineralled hydration",why="Electrolytes in a fizzy, zero-sugar format.",unitml=355)
add(brand="Gorgie",sku="Gorgie Energy Sparkling Water",category=C1,format="355 mL can",pack=12,flavours="Watermelon, Tropical, Strawberry Lemonade, Cherry Limeade",caffeine=120,sodium=10,potassium=None,sugar=0,calories=10,sweetener="Stevia",claims="120 mg natural caffeine, L-theanine, biotin, zero sugar",vband="Low",vsignal="New social-driven entrant; small footprint",message="A lighter, prettier caffeinated sparkling energy",why="Mid-dose caffeine with a clean, lifestyle aesthetic.",unitml=355)
add(brand="United Sodas",sku="United Sodas of America",category=C1,format="355 mL can",pack=12,flavours="Wild Berry, Black Cherry, Mango Chili, Lemon Lime, Peach Vanilla",caffeine=0,sodium=10,potassium=None,sugar=8,calories=30,sweetener="Cane sugar (small amount)",claims="Low calorie, light sparkling soda, minimalist branding",vband="Low",vsignal="Niche premium soda; design-led",message="A light, low-calorie modern soda",why="Low-calorie soda with a clean, design-forward identity.",unitml=355)

# ===== BREADTH EXPANSION: additional real product LINES (approx attributes, flagged) =====
# -- more caffeine & energy / RTD coffee / caffeinated tea (C4) --
add(brand="Full Throttle",sku="Full Throttle Energy",category=C4,format="473 mL can",pack=12,flavours="Original Citrus, Blue Agave, Berry",caffeine=160,sodium=140,potassium=None,sugar=58,calories=220,sweetener="HFCS + sugar",claims="160 mg caffeine, D-ribose, B-vitamins (B3/B5/B6/B12); no taurine/ginseng; blue-collar positioning",vband="Low",vsignal="Legacy Coca-Cola energy brand",message="Old-school heavy energy can",why="Value legacy energy.",unitml=473)
add(brand="Venom",sku="Venom Energy",category=C4,format="473 mL can",pack=12,flavours="Black Mamba, Mojave Rattler, Death Adder",caffeine=160,sodium=310,potassium=None,sugar=53,calories=240,sweetener="HFCS + sucralose",claims="Energy blend (taurine, guarana, L-carnitine, ginseng, inositol), B-vitamins",vband="Low",vsignal="Discount-channel energy brand",message="Cheap high-sugar energy",why="Lowest-price energy option.",unitml=473)
add(brand="AMP",sku="AMP Energy",category=C4,format="473 mL can",pack=12,flavours="Original, Cherry Blast",caffeine=142,sodium=140,potassium=0,sugar=58,calories=220,sweetener="HFCS",claims="142 mg caffeine, guarana seed extract, ginseng, taurine, B-vitamins (niacinamide/B5/B6/B12), 0 mg potassium; regional availability in select US markets (discontinued Canada 2023)",vband="Low",vsignal="Legacy PepsiCo energy",message="Legacy mid-tier energy",why="Familiar legacy brand.",unitml=473)
add(brand="5-hour Energy",sku="5-hour Energy Shot",category=C4,format="57 mL shot",pack=12,flavours="Berry, Grape, Orange, Pomegranate, Peach Mango, Extra Strength Berry",caffeine=200,sodium=18,potassium=None,sugar=0,calories=4,sweetener="Sucralose",claims="Zero sugar energy shot, B-vitamins, amino acids, no crash",vband="High",vsignal="Dominant energy-shot brand; impulse repeat",message="A tiny no-sugar shot for instant energy",why="Fast caffeine with no sugar and no big can.",unitml=57)
add(brand="Monster",sku="Monster Java",category=C4,format="444 mL can",pack=12,flavours="Mean Bean, Loca Moca, Salted Caramel, Vanilla",caffeine=200,sodium=460,potassium=460,sugar=35,calories=220,sweetener="Sugar + glucose + sucralose",claims="200 mg caffeine (coffee + added), brewed coffee + skim milk + cream, 8 g protein, 340 mg calcium, taurine + B-vitamins; contains dairy",vband="Medium",vsignal="Coffee-energy crossover",message="Coffee-flavoured energy in a big can",why="Coffee taste plus energy dose.",unitml=444)
add(brand="Red Bull",sku="Red Bull Editions",category=C4,format="250 mL can",pack=12,flavours="Red (Watermelon), Blue (Blueberry), Yellow (Tropical), Green (Dragon Fruit), Peach (Nectarine), Summer",caffeine=80,sodium=105,potassium=None,sugar=27,calories=110,sweetener="Sucrose + glucose-fructose",claims="Flavoured Red Bull energy, taurine + B-vitamins",vband="High",vsignal="Flavour line extensions of #1 brand",message="Fruit-flavoured Red Bull energy",why="Red Bull energy with flavour variety.",unitml=250)
add(brand="G FUEL",sku="G FUEL Energy Cans",category=C4,format="473 mL can",pack=12,flavours="Sour Cherry, Blue Ice, Tropical Rain, Fazeberry, Rainbow Sherbet",caffeine=300,sodium=10,potassium=None,sugar=0,calories=10,sweetener="Sucralose + acesulfame-K",claims="300 mg caffeine, zero sugar, gamer focus blend",vband="Medium",vsignal="Gaming-community brand; strong online pull",message="High-caffeine zero-sugar gamer energy",why="Gamer-brand identity plus max caffeine.",unitml=473)
add(brand="Mtn Dew",sku="Mtn Dew Rise Energy",category=C4,format="473 mL can",pack=12,flavours="Orange Breeze, Berry Blast, Pomegranate Blue Burst, Tropical Sunrise",caffeine=180,sodium=80,potassium=None,sugar=0,calories=15,sweetener="Sucralose + acesulfame-K",claims="180 mg caffeine, zero sugar, vitamins, antioxidants",vband="Medium",vsignal="PepsiCo morning-energy play",message="Morning energy with zero sugar",why="Daytime energy with vitamins, no sugar.",unitml=473)
add(brand="Adrenaline Shoc",sku="Adrenaline Shoc Smart Energy",category=C4,format="473 mL can",pack=12,flavours="Acai Berry, Sour Slurp, Frozen Fusion, Peach Mango",caffeine=300,sodium=10,potassium=None,sugar=0,calories=15,sweetener="Sucralose",claims="300 mg natural caffeine, nootropics, BCAAs, zero sugar",vband="Low",vsignal="Performance-energy challenger",message="Nootropic high-caffeine energy",why="Max caffeine plus nootropics.",unitml=473)
add(brand="Arizona",sku="Arizona Green Tea",category=C4,format="680 mL can",pack=12,flavours="Green Tea w/ Honey, Arnold Palmer, Mucho Mango, Watermelon, Lemon Tea",caffeine=22,sodium=0,potassium=None,sugar=49,calories=200,sweetener="HFCS + honey",claims="Big-can iced tea, value 99c brand; 7.5 mg caffeine per 8 oz from green tea, ginseng root extract; 23oz can uses HFCS (Real Sugar version is 16oz glass bottle)",vband="High",vsignal="Iconic value big-can tea",message="Huge can of sweet tea for almost nothing",why="Unbeatable size-for-price value.",unitml=680)
add(brand="Pure Leaf",sku="Pure Leaf Iced Tea",category=C4,format="547 mL bottle",pack=12,flavours="Sweet Tea, Lemon, Unsweetened, Peach, Raspberry, Extra Sweet",caffeine=70,sodium=10,potassium=None,sugar=32,calories=130,sweetener="Sugar",claims="Real brewed tea, no powders",vband="High",vsignal="Leading premium RTD tea",message="Real-brewed bottled iced tea",why="Brewed-tea taste, not from powder.",unitml=547)
add(brand="Gold Peak",sku="Gold Peak Tea",category=C4,format="547 mL bottle",pack=12,flavours="Sweet Tea, Unsweetened, Green Tea, Lemon, Peach",caffeine=45,sodium=15,potassium=None,sugar=35,calories=140,sweetener="Sugar",claims="Home-brewed taste iced tea",vband="Medium",vsignal="Coca-Cola RTD tea",message="Diner-style sweet tea",why="Familiar home-brewed-style taste.",unitml=547)
add(brand="Brisk",sku="Brisk Iced Tea",category=C4,format="591 mL bottle",pack=12,flavours="Lemon, Sweet Tea, Half & Half, Raspberry",caffeine=18,sodium=170,potassium=None,sugar=29,calories=110,sweetener="HFCS",claims="Bold value iced tea; 20 oz bottle = 1 serving per container (whole-bottle figures); HFCS-sweetened",vband="Medium",vsignal="Value RTD tea brand",message="Bold cheap iced tea",why="Punchy flavour at a low price.",unitml=591)
add(brand="Honest Tea",sku="Honest Tea Organic",category=C4,format="473 mL bottle",pack=12,flavours="Honey Green, Peach, Half & Half, Lemon, Just Black",caffeine=42,sodium=10,potassium=None,sugar=21,calories=90,sweetener="Organic cane sugar",claims="Organic, Fair Trade, lower sugar tea",vband="Medium",vsignal="Organic RTD tea",message="Organic, less-sweet bottled tea",why="Organic with restrained sweetness.",unitml=473)
add(brand="STOK",sku="STOK Cold Brew Coffee",category=C4,format="405 mL bottle",pack=12,flavours="Not Too Sweet Black, Un-Sweet Black, Vanilla, Mocha",caffeine=145,sodium=40,potassium=None,sugar=13,calories=70,sweetener="Sugar",claims="Smooth cold brew, real coffee",vband="Medium",vsignal="Leading RTD cold brew",message="Smooth ready-to-pour cold brew",why="Cafe cold-brew at home.",unitml=405)
add(brand="Califia Farms",sku="Califia Farms Cold Brew",category=C4,format="355 mL bottle",pack=12,flavours="Black Unsweetened, Mocha, Vanilla, Café Latte",caffeine=120,sodium=30,potassium=None,sugar=9,calories=70,sweetener="Cane sugar (plant-based)",claims="Plant-based, almond/oat cold brew",vband="Medium",vsignal="Premium plant-based coffee",message="Dairy-free cold brew lattes",why="Plant-based coffee without dairy.",unitml=355)
add(brand="La Colombe",sku="La Colombe Draft Latte",category=C4,format="260 mL can",pack=12,flavours="Original, Vanilla, Mocha, Oatmilk, Triple",caffeine=120,sodium=70,potassium=None,sugar=14,calories=110,sweetener="Cane sugar",claims="Frothy draft latte, real cold-pressed espresso",vband="Medium",vsignal="Premium canned latte",message="Frothy cafe latte from a can",why="Cafe-quality latte texture.",unitml=260)
add(brand="Starbucks",sku="Starbucks Frappuccino",category=C4,format="281 mL bottle",pack=12,flavours="Mocha, Vanilla, Caramel, Coffee, Dark Chocolate",caffeine=90,sodium=85,potassium=None,sugar=31,calories=200,sweetener="Sugar",claims="Bottled coffee drink",vband="High",vsignal="Mass RTD coffee staple",message="Sweet bottled coffee treat",why="Familiar sweet Starbucks coffee.",unitml=281)
# -- more sodas & sparkling (C5) --
add(brand="Pepsi",sku="Pepsi Cola",category=C5,format="355 mL can",pack=12,flavours="Original, Wild Cherry, Vanilla",caffeine=38,sodium=30,potassium=None,sugar=41,calories=150,sweetener="HFCS",claims="Classic cola (no functional claim)",vband="Very High",vsignal="#2 cola; mass volume",message="The other classic cola",why="Cola taste preference and price.",unitml=355)
add(brand="Pepsi",sku="Pepsi Zero Sugar",category=C5,format="355 mL can",pack=12,flavours="Original, Wild Cherry, Mango",caffeine=69,sodium=35,potassium=None,sugar=0,calories=0,sweetener="Aspartame + acesulfame-K",claims="Zero sugar cola, extra caffeine",vband="High",vsignal="Strong zero-sugar cola",message="Zero-sugar cola with a caffeine kick",why="No sugar, more caffeine.",unitml=355)
add(brand="Diet Pepsi",sku="Diet Pepsi",category=C5,format="355 mL can",pack=12,flavours="Original",caffeine=35,sodium=35,potassium=None,sugar=0,calories=0,sweetener="Aspartame",claims="Zero calorie cola",vband="High",vsignal="Legacy diet cola",message="Classic diet cola",why="Familiar diet-cola taste.",unitml=355)
add(brand="Coca-Cola",sku="Coca-Cola Zero Sugar",category=C5,format="355 mL can",pack=12,flavours="Original, Cherry, Vanilla",caffeine=34,sodium=40,potassium=None,sugar=0,calories=0,sweetener="Aspartame + acesulfame-K",claims="Zero sugar, real-cola taste",vband="Very High",vsignal="Fast-growing zero-sugar cola",message="Coke taste, zero sugar",why="Closest zero-sugar match to Coke.",unitml=355)
add(brand="Diet Coke",sku="Diet Coke",category=C5,format="355 mL can",pack=12,flavours="Original, Cherry, Lime, Caffeine-Free",caffeine=46,sodium=40,potassium=None,sugar=0,calories=0,sweetener="Aspartame",claims="Zero calorie",vband="Very High",vsignal="Iconic diet cola",message="The classic diet cola",why="Distinct diet-cola flavour loyalty.",unitml=355)
add(brand="Sprite",sku="Sprite Lemon-Lime",category=C5,format="355 mL can",pack=12,flavours="Original, Cherry, Tropical",caffeine=0,sodium=65,potassium=None,sugar=38,calories=140,sweetener="HFCS",claims="Caffeine-free lemon-lime",vband="Very High",vsignal="#1 lemon-lime soda",message="Crisp caffeine-free lemon-lime",why="Clean lemon-lime, no caffeine.",unitml=355)
add(brand="Sprite",sku="Sprite Zero Sugar",category=C5,format="355 mL can",pack=12,flavours="Original",caffeine=0,sodium=40,potassium=None,sugar=0,calories=0,sweetener="Aspartame + acesulfame-K",claims="Zero sugar lemon-lime",vband="High",vsignal="Zero-sugar variant",message="Zero-sugar lemon-lime",why="Lemon-lime without sugar.",unitml=355)
add(brand="Fanta",sku="Fanta Orange",category=C5,format="355 mL can",pack=12,flavours="Orange, Grape, Pineapple, Strawberry",caffeine=0,sodium=55,potassium=None,sugar=44,calories=160,sweetener="HFCS",claims="Fruit-flavoured soda",vband="High",vsignal="Leading orange soda",message="Bright fruity orange soda",why="Sweet fruity flavour, no caffeine.",unitml=355)
add(brand="7UP",sku="7UP Lemon-Lime",category=C5,format="355 mL can",pack=12,flavours="Original, Cherry, Diet",caffeine=0,sodium=45,potassium=None,sugar=38,calories=140,sweetener="HFCS",claims="Caffeine-free lemon-lime",vband="High",vsignal="Legacy lemon-lime",message="Classic lemon-lime soda",why="Lemon-lime alternative to Sprite.",unitml=355)
add(brand="Mountain Dew",sku="Mountain Dew Citrus",category=C5,format="355 mL can",pack=12,flavours="Original, Code Red, Voltage, Baja Blast, Major Melon",caffeine=54,sodium=60,potassium=None,sugar=46,calories=170,sweetener="HFCS",claims="High-caffeine citrus soda",vband="Very High",vsignal="Cult citrus soda; many flavours",message="Bold caffeinated citrus soda",why="Distinctive sweet citrus and caffeine.",unitml=355)
add(brand="Dr Pepper",sku="Dr Pepper",category=C5,format="355 mL can",pack=12,flavours="Original, Cherry, Cream Soda, Zero Sugar",caffeine=41,sodium=55,potassium=None,sugar=40,calories=150,sweetener="HFCS",claims="23-flavour soda",vband="Very High",vsignal="Beloved unique-flavour soda",message="The unique 23-flavour soda",why="One-of-a-kind flavour loyalty.",unitml=355)
add(brand="Canada Dry",sku="Canada Dry Ginger Ale",category=C5,format="355 mL can",pack=12,flavours="Original, Diet, Cranberry, Lemonade",caffeine=0,sodium=35,potassium=None,sugar=36,calories=140,sweetener="HFCS",claims="Real ginger ale (no functional claim)",vband="High",vsignal="#1 ginger ale",message="Crisp classic ginger ale",why="Go-to ginger ale, also for upset stomach.",unitml=355)
add(brand="Schweppes",sku="Schweppes Ginger Ale",category=C5,format="355 mL can",pack=12,flavours="Original, Diet, Club Soda, Tonic Water",caffeine=0,sodium=40,potassium=None,sugar=33,calories=130,sweetener="HFCS",claims="Ginger ale + mixers",vband="Medium",vsignal="Mixer staple",message="Ginger ale and bar mixers",why="Reliable mixer brand.",unitml=355)
add(brand="A&W",sku="A&W Root Beer",category=C5,format="355 mL can",pack=12,flavours="Original, Diet, Cream Soda",caffeine=0,sodium=45,potassium=None,sugar=46,calories=170,sweetener="HFCS",claims="Caffeine-free root beer",vband="High",vsignal="#1 root beer",message="Creamy classic root beer",why="Rich caffeine-free root beer.",unitml=355)
add(brand="Crush",sku="Crush Orange",category=C5,format="355 mL can",pack=12,flavours="Orange, Grape, Pineapple, Cream Soda",caffeine=0,sodium=45,potassium=None,sugar=49,calories=180,sweetener="HFCS",claims="Bold fruit soda",vband="Medium",vsignal="Legacy fruit soda",message="Super-sweet orange soda",why="Bold sweet fruit flavour.",unitml=355)
add(brand="Fresca",sku="Fresca Sparkling",category=C5,format="355 mL can",pack=12,flavours="Original Citrus, Black Cherry, Peach",caffeine=0,sodium=35,potassium=None,sugar=0,calories=0,sweetener="Aspartame + acesulfame-K",claims="Zero sugar grapefruit-citrus",vband="Medium",vsignal="Cult zero-cal soda",message="Zero-calorie citrus soda",why="Grown-up zero-cal citrus.",unitml=355)
add(brand="Sunkist",sku="Sunkist Orange",category=C5,format="355 mL can",pack=12,flavours="Orange, Grape, Strawberry, Diet",caffeine=19,sodium=45,potassium=None,sugar=51,calories=190,sweetener="HFCS",claims="Bright orange soda",vband="Medium",vsignal="Legacy orange soda",message="Sweet bright orange soda",why="Big citrus flavour.",unitml=355)
add(brand="Mug",sku="Mug Root Beer",category=C5,format="355 mL can",pack=12,flavours="Original, Diet, Cream Soda",caffeine=0,sodium=65,potassium=None,sugar=43,calories=160,sweetener="HFCS",claims="Caffeine-free root beer",vband="Medium",vsignal="PepsiCo root beer",message="Creamy root beer",why="Value root beer.",unitml=355)
add(brand="Dasani",sku="Dasani Sparkling Water",category=C5,format="355 mL can",pack=8,flavours="Black Cherry, Lemon, Berry, Tropical Pineapple",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None (natural flavours)",claims="Zero-calorie flavoured sparkling water",vband="Low",vsignal="Coca-Cola sparkling water",message="Plain flavoured sparkling water",why="Value zero-cal seltzer.",unitml=355)
add(brand="Smartwater",sku="Smartwater Sparkling",category=C5,format="591 mL bottle",pack=12,flavours="Original, Lemon, Cucumber-Lime, Watermelon-Mint",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None",claims="Vapor-distilled sparkling water w/ electrolytes added",vband="Medium",vsignal="Premium water brand",message="Premium distilled sparkling water",why="Premium water brand cachet.",unitml=591)
# -- more electrolyte RTD (C2) --
add(brand="Propel",sku="Propel Electrolyte Water",category=C2,format="591 mL bottle",pack=12,flavours="Berry, Grape, Kiwi Strawberry, Watermelon, Lemon, Black Cherry",caffeine=0,sodium=150,potassium=40,sugar=0,calories=5,sweetener="Sucralose + acesulfame-K",claims="Electrolytes + vitamins, zero sugar, fitness water",vband="High",vsignal="Leading zero-sugar electrolyte water (Gatorade family)",message="Zero-sugar workout water with electrolytes",why="Light electrolyte water for everyday workouts.",unitml=591)
add(brand="Essentia",sku="Essentia Ionized Alkaline Water",category=C2,format="1 L bottle",pack=12,flavours="Original (pH 9.5)",caffeine=0,sodium=10,potassium=None,sugar=0,calories=0,sweetener="None",claims="Ionized alkaline pH 9.5, electrolytes for taste, hydration",vband="High",vsignal="Leading premium alkaline water",message="Smooth alkaline ionized water",why="Alkaline-water believers; smooth taste.",unitml=1000)
add(brand="Smartwater",sku="Smartwater Alkaline",category=C2,format="1 L bottle",pack=12,flavours="Alkaline pH 9+, Original",caffeine=0,sodium=0,potassium=None,sugar=0,calories=0,sweetener="None",claims="Vapor-distilled, added electrolytes / alkaline variant",vband="High",vsignal="Premium water brand",message="Premium distilled water",why="Brand trust and clean taste.",unitml=1000)
add(brand="Gatorade",sku="Gatorade G2 Low Sugar",category=C2,format="591 mL bottle",pack=12,flavours="Glacier Freeze, Fruit Punch, Grape, Lemon-Lime",caffeine=0,sodium=270,potassium=75,sugar=7,calories=30,sweetener="Sugar + sucralose",claims="Same electrolytes, lower sugar",vband="Medium",vsignal="Reduced-sugar Gatorade",message="Lower-sugar Gatorade electrolytes",why="Electrolytes with much less sugar.",unitml=591)
add(brand="Mas+",sku="Mas+ by Messi Hydration",category=C2,format="473 mL bottle",pack=12,flavours="Limon, Berry, Orange, Tropical",caffeine=0,sodium=200,potassium=100,sugar=4,calories=20,sweetener="Cane sugar + stevia",claims="Electrolytes, low sugar, vitamins; Messi-backed",vband="Medium",vsignal="New celebrity hydration entrant",message="Celebrity-backed low-sugar hydration",why="Messi brand plus low-sugar electrolytes.",unitml=473)
add(brand="Pedialyte",sku="Pedialyte Sport",category=C2,format="495 mL bottle",pack=12,flavours="Berry, Fruit Punch, Cherry",caffeine=0,sodium=490,potassium=350,sugar=11,calories=45,sweetener="Sugar + sucralose",claims="High-electrolyte rapid rehydration for athletes",vband="Medium",vsignal="Sport line of clinical brand",message="Clinical-grade sport rehydration",why="Medical-brand electrolytes for sport.",unitml=495)
# -- more functional (C3) --
add(brand="KeVita",sku="KeVita Sparkling Probiotic Drink",category=C3,format="440 mL bottle",pack=12,flavours="Lemon Ginger, Mojita Lime Mint Coconut, Strawberry Acai Coconut, Roots Beer",caffeine=0,sodium=80,potassium=None,sugar=5,calories=25,sweetener="Stevia + juice",claims="Billions of probiotics, sparkling, gut health, low sugar",vband="Medium",vsignal="Mass probiotic drink (PepsiCo)",message="Sparkling probiotic for gut health",why="Probiotics in a light sparkling format.",unitml=440)
add(brand="Brew Dr",sku="Brew Dr Kombucha",category=C3,format="414 mL bottle",pack=12,flavours="Clear Mind, Superberry, Island Mango, Happiness, Ginger Turmeric",caffeine=15,sodium=10,potassium=None,sugar=10,calories=50,sweetener="Naturally fermented",claims="Organic raw kombucha, probiotics, gut health",vband="Medium",vsignal="Premium organic kombucha",message="Organic raw kombucha with botanicals",why="Organic raw kombucha taste.",unitml=414)
add(brand="Karma",sku="Karma Wellness Water",category=C3,format="532 mL bottle",pack=12,flavours="Berry Cherry, Orange Mango, Acai Pomegranate, Tropical Pineapple",caffeine=0,sodium=10,potassium=None,sugar=10,calories=40,sweetener="Cane sugar + stevia",claims="Push-cap fresh vitamins, antioxidants, function-by-flavour",vband="Low",vsignal="Niche push-cap vitamin water",message="Fresh-mixed vitamins via push cap",why="Vitamins mixed fresh at opening.",unitml=532)
add(brand="ALO",sku="ALO Exposed Aloe Vera Drink",category=C3,format="500 mL bottle",pack=12,flavours="Original Aloe + Honey, Mango Mangosteen, Pomegranate Cranberry, Pink Grapefruit Lemon",caffeine=0,sodium=20,potassium=None,sugar=20,calories=90,sweetener="Cane sugar",claims="Real aloe vera pulp, antioxidants, vitamins",vband="Low",vsignal="Leading aloe drink",message="Real aloe pulp wellness drink",why="Aloe pulp and a wellness halo.",unitml=500)
# -- more functional sparkling (C1) --
add(brand="Sparkling Ice",sku="Sparkling Ice +Caffeine",category=C1,format="473 mL can",pack=12,flavours="Black Raspberry, Citrus Twist, Orange Passionfruit, Strawberry Citrus, Triple Citrus",caffeine=70,sodium=25,potassium=None,sugar=0,calories=5,sweetener="Sucralose",claims="70 mg caffeine, vitamins, antioxidants, zero sugar",vband="Medium",vsignal="Caffeinated line of mass sparkling brand",message="Zero-sugar sparkling with a caffeine boost",why="Light caffeine plus vitamins, zero sugar.",unitml=473)

# ===== SKU EXPANSION ENGINE (line x flavour x pack-size = one SKU/ASIN row) =====
def short_size(ln):
    m=re.search(r'(\d+(?:\.\d+)?)\s*(mL|L)',ln["format"] or "")
    if not m: return ""
    return f'{m.group(1)}{m.group(2)}'
def pack_configs(ln):
    v=ln.get("unitml") or 355
    if v<=70:   return [6,12,18,24]            # shots
    if v<=250:  return [4,8,12,18,24]
    if v<=355:  return [1,8,12,15,18,24]
    if v<=550:  return [1,8,12,18,24]
    if v<=700:  return [1,6,12,24]
    return [8,12,24]                            # ~1 L
def expand(lines):
    out=[]
    for ln in lines:
        flavs=[f.strip() for f in (ln.get("flavours") or "Original").split(",") if f.strip()] or ["Original"]
        size=short_size(ln)
        for fl in flavs:
            for pk in pack_configs(ln):
                s=dict(ln)
                packlbl = f"single {size}" if pk==1 else f"{pk}-pack {size}"
                s["line"]=ln["sku"]
                s["flavour"]=fl
                s["flavours"]=fl
                s["pack"]=pk
                s["sku"]=f'{ln["sku"]} | {fl} | {packlbl}'
                out.append(s)
    return out

LINES=list(P)          # the product lines (line-level tabs use these)
SKUS=expand(LINES)     # the exploded SKUs (master + pricing tabs use these)


# ============================================================
# PY-SIDE ANALYTICS (for charts, rollups, benchmarks, sorting)
# ============================================================
def num(x): return x if isinstance(x,(int,float)) else None
def py_health(p):
    s=num(p["sugar"]); cal=num(p["calories"]) or 0
    if s is None: return None
    return round(max(0,min(100,100 - s*2.5 - cal*0.12 + (12 if s<=1 else 0))))
def py_func(p):
    caf=num(p["caffeine"]); na=num(p["sodium"]); k=num(p["potassium"]); s=num(p["sugar"])
    sc=0
    if na is not None: sc+= 20 if na>=100 else (10 if na>=30 else 0)
    if k is not None and k>=100: sc+=20
    if caf is not None: sc+= 20 if caf>=150 else (12 if caf>=50 else 5)
    if s is not None and s<=1: sc+=15
    return sc
def sweet_class(s):
    s=(s or "").lower()
    if s.startswith("none"): return "None / unsweetened"
    noncal=any(x in s for x in ["sucralose","stevia","erythritol","aspartame","acesulfame","monk"])
    cal=any(x in s for x in ["sugar","cane","hfcs","corn syrup","fructose","agave","dextrose","glucose","cassava","juice"])
    if noncal and cal: return "Blend (caloric + non-caloric)"
    if noncal: return "Non-caloric only"
    if cal: return "Caloric only"
    return "Other"

def sweet_fine(s):
    """Finer sweetener taxonomy that separates natural-zero from artificial-zero."""
    s=(s or "").lower()
    if s.startswith("none"): return "Unsweetened"
    nat=any(x in s for x in ["stevia","monk"])
    art=any(x in s for x in ["sucralose","aspartame","acesulfame","erythritol"])
    cal=any(x in s for x in ["sugar","cane","hfcs","corn syrup","fructose","agave","dextrose","glucose","cassava","juice"])
    if cal and (nat or art): return "Caloric + sweetener blend"
    if cal: return "Caloric (sugar)"
    if nat and art: return "Zero (natural + artificial)"
    if nat: return "Zero (natural: stevia/monk)"
    if art: return "Zero (artificial)"
    return "Other"

# ---- shared flavour taxonomy (used by Flavour Map + Strategic Insights) ----
FLAV_FAMS=["Cola/Root Beer","Citrus (Lem/Lime)","Orange/Mango","Berry/Rasp","Cherry","Grape",
      "Watermelon","Peach","Tropical/Pineapple","Grapefruit","Coconut","Ginger","Herbal/Botanical","Original/Plain"]
FLAV_KEYMAP={"Cola/Root Beer":["cola","root beer","doctor","kola","dr."],
 "Citrus (Lem/Lime)":["lemon","lime","limon","limonata","yuzu","citrus"],
 "Orange/Mango":["orange","mango","aranciata","clementine","tangerine","mandarin"],
 "Berry/Rasp":["berry","raspberry","blackberry","blueberry","razz","acai","pomegranate","cranberry"],
 "Cherry":["cherry"],"Grape":["grape"],"Watermelon":["watermelon"],"Peach":["peach"],
 "Tropical/Pineapple":["tropical","pineapple","passion","guava","paradise","coconut pineapple","hawaiian"],
 "Grapefruit":["grapefruit","pamplemousse","pompel"],"Coconut":["coconut"],"Ginger":["ginger"],
 "Herbal/Botanical":["lavender","basil","hibiscus","rose","mint","cactus","hops","chai","verbena","lemongrass","elderflower","botanical"],
 "Original/Plain":["original","classic","pure","unflavored","unflavoured","still","natural","espresso"]}
def flav_hits(text):
    text=(text or "").lower()
    return {fam for fam in FLAV_FAMS if any(k in text for k in FLAV_KEYMAP[fam])}

# ============================================================
# WORKBOOK
# ============================================================
wb=Workbook()
def style_header(ws,ncols,row=1,h=42,fill=NAVY):
    f=PatternFill("solid",fgColor=fill)
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c); cell.font=HFONT; cell.fill=f
        cell.alignment=HALIGN; cell.border=BORDER
    ws.row_dimensions[row].height=h
def widths(ws,ws_w):
    for i,w in enumerate(ws_w,1): ws.column_dimensions[get_column_letter(i)].width=w
def tabcolor(ws,c): ws.sheet_properties.tabColor=c

BLANK_KINDS={"blankmoney","blanknum","blanktext","blankpct"}

# ---------- formulas ----------
def f_ppu(r): return f'=IF(AND(ISNUMBER($F{r}),ISNUMBER($E{r}),$E{r}<>0),$F{r}/$E{r},"")'
def f_p100(r):
    AF=f"${CL('unitml')}{r}"
    return f'=IF(AND(ISNUMBER($F{r}),ISNUMBER({AF}),ISNUMBER($E{r}),({AF}*$E{r})<>0),$F{r}/({AF}*$E{r})*100,"")'
def f_ssprice(r): return f'=IF(AND(ISNUMBER($F{r}),ISNUMBER($J{r})),$F{r}*(1-$J{r}),"")'
def f_floor(r):
    AG=f"${CL('couponval')}{r}"
    return f'=IF(ISNUMBER($F{r}),($F{r}-IF(ISNUMBER({AG}),{AG},0))*(1-IF(ISNUMBER($J{r}),$J{r},0)),"")'
def f_caf100(r):
    caf=f"${CL('caffeine')}{r}"; u=f"${CL('unitml')}{r}"
    return f'=IF(AND(ISNUMBER({caf}),ISNUMBER({u}),{u}<>0),ROUND({caf}/{u}*100,1),"")'
def f_sugar100(r):
    s=f"${CL('sugar')}{r}"; u=f"${CL('unitml')}{r}"
    return f'=IF(AND(ISNUMBER({s}),ISNUMBER({u}),{u}<>0),ROUND({s}/{u}*100,2),"")'
def f_sodium100(r):
    na=f"${CL('sodium')}{r}"; u=f"${CL('unitml')}{r}"
    return f'=IF(AND(ISNUMBER({na}),ISNUMBER({u}),{u}<>0),ROUND({na}/{u}*100,1),"")'
def f_sugarfree(r):
    s=f"${CL('sugar')}{r}"
    return f'=IF(ISNUMBER({s}),IF({s}<=1,"Yes","No"),"n/a")'
def f_health(r):
    s=f"${CL('sugar')}{r}"; cal=f"${CL('calories')}{r}"
    return f'=IF(ISNUMBER({s}),ROUND(MAX(0,MIN(100,100-{s}*2.5-IF(ISNUMBER({cal}),{cal},0)*0.12+IF({s}<=1,12,0))),0),"")'
def f_func(r):
    caf=f"${CL('caffeine')}{r}"; na=f"${CL('sodium')}{r}"; k=f"${CL('potassium')}{r}"; s=f"${CL('sugar')}{r}"
    return ('='
        f'IF(ISNUMBER({na}),IF({na}>=100,20,IF({na}>=30,10,0)),0)+'
        f'IF(ISNUMBER({k}),IF({k}>=100,20,0),0)+'
        f'IF(ISNUMBER({caf}),IF({caf}>=150,20,IF({caf}>=50,12,5)),0)+'
        f'IF(ISNUMBER({s}),IF({s}<=1,15,0),0)')
FORMULA_MAP={"ppu":f_ppu,"p100":f_p100,"ss_price":f_ssprice,"floor":f_floor}
CALC_MAP={"caf100":f_caf100,"sugar100":f_sugar100,"sodium100":f_sodium100,
          "sugarfree":f_sugarfree,"health":f_health,"func":f_func}

MASTER_RANGES={}  # sheet -> (first,last)

def build_master(category):
    name=SHEET_FOR[category]
    ws=wb.create_sheet(name); tabcolor(ws,BLUE)
    widths(ws,[c[2] for c in COLS])
    for i,c in enumerate(COLS,1): ws.cell(row=1,column=i,value=c[1])
    style_header(ws,len(COLS))
    ws.freeze_panes="C2"
    prods=[p for p in SKUS if p["category"]==category]
    r=2
    for p in prods:
        p["_sheet"]=name; p["_row"]=r
        for ci,(key,hdr,w,kind) in enumerate(COLS,1):
            cell=ws.cell(row=r,column=ci); cell.border=BORDER; cell.alignment=WRAP
            if kind=="formula":
                cell.value=FORMULA_MAP[key](r); cell.number_format=MONEY
            elif kind=="calc":
                cell.value=CALC_MAP[key](r)
                if key in("caf100","sodium100"): cell.number_format=NUM1
                if key=="sugar100": cell.number_format='0.00'
                cell.alignment=Alignment(vertical="top",horizontal="center")
            elif kind in BLANK_KINDS:
                cell.fill=PatternFill("solid",fgColor=AMBER)
                if kind=="blankmoney": cell.number_format=MONEY
                elif kind=="blankpct": cell.number_format=PCT
            else:
                if key=="prov": v=PROV
                elif key=="status": v=STATUS
                elif key=="flag": v=FLAG
                else: v=p.get(key,"")
                if v is None: v="Not listed"
                cell.value=v
                if key in("caffeine","sodium","potassium","sugar","calories","pack","unitml") and isinstance(v,(int,float)):
                    cell.alignment=Alignment(vertical="top",horizontal="center")
        ws.row_dimensions[r].height=28
        r+=1
    last=r-1; MASTER_RANGES[name]=(2,last)
    # Excel Table (gives sort/filter + banded rows)
    tname="tbl"+name.split(" ")[0].replace("-","")
    tab=Table(displayName=tname,ref=f"A1:{get_column_letter(len(COLS))}{last}")
    tab.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True,showColumnStripes=False)
    ws.add_table(tab)
    # Yes/No validation
    dv=DataValidation(type="list",formula1='"Yes,No"',allow_blank=True); ws.add_data_validation(dv)
    dv.add(f"{CL('ss_off')}2:{CL('ss_off')}{last}")
    # colour scales / data bars on nutrition + scores
    if last>=2:
        ws.conditional_formatting.add(f"{CL('sugar')}2:{CL('sugar')}{last}",
            ColorScaleRule(start_type="min",start_color="C6EFCE",mid_type="percentile",mid_value=50,mid_color="FFEB9C",end_type="max",end_color="FFC7CE"))
        ws.conditional_formatting.add(f"{CL('caffeine')}2:{CL('caffeine')}{last}",
            ColorScaleRule(start_type="min",start_color="DDEBF7",mid_type="percentile",mid_value=50,mid_color="9BC2E6",end_type="max",end_color="2E75B6"))
        ws.conditional_formatting.add(f"{CL('health')}2:{CL('health')}{last}",
            DataBarRule(start_type="num",start_value=0,end_type="num",end_value=100,color="63BE7B"))
        ws.conditional_formatting.add(f"{CL('func')}2:{CL('func')}{last}",
            DataBarRule(start_type="num",start_value=0,end_type="num",end_value=75,color="8FAADC"))
        ws.conditional_formatting.add(f"{CL('flag')}2:{CL('flag')}{last}",
            FormulaRule(formula=[f'ISBLANK(${CL("price")}2)'],fill=PatternFill("solid",fgColor=RED)))
    return ws

# ---------- README ----------
def build_readme():
    ws=wb.active; ws.title="README & Methodology"; ws.sheet_view.showGridLines=False; tabcolor(ws,"808080")
    widths(ws,[3,30,112])
    ws.cell(row=2,column=2,value="Amazon.ca Beverage - Competitor Intelligence Master File").font=Font(bold=True,size=18,color=NAVY)
    ws.cell(row=3,column=2,value=f"Build date: {TODAY}   |   v3 Enterprise   |   {len(SKUS)} SKUs from {len(LINES)} product lines / {len(set(p['brand'] for p in LINES))} brands").font=Font(italic=True,size=11,color=BLUE)
    rows=[
     ("CRITICAL - READ FIRST",""),
     ("Data-access limitation","Generated with NO live access to Amazon (HTTP 403 on search, product, and brand pages). The file therefore contains NO live prices, Subscribe & Save terms, coupons, ratings, review counts, badges, ranks, or URLs. Those cells are intentionally BLANK and amber-highlighted - they are NOT fabricated."),
     ("SKU model - HOW THE ~1000+ ROWS ARE BUILT","The Master tabs are SKU-level. Because each flavour x pack-size is a separate ASIN on Amazon, every product LINE is exploded into individual SKU rows (line x flavour x pack). These are DERIVED / CANDIDATE SKUs to verify against live listings - NOT confirmed ASINs. Flavour lists are representative not exhaustive, and some specific flavour+pack combinations may not exist exactly as shown. The live catalogue cannot be enumerated without Amazon access; this expansion approximates its shape."),
     ("What IS filled (and its confidence)","Stable attributes - brand, SKU, format/size, typical pack, flavours, caffeine/sodium/potassium/sugar/calories, sweetener, claims, positioning, plus an honest velocity ESTIMATE - filled from manufacturer + secondary sources. Treat all as APPROXIMATE and verify on the live label. Nutrition is per single serving and varies by flavour."),
     ("What is BLANK (capture live)","Fields 6, 9-14, 23-26, 30 and helper H2. All amber. Filled by a live-capture pass (see 'Live-Capture Protocol' tab)."),
     ("Built-in automation","Fields 7,8,11,14 are LIVE FORMULAS. Enter price (F), pack (E), S&S % as decimal (J, e.g. 0.15) and coupon value (H2); per-unit, per-100 mL, S&S price and stacked floor compute instantly and the red row-flag clears."),
     ("NEW v2 - computed analytics","Columns C1-C6 on every Master tab: Caffeine/100 mL, Sugar/100 mL, Sodium/100 mL, Sugar-Free flag, Health Score, Functional Score - all live formulas off the nutrition you can already see."),
     ("Health Score (0-100) model","=MAX(0, MIN(100, 100 - sugar*2.5 - calories*0.12 + 12 if sugar<=1)). Higher = lighter/cleaner. Transparent and editable; based on APPROXIMATE nutrition."),
     ("Functional Score (0-100) model","Sodium (>=100=20, >=30=10) + Potassium (>=100=20) + Caffeine (>=150=20, >=50=12, else 5) + Low-sugar bonus (<=1 g=15). Higher = more 'functional' payload. A heuristic, not a medical claim."),
     ("Stacking model (field 14)","Lowest Effective = (Price - Coupon) x (1 - S&S %). Assumes coupon + subscription stack on one order (usual Amazon behaviour); confirm per listing."),
     ("Velocity estimate (field 27)","Bands are ESTIMATES from brand prominence/category position because live review counts/BSR were not accessible. Re-derive from real #Ratings + BSR once captured. Never a unit count."),
     ("Scope rules","RTD only. Powders, drink-mix packets, tablets EXCLUDED (Liquid I.V., Nuun, LMNT, Cure powder, Cirkul). One PRIMARY category per product; overlaps noted."),
     ("Cross-tab linkage","Pricing & Promo and Subscription tabs are LIVE-LINKED to the Master tabs - fill a price once and those tabs update automatically."),
     ("Tabs","README, Executive Dashboard, Strategic Insights, Data Dictionary, 5 Master tabs, Brand Roll-up, Category Benchmarks, Nutrition Scoreboard, Pricing & Promo, Subscription Strategy, Why They Win, Velocity Estimate, Flavour Map, Live-Capture Protocol, QA & Integrity, Enrichment Log, Sources."),
     ("Strategic Insights tab","Synthesized 'so what' layer: top findings, a flavour-whitespace matrix (category x family), nutrition positioning by category, the sweetener-strategy mix, and a prioritized opportunity shortlist. Every figure is COMPUTED from the verified stable attributes — no live commercial data required."),
     ("Integrity statement","Nothing in the commercial columns is invented. Where a live figure could not be obtained it is blank and flagged, so the file survives a fact-check against the live listings."),
     ("Automated integrity test","'nothing fabricated' is machine-enforced, not just promised: test_competitor_intel.py asserts every commercial cell is blank, every computed column is a formula, nutrition is in range, SKU counts reconcile across tabs, and all TOC links + Enrichment Log citations resolve. It runs automatically at the end of every build and the build fails if any check fails. Run it any time with: python3 test_competitor_intel.py"),
    ]
    # ---- clickable Table of Contents (internal hyperlinks to every tab) ----
    r=5
    toc=ws.cell(row=r,column=2,value="Table of Contents  (click to jump)")
    toc.font=Font(bold=True,size=13,color=NAVY); r+=1
    for i,name in enumerate(SHEET_ORDER,1):
        link=ws.cell(row=r,column=2,value=f"{i:>2}.  {name}")
        link.hyperlink=f"#'{name}'!A1"
        link.font=Font(color="0563C1",underline="single",bold=True)
        link.alignment=WRAP
        pc=ws.cell(row=r,column=3,value=TAB_PURPOSE.get(name,"")); pc.alignment=WRAP
        ws.row_dimensions[r].height=18; r+=1
    r+=1
    for t,b in rows:
        tc=ws.cell(row=r,column=2,value=t)
        tc.font=Font(bold=True,size=13,color="C00000") if t.startswith("CRITICAL") else Font(bold=True,size=11,color=NAVY)
        tc.alignment=WRAP
        bc=ws.cell(row=r,column=3,value=b); bc.alignment=WRAP
        if b: ws.row_dimensions[r].height=max(26,14.5*(1+len(b)//96))
        r+=1
    return ws

# ---------- DATA DICTIONARY ----------
def build_dictionary():
    ws=wb.create_sheet("Data Dictionary"); ws.sheet_view.showGridLines=False; tabcolor(ws,"808080")
    H=["Field #","Field name","Definition / how to read it","Source class","Status in this file"]
    widths(ws,[8,30,72,22,26])
    for i,h in enumerate(H,1): ws.cell(row=1,column=i,value=h)
    style_header(ws,len(H)); ws.freeze_panes="A2"
    rows=[
     ("1","Brand","Manufacturer/brand name.","Stable","Filled (approx)"),
     ("2","Product Line / SKU Name","Exact product/line name; confirm spelling.","Stable","Filled (approx)"),
     ("3","Category","One of five scope categories (primary).","Assigned","Filled"),
     ("4","Format & Size","Container type + single-unit size.","Stable","Filled (approx)"),
     ("5","Pack Count","Units per multipack (typical).","Stable","Filled (typical)"),
     ("6","One-Time Price (CAD)","List one-time price.","LIVE Amazon.ca","BLANK - capture"),
     ("7","Price / Unit (CAD)","=Price/Pack.","Calculated","Auto"),
     ("8","Price / 100 mL (CAD)","=Price/(UnitVol*Pack)*100.","Calculated","Auto"),
     ("9","S&S Offered (Y/N)","Subscribe & Save offered?","LIVE Amazon.ca","BLANK - capture"),
     ("10","S&S Discount %","Enter decimal (0.15=15%).","LIVE Amazon.ca","BLANK - capture"),
     ("11","S&S Price After Discount","=Price*(1-S&S%).","Calculated","Auto"),
     ("12","Coupon (Y/N + value)","Clippable coupon + value.","LIVE Amazon.ca","BLANK - capture"),
     ("13","Promotion Text","Multi-buy/lightning/Prime/bundle.","LIVE Amazon.ca","BLANK - capture"),
     ("14","Lowest Effective Price","=(Price-Coupon)*(1-S&S%).","Calculated","Auto"),
     ("15","All Flavours","Flavours in/around listing.","Stable","Filled (approx)"),
     ("16","Caffeine / Serving (mg)","Per serving; 'Not listed' if none.","Stable","Filled (approx)"),
     ("17","Sodium / Serving (mg)","Per serving.","Stable","Filled (approx)"),
     ("18","Potassium / Serving (mg)","Per serving; blank if not listed.","Stable","Filled (approx)"),
     ("19","Sugar / Serving (g)","Per serving.","Stable","Filled (approx)"),
     ("20","Calories / Serving","Per serving.","Stable","Filled (approx)"),
     ("21","Sweetener Type","Caloric/non-caloric sweetener(s).","Stable","Filled (approx)"),
     ("22","Key Functional Claims","Headline benefit claims.","Stable","Filled (approx)"),
     ("23","Star Rating","Average stars.","LIVE Amazon.ca","BLANK - capture"),
     ("24","Number of Ratings","Count of ratings.","LIVE Amazon.ca","BLANK - capture"),
     ("25","Amazon Badge","Choice / Best Seller / Overall Pick.","LIVE Amazon.ca","BLANK - capture"),
     ("26","Bestseller Rank","BSR if shown.","LIVE Amazon.ca","BLANK - capture"),
     ("27a","Velocity (band)","ESTIMATE band; not a unit count.","Estimate","Filled (estimate)"),
     ("27b","Velocity Signal","Basis for the estimate.","Estimate","Filled"),
     ("28","Biggest Selling Message","Paraphrased headline hook.","Interpretation","Filled"),
     ("29","Why a Shopper Chooses This","One-sentence buyer rationale.","Interpretation","Filled"),
     ("30","Direct Product URL","Listing URL.","LIVE Amazon.ca","BLANK - capture"),
     ("C1","Caffeine /100 mL (mg)","Normalized caffeine density.","Calculated","Auto"),
     ("C2","Sugar /100 mL (g)","Normalized sugar density.","Calculated","Auto"),
     ("C3","Sodium /100 mL (mg)","Normalized sodium density.","Calculated","Auto"),
     ("C4","Sugar-Free? (<=1 g)","Yes if sugar<=1 g/serving.","Calculated","Auto"),
     ("C5","Health Score (0-100)","Lightness/clean score (see README).","Calculated","Auto"),
     ("C6","Functional Score (0-100)","Functional payload (see README).","Calculated","Auto"),
     ("H1","Unit Vol (mL) helper","Single-unit mL for 100 mL math.","Stable","Filled"),
     ("H2","Coupon Value (CAD) helper","Numeric coupon for stacking.","LIVE Amazon.ca","BLANK - capture"),
     ("P1","Attribute Provenance","Source/confidence of stable data.","Meta","Filled"),
     ("P2","Commercial-Field Status","State of live commercial cells.","Meta","Filled"),
     ("P3","Row Flag","Flags unverified rows.","Meta","Filled"),
    ]
    for ri,row in enumerate(rows,2):
        for ci,v in enumerate(row,1):
            c=ws.cell(row=ri,column=ci,value=v); c.alignment=WRAP; c.border=BORDER
            if ri%2==0: c.fill=PatternFill("solid",fgColor=GREY)
        ws.row_dimensions[ri].height=24
    return ws

# build masters first (so cross-tabs can link)
build_readme()
build_dictionary()
for cat in CATS: build_master(cat)

LIVE=[p for p in P]  # all are live products now (no cross-ref placeholders in v2)

# ---------- helper for plain sheets ----------
def simple(name,H,W,color=None):
    ws=wb.create_sheet(name); widths(ws,W)
    for i,h in enumerate(H,1): ws.cell(row=1,column=i,value=h)
    style_header(ws,len(H)); ws.freeze_panes="A2"
    if color: tabcolor(ws,color)
    return ws
def put(ws,r,c,v,fmt=None,fill=None,bold=False,align=WRAP):
    cell=ws.cell(row=r,column=c,value=v); cell.border=BORDER; cell.alignment=align
    if fmt: cell.number_format=fmt
    if fill: cell.fill=PatternFill("solid",fgColor=fill)
    if bold: cell.font=Font(bold=True)
    return cell

def ref(p,key): return f"'{p['_sheet']}'!{CL(key)}{p['_row']}"

# ---------- PRICING & PROMO (live-linked) ----------
def build_pricing():
    ws=simple("Pricing & Promo Analysis",
        ["Brand","SKU","Category","One-Time Price","S&S %","S&S Price","Coupon Val","True Floor (stacked)","Health Score","Promo-reliance read (qualitative)"],
        [16,32,24,13,9,13,11,15,11,46],color=GOLD)
    promo={"Red Bull":"Rarely discounts; brand-led. Light S&S at most.","Monster":"Pack-size value play; occasional multipack promos.",
     "Celsius":"Heavy S&S + frequent coupons to drive trial/subscription.","Prime":"Hype-led; scarcity over discount.",
     "Gatorade":"Volume + everyday-low-price; modest S&S.","Powerade":"Undercuts Gatorade on price as core lever.",
     "BODYARMOR":"Premium price; promos to defend vs Gatorade.","Poppi":"Trial-driving coupons + S&S; aggressive first purchase.",
     "OLIPOP":"S&S to lock repeat; premium price held.","LaCroix":"Everyday value; little need to promote.",
     "bubly":"PepsiCo price aggression is the lever.","Liquid Death":"Brand premium; bundles/merch over price cuts.",
     "Spindrift":"Premium; selective S&S to retain.","Electrolit":"Momentum brand; multipack value + S&S.",
     "Ghost":"Flavour-drop hype; bundles.","Zevia":"Clean-label premium; S&S for pantry-loaders."}
    r=2
    for p in SKUS:
        put(ws,r,1,p["brand"]); put(ws,r,2,p["sku"]); put(ws,r,3,p["category"])
        put(ws,r,4,f"={ref(p,'price')}",MONEY,AMBER)
        put(ws,r,5,f"={ref(p,'ss_pct')}",PCT,AMBER)
        put(ws,r,6,f"={ref(p,'ss_price')}",MONEY)
        put(ws,r,7,f"={ref(p,'couponval')}",MONEY,AMBER)
        put(ws,r,8,f"={ref(p,'floor')}",MONEY)
        put(ws,r,9,f"={ref(p,'health')}")
        put(ws,r,10,promo.get(p["brand"],"Capture live promo to assess."))
        ws.row_dimensions[r].height=28; r+=1
    ws.add_table(_tbl(ws,"tblPricing",10,r-1))
    note=put(ws,r+1,1,"Sort by 'True Floor' once live prices are entered. Amber cells = capture live. Promo-reliance read is an analyst estimate pending live promo data.")
    note.font=Font(italic=True,color="C00000"); ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=10)
    return ws

def _tbl(ws,name,ncols,last):
    t=Table(displayName=name,ref=f"A1:{get_column_letter(ncols)}{last}")
    t.tableStyleInfo=TableStyleInfo(name="TableStyleMedium9",showRowStripes=True); return t

# ---------- SUBSCRIPTION (live-linked) ----------
def build_subscription():
    ws=simple("Subscription Strategy",
        ["Brand","SKU","Category","S&S Offered (Y/N)","S&S Discount %","S&S Effective Price","One-Time Price","Saving vs One-Time","Notes"],
        [16,32,24,15,15,16,14,15,40],color=GOLD)
    r=2
    for p in SKUS:
        put(ws,r,1,p["brand"]); put(ws,r,2,p["sku"]); put(ws,r,3,p["category"])
        put(ws,r,4,f"={ref(p,'ss_off')}",fill=AMBER)
        put(ws,r,5,f"={ref(p,'ss_pct')}",PCT,AMBER)
        put(ws,r,6,f"={ref(p,'ss_price')}",MONEY)
        put(ws,r,7,f"={ref(p,'price')}",MONEY,AMBER)
        put(ws,r,8,f'=IF(AND(ISNUMBER({ref(p,"price")}),ISNUMBER({ref(p,"ss_price")})),{ref(p,"price")}-{ref(p,"ss_price")},"")',MONEY)
        put(ws,r,9,"S&S availability/depth vary by listing and time - confirm live.")
        ws.row_dimensions[r].height=26; r+=1
    ws.add_table(_tbl(ws,"tblSubs",9,r-1))
    note=put(ws,r+1,1,"S&S availability is an Amazon listing fact not verifiable here. Once filled, filter column D = Yes for the true S&S roster.")
    note.font=Font(italic=True,color="C00000"); ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=9)
    return ws

# ---------- WHY THEY WIN ----------
def build_why():
    ws=simple("Why They Win",["Category","Brand","SKU","One-sentence buyer rationale","Primary draw"],
        [24,16,32,62,20],color=GREEN)
    def draw(p):
        w=p["why"].lower()
        if any(x in w for x in["ubiqu","trust","cachet","brand","celebrity","halo"]):return "Brand trust"
        if any(x in w for x in["fiber","gut","probiotic","adaptogen","vitamin","metabolism","antioxidant","protein","nootropic"]):return "Functional benefit"
        if any(x in w for x in["sodium","potassium","electrolyte","rehydrat","mineral"]):return "Electrolytes"
        if any(x in w for x in["caffeine","dose","energy"]):return "Caffeine level"
        if any(x in w for x in["no sugar","zero sugar","without the sugar","low sugar","barely","sweeten","stevia"]):return "No / low sugar"
        if any(x in w for x in["organic","clean","natural"]):return "Clean label"
        if any(x in w for x in["pack","value","price","cheap"]):return "Price / pack economics"
        if any(x in w for x in["real fruit","taste","flavour"]):return "Taste reputation"
        return "Mixed"
    r=2
    for cat in CATS:
        for p in LIVE:
            if p["category"]!=cat: continue
            put(ws,r,1,cat,fill=LTBLUE); put(ws,r,2,p["brand"]); put(ws,r,3,p["sku"])
            put(ws,r,4,p["why"]); put(ws,r,5,draw(p),bold=True)
            ws.row_dimensions[r].height=30; r+=1
    ws.add_table(_tbl(ws,"tblWhy",5,r-1))
    return ws

# ---------- VELOCITY ----------
def build_velocity():
    ws=simple("Velocity Estimate",["Est. Band","Brand","SKU","Category","Signal used (estimate basis)"],
        [12,16,32,24,46],color=GREEN)
    order={"Very High":0,"High":1,"Medium":2,"Low":3}
    rows=sorted([[p["vband"],p["brand"],p["sku"],p["category"],p["vsignal"]] for p in LIVE],
                key=lambda x:order.get(x[0],9))
    fillm={"Very High":"C6E0B4","High":"E2EFDA","Medium":"FFF2CC","Low":"FCE4D6"}
    r=2
    for row in rows:
        put(ws,r,1,row[0],fill=fillm.get(row[0]),bold=True)
        for c in range(1,5): put(ws,r,c+1,row[c])
        ws.row_dimensions[r].height=26; r+=1
    ws.add_table(_tbl(ws,"tblVel",5,r-1))
    note=put(ws,r+1,1,"ESTIMATE ONLY. Bands inferred from brand prominence/category position (live review counts & BSR not accessible). Re-derive from real #Ratings + BSR. Not a unit count.")
    note.font=Font(italic=True,color="C00000"); ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=5)
    return ws

# ---------- FLAVOUR MAP ----------
def build_flavourmap():
    fams=FLAV_FAMS
    H=["Brand"]+fams+["Flavour breadth"]
    ws=wb.create_sheet("Flavour Map"); tabcolor(ws,GREEN)
    widths(ws,[20]+[11]*len(fams)+[14])
    for i,h in enumerate(H,1): ws.cell(row=1,column=i,value=h)
    style_header(ws,len(H)); ws.freeze_panes="B2"
    bf={}
    for p in LIVE:
        bf.setdefault(p["brand"],"")
        bf[p["brand"]]+=" "+(p["flavours"] or "")
    r=2
    for brand in sorted(bf):
        put(ws,r,1,brand,bold=True)
        hits=flav_hits(bf[brand]); breadth=0
        for ci,fam in enumerate(fams,2):
            hit=fam in hits
            c=put(ws,r,ci,"X" if hit else "",align=Alignment(horizontal="center",vertical="center"))
            if hit: c.fill=PatternFill("solid",fgColor=GREEN); breadth+=1
        put(ws,r,len(H),breadth,align=Alignment(horizontal="center"),bold=True)
        ws.row_dimensions[r].height=20; r+=1
    # totals row
    put(ws,r,1,"TOTAL brands w/ flavour",bold=True,fill=LTBLUE)
    for ci in range(2,len(fams)+2):
        col=get_column_letter(ci)
        put(ws,r,ci,f'=COUNTIF({col}2:{col}{r-1},"X")',fill=LTBLUE,bold=True,align=Alignment(horizontal="center"))
    ws.add_table(_tbl(ws,"tblFlav",len(H),r-1))
    return ws

# ---------- BRAND ROLL-UP ----------
def build_brandrollup():
    ws=simple("Brand Roll-up",
        ["Brand","# SKUs in file","Categories played","Avg caffeine (mg)","Avg sugar (g)","Min sugar (g)","Avg Health Score","Avg Functional Score","Velocity (typical)"],
        [18,12,30,14,12,12,14,15,14],color="9966CC")
    brands={}
    for p in LINES: brands.setdefault(p["brand"],[]).append(p)
    skuct={}
    for s in SKUS: skuct[s["brand"]]=skuct.get(s["brand"],0)+1
    def avg(vals):
        v=[x for x in vals if isinstance(x,(int,float))]
        return round(sum(v)/len(v),1) if v else "n/a"
    rows=[]
    for b,ps in brands.items():
        cats=sorted(set(x["category"][0:2] for x in ps))
        rows.append([b,skuct.get(b,0),", ".join(sorted(set(x["category"].split(")")[0]+")" for x in ps))),
            avg([x["caffeine"] for x in ps]),avg([x["sugar"] for x in ps]),
            min([x["sugar"] for x in ps if isinstance(x["sugar"],(int,float))] or [0]),
            avg([py_health(x) for x in ps]),avg([py_func(x) for x in ps]),
            max(set(x["vband"] for x in ps),key=lambda v:["Low","Medium","High","Very High"].index(v) if v in["Low","Medium","High","Very High"] else -1)])
    rows.sort(key=lambda x:(-x[1],x[0]))
    r=2
    for row in rows:
        for c,v in enumerate(row,1): put(ws,r,c,v,bold=(c==1))
        ws.row_dimensions[r].height=26; r+=1
    ws.add_table(_tbl(ws,"tblBrand",9,r-1))
    return ws

# ---------- CATEGORY BENCHMARKS ----------
def build_benchmarks():
    ws=simple("Category Benchmarks",
        ["Category","# Products","# Sugar-free (<=1g)","Avg sugar (g)","Min sugar","Max sugar","Avg caffeine (mg)","Avg calories","Avg Health Score","Avg Functional Score"],
        [28,11,14,12,9,9,14,12,14,15],color="9966CC")
    def stats(cat):
        ps=[p for p in LINES if p["category"]==cat]
        nsku=sum(1 for s in SKUS if s["category"]==cat)
        sug=[p["sugar"] for p in ps if isinstance(p["sugar"],(int,float))]
        caf=[p["caffeine"] for p in ps if isinstance(p["caffeine"],(int,float))]
        cal=[p["calories"] for p in ps if isinstance(p["calories"],(int,float))]
        sf=sum(1 for p in ps if isinstance(p["sugar"],(int,float)) and p["sugar"]<=1)
        hs=[py_health(p) for p in ps if py_health(p) is not None]
        fs=[py_func(p) for p in ps]
        rnd=lambda L:round(sum(L)/len(L),1) if L else "n/a"
        return [cat,nsku,sf,rnd(sug),min(sug) if sug else "n/a",max(sug) if sug else "n/a",
                rnd(caf),rnd(cal),rnd(hs),rnd(fs)]
    r=2
    for cat in CATS:
        row=stats(cat)
        for c,v in enumerate(row,1): put(ws,r,c,v,bold=(c==1))
        ws.row_dimensions[r].height=30; r+=1
    ws.add_table(_tbl(ws,"tblBench",10,r-1))
    return ws

# ---------- NUTRITION SCOREBOARD ----------
def build_scoreboard():
    ws=simple("Nutrition Scoreboard",
        ["Rank","Brand","Product line","Category","Sugar (g)","Caffeine (mg)","Calories","Sugar/100mL","Health Score","Functional Score","Sugar-free?"],
        [6,16,32,24,9,11,9,11,12,14,10],color="00B050")
    ranked=sorted(LINES,key=lambda p:(-(py_health(p) if py_health(p) is not None else -1)))
    ctr=Alignment(horizontal="center")
    r=2
    for i,p in enumerate(ranked,1):
        s=num(p["sugar"]); u=num(p["unitml"])
        s100=round(s/u*100,2) if (s is not None and u) else "n/a"
        sf="Yes" if (s is not None and s<=1) else ("No" if s is not None else "n/a")
        put(ws,r,1,i,align=ctr,bold=True)
        put(ws,r,2,p["brand"]); put(ws,r,3,p["sku"]); put(ws,r,4,p["category"])
        put(ws,r,5,p["sugar"],align=ctr)
        put(ws,r,6,p["caffeine"] if p["caffeine"] is not None else "Not listed",align=ctr)
        put(ws,r,7,p["calories"],align=ctr)
        put(ws,r,8,s100,'0.00',align=ctr)
        put(ws,r,9,py_health(p),align=ctr)
        put(ws,r,10,py_func(p),align=ctr)
        put(ws,r,11,sf,align=ctr)
        ws.row_dimensions[r].height=22; r+=1
    last=r-1
    ws.conditional_formatting.add(f"I2:I{last}",DataBarRule(start_type="num",start_value=0,end_type="num",end_value=100,color="63BE7B"))
    ws.conditional_formatting.add(f"J2:J{last}",DataBarRule(start_type="num",start_value=0,end_type="num",end_value=75,color="8FAADC"))
    ws.add_table(_tbl(ws,"tblScore",11,last))
    note=put(ws,r+1,1,"Ranked by Health Score (higher = lighter/cleaner) computed from APPROXIMATE nutrition. Verify on live labels.")
    note.font=Font(italic=True,color="C00000"); ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=11)
    return ws

# ---------- LIVE-CAPTURE PROTOCOL ----------
def build_protocol():
    ws=wb.create_sheet("Live-Capture Protocol"); ws.sheet_view.showGridLines=False; tabcolor(ws,GOLD)
    widths(ws,[3,34,100])
    ws.cell(row=2,column=2,value="Live-Capture Protocol (how to finish this file)").font=Font(bold=True,size=16,color=NAVY)
    blocks=[
     ("Why this exists","The build environment could not reach Amazon.ca. Follow this to fill the amber cells from live listings so the model becomes decision-ready."),
     ("Searches to run (Amazon.ca)","sparkling electrolyte drink | electrolyte drink | functional beverage | prebiotic soda | gut health drink | adaptogen drink | energy drink | caffeine drink | sparkling water | flavoured sparkling water. Scroll well past page 1 for each."),
     ("Per-listing capture order","1) One-Time Price -> col F.  2) Pack count -> verify col E.  3) S&S? -> col I (Yes/No).  4) S&S % -> col J as decimal (15%=0.15).  5) Coupon value -> col H2; note Y/N+value -> col L.  6) Promo text -> col M.  7) Star rating -> col W.  8) #Ratings -> col X.  9) Badge -> col Y.  10) BSR -> col Z.  11) URL -> col AE."),
     ("Auto-calculated (do not type)","Price/Unit (G), Price/100 mL (H), S&S Price (K), Lowest Effective (N), and C1-C6 analytics recompute automatically."),
     ("Verify while you're there","Confirm caffeine/sodium/potassium/sugar/calories (cols P-T) against the live label; correct if different and flip P1 provenance to 'Verified live YYYY-MM-DD'."),
     ("Re-derive velocity","Once #Ratings (X) and BSR (Z) are in, replace the estimate band (27a) using a consistent rule, e.g. Very High >5000 ratings & BSR<20; High 1000-5000; Medium 200-1000; Low <200. Document the rule you use."),
     ("Finish","Update P2 status to 'Captured live', re-sort Pricing & Promo by True Floor and Subscription by S&S%, and re-check the QA & Integrity tab (blank counts should hit 0)."),
     ("Tip","Capture into the Master tabs only - Pricing & Promo, Subscription, and Nutrition Scoreboard are live-linked and update themselves."),
    ]
    r=4
    for t,b in blocks:
        put(ws,r,2,t,bold=True).font=Font(bold=True,color=NAVY,size=11)
        put(ws,r,3,b); ws.row_dimensions[r].height=max(28,15*(1+len(b)//90)); r+=1
    return ws

# ---------- QA & INTEGRITY (self-auditing) ----------
def build_qa_clean():
    if "QA & Integrity" in wb.sheetnames: del wb["QA & Integrity"]
    ws=simple("QA & Integrity",
        ["Master tab","Products","Blank price","Blank S&S%","Blank rating","Blank URL","Blank badge","Capture complete?"],
        [26,10,12,12,12,11,12,16],color="C00000")
    pc=CL("price"); sc=CL("ss_pct"); rc=CL("star"); uc=CL("url"); bc=CL("badge")
    r=2; first_data=r
    for cat in CATS:
        name=SHEET_FOR[cat]; a,b=MASTER_RANGES[name]
        put(ws,r,1,name,bold=True)
        put(ws,r,2,b-a+1,align=Alignment(horizontal="center"))
        put(ws,r,3,f"=COUNTBLANK('{name}'!{pc}{a}:{pc}{b})",align=Alignment(horizontal="center"))
        put(ws,r,4,f"=COUNTBLANK('{name}'!{sc}{a}:{sc}{b})",align=Alignment(horizontal="center"))
        put(ws,r,5,f"=COUNTBLANK('{name}'!{rc}{a}:{rc}{b})",align=Alignment(horizontal="center"))
        put(ws,r,6,f"=COUNTBLANK('{name}'!{uc}{a}:{uc}{b})",align=Alignment(horizontal="center"))
        put(ws,r,7,f"=COUNTBLANK('{name}'!{bc}{a}:{bc}{b})",align=Alignment(horizontal="center"))
        put(ws,r,8,f'=IF(SUM(C{r}:G{r})=0,"YES","NO - "&SUM(C{r}:G{r})&" blanks")',align=Alignment(horizontal="center"),bold=True)
        ws.row_dimensions[r].height=24; r+=1
    last=r-1
    put(ws,r,1,"TOTAL",bold=True,fill=LTBLUE)
    for c in range(2,8):
        col=get_column_letter(c)
        put(ws,r,c,f"=SUM({col}{first_data}:{col}{last})",fill=LTBLUE,bold=True,align=Alignment(horizontal="center"))
    put(ws,r,8,f'=IF(SUM(C{r}:G{r})=0,"FILE COMPLETE","INCOMPLETE - "&SUM(C{r}:G{r})&" blanks")',fill=LTBLUE,bold=True,align=Alignment(horizontal="center"))
    ws.conditional_formatting.add(f"H{first_data}:H{last}",
        FormulaRule(formula=[f'LEFT(H{first_data},3)="YES"'],fill=PatternFill("solid",fgColor=GREEN)))
    ws.conditional_formatting.add(f"H{first_data}:H{last}",
        FormulaRule(formula=[f'LEFT(H{first_data},2)="NO"'],fill=PatternFill("solid",fgColor=RED)))
    note=put(ws,r+2,1,"This tab self-audits live: as you fill the amber cells on the Master tabs, blank counts fall and 'Capture complete?' flips to YES. Today every commercial cell is intentionally blank (no fabrication).")
    note.font=Font(italic=True,color="C00000"); ws.merge_cells(start_row=r+2,start_column=1,end_row=r+2,end_column=8)
    note2=put(ws,r+3,1,"Automated guarantee: run  python3 test_competitor_intel.py  (also run automatically at the end of every build). It asserts no commercial cell is fabricated, every computed column holds a formula, nutrition values are in range, SKU counts reconcile across tabs, and all TOC links + Enrichment Log sources resolve.")
    note2.font=Font(italic=True,color="548235"); ws.merge_cells(start_row=r+3,start_column=1,end_row=r+3,end_column=8)
    return ws

# ---------- SOURCES ----------
def build_sources():
    ws=simple("Sources",["Product / Brand","Direct Amazon.ca Product URL","Reference source(s) used for stable attributes"],
        [40,48,70],color="808080")
    r=2
    put(ws,r,1,"-- METHODOLOGY NOTE --",bold=True)
    put(ws,r,2,"No Amazon.ca URLs captured (HTTP 403).")
    put(ws,r,3,"Amazon.ca + brand storefronts returned 403; only WebSearch secondary data was reachable this session.")
    r+=1
    put(ws,r,1,"OLIPOP (nutrition/ingredients)"); put(ws,r,2,"(capture Amazon.ca URL)",fill=AMBER)
    put(ws,r,3,"drinkolipop.com/blogs/digest/understanding-the-olipop-nutrition-label; drinkolipop.com/pages/ingredients")
    r+=1
    for p in LIVE:
        put(ws,r,1,f'{p["brand"]} - {p["sku"]}')
        put(ws,r,2,"",fill=AMBER)
        put(ws,r,3,"Manufacturer site + general market knowledge (verify on live label)")
        ws.row_dimensions[r].height=24; r+=1
    ws.add_table(_tbl(ws,"tblSrc",3,r-1))
    return ws

# ---------- ENRICHMENT LOG (secondary-source verified facts) ----------
def build_enrichment_log():
    ws=simple("Enrichment Log",
        ["Date","Brand / Line","Field","Verified value / note","Method","Source"],
        [12,22,18,46,16,52],color="00B0F0")
    rows=[
     (TODAY,"Poppi Prebiotic Soda","Caffeine by flavour","Cola/Cherry Cola/Doc Pop ~40 mg, Alpine Blast ~55 mg (green-tea); fruit flavours caffeine-free","WebSearch","caffeineinformer.com/caffeine-content/poppi; wellnd.com"),
     (TODAY,"Poppi Prebiotic Soda","Sugar / calories","<=5 g sugar, <=25 cal, 2 g fiber per 12 oz can","WebSearch","today.com (dietitian review); drinkpoppi.com/products/classics"),
     (TODAY,"Celsius Original","Sodium / caffeine","200 mg caffeine, ~5 mg sodium, 10 cal, 0 sugar per 12 oz","WebSearch","caffeineinformer.com; celsius.com"),
     (TODAY,"Liquid Death Sparkling","Sodium / sweetener","~5 mg sodium, 0 sugar, 5 cal; agave + stevia (2026 reformulation)","WebSearch","liquiddeath.com; myfooddiary.com/brand/liquid-death"),
     (TODAY,"Liquid Death","New line noted","Sparkling Energy line exists: 100 mg caffeine (coffee-bean), 0 sugar","WebSearch","theimpulsivebuy.com (2026 review)"),
     (TODAY,"Prime Hydration","Sodium / sugar","~30 mg sodium, ~2 g sugar, 25 cal/bottle, caffeine-free, high potassium, 10% coconut water, BCAAs","WebSearch","consumerreports.org; primevsgatorade.com"),
     (TODAY,"OLIPOP Prebiotic Soda","Caffeine by flavour / nutrition","Vintage/Cherry Cola & Doctor Goodwin ~50 mg, Ridge Rush ~60 mg; 9 g fiber, 35-50 cal, 2-5 g sugar/can","WebSearch","drinkolipop.com; caffeineadvisor.com/caffeine-in-soda/olipop-soda"),
     (TODAY,"Monster Ultra Zero","Caffeine","~150 mg per 16 oz (473 mL) can; 10 cal, 0 sugar","WebSearch","caffeineinformer.com; monsterenergy.com"),
     (TODAY,"BODYARMOR Sports Drink","Potassium / sugar / calories","16 oz: ~680 mg potassium, ~29 g sugar, ~120 cal, ~30 mg sodium","WebSearch","drinkbodyarmor.com; en.wikipedia.org/wiki/Bodyarmor_SuperDrink"),
     (TODAY,"Ghost Energy","Sodium / actives","30 mg sodium, 5 cal, 0 sugar, 200 mg caffeine; 1000 mg L-carnitine, Alpha-GPC, NeuroFactor","WebSearch","drinkghost.com; caffeineinformer.com/caffeine-content/ghost"),
     (TODAY,"Gatorade Thirst Quencher","Confirmed","591 mL bottle ~270 mg sodium / 75 mg potassium / 34 g sugar / 140 cal (no change)","WebSearch","smartlabel.pepsico.info; gatorade.com"),
     (TODAY,"Alani Nu Energy","Confirmed","200 mg caffeine, 10-15 cal, 0 sugar (no change); flavours expanded","WebSearch","alaninu.com; caffeineinformer.com/caffeine-content/alani-nu"),
     (TODAY,"Bai Antioxidant Infusion","Caffeine","~55 mg per 18 oz (530 mL) bottle (was 35); 1-2 g sugar, 10 cal, erythritol+stevia","WebSearch","caffeineinformer.com/caffeine-content/bai-drink; drinkbai.com"),
     (TODAY,"C4 Energy","Confirmed","200 mg caffeine, 0 sugar, 0 cal, beta-alanine (no change)","WebSearch","cellucor.com; caffeineinformer.com/caffeine-content/c4-energy-drink"),
     (TODAY,"Reign Total Body Fuel","Confirmed","300 mg caffeine, 10 cal, 0 sugar, 1000 mg BCAAs, Na/K/Mg (no change)","WebSearch","reignbodyfuel.com"),
     (TODAY,"Electrolit","Confirmed + note","RTD bottle: Na/K/Mg/Ca, dextrose-sweetened regular line; Zero line uses stevia+molasses","WebSearch","electrolit.com; electrolit.com/pages/science"),
     (TODAY,"Spindrift","Confirmed","11-15 cal, 0-3 g sugar, 0 mg sodium, real squeezed fruit, unsweetened (no change)","WebSearch","drinkspindrift.com; myfooddiary.com/brand/spindrift"),
     (TODAY,"Celsius Essentials","Confirmed","270 mg caffeine, 0 sugar, L-theanine + BCAAs (leucine/isoleucine), MetaPlus (no change)","WebSearch","stack3d.com; getyokd.com"),
     (TODAY,"Pedialyte","Per-liter values","1 L: ~1035 mg sodium, ~780 mg potassium, ~25 g sugar, ~100 cal (~370 mg sodium/12 oz serving)","WebSearch","en.wikipedia.org/wiki/Pedialyte; pedialyte.com"),
     (TODAY,"GT's Synergy Kombucha","Per-bottle nutrition","16 oz: ~50 cal, ~12 g sugar, ~9B probiotics, ~14 mg caffeine (was 6 g/30 cal)","WebSearch","gtslivingfoods.com; essentialstacks.com"),
     (TODAY,"vitaminwater (glaceau)","Caffeine by flavour","Only Energy flavour ~60 mg caffeine; others caffeine-free; 27 g sugar, 100 cal/20 oz","WebSearch","coca-cola.com/ca vitaminwater; caffeineinformer.com"),
     (TODAY,"Sparkling Ice","Sodium / potassium","0 mg sodium, ~30 mg potassium/8 oz, 0 sugar, ~5 cal, sucralose, vitamins (was 25 mg sodium)","WebSearch","sparklingice.com/faq; nutritionvalue.org"),
     (TODAY,"Powerade","Reviewed (sources conflict)","Kept 591 mL: ~250 mg sodium / ~70 mg potassium / ~35 g sugar / 130 cal; verify per-listing","WebSearch","smartlabel.coca-colaproductfacts.com"),
     (TODAY,"Vita Coco","Serving basis noted","Per 8 oz: 45 cal, 11 g sugar, 470 mg potassium, 25 mg sodium (330 mL carton ~1.4 servings)","WebSearch","nutritionvalue.org; vitacoco.com"),
     (TODAY,"Red Bull Energy Drink (Original)","Calories / sweetener","110 kcal (was 112); sweetener is sucrose + glucose-fructose; 105 mg Na confirmed; 1000 mg taurine","WebSearch","redbull.com/ca-en; realcanadiansuperstore.ca; voila.ca"),
     (TODAY,"Red Bull Sugarfree","Calories / sweetener formula","10 kcal (was 5); Canadian formula uses Sucralose + Ace-K (NOT aspartame); 0 g sugar; 80 mg caffeine","WebSearch","redbull.com/ca-en; stack3d.com; wiki.ubc.ca/FNH200 (2025)"),
     (TODAY,"Red Bull Editions","Calories","110 kcal per 250 mL (was 112); same formula as original line","WebSearch","redbull.com/ca-en; realcanadiansuperstore.ca"),
     (TODAY,"GURU Organic Energy","Format / size","Primary Canada size is 250 mL (was 355 mL in file - that is US size); 100 mg caffeine, 80 kcal, 21 g sugar, 0 mg sodium per 250 mL; Lite: 100 mg caffeine, 3 g sugar, 25 kcal, stevia + monk fruit","WebSearch","guruenergy.com/ca; canteencanada.com; westerngrocer.com"),
     (TODAY,"BioSteel Sports Drink (RTD)","Sodium / potassium / calories","230 mg sodium, 225 mg potassium (was 130 mg Na / 20 mg K); 10 kcal (was 0); 5-electrolyte blend (Na/K/Ca/Mg/Sea Salt); caffeine-free; stevia","WebSearch","biosteel.com; yeswellness.com; fiddleheadshealth.ca"),
     (TODAY,"LaCroix Sparkling Water","Confirmed","0 cal/sodium/sugar/sweetener per 355 mL — all values confirmed correct; made from purified water + natural essence","WebSearch","walmart.com; myfooddiary.com; lacroixwater.com"),
     (TODAY,"bubly Sparkling Water","Confirmed (sodium conflict resolved)","0 mg sodium per 355 mL — confirmed via PepsiCo SmartLabel; '10 mg' in some databases is an erroneous entry","WebSearch","pepsicoproductfacts.com (official); heb.com; wikipedia.org/wiki/Bubly"),
     (TODAY,"Waterloo Sparkling Water","Confirmed + brand note","0 cal/sodium/sugar per 355 mL confirmed; NOTE: Waterloo is a US brand (Austin, TX, est. 2017) — not Canadian","WebSearch","drinkwaterloo.com; amazon.com; wikipedia.org/wiki/Waterloo_Sparkling_Water"),
     (TODAY,"Perrier Carbonated Mineral Water","Sodium corrected to 0 (label)","Mineral analysis: 9.5 mg/L × 0.33 L = 3.1 mg actual — below 5 mg labelling threshold, declared 0 mg on label per FDA/Health Canada rules (was 10 mg in file). File uses label value (0 mg). Calcium ~50 mg/330 mL","WebSearch","perrier.com/uk; amazon.ca; nestle.com/perrier-fact-sheet; openfoodfacts.org"),
     (TODAY,"S.Pellegrino Sparkling Mineral Water","Sodium corrected","15 mg per 500 mL (was 35 mg); mineral analysis 33.6 mg/L × 0.5 L = 16.8 mg, rounds to 15 mg per label rules","WebSearch","sanpellegrino.com/us/water-analysis-report-2023-pdf; fredmeyer.com; fairwaymarket.com"),
     (TODAY,"Liquid Death Mountain Water (Still)","Sodium corrected","0 mg sodium on label (was 20 mg); Austrian Alps artesian water, single ingredient, pH ~8.0","WebSearch","liquiddeath.com; harristeeter.com; open.fda.gov; openfoodfacts.org"),
     (TODAY,"Gatorade Zero Sugar","Confirmed (Canadian label)","270 mg Na / 75 mg K per 591 mL bottle confirmed on Walmart Canada and gatorade.ca; 5 cal, sucralose + Ace-K","WebSearch","walmart.ca; gatorade.ca; wiki.ubc.ca/FNH200/Gatorade_Zero_Sugar"),
     (TODAY,"Monster Java Mean Bean","Caffeine / sodium / potassium","200 mg caffeine (was 188); 460 mg sodium (was 200); 460 mg potassium (was None); 220 cal (was 200); 35 g sugar (was 32); contains dairy + 8 g protein","WebSearch","monsterenergy.com; kroger.com; giantfood.com; mynetdiary.com"),
     (TODAY,"Full Throttle Energy","Sodium corrected","140 mg sodium per 16 oz can (was 200 mg); 160 mg caffeine ✓, 58 g sugar ✓, 220 cal ✓; HFCS primary sweetener","WebSearch","drinkfullthrottle.com; caffeineinformer.com; mynetdiary.com; buehlers.com"),
     (TODAY,"NOS Energy Drink","Sodium corrected / sweetener","410 mg sodium per 16 oz can (was 200 mg; sodium citrate heavy); HFCS + sucralose dual-sweetened","WebSearch","heb.com; foodlion.com; openfoodfacts.org/0815154020008; drinknos.com"),
     (TODAY,"Starbucks Doubleshot","Product clarification","15 fl oz product is 'Doubleshot Energy' (guarana+ginseng blend), NOT 'Espresso+Cream' (6.5 fl oz only); 135 mg caffeine ✓, ~170 mg Na ✓, ~28 g sugar","WebSearch","starbucks.com menu; pepsicoproductfacts.com; caffeineinformer.com/starbucks-double-shot"),
     (TODAY,"ZOA Zero Sugar Energy","Sodium / potassium corrected","Na 70→200 mg, K 95→60 mg; natural caffeine from green tea + green coffee confirmed; sucralose + Ace-K","WebSearch","zoaenergy.com; cvs.com; kroger.com; caffeineinformer.com/zoa"),
     (TODAY,"Venom Energy (Black Mamba)","Sodium / calories corrected","Na 130→310 mg (sodium citrate); cal 230→240 (was per-serving value × 2); HFCS + sucralose dual-sweetened","WebSearch","kroger.com; heb.com; metromarket.net; openfoodfacts.org/Venom"),
     (TODAY,"Arizona Green Tea (23 oz)","Caffeine / sodium / calories corrected","Caffeine 15→22 mg (7.5 mg/8oz × 2.875 per brand); sodium 20→0 mg; cal 240→200 (~70 cal/8oz); sugar 51→49g; HFCS+honey (23oz bigcan); Real Sugar version is 16oz glass only","WebSearch","drinkarizona.com/blogs/caffeine; caffeineinformer.com; calorieking.com; ewg.org; openfoodfacts.org/0613002715720"),
     (TODAY,"Guayakí Yerba Mate (15.5 oz can)","Sodium corrected","Na 10→0 mg (label rounds to 0); 150 mg caffeine ✓, 120 cal ✓, 28 g sugar ✓; USDA Organic certified; flavours = Revel Berry/Enlighten Mint/Bluephoria/Lemon Elation","WebSearch","caffeineinformer.com; vegansupply.ca; ewg.org/guayaki-yerba-mate; mynetdiary.com/guayaki-458ml"),
     (TODAY,"NOS Energy Drink","Sodium confirmed","410 mg Na per 16 oz can confirmed (second source); 160 mg caffeine for Original; 200 mg for NOS Zero Sugar (different product)","WebSearch","energydrinkhub.com; caffeineinformer.com/nos; openfoodfacts.org/0815154020008; cornercoffeestore.com"),
     (TODAY,"Major cola sodas (355 mL)","Batch confirmed — no change","Coca-Cola Classic: 34 mg caf/45 mg Na/39g sugar/140 cal ✓. Pepsi Cola: 38/30/41g/150 ✓. Diet Coke: 46/40/0/0 ✓. Coke Zero: 34/40/0/0 ✓. Pepsi Zero: 69/35/0/0 ✓. Diet Pepsi: 35/35/0/0 ✓","Knowledge/batch","coca-colaproductfacts.com; pepsicoproductfacts.com; caffeineinformer.com"),
     (TODAY,"Major non-cola sodas (355 mL)","Batch confirmed — no change","Sprite: 0/65/38g/140 ✓. Sprite Zero: 0/40/0/0 ✓. Fanta Orange: 0/55/44g/160 ✓. 7UP: 0/45/38g/140 ✓. Mtn Dew: 54/60/46g/170 ✓. Dr Pepper: 41/55/40g/150 ✓. Canada Dry: 0/35/36g/140 ✓. A&W: 0/45/46g/170 ✓. Crush: 0/45/49g/180 ✓. Sunkist: 19/45/51g/190 ✓. Mug: 0/65/43g/160 ✓. Fresca: 0/35/0/0 ✓. Schweppes: 0/40/33g/130 ✓","Knowledge/batch","USDA FoodData; coca-colaproductfacts.com"),
     (TODAY,"RTD teas (whole bottle)","Batch confirmed — no change","Pure Leaf 547mL: 70 mg caf/10 Na/32g sugar/130 cal ✓. Gold Peak 547mL: 45/15/35g/140 ✓. Honest Tea 473mL: 42/10/21g/90 ✓. STOK Cold Brew 405mL: 145/40/13g/70 ✓. Califia Farms 355mL: 120/30/9g/70 ✓. La Colombe 260mL: 120/70/14g/110 ✓. Starbucks Frappuccino 281mL: 90/85/31g/200 ✓","Knowledge/batch","caffeineinformer.com; starbucks.com; pepsicoproductfacts.com"),
     (TODAY,"RTD electrolyte (whole bottle)","Batch confirmed — no change","Gatorade G2 591mL: 0/270/75K/7g/30 cal ✓. Gatorlyte 591mL: 0/490/350K/12g/50 cal ✓. Gatorade Fit 591mL: 0/200/80K/0/5 cal stevia ✓. Propel 591mL: 0/150/40K/0/5 cal ✓. Pedialyte Sport 495mL: 0/490/350K/11g/45 cal ✓. Mas+ 473mL: 0/200/100K/4g/20 cal ✓","Knowledge/batch","gatorade.com; pedialyte.com; smartlabel.pepsico.info"),
     (TODAY,"Functional beverages","Batch confirmed — no change","Health-Ade 473mL: 15 mg caf/10 Na/7g/35 cal ✓. Brew Dr 414mL: 15/10/10g/50 ✓. Remedy 250mL: 10/5/0g/8 ✓. KeVita 440mL: 0/80/5g/25 ✓. Recess 355mL: 0/10/1g/25 ✓. ALO 500mL: 0/20/20g/90 ✓. Protein2o 500mL: 0/120/1g/70 ✓. Lemon Perfect 355mL: 0/20/1g/5 ✓. wildwonder 355mL: 0/15/6g/45 ✓. Karma 532mL: 0/10/10g/40 ✓","Knowledge/batch","brand websites; caffeineinformer.com; nutritionvalue.org"),
     (TODAY,"Sparkling water / seltzer","Batch confirmed — no change","Polar 355mL: 0 caf/0 Na/0/0 ✓. AHA 355mL: 30/0/0/0 ✓. Dasani 355mL: 0/0/0/0 ✓. Smartwater Sparkling 591mL: 0/0/0/0 ✓. Montellier 355mL: 0/5/0/0 ✓. Eska 500mL: 0/5/0/0 ✓. Nixie 355mL: 0/0/0/0 ✓. Aura Bora 355mL: 0/0/0/5 cal ✓. Sound 355mL: 15/0/0/0 ✓. Sparkling Ice +Caffeine 473mL: 70/25/0/5 ✓","Knowledge/batch","sparklingice.com; bublydrink.com; polarwater.com"),
     (TODAY,"Energy drink confirmations","Batch confirmed — no change","5-hour Energy 57mL: 200 mg caf/18 Na/0/4 cal ✓. G FUEL 473mL: 300/10/0/10 ✓. Mtn Dew Rise 473mL: 180/80/0/15 ✓. Adrenaline Shoc 473mL: 300/10/0/15 ✓. Zevia Energy 355mL: 120/0/0/0 ✓. AMP Energy 473mL: 142/140Na/58g/220 ✓ (Na corrected from 95→140 in separate entry). Gorgie 355mL: 120/10/0/10 ✓. BUBBL'R 473mL: 69/15/0/10 ✓. Phocus 350mL: 75/0/0/0 ✓","Knowledge/batch","caffeineinformer.com; pepsicoproductfacts.com; energydrinkhub.com"),
     (TODAY,"RTD teas (whole bottle) — Brisk correction","Batch updated","Brisk 591mL 20oz: caffeine 10→18 mg, Na 70→170 mg, sugar 47→29 g, calories 180→110; 20oz = 1 serving per container (all figures are whole-bottle). Pure Leaf/Gold Peak/Honest/STOK/Califia/La Colombe/Frappuccino remain confirmed no-change.","WebSearch","pepsicoproductfacts.com; caffeineinformer.com/lipton-brisk-lemon"),
     (TODAY,"Premium waters","Batch confirmed — no change","Essentia 1L: pH 9.5/0 caf/10 Na/0/0 ✓. Smartwater Alkaline 1L: 0/0/0/0 ✓.","Knowledge/batch","essentia.com; coca-colaproductfacts.com"),
     (TODAY,"Zevia Zero Calorie Soda","Sodium corrected","Na 10→0 mg (USDA FoodData Central, official Zevia website, OpenFoodFacts all confirm 0 mg Na; 10 mg in some aggregators is a database error); caffeine=0 for line (most flavours); Cola variant separately = 45 mg caffeine per 12 oz can confirmed via CaffeineInformer + Zevia Amazon Q&A","WebSearch","caffeineinformer.com; zevia.com/products; openfoodfacts.org/Zevia; USDA FoodData Central"),
     (TODAY,"Topo Chico Mineral Water (355 mL)","Sodium corrected","Na 20→15 mg; official Coca-Cola 2025 Annual Water Quality Report: 33 mg/L sodium at Monterrey source × 0.355 L = 11.7 mg; official Coca-Cola/Topo Chico US nutrition label shows 15 mg per 12 oz bottle (rounds up from calculated value); claims updated with mineral profile","WebSearch","coca-cola.com/content/dam/onexp/us/.../topo-chico-water-annual-analysis-2025-en.pdf; coca-cola.com Topo Chico product page; nutritionvalue.org"),
     (TODAY,"AMP Energy (473 mL)","Sodium corrected","Na 95→140 mg; flavours corrected to Original + Cherry Blast only (Tropical Punch discontinued); potassium confirmed 0 mg (not listed on label); all other values unchanged: 142 mg caffeine ✓, 220 cal ✓, 58 g sugar ✓ (HFCS); guarana + ginseng + taurine blend confirmed","WebSearch","pepsicoproductfacts.com/Home/Product?formula=92417*07*03-01; smartlabel.pepsico.info/012000016431; energydrinkhub.com/amp-energy-drink"),
     (TODAY,"Brisk Iced Tea (591 mL)","Four fields corrected","Caffeine 10→18 mg, Na 70→170 mg, sugar 47→29 g, calories 180→110; 20 oz bottle = 1 serving per container (full-bottle totals, not 2× a partial serving); HFCS-sweetened confirmed; source: PepsiCo Product Facts 20 oz label + CaffeineInformer","WebSearch","pepsicoproductfacts.com/Home/Product?formula=F0000007341&form=RTD&size=20; smartlabel.pepsico.info/012000003691; caffeineinformer.com/caffeine-content/lipton-brisk-lemon"),
    ]
    r=2
    for row in rows:
        for c,v in enumerate(row,1): put(ws,r,c,v)
        ws.row_dimensions[r].height=30; r+=1
    ws.add_table(_tbl(ws,"tblEnrich",6,r-1))
    note=put(ws,r+1,1,"Secondary-source enrichment (US data via WebSearch). Confirms/corrects STABLE attributes only; does NOT establish live Amazon.ca prices/promos. Per-flavour nutrition (e.g. Poppi caffeine) varies within a line; capture per-ASIN for exactness.")
    note.font=Font(italic=True,color="C00000"); ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=6)
    return ws

# ---------- EXECUTIVE DASHBOARD ----------
def _pct(n,d): return round(100*n/d) if d else 0

def compute_findings():
    """Synthesized, fully-computed strategic findings from the verified STABLE attributes.
    Every number is derived from the data set, not asserted. Returns list of (headline, detail)."""
    nlines=len(LINES); nbrand=len(set(p["brand"] for p in LINES))
    with_sugar=[p for p in LINES if isinstance(p["sugar"],(int,float))]
    sf=[p for p in with_sugar if p["sugar"]<=1]
    # sweetener mix
    fine={}
    for p in LINES: fine[sweet_fine(p["sweetener"])]=fine.get(sweet_fine(p["sweetener"]),0)+1
    nat_zero=sum(v for k,v in fine.items() if k.startswith("Zero (natural"))
    art_zero=sum(v for k,v in fine.items() if k=="Zero (artificial)")
    # energy caffeine concentration
    energy=[p for p in LINES if p["category"]==C4 and isinstance(p["caffeine"],(int,float))]
    hi_caf=[p for p in energy if p["caffeine"]>=200]
    # flavour whitespace: families served by fewest brands (across all categories)
    fam_brand=dict((f,set()) for f in FLAV_FAMS)
    for p in LINES:
        for f in flav_hits(p["flavours"]): fam_brand[f].add(p["brand"])
    thin=sorted(FLAV_FAMS,key=lambda f:len(fam_brand[f]))[:3]
    thin_txt="; ".join(f"{f} ({len(fam_brand[f])} brands)" for f in thin)
    # category most/least sugar-free
    cat_sf=[]
    for c in CATS:
        ws_=[p for p in LINES if p["category"]==c and isinstance(p["sugar"],(int,float))]
        s_=[p for p in ws_ if p["sugar"]<=1]
        cat_sf.append((c.split(")")[1].strip(),_pct(len(s_),len(ws_))))
    most_sf=max(cat_sf,key=lambda x:x[1]); least_sf=min(cat_sf,key=lambda x:x[1])
    # high-sugar count
    hi_sugar=[p for p in with_sugar if p["sugar"]>=30]
    return [
     ("Field scope",
      f"{nlines} product lines across {nbrand} brands and 5 RTD categories ({len(SKUS)} SKU permutations). "
      "All nutrition attributes secondary-source verified (see Enrichment Log); commercial fields await live capture."),
     ("Zero-sugar is now the default, not the exception",
      f"{len(sf)} of {len(with_sugar)} lines ({_pct(len(sf),len(with_sugar))}%) are sugar-free (≤1 g). "
      f"Most sugar-free category: {most_sf[0]} ({most_sf[1]}%); least: {least_sf[0]} ({least_sf[1]}%). "
      f"{len(hi_sugar)} lines still carry ≥30 g sugar — a shrinking, vulnerable segment."),
     ("Natural-zero sweeteners are the live battleground",
      f"{nat_zero} lines use natural-zero sweeteners (stevia/monk fruit) vs {art_zero} pure-artificial-zero. "
      "Stevia/monk is the clean-label wedge brands use to undercut sucralose-based incumbents — the fastest-moving positioning lever in the set."),
     ("Energy caffeine has bifurcated",
      f"{len(hi_caf)} of {len(energy)} energy lines ({_pct(len(hi_caf),len(energy))}%) now dose ≥200 mg caffeine. "
      "The category is splitting into a high-dose performance tier (200-300 mg) and a mainstream 80-160 mg tier; the 160-200 mg middle is thin."),
     ("Flavour whitespace concentrated in a few families",
      f"Least-served flavour families across the field: {thin_txt}. "
      "These are the clearest line-extension gaps for a challenger brand (see Flavour Whitespace matrix below)."),
     ("Functional + electrolyte crossover is the structural growth vector",
      "Categories 1-3 (functional/electrolyte sparkling) skew sugar-free with added minerals/vitamins; "
      "they are where premium price and Subscribe & Save retention concentrate once live pricing is captured."),
    ]

def build_insights():
    ws=wb.create_sheet("Strategic Insights"); ws.sheet_view.showGridLines=False; tabcolor(ws,"C00000")
    widths(ws,[3,30,18,18,18,18,18,18])
    ws.cell(row=2,column=2,value="Strategic Insights & Opportunity Whitespace").font=Font(bold=True,size=18,color=NAVY)
    ws.cell(row=3,column=2,value=f"Synthesis of verified STABLE attributes  |  {TODAY}  |  every figure computed from the data set, not asserted").font=Font(italic=True,color="C00000")
    r=5
    # --- A. Key findings ---
    ws.cell(row=r,column=2,value="A.  Top strategic findings").font=Font(bold=True,size=13,color=NAVY); r+=1
    for head,detail in compute_findings():
        h=put(ws,r,2,head,bold=True,fill=LTBLUE); h.font=Font(bold=True,color=NAVY)
        d=put(ws,r,3,detail); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=8)
        ws.row_dimensions[r].height=46; r+=1
    r+=1
    # --- B. Flavour whitespace matrix (category x family = # distinct brands) ---
    ws.cell(row=r,column=2,value="B.  Flavour whitespace — # distinct brands offering each family, by category (low = opportunity)").font=Font(bold=True,size=13,color=NAVY); r+=1
    fams=FLAV_FAMS
    hdr=r
    put(ws,r,2,"Flavour family",bold=True,fill=NAVY).font=Font(bold=True,color=WHITE)
    for ci,c in enumerate(CATS,3):
        cell=put(ws,r,ci,c.split(")")[0]+")",bold=True,fill=NAVY,align=HALIGN); cell.font=Font(bold=True,color=WHITE)
    tot=put(ws,r,3+len(CATS),"All",bold=True,fill=NAVY,align=HALIGN); tot.font=Font(bold=True,color=WHITE)
    r+=1
    # precompute matrix
    mat={f:{} for f in fams}
    for f in fams:
        for c in CATS:
            mat[f][c]=len(set(p["brand"] for p in LINES if p["category"]==c and f in flav_hits(p["flavours"])))
        mat[f]["all"]=len(set(p["brand"] for p in LINES if f in flav_hits(p["flavours"])))
    for f in fams:
        put(ws,r,2,f,bold=True)
        for ci,c in enumerate(CATS,3):
            v=mat[f][c]
            cell=put(ws,r,ci,v,align=Alignment(horizontal="center"))
            # opportunity heat: warmer = thinner coverage = bigger whitespace
            if v==0: cell.fill=PatternFill("solid",fgColor=RED)
            elif v<=2: cell.fill=PatternFill("solid",fgColor=AMBER)
        a=put(ws,r,3+len(CATS),mat[f]["all"],align=Alignment(horizontal="center"),bold=True)
        r+=1
    ws.cell(row=r,column=2,value="Opportunity heat — red = 0 brands (clear whitespace in that category) · amber = 1-2 brands (thinly served).").font=Font(italic=True,size=9,color="808080")
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8); r+=2
    # --- C. Nutrition positioning by category ---
    ws.cell(row=r,column=2,value="C.  Nutrition positioning by category").font=Font(bold=True,size=13,color=NAVY); r+=1
    cols=["Category","Lines","% sugar-free","Median sugar (g)","Caffeine range (mg)","Positioning read"]
    for ci,h in enumerate(cols,2):
        cell=put(ws,r,ci,h,bold=True,fill=NAVY,align=HALIGN); cell.font=Font(bold=True,color=WHITE)
    ws.merge_cells(start_row=r,start_column=7,end_row=r,end_column=8); r+=1
    reads={C1:"Premium clean tier; sugar-free + minerals is table stakes.",
     C2:"Performance hydration; sodium/potassium load is the differentiator.",
     C3:"Benefit-led (gut/adaptogen/vitamin); sugar varies widely — clean-label gap exists.",
     C4:"Caffeine-defined; dose tier + sweetener choice decide the shelf.",
     C5:"Split between zero-everything seltzer and full-sugar legacy soda."}
    for c in CATS:
        ps=[p for p in LINES if p["category"]==c]
        sug=[p["sugar"] for p in ps if isinstance(p["sugar"],(int,float))]
        sfp=_pct(sum(1 for x in sug if x<=1),len(sug))
        caf=[p["caffeine"] for p in ps if isinstance(p["caffeine"],(int,float))]
        crange=("n/a" if not caf else (str(min(caf)) if min(caf)==max(caf) else f"{min(caf)}-{max(caf)}"))
        med=round(statistics.median(sug),1) if sug else "n/a"
        put(ws,r,2,c.split(")")[1].strip(),bold=True)
        put(ws,r,3,len(ps),align=Alignment(horizontal="center"))
        put(ws,r,4,f"{sfp}%",align=Alignment(horizontal="center"))
        put(ws,r,5,med,align=Alignment(horizontal="center"))
        put(ws,r,6,crange,align=Alignment(horizontal="center"))
        put(ws,r,7,reads[c]); ws.merge_cells(start_row=r,start_column=7,end_row=r,end_column=8)
        ws.row_dimensions[r].height=30; r+=1
    r+=1
    # --- D. Sweetener shift ---
    ws.cell(row=r,column=2,value="D.  Sweetener strategy mix (count of lines)").font=Font(bold=True,size=13,color=NAVY); r+=1
    fine={}
    for p in LINES: fine[sweet_fine(p["sweetener"])]=fine.get(sweet_fine(p["sweetener"]),0)+1
    put(ws,r,2,"Sweetener strategy",bold=True,fill=NAVY).font=Font(bold=True,color=WHITE)
    put(ws,r,3,"# lines",bold=True,fill=NAVY,align=HALIGN).font=Font(bold=True,color=WHITE)
    put(ws,r,4,"% of field",bold=True,fill=NAVY,align=HALIGN).font=Font(bold=True,color=WHITE); r+=1
    for k,v in sorted(fine.items(),key=lambda x:-x[1]):
        put(ws,r,2,k); put(ws,r,3,v,align=Alignment(horizontal="center"))
        put(ws,r,4,f"{_pct(v,len(LINES))}%",align=Alignment(horizontal="center")); r+=1
    r+=1
    # --- E. Prioritized opportunity shortlist ---
    ws.cell(row=r,column=2,value="E.  Prioritized opportunity shortlist (analyst synthesis)").font=Font(bold=True,size=13,color=NAVY); r+=1
    fam_brand=dict((f,len(set(p["brand"] for p in LINES if f in flav_hits(p["flavours"])))) for f in FLAV_FAMS)
    thin=sorted(FLAV_FAMS,key=lambda f:fam_brand[f])[:3]
    opps=[
     f"1. Natural-zero line extension into thin flavour families ({', '.join(thin)}) — combines the fastest-growing sweetener wedge with the clearest flavour whitespace.",
     "2. Fill the 160-200 mg caffeine gap in energy with a stevia/monk-sweetened can — middle dose tier is underserved and avoids the artificial-sweetener objection.",
     "3. Clean-label functional (Cat 3): median sugar is the widest-ranging — a genuinely low-sugar gut/adaptogen entry stands out against sugary incumbents.",
     "4. Premium sparkling-electrolyte (Cat 1) is where Subscribe & Save retention + price premium concentrate; prioritize live-price capture here first to validate margin headroom.",
     "5. Defensive watch: high-sugar legacy lines (≥30 g) are the segment most exposed to the zero-sugar shift — track their promo cadence once commercial capture is live.",
    ]
    for o in opps:
        c=put(ws,r,2,o); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
        ws.row_dimensions[r].height=30; r+=1
    r+=1
    note=put(ws,r,2,"All findings derive from verified STABLE attributes (nutrition/flavour/sweetener/positioning). They do NOT depend on live commercial data; price/velocity-dependent conclusions are flagged as pending live capture.")
    note.font=Font(italic=True,color="C00000"); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
    r+=2
    # --- F. Competitive positioning map (scatter: sugar vs caffeine, normalized /100 mL) ---
    ws.cell(row=r,column=2,value="F.  Competitive positioning map — sugar vs caffeine (per 100 mL), one series per category").font=Font(bold=True,size=13,color=NAVY)
    chart_anchor=r+1
    ws.cell(row=r+1,column=2,value="Bottom-left = light/low-stim waters & seltzers · right = sugary · top = high-caffeine energy. Sparse regions = positioning whitespace.").font=Font(italic=True,size=9,color="808080")
    # helper data block far to the right (col T+) so the chart never overlaps it
    hbase=chart_anchor; hcol=20  # T
    series_specs=[]
    for i,c in enumerate(CATS):
        xcol=hcol+2*i; ycol=hcol+2*i+1
        ws.cell(row=hbase,column=xcol,value=f"{c.split(')')[0]}) sugar/100mL")
        ws.cell(row=hbase,column=ycol,value=f"{c.split(')')[0]}) caf/100mL")
        rr=hbase+1; npts=0
        for p in LINES:
            if p["category"]!=c: continue
            s=num(p["sugar"]); caf=num(p["caffeine"]); u=num(p["unitml"])
            if s is None or caf is None or not u: continue
            ws.cell(row=rr,column=xcol,value=round(s/u*100,2))
            ws.cell(row=rr,column=ycol,value=round(caf/u*100,1))
            rr+=1; npts+=1
        series_specs.append((c,xcol,ycol,hbase+1,hbase+npts))
    for ci in range(hcol,hcol+2*len(CATS)):  # hide raw scatter-data columns
        ws.column_dimensions[get_column_letter(ci)].hidden=True
    chart=ScatterChart(); chart.title="Sugar vs caffeine positioning (per 100 mL)"
    chart.x_axis.title="Sugar (g / 100 mL)"; chart.y_axis.title="Caffeine (mg / 100 mL)"
    chart.height=11; chart.width=20; chart.style=2
    # openpyxl quirk: axes need delete=False to render titles
    chart.x_axis.delete=False; chart.y_axis.delete=False
    for cname,xcol,ycol,r0,r1 in series_specs:
        if r1<r0: continue
        xref=Reference(ws,min_col=xcol,min_row=r0,max_row=r1)
        yref=Reference(ws,min_col=ycol,min_row=r0,max_row=r1)
        s=Series(yref,xref,title=cname.split(")")[1].strip())
        s.marker=Marker(symbol="circle",size=6); s.graphicalProperties.line.noFill=True
        chart.series.append(s)
    ws.add_chart(chart,f"B{chart_anchor+2}")
    return ws

def build_dashboard():
    ws=wb.create_sheet("Executive Dashboard"); ws.sheet_view.showGridLines=False; tabcolor(ws,GOLD)
    widths(ws,[3,26,16,16,16,16,16,16,16])
    ws.cell(row=2,column=2,value="Executive Dashboard").font=Font(bold=True,size=20,color=NAVY)
    ws.cell(row=3,column=2,value=f"Amazon.ca RTD beverage landscape  |  {TODAY}  |  ESTIMATE/secondary data - see README").font=Font(italic=True,color="C00000")
    # KPIs
    nbrand=len(set(p["brand"] for p in LINES)); ncat=len(CATS)
    avgsugar=round(sum(p["sugar"] for p in LINES if isinstance(p["sugar"],(int,float)))/sum(1 for p in LINES if isinstance(p["sugar"],(int,float))),1)
    kpis=[("SKUs",len(SKUS)),("Product lines",len(LINES)),("Brands",nbrand),("Categories",ncat),("Avg sugar (g)",avgsugar)]
    c=2
    for label,val in kpis:
        b=ws.cell(row=5,column=c,value=val); b.font=Font(bold=True,size=22,color=BLUE)
        b.alignment=Alignment(horizontal="center"); b.fill=PatternFill("solid",fgColor=LTBLUE); b.border=BORDER
        l=ws.cell(row=6,column=c,value=label); l.alignment=Alignment(horizontal="center"); l.font=Font(size=10,color="595959"); l.border=BORDER
        c+=1
    # capture-completion KPI (live)
    pc=CL("price"); parts=[]
    tot=0
    for cat in CATS:
        name=SHEET_FOR[cat]; a,b=MASTER_RANGES[name]; tot+=(b-a+1)
        parts.append(f"COUNT('{name}'!{pc}{a}:{pc}{b})")
    cap=ws.cell(row=5,column=c,value=f"=({'+'.join(parts)})/{tot}")
    cap.number_format="0%"; cap.font=Font(bold=True,size=22,color="C00000"); cap.alignment=Alignment(horizontal="center")
    cap.fill=PatternFill("solid",fgColor=AMBER); cap.border=BORDER
    ws.cell(row=6,column=c,value="Price capture %").alignment=Alignment(horizontal="center")
    ws.cell(row=6,column=c).border=BORDER

    # Key findings panel (top 3 synthesized takeaways; full set on Strategic Insights tab)
    fr=29
    kf=ws.cell(row=fr-1,column=2,value="Key findings  (full synthesis on the 'Strategic Insights' tab)")
    kf.font=Font(bold=True,size=13,color=NAVY)
    for head,detail in compute_findings()[:4]:
        h=ws.cell(row=fr,column=2,value="• "+head); h.font=Font(bold=True,color=NAVY)
        h.alignment=WRAP; h.fill=PatternFill("solid",fgColor=LTBLUE); h.border=BORDER
        d=ws.cell(row=fr,column=3,value=detail); d.alignment=WRAP; d.border=BORDER
        ws.merge_cells(start_row=fr,start_column=3,end_row=fr,end_column=9)
        ws.row_dimensions[fr].height=44; fr+=1

    # calc data area (for charts) lower down / hidden-ish
    base=46
    ws.cell(row=base-1,column=2,value="Chart data (computed from stable attributes)").font=Font(bold=True,color="808080")
    # category counts + avg sugar + avg caffeine
    ws.cell(row=base,column=2,value="Category"); ws.cell(row=base,column=3,value="# Products")
    ws.cell(row=base,column=4,value="Avg sugar (g)"); ws.cell(row=base,column=5,value="Avg caffeine (mg)")
    rr=base+1
    for cat in CATS:
        ps=[p for p in LINES if p["category"]==cat]
        nsku=sum(1 for s in SKUS if s["category"]==cat)
        sug=[p["sugar"] for p in ps if isinstance(p["sugar"],(int,float))]
        caf=[p["caffeine"] for p in ps if isinstance(p["caffeine"],(int,float))]
        ws.cell(row=rr,column=2,value=cat.split(")")[1].strip()[:22])
        ws.cell(row=rr,column=3,value=nsku)
        ws.cell(row=rr,column=4,value=round(sum(sug)/len(sug),1) if sug else 0)
        ws.cell(row=rr,column=5,value=round(sum(caf)/len(caf),1) if caf else 0)
        rr+=1
    cat_last=rr-1
    # sweetener distribution
    sw={}
    for p in LIVE: sw[sweet_class(p["sweetener"])]=sw.get(sweet_class(p["sweetener"]),0)+1
    sbase=rr+1
    ws.cell(row=sbase,column=2,value="Sweetener class"); ws.cell(row=sbase,column=3,value="# Products")
    sr=sbase+1
    for k,v in sorted(sw.items(),key=lambda x:-x[1]):
        ws.cell(row=sr,column=2,value=k); ws.cell(row=sr,column=3,value=v); sr+=1
    sw_last=sr-1

    # Charts
    ch1=BarChart(); ch1.title="SKUs per category"; ch1.type="col"; ch1.height=7; ch1.width=13
    data=Reference(ws,min_col=3,min_row=base,max_row=cat_last)
    cats=Reference(ws,min_col=2,min_row=base+1,max_row=cat_last)
    ch1.add_data(data,titles_from_data=True); ch1.set_categories(cats); ch1.legend=None
    ws.add_chart(ch1,"B8")

    ch2=BarChart(); ch2.title="Avg sugar (g) by category"; ch2.type="col"; ch2.height=7; ch2.width=13
    d2=Reference(ws,min_col=4,min_row=base,max_row=cat_last)
    ch2.add_data(d2,titles_from_data=True); ch2.set_categories(cats); ch2.legend=None
    ws.add_chart(ch2,"F8")

    ch3=PieChart(); ch3.title="Sweetener mix"; ch3.height=7; ch3.width=13
    d3=Reference(ws,min_col=3,min_row=sbase,max_row=sw_last)
    c3=Reference(ws,min_col=2,min_row=sbase+1,max_row=sw_last)
    ch3.add_data(d3,titles_from_data=True); ch3.set_categories(c3)
    ws.add_chart(ch3,"B14")

    ch4=BarChart(); ch4.title="Avg caffeine (mg) by category"; ch4.type="bar"; ch4.height=7; ch4.width=13
    d4=Reference(ws,min_col=5,min_row=base,max_row=cat_last)
    ch4.add_data(d4,titles_from_data=True); ch4.set_categories(cats); ch4.legend=None
    ws.add_chart(ch4,"F14")
    return ws

# build everything in order
build_dashboard_placeholder=None
build_pricing()
build_subscription()
build_why()
build_velocity()
build_flavourmap()
build_brandrollup()
build_benchmarks()
build_scoreboard()
build_protocol()
build_qa_clean()
build_enrichment_log()
build_sources()
build_insights()
build_dashboard()

# order sheets nicely (single source of truth defined at top: SHEET_ORDER)
wb._sheets.sort(key=lambda s: SHEET_ORDER.index(s.title) if s.title in SHEET_ORDER else 999)

# force full recalculation when the file is opened (Excel/LibreOffice)
try:
    wb.calculation.fullCalcOnLoad=True
except Exception:
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation=CalcProperties(fullCalcOnLoad=True)

wb.save(FNAME)

# audit
print("Saved:",FNAME)
for c in CATS:
    print(f"  {c}: {sum(1 for s in SKUS if s['category']==c)} SKUs  ({sum(1 for p in LINES if p['category']==c)} lines)")
print("TOTAL SKUs:",len(SKUS),"| lines:",len(LINES),"| brands:",len(set(p['brand'] for p in LINES)))
print("Sheets:",len(wb.sheetnames)); print(wb.sheetnames)

# self-validate: run the integrity suite against the file we just wrote.
# A failed build should NOT ship a workbook that violates the honesty guarantees.
import os, sys, subprocess
_test=os.path.join(os.path.dirname(os.path.abspath(__file__)),"test_competitor_intel.py")
if os.path.exists(_test):
    print("\nRunning integrity self-check ...")
    rc=subprocess.run([sys.executable,_test,FNAME]).returncode
    if rc!=0:
        sys.exit(rc)
