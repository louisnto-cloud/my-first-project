#!/usr/bin/env python3
"""Loop batch I4+I9:
 I4 print polish — set print_area to used range on every sheet; freeze header rows on table tabs.
 I9 Executive Brief (1-page) — printable consolidated verdict + numbers + moves + risks."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; TEAL="2E7D8A"; STEEL="3E5C76"; LIGHT="EAF1F4"; AMBER="FFF2CC"; AMBERHEAD="B45309"; GREEN="E2EFDA"; GREENHEAD="1E7D32"; RED="F8D7DA"; REDH="9C2A2A"; WHITE="FFFFFF"; GREY="595959"; GOLD="C9A227"; LINKBLUE="1155CC"
def F(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top"); CTR=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

# ---------- I4: print areas + freeze panes ----------
for ws in wb.worksheets:
    try:
        if ws.max_row>=1 and ws.max_column>=1:
            ws.print_area=f"A1:{ws.cell(row=1,column=min(ws.max_column,26)).column_letter}{ws.max_row}"
    except Exception: pass
# freeze header on the big table tabs (title rows 1-3 -> freeze A4) where safe
freeze_tabs={"08 Extended Landscape":"A4","MÜV Peer Set":"A4","15 Outcomes & M&A":"A4",
             "16 SWOT":"A4","22 Risk Register":"A4","24 Sources":"A4"}
for name,fp in freeze_tabs.items():
    if name in wb.sheetnames:
        try: wb[name].freeze_panes=fp
        except Exception: pass

# ---------- I9: Executive Brief (1-page) ----------
if "Executive Brief 1-page" in wb.sheetnames: del wb["Executive Brief 1-page"]
eb=wb.create_sheet("Executive Brief 1-page")
eb.sheet_view.showGridLines=False
eb.column_dimensions["A"].width=2
for col in "BCDE": eb.column_dimensions[col].width=24
eb.merge_cells("B2:E2"); t=eb["B2"]; t.value="EXECUTIVE BRIEF — MÜV Canada Launch"; t.font=F(bold=True,color=WHITE,size=16); t.fill=fill(NAVY); t.alignment=CTR; eb.row_dimensions[2].height=30
eb.merge_cells("B3:E3"); v=eb["B3"]; v.value="VERDICT: GO — conditional.  Launch as daily-wellness sparkling hydration, made in Canada."; v.font=F(bold=True,color=WHITE,size=11); v.fill=fill(GREENHEAD); v.alignment=CTR; eb.row_dimensions[3].height=22

def block(r,title,color):
    eb.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
    c=eb.cell(row=r,column=2,value=title); c.font=F(bold=True,color=WHITE,size=10); c.fill=fill(color); c.alignment=LEFT; eb.row_dimensions[r].height=18
    return r+1
def line(r,text,bold=False):
    eb.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
    c=eb.cell(row=r,column=2,value=text); c.font=F(bold=bold,size=9); c.alignment=WRAP
    eb.row_dimensions[r].height=14
    return r+1

r=5
r=block(r,"WHAT MÜV IS",STEEL)
r=line(r,"Ready-to-drink sparkling electrolyte CAN (SKU 4338, ~$14.99): carbonated water + Fibersol prebiotic fibre, magnesium glycinate, stevia; 0 sugar, caffeine-free. Food-regulated in Canada.")
r=block(r,"WHY NOW (category)",STEEL)
r=line(r,"US powdered hydration $1.5B in 2024, +20% YoY, 4 straight yrs double-digit (Circana). Liquid I.V. #1, entered Canada 2023. Flavored SPARKLING electrolyte is genuine white space.")
r=block(r,"THE PRICE / SODIUM LADDER",STEEL)
r=line(r,"LMNT 1,000mg / ~$1.50 (performance)  ·  Liquid I.V. 500mg / ~$1.56 (daily)  ·  Nuun 300mg / ~$0.75 (value)  ·  MÜV sparkling can $14.99/pack (daily-wellness sparkling).")
r=block(r,"RECOMMENDATION — 5 MOVES",GREENHEAD)
for txt in ["1. Beachhead: Costco Canada (roadshow sampling) + Amazon.ca / Well.ca / DTC — channels Organika already holds.",
            "2. Position: daily-wellness sparkling, made in Canada — between Liquid I.V. and LMNT.",
            "3. Label to the Supplemented-Food framework from day one (SFFt; cautions if triggered).",
            "4. Keep collagen-beauty claims on the NHP powder line, NOT the can.",
            "5. Borrow LMNT's sample funnel + #WaterTok creators; lean on BC manufacturing for margin."]:
    r=line(r,txt)
r=block(r,"TOP RISKS",REDH)
for txt in ["Collagen efficacy/claims liability → dose to efficacy, food-safe language.",
            "Strategic-owned rivals (Nestlé/Unilever/PepsiCo) → compete on Canadian-made + niche beachhead.",
            "Over-distribution before PMF → prove repeat/velocity before national grocery."]:
    r=line(r,txt)
r=block(r,"REGULATORY (Canada)",AMBERHEAD)
r=line(r,"Sport-electrolytes reclassifying NHP → Supplemented Foods by 2027-12-31; new products comply now. Format is not the decider — claims are. Collagen beauty = NHP-only.")
r+=1
eb.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
f=eb.cell(row=r,column=2,value="Confidence-tagged throughout; MÜV on-can sodium TBD (confirm with brand). Full detail via 🏠 Start Here.")
f.font=F(italic=True,size=8,color=GREY); f.alignment=LEFT
h=eb.cell(row=1,column=13,value="🏠 Home"); h.hyperlink="#'🏠 Start Here'!A1"; h.font=F(bold=True,color=LINKBLUE,underline="single",size=8); h.alignment=Alignment(horizontal="right")
eb.sheet_properties.tabColor=GOLD
eb.page_setup.orientation="portrait"; eb.page_setup.fitToWidth=1; eb.page_setup.fitToHeight=1; eb.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
eb.print_area=f"A1:E{r}"

# place Executive Brief right after Start Here; add to contents
sc=wb["Executive Brief 1-page"]; wb._sheets.remove(sc); wb._sheets.insert(1,sc)
toc=wb["01 Contents"]
rr=toc.max_row+1
a=toc.cell(row=rr,column=1,value="Executive Brief 1-page"); a.hyperlink="#'Executive Brief 1-page'!A1"; a.font=F(bold=True,color=LINKBLUE,underline="single",size=10); a.fill=fill(GOLD); a.border=BORDER; a.alignment=LEFT
b=toc.cell(row=rr,column=2,value="One printable page: verdict, numbers, 5 moves, risks"); b.font=F(size=10); b.border=BORDER; b.alignment=LEFT

wb.save(PATH)
print("I4 print/freeze ok; I9 Executive Brief added. sheets:",len(wb.sheetnames))
