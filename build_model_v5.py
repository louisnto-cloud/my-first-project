"""
Costco UK D93 Portfolio ROI Model — v5

Vs v4:
  - Reverted to the v2 colour scheme (navy title bars, blue section
    headers, yellow inputs, light-blue KPIs, Calibri 11pt). The
    finance-analyst grey look was harder to navigate.
  - Removed the Override column. Every value cell on the Inputs sheet
    is now a single yellow editable cell. Currently-calculated cells
    (GM/unit, GM %, Base Units/Week, Demo Total, End Cap Total) keep
    their formula by default — just type over the cell to override.
  - No hardcoded constants in formulas (factor, end-cap booking cycle,
    discount %, etc. are all surfaced as named inputs).
  - Carries forward from earlier versions: Start Date input + Wk-
    commencing date row, named ranges, the 26-week Bee Propolis
    market-test detail across BEST/IDEAL/WORST scenarios, and the
    Portfolio tab with the 7 D93 SKUs.
"""

import openpyxl
from datetime import date
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

# -----------------------------------------------------------------------------
# COLOURS (v2 palette)
# -----------------------------------------------------------------------------
NAVY = "1F4E78"
BLUE = "2E75B6"
LIGHT_BLUE = "DDEBF7"
YELLOW = "FFF2CC"
GREY = "F2F2F2"
WHITE = "FFFFFF"
GREEN = "C6EFCE"
RED = "FFC7CE"
DARK_GREY = "595959"

thin = Side(border_style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

f_title       = Font(name="Calibri", size=18, bold=True, color=WHITE)
f_section     = Font(name="Calibri", size=12, bold=True, color=WHITE)
f_subhead     = Font(name="Calibri", size=11, bold=True, color=NAVY)
f_label       = Font(name="Calibri", size=11, bold=True)
f_input       = Font(name="Calibri", size=11, color="0070C0", bold=True)
f_calc        = Font(name="Calibri", size=11)
f_note        = Font(name="Calibri", size=10, italic=True, color=DARK_GREY)
f_kpi_val     = Font(name="Calibri", size=14, bold=True, color=NAVY)
f_white_bold  = Font(name="Calibri", size=11, bold=True, color=WHITE)

fill_title   = PatternFill("solid", fgColor=NAVY)
fill_section = PatternFill("solid", fgColor=BLUE)
fill_input   = PatternFill("solid", fgColor=YELLOW)
fill_calc    = PatternFill("solid", fgColor=GREY)
fill_kpi     = PatternFill("solid", fgColor=LIGHT_BLUE)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
right  = Alignment(horizontal="right", vertical="center")
top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

FMT_GBP   = '_-£* #,##0_-;[Red]-£* #,##0_-;_-£* "-"??_-;_-@_-'
FMT_GBP_D = '_-£* #,##0.00_-;[Red]-£* #,##0.00_-;_-£* "-"??_-;_-@_-'
FMT_INT   = "#,##0"
FMT_PCT   = "0.0%"
FMT_MULT  = '0.0"x"'
FMT_DATE  = "dd-mmm"
FMT_DATE_L = "dd-mmm-yyyy"

WEEKS = 26
FIRST_WEEK_NUM = 1
DEFAULT_START_DATE = date(2026, 1, 12)

SCENARIOS = [
    {"key": "BEST",  "name": "BEST – x2 Demo (Wks 5-6) + TPD",
     "promo": {"Demo":[5,6],"End Cap":[],"Fence":[],"TPD":[5,6],"Markdown":[],"Advert":[]},
     "lift_overrides": {5:3.0, 6:3.0},
     "rationale": "Two demo weekends with TPD. Demo lifts ~3x in-week. No end cap, no markdown."},
    {"key": "IDEAL", "name": "IDEAL – x2 Demo + End Cap (Wks 5-6) + TPD",
     "promo": {"Demo":[5,6],"End Cap":[5,6],"Fence":[],"TPD":[5,6],"Markdown":[],"Advert":[]},
     "lift_overrides": {5:3.6, 6:3.6},
     "rationale": "Demo + end cap + TPD stacked for two weeks. Combined velocity ~3.6x base."},
    {"key": "WORST", "name": "WORST – Demo+EndCap+TPD then Markdown (Wks 18-20)",
     "promo": {"Demo":[5,6],"End Cap":[5,6],"Fence":[],"TPD":[5,6],"Markdown":[18,19,20],"Advert":[]},
     "lift_overrides": {5:3.0, 6:3.0, 18:1.0, 19:0.3, 20:8.7},
     "rationale": "Promo plan as IDEAL but residual inventory clears at 50% off over Wks 18-20."},
]

PROMO_TYPES = ["Demo", "End Cap", "Fence", "TPD", "Markdown", "Advert"]

PORTFOLIO_SKUS = [
    {"rank":1,"code":"UK-D93-#1","name":"Bee Propolis Spray 30ml x2",
     "pack":"30ml x 2","sp":15.99,"cost":7.29,"units":17000,"promo":25000,
     "rationale":"Lead pitch — strong GP%, supported by demo + TPD."},
    {"rank":2,"code":"UK-D93-#2","name":"Mag 8-in-1 90's/120's",
     "pack":None,"sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":3,"code":"UK-D93-#3","name":"O1 450g",
     "pack":None,"sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":4,"code":"UK-D93-#4","name":"Belli Bliss Raspberry, 450g",
     "pack":None,"sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":5,"code":"UK-D93-#5","name":"Daily Boost",
     "pack":None,"sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":6,"code":"UK-D93-#6","name":"Berberine 90's/120's",
     "pack":None,"sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":7,"code":"UK-D93-#7","name":"L-theanine Capsules 90's/120's",
     "pack":None,"sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
]

# -----------------------------------------------------------------------------
wb = openpyxl.Workbook()
wb.remove(wb.active)


def wc(ws, row, col, value, *, font=None, fill=None, align=None,
       border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font is not None: c.font = font
    if fill is not None: c.fill = fill
    if align is not None: c.alignment = align
    if border is not None: c.border = border
    if number_format is not None: c.number_format = number_format
    return c


def section_header(ws, row, text, span_start=2, span_end=4):
    ws.merge_cells(start_row=row, start_column=span_start, end_row=row, end_column=span_end)
    c = ws.cell(row=row, column=span_start, value=text)
    c.font = f_section
    c.fill = fill_section
    c.alignment = left
    ws.row_dimensions[row].height = 22


# =============================================================================
# README
# =============================================================================
rd = wb.create_sheet("README")
rd.sheet_view.showGridLines = False
rd.column_dimensions["A"].width = 4
rd.column_dimensions["B"].width = 110

wc(rd, 2, 2, "Costco UK — D93 Portfolio ROI Model",
   font=Font(name="Calibri", size=20, bold=True, color=NAVY), align=left)
wc(rd, 3, 2, "v5.0 · Bee Propolis market-test (baseline) + 1-year portfolio roll-up",
   font=Font(name="Calibri", size=12, italic=True, color=DARK_GREY), align=left)

wc(rd, 5, 2, "HOW TO USE", font=Font(name="Calibri", size=13, bold=True, color=WHITE),
   fill=fill_section, align=left)

lines = [
    "",
    "1.  Two views in one workbook:",
    "      •  Bee Propolis deep-dive — 26-week market test with BEST / IDEAL / WORST scenarios.",
    "      •  Portfolio roll-up — annual Y1 numbers across all seven D93 SKUs with blended margin & ROI.",
    "",
    "2.  Only the YELLOW cells are editable. Every yellow cell is an input you can change.",
    "      •  Some yellow cells contain a formula by default (e.g. Gross Margin per unit = SP − Cost).",
    "      •  Type your own number into any yellow cell to override — the formula is replaced.",
    "      •  No separate 'override' column — one column, one place to change anything.",
    "",
    "3.  Change the 'Test Start Date (W1)' on Inputs and every week column on every sheet shifts dates.",
    "",
    "4.  On the Portfolio sheet, fill in SP / Cost / Y1 Units / Promo Spend for SKUs 2-7 from Asana.",
    "    The bottom row computes the blended GM%, Net Margin %, and ROI on Promo Spend across SKUs.",
    "",
    "",
]
r = 6
for ln in lines:
    wc(rd, r, 2, ln, font=Font(name="Calibri", size=11), align=top_wrap)
    r += 1

wc(rd, r, 2, "KEY CONCEPTS", font=f_subhead); r += 1
concepts = [
    ("Base Demand", "Organic Costco velocity with no promo. ~500 units / 29 weeks × 0.8 sell-through, "
     "scaled to 20 warehouses → ~13.8 units / wh / week."),
    ("Lift Multiplier", "Sales multiplier vs base for a given week. 1.0x = base, 3.0x = demo, "
     "3.6x = demo + end cap, 8.0x+ = clearance markdown."),
    ("TPD", "20% off MSRP = 25% deduction on net (Costco takes a 20% retail margin: 0.20 / 0.80 = 0.25)."),
    ("Demo cost", "£199 / warehouse / weekend × 1.86 staffing-and-POS factor."),
    ("End cap", "£850 / warehouse / 2-week cycle × 1.86 factor. Per-week cost = total ÷ 2."),
    ("ROI on Promo Spend", "Net Profit ÷ Promo Spend. What every £1 of promo investment returned."),
    ("Blended Margin / ROI", "On the Portfolio total row: Σ GP ÷ Σ Rev, Σ Net ÷ Σ Promo. Shows how "
     "high-margin SKUs offset lower-margin SKUs in the overall pitch."),
]
for label, desc in concepts:
    wc(rd, r, 2, f"•  {label} — {desc}", font=Font(name="Calibri", size=11), align=top_wrap)
    rd.row_dimensions[r].height = 38
    r += 1

r += 1
wc(rd, r, 2, "COLOUR KEY", font=f_subhead); r += 1
for label, color, desc in [
    ("Yellow",     YELLOW,     "Editable input — change me"),
    ("Light Blue", LIGHT_BLUE, "Summary KPI"),
    ("Grey",       GREY,       "Reference / structure (you generally won't edit)"),
]:
    c = wc(rd, r, 2, f"   {label}: {desc}", align=left,
           fill=PatternFill("solid", fgColor=color),
           font=Font(name="Calibri", size=11))
    r += 1


# =============================================================================
# INPUTS — single yellow column
# =============================================================================
inp = wb.create_sheet("Inputs")
inp.sheet_view.showGridLines = False
inp.column_dimensions["A"].width = 2
inp.column_dimensions["B"].width = 38
inp.column_dimensions["C"].width = 16
inp.column_dimensions["D"].width = 78

inp.merge_cells("B2:D2")
wc(inp, 2, 2, "INPUTS & ASSUMPTIONS — change any yellow cell",
   font=f_title, fill=fill_title, align=center)
inp.row_dimensions[2].height = 32

# Column labels strip
wc(inp, 3, 2, "Field",  font=f_subhead, fill=fill_calc, align=left,  border=box)
wc(inp, 3, 3, "Value",  font=f_subhead, fill=fill_calc, align=center, border=box)
wc(inp, 3, 4, "Notes / justification", font=f_subhead, fill=fill_calc, align=left, border=box)


def row(ws, r, label, value, note, fmt=None, is_formula=False):
    """One input row. Every value cell is yellow & editable. If is_formula=True,
    the value is a formula string that the cell holds by default — the user
    can overtype to override."""
    wc(ws, r, 2, label, font=f_label, align=left)
    cell_value = f"={value}" if is_formula else value
    cell_font = f_calc if is_formula else f_input
    wc(ws, r, 3, cell_value, font=cell_font, fill=fill_input,
       align=right, border=box, number_format=fmt)
    wc(ws, r, 4, note, font=f_note, align=top_wrap)
    ws.row_dimensions[r].height = 24


# 1 — Product & Distribution
section_header(inp, 5, "1.  Product & Distribution")
row(inp, 6, "SKU", "Bee Propolis Spray 30ml x 2", "Costco UK market test SKU.")
row(inp, 7, "Test Start Date (W1)", DEFAULT_START_DATE,
    "Week-commencing date for W1. Drives every date header.", fmt=FMT_DATE_L)
row(inp, 8, "Test Warehouses", 20, "Number of Costco UK warehouses in test.", fmt=FMT_INT)
row(inp, 9, "Test Period (Weeks)", WEEKS,
    f"Forecast horizon ({WEEKS} weeks, W1 to W{WEEKS}).", fmt=FMT_INT)

# 2 — Pricing & Cost
section_header(inp, 11, "2.  Pricing & Cost (per unit, £)")
row(inp, 12, "Selling Price (MSRP)", 15.99, "Retail per pack.", fmt=FMT_GBP_D)
row(inp, 13, "Cost Price (landed)", 7.29, "Brand cost delivered.", fmt=FMT_GBP_D)
row(inp, 14, "Gross Margin per unit", "C12-C13",
    "Default = SP − Cost. Type over to override (display-only — Forecast uses SP and Cost).",
    fmt=FMT_GBP_D, is_formula=True)
row(inp, 15, "Gross Margin %", "IFERROR((C12-C13)/C12,0)",
    "Default = GP ÷ SP. Type over to override (display-only).",
    fmt=FMT_PCT, is_formula=True)

# 3 — Base Demand
section_header(inp, 17, "3.  Base Demand (no promo)")
row(inp, 18, "Units / Warehouse / Week", 13.79,
    "From prior 30ml SKU run-rate (500 × 0.8 / 29).", fmt="#,##0.00")
row(inp, 19, "Base Units / Week (all warehouses)", "C18*C8",
    "Default = per-wh rate × warehouses. Type over to set the all-wh weekly base directly.",
    fmt=FMT_INT, is_formula=True)

# 4 — Promo Economics
section_header(inp, 21, "4.  Promo Economics")
row(inp, 22, "Demo £ / warehouse / week", 199, "Costco standard demo rate.", fmt=FMT_GBP)
row(inp, 23, "Demo cost factor", 1.86, "Staffing + POS overhead multiplier.", fmt=FMT_MULT)
row(inp, 24, "Demo Total £ / week", "C22*C8*C23",
    "Default = rate × warehouses × factor. Type over to set total directly.",
    fmt=FMT_GBP, is_formula=True)
row(inp, 25, "End Cap £ / warehouse / booking", 850,
    "Costco standard end cap rental per warehouse per booking.", fmt=FMT_GBP)
row(inp, 26, "End Cap cost factor", 1.86, "Same as demo.", fmt=FMT_MULT)
row(inp, 27, "End Cap booking cycle (weeks)", 2,
    "End cap is typically booked across 2 weeks; per-week cost = total ÷ this.", fmt=FMT_INT)
row(inp, 28, "End Cap Total £ / week", "C25*C8*C26/C27",
    "Default = rate × warehouses × factor ÷ booking weeks. Type over to set directly.",
    fmt=FMT_GBP, is_formula=True)
row(inp, 29, "Fence £ / week (when active)", 0, "Not currently budgeted.", fmt=FMT_GBP)
row(inp, 30, "Advertising £ / week (when active)", 0, "Not currently budgeted.", fmt=FMT_GBP)
row(inp, 31, "TPD discount % (of net)", 0.25,
    "20% off MSRP = 25% deduction on net.", fmt=FMT_PCT)
row(inp, 32, "Markdown discount %", 0.50, "Clearance pricing.", fmt=FMT_PCT)

# Named ranges
wb.defined_names["StartDate"]  = DefinedName("StartDate",  attr_text="Inputs!$C$7")
wb.defined_names["WHs"]        = DefinedName("WHs",        attr_text="Inputs!$C$8")
wb.defined_names["Period"]     = DefinedName("Period",     attr_text="Inputs!$C$9")
wb.defined_names["Price"]      = DefinedName("Price",      attr_text="Inputs!$C$12")
wb.defined_names["Cost"]       = DefinedName("Cost",       attr_text="Inputs!$C$13")
wb.defined_names["BaseUnits"]  = DefinedName("BaseUnits",  attr_text="Inputs!$C$19")
wb.defined_names["DemoCost"]   = DefinedName("DemoCost",   attr_text="Inputs!$C$24")
wb.defined_names["EndCapCost"] = DefinedName("EndCapCost", attr_text="Inputs!$C$28")
wb.defined_names["FenceCost"]  = DefinedName("FenceCost",  attr_text="Inputs!$C$29")
wb.defined_names["AdvertCost"] = DefinedName("AdvertCost", attr_text="Inputs!$C$30")
wb.defined_names["TPDPct"]     = DefinedName("TPDPct",     attr_text="Inputs!$C$31")
wb.defined_names["MdPct"]      = DefinedName("MdPct",      attr_text="Inputs!$C$32")

# --- Scenario activity grids -------------------------------------------------
section_header(inp, 34, "5.  Scenario activity & demand lift", span_start=2, span_end=4)
wc(inp, 35, 2,
   "Toggle 1 = promo active, 0 = off per week. Lift is the demand multiplier vs base. "
   "Dates auto-fill from Test Start Date. All yellow cells are editable.",
   font=f_note, align=top_wrap)
inp.merge_cells("B35:D35")
inp.row_dimensions[35].height = 30

WEEK_START_COL = 6
WEEK_END_COL = WEEK_START_COL + WEEKS - 1
for col in range(WEEK_START_COL, WEEK_END_COL + 1):
    inp.column_dimensions[get_column_letter(col)].width = 7.5

scenario_grid_start = {}
current = 37
for sc in SCENARIOS:
    # Title bar
    inp.merge_cells(start_row=current, start_column=2,
                    end_row=current, end_column=WEEK_END_COL)
    wc(inp, current, 2, sc["name"], font=f_white_bold, fill=fill_title, align=left)
    inp.row_dimensions[current].height = 22
    current += 1

    # Rationale
    inp.merge_cells(start_row=current, start_column=2,
                    end_row=current, end_column=WEEK_END_COL)
    wc(inp, current, 2, f"Rationale: {sc['rationale']}", font=f_note, align=top_wrap)
    inp.row_dimensions[current].height = 30
    current += 1

    # Week header
    wc(inp, current, 2, "Week", font=f_label, fill=fill_calc, align=left)
    for i in range(WEEKS):
        wc(inp, current, WEEK_START_COL + i, f"W{FIRST_WEEK_NUM + i}",
           font=f_white_bold, fill=PatternFill("solid", fgColor=BLUE),
           align=center, border=box)
    current += 1

    # Date sub-header
    wc(inp, current, 2, "Wk-commencing", font=f_label, fill=fill_calc, align=left)
    for i in range(WEEKS):
        wc(inp, current, WEEK_START_COL + i, f"=StartDate+7*{i}",
           font=Font(name="Calibri", size=9, color=DARK_GREY),
           fill=fill_calc, align=center, border=box, number_format=FMT_DATE)
    current += 1

    scenario_grid_start[sc["key"]] = current

    for promo in PROMO_TYPES:
        wc(inp, current, 2, promo, font=f_label, align=left)
        active = set(sc["promo"].get(promo, []))
        for i in range(WEEKS):
            wknum = FIRST_WEEK_NUM + i
            wc(inp, current, WEEK_START_COL + i,
               1 if wknum in active else 0,
               font=f_input, fill=fill_input, align=center, border=box,
               number_format="0;;;@")
        current += 1

    wc(inp, current, 2, "Lift Multiplier (x base)", font=f_label, align=left)
    for i in range(WEEKS):
        wknum = FIRST_WEEK_NUM + i
        wc(inp, current, WEEK_START_COL + i,
           sc["lift_overrides"].get(wknum, 1.0),
           font=f_input, fill=fill_input, align=center, border=box,
           number_format=FMT_MULT)
    current += 1
    current += 1

inp.freeze_panes = "F5"


# =============================================================================
# FORECAST
# =============================================================================
fc = wb.create_sheet("Forecast")
fc.sheet_view.showGridLines = False
fc.column_dimensions["A"].width = 2
fc.column_dimensions["B"].width = 32
for col in range(WEEK_START_COL, WEEK_END_COL + 1):
    fc.column_dimensions[get_column_letter(col)].width = 11
TOTAL_COL = WEEK_END_COL + 1
fc.column_dimensions[get_column_letter(TOTAL_COL)].width = 14

fc.merge_cells(start_row=2, start_column=2, end_row=2, end_column=TOTAL_COL)
wc(fc, 2, 2, "FORECAST — Weekly P&L by Scenario",
   font=f_title, fill=fill_title, align=center)
fc.row_dimensions[2].height = 32

PNL_ROWS = [
    ("POS Units",                "units"),
    ("Selling Price",            "price"),
    ("Revenue",                  "rev"),
    ("Cost Price",               "cost"),
    ("Raw COGS",                 "cogs"),
    ("— Demo Spend",             "demo"),
    ("— End Cap Spend",          "endcap"),
    ("— Fence Spend",            "fence"),
    ("— TPD Spend",              "tpd"),
    ("— Markdown Spend",         "markdown"),
    ("— Advertising Spend",      "advert"),
    ("Total Promo Spend",        "promo_total"),
    ("Gross Profit",             "gp"),
    ("Net Profit (after promo)", "net"),
]
PROMO_ROW_OFFSET = {p: i for i, p in enumerate(PROMO_TYPES)}
LIFT_OFFSET = len(PROMO_TYPES)

fc_row = 4
scenario_summary_rows = {}

for sc in SCENARIOS:
    fc.merge_cells(start_row=fc_row, start_column=2, end_row=fc_row, end_column=TOTAL_COL)
    wc(fc, fc_row, 2, sc["name"], font=f_white_bold, fill=fill_title, align=left)
    fc.row_dimensions[fc_row].height = 22
    fc_row += 1

    # Week header
    wc(fc, fc_row, 2, "Week", font=f_label, fill=fill_calc, align=left)
    for i in range(WEEKS):
        wc(fc, fc_row, WEEK_START_COL + i, f"W{FIRST_WEEK_NUM + i}",
           font=f_white_bold, fill=PatternFill("solid", fgColor=BLUE),
           align=center, border=box)
    wc(fc, fc_row, TOTAL_COL, "TOTAL",
       font=f_white_bold, fill=fill_title, align=center, border=box)
    fc_row += 1

    # Date sub-header
    wc(fc, fc_row, 2, "Wk-commencing", font=f_label, fill=fill_calc, align=left)
    for i in range(WEEKS):
        wc(fc, fc_row, WEEK_START_COL + i, f"=StartDate+7*{i}",
           font=Font(name="Calibri", size=9, color=DARK_GREY),
           fill=fill_calc, align=center, border=box, number_format=FMT_DATE)
    wc(fc, fc_row, TOTAL_COL, "", fill=fill_calc, border=box)
    fc_row += 1

    scenario_pnl_rows = {}
    grid_top = scenario_grid_start[sc["key"]]
    promo_rows_on_inp = {p: grid_top + PROMO_ROW_OFFSET[p] for p in PROMO_TYPES}
    lift_row_on_inp = grid_top + LIFT_OFFSET

    for label, key in PNL_ROWS:
        scenario_pnl_rows[key] = fc_row
        bold = key in ("rev", "cogs", "promo_total", "gp", "net", "units")
        navy_bold = key in ("gp", "net")
        lbl_font = (Font(name="Calibri", size=11, bold=True, color=NAVY)
                    if navy_bold else (f_label if bold else f_calc))
        wc(fc, fc_row, 2, label, font=lbl_font, align=left)

        for i in range(WEEKS):
            wcol = get_column_letter(WEEK_START_COL + i)
            if key == "units":
                formula = f"=ROUND(BaseUnits*Inputs!{wcol}{lift_row_on_inp},0)"
            elif key == "price":
                formula = "=Price"
            elif key == "rev":
                formula = f"={wcol}{scenario_pnl_rows['units']}*{wcol}{scenario_pnl_rows['price']}"
            elif key == "cost":
                formula = "=Cost"
            elif key == "cogs":
                formula = f"={wcol}{scenario_pnl_rows['units']}*{wcol}{scenario_pnl_rows['cost']}"
            elif key == "demo":
                formula = f"=Inputs!{wcol}{promo_rows_on_inp['Demo']}*DemoCost"
            elif key == "endcap":
                formula = f"=Inputs!{wcol}{promo_rows_on_inp['End Cap']}*EndCapCost"
            elif key == "fence":
                formula = f"=Inputs!{wcol}{promo_rows_on_inp['Fence']}*FenceCost"
            elif key == "tpd":
                formula = (f"=Inputs!{wcol}{promo_rows_on_inp['TPD']}"
                           f"*{wcol}{scenario_pnl_rows['units']}"
                           f"*{wcol}{scenario_pnl_rows['price']}*TPDPct")
            elif key == "markdown":
                formula = (f"=Inputs!{wcol}{promo_rows_on_inp['Markdown']}"
                           f"*{wcol}{scenario_pnl_rows['units']}"
                           f"*{wcol}{scenario_pnl_rows['price']}*MdPct")
            elif key == "advert":
                formula = f"=Inputs!{wcol}{promo_rows_on_inp['Advert']}*AdvertCost"
            elif key == "promo_total":
                formula = (f"=SUM({wcol}{scenario_pnl_rows['demo']}:"
                           f"{wcol}{scenario_pnl_rows['advert']})")
            elif key == "gp":
                formula = f"={wcol}{scenario_pnl_rows['rev']}-{wcol}{scenario_pnl_rows['cogs']}"
            elif key == "net":
                formula = f"={wcol}{scenario_pnl_rows['gp']}-{wcol}{scenario_pnl_rows['promo_total']}"
            else:
                formula = ""
            num_fmt = (FMT_INT if key == "units"
                       else FMT_GBP_D if key in ("price","cost")
                       else FMT_GBP)
            wc(fc, fc_row, WEEK_START_COL + i, formula,
               font=f_calc, fill=fill_calc, align=right, border=box,
               number_format=num_fmt)

        first_w = get_column_letter(WEEK_START_COL)
        last_w = get_column_letter(WEEK_END_COL)
        if key in ("price","cost"):
            tot_formula = "=Price" if key == "price" else "=Cost"
        else:
            tot_formula = f"=SUM({first_w}{fc_row}:{last_w}{fc_row})"
        wc(fc, fc_row, TOTAL_COL, tot_formula,
           font=Font(name="Calibri", size=11, bold=True, color=NAVY),
           fill=fill_kpi, align=right, border=box,
           number_format=(FMT_INT if key == "units"
                          else FMT_GBP_D if key in ("price","cost")
                          else FMT_GBP))
        fc_row += 1

    scenario_summary_rows[sc["key"]] = scenario_pnl_rows
    fc_row += 2

fc.freeze_panes = "C4"


# =============================================================================
# SUMMARY
# =============================================================================
sm = wb.create_sheet("Summary")
sm.sheet_view.showGridLines = False
sm.column_dimensions["A"].width = 2
sm.column_dimensions["B"].width = 36
for i in range(len(SCENARIOS)):
    sm.column_dimensions[get_column_letter(3 + i)].width = 22

end_col = 2 + len(SCENARIOS)
sm.merge_cells(start_row=2, start_column=2, end_row=2, end_column=end_col)
wc(sm, 2, 2, "ROI SUMMARY — Bee Propolis Scenario Comparison",
   font=f_title, fill=fill_title, align=center)
sm.row_dimensions[2].height = 32

sm.merge_cells(start_row=3, start_column=2, end_row=3, end_column=end_col)
wc(sm, 3, 2,
   f'="Test window: W1 ("&TEXT(StartDate,"dd-mmm-yyyy")&")  →  W"&Period&" ("&TEXT(StartDate+7*(Period-1),"dd-mmm-yyyy")&")"',
   font=Font(name="Calibri", size=11, italic=True, color=DARK_GREY),
   fill=fill_calc, align=center)

hr = 5
wc(sm, hr, 2, "Metric", font=f_white_bold, fill=fill_section, align=left)
for i, sc in enumerate(SCENARIOS):
    wc(sm, hr, 3 + i, sc["key"], font=f_white_bold, fill=fill_section, align=center)
sm.row_dimensions[hr].height = 24

TC = get_column_letter(TOTAL_COL)
KPIS = [
    ("Units Sold","units",FMT_INT,False),
    ("Revenue","rev",FMT_GBP,False),
    ("Raw COGS","cogs",FMT_GBP,False),
    ("Gross Profit","gp",FMT_GBP,True),
    ("Gross Margin %","gp_pct",FMT_PCT,True),
    ("Demo Spend","demo",FMT_GBP,False),
    ("End Cap Spend","endcap",FMT_GBP,False),
    ("Fence Spend","fence",FMT_GBP,False),
    ("TPD Spend","tpd",FMT_GBP,False),
    ("Markdown Spend","markdown",FMT_GBP,False),
    ("Advertising Spend","advert",FMT_GBP,False),
    ("Total Promo Spend","promo_total",FMT_GBP,True),
    ("Net Profit","net",FMT_GBP,True),
    ("Net Margin %","net_pct",FMT_PCT,True),
    ("ROI on Promo Spend","roi",FMT_PCT,True),
    ("Avg Units / WH / Week","uphpw",FMT_INT,False),
]

sr = hr + 1
kpi_row_idx = {}
for label, key, fmt, bold in KPIS:
    kpi_row_idx[key] = sr
    wc(sm, sr, 2, label,
       font=Font(name="Calibri",size=11,bold=bold,color=NAVY if bold else "000000"),
       fill=fill_kpi if bold else None, align=left, border=box)
    for i, sc in enumerate(SCENARIOS):
        rows = scenario_summary_rows[sc["key"]]
        if key in rows:
            formula = f"=Forecast!{TC}{rows[key]}"
        elif key == "gp_pct":
            formula = f"=IFERROR(Forecast!{TC}{rows['gp']}/Forecast!{TC}{rows['rev']},0)"
        elif key == "net_pct":
            formula = f"=IFERROR(Forecast!{TC}{rows['net']}/Forecast!{TC}{rows['rev']},0)"
        elif key == "roi":
            formula = f"=IFERROR(Forecast!{TC}{rows['net']}/Forecast!{TC}{rows['promo_total']},0)"
        elif key == "uphpw":
            formula = f"=IFERROR(Forecast!{TC}{rows['units']}/WHs/Period,0)"
        else:
            formula = ""
        wc(sm, sr, 3 + i, formula,
           font=f_kpi_val if bold else f_calc,
           fill=fill_kpi if bold else None,
           align=right, border=box, number_format=fmt)
    sm.row_dimensions[sr].height = 22 if not bold else 26
    sr += 1

last_col_letter = get_column_letter(2 + len(SCENARIOS))
for k in ("net","roi"):
    r = kpi_row_idx[k]
    rng = f"C{r}:{last_col_letter}{r}"
    sm.conditional_formatting.add(rng,
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=RED),
                   font=Font(bold=True, color="9C0006")))
    sm.conditional_formatting.add(rng,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=PatternFill("solid", fgColor=GREEN),
                   font=Font(bold=True, color="006100")))

sr += 2
sm.merge_cells(start_row=sr, start_column=2, end_row=sr, end_column=end_col)
wc(sm, sr, 2, "Why each scenario?", font=f_section, fill=fill_section, align=left)
sm.row_dimensions[sr].height = 22
sr += 1
for sc in SCENARIOS:
    wc(sm, sr, 2, sc["key"],
       font=Font(name="Calibri",size=11,bold=True,color=NAVY), align=top_wrap)
    sm.merge_cells(start_row=sr, start_column=3, end_row=sr, end_column=end_col)
    wc(sm, sr, 3, sc["rationale"], font=f_note, align=top_wrap)
    sm.row_dimensions[sr].height = 38
    sr += 1

sm.freeze_panes = "C6"


# =============================================================================
# PORTFOLIO
# =============================================================================
pf = wb.create_sheet("Portfolio")
pf.sheet_view.showGridLines = False
pf.column_dimensions["A"].width = 2
widths = {"B":5,"C":12,"D":32,"E":12,"F":9,"G":9,"H":9,"I":8,"J":11,"K":13,
          "L":13,"M":13,"N":12,"O":13,"P":8,"Q":8,"R":40}
for col, w in widths.items():
    pf.column_dimensions[col].width = w

pf.merge_cells("B2:R2")
wc(pf, 2, 2, "PORTFOLIO — D93 Pitch SKUs (Annual Y1 Roll-Up)",
   font=f_title, fill=fill_title, align=center)
pf.row_dimensions[2].height = 32

pf.merge_cells("B3:R3")
wc(pf, 3, 2,
   "Bee Propolis pre-filled at indicative annual numbers. For SKUs 2-7, "
   "paste SP / Cost / Y1 Units / Promo Spend from Asana into the yellow cells. "
   "Bottom row blends across SKUs.",
   font=Font(name="Calibri",size=11,italic=True,color=DARK_GREY), align=top_wrap)
pf.row_dimensions[3].height = 30

HDR_ROW = 5
hdrs = [("B","#"),("C","Code"),("D","SKU"),("E","Pack"),
        ("F","SP £"),("G","Cost £"),("H","GM/u £"),("I","GM %"),
        ("J","Y1 Units"),("K","Revenue £"),("L","COGS £"),("M","GP £"),
        ("N","Promo £"),("O","Net Profit £"),("P","Net %"),
        ("Q","ROI %"),("R","Rationale / status")]
for col_letter, txt in hdrs:
    al = left if col_letter in ("C","D","E","R") else center
    wc(pf, HDR_ROW, column_index_from_string(col_letter), txt,
       font=f_white_bold, fill=fill_section, align=al, border=box)

sku_start_row = HDR_ROW + 1
for idx, sku in enumerate(PORTFOLIO_SKUS):
    r = sku_start_row + idx
    wc(pf, r, 2, sku["rank"], font=f_label, fill=fill_calc, align=center, border=box,
       number_format=FMT_INT)
    wc(pf, r, 3, sku["code"], font=f_label, fill=fill_calc, align=left, border=box)
    wc(pf, r, 4, sku["name"], font=f_label, fill=fill_calc, align=left, border=box)
    wc(pf, r, 5, sku["pack"], font=f_input, fill=fill_input, align=left, border=box)
    wc(pf, r, 6, sku["sp"],   font=f_input, fill=fill_input, align=right, border=box,
       number_format=FMT_GBP_D)
    wc(pf, r, 7, sku["cost"], font=f_input, fill=fill_input, align=right, border=box,
       number_format=FMT_GBP_D)
    wc(pf, r, 8, f"=IFERROR(F{r}-G{r},0)",        font=f_calc, fill=fill_calc, align=right, border=box, number_format=FMT_GBP_D)
    wc(pf, r, 9, f"=IFERROR((F{r}-G{r})/F{r},0)", font=f_calc, fill=fill_calc, align=right, border=box, number_format=FMT_PCT)
    wc(pf, r,10, sku["units"], font=f_input, fill=fill_input, align=right, border=box, number_format=FMT_INT)
    wc(pf, r,11, f"=IFERROR(J{r}*F{r},0)", font=f_calc, fill=fill_calc, align=right, border=box, number_format=FMT_GBP)
    wc(pf, r,12, f"=IFERROR(J{r}*G{r},0)", font=f_calc, fill=fill_calc, align=right, border=box, number_format=FMT_GBP)
    wc(pf, r,13, f"=IFERROR(K{r}-L{r},0)",
       font=Font(name="Calibri",size=11,bold=True,color=NAVY),
       fill=fill_kpi, align=right, border=box, number_format=FMT_GBP)
    wc(pf, r,14, sku["promo"], font=f_input, fill=fill_input, align=right, border=box, number_format=FMT_GBP)
    wc(pf, r,15, f"=IFERROR(M{r}-N{r},0)",
       font=Font(name="Calibri",size=11,bold=True,color=NAVY),
       fill=fill_kpi, align=right, border=box, number_format=FMT_GBP)
    wc(pf, r,16, f"=IFERROR(O{r}/K{r},0)", font=f_calc, fill=fill_calc, align=right, border=box, number_format=FMT_PCT)
    wc(pf, r,17, f"=IFERROR(O{r}/N{r},0)", font=f_calc, fill=fill_calc, align=right, border=box, number_format=FMT_PCT)
    wc(pf, r,18, sku["rationale"], font=f_input, fill=fill_input, align=left, border=box)
    pf.row_dimensions[r].height = 20

last_sku_row = sku_start_row + len(PORTFOLIO_SKUS) - 1
tr = last_sku_row + 1

wc(pf, tr, 2, "", fill=fill_title, border=box)
wc(pf, tr, 3, "", fill=fill_title, border=box)
wc(pf, tr, 4, "PORTFOLIO TOTAL / BLENDED",
   font=f_white_bold, fill=fill_title, align=left, border=box)
wc(pf, tr, 5, "", fill=fill_title, border=box)
wc(pf, tr, 6, "", fill=fill_title, border=box)
wc(pf, tr, 7, "", fill=fill_title, border=box)
wc(pf, tr, 8, "", fill=fill_title, border=box)
wc(pf, tr, 9, f"=IFERROR(SUM(M{sku_start_row}:M{last_sku_row})/SUM(K{sku_start_row}:K{last_sku_row}),0)",
   font=f_white_bold, fill=fill_title, align=right, border=box, number_format=FMT_PCT)
for col_letter, fmt in [("J",FMT_INT),("K",FMT_GBP),("L",FMT_GBP),("M",FMT_GBP),
                       ("N",FMT_GBP),("O",FMT_GBP)]:
    wc(pf, tr, column_index_from_string(col_letter),
       f"=SUM({col_letter}{sku_start_row}:{col_letter}{last_sku_row})",
       font=f_white_bold, fill=fill_title, align=right, border=box, number_format=fmt)
wc(pf, tr,16, f"=IFERROR(O{tr}/K{tr},0)",
   font=f_white_bold, fill=fill_title, align=right, border=box, number_format=FMT_PCT)
wc(pf, tr,17, f"=IFERROR(O{tr}/N{tr},0)",
   font=f_white_bold, fill=fill_title, align=right, border=box, number_format=FMT_PCT)
wc(pf, tr,18, "Blended across SKUs",
   font=f_white_bold, fill=fill_title, align=left, border=box)
pf.row_dimensions[tr].height = 26

nr = tr + 2
pf.merge_cells(start_row=nr, start_column=2, end_row=nr, end_column=18)
wc(pf, nr, 2,
   "Blended GM% and Net% are revenue-weighted (Σ GP ÷ Σ Rev, Σ Net ÷ Σ Rev). "
   "Blended ROI is Σ Net Profit ÷ Σ Promo Spend. Yellow cells are inputs — fill in "
   "from Asana for rows 2-7.",
   font=f_note, align=top_wrap)
pf.row_dimensions[nr].height = 32

pf.freeze_panes = "B6"


# =============================================================================
out = "/home/user/my-first-project/Costco_UK_D93_Portfolio_ROI_v5.xlsx"
wb.save(out)
print(f"Saved: {out}")
