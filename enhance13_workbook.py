#!/usr/bin/env python3
"""Loop batch I16 + I20:
 I16 Data Confidence Scoreboard — counts ✅/◐/⚠ (and HARD/DIRECTIONAL/FLAG) across all tabs, with chart.
 I20 Cover refresh — date, decision line, and status reflect current state (38+ tabs, Start Here, Exec Brief)."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.properties import PageSetupProperties
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; STEEL="3E5C76"; LIGHT="EAF1F4"; AMBER="FFF2CC"; AMBERHEAD="B45309"; GREEN="E2EFDA"; GREENHEAD="1E7D32"; GREY="595959"; GOLD="C9A227"; WHITE="FFFFFF"; LINKBLUE="1155CC"
def F(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CTR=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

# ---------- I16: count confidence markers ----------
hard=direc=flag=0
for ws in wb.worksheets:
    if ws.title in ("Data Confidence",): continue
    for row in ws.iter_rows():
        for c in row:
            if not isinstance(c.value,str): continue
            v=c.value
            hard  += v.count("✅") + v.count("HARD")
            direc += v.count("◐") + v.count("DIRECTIONAL")
            flag  += v.count("⚠") + v.count("FLAG")
total=hard+direc+flag or 1

if "Data Confidence" in wb.sheetnames: del wb["Data Confidence"]
dc=wb.create_sheet("Data Confidence")
dc.sheet_view.showGridLines=False
dc.merge_cells("A1:D1"); t=dc["A1"]; t.value="Data Confidence Scoreboard"; t.font=F(bold=True,color=WHITE,size=15); t.fill=fill(NAVY); t.alignment=CTR; dc.row_dimensions[1].height=26
dc.merge_cells("A2:D2"); s=dc["A2"]; s.value="Tally of confidence tags across the whole dossier — how much rests on hard evidence vs estimates."; s.font=F(italic=True,color=GREY,size=10); s.alignment=CTR
hdr=["Confidence tier","Marker","Count","Share"]
for j,h in enumerate(hdr):
    x=dc.cell(row=4,column=1+j,value=h); x.font=F(bold=True,color=WHITE,size=10); x.fill=fill(STEEL); x.alignment=CTR; x.border=BORDER
rows=[("Hard (primary/scanner/regulatory)","✅ HARD",hard,GREEN),
      ("Directional (research-firm/single-source)","◐ DIRECTIONAL",direc,LIGHT),
      ("Flag (estimate/verify/contested)","⚠ FLAG",flag,AMBER)]
r=5
for name,mark,cnt,color in rows:
    dc.cell(row=r,column=1,value=name).font=F(size=10,bold=True); dc.cell(row=r,column=1).border=BORDER; dc.cell(row=r,column=1).alignment=LEFT; dc.cell(row=r,column=1).fill=fill(color)
    dc.cell(row=r,column=2,value=mark).font=F(size=10); dc.cell(row=r,column=2).border=BORDER; dc.cell(row=r,column=2).alignment=CTR
    dc.cell(row=r,column=3,value=cnt).font=F(size=11,bold=True); dc.cell(row=r,column=3).border=BORDER; dc.cell(row=r,column=3).alignment=CTR
    sh=dc.cell(row=r,column=4,value=cnt/total); sh.number_format='0%'; sh.font=F(size=10); sh.border=BORDER; sh.alignment=CTR
    r+=1
dc.cell(row=r,column=1,value="TOTAL tags").font=F(bold=True); dc.cell(row=r,column=3,value=total).font=F(bold=True); dc.cell(row=r,column=3).alignment=CTR
dc.column_dimensions["A"].width=40; dc.column_dimensions["B"].width=18; dc.column_dimensions["C"].width=10; dc.column_dimensions["D"].width=10
# chart
ch=BarChart(); ch.type="col"; ch.title="Confidence tag distribution"; ch.height=7; ch.width=13
ch.add_data(Reference(dc,min_col=3,min_row=4,max_row=7),titles_from_data=True)
ch.set_categories(Reference(dc,min_col=2,min_row=5,max_row=7)); ch.legend=None
ch.dataLabels=DataLabelList(); ch.dataLabels.showVal=True
dc.add_chart(ch,"A"+str(r+2))
note=dc.cell(row=r+16,column=1,value="Read: a healthy dossier leans on hard/primary evidence for its load-bearing claims (verified on the Verification Log) while transparently flagging estimates. Market-size projections and all financial inputs are ◐/⚠ by nature — never quote them as fact.")
dc.merge_cells(start_row=r+16,start_column=1,end_row=r+16,end_column=4); note.font=F(italic=True,size=9,color=GREY); note.fill=fill(LIGHT); note.alignment=LEFT; note.border=BORDER; dc.row_dimensions[r+16].height=44
h=dc.cell(row=1,column=13,value="🏠 Home"); h.hyperlink="#'🏠 Start Here'!A1"; h.font=F(bold=True,color=LINKBLUE,underline="single",size=8); h.alignment=Alignment(horizontal="right")
dc.sheet_properties.tabColor=GOLD
dc.page_setup.orientation="landscape"; dc.page_setup.fitToWidth=1; dc.page_setup.fitToHeight=0; dc.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
# place before 24 Sources; contents
s2=wb["Data Confidence"]; wb._sheets.remove(s2); idx=wb.sheetnames.index("Verification Log"); wb._sheets.insert(idx,s2)
toc=wb["01 Contents"]; row=toc.max_row+1
a=toc.cell(row=row,column=1,value="Data Confidence"); a.hyperlink="#'Data Confidence'!A1"; a.font=F(bold=True,color=LINKBLUE,underline="single",size=10); a.fill=fill(GOLD); a.border=BORDER; a.alignment=LEFT
b=toc.cell(row=row,column=2,value="Scoreboard: ✅/◐/⚠ tag counts across the dossier"); b.font=F(size=10); b.border=BORDER; b.alignment=LEFT

# ---------- I20: Cover refresh (safe single-cell updates) ----------
cov=wb["00 Cover"]
cov["C8"]="Whether/how to launch MÜV — a Canadian sparkling electrolyte RTD (can) — in Canada"
cov["C12"]="2026-07-18"
cov["C13"]=("Working draft — 38+ tabs. Start on 🏠 Start Here; one-page summary on Executive Brief 1-page. "
            "Interactive: Financial Model · Scenario Comparison · Go-NoGo. Verify amber-flagged items before external use.")

wb.save(PATH)
print(f"I16 scoreboard (hard {hard}, direc {direc}, flag {flag}); I20 cover refreshed. sheets:",len(wb.sheetnames))
