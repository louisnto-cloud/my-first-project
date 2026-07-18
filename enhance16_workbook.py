#!/usr/bin/env python3
"""Loop batch I23+I24 (+dashboard thesis line):
 I23 per-scenario one-line takeaway on Scenario Comparison.
 I24 glossary cross-link note to Canada Regulatory v2.
 +  one-line thesis banner on the Dashboard (empty row 3)."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; STEEL="3E5C76"; LIGHT="EAF1F4"; AMBER="FFF2CC"; GREEN="E2EFDA"; GREENHEAD="1E7D32"; GREY="595959"; WHITE="FFFFFF"
def F(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

# ---------- I23: per-scenario takeaways on Scenario Comparison ----------
sc=wb["Scenario Comparison"]
r=sc.max_row+2
sc.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
h=sc.cell(row=r,column=1,value="TAKEAWAY BY SCENARIO"); h.font=F(bold=True,color=WHITE,size=10); h.fill=fill(STEEL); h.alignment=Alignment(horizontal="left",vertical="center",indent=1)
r+=1
takeaways=[
 ("Conservative","Downside case — thin margins, later break-even. Validates the floor: even here MÜV has a defined path, not a hole.",AMBER),
 ("Base","Planning anchor — healthy contribution and break-even within the Year-1 door plan. The realistic case to underwrite.",GREEN),
 ("Aggressive","Upside — if the sparkling wedge + Costco beachhead over-deliver on velocity. Treat as ceiling, not plan.",LIGHT),
]
for name,txt,color in takeaways:
    a=sc.cell(row=r,column=1,value=name); a.font=F(bold=True,size=10); a.fill=fill(color); a.border=BORDER; a.alignment=Alignment(vertical="top")
    sc.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
    b=sc.cell(row=r,column=2,value=txt); b.font=F(size=10); b.border=BORDER; b.alignment=Alignment(wrap_text=True,vertical="top")
    sc.row_dimensions[r].height=30; r+=1

# ---------- I24: glossary cross-link ----------
gl=wb["25 Glossary"]
r=gl.max_row+2
gl.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
c=gl.cell(row=r,column=1,value="→ For full detail on the Supplemented-Foods pathway, SFFt, SFCI and the electrolyte→food reclassification, see the ‘Canada Regulatory v2’ tab.")
c.font=F(italic=True,size=9,color=GREENHEAD); c.fill=fill(GREEN); c.alignment=Alignment(wrap_text=True,vertical="center"); c.border=BORDER
gl.row_dimensions[r].height=28

# ---------- Dashboard thesis banner (empty row 3) ----------
d=wb["Dashboard"]
d.merge_cells(start_row=3,start_column=1,end_row=3,end_column=12)
t=d.cell(row=3,column=1,value="THESIS: MÜV wins the daily-wellness SPARKLING electrolyte lane, made in Canada — food-regulated, priced with the electrolyte peers, GTM borrowed from the soda winners.")
t.font=F(bold=True,color=WHITE,size=9); t.fill=fill(GREENHEAD); t.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); d.row_dimensions[3].height=18

wb.save(PATH)
print("I23 takeaways + I24 glossary link + Dashboard thesis banner added.")
