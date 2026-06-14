# -*- coding: utf-8 -*-
"""Builds Organika RTD Community Partnerships Tracker_v2.xlsx, styled to match the BC Tracker.

v2 changes per Louis:
- No proposed activations anywhere. Activation Type, Activation Date and idea text all blank.
  The Activation Calendar stays live but starts empty and fills as the team books real dates.
- Sales Team is Maddie only. Maddie is the one owner on the dropdown and is prefilled as
  Primary Owner on every researched partner row.
- New Suggested Events tab: an empty, structured parking lot for event ideas farther out.
- Review pass: auto numbering and auto type label formulas, List Health block on the
  Dashboard, % to Target on Type Summary, stale amber rule applies to P1 rows only,
  corrected apostrophes in three partner names, rewritten Guide.
"""
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.properties import PageSetupProperties

OUT = "/home/user/my-first-project/Organika RTD Community Partnerships Tracker_v10.xlsx"

# ---------- palette (refined, premium, mostly tonal greens + warm neutrals) ----------
C_TITLE   = "FF22413A"   # deepest evergreen, title bands and big numbers
C_HEADER  = "FF2E5A4E"   # deep green, table headers and section labels
C_TOTAL   = "FFEFF4F1"   # total row tint
C_DATA    = "FF2C3A34"   # body text
C_SUB     = "FF73827B"   # secondary text, soft grey green
WHITE     = "FFFFFFFF"
ACCENT    = "FF6FA392"   # soft green accent line
CARD_BRD  = "FFE5EBE8"   # hairline
# status / state fills (kept for meaning, muted so they read calm not candy)
F_GREEN   = "FFE6F5EC"
F_AMBER   = "FFFBF2DA"
F_BLUE    = "FFE6EEF8"
F_PALEBLU = "FFEDF4FB"
F_RED     = "FFFBE9E9"
F_GREY    = "FFF5F7F6"
F_SLATE   = "FFECEFF2"
F_P1      = "FFFBF0E0"
F_P2      = "FFE9EFF8"
F_P3      = "FFF3F6F4"

BORDER_CLR = "FFE5EBE8"
_hair = Side(style="thin", color=CARD_BRD)
BORD = Border(bottom=_hair)                                   # hairline horizontal rule only
HEADBORD = Border(bottom=Side(style="medium", color=C_HEADER))

FONTNAME = "Helvetica Neue"
def font(sz=12, b=False, color=C_DATA):
    return Font(name=FONTNAME, size=sz, bold=b, color=color)
def fill(c):
    return PatternFill(fill_type="solid", fgColor=c)
A_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_L = Alignment(horizontal="left", vertical="center", wrap_text=False, indent=1)
A_CL = Alignment(horizontal="center", vertical="center", wrap_text=False)
A_LW = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

FMT_MONEY = "$#,##0"
FMT_CENTS = "$#,##0.00"
FMT_PCT   = "0.0%"
FMT_DATE  = "yyyy/mm/dd"
FMT_INT   = "0"

wb = openpyxl.Workbook()

# =====================================================================
# TYPE TAB COLUMNS
# =====================================================================
HEADERS = ["#","Partner Name","Partnership Type","City","Neighbourhood","Priority","Status",
"Primary Owner","Contact Name","Role","Contact Email","Contact Phone","Instagram or Website",
"Audience Size","Audience Source","Source","Warm","Last Contacted","Days Since Activity",
"Next Action","Next Action Date","Activation Type","Activation Date","Cases Committed",
"Cases Delivered","Cost","Cost Per Can","Contra Value","Deliverables Promised",
"Deliverables Received","What They Want","Risks","Nearby Retail Doors","In BC Tracker?",
"Raspberry 4338","Lemon Lime 4336","Pineapple Passion Fruit 4340","Notes"]
NCOL = len(HEADERS)                       # 38
LASTCOL = get_column_letter(NCOL)         # AL
DATA_ROWS = 30                            # rows 3..32 per type tab
FIRST = 3
LASTROW = FIRST + DATA_ROWS - 1           # 32

WIDTHS = {1:5,2:30,3:20,4:14,5:18,6:10,7:16,8:16,9:20,10:16,11:26,12:15,13:26,14:16,15:30,
16:14,17:12,18:15,19:13,20:28,21:15,22:18,23:15,24:13,25:13,26:12,27:13,28:14,29:26,30:26,
31:18,32:18,33:22,34:13,35:15,36:15,37:22,38:44}

# key -> column number
KEYCOL = {"name":2,"city":4,"hood":5,"prio":6,"status":7,"owner":8,"contact":9,"role":10,
"email":11,"phone":12,"ig":13,"aud":14,"audsrc":15,"source":16,"warm":17,"last":18,
"nextact":20,"nextdate":21,"acttype":22,"actdate":23,"cases_c":24,"cases_d":25,"cost":26,
"contra":28,"deliv_p":29,"deliv_r":30,"want":31,"risks":32,"doors":33,"inbc":34,
"rasp":35,"lime":36,"pine":37,"notes":38}
DATECOLS = {18,21,23}
MONEYCOLS = {26,28}
CENTERCOLS = {1,4,6,7,8,10,12,14,16,17,18,19,21,22,23,24,25,26,27,28,34,35,36,37}

# =====================================================================
# PARTNER DATA  (60 verified rows, Tevah first on Wellness & Recovery)
# Notes hold research facts only. No activation is proposed anywhere.
# =====================================================================
TYPES = ["Run Clubs","Gyms & Studios","Events & Festivals","Sports Teams & Leagues",
         "Campus & Student Groups","Wellness & Recovery","Ambassadors & Creators","Charity & Causes"]

PARTNERS = {
'Run Clubs':[
 {'name': 'East Van Run Crew', 'city': 'Vancouver', 'hood': 'East Vancouver', 'prio': 'P1', 'ig': '@eastvanruncrew', 'aud': '11,000 IG', 'audsrc': 'Instagram @eastvanruncrew follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Monday night runs starting from a rotating East Van brewery.'},
 {'name': 'Social Run Club YVR', 'city': 'Vancouver', 'hood': 'Yaletown', 'prio': 'P1', 'ig': '@socialrunclub.yvr', 'aud': '12,000 IG', 'audsrc': 'Instagram @socialrunclub.yvr follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Saturday 9:30am Form Focus run that ends with a coffee social.'},
 {'name': 'Slow Girls Run Club', 'city': 'Vancouver', 'hood': 'Downtown', 'prio': 'P1', 'ig': '@slowgirlsrunclub', 'aud': '9,400 IG', 'audsrc': 'Instagram @slowgirlsrunclub follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Saturday 8am Social Saturdays 5km. Women focused community.'},
 {'name': 'One Run Club', 'city': 'Vancouver', 'hood': 'West End', 'prio': 'P2', 'ig': '@onerunclubvan', 'aud': '7,950 IG', 'audsrc': 'Instagram @onerunclubvan follower count, web search 2026. About 1,536 active members on Heylo', 'source': 'Verified Web', 'notes': 'Wednesday 6:30pm run with a social wrap at English Bay.'},
 {'name': 'Striderz Run Club', 'city': 'Vancouver', 'hood': 'Olympic Village', 'prio': 'P2', 'ig': '@striderzrunclub', 'aud': '1,120 IG', 'audsrc': 'Instagram @striderzrunclub follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Free Sunday 8:30am social run from Athletes Way, all paces regroup.'},
 {'name': "North Shore Women's Trail Running Club", 'city': 'North Vancouver', 'hood': 'Lynn Valley', 'prio': 'P2', 'ig': '@northshore__runclub', 'aud': '2,930 IG', 'audsrc': 'Instagram @northshore__runclub follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Trail runs meeting at the End of the Line General Store.'},
 {'name': 'North Burnaby Runners', 'city': 'Burnaby', 'hood': 'North Burnaby', 'prio': 'P2', 'ig': '@northburnabyrunners', 'aud': '2,370 IG', 'audsrc': 'Instagram @northburnabyrunners follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Wednesday 6:30pm social run from Dageraad Brewing.'},
 {'name': 'Flight Crew Run Club', 'city': 'Vancouver', 'hood': 'Kitsilano', 'prio': 'P2', 'ig': '@flightcrewrunclub', 'aud': '4,426 IG', 'audsrc': 'Instagram @flightcrewrunclub follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Thursday 6:15pm Kitsilano road run with 3km, 5km and 10km options. Run by Vancouver Running Company.'},
 {'name': 'Capital City Run Crew', 'city': 'Victoria', 'hood': 'Downtown', 'prio': 'P2', 'ig': '@capitalcityruncrew', 'aud': '2,820 IG', 'audsrc': 'Instagram @capitalcityruncrew follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Thursday evening downtown run.'},
 {'name': 'Kelowna Running Club', 'city': 'Kelowna', 'hood': 'Waterfront Park', 'prio': 'P3', 'ig': '@kelownarunningclub', 'aud': '990 IG', 'audsrc': 'Instagram @kelownarunningclub follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Saturday long runs from rotating spots plus Tuesday speed work. 2026 schedule on the club site.'},
 {'name': 'Notorious Run Club', 'city': 'Victoria', 'hood': 'Inner Harbour', 'prio': 'P2', 'ig': '@notoriousrunclub', 'aud': '1,703 IG', 'audsrc': 'Instagram @notoriousrunclub follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Saturday 9am 5km social run downtown.'},
 {'name': 'NRG Kelowna', 'city': 'Kelowna', 'hood': 'Downtown', 'prio': 'P2', 'ig': '@nrgkelowna', 'aud': '4,785 IG', 'audsrc': 'Instagram @nrgkelowna follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Run club that meets at Red Bird Brewing.'},
 {'name': 'the girls vancouver', 'city': 'Vancouver', 'prio': 'P2', 'ig': '@thegirlsvancouver', 'aud': '1,807 IG', 'audsrc': 'Instagram @thegirlsvancouver follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Womens run and social community.'},
 {'name': 'Victoria Queer Run Club', 'city': 'Victoria', 'hood': 'Dallas Road', 'prio': 'P3', 'ig': '@vicqueerrunclub', 'aud': '1,357 IG', 'audsrc': 'Instagram @vicqueerrunclub follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Monday 6pm run, inclusive community.'},
],
'Gyms & Studios':[
 {'name': 'RIDE Cycle Club', 'city': 'Vancouver', 'hood': 'Yaletown', 'prio': 'P1', 'ig': '@ridecycleclub', 'aud': '23,000 IG', 'audsrc': 'Instagram @ridecycleclub follower count, web search 2026. Account covers Vancouver and Toronto', 'source': 'Verified Web', 'notes': 'Spin studio on Hamilton Street, part of a multi city brand.'},
 {'name': 'The Hive Bouldering Gym', 'city': 'Vancouver', 'hood': 'Strathcona', 'prio': 'P1', 'ig': '@hiveclimbing', 'aud': '21,000 IG', 'audsrc': 'Instagram @hiveclimbing follower count, web search 2026. Brand account, several BC sites', 'source': 'Verified Web', 'notes': 'Bouldering gym on Industrial Avenue, brand has several BC locations.'},
 {'name': 'Progression Bouldering', 'city': 'Vancouver', 'hood': 'Mount Pleasant', 'prio': 'P1', 'ig': '@progressionbouldering', 'aud': '9,700 IG', 'audsrc': 'Instagram @progressionbouldering follower count, web search 2026', 'source': 'Verified Web', 'notes': '18,000 sq ft bouldering gym with a licensed cafe on site.'},
 {'name': 'Tantra Fitness', 'city': 'Vancouver', 'hood': 'Kitsilano', 'prio': 'P2', 'ig': '@tantrafitness', 'aud': '17,000 IG', 'audsrc': 'Instagram @tantrafitness follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Pole and aerial fitness across three Vancouver locations.'},
 {'name': '604 Athletics', 'city': 'Vancouver', 'hood': 'Mount Pleasant', 'prio': 'P2', 'ig': '@604_athletics', 'aud': '6,148 IG', 'audsrc': 'Instagram @604_athletics follower count, web search 2026', 'source': 'Verified Web', 'notes': 'CrossFit box on Main Street, runs HYROX prep classes.'},
 {'name': 'FAR Studio', 'city': 'Vancouver', 'hood': 'Gastown', 'prio': 'P2', 'ig': '@farstudiogym', 'aud': '2,458 IG', 'audsrc': 'Instagram @farstudiogym follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Kickboxing and strength studio on Powell Street.'},
 {'name': 'The Lab Victoria', 'city': 'Victoria', 'hood': 'Downtown', 'prio': 'P2', 'ig': '@thelabvictoria', 'aud': '3,564 IG', 'audsrc': 'Instagram @thelabvictoria follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Yoga and pilates studio on Fort Street.'},
 {'name': 'CrossFit BC', 'city': 'Vancouver', 'hood': 'Olympic Village', 'prio': 'P3', 'ig': '@crossfitbc', 'aud': '6,164 IG', 'audsrc': 'Instagram @crossfitbc follower count, web search 2026', 'source': 'Verified Web', 'notes': 'CrossFit box on East 1st Avenue.'},
 {'name': 'CrossFit Zone', 'city': 'Victoria', 'hood': 'Downtown', 'prio': 'P3', 'ig': '@crossfit_zone_', 'aud': '2,250 IG', 'audsrc': 'Instagram @crossfit_zone_ follower count, web search 2026', 'source': 'Verified Web', 'notes': 'One of the oldest CrossFit boxes in Victoria, operating since 2008.'},
 {'name': 'CrossFit Okanagan', 'city': 'Kelowna', 'hood': 'Kelowna', 'prio': 'P3', 'ig': '@crossfitokanagan', 'aud': '4,711 IG', 'audsrc': 'Instagram @crossfitokanagan follower count, web search 2026', 'source': 'Verified Web', 'notes': 'CrossFit box, also runs a ForeverFit 55 plus program.'},
 {'name': 'Sweat Studios', 'city': 'Kelowna', 'hood': 'Downtown', 'prio': 'P2', 'ig': '@sweatkelowna', 'aud': '5,330 IG', 'audsrc': 'Instagram @sweatkelowna follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Multi format studio: barre, cycle, pilates, yoga and boot camp at 529 Lawrence Ave. Also home to BioShack recovery.'},
 {'name': 'Quantum Yoga and Pilates', 'city': 'Victoria', 'hood': 'Downtown', 'prio': 'P3', 'ig': '@quantumyogapilates', 'aud': '3,148 IG', 'audsrc': 'Instagram @quantumyogapilates follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Yoga and pilates studio downtown.'},
 {'name': 'F45 Training Downtown Victoria', 'city': 'Victoria', 'hood': 'Downtown', 'prio': 'P3', 'ig': '@f45_training_downtownvic', 'aud': '1,912 IG', 'audsrc': 'Instagram @f45_training_downtownvic follower count, web search 2026', 'source': 'Verified Web', 'notes': 'HIIT studio at 595 Pandora Ave.'},
],
'Events & Festivals':[
 {'name': 'Richmond Night Market', 'city': 'Richmond', 'hood': 'Bridgeport', 'prio': 'P1', 'ig': 'richmondnightmarket.com', 'aud': '1,000,000+ per year', 'audsrc': 'vancouversbestplaces and official market materials, over a million visitors annually', 'source': 'Verified Web', 'notes': 'Runs Apr 24 to Sep 20 2026, Friday to Sunday evenings, near Bridgeport station.'},
 {'name': 'Vancouver Pride Parade and Festival', 'city': 'Vancouver', 'hood': 'West End', 'prio': 'P1', 'ig': '@vancouverpride', 'aud': '600,000+ attendees, press estimate', 'audsrc': 'misterbandb and Destination Vancouver listings cite 600,000 plus, press estimate not official', 'source': 'Verified Web', 'notes': 'Pride week Jul 25 to Aug 2 2026, main parade Sunday Aug 2 ending at Sunset Beach.'},
 {'name': 'Concord Pacific Dragon Boat Summer Regatta', 'city': 'Vancouver', 'hood': 'False Creek', 'prio': 'P1', 'ig': 'dragonboatbc.ca', 'aud': 'TBD', 'audsrc': '', 'source': 'Verified Web', 'notes': 'One day regatta Aug 22 2026 at False Creek. Downsized for 2026 due to the FIFA security cordon, full festival returns 2027.'},
 {'name': 'Italian Day on The Drive', 'city': 'Vancouver', 'hood': 'Commercial Drive', 'prio': 'P2', 'ig': '@italiandayonthedrive', 'aud': '300,000 attendees, press estimate', 'audsrc': 'Vancouver Is Awesome 2026 coverage cites 300,000, press estimate not official', 'source': 'Verified Web', 'notes': 'Sunday Jun 14 2026, noon to 8pm, 14 blocks of Commercial Drive.'},
 {'name': 'Khatsahlano Street Party', 'city': 'Vancouver', 'hood': 'Kitsilano', 'prio': 'P2', 'ig': '@khatsahlano', 'aud': '200,000 attendees', 'audsrc': 'CTV News and vancouversbestplaces, up to 200,000 at the 2025 edition', 'source': 'Verified Web', 'notes': 'Saturday Jul 11 2026 on West 4th Ave, 10 blocks, free street festival.'},
 {'name': 'Summer Lights in English Bay', 'city': 'Vancouver', 'hood': 'West End', 'prio': 'P2', 'ig': 'vancouver.ca', 'aud': 'TBD', 'audsrc': '', 'source': 'Verified Web', 'notes': 'Friday Jul 31 2026 fireworks at English Bay, BC Day long weekend. City replacement for Celebration of Light.'},
 {'name': 'Canada Dry Victoria Dragon Boat Festival', 'city': 'Victoria', 'hood': 'Inner Harbour', 'prio': 'P2', 'ig': 'victoriadragonboatfestival.com', 'aud': '30 plus teams', 'audsrc': 'Westerly News and Parksville Qualicum News Jun 2026, more than 30 teams', 'source': 'Verified Web', 'notes': 'Saturday Jun 20 2026 at the Inner Harbour, more than 30 teams. Moved to June for 2026 only, back to August in 2027.'},
 {'name': 'Kelowna Wine Country Half Marathon', 'city': 'Kelowna', 'hood': 'Waterfront Park', 'prio': 'P2', 'ig': 'kelownamarathon.ca', 'aud': '1,800 runners, capacity', 'audsrc': 'Event materials note registration normally limited to about 1,800 runners, confirm on site', 'source': 'Verified Web', 'notes': 'Saturday Jun 13 2026, finish at Waterfront Park with a post run festival.'},
 {'name': 'Harmony Arts Festival', 'city': 'West Vancouver', 'hood': 'Ambleside', 'prio': 'P1', 'ig': 'harmonyarts.ca', 'aud': '140,000 attendees', 'audsrc': 'District of West Vancouver festival page, as many as 140,000 attend', 'source': 'Verified Web', 'notes': 'Runs Jul 31 to Aug 9 2026 on the Ambleside waterfront.'},
 {'name': 'Parks Alive Kelowna', 'city': 'Kelowna', 'hood': 'City Park', 'prio': 'P2', 'ig': 'festivalskelowna.com', 'aud': '24,000 plus attendees', 'audsrc': 'Castanet and Festivals Kelowna, over 24,000 guests annually', 'source': 'Verified Web', 'notes': 'Free music and events series by Festivals Kelowna, opens Jul 3 2026, runs nine weeks.'},
 {'name': 'Victoria Pride Festival', 'city': 'Victoria', 'hood': 'James Bay', 'prio': 'P2', 'ig': 'victoriapridesociety.org', 'aud': '10,000 festival, 80,000 spectators', 'audsrc': 'Tourism Victoria, upwards of 10,000 at the festival and 80,000 plus along the parade', 'source': 'Verified Web', 'notes': 'Festival in the Park at MacDonald Park Jul 12 2026.'},
 {'name': 'RBC GranFondo Whistler', 'city': 'Vancouver', 'hood': 'Sea to Sky', 'prio': 'P2', 'ig': 'rbcgranfondo.com', 'aud': '5,000 cyclists', 'audsrc': 'event listings, about 5,000 cyclists ride the 122km route', 'source': 'Verified Web', 'notes': '122km ride from downtown Vancouver to Whistler Sep 12 2026.'},
 {'name': 'Kelowna International Dragon Boat Festival', 'city': 'Kelowna', 'hood': 'Tugboat Bay', 'prio': 'P3', 'ig': 'kelownadragonboatclub.com', 'aud': 'TBD', 'audsrc': '', 'source': 'Verified Web', 'notes': 'Jul 11 2026 at Tugboat Bay, supports the Central Okanagan Food Bank.'},
],
'Sports Teams & Leagues':[
 {'name': 'Urban Rec Vancouver', 'city': 'Vancouver', 'hood': 'Mount Pleasant', 'prio': 'P1', 'ig': '@urbanrec', 'aud': '56,000 members, self stated', 'audsrc': 'Urban Rec Vancouver website states over 56,000 members, brand self stated figure', 'source': 'Verified Web', 'notes': 'Largest sport and social club in Western Canada. Leagues across Vancouver, Richmond and Burnaby.'},
 {'name': 'Vancouver Dodgeball League', 'city': 'Vancouver', 'hood': 'Multiple', 'prio': 'P2', 'ig': '@vdldodgeball', 'aud': '2,000+ players, 260+ teams', 'audsrc': 'VDL sources describe over 260 teams and 2,000 plus players at peak', 'source': 'Verified Web', 'notes': 'Non profit dodgeball league playing in school and community gyms.'},
 {'name': 'Vancouver Pickleball Association', 'city': 'Vancouver', 'hood': 'Multiple', 'prio': 'P2', 'ig': '@vancouverpickleballassociation', 'aud': '1,065 IG', 'audsrc': 'Instagram @vancouverpickleballassociation follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Box league runs Apr to Sep 2026 across community centres and outdoor courts.'},
 {'name': 'False Creek Racing Canoe Club', 'city': 'Vancouver', 'hood': 'Granville Island', 'prio': 'P2', 'ig': '@falsecreekcanoeclub', 'aud': '956 IG', 'audsrc': 'Instagram @falsecreekcanoeclub follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Dragon boat, outrigger and sprint paddling club based on Granville Island.'},
 {'name': 'Victoria Sport and Social Club', 'city': 'Victoria', 'hood': 'Multiple', 'prio': 'P2', 'ig': '@vicsportnsocial', 'aud': 'TBD', 'audsrc': '', 'source': 'Verified Web', 'notes': 'Mixed adult leagues, 9 plus sports running for summer 2026.'},
 {'name': 'Pickleball Kelowna Club', 'city': 'Kelowna', 'hood': 'Parkinson Rec Centre', 'prio': 'P2', 'ig': '@pickleballkelownaclub', 'aud': '700 members', 'audsrc': 'Instagram @pickleballkelownaclub states the club is home to 700 members', 'source': 'Verified Web', 'notes': 'Plays on 12 fenced outdoor courts, May to Sep season, hosts the Kelowna Open.'},
 {'name': 'Urban Rec Victoria', 'city': 'Victoria', 'hood': 'Multiple', 'prio': 'P2', 'ig': 'victoria.urbanrec.ca', 'aud': 'TBD', 'audsrc': '', 'source': 'Verified Web', 'notes': 'Ten coed adult leagues across Greater Victoria, year round.'},
 {'name': 'Kelowna Rowing Club', 'city': 'Kelowna', 'hood': 'Waterfront Park', 'prio': 'P3', 'ig': '@kelownarowingclub', 'aud': '70 plus rowers', 'audsrc': 'kelownarowing.com, membership over 70 rowers', 'source': 'Verified Web', 'notes': 'Rows on Okanagan Lake, learn to row and competitive crews.'},
],
'Campus & Student Groups':[
 {'name': 'UBC Thunderbirds Athletics', 'city': 'Vancouver', 'hood': 'UBC', 'prio': 'P1', 'ig': '@ubctbirds', 'aud': '33,000 IG', 'audsrc': 'Instagram @ubctbirds follower count, web search 2026. 26 varsity teams per gothunderbirds.ca', 'source': 'Verified Web', 'notes': '26 varsity teams across 15 sports at the Point Grey campus.'},
 {'name': 'AMS of UBC', 'city': 'Vancouver', 'hood': 'UBC', 'prio': 'P1', 'ig': 'ams.ubc.ca', 'aud': '60,000 students', 'audsrc': 'AMS represents more than 60,000 students per ams.ubc.ca, operates 200 plus clubs', 'source': 'Verified Web', 'notes': 'Student society running the Nest student union building and 200 plus clubs.'},
 {'name': 'SFU Recreation', 'city': 'Burnaby', 'hood': 'Burnaby Mountain', 'prio': 'P2', 'ig': '@sfurecreation', 'aud': '4,500 IG', 'audsrc': 'Instagram @sfurecreation follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Intramural leagues open to students, staff and faculty.'},
 {'name': 'UVic Vikes Recreation', 'city': 'Victoria', 'hood': 'UVic', 'prio': 'P2', 'ig': 'vikesrec.ca', 'aud': '22,000 students', 'audsrc': 'UVic enrolment over 22,000 students per uvic.ca, Vikes Recreation open to all students', 'source': 'Verified Web', 'notes': 'Campus recreation programs based at the CARSA centre.'},
 {'name': "Students' Union Okanagan of UBC", 'city': 'Kelowna', 'hood': 'UBC Okanagan', 'prio': 'P2', 'ig': '@suo_ubc', 'aud': '12,000 students', 'audsrc': 'SUO represents over 12,000 students per suo.ca', 'source': 'Verified Web', 'notes': 'Runs The Well student pub and The Green Bean cafe. New recreation facility broke ground May 2026.'},
 {'name': 'Camosun College Student Society', 'city': 'Victoria', 'hood': 'Lansdowne', 'prio': 'P3', 'ig': 'camosunstudent.org', 'aud': '9,000+ students', 'audsrc': 'CCSS represents the 9,000 plus students of Camosun College per camosunstudent.org', 'source': 'Verified Web', 'notes': 'Student society across the Lansdowne and Interurban campuses.'},
 {'name': 'Simon Fraser Student Society', 'city': 'Burnaby', 'hood': 'Burnaby Mountain', 'prio': 'P2', 'ig': 'sfss.ca', 'aud': 'TBD', 'audsrc': '', 'source': 'Verified Web', 'notes': 'Represents the SFU undergraduate body, runs the Student Union Building.'},
 {'name': 'Okanagan College Students Union', 'city': 'Kelowna', 'prio': 'P2', 'ig': '@ocsu.kelowna', 'aud': '5,000 plus students', 'audsrc': 'ocsu.ca, membership of over 5,000 Okanagan College students', 'source': 'Verified Web', 'notes': 'Student union at the Kelowna campus, 1000 KLO Road.'},
 {'name': 'BCIT Student Association', 'city': 'Burnaby', 'prio': 'P3', 'ig': '@bcitsa', 'aud': 'TBD', 'audsrc': '', 'source': 'Verified Web', 'notes': 'Serves all BCIT students, Burnaby campus at 3700 Willingdon Ave.'},
],
'Wellness & Recovery':[
 {'name': 'Tevah Wellness', 'city': 'Vancouver', 'hood': 'Yaletown', 'prio': 'P1', 'source': 'Verified Web', 'inbc': 'Yes', 'aud': '7,690 IG', 'audsrc': 'Instagram @tevahwellness follower count, web search 2026. Address 955 Pacific Blvd', 'notes': 'Imported from the BC Tracker Community tab so it is tracked once. Sauna and recovery. Address 955 Pacific Blvd.', 'ig': '@tevahwellness'},
 {'name': 'HAVN Saunas', 'city': 'Victoria', 'hood': 'Inner Harbour', 'prio': 'P1', 'ig': '@havn.saunas', 'aud': '34,000 IG', 'audsrc': 'Instagram @havn.saunas follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Floating sauna barge on the Inner Harbour with cold plunge circuits.'},
 {'name': 'Kolm Kontrast Nordic Spa', 'city': 'Vancouver', 'hood': 'Cambie Village', 'prio': 'P1', 'ig': '@kolmkontrast', 'aud': '12,000 IG', 'audsrc': 'Instagram @kolmkontrast follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Nordic spa at 525 W 8th Ave with heat and cold cycles.'},
 {'name': 'RITUAL Nordic Spa', 'city': 'Victoria', 'hood': 'Harris Green', 'prio': 'P2', 'ig': '@ritualnordicspa', 'aud': '9,400 IG', 'audsrc': 'Instagram @ritualnordicspa follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Nordic spa with an onsite cafe and salt lounge.'},
 {'name': 'Float House Vancouver', 'city': 'Vancouver', 'hood': 'Gastown', 'prio': 'P2', 'ig': '@float_house', 'aud': '9,000 IG', 'audsrc': 'Instagram @float_house follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Float and cold plunge studio at 70 W Cordova St.'},
 {'name': 'Regen Recovery', 'city': 'Vancouver', 'hood': 'Downtown', 'prio': 'P2', 'ig': '@regenrecovery', 'aud': '3,900 IG', 'audsrc': 'Instagram @regenrecovery follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Recovery lounge with sauna, cold plunge and IV services.'},
 {'name': 'BioShack', 'city': 'Kelowna', 'hood': 'Kelowna', 'prio': 'P3', 'ig': '@bioshack.kelowna', 'aud': 'TBD', 'audsrc': '', 'source': 'Verified Web', 'notes': 'Self led contrast therapy suite at Sweat Studios. Brand account about 2,549 followers, the Kelowna only count was not retrievable.'},
 {'name': 'Loyly Floating Sauna Kelowna', 'city': 'Kelowna', 'hood': 'City Park', 'prio': 'P1', 'ig': '@loyly.kelowna', 'aud': '14,000 IG', 'audsrc': 'Instagram @loyly.kelowna follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Floating sauna and cold plunge on Okanagan Lake at the downtown marina.'},
],
'Ambassadors & Creators':[
 {'name': 'Angela Liguori', 'city': 'Vancouver', 'prio': 'P1', 'ig': '@angelaliggs', 'aud': '908,000 IG', 'audsrc': 'Instagram @angelaliggs follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Outdoor and hiking creator covering BC trails.'},
 {'name': 'Bailey Campbell', 'city': 'Kelowna', 'prio': 'P1', 'ig': '@basicswithbails', 'aud': '641,000 IG', 'audsrc': 'Instagram @basicswithbails follower count, web search 2026. A tracker listed 394,700, reverify', 'source': 'Verified Web', 'notes': 'Kelowna Foodie. Recipe and lifestyle content.'},
 {'name': 'Twin Coast', 'city': 'Vancouver', 'prio': 'P1', 'ig': '@twincoast', 'aud': '537,000 IG', 'audsrc': 'Instagram @twincoast follower count, web search 2026. Also large on TikTok and YouTube', 'source': 'Verified Web', 'notes': 'Twin sisters, plant based recipe creators with a cookbook brand.'},
 {'name': 'Vancouver Dietitians', 'city': 'Vancouver', 'hood': 'South Granville', 'prio': 'P2', 'ig': '@vancouverdietitians', 'aud': '64,000 IG', 'audsrc': 'Instagram @vancouverdietitians follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Registered dietitian duo with a South Granville clinic.'},
 {'name': 'Cam Lee', 'city': 'Vancouver', 'prio': 'P2', 'ig': '@camleeyoga', 'aud': '47,000 IG', 'audsrc': 'Instagram @camleeyoga follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Yoga and wellness creator.'},
 {'name': 'Caroline Doucet', 'city': 'Vancouver', 'prio': 'P2', 'ig': '@nourishedbycaroline', 'aud': '38,000 IG', 'audsrc': 'Instagram @nourishedbycaroline follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Registered dietitian, plant based recipes, non diet approach.'},
 {'name': 'Trudy Leung', 'city': 'Vancouver', 'prio': 'P2', 'ig': '@missvancityfoodie', 'aud': '37,000 IG', 'audsrc': 'Instagram @missvancityfoodie follower count, web search 2026', 'source': 'Verified Web', 'notes': 'Miss Vancity Foodie. Local food and lifestyle content.'},
 {'name': 'Hilary Ann Yang', 'city': 'Vancouver', 'prio': 'P2', 'ig': '@thehilaryann', 'aud': '31,000 IG', 'audsrc': 'Instagram @thehilaryann follower count, web search 2026', 'source': 'Verified Web', 'notes': "Trail running creator tied to Squamish, runs a women's trail running community."},
 {'name': 'The Official Vancouver Fitness', 'city': 'Vancouver', 'prio': 'P3', 'ig': '@vancouver.fitness', 'aud': '24,000 IG', 'audsrc': 'Instagram @vancouver.fitness publicly visible follower count, web search 2026', 'source': 'Pending', 'notes': 'Vancouver fitness community account. Reconfirm the live follower count before outreach.'},
],
'Charity & Causes':[
 {'name': "BC Children's Hospital Foundation", 'city': 'Vancouver', 'hood': 'Queen Elizabeth Park', 'prio': 'P1', 'ig': 'fundraise.bcchf.ca', 'aud': '10,000 runners target 2026', 'audsrc': 'runguides and BCCHF pages, 2026 aims to welcome more than 10,000 runners', 'source': 'Verified Web', 'notes': 'Hosts RBC Race for the Kids, Jun 8 2026 at Queen Elizabeth Park. 5k plus 2k fun run and a family carnival.'},
 {'name': 'Greater Vancouver Food Bank', 'city': 'Burnaby', 'hood': 'Swangard Stadium', 'prio': 'P2', 'ig': 'foodbank.bc.ca', 'aud': 'TBD', 'audsrc': '', 'source': 'Verified Web', 'notes': 'Hosts Foodstock, Jun 23 2026 at Swangard Stadium. A 19 plus outdoor music and food festival.'},
 {'name': 'Backpack Buddies', 'city': 'Vancouver', 'hood': 'Multiple', 'prio': 'P2', 'ig': '@backpackbuddiesbc', 'aud': '4,347 IG, 6,800 kids weekly', 'audsrc': 'Web search, 4,347 Instagram followers and over 6,800 kids reached weekly, 1.8 million meals in 2025 and 2026', 'source': 'Verified Web', 'notes': 'Hosts the Birdies and Buddies golf tournament Jun 8 2026. School food programs across 83 BC communities.'},
 {'name': 'Victoria Hospice', 'city': 'Victoria', 'hood': 'Oak Bay', 'prio': 'P3', 'ig': 'victoriahospice.org', 'aud': 'TBD', 'audsrc': '', 'source': 'Verified Web', 'notes': 'Hosts Hike for Hospice, May 3 2026 at Willows Beach. A Goddess Run charity of choice.'},
 {'name': 'Wild One Run for Youth Mental Health', 'city': 'Kelowna', 'hood': 'Wilden', 'prio': 'P3', 'ig': 'wildonerun.ca', 'aud': 'TBD', 'audsrc': '', 'source': 'Verified Web', 'notes': 'Trail run and walk Oct 3 2026 on the Wilden trails. Proceeds to Foundry Kelowna.'},
 {'name': "Ryder Hesjedal's Tour de Victoria", 'city': 'Victoria', 'hood': 'Downtown', 'prio': 'P2', 'ig': 'tourdevictoria.com', 'aud': 'TBD', 'audsrc': '', 'source': 'Verified Web', 'notes': 'Mass participation cycling event Aug 15 2026, eight distances, 15th annual.'},
],
}

# =====================================================================
# LOOKUPS DATA
# =====================================================================
OWNERS = ["Maddie"]
PRIORITY = ["P1","P2","P3"]
STATUS = ["Open","Outreach Sent","In Conversation","Proposal Sent","Agreed","Activated",
"Repeat Partner","Lost","On Hold"]
WARM = ["Warm","Cold","Past Contact","Unknown"]
SOURCE = ["Verified Web","Verified Phone","Verified In Person","Pending","Unverified"]
ACTIVATION = ["Sampling","Event Booth","Sponsorship","Ambassador","Contra Product",
"Co Branded Content","Hydration Station"]
WANT = ["Free Product","Sampling Budget","Sponsorship Fee","Social Posts From Us",
"Co Branded Giveaway","Discount Code","Long Term Deal"]
RISKS = ["Competitor Locked In","Audience Mismatch","High Fee","Slow To Respond",
"One Off Only","Logistics"]
INBC = ["Yes","No"]
SKU = ["Requested","Sampled","Stocked"]
ROLE = ["Owner or Founder","Director","Manager","General Manager","Marketing Lead",
"Community Manager","Events Lead","Coach","Run Lead","Studio Owner","Coordinator",
"President","Board Member","Sponsorship Lead","Athletic Director","Creator"]
CITY = ["Vancouver","Burnaby","North Vancouver","West Vancouver","Richmond","Surrey",
"Langley","Coquitlam","New Westminster","Victoria","Kelowna","Nanaimo","Online"]
HOOD = ["Cambie Village","Coal Harbour","Commercial Drive","Downtown","Dunbar","East Vancouver",
"Fairview","False Creek","Gastown","Granville Island","Harris Green","Inner Harbour",
"Jericho Beach","Kitsilano","Lansdowne","Lower Lonsdale","Lynn Valley","Main Street",
"Mount Pleasant","Multiple","North Burnaby","Oak Bay","Olympic Village","Parkinson Rec Centre",
"Queen Elizabeth Park","South Granville","Strathcona","Swangard Stadium","UBC","UBC Okanagan",
"UVic","Bridgeport","Burnaby Mountain","Waterfront Park","West End","Wilden","Willows Beach","Yaletown","Ambleside","City Park","Dallas Road","James Bay","Sea to Sky","Tugboat Bay"]
EVENTSTATUS = ["Idea","Scoping","Approved","Moved To Events Tab","Passed"]
ACTSTATUS = ["Idea","Pitched","Booked","Confirmed","Delivered","Cancelled"]
ASSETS = ["Samples","Branded Cooler","Table","Tent","Banner","Signage","Staff","Ice","Swag","Permit","Demo Kit"]
BUDGETTIER = [0,250,500,1000,2500,5000]

PRANK = {"P1":0,"P2":1,"P3":2}
for _t in PARTNERS:
    PARTNERS[_t] = sorted(PARTNERS[_t], key=lambda d: PRANK.get(d.get("prio","P3"),3))
PARTNER_NAMES = sorted({d["name"] for _t in PARTNERS for d in PARTNERS[_t]})

LU_COLS = [   # (header, list, defined_name)
 ("Owner", OWNERS, "LU_Owner"),
 ("Partnership Type", TYPES, "LU_Type"),
 ("Priority", PRIORITY, "LU_Priority"),
 ("Status", STATUS, "LU_Status"),
 ("Warm", WARM, "LU_Warm"),
 ("Source", SOURCE, "LU_Source"),
 ("Activation Type", ACTIVATION, "LU_Activation"),
 ("What They Want", WANT, "LU_Want"),
 ("Risks", RISKS, "LU_Risks"),
 ("In BC Tracker", INBC, "LU_InBC"),
 ("SKU Status", SKU, "LU_SKU"),
 ("Role", ROLE, "LU_Role"),
 ("City", CITY, "LU_City"),
 ("Neighbourhood", HOOD, "LU_Neighbourhood"),
 ("Event Status", EVENTSTATUS, "LU_EventStatus"),
 ("Activation Status", ACTSTATUS, "LU_ActStatus"),
 ("Assets Needed", ASSETS, "LU_Assets"),
 ("Budget Tier", BUDGETTIER, "LU_BudgetTier"),
 ("Partner", PARTNER_NAMES, "LU_Partners"),
]

# =====================================================================
# helpers for styling
# =====================================================================
def title_row(ws, text, span_last):
    ws.merge_cells(f"A1:{span_last}1")
    c = ws.cell(1,1,text)
    c.font = font(22, True, WHITE); c.fill = fill(C_TITLE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 40

def subtitle(ws, row, text):
    c = ws.cell(row,1,text); c.font = font(11, False, C_SUB); c.alignment = A_L
    ws.row_dimensions[row].height = 22

def hcell(ws, row, col, text):
    c = ws.cell(row,col,text); c.font = font(11, True, WHITE); c.fill = fill(C_HEADER)
    c.alignment = A_C; c.border = HEADBORD; return c

BAND = "FFF6F9F7"   # whisper of green — kept for reference, no longer applied to rows
def band(ws, first, last, ncol):
    pass  # v10: removed alternating fills for a cleaner, more premium look

def printsetup(ws, title_rows="1:2"):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    if title_rows: ws.print_title_rows = title_rows

def linkcell(cell, value):
    tgt = ("https://instagram.com/"+value[1:]) if value.startswith("@") else (value if value.startswith("http") else "https://"+value)
    cell.hyperlink = tgt
    cell.font = Font(name=FONTNAME, size=12, color=ACCENT, underline="single")

print("building lookups...")
# ---------------- LOOKUPS ----------------
lu = wb.active; lu.title = "Lookups"
title_row(lu, "Lookups", "S")
for i,(hdr,vals,nm) in enumerate(LU_COLS):
    col = i+1
    hc = lu.cell(2,col,hdr); hc.font = font(12, True, WHITE); hc.fill = fill(C_HEADER)
    hc.alignment = A_C; hc.border = BORD
    lu.column_dimensions[get_column_letter(col)].width = 22
    for j,v in enumerate(vals):
        cc = lu.cell(3+j,col,v); cc.font = font(11, False, C_DATA); cc.alignment = A_L
    # defined name (dynamic OFFSET range)
    L = get_column_letter(col)
    dn = DefinedName(nm, attr_text=f"OFFSET(Lookups!${L}$3,0,0,MAX(1,COUNTA(Lookups!${L}$3:${L}$1000)),1)")
    wb.defined_names[nm] = dn
lu.freeze_panes = "A3"; lu.row_dimensions[2].height = 30
lu.sheet_properties.tabColor = "B7C9C1"

# =====================================================================
# TYPE TABS
# =====================================================================
def add_validations(ws):
    def mk(name):
        dv = DataValidation(type="list", formula1=name, allow_blank=True, showErrorMessage=False); ws.add_data_validation(dv); return dv
    rng = lambda col: f"{col}{FIRST}:{col}{LASTROW}"
    mk("LU_Type").add(rng("C"))
    mk("LU_City").add(rng("D"))
    mk("LU_Neighbourhood").add(rng("E"))
    mk("LU_Priority").add(rng("F"))
    mk("LU_Status").add(rng("G"))
    mk("LU_Owner").add(rng("H"))
    mk("LU_Role").add(rng("J"))
    mk("LU_Source").add(rng("P"))
    mk("LU_Warm").add(rng("Q"))
    mk("LU_Activation").add(rng("V"))
    mk("LU_Want").add(rng("AE"))
    mk("LU_Risks").add(rng("AF"))
    mk("LU_InBC").add(rng("AH"))
    sku = mk("LU_SKU"); sku.add(f"AI{FIRST}:AK{LASTROW}")

def add_cond_formats(ws):
    cf = ws.conditional_formatting.add
    # Status colours G
    smap = [("Activated",F_GREEN),("Repeat Partner",F_GREEN),("Agreed",F_AMBER),
            ("Proposal Sent",F_AMBER),("In Conversation",F_AMBER),("Outreach Sent",F_PALEBLU),
            ("Open",F_GREY),("Lost",F_RED),("On Hold",F_SLATE)]
    for val,clr in smap:
        cf(f"G{FIRST}:G{LASTROW}", CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill(clr)))
    # Priority F
    for val,clr in [("P1",F_P1),("P2",F_P2),("P3",F_P3)]:
        cf(f"F{FIRST}:F{LASTROW}", CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill(clr)))
    # Warm Q
    cf(f"Q{FIRST}:Q{LASTROW}", CellIsRule(operator="equal", formula=['"Warm"'], fill=fill(F_AMBER)))
    # Source P
    for val,clr in [("Verified Web",F_GREEN),("Verified Phone",F_GREEN),("Verified In Person",F_GREEN),
                    ("Pending",F_AMBER),("Unverified",F_P3)]:
        cf(f"P{FIRST}:P{LASTROW}", CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill(clr)))
    # Days since activity S : red over 14 for any row, amber over 7 on P1 rows
    cf(f"S{FIRST}:S{LASTROW}", CellIsRule(operator="greaterThan", formula=["14"], fill=fill(F_RED)))
    cf(f"S{FIRST}:S{LASTROW}", FormulaRule(formula=[f'AND($F{FIRST}="P1",$S{FIRST}>7)'], fill=fill(F_AMBER)))
    # Next Action Date past due U
    cf(f"U{FIRST}:U{LASTROW}", FormulaRule(formula=[f'AND($U{FIRST}<>"",$U{FIRST}<TODAY())'], fill=fill(F_RED)))
    # SKU AI:AK
    for val,clr in [("Stocked",F_GREEN),("Sampled",F_BLUE),("Requested",F_PALEBLU)]:
        cf(f"AI{FIRST}:AK{LASTROW}", CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill(clr)))

def build_type_tab(tname):
    ws = wb.create_sheet(tname)
    title_row(ws, tname, LASTCOL)
    # headers
    for c,h in enumerate(HEADERS, start=1):
        hcell(ws, 2, c, h)
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS[c]
    ws.row_dimensions[2].height = 52
    rows = PARTNERS.get(tname, [])
    for i in range(DATA_ROWS):
        r = FIRST + i
        ws.row_dimensions[r].height = 26
        # default styling for every cell in row
        for c in range(1, NCOL+1):
            cell = ws.cell(r,c)
            cell.font = font(); cell.border = BORD
            cell.alignment = A_CL if c in CENTERCOLS else A_L
        d = rows[i] if i < len(rows) else None
        if d:
            for k,col in KEYCOL.items():
                if k in d and d[k] not in (None,""):
                    ws.cell(r,col, d[k])
            # defaults for researched rows
            if "status" not in d: ws.cell(r,7,"Open")
            if "owner" not in d: ws.cell(r,8,"Maddie")
            if "warm" not in d: ws.cell(r,17,"Cold")
            if "inbc" not in d: ws.cell(r,34,"No")
            if d.get("ig"): linkcell(ws.cell(r,13), d["ig"])   # clickable Instagram or website
        # live formulas on every row so new entries behave the same
        ws.cell(r,1, f'=IF($B{r}="","",ROW()-2)')                                   # auto number
        ws.cell(r,3, f'=IF($B{r}="","","{tname}")')                                 # auto type label
        ws.cell(r,19, f'=IF(ISBLANK(R{r}),"",TODAY()-R{r})')                        # Days Since Activity
        ws.cell(r,27, f'=IF(OR($Z{r}="",$Y{r}="",$Y{r}=0),"",$Z{r}/($Y{r}*24))')    # Cost Per Can
        # number formats
        for c in DATECOLS: ws.cell(r,c).number_format = FMT_DATE
        for c in MONEYCOLS: ws.cell(r,c).number_format = FMT_MONEY
        ws.cell(r,27).number_format = FMT_CENTS
        ws.cell(r,19).number_format = FMT_INT
    band(ws, FIRST, LASTROW, NCOL)
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{LASTCOL}{LASTROW}"
    ws.sheet_properties.tabColor = "4F8A78"
    printsetup(ws)
    add_validations(ws)
    add_cond_formats(ws)
    ws.column_dimensions.group('V','AG', hidden=True)
    return ws

print("building type tabs...")
for t in TYPES:
    build_type_tab(t)

# =====================================================================
# MASTER LIST  (read only, auto pulls every partner from every type tab)
# =====================================================================
print("building master list...")
SRCCOLS = [c for c in range(2, NCOL+1) if c != 3]   # all type-tab cols except # and Partnership Type
ML_LAST = get_column_letter(1 + len(SRCCOLS))        # AK (37 cols)
SRCFMT = {18:FMT_DATE,21:FMT_DATE,23:FMT_DATE,26:FMT_MONEY,28:FMT_MONEY,27:FMT_CENTS,19:FMT_INT}
def ml_letter(src):  # master-list column letter for a given source col
    return get_column_letter(2 + SRCCOLS.index(src))
ML_ACT = ml_letter(23)  # Activation Date column on Master List

ml = wb.create_sheet("Master List")
title_row(ml, "Master List", ML_LAST)
ml.cell(2,1,"Partnership Type"); h = ml.cell(2,1)
h.font = font(12, True, WHITE); h.fill = fill(C_HEADER); h.alignment = A_C; h.border = BORD
ml.column_dimensions["A"].width = 20
for m,src in enumerate(SRCCOLS, start=2):
    hcell(ml, 2, m, HEADERS[src-1])
    ml.column_dimensions[get_column_letter(m)].width = WIDTHS[src]
ml.row_dimensions[2].height = 52
mr = 3
for t in TYPES:
    for tr in range(FIRST, LASTROW+1):
        ml.row_dimensions[mr].height = 24
        a = ml.cell(mr,1, "=IF('{T}'!B{r}=\"\",\"\",\"{T}\")".format(T=t, r=tr))
        a.font = font(); a.border = BORD; a.alignment = A_L
        for m,src in enumerate(SRCCOLS, start=2):
            L = get_column_letter(src)
            cell = ml.cell(mr,m, "=IF('{T}'!{L}{r}=\"\",\"\",'{T}'!{L}{r})".format(T=t, L=L, r=tr))
            cell.font = font(); cell.border = BORD
            cell.alignment = A_CL if src in CENTERCOLS else A_L
            if src in SRCFMT: cell.number_format = SRCFMT[src]
        ml.cell(mr,38, '=IF(AND($A{r}<>"",$F{r}<>"Lost",$F{r}<>"On Hold",$F{r}<>"Activated",$F{r}<>"Repeat Partner"),IF(ISNUMBER($T{r}),$T{r}+ROW()/1000000,90000+IF($E{r}="P1",1,IF($E{r}="P2",2,3))*10+ROW()/1000000),"")'.format(r=mr)).font=font(9,color="FFB7C9C1")
        mr += 1
ML_ROWS_END = mr-1
band(ml, 3, ML_ROWS_END, 1 + len(SRCCOLS))
ml.freeze_panes = "C3"; ml.auto_filter.ref = "A2:{}{}".format(ML_LAST, ML_ROWS_END)
ml.column_dimensions[get_column_letter(38)].hidden = True
ml.sheet_properties.tabColor = "6FA392"; printsetup(ml)
MLR = "'Master List'!"   # shorthand for formulas, ranges $3:$500

# =====================================================================
# ACTIVATIONS  (manual log, one row per activation, menus allow free entry)
# =====================================================================
print("building activations...")
act = wb.create_sheet("Activations")
ACT_H = ["#","Date","Partner","Partnership Type","City","Owner","Status","Activation Type",
"Assets Needed","Budget","Raspberry 4338 Cans","Lemon Lime 4336 Cans","Pineapple Passion Fruit 4340 Cans",
"Total Cans","Cost Per Can","Notes"]
ACT_LASTCOL = get_column_letter(len(ACT_H))   # O
title_row(act, "Activations", ACT_LASTCOL)
subtitle(act, 2, "Log each activation here, one row each. Pick from a menu or type your own. Click the Date filter to sort soonest first.")
actw={1:5,2:13,3:30,4:20,5:14,6:14,7:14,8:18,9:26,10:12,11:13,12:13,13:20,14:12,15:13,16:60}
for c,htext in enumerate(ACT_H, start=1):
    hcell(act,2,c,htext); act.column_dimensions[get_column_letter(c)].width=actw[c]
act.row_dimensions[2].height=52
ACT_FIRST=3; ACT_LAST=62
ACT_CENTER={1,2,5,6,7,8,10,11,12,13,14,15}
for i in range(ACT_LAST-ACT_FIRST+1):
    r=ACT_FIRST+i
    act.row_dimensions[r].height=26
    for c in range(1,len(ACT_H)+1):
        cell=act.cell(r,c); cell.font=font(); cell.border=BORD
        cell.alignment=A_CL if c in ACT_CENTER else A_L
    act.cell(r,1, f'=IF($C{r}="","",ROW()-2)')                          # auto number
    act.cell(r,14, f'=IF(COUNT($K{r}:$M{r})=0,"",SUM($K{r}:$M{r}))')    # Total Cans
    act.cell(r,15, f'=IF(OR($J{r}="",$N{r}="",$N{r}=0),"",$J{r}/$N{r})')   # Cost Per Can
    act.cell(r,2).number_format=FMT_DATE
    act.cell(r,10).number_format=FMT_MONEY
    for c in (11,12,13,14): act.cell(r,c).number_format=FMT_INT
    act.cell(r,15).number_format=FMT_CENTS
def actdv(name,col):
    dv=DataValidation(type="list",formula1=name,allow_blank=True,showErrorMessage=False)
    act.add_data_validation(dv); dv.add(f"{col}{ACT_FIRST}:{col}{ACT_LAST}")
actdv("LU_Type","D"); actdv("LU_City","E"); actdv("LU_Owner","F"); actdv("LU_ActStatus","G")
actdv("LU_Partners","C"); actdv("LU_Activation","H"); actdv("LU_Assets","I"); actdv("LU_BudgetTier","J")
acf=act.conditional_formatting.add
for val,clr in [("Idea",F_GREY),("Pitched",F_PALEBLU),("Booked",F_AMBER),("Confirmed",F_BLUE),("Delivered",F_GREEN),("Cancelled",F_RED)]:
    acf(f"G{ACT_FIRST}:G{ACT_LAST}", CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill(clr)))
acf(f"B{ACT_FIRST}:B{ACT_LAST}", FormulaRule(formula=[f'AND($B{ACT_FIRST}<>"",$B{ACT_FIRST}<TODAY())'], fill=fill(F_SLATE)))
acf(f"B{ACT_FIRST}:B{ACT_LAST}", FormulaRule(formula=[f'AND($B{ACT_FIRST}<>"",$B{ACT_FIRST}>=TODAY(),$B{ACT_FIRST}<=TODAY()+14)'], fill=fill(F_AMBER)))
band(act, ACT_FIRST, ACT_LAST, len(ACT_H))
act.freeze_panes="C3"; act.auto_filter.ref=f"A2:{ACT_LASTCOL}{ACT_LAST}"
act.sheet_properties.tabColor="246B5A"; printsetup(act)

# =====================================================================
# ACTION LIST  (live worklist, most urgent first, from the Master List)
# =====================================================================
print("building action list...")
al = wb.create_sheet("Action List")
AL_H=["#","Partner","Type","Priority","Status","Next Action","Next Action Date","Days Since","Owner"]
title_row(al,"Action List","I")
subtitle(al,2,"Most urgent first. Partners with a Next Action Date coming due lead the list, then P1, then P2.")
alw={1:5,2:36,3:20,4:10,5:16,6:34,7:16,8:12,9:14}
for c,htext in enumerate(AL_H,start=1):
    hcell(al,2,c,htext); al.column_dimensions[get_column_letter(c)].width=alw[c]
al.row_dimensions[2].height=46
AL_FIRST=3; AL_LAST=82
AL_CENTER={1,4,5,7,8,9}
ALMAP={2:"B",3:"A",4:"E",5:"F",6:"S",7:"T",8:"R",9:"G"}   # Action List col -> Master List col
for i in range(AL_LAST-AL_FIRST+1):
    r=AL_FIRST+i
    al.row_dimensions[r].height=26
    al.cell(r,11,"=IFERROR(MATCH(SMALL({M}$AL$3:$AL$500,ROW()-2),{M}$AL$3:$AL$500,0),\"\")".format(M=MLR))
    for c,col in ALMAP.items():
        cell=al.cell(r,c,"=IFERROR(INDEX({M}${col}$3:${col}$500,$K{r}),\"\")".format(M=MLR,col=col,r=r))
        if c == 2:
            cell.font=font(13,True,C_DATA)
        else:
            cell.font=font()
        cell.border=BORD
        cell.alignment=A_CL if c in AL_CENTER else A_L
    al.cell(r,1,'=IF($B{r}="","",ROW()-2)'.format(r=r))
    al.cell(r,7).number_format=FMT_DATE
    al.cell(r,8).number_format=FMT_INT
    al.cell(r,11).font=font(9,color="FFB7C9C1")
al.column_dimensions["K"].hidden=True
al.freeze_panes="C3"; al.auto_filter.ref="A2:I{}".format(AL_LAST)
al.sheet_properties.tabColor="3E7C68"; printsetup(al)
alcf=al.conditional_formatting.add
for val,clr in [("P1",F_P1),("P2",F_P2),("P3",F_P3)]:
    alcf("D{}:D{}".format(AL_FIRST,AL_LAST), CellIsRule(operator="equal", formula=['"{}"'.format(val)], fill=fill(clr)))
for val,clr in [("Activated",F_GREEN),("Repeat Partner",F_GREEN),("Agreed",F_AMBER),("Proposal Sent",F_AMBER),("In Conversation",F_AMBER),("Outreach Sent",F_PALEBLU),("Open",F_GREY),("Lost",F_RED),("On Hold",F_SLATE)]:
    alcf("E{}:E{}".format(AL_FIRST,AL_LAST), CellIsRule(operator="equal", formula=['"{}"'.format(val)], fill=fill(clr)))
alcf("G{}:G{}".format(AL_FIRST,AL_LAST), FormulaRule(formula=['AND($G{}<>"",$G{}<TODAY())'.format(AL_FIRST,AL_FIRST)], fill=fill(F_RED)))
alcf("H{}:H{}".format(AL_FIRST,AL_LAST), FormulaRule(formula=['AND($D{}="P1",$H{}>7)'.format(AL_FIRST,AL_FIRST)], fill=fill(F_AMBER)))


# =====================================================================
# DASHBOARD
# =====================================================================
print("building dashboard...")
dash = wb.create_sheet("Dashboard")
dash.sheet_view.showGridLines = False
title_row(dash, "Organika RTD  ·  Community Partnerships", "N")
subtitle(dash, 2, "Live numbers. Everything here updates the moment you save a change on any tab.")
# top-right countdown chip in the title band
dash.merge_cells("A2:K2")
dash.merge_cells("L2:N2")
for c in range(1,15): dash.column_dimensions[get_column_letter(c)].width = 12.3
ds=dash.cell(2,12,'="As of  "&TEXT(TODAY(),"yyyy/mm/dd")')
ds.font=font(11,False,C_SUB); ds.alignment=Alignment(horizontal="right",vertical="center")
ML=MLR
ACT="Activations!"
# hidden helper data that feeds the charts (columns P,Q)
for c in ("P","Q","R","S"): dash.column_dimensions[c].hidden=True
def hset(r,lbl,f):
    dash.cell(r,16,lbl).font=font(10,False,C_SUB); dash.cell(r,17,f)
for i,s in enumerate(STATUS): hset(4+i,s,'=COUNTIF({M}$F$3:$F$500,"{s}")'.format(M=ML,s=s))
for i,t in enumerate(TYPES): hset(14+i,t,"=COUNTA('{t}'!$B$3:$B$32)".format(t=t))
for i,p in enumerate(PRIORITY): hset(23+i,p,'=COUNTIF({M}$E$3:$E$500,"{p}")'.format(M=ML,p=p))
# KPI cards
def kpi(col,label,formula,numfmt=None):
    L1=get_column_letter(col); L2=get_column_letter(col+1)
    for r in range(4,7):                       # a clean white card with an accent top rule
        for c in (col,col+1):
            cell=dash.cell(r,c); cell.fill=fill(WHITE); sides={}
            if r==4: sides["top"]=Side(style="medium",color=ACCENT)
            if r==6: sides["bottom"]=Side(style="thin",color=CARD_BRD)
            if c==col: sides["left"]=Side(style="thin",color=CARD_BRD)
            if c==col+1: sides["right"]=Side(style="thin",color=CARD_BRD)
            cell.border=Border(**sides)
    dash.merge_cells("{}4:{}4".format(L1,L2)); dash.merge_cells("{}5:{}6".format(L1,L2))
    a=dash.cell(4,col,label); a.font=font(10,False,C_SUB); a.alignment=Alignment(horizontal="left",vertical="center",indent=2)
    b=dash.cell(5,col,formula); b.font=font(32,True,C_TITLE); b.alignment=Alignment(horizontal="left",vertical="center",indent=2)
    if numfmt: b.number_format=numfmt
dash.row_dimensions[4].height=22; dash.row_dimensions[5].height=36; dash.row_dimensions[6].height=12
kpi(1,"Total Partners","=SUM(Q14:Q21)")
kpi(3,"P1 Partners","=Q23")
kpi(5,"Active Conversations",'=COUNTIF({M}$F$3:$F$500,"Outreach Sent")+COUNTIF({M}$F$3:$F$500,"In Conversation")+COUNTIF({M}$F$3:$F$500,"Proposal Sent")+COUNTIF({M}$F$3:$F$500,"Agreed")'.format(M=ML))
kpi(7,"Activations Booked",'=COUNTIF({A}$G$3:$G$500,"Booked")+COUNTIF({A}$G$3:$G$500,"Confirmed")+COUNTIF({A}$G$3:$G$500,"Delivered")'.format(A=ACT))
kpi(9,"Cans Sampled","=SUM({A}$N$3:$N$500)".format(A=ACT),"#,##0")
kpi(11,"Cost Per Can","=IFERROR(SUM({A}$J$3:$J$500)/SUM({A}$N$3:$N$500),0)".format(A=ACT),FMT_CENTS)
kpi(13,"Days to Costco Road Show","=MAX(0,DATE(2026,8,1)-TODAY())","#,##0")
# section dividers and table blocks
def divider(r,text):
    dash.merge_cells(start_row=r,start_column=1,end_row=r,end_column=14)
    c=dash.cell(r,1,text); c.font=font(11,True,WHITE); c.fill=fill(C_HEADER); c.alignment=A_L
    dash.row_dimensions[r].height=20
def block(r,c,title,rows):
    dash.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+2)
    h=dash.cell(r,c,title); h.font=font(11,True,WHITE); h.fill=fill(C_HEADER); h.alignment=A_L; h.border=BORD
    dash.cell(r,c+1).fill=fill(C_HEADER); dash.cell(r,c+1).border=BORD
    dash.cell(r,c+2).fill=fill(C_HEADER); dash.cell(r,c+2).border=BORD
    for i,(lbl,f,fmt) in enumerate(rows):
        rr=r+1+i
        dash.merge_cells(start_row=rr,start_column=c,end_row=rr,end_column=c+1)
        a=dash.cell(rr,c,lbl); a.font=font(11,False,C_DATA); a.alignment=A_L; a.border=BORD
        dash.cell(rr,c+1).border=BORD
        b=dash.cell(rr,c+2,f); b.font=font(11,False,C_DATA); b.alignment=A_CL; b.border=BORD
        if fmt: b.number_format=fmt
divider(8,"Snapshot")
def style_series(ser,clr): ser.graphicalProperties=GraphicalProperties(solidFill=clr)
ch1=BarChart(); ch1.type="bar"; ch1.title="Pipeline by Stage"; ch1.legend=None; ch1.height=6.5; ch1.width=8.8
ch1.add_data(Reference(dash,min_col=17,min_row=4,max_row=12)); ch1.set_categories(Reference(dash,min_col=16,min_row=4,max_row=12))
style_series(ch1.series[0],C_HEADER[2:]); ch1.dataLabels=DataLabelList(); ch1.dataLabels.showVal=True
dash.add_chart(ch1,"A9")
ch2=BarChart(); ch2.type="bar"; ch2.title="Partners by Type"; ch2.legend=None; ch2.height=6.5; ch2.width=8.8
ch2.add_data(Reference(dash,min_col=17,min_row=14,max_row=21)); ch2.set_categories(Reference(dash,min_col=16,min_row=14,max_row=21))
style_series(ch2.series[0],"6FA392"); ch2.dataLabels=DataLabelList(); ch2.dataLabels.showVal=True
dash.add_chart(ch2,"F9")
ch3=DoughnutChart(); ch3.title="Priority Split"; ch3.height=6.5; ch3.width=8.8
ch3.add_data(Reference(dash,min_col=17,min_row=23,max_row=25)); ch3.set_categories(Reference(dash,min_col=16,min_row=23,max_row=25))
pts=[DataPoint(idx=i) for i in range(3)]
for dp,clr in zip(pts,["2E5A4E","4F8A78","8FB3A6"]): dp.graphicalProperties=GraphicalProperties(solidFill=clr)
ch3.series[0].data_points=pts; ch3.dataLabels=DataLabelList(); ch3.dataLabels.showVal=True
dash.add_chart(ch3,"K9")
divider(26,"Detail")
METRO=["Vancouver","Burnaby","North Vancouver","West Vancouver","Richmond","Surrey","Langley","Coquitlam","New Westminster"]
block(27,1,"Region",[("Metro Vancouver","="+"+".join('COUNTIF({M}$C$3:$C$500,"{c}")'.format(M=ML,c=c) for c in METRO),None),
    ("Victoria",'=COUNTIF({M}$C$3:$C$500,"Victoria")'.format(M=ML),None),
    ("Kelowna",'=COUNTIF({M}$C$3:$C$500,"Kelowna")'.format(M=ML),None)])
block(27,5,"Activations Ahead",[(lbl,'=COUNTIFS({A}$B$3:$B$500,">="&TODAY(),{A}$B$3:$B$500,"<="&TODAY()+{n})'.format(A=ACT,n=n),None) for lbl,n in [("Next 30 Days",30),("Next 60 Days",60),("Next 90 Days",90)]])
block(27,9,"Cans by Flavour",[("Raspberry","=SUM({A}$K$3:$K$500)".format(A=ACT),"#,##0"),
    ("Lemon Lime","=SUM({A}$L$3:$L$500)".format(A=ACT),"#,##0"),
    ("Pineapple","=SUM({A}$M$3:$M$500)".format(A=ACT),"#,##0"),
    ("Total Cans","=SUM({A}$N$3:$N$500)".format(A=ACT),"#,##0")])
block(33,1,"List Health",[("Overdue Actions",'=COUNTIF({M}$T$3:$T$500,"<"&TODAY())'.format(M=ML),None),
    ("Stale P1 Rows",'=COUNTIFS({M}$E$3:$E$500,"P1",{M}$R$3:$R$500,">7")'.format(M=ML),None),
    ("Verified",'=COUNTIF({M}$O$3:$O$500,"Verified Web")+COUNTIF({M}$O$3:$O$500,"Verified Phone")+COUNTIF({M}$O$3:$O$500,"Verified In Person")'.format(M=ML),None),
    ("Pending",'=COUNTIF({M}$O$3:$O$500,"Pending")'.format(M=ML),None)])
block(33,5,"Budget",[("Total Budget","=Budget!B14",FMT_MONEY),("Committed","=Budget!C14",FMT_MONEY),
    ("Spent","=Budget!D14",FMT_MONEY),("Remaining","=Budget!E14",FMT_MONEY),("% Spent","=Budget!F14",FMT_PCT)])
dash.cell(40,1,"Everything here updates on its own. Edit a partner on its type tab, log an activation on the Activations tab, and these tiles and charts refresh.").font=font(11,False,C_SUB)
# conditional: Days to Costco tile turns amber under 30 days, rose under 7
from openpyxl.formatting.rule import FormulaRule as _FR
dash.conditional_formatting.add("M5:N6", _FR(formula=["MAX(0,DATE(2026,8,1)-TODAY())<30"], fill=fill(F_AMBER)))
dash.conditional_formatting.add("M5:N6", _FR(formula=["MAX(0,DATE(2026,8,1)-TODAY())<7"],  fill=fill(F_RED)))
dash.sheet_properties.tabColor="2E5A4E"

# =====================================================================
# TYPE SUMMARY
# =====================================================================
print("building type summary...")
TARGETS = {"Run Clubs":10,"Gyms & Studios":10,"Events & Festivals":8,"Sports Teams & Leagues":6,
"Campus & Student Groups":6,"Wellness & Recovery":6,"Ambassadors & Creators":8,"Charity & Causes":4}
ts = wb.create_sheet("Type Summary")
ts.sheet_view.showGridLines = False
title_row(ts, "Type Summary", "K")
TS_HDR=["Partnership Type","Draft Target","Partners","P1","P2","P3","In Conversation","Agreed","Activated","Activations","% to Target"]
for c,htext in enumerate(TS_HDR,start=1): hcell(ts,2,c,htext)
tsw={1:24,2:14,3:12,4:8,5:8,6:8,7:16,8:12,9:12,10:16,11:12}
for c,w in tsw.items(): ts.column_dimensions[get_column_letter(c)].width=w
ts.row_dimensions[2].height=50
for i,t in enumerate(TYPES):
    r=3+i
    vals=[t,TARGETS[t],
      "=COUNTA('{t}'!$B$3:$B$32)".format(t=t),
      "=COUNTIF('{t}'!$F$3:$F$32,\"P1\")".format(t=t),
      "=COUNTIF('{t}'!$F$3:$F$32,\"P2\")".format(t=t),
      "=COUNTIF('{t}'!$F$3:$F$32,\"P3\")".format(t=t),
      "=COUNTIF('{t}'!$G$3:$G$32,\"In Conversation\")".format(t=t),
      "=COUNTIF('{t}'!$G$3:$G$32,\"Agreed\")".format(t=t),
      "=COUNTIF('{t}'!$G$3:$G$32,\"Activated\")+COUNTIF('{t}'!$G$3:$G$32,\"Repeat Partner\")".format(t=t),
      "=COUNTIF(Activations!$D$3:$D$500,\"{t}\")".format(t=t),
      "=IFERROR(C{r}/B{r},0)".format(r=r)]
    for c,v in enumerate(vals,start=1):
        cell=ts.cell(r,c,v); cell.font=font(); cell.border=BORD
        cell.alignment=A_L if c==1 else A_CL
        if c==11: cell.number_format=FMT_PCT
tr=3+len(TYPES)
ts.cell(tr,1,"Total").font=font(12,True,C_DATA)
ts.cell(tr,1).fill=fill(C_TOTAL); ts.cell(tr,1).border=BORD
for c in range(2,11):
    L=get_column_letter(c); cell=ts.cell(tr,c,"=SUM({L}3:{L}{e})".format(L=L,e=tr-1))
    cell.font=font(12,True,C_DATA); cell.fill=fill(C_TOTAL); cell.alignment=A_CL; cell.border=BORD
ck=ts.cell(tr,11,"=IFERROR(C{r}/B{r},0)".format(r=tr))
ck.font=font(12,True,C_DATA); ck.fill=fill(C_TOTAL); ck.alignment=A_CL; ck.border=BORD; ck.number_format=FMT_PCT
# colour cues on % to Target
tscf = ts.conditional_formatting.add
tscf("K3:K{}".format(tr), CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=fill(F_GREEN)))
tscf("K3:K{}".format(tr), CellIsRule(operator="greaterThanOrEqual", formula=["0.5"], fill=fill(F_AMBER)))
ts.cell(tr+2,1,"Targets are a starting draft for Louis to confirm. Edit the Target column any time.").font=font(11,False,C_SUB)
ts.freeze_panes="A3"; ts.sheet_properties.tabColor="3E7C68"

# =====================================================================
# BUDGET
# =====================================================================
print("building budget...")
bud = wb.create_sheet("Budget")
bud.sheet_view.showGridLines = False
title_row(bud, "Budget", "F")
B_HDR=["Partnership Type","Total Budget","Committed","Spent","Remaining","% Spent"]
for c,htext in enumerate(B_HDR,start=1): hcell(bud,2,c,htext)
budw={1:24,2:16,3:16,4:16,5:16,6:12}
for c,w in budw.items(): bud.column_dimensions[get_column_letter(c)].width=w
bud.row_dimensions[2].height=50
for i,t in enumerate(TYPES):
    r=3+i
    cells=[t,0,0,"=SUMIF(Activations!$D$3:$D$500,$A{r},Activations!$J$3:$J$500)".format(r=r),"=B{r}-D{r}".format(r=r),"=IFERROR(D{r}/B{r},0)".format(r=r)]
    for c,v in enumerate(cells,start=1):
        cell=bud.cell(r,c,v); cell.font=font(); cell.border=BORD
        cell.alignment=A_L if c==1 else A_CL
        if c in (2,3,4,5): cell.number_format=FMT_MONEY
        if c==6: cell.number_format=FMT_PCT
br=3+len(TYPES)
bud.cell(br,1,"Total").font=font(12,True,C_DATA); bud.cell(br,1).fill=fill(C_TOTAL); bud.cell(br,1).border=BORD
for c in range(2,6):
    L=get_column_letter(c); cell=bud.cell(br,c,"=SUM({L}3:{L}{e})".format(L=L,e=br-1))
    cell.font=font(12,True,C_DATA); cell.fill=fill(C_TOTAL); cell.alignment=A_CL; cell.border=BORD; cell.number_format=FMT_MONEY
ce=bud.cell(br,6,"=IFERROR(D{b}/B{b},0)".format(b=br)); ce.font=font(12,True,C_DATA); ce.fill=fill(C_TOTAL); ce.alignment=A_CL; ce.border=BORD; ce.number_format=FMT_PCT
bud.cell(br+2,1,"Total Budget and Committed are yours to set and start at zero. Spent adds up the Budget column on the Activations tab by type.").font=font(11,False,C_SUB)
bud.freeze_panes="A3"; bud.sheet_properties.tabColor="8FB3A6"

# =====================================================================
# SUGGESTED EVENTS  (empty parking lot for ideas farther down the line)
# =====================================================================
print("building suggested events...")
se = wb.create_sheet("Suggested Events")
SE_HDRS = ["#","Event","City","Venue or Neighbourhood","Expected Window","Audience Size",
"Audience Source","Source","Priority","Event Status","Owner","Why It Fits","Next Step","Notes"]
SE_LAST = get_column_letter(len(SE_HDRS))   # N
title_row(se, "Suggested Events", SE_LAST)
sew = {1:5,2:30,3:14,4:24,5:18,6:16,7:30,8:14,9:10,10:20,11:14,12:32,13:28,14:40}
for c,htext in enumerate(SE_HDRS, start=1):
    hcell(se, 2, c, htext)
    se.column_dimensions[get_column_letter(c)].width = sew[c]
se.row_dimensions[2].height = 52
SE_CENTER = {1,3,5,6,8,9,10,11}
for i in range(DATA_ROWS):
    r = FIRST + i
    se.row_dimensions[r].height = 26
    for c in range(1, len(SE_HDRS)+1):
        cell = se.cell(r,c)
        cell.font = font(); cell.border = BORD
        cell.alignment = A_CL if c in SE_CENTER else A_L
    se.cell(r,1, f'=IF($B{r}="","",ROW()-2)')
# dropdowns
def se_dv(name, colrange):
    dv = DataValidation(type="list", formula1=name, allow_blank=True, showErrorMessage=False)
    se.add_data_validation(dv); dv.add(colrange)
se_dv("LU_City", f"C{FIRST}:C{LASTROW}")
se_dv("LU_Source", f"H{FIRST}:H{LASTROW}")
se_dv("LU_Priority", f"I{FIRST}:I{LASTROW}")
se_dv("LU_EventStatus", f"J{FIRST}:J{LASTROW}")
se_dv("LU_Owner", f"K{FIRST}:K{LASTROW}")
# conditional formatting
secf = se.conditional_formatting.add
for val,clr in [("Idea",F_GREY),("Scoping",F_PALEBLU),("Approved",F_AMBER),
                ("Moved To Events Tab",F_GREEN),("Passed",F_RED)]:
    secf(f"J{FIRST}:J{LASTROW}", CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill(clr)))
for val,clr in [("P1",F_P1),("P2",F_P2),("P3",F_P3)]:
    secf(f"I{FIRST}:I{LASTROW}", CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill(clr)))
for val,clr in [("Verified Web",F_GREEN),("Verified Phone",F_GREEN),("Verified In Person",F_GREEN),
                ("Pending",F_AMBER),("Unverified",F_P3)]:
    secf(f"H{FIRST}:H{LASTROW}", CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill(clr)))
band(se, FIRST, LASTROW, len(SE_HDRS))
se.freeze_panes = "C3"
se.auto_filter.ref = f"A2:{SE_LAST}{LASTROW}"
se.sheet_properties.tabColor = "8FB3A6"; printsetup(se)
subtitle(se, LASTROW+2, "A place to park event ideas for later. It starts empty. When you green light one, add it as a partner on the Events & Festivals tab and log the activation on the Activations tab.")

# =====================================================================
# SALES TEAM  (Maddie only, per Louis)
# =====================================================================
print("building sales team...")
st = wb.create_sheet("Sales Team")
title_row(st, "Sales Team", "F")
ST_HDR=["Rep","Email","Cell","Region","Role on This File","On Owner Dropdown?"]
for c,htext in enumerate(ST_HDR,start=1): hcell(st,2,c,htext)
SALES=[
("Maddie","TBD verify","TBD verify","BC","Community partnerships owner per Louis","Yes"),
]
stw={1:22,2:30,3:16,4:10,5:40,6:22}
for c,w in stw.items(): st.column_dimensions[get_column_letter(c)].width=w
st.row_dimensions[2].height=50
for i,row in enumerate(SALES):
    r=3+i
    for c,v in enumerate(row,start=1):
        cell=st.cell(r,c,v); cell.font=font(); cell.border=BORD
        cell.alignment=A_L if c in (1,2,5) else A_CL
subtitle(st, 5, "Maddie runs this file. To add a rep later, add a row here and add the name to the Owner column on the Lookups tab so it appears on the Primary Owner menu.")
st.freeze_panes="A3"; st.sheet_properties.tabColor="B7C9C1"

# =====================================================================
# GUIDE
# =====================================================================
print("building guide...")
gd = wb.create_sheet("Guide")
gd.sheet_view.showGridLines = False
title_row(gd, "Guide", "D")
gd.column_dimensions["A"].width=32; gd.column_dimensions["B"].width=84
gd.column_dimensions["C"].width=4; gd.column_dimensions["D"].width=4
def gsec(r,txt):
    c=gd.cell(r,1,txt); gd.merge_cells("A{r}:D{r}".format(r=r))
    c.font=font(13,True,WHITE); c.fill=fill(C_HEADER); c.alignment=A_L; gd.row_dimensions[r].height=24
def gline(r,a,b=None):
    ca=gd.cell(r,1,a); ca.font=font(11,True,C_DATA) if b else font(11,False,C_DATA); ca.alignment=A_LW
    if b is not None:
        cb=gd.cell(r,2,b); cb.font=font(11,False,C_DATA); cb.alignment=A_LW
    gd.row_dimensions[r].height=16
gd.cell(2,1,"How the file works. You type into the green tabs or pick from a menu.").font=font(11,False,C_SUB)
r=4
gsec(r,"What this file is for"); r+=1
gline(r,"This is the sister file to the BC Tracker. The BC Tracker tracks retail doors. This one tracks community partnerships and the activations that come from them."); r+=1
gline(r,"If it sells cans on a shelf, it goes in the BC Tracker. If it samples, sponsors, or makes content, it goes here."); r+=1
gline(r,"A partner can be in both files. The In BC Tracker column flags the ones that are."); r+=2
gsec(r,"Working a partner"); r+=1
gline(r,"Open a green type tab. Each tab is one kind of partner: run clubs, gyms, events, and so on."); r+=1
gline(r,"A few detail columns are folded away under the plus sign above column V. Click it if you need cost, contra or deliverables."); r+=1
gline(r,"Keep Status, Last Contacted, Next Action and Next Action Date current as you go."); r+=1
gline(r,"Click a partner's Instagram or website link to open their page."); r+=1
gline(r,"Nothing about activations is filled in for you. You decide what runs and when."); r+=2
gsec(r,"Booking an activation"); r+=1
gline(r,"Go to the Activations tab and add a row."); r+=1
gline(r,"Pick the partner, date, status, assets and budget from the menus, or type your own value if it is not on the list."); r+=1
gline(r,"Log cans used by flavour. Total Cans and Cost Per Can fill in on their own."); r+=1
gline(r,"The Dashboard, Budget and Type Summary all read from this tab."); r+=2
gsec(r,"Reading the colours"); r+=1
gline(r,"Green is done: Activated, Repeat Partner, Delivered."); r+=1
gline(r,"Amber is in progress: In Conversation, Proposal Sent, Agreed, Booked."); r+=1
gline(r,"Pale blue is early days. Grey is Open or On Hold. Rose is Lost or Cancelled."); r+=1
gline(r,"A Next Action Date in the past turns rose. A P1 partner with no contact for over 7 days turns amber."); r+=1
gline(r,"P1 is warm orange, P2 is blue, P3 is grey, the same as the BC Tracker."); r+=2
gsec(r,"How sure we are about each partner"); r+=1
gline(r,"Verified Web","Confirmed on an official site or public Instagram in 2026."); r+=1
gline(r,"Pending","Real, but one detail such as a follower count needs a second look. Reverify before outreach."); r+=1
gline(r,"Unverified","Heard of but not confirmed. Treat as a lead only."); r+=1
gline(r,"Anything we could not confirm is left blank or TBD."); r+=2
gsec(r,"The three flavours"); r+=1
gline(r,"Raspberry is 4338. Lemon Lime is 4336. Pineapple Passion Fruit is 4340. 355ml cans."); r+=2
gsec(r,"The tabs"); r+=1
for lbl,desc in [
 ("Start Here","The front door. Tap a card to jump to where you need to go."),
 ("Dashboard","The numbers and charts at a glance. Read only."),
 ("Action List","Who to work next, most urgent first. Read only."),
 ("Activations","Where you log each activation. Cans, budget, status, assets, notes."),
 ("Master List","Every partner in one place. Read only, builds itself from the type tabs."),
 ("Type Summary","Targets against actuals per type. Targets are a draft to confirm."),
 ("Budget","Set a budget per type. Spent fills in from the Activations tab."),
 ("Type tabs","Where you work the partners. One per type."),
 ("Suggested Events","A parking lot for event ideas. Starts empty."),
 ("Lookups","The lists behind the menus. Add a value and it shows up everywhere."),
 ("Sales Team","Maddie owns this file for now."),
]:
    gline(r,lbl,desc); r+=1
gd.sheet_properties.tabColor="B7C9C1"


# =====================================================================
# START HERE  v10 — 2x2 card grid, live hero stats, Apple-level spacing
# =====================================================================
print("building start here...")
from openpyxl.worksheet.hyperlink import Hyperlink
home = wb.create_sheet("Start Here")
home.sheet_view.showGridLines = False

# column grid: A=narrow margin, B-I=left half, J=gap, K-Q=right half
home.column_dimensions["A"].width = 2
for c in range(2, 10):  home.column_dimensions[get_column_letter(c)].width = 11.2   # B-I
home.column_dimensions["J"].width = 2.4
for c in range(11, 18): home.column_dimensions[get_column_letter(c)].width = 11.2   # K-Q

# ---- HERO BAND rows 2-5 ----
for r in range(2, 6):
    for col in range(2, 18):
        home.cell(r, col).fill = fill(C_TITLE)
home.row_dimensions[1].height = 8    # small top margin
home.merge_cells("B2:Q3")
h = home.cell(2, 2, "Community Partnerships")
h.font = Font(name=FONTNAME, size=30, bold=True, color=WHITE)
h.alignment = Alignment(horizontal="left", vertical="center", indent=2)
home.row_dimensions[2].height = 44
home.row_dimensions[3].height = 14

# subtitle row with live partner count and countdown
home.merge_cells("B4:L4")
sst = home.cell(4, 2, "Organika RTD    ·    Find partners, book activations, win the summer 2026 Costco road show.")
sst.font = Font(name=FONTNAME, size=11, bold=False, color="FFBFD2CB")
sst.alignment = Alignment(horizontal="left", vertical="center", indent=2)

# days-to-Costco chip in top-right of hero
home.merge_cells("M4:Q4")
chip = home.cell(4, 13, '=TEXT(MAX(0,DATE(2026,8,1)-TODAY()),"0")&" days to Costco"')
chip.font = Font(name=FONTNAME, size=11, bold=True, color=WHITE)
chip.alignment = Alignment(horizontal="right", vertical="center")
home.row_dimensions[4].height = 28

# accent rule row 5
for col in range(2, 18): home.cell(5, col).fill = fill(ACCENT)
home.row_dimensions[5].height = 4

# spacer row 6
home.row_dimensions[6].height = 20

# ---- NAV CARDS: 2×2 grid ----
# Card helper — c0 and c1 are start/end columns, r0 and r1 are start/end rows
def card2(c0, c1, r0, r1, title, desc, target, clr):
    for rr in range(r0, r1+1):
        for cc in range(c0, c1+1):
            cell = home.cell(rr, cc)
            cell.fill = fill(WHITE)
            sides = {}
            if rr == r0:   sides["top"]    = Side(style="thin",  color=CARD_BRD)
            if rr == r1:   sides["bottom"] = Side(style="thin",  color=CARD_BRD)
            if cc == c0:   sides["left"]   = Side(style="thick", color=clr)
            if cc == c1:   sides["right"]  = Side(style="thin",  color=CARD_BRD)
            cell.border = Border(**sides)
    # title row spans top half
    mid = r0 + (r1 - r0) // 2
    home.merge_cells(start_row=r0,   start_column=c0, end_row=mid,  end_column=c1)
    home.merge_cells(start_row=mid+1,start_column=c0, end_row=r1,   end_column=c1)
    t = home.cell(r0, c0, title + "    ›")
    t.font = Font(name=FONTNAME, size=15, bold=True, color=C_TITLE)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=False)
    t.hyperlink = Hyperlink(ref=t.coordinate, location="'%s'!A1" % target, display=title)
    d = home.cell(mid+1, c0, desc)
    d.font = Font(name=FONTNAME, size=10, bold=False, color=C_SUB)
    d.alignment = Alignment(horizontal="left", vertical="top", indent=2, wrap_text=True)
    d.hyperlink = Hyperlink(ref=d.coordinate, location="'%s'!A1" % target, display=desc)

# row heights for the 2×2 grid (rows 7-11 = top row, rows 13-17 = bottom row)
for r in range(7, 18):
    home.row_dimensions[r].height = 24
home.row_dimensions[7].height  = 14   # top padding inside card
home.row_dimensions[9].height  = 8    # divider between title and desc
home.row_dimensions[11].height = 14   # bottom padding
home.row_dimensions[12].height = 10   # gap between rows
home.row_dimensions[13].height = 14
home.row_dimensions[15].height = 8
home.row_dimensions[17].height = 14

# Top row: cols B-I (2-9) and K-Q (11-17)
card2(2,  9,  7, 11, "Who to contact",    "Your worklist, sorted by urgency. The partners who need attention today are at the top.", "Action List", "FF2E5A4E")
card2(11, 17, 7, 11, "Log an activation", "Record every sampling, event booth, and sponsorship here. Budget and cans track themselves.", "Activations", "FF3E7C68")
# Bottom row
card2(2,  9, 13, 17, "See the picture",   "Live totals, pipeline charts, and budget at a glance. Updates the moment you log anything.", "Dashboard",   "FF6FA392")
card2(11, 17,13, 17, "Browse all partners","All 80 partners in one scrollable view. Filter by city, type, or status.", "Master List", "FF8FB3A6")

# ---- HOW IT WORKS ----
home.row_dimensions[19].height = 16   # spacer
hw = home.cell(20, 2, "How it works")
hw.font = Font(name=FONTNAME, size=13, bold=True, color=C_TITLE)
hw.alignment = Alignment(horizontal="left", vertical="center", indent=2)
home.row_dimensions[20].height = 26

steps = [
    "1.    Open the Action List. The most urgent partner is always at the top.",
    "2.    Open that partner on its green tab. Update Status, Next Action and Next Action Date.",
    "3.    When something is confirmed, log it on the Activations tab. Everything else fills itself in.",
]
for i, sline in enumerate(steps):
    cs = home.cell(21+i, 2, sline)
    cs.font = font(11, False, C_DATA)
    cs.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    home.row_dimensions[21+i].height = 22

home.row_dimensions[25].height = 10
ft = home.cell(26, 2, "You only ever type into the green partner tabs and the Activations tab. Every other tab reads itself.")
ft.font = font(10, False, C_SUB)
ft.alignment = Alignment(horizontal="left", vertical="center", indent=2)
home.row_dimensions[26].height = 20

home.sheet_properties.tabColor = "22413A"

# =====================================================================
# ORDER + SAVE
# =====================================================================
for ws in wb.worksheets:
    printsetup(ws, "1:2" if ws.title not in ("Dashboard","Guide","Start Here") else None)
for _nm in ("Dashboard","Type Summary"):
    wb[_nm].protection.sheet = True
    wb[_nm].protection.selectLockedCells = False
_mp = wb["Master List"].protection
_mp.sheet = True; _mp.autoFilter = False; _mp.sort = False; _mp.selectLockedCells = False
order = ["Start Here","Dashboard","Action List","Activations","Master List","Type Summary","Budget"] + TYPES + ["Suggested Events","Lookups","Sales Team","Guide"]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active = 0
try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation = CalcProperties(fullCalcOnLoad=True)
wb.save(OUT)
print("SAVED:", OUT)
print("master list rows:", ML_ROWS_END, " sheets:", [s.title for s in wb.worksheets])
