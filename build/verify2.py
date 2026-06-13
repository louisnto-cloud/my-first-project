"""Verify the new Scenarios sheet + Summary additions against the python reference."""
import sys
from openpyxl.workbook.defined_name import DefinedNameDict
if not hasattr(DefinedNameDict, "definedName"):
    DefinedNameDict.definedName = property(lambda s: list(s.values()))
from pycel import ExcelCompiler
import openpyxl, muv_config as m

xl = ExcelCompiler("../MUV_RTD_Forecast_v1.xlsx")
wb = openpyxl.load_workbook("../MUV_RTD_Forecast_v1.xlsx")
sm = wb["Summary"]
fails = 0
def chk(name, got, exp, tol=2.0):
    global fails
    got = got if isinstance(got, (int, float)) else float("nan")
    ok = abs(got - exp) <= tol
    fails += 0 if ok else 1
    print(f"  {name:36s} xl={got:14,.4f} ref={exp:14,.4f} {'OK' if ok else 'FAIL'}")

S = m.compute()
online = S["by_channel"]["Online"]; phys = S["gross"] - online

print("Scenarios — live helpers & Path to $3M:")
chk("B4 AnnualGross", xl.evaluate("Scenarios!B4"), S["gross"])
chk("B5 Online", xl.evaluate("Scenarios!B5"), online)
chk("B6 Physical", xl.evaluate("Scenarios!B6"), phys)
req = m.required_single_lever()
chk("Path price index (C13)", xl.evaluate("Scenarios!C13"), req["price_index"], 1e-5)
chk("Path velocity (C14)", xl.evaluate("Scenarios!C14"), req["velocity_uplift"], 1e-5)
chk("Path online (C15)", xl.evaluate("Scenarios!C15"), req["online_growth"], 1e-5)
chk("Path doors (C16)", xl.evaluate("Scenarios!C16"), req["calib_scalar"], 1e-5)

print("Scenario comparison (closed-form == full engine):")
for i, (nm, lev) in enumerate(m.SCENARIOS.items()):
    chk(f"Scenario {nm} gross (E{20+i})", xl.evaluate(f"Scenarios!E{20+i}"), m.compute(lev)["gross"])

print("Sensitivity grid:")
chk("calib0.36 x price0 (C29)", xl.evaluate("Scenarios!C29"), S["gross"])
chk("calib0.42 x price+5% (D31)", xl.evaluate("Scenarios!D31"), (phys*(0.42/0.36)+online)*1.05)
chk("calib0.30 x price-5% (B27)", xl.evaluate("Scenarios!B27"), (phys*(0.30/0.36)+online)*0.95)

print("Summary — By Store Tier & Volume:")
# locate tier table (col D label 'Convenience A') and volume table (col H label 'Convenience')
def find_row(col, text):
    for r in range(1, sm.max_row+1):
        if sm.cell(r, col).value == text: return r
    return None
tr = find_row(4, "Convenience A")
conv_gross = S["by_channel"]["Convenience"]
chk("Tier Convenience A (E)", xl.evaluate(f"Summary!E{tr}"), conv_gross*450/1370)
tier_tot = sum(xl.evaluate(f"Summary!E{tr+i}") for i in range(18))
chk("Tier table total == physical", tier_tot, phys)
vr = find_row(8, "Convenience")
vol_tot = sum(xl.evaluate(f"Summary!I{vr+i}") for i in range(len(m.CHANNELS)))
chk("Volume total == total cases", vol_tot, S["cases"])

print("\n", "ALL NEW-FEATURE CHECKS PASSED" if fails == 0 else f"*** {fails} FAILED ***")
sys.exit(1 if fails else 0)
