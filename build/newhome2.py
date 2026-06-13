print("building start here...")
from openpyxl.worksheet.hyperlink import Hyperlink
home = wb.create_sheet("Start Here")
home.sheet_view.showGridLines = False
for c in range(1,18): home.column_dimensions[get_column_letter(c)].width = 9.6
home.column_dimensions["A"].width = 3
# hero band (deep evergreen) with a thin accent rule beneath
for r in (2,3,4):
    for col in range(2,18): home.cell(r,col).fill=fill(C_TITLE)
home.merge_cells("B2:Q3")
h=home.cell(2,2,"Community Partnerships"); h.font=Font(name=FONTNAME,size=28,bold=True,color=WHITE)
h.alignment=Alignment(horizontal="left",vertical="center",indent=2)
home.merge_cells("B4:Q4")
sst=home.cell(4,2,"Organika RTD    ·    Find partners, book activations, win the summer 2026 Costco road show.")
sst.font=Font(name=FONTNAME,size=12,bold=False,color="FFBFD2CB")
sst.alignment=Alignment(horizontal="left",vertical="center",indent=2)
for col in range(2,18): home.cell(5,col).fill=fill(ACCENT)
home.row_dimensions[2].height=38; home.row_dimensions[3].height=16; home.row_dimensions[4].height=26
home.row_dimensions[5].height=4; home.row_dimensions[6].height=16
# nav cards: white cards with a coloured left accent and an arrow cue
def card(c0, title, desc, target, clr):
    c1=c0+2
    for rr in range(7,11):
        for cc in range(c0,c1+1):
            cell=home.cell(rr,cc); cell.fill=fill(WHITE); sides={}
            if rr==7: sides["top"]=Side(style="thin",color=CARD_BRD)
            if rr==10: sides["bottom"]=Side(style="thin",color=CARD_BRD)
            if cc==c1: sides["right"]=Side(style="thin",color=CARD_BRD)
            if cc==c0: sides["left"]=Side(style="thick",color=clr)
            cell.border=Border(**sides)
    home.merge_cells(start_row=7,start_column=c0,end_row=8,end_column=c1)
    home.merge_cells(start_row=9,start_column=c0,end_row=10,end_column=c1)
    t=home.cell(7,c0,title+"    ›"); t.font=Font(name=FONTNAME,size=14,bold=True,color=C_TITLE)
    t.alignment=Alignment(horizontal="left",vertical="center",indent=2,wrap_text=True)
    t.hyperlink=Hyperlink(ref=t.coordinate, location="'%s'!A1"%target, display=title)
    d=home.cell(9,c0,desc); d.font=Font(name=FONTNAME,size=10,bold=False,color=C_SUB)
    d.alignment=Alignment(horizontal="left",vertical="top",indent=2,wrap_text=True)
    d.hyperlink=Hyperlink(ref=d.coordinate, location="'%s'!A1"%target, display=desc)
home.row_dimensions[7].height=22; home.row_dimensions[8].height=16; home.row_dimensions[9].height=18; home.row_dimensions[10].height=22
card(2,"Who to contact","Your worklist, most urgent first.","Action List","FF2E5A4E")
card(6,"Log an activation","Samples, budget and cans by flavour.","Activations","FF3E7C68")
card(10,"See the picture","Totals, charts and budget.","Dashboard","FF6FA392")
card(14,"Browse all partners","Every partner in one place.","Master List","FF8FB3A6")
# how it works
home.row_dimensions[12].height=12
hw=home.cell(13,2,"How it works"); hw.font=Font(name=FONTNAME,size=14,bold=True,color=C_TITLE)
hw.alignment=Alignment(horizontal="left",vertical="center",indent=2)
for i,sline in enumerate([
 "1.    Open the Action List to see who to reach out to first.",
 "2.    Work a partner on its green tab. Keep status and next steps current.",
 "3.    When something is booked, log it on the Activations tab. Everything else updates itself.",
]):
    cs=home.cell(14+i,2,sline); cs.font=font(11,False,C_DATA)
    cs.alignment=Alignment(horizontal="left",vertical="center",indent=2); home.row_dimensions[14+i].height=20
ft=home.cell(18,2,"You only ever type into the green partner tabs and the Activations tab. The other tabs read themselves.")
ft.font=font(11,False,C_SUB); ft.alignment=Alignment(horizontal="left",vertical="center",indent=2)
home.sheet_properties.tabColor="22413A"
