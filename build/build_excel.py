"""
Build MUV_RTD_Forecast_v1.xlsx - a fully formula-driven annual sales forecast.

Every result cell is a live formula tracing back to an editable input. The math
mirrors build/muv_config.py exactly (verified by LibreOffice recalc in verify.py).

Engine grain: week x channel x province.
  - Store tiers (A/B/C) are editable inputs that ROLL UP per channel via SUMIF
    (sum of doors*velocity), preserving exact totals.
  - Flavours/SKUs are handled by share-allocation with BLENDED cogs & price index
    (Total = cases * sum(share*cogs) is exact), so adding a flavour or SKU is
    literally adding a row to the Flavours/SKUs tables - no new columns.
"""
from __future__ import annotations
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList

import muv_config as C

OUT = "../MUV_RTD_Forecast_v1.xlsx"

# ------------------------------------------------------------------ styling
NAVY = "1F2A44"; BLUE = "2E5CB8"; TEAL = "0E7C7B"; LIGHT = "EAF0FB"
INPUT = "FFF6D5"; INPUTBRD = "E0C76A"; GREY = "F3F4F6"; GREEN = "C6EFCE"
GREENF = "006100"; RED = "FFC7CE"; REDF = "9C0006"; AMBER = "FFEB9C"
WHITE = "FFFFFF"; CALCHDR = "33415C"

thin = Side(style="thin", color="D0D5DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def font(sz=10, b=False, color="222B36", it=False):
    return Font(name="Calibri", size=sz, bold=b, color=color, italic=it)
def fill(c): return PatternFill("solid", fgColor=c)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def right(): return Alignment(horizontal="right", vertical="center")

CUR = '_-#,##0_-;[Red]-#,##0_-'           # currency, no decimals
CUR2 = '#,##0.00'
NUM1 = '#,##0.0'
NUM0 = '#,##0'
PCT1 = '0.0%'
RT3 = '0.000'
DATEF = 'yyyy-mm-dd'

wb = Workbook()

def dn(name, sheet, ref):
    """Add a workbook-global defined name -> 'Sheet'!ref ."""
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")

def hdr_cell(ws, r, c, text, bg=BLUE, fg=WHITE, sz=10, al=None):
    cell = ws.cell(r, c, text)
    cell.font = font(sz, True, fg); cell.fill = fill(bg)
    cell.alignment = al or center(); cell.border = border
    return cell

def title_row(ws, r, c, ncols, text, bg=NAVY):
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+ncols-1)
    cell = ws.cell(r, c, text)
    cell.font = font(12, True, WHITE); cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for cc in range(c, c+ncols):
        ws.cell(r, cc).fill = fill(bg)
    ws.row_dimensions[r].height = 22

def write_table(ws, top, left_c, title, headers, rows, *, col_fmt=None,
                input_cols=None, spare=0, note=None, widths=None):
    """Generic titled table. Returns info dict with column letters & row range."""
    col_fmt = col_fmt or {}
    input_cols = input_cols or set()
    n = len(headers)
    title_row(ws, top, left_c, n, title)
    hr = top + 1
    for j, h in enumerate(headers):
        hdr_cell(ws, hr, left_c + j, h)
    first = hr + 1
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            cell = ws.cell(first + i, left_c + j, val)
            cell.font = font(10)
            cell.border = border
            cell.alignment = right() if isinstance(val, (int, float)) else left()
            if j in col_fmt:
                cell.number_format = col_fmt[j]
            if j in input_cols:
                cell.fill = fill(INPUT)
    last = first + len(rows) - 1
    # spare editable rows (for "add a row" extensibility)
    for s in range(spare):
        for j in range(n):
            cell = ws.cell(last + 1 + s, left_c + j)
            cell.border = border
            if j in col_fmt:
                cell.number_format = col_fmt[j]
            if j in input_cols:
                cell.fill = fill(INPUT)
    last_with_spare = last + spare
    if note:
        nr = last_with_spare + 1
        ws.cell(nr, left_c, note).font = font(8, False, "667085", it=True)
    cols = {h: get_column_letter(left_c + j) for j, h in enumerate(headers)}
    if widths:
        for j, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(left_c + j)].width = w
    return {"hdr": hr, "first": first, "last": last, "last_spare": last_with_spare,
            "cols": cols, "title": top}

# ============================================================ SETTINGS sheet
ws = wb.active; ws.title = "Settings"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 16
for col in "DEF": ws.column_dimensions[col].width = 14
title_row(ws, 1, 1, 6, "MÜV RTD  •  Annual Sales Forecast — Settings & Global Drivers")
ws.cell(2, 1, "Yellow cells are inputs. Change any of them and the whole model recalculates.").font = font(9, False, "667085", it=True)

# Settings key/value block
srow = 4
hdr_cell(ws, srow, 1, "Setting"); hdr_cell(ws, srow, 2, "Value"); hdr_cell(ws, srow, 3, "Notes", bg=BLUE)
settings = [
    ("Fiscal year start date", C.SETTINGS["fiscal_start"], "FiscalStart", DATEF,
     "Week 1 begins here. FY ends end of August."),
    ("Number of weeks (52 or 53)", C.SETTINGS["n_weeks"], "NWeeks", NUM0,
     "Toggle 53-week year. Week 53 auto-activates."),
    ("Gross revenue basis", C.SETTINGS["gross_basis"], "GrossBasis", None,
     "Wholesale (default) or Sell-through. Moves the goal math the most."),
    ("$ Target (goal)", C.SETTINGS["target_revenue"], "TargetRev", CUR,
     "Annual gross revenue goal."),
    ("Calibration scalar (doors)", C.SETTINGS["calib_scalar"], "CalibScalar", '0.0000',
     "Scales all door counts. Calibrates base case into the $2.7–3.3M band."),
    ("Velocity uplift %", C.SETTINGS["velocity_uplift"], "VelocityUplift", PCT1,
     "Global +% on velocity (mirrors dashboard slider)."),
    ("Door ramp speed ×", C.SETTINGS["ramp_speed"], "RampSpeed", '0.00',
     "Multiplier on ramp progress (>1 = faster to full)."),
    ("Price index %", C.SETTINGS["price_index"], "PriceIndex", PCT1,
     "Global +% on all prices."),
    ("Seasonality amplitude ×", C.SETTINGS["season_amp"], "SeasonAmp", '0.00',
     "Scales seasonal swing around 1.0."),
    ("Online growth %", C.SETTINGS["online_growth"], "OnlineGrowth", PCT1,
     "Global +% on online units/week."),
]
r = srow + 1
for label, val, name, fmt, note in settings:
    ws.cell(r, 1, label).font = font(10, True); ws.cell(r, 1).border = border
    cell = ws.cell(r, 2, val); cell.fill = fill(INPUT); cell.border = border
    cell.font = font(10, True); cell.alignment = center()
    if fmt: cell.number_format = fmt
    ws.cell(r, 3, note).font = font(8, False, "667085", it=True)
    ws.cell(r, 3).alignment = left()
    dn(name, "Settings", f"$B${r}")
    r += 1
settings_end = r - 1

# Flavours
fl = write_table(ws, settings_end + 2, 1, "Flavours  (add a flavour = add a row; shares must sum to 100%)",
                 ["Flavour", "Share %"],
                 [[f["name"], f["share"]] for f in C.FLAVOURS],
                 col_fmt={1: PCT1}, input_cols={0, 1}, spare=2,
                 note="Shares are a fraction (0.40 = 40%). Check sum below = 100%.")
dn("FlavName", "Settings", f"${fl['cols']['Flavour']}${fl['first']}:${fl['cols']['Flavour']}${fl['last_spare']}")
dn("FlavShare", "Settings", f"${fl['cols']['Share %']}${fl['first']}:${fl['cols']['Share %']}${fl['last_spare']}")
chk_r = fl["last_spare"] + 2
ws.cell(chk_r, 1, "Flavour share check (=100%):").font = font(9, True)
chk = ws.cell(chk_r, 2, "=SUM(FlavShare)"); chk.number_format = PCT1; chk.font = font(9, True)
ws.conditional_formatting.add(f"B{chk_r}",
    CellIsRule(operator="equal", formula=["1"], fill=fill(GREEN), font=font(9, True, GREENF)))
ws.conditional_formatting.add(f"B{chk_r}",
    CellIsRule(operator="notEqual", formula=["1"], fill=fill(RED), font=font(9, True, REDF)))

# Pack formats (blended cogs/price computed from SKU table below)
pf_top = chk_r + 2
pf = write_table(ws, pf_top, 1,
    "Pack Formats",
    ["Format", "Cans/pack", "Packs/case", "Cans/case", "Blended COGS/case", "Blended price idx"],
    [["4-pack", 4, 6, 24, None, None],
     ["12-pack", 12, 2, 24, None, None]],
    col_fmt={1: NUM0, 2: NUM0, 3: NUM0, 4: CUR2, 5: '0.000'},
    input_cols={0, 1, 2, 3},
    note="Blended cols are computed from the SKUs table (share-weighted).")
dn("FmtName", "Settings", f"${pf['cols']['Format']}${pf['first']}:${pf['cols']['Format']}${pf['last']}")
dn("FmtCansPerPack", "Settings", f"${pf['cols']['Cans/pack']}${pf['first']}:${pf['cols']['Cans/pack']}${pf['last']}")
dn("FmtCansPerCase", "Settings", f"${pf['cols']['Cans/case']}${pf['first']}:${pf['cols']['Cans/case']}${pf['last']}")
dn("FmtBlendedCogs", "Settings", f"${pf['cols']['Blended COGS/case']}${pf['first']}:${pf['cols']['Blended COGS/case']}${pf['last']}")
dn("FmtBlendedPriceIdx", "Settings", f"${pf['cols']['Blended price idx']}${pf['first']}:${pf['cols']['Blended price idx']}${pf['last']}")

# SKUs
sk_top = pf["last"] + 3
sku_rows = [[s["flavour"], s["format"], s["cogs_per_case"], s["price_index"], None] for s in C.SKUS]
sk = write_table(ws, sk_top, 1,
    "SKUs  (flavour × format; add a SKU = add a row)",
    ["Flavour", "Format", "COGS/case", "Price idx", "Share (auto)"],
    sku_rows, col_fmt={2: CUR2, 3: '0.000', 4: PCT1}, input_cols={0, 1, 2, 3}, spare=4,
    note="COGS per 24-can case. Price idx = per-SKU price multiplier (1.00 = neutral). Share auto-pulls from Flavours.")
dn("SkuFlav", "Settings", f"${sk['cols']['Flavour']}${sk['first']}:${sk['cols']['Flavour']}${sk['last_spare']}")
dn("SkuFmt", "Settings", f"${sk['cols']['Format']}${sk['first']}:${sk['cols']['Format']}${sk['last_spare']}")
dn("SkuCogs", "Settings", f"${sk['cols']['COGS/case']}${sk['first']}:${sk['cols']['COGS/case']}${sk['last_spare']}")
dn("SkuPriceIdx", "Settings", f"${sk['cols']['Price idx']}${sk['first']}:${sk['cols']['Price idx']}${sk['last_spare']}")
dn("SkuShare", "Settings", f"${sk['cols']['Share (auto)']}${sk['first']}:${sk['cols']['Share (auto)']}${sk['last_spare']}")
# fill SKU share auto column (real data rows only; copy down when adding rows)
shcol = sk["cols"]["Share (auto)"]; flcol = sk["cols"]["Flavour"]
for i in range(sk["first"], sk["last"] + 1):
    cell = ws.cell(i, column_index_from_string(shcol))
    cell.value = f'=IFERROR(INDEX(FlavShare,MATCH({flcol}{i},FlavName,0)),"")'
    cell.number_format = PCT1
# blended cogs/price formulas in formats table (share-weighted over SKUs of that format)
fmtcol = pf["cols"]["Format"]
for frow in range(pf["first"], pf["last"] + 1):
    bc = ws.cell(frow, column_index_from_string(pf["cols"]["Blended COGS/case"]))
    bc.value = f'=SUMPRODUCT((SkuFmt={fmtcol}{frow})*SkuCogs*SkuShare)'
    bc.number_format = CUR2
    bp = ws.cell(frow, column_index_from_string(pf["cols"]["Blended price idx"]))
    bp.value = f'=SUMPRODUCT((SkuFmt={fmtcol}{frow})*SkuPriceIdx*SkuShare)'
    bp.number_format = '0.000'

# data validation on Settings
dv_basis = DataValidation(type="list", formula1='"Wholesale,Sell-through"', allow_blank=False)
dv_weeks = DataValidation(type="list", formula1='"52,53"', allow_blank=False)
ws.add_data_validation(dv_basis); ws.add_data_validation(dv_weeks)
dv_basis.add(ws.cell(srow + 3, 2))  # gross basis row
dv_weeks.add(ws.cell(srow + 2, 2))  # n weeks row
dv_fmt = DataValidation(type="list", formula1='"4-pack,12-pack"', allow_blank=True)
ws.add_data_validation(dv_fmt)
dv_fmt.add(f"{sk['cols']['Format']}{sk['first']}:{sk['cols']['Format']}{sk['last_spare']}")

print("Settings sheet done")

# ============================================================ ASSUMPTIONS sheet
wa = wb.create_sheet("Assumptions")
wa.sheet_view.showGridLines = False
title_row(wa, 1, 1, 7, "MÜV RTD  •  Assumptions — Channels, Tiers, Geography, Seasonality")
wa.cell(2, 1, "SEED PLACEHOLDERS (illustrative). Yellow = inputs. Replace with real numbers later.").font = font(9, False, "B23B3B", it=True)

# ---- Channels ----
ch_rows = [[c["name"], c["format"], c["launch"], c["weeks_to_full"], c["ramp"],
            (C.BASE_CP.get(c["name"]) if c["name"] != "Online" else None), None]
           for c in C.CHANNELS]
ch = write_table(wa, 4, 1,
    "Channels  (add a channel = add a row; logic is generic — no new code needed)",
    ["Channel", "Format", "Launch wk", "Weeks to full", "Ramp shape", "Base CP/case", "Doors×Vel (auto)"],
    ch_rows, col_fmt={2: NUM0, 3: NUM0, 5: CUR2, 6: NUM1},
    input_cols={0, 1, 2, 3, 4, 5}, spare=5,
    widths=[14, 9, 9, 11, 10, 11, 14],
    note="Base CP = wholesale case price before province modifier. Online uses Online price (Geography) instead. Doors×Vel auto-rolls from Tiers.")
for nm, key in [("Channel","ChName"),("Format","ChFormat"),("Launch wk","ChLaunch"),
                ("Weeks to full","ChWeeksFull"),("Ramp shape","ChRamp"),
                ("Base CP/case","ChBaseCP"),("Doors×Vel (auto)","ChDoorVel")]:
    col = ch["cols"][nm]
    dn(key, "Assumptions", f"${col}${ch['first']}:${col}${ch['last_spare']}")

# ---- Tiers ----
tier_rows = [[t["channel"], t["tier"], t["doors"], t["velocity"], None, None, None]
             for t in C.TIERS]
ti = write_table(wa, ch["last_spare"] + 3, 1,
    "Store Tiers per physical channel  (add a tier = add a row; rolls up into the channel above)",
    ["Channel", "Tier", "Doors @ full", "Velocity (cases/door/wk)", "Doors×Vel (auto)",
     "Cal. cases/wk (auto)", "% of channel (auto)"],
    tier_rows, col_fmt={2: NUM0, 3: NUM1, 4: NUM1, 5: NUM1, 6: PCT1},
    input_cols={0, 1, 2, 3}, spare=8,
    widths=[14, 6, 11, 14, 13, 13, 13],
    note="1 case = 24 cans. Base velocity unit = cases per door per week. Higher tiers sell faster.")
for nm, key in [("Channel","TierChannel"),("Tier","TierTier"),("Doors @ full","TierDoors"),
                ("Velocity (cases/door/wk)","TierVel"),("Doors×Vel (auto)","TierDoorVel")]:
    col = ti["cols"][nm]
    dn(key, "Assumptions", f"${col}${ti['first']}:${col}${ti['last_spare']}")
# tier auto columns
dvcol = ti["cols"]["Doors×Vel (auto)"]; dcol = ti["cols"]["Doors @ full"]; vcol = ti["cols"]["Velocity (cases/door/wk)"]
calcol = ti["cols"]["Cal. cases/wk (auto)"]; pctcol = ti["cols"]["% of channel (auto)"]; tchcol = ti["cols"]["Channel"]
for i in range(ti["first"], ti["last"] + 1):
    wa.cell(i, column_index_from_string(dvcol)).value = f'=IF({dcol}{i}="","",{dcol}{i}*{vcol}{i})'
    wa.cell(i, column_index_from_string(dvcol)).number_format = NUM1
    wa.cell(i, column_index_from_string(calcol)).value = f'=IF({dcol}{i}="","",{dvcol}{i}*CalibScalar*(1+VelocityUplift))'
    wa.cell(i, column_index_from_string(calcol)).number_format = NUM1
    wa.cell(i, column_index_from_string(pctcol)).value = f'=IFERROR({dvcol}{i}/SUMIF(TierChannel,{tchcol}{i},TierDoorVel),"")'
    wa.cell(i, column_index_from_string(pctcol)).number_format = PCT1
# channel Doors×Vel rollup (now that tier ranges exist)
chdvcol = ch["cols"]["Doors×Vel (auto)"]; chnmcol = ch["cols"]["Channel"]
for i in range(ch["first"], ch["last"] + 1):
    wa.cell(i, column_index_from_string(chdvcol)).value = f'=SUMIF(TierChannel,{chnmcol}{i},TierDoorVel)'
    wa.cell(i, column_index_from_string(chdvcol)).number_format = NUM1

# ---- Provinces / Geography ----
prov_rows = [[p["code"], p["modifier"], p["srp_4pack"], p["online_price"]] for p in C.PROVINCES]
pr = write_table(wa, ti["last_spare"] + 3, 1,
    "Geography — Provinces  (add a province/territory = add a row)",
    ["Province", "Price modifier", "Retail SRP / 4-pack", "Online price / 12-pack"],
    prov_rows, col_fmt={1: PCT1, 2: CUR2, 3: CUR2}, input_cols={0, 1, 2, 3}, spare=4,
    widths=[10, 13, 16, 18],
    note="Price modifier (±5%) adjusts wholesale CP by province. SRP = retail per 4-pack. Online price = retail per 12-pack.")
for nm, key in [("Province","ProvCode"),("Price modifier","ProvMod"),
                ("Retail SRP / 4-pack","ProvSRP"),("Online price / 12-pack","ProvOnline")]:
    col = pr["cols"][nm]
    dn(key, "Assumptions", f"${col}${pr['first']}:${col}${pr['last_spare']}")

# ---- Province split (long table channel|province) ----
split_rows = []
for c in C.CHANNELS:
    for p in C.PROVINCES:
        split_rows.append([c["name"], p["code"], C.PROVINCE_SPLIT[c["name"]][p["code"]] / 100.0, None])
ps = write_table(wa, pr["last_spare"] + 3, 1,
    "Province split per channel  (each channel's doors distributed across provinces; sums to 100% per channel)",
    ["Channel", "Province", "Split %", "Key (auto)"],
    split_rows, col_fmt={2: PCT1}, input_cols={0, 1, 2}, spare=20,
    widths=[14, 10, 10, 18],
    note="Split is a fraction (0.38 = 38%). Key auto = Channel|Province (used by the engine). Adding a province? add a row per channel.")
for nm, key in [("Channel","PSChannel"),("Province","PSProv"),("Split %","PSPct"),("Key (auto)","PSKey")]:
    col = ps["cols"][nm]
    dn(key, "Assumptions", f"${col}${ps['first']}:${col}${ps['last_spare']}")
pskeycol = ps["cols"]["Key (auto)"]; pschcol = ps["cols"]["Channel"]; psprcol = ps["cols"]["Province"]
for i in range(ps["first"], ps["last"] + 1):
    wa.cell(i, column_index_from_string(pskeycol)).value = f'={pschcol}{i}&"|"&{psprcol}{i}'
# split sum check per channel
sc_top = ps["last_spare"] + 2
wa.cell(sc_top, 1, "Split sum check (each must = 100%)").font = font(9, True)
for k, c in enumerate(C.CHANNELS):
    rr = sc_top + 1 + k
    wa.cell(rr, 1, c["name"]).font = font(9)
    cell = wa.cell(rr, 2, f'=SUMIF(PSChannel,"{c["name"]}",PSPct)')
    cell.number_format = PCT1; cell.font = font(9)
    wa.conditional_formatting.add(f"B{rr}",
        CellIsRule(operator="between", formula=["0.999", "1.001"], fill=fill(GREEN), font=font(9, True, GREENF)))
    wa.conditional_formatting.add(f"B{rr}",
        CellIsRule(operator="notBetween", formula=["0.999", "1.001"], fill=fill(RED), font=font(9, True, REDF)))

# ---- Seasonality (53 weeks) ----
seas_rows = [[w, C.SEASONALITY[w - 1]] for w in range(1, 54)]
se = write_table(wa, 4, 9,
    "Seasonality  (53 weekly indices, summer-peaked, average ≈ 1.0)",
    ["Week", "Index"], seas_rows, col_fmt={0: NUM0, 1: RT3}, input_cols={1},
    widths=[8, 10], note="Index 1.0 = average week. Engine applies (1+(idx-1)×amplitude).")
dn("SeasonIdx", "Assumptions", f"${se['cols']['Index']}${se['first']}:${se['cols']['Index']}${se['last']}")
# seasonality avg check
savg = se["last"] + 2
wa.cell(savg, 9, "Avg (wk1-52):").font = font(9, True)
cell = wa.cell(savg, 10, "=AVERAGE(INDEX(SeasonIdx,1):INDEX(SeasonIdx,52))"); cell.number_format = RT3; cell.font = font(9, True)

# ---- Online units curve (53 weeks) ----
on_rows = [[w, C.ONLINE_UNITS[w - 1]] for w in range(1, 54)]
on = write_table(wa, 4, 12,
    "Online 12-pack units/week  (ramps 50 → 600)",
    ["Week", "Units/wk"], on_rows, col_fmt={0: NUM0, 1: NUM0}, input_cols={1},
    widths=[8, 10], note="Total 12-packs ordered online per week (before province split, growth, seasonality).")
dn("OnlineUnitsCol", "Assumptions", f"${on['cols']['Units/wk']}${on['first']}:${on['cols']['Units/wk']}${on['last']}")

# data validation: ramp shape + format on channels table, format on provinces? (none)
dv_ramp = DataValidation(type="list", formula1='"Linear,S-curve"', allow_blank=True)
wa.add_data_validation(dv_ramp)
dv_ramp.add(f"{ch['cols']['Ramp shape']}{ch['first']}:{ch['cols']['Ramp shape']}{ch['last_spare']}")
dv_cfmt = DataValidation(type="list", formula1='"4-pack,12-pack"', allow_blank=True)
wa.add_data_validation(dv_cfmt)
dv_cfmt.add(f"{ch['cols']['Format']}{ch['first']}:{ch['cols']['Format']}{ch['last_spare']}")

print("Assumptions sheet done")

# ============================================================ CALC sheet (engine)
wc = wb.create_sheet("Calc")
wc.sheet_view.showGridLines = False
COLS = ["Week", "Date", "Month", "Quarter", "Channel", "Format", "Province", "Active",
        "Launch", "WksFull", "RampShape", "RampX", "Ramp", "Season", "ProvSplit",
        "DoorVelEff", "OnlineUnits", "BrandCases", "Cans", "ConsumerUnits", "UnitPrice",
        "Wholesale$", "Sellthru$", "Gross$", "COGS$", "Margin$"]
# header
for j, h in enumerate(COLS, start=1):
    hdr_cell(wc, 1, j, h, bg=CALCHDR)
wc.freeze_panes = "E2"
# column widths
widths = [6, 11, 10, 7, 12, 8, 8, 6, 6, 7, 9, 6, 6, 7, 8, 9, 9, 10, 9, 10, 9, 11, 11, 11, 10, 11]
for j, w in enumerate(widths, start=1):
    wc.column_dimensions[get_column_letter(j)].width = w
# letter map
L = {h: get_column_letter(j) for j, h in enumerate(COLS, start=1)}
fmt_cur = {"Wholesale$", "Sellthru$", "Gross$", "COGS$", "Margin$", "UnitPrice"}
fmt_num1 = {"BrandCases", "DoorVelEff", "OnlineUnits", "Cans", "ConsumerUnits"}
fmt_rt3 = {"RampX", "Ramp", "Season", "ProvSplit"}

combos = [(c["name"], c["name"] == "Online") for c in C.CHANNELS for p in C.PROVINCES]
prov_of = [p["code"] for c in C.CHANNELS for p in C.PROVINCES]

r = 2
for (cname, is_online), pcode in zip(combos, prov_of):
    for w in range(1, 54):
        f = {}
        f["Week"] = w
        f["Date"] = f'=FiscalStart+({L["Week"]}{r}-1)*7'
        f["Month"] = f'=TEXT({L["Date"]}{r},"mmm yyyy")'
        f["Quarter"] = f'="Q"&(INT(({L["Week"]}{r}-1)/13)+1)'
        f["Channel"] = cname
        f["Format"] = f'=IFERROR(INDEX(ChFormat,MATCH({L["Channel"]}{r},ChName,0)),"")'
        f["Province"] = pcode
        f["Active"] = f'=IF({L["Week"]}{r}<=NWeeks,1,0)'
        f["Launch"] = f'=INDEX(ChLaunch,MATCH({L["Channel"]}{r},ChName,0))'
        f["WksFull"] = f'=INDEX(ChWeeksFull,MATCH({L["Channel"]}{r},ChName,0))'
        f["RampShape"] = f'=INDEX(ChRamp,MATCH({L["Channel"]}{r},ChName,0))'
        f["RampX"] = f'=MIN(1,MAX(0,({L["Week"]}{r}-{L["Launch"]}{r}+1)*RampSpeed/{L["WksFull"]}{r}))'
        f["Ramp"] = (f'=IF({L["Week"]}{r}<{L["Launch"]}{r},0,'
                     f'IF({L["RampShape"]}{r}="S-curve",{L["RampX"]}{r}*{L["RampX"]}{r}*(3-2*{L["RampX"]}{r}),{L["RampX"]}{r}))')
        f["Season"] = f'=1+(INDEX(SeasonIdx,{L["Week"]}{r})-1)*SeasonAmp'
        f["ProvSplit"] = f'=IFERROR(INDEX(PSPct,MATCH({L["Channel"]}{r}&"|"&{L["Province"]}{r},PSKey,0)),0)'
        if is_online:
            f["DoorVelEff"] = 0
            f["OnlineUnits"] = (f'=INDEX(OnlineUnitsCol,{L["Week"]}{r})*{L["ProvSplit"]}{r}*(1+OnlineGrowth)'
                                f'*{L["Season"]}{r}*{L["Active"]}{r}*IF({L["Week"]}{r}>={L["Launch"]}{r},1,0)')
            f["BrandCases"] = (f'={L["OnlineUnits"]}{r}*INDEX(FmtCansPerPack,MATCH({L["Format"]}{r},FmtName,0))'
                               f'/INDEX(FmtCansPerCase,MATCH({L["Format"]}{r},FmtName,0))')
            f["UnitPrice"] = f'=INDEX(ProvOnline,MATCH({L["Province"]}{r},ProvCode,0))*(1+PriceIndex)'
            f["Wholesale$"] = f'={L["OnlineUnits"]}{r}*{L["UnitPrice"]}{r}*INDEX(FmtBlendedPriceIdx,MATCH({L["Format"]}{r},FmtName,0))'
            f["Sellthru$"] = f'={L["Wholesale$"]}{r}'
        else:
            f["DoorVelEff"] = f'=INDEX(ChDoorVel,MATCH({L["Channel"]}{r},ChName,0))*CalibScalar*(1+VelocityUplift)'
            f["OnlineUnits"] = 0
            f["BrandCases"] = f'={L["DoorVelEff"]}{r}*{L["ProvSplit"]}{r}*{L["Ramp"]}{r}*{L["Season"]}{r}*{L["Active"]}{r}'
            f["UnitPrice"] = (f'=INDEX(ChBaseCP,MATCH({L["Channel"]}{r},ChName,0))'
                              f'*(1+INDEX(ProvMod,MATCH({L["Province"]}{r},ProvCode,0)))*(1+PriceIndex)')
            f["Wholesale$"] = f'={L["BrandCases"]}{r}*{L["UnitPrice"]}{r}*INDEX(FmtBlendedPriceIdx,MATCH({L["Format"]}{r},FmtName,0))'
            f["Sellthru$"] = f'={L["ConsumerUnits"]}{r}*INDEX(ProvSRP,MATCH({L["Province"]}{r},ProvCode,0))*(1+PriceIndex)'
        f["Cans"] = f'={L["BrandCases"]}{r}*INDEX(FmtCansPerCase,MATCH({L["Format"]}{r},FmtName,0))'
        f["ConsumerUnits"] = f'={L["Cans"]}{r}/INDEX(FmtCansPerPack,MATCH({L["Format"]}{r},FmtName,0))'
        f["Gross$"] = f'=IF(GrossBasis="Wholesale",{L["Wholesale$"]}{r},{L["Sellthru$"]}{r})'
        f["COGS$"] = f'={L["BrandCases"]}{r}*INDEX(FmtBlendedCogs,MATCH({L["Format"]}{r},FmtName,0))'
        f["Margin$"] = f'={L["Gross$"]}{r}-{L["COGS$"]}{r}'
        for h in COLS:
            cell = wc.cell(r, column_index_from_string(L[h]), f[h])
            if h == "Date": cell.number_format = DATEF
            elif h in fmt_cur: cell.number_format = CUR
            elif h in fmt_num1: cell.number_format = NUM1
            elif h in fmt_rt3: cell.number_format = RT3
        r += 1
calc_last = r - 1
calc_spare = calc_last + 350   # room for added channels/provinces (blank rows sum to 0)
# Calc named ranges (data + spare)
for h in COLS:
    dn("Calc" + h.replace("$", "").replace("Units", "Units"), "Calc",
       f"${L[h]}$2:${L[h]}${calc_spare}")
print(f"Calc sheet done: {calc_last-1} rows")

# ============================================================ SUMMARY sheet
wm = wb.create_sheet("Summary")
wm.sheet_view.showGridLines = False
for col, wd in {"A": 26, "B": 16, "C": 3, "D": 14, "E": 14, "F": 9, "G": 3,
                "H": 16, "I": 14, "J": 9, "K": 3, "L": 12, "M": 14, "N": 9, "O": 3,
                "P": 7, "Q": 13, "R": 14, "S": 13, "T": 13}.items():
    wm.column_dimensions[col].width = wd
title_row(wm, 1, 1, 14, "MÜV RTD  •  Summary & Goal Tracking")

# ---- KPI panel ----
def kpi(r, label, formula, fmt=CUR, big=False, bold=True):
    lc = wm.cell(r, 1, label); lc.font = font(11 if big else 10, bold); lc.alignment = left()
    lc.border = border; lc.fill = fill(GREY if not big else LIGHT)
    vc = wm.cell(r, 2, formula)
    if fmt: vc.number_format = fmt
    vc.font = font(16 if big else 11, True, NAVY if big else "222B36"); vc.alignment = right()
    vc.border = border; vc.fill = fill(LIGHT if big else WHITE)
    return r

hdr_cell(wm, 3, 1, "KEY METRICS", bg=TEAL); hdr_cell(wm, 3, 2, "", bg=TEAL)
kpi(4, "Gross revenue basis", "=GrossBasis", fmt=None, bold=True)
kpi(5, "ANNUAL GROSS REVENUE", "=SUM(CalcGross)", big=True)
kpi(6, "Target (goal)", "=TargetRev")
kpi(7, "Gap to target", "=B5-B6")
kpi(8, "% to goal", "=B5/B6", fmt=PCT1)
kpi(9, "Status", '=IF(B5>=B6,"✓ ON / ABOVE TARGET","▼ BELOW TARGET — see levers")', fmt=None)
kpi(10, "Total volume (cases)", "=SUM(CalcBrandCases)", fmt=NUM0)
kpi(11, "Total volume (cans)", "=SUM(CalcCans)", fmt=NUM0)
kpi(12, "Total consumer units", "=SUM(CalcConsumerUnits)", fmt=NUM0)
kpi(13, "Avg weekly revenue", "=B5/NWeeks")
kpi(14, "Gross margin $", "=B5-SUM(CalcCOGS)")
kpi(15, "Gross margin %", "=B14/B5", fmt=PCT1)
kpi(16, "Wholesale total (ref)", "=SUM(CalcWholesale)")
kpi(17, "Sell-through total (ref)", "=SUM(CalcSellthru)")
# conditional format the goal cells: green >= target, amber otherwise
for cellref in ("B5", "B8", "B9"):
    wm.conditional_formatting.add(cellref,
        FormulaRule(formula=["$B$5>=$B$6"], fill=fill(GREEN), font=font(11, True, GREENF)))
    wm.conditional_formatting.add(cellref,
        FormulaRule(formula=["$B$5<$B$6"], fill=fill(AMBER), font=font(11, True, "7A5B00")))

# ---- breakdown helper ----
def breakdown(top, left_c, title, label_value_pairs, vfmt=CUR, pct=True, total=True):
    title_row(wm, top, left_c, 3 if pct else 2, title)
    hr = top + 1
    hdr_cell(wm, hr, left_c, "Item"); hdr_cell(wm, hr, left_c + 1, "Gross $")
    if pct: hdr_cell(wm, hr, left_c + 2, "% mix")
    rr = hr + 1
    first = rr
    for label, formula in label_value_pairs:
        wm.cell(rr, left_c, label).font = font(10); wm.cell(rr, left_c).border = border
        wm.cell(rr, left_c).alignment = left()
        v = wm.cell(rr, left_c + 1, formula); v.number_format = vfmt; v.border = border; v.font = font(10)
        if pct:
            pc = wm.cell(rr, left_c + 2, f"={get_column_letter(left_c+1)}{rr}/$B$5")
            pc.number_format = PCT1; pc.border = border; pc.font = font(10)
        rr += 1
    last = rr - 1
    if total:
        wm.cell(rr, left_c, "Total").font = font(10, True); wm.cell(rr, left_c).border = border
        tc = wm.cell(rr, left_c + 1, f"=SUM({get_column_letter(left_c+1)}{first}:{get_column_letter(left_c+1)}{last})")
        tc.number_format = vfmt; tc.font = font(10, True); tc.border = border; tc.fill = fill(GREY)
        if pct:
            pc = wm.cell(rr, left_c + 2, f"={get_column_letter(left_c+1)}{rr}/$B$5"); pc.number_format = PCT1
            pc.font = font(10, True); pc.border = border; pc.fill = fill(GREY)
    return {"first": first, "last": last, "lab": get_column_letter(left_c),
            "val": get_column_letter(left_c + 1)}

# By channel
chan_pairs = [(c["name"], f'=SUMIF(CalcChannel,"{c["name"]}",CalcGross)') for c in C.CHANNELS]
b_chan = breakdown(3, 4, "By Channel", chan_pairs)
# By province
prov_pairs = [(p["code"], f'=SUMIF(CalcProvince,"{p["code"]}",CalcGross)') for p in C.PROVINCES]
b_prov = breakdown(b_chan["last"] + 3, 4, "By Province", prov_pairs)
# By flavour (= annual gross x share)
flav_pairs = [(f["name"], f'=$B$5*INDEX(FlavShare,MATCH("{f["name"]}",FlavName,0))') for f in C.FLAVOURS]
b_flav = breakdown(3, 8, "By Flavour", flav_pairs)
# By SKU (= gross by format x flavour share)
sku_pairs = []
for f in C.FLAVOURS:
    for fmt in ("4-pack", "12-pack"):
        sku_pairs.append((f'{f["name"]} {fmt}',
            f'=SUMIF(CalcFormat,"{fmt}",CalcGross)*INDEX(FlavShare,MATCH("{f["name"]}",FlavName,0))'))
b_sku = breakdown(b_flav["last"] + 3, 8, "By SKU (flavour × format)", sku_pairs)
# By quarter
q_pairs = [(f"Q{q}", f'=SUMIF(CalcQuarter,"Q{q}",CalcGross)') for q in (1, 2, 3, 4)]
b_q = breakdown(3, 12, "By Fiscal Quarter", q_pairs)
# By month
months = []
for w in range(1, 54):
    d = C.SETTINGS["fiscal_start"] + dt.timedelta(days=7 * (w - 1))
    lab = d.strftime("%b %Y")
    if lab not in months: months.append(lab)
mon_pairs = [(m, f'=SUMIF(CalcMonth,"{m}",CalcGross)') for m in months]
b_mon = breakdown(b_q["last"] + 3, 12, "By Calendar Month", mon_pairs, total=False)

# ---- Weekly trend (chart data) ----
wt_top = 3
title_row(wm, wt_top, 16, 5, "Weekly Trend vs Goal")
hr = wt_top + 1
for j, h in enumerate(["Week", "Weekly $", "Cumulative $", "$3M Goal", "Goal pace"]):
    hdr_cell(wm, hr, 16 + j, h)
wk_first = hr + 1
for i in range(53):
    rr = wk_first + i
    w = i + 1
    wm.cell(rr, 16, w).font = font(9); wm.cell(rr, 16).border = border
    c1 = wm.cell(rr, 17, f"=SUMIF(CalcWeek,P{rr},CalcGross)")
    c1.number_format = CUR; c1.font = font(9); c1.border = border
    cum = wm.cell(rr, 18,
        (f"=R{rr-1}+Q{rr}" if i > 0 else f"=Q{rr}")); cum.number_format = CUR; cum.font = font(9); cum.border = border
    tgt = wm.cell(rr, 19, "=TargetRev"); tgt.number_format = CUR; tgt.font = font(9, False, "B23B3B"); tgt.border = border
    pace = wm.cell(rr, 20, f"=TargetRev*P{rr}/NWeeks"); pace.number_format = CUR; pace.font = font(9, False, "999999"); pace.border = border
wk_last = wk_first + 52
print("Summary sheet done")

# ============================================================ CHARTS sheet
wch = wb.create_sheet("Charts")
wch.sheet_view.showGridLines = False
title_row(wch, 1, 1, 10, "MÜV RTD  •  Charts")

def add_bar(title, lab_col, val_col, first, last, anchor, color=BLUE, horizontal=False):
    chart = BarChart(); chart.type = "bar" if horizontal else "col"
    chart.title = title; chart.legend = None; chart.height = 7.5; chart.width = 13
    data = Reference(wm, min_col=column_index_from_string(val_col), min_row=first, max_row=last)
    cats = Reference(wm, min_col=column_index_from_string(lab_col), min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.y_axis.numFmt = '#,##0'; chart.y_axis.majorGridlines = None
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = color
    wch.add_chart(chart, anchor)

def add_pie(title, lab_col, val_col, first, last, anchor):
    chart = PieChart(); chart.title = title; chart.height = 7.5; chart.width = 9
    data = Reference(wm, min_col=column_index_from_string(val_col), min_row=first, max_row=last)
    cats = Reference(wm, min_col=column_index_from_string(lab_col), min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=False); chart.set_categories(cats)
    chart.dataLabels = DataLabelList(); chart.dataLabels.showPercent = True
    wch.add_chart(chart, anchor)

# Cumulative vs goal (line)
line = LineChart(); line.title = "Cumulative Gross vs $3M Goal"; line.height = 8; line.width = 18
line.y_axis.numFmt = '#,##0'
data = Reference(wm, min_col=18, min_row=wt_top + 1, max_col=20, max_row=wk_last)  # Cum, Goal, Pace + header
line.add_data(data, titles_from_data=True)
line.set_categories(Reference(wm, min_col=16, min_row=wk_first, max_row=wk_last))
line.x_axis.title = "Week"; line.y_axis.title = "CAD"
for s, col in zip(line.series, [TEAL, "B23B3B", "BBBBBB"]):
    s.graphicalProperties.line.solidFill = col
    s.graphicalProperties.line.width = 28000
line.series[1].graphicalProperties.line.dashStyle = "dash"
line.series[2].graphicalProperties.line.dashStyle = "sysDot"
wch.add_chart(line, "A3")

# Weekly revenue (col)
wbar = BarChart(); wbar.type = "col"; wbar.title = "Weekly Gross Revenue"; wbar.legend = None
wbar.height = 8; wbar.width = 18; wbar.y_axis.numFmt = '#,##0'
wbar.add_data(Reference(wm, min_col=17, min_row=wk_first, max_row=wk_last), titles_from_data=False)
wbar.set_categories(Reference(wm, min_col=16, min_row=wk_first, max_row=wk_last))
if wbar.series: wbar.series[0].graphicalProperties.solidFill = BLUE
wch.add_chart(wbar, "A20")

add_bar("Gross by Channel", b_chan["lab"], b_chan["val"], b_chan["first"], b_chan["last"], "L3", color=BLUE, horizontal=True)
add_bar("Gross by Province", b_prov["lab"], b_prov["val"], b_prov["first"], b_prov["last"], "L20", color=TEAL, horizontal=True)
add_pie("Flavour Mix", b_flav["lab"], b_flav["val"], b_flav["first"], b_flav["last"], "U3")
add_bar("Gross by SKU", b_sku["lab"], b_sku["val"], b_sku["first"], b_sku["last"], "U20", color="6A4FB3", horizontal=True)
add_bar("Gross by Month", b_mon["lab"], b_mon["val"], b_mon["first"], b_mon["last"], "L37", color="0E7C7B")
add_pie("Channel Mix", b_chan["lab"], b_chan["val"], b_chan["first"], b_chan["last"], "U37")
print("Charts sheet done")

# ============================================================ README sheet
wr = wb.create_sheet("ReadMe")
wr.sheet_view.showGridLines = False
wr.column_dimensions["A"].width = 4
wr.column_dimensions["B"].width = 120
title_row(wr, 1, 1, 2, "MÜV RTD Forecast — How to use & extend")
readme = [
    ("h", "WHAT THIS IS"),
    ("p", "A fully formula-driven annual sales forecast for MÜV (sparkling electrolyte RTD), launching in Canada FY2027 (starts 1 Sep 2027)."),
    ("p", "Every result is a live formula tracing back to an input cell. Change any yellow input and the whole model recalculates."),
    ("h", "TABS"),
    ("p", "Settings   – fiscal start, 52/53 weeks, gross-revenue basis, $3M target, calibration scalar, and 5 global driver levers; Flavours, Pack formats, SKUs."),
    ("p", "Assumptions– Channels, Store tiers, Geography (provinces), Province split, Seasonality, Online units curve. SEED PLACEHOLDERS — replace with real data."),
    ("p", "Calc       – the engine: one row per Week × Channel × Province (53 × 70 = 3,710 rows). Tiers roll up; flavours/SKUs split by share."),
    ("p", "Summary    – KPIs, the $3M goal, and breakdowns by channel / province / flavour / SKU / quarter / month, plus the weekly trend."),
    ("p", "Charts     – cumulative-vs-goal, weekly revenue, channel/province/SKU/month bars, flavour & channel mix."),
    ("h", "GROSS REVENUE BASIS  (Settings!B6)"),
    ("p", "Wholesale (default) = case price × cases (physical) + online direct revenue. Sell-through = consumer units × retail SRP (physical) + online direct."),
    ("p", "This single choice moves the goal math the most. The label on the Summary always states the active basis."),
    ("h", "HOW TO EXTEND  (add a row — formulas follow)"),
    ("p", "• Add a FLAVOUR: add a row to Settings→Flavours (name + share; keep shares summing to 100%) AND a row per format in Settings→SKUs. Breakdowns update."),
    ("p", "• Add a SKU: add a row to Settings→SKUs (flavour, format, COGS, price idx). Blended COGS/price recompute automatically."),
    ("p", "• Add a CHANNEL: add a row to Assumptions→Channels (+ its tier rows + a province-split row per province). Then copy a 53-week block in Calc for each new"),
    ("p", "  channel×province and add a 'By Channel' row in Summary. Channel logic is generic — no formula rewrite needed."),
    ("p", "• Add a TIER: add a row to Assumptions→Store Tiers (channel, tier, doors, velocity). It auto-rolls into that channel's Doors×Vel. Nothing else to touch."),
    ("p", "• Add a PROVINCE/TERRITORY: add a row to Assumptions→Geography, a split row per channel in Province split, then Calc blocks + a Summary 'By Province' row."),
    ("p", "Spare yellow rows are pre-formatted under each table and the named ranges already extend over them, so SUMIF/SUMPRODUCT pick up new rows automatically."),
    ("p", "When you fill a spare row, copy the grey '(auto)' formula cells (e.g. Doors×Vel, Key, Share) down from the row above so the derived columns populate."),
    ("h", "CALIBRATION"),
    ("p", "Raw seed anchors overshoot $3M (~$7.1M). The calibration scalar (Settings) scales all door counts. It is set to 0.36 so the base case lands just under"),
    ("p", "the goal (~$2.96M, 98.7%) — inside the $2.7–3.3M band — leaving a real gap so the goal tracking and levers are meaningful. Edit it freely."),
    ("h", "GLOBAL LEVERS (Settings) — mirror the HTML dashboard sliders"),
    ("p", "Velocity uplift %, Door ramp speed ×, Price index %, Seasonality amplitude ×, Online growth %. All neutral by default."),
    ("h", "QA INVARIANTS (built in)"),
    ("p", "Province splits sum to 100% per channel (check block on Assumptions). Flavour shares sum to 100% (check on Settings). Ramp = 0 before launch, exactly"),
    ("p", "1.0 at full. No negative or pre-launch volume. 1 case = 24 cans everywhere. Excel totals match the HTML dashboard within rounding."),
]
rr = 3
for kind, text in readme:
    cell = wr.cell(rr, 2, text)
    if kind == "h":
        cell.font = font(11, True, BLUE)
    else:
        cell.font = font(10); cell.alignment = left()
    rr += 1
print("ReadMe sheet done")

# order: open on Summary
wb.active = wb.sheetnames.index("Summary")
wb.save(OUT)
print(f"SAVED {OUT}")
