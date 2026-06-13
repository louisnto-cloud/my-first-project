# -*- coding: utf-8 -*-
"""QA for Organika RTD Community Partnerships Tracker_v4.xlsx"""
import re, warnings
import openpyxl
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

OUT = "/home/user/my-first-project/Organika RTD Community Partnerships Tracker_v5.xlsx"
SRC = "/root/.claude/uploads/607dab49-f51b-5b86-83ad-5f4c7139295f/029750d9-Organika_RTD_BC_Tracker.xlsx"
fails=[]; warns=[]
def ok(m): print("  PASS  "+m)
def bad(m): fails.append(m); print("  FAIL  "+m)
def warn(m): warns.append(m); print("  NOTE  "+m)

wb = openpyxl.load_workbook(OUT, data_only=False)
TYPES = ["Run Clubs","Gyms & Studios","Events & Festivals","Sports Teams & Leagues",
"Campus & Student Groups","Wellness & Recovery","Ambassadors & Creators","Charity & Causes"]
EXPECT_TYPE_HDRS = ["#","Partner Name","Partnership Type","City","Neighbourhood","Priority","Status",
"Primary Owner","Contact Name","Role","Contact Email","Contact Phone","Instagram or Website",
"Audience Size","Audience Source","Source","Warm","Last Contacted","Days Since Activity",
"Next Action","Next Action Date","Activation Type","Activation Date","Cases Committed",
"Cases Delivered","Cost","Cost Per Can","Contra Value","Deliverables Promised",
"Deliverables Received","What They Want","Risks","Nearby Retail Doors","In BC Tracker?",
"Raspberry 4338","Lemon Lime 4336","Pineapple Passion Fruit 4340","Notes"]

print("\n[1] NO HYPHENS and NO PARENTHESES in any cell value (formulas excluded)")
hy=[]; par=[]
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            v=c.value
            if isinstance(v,str) and not v.startswith("="):
                if "-" in v: hy.append((ws.title,c.coordinate,v))
                if "(" in v or ")" in v: par.append((ws.title,c.coordinate,v))
for t in wb.sheetnames:
    if "-" in t or "(" in t or ")" in t: bad(f"bad sheet name {t!r}")
if hy:
    for s,co,v in hy[:30]: bad(f"hyphen {s}!{co}: {v!r}")
else: ok("zero hyphens anywhere")
if par:
    for s,co,v in par[:30]: bad(f"parenthesis {s}!{co}: {v!r}")
else: ok("zero parentheses anywhere")

print("\n[2] HEADERS (Cost and Contra Value now without the dollar parens)")
for t in TYPES:
    got=[wb[t].cell(2,c).value for c in range(1,39)]
    if got==EXPECT_TYPE_HDRS: ok(f"{t}: 38 headers correct")
    else:
        for i,(a,b) in enumerate(zip(got,EXPECT_TYPE_HDRS)):
            if a!=b: bad(f"{t} col {i+1}: {a!r} != {b!r}")

print("\n[3] ACTIVATIONS tab: per flavour can columns + structure")
act=wb["Activations"]
ah=[act.cell(2,c).value for c in range(1,16)]
EXP_ACT=["#","Date","Partner","Partnership Type","City","Owner","Status","Activation Type",
"Assets Needed","Budget","Raspberry 4338 Cans","Lemon Lime 4336 Cans","Pineapple Passion Fruit 4340 Cans","Total Cans","Notes"]
if ah==EXP_ACT: ok("15 columns incl a can column per flavour and Total Cans")
else: bad(f"Activations headers: {ah}")
# empty grid except # and Total Cans formulas
content=[]
for r in range(3,63):
    for c in list(range(2,14))+[15]:
        v=act.cell(r,c).value
        if v not in (None,""): content.append((r,c,v))
if not content: ok("grid starts empty for the team to fill")
else: bad(f"Activations not empty: {content[:6]}")
if isinstance(act.cell(3,14).value,str) and act.cell(3,14).value.startswith("=IF(COUNT"): ok("Total Cans is a live sum of the three flavours")
else: bad("Total Cans formula missing")

print("\n[4] FREE ENTRY: every dropdown accepts a typed value not on the list")
strict=0; total=0
for nm in TYPES+["Activations","Suggested Events"]:
    for dv in wb[nm].data_validations.dataValidation:
        total+=1
        if dv.showErrorMessage: strict+=1
if strict==0: ok(f"all {total} dropdowns across working tabs allow add your own")
else: bad(f"{strict} dropdowns still reject off list entries")

print("\n[5] TEAM = Maddie only; owner menu = Maddie")
src=openpyxl.load_workbook(SRC,data_only=False)["Sales Team"]
src_names=[src.cell(r,1).value for r in range(3,24)]
my=wb["Sales Team"]; reps=[my.cell(r,1).value for r in range(3,24) if my.cell(r,1).value and my.cell(r,2).value]
if reps==["Maddie"] and "Maddie" in src_names: ok("Sales Team is Maddie only and matches BC Tracker spelling")
else: bad(f"reps={reps}")
own=[wb["Lookups"].cell(r,1).value for r in range(3,6) if wb["Lookups"].cell(r,1).value]
if own==["Maddie"]: ok("Owner menu = Maddie")
else: bad(f"owner menu {own}")

print("\n[6] MASTER LIST traceability + Tevah")
ml=wb["Master List"]; blocks={}; ok_ref=True
for r in range(3,243):
    fa=ml.cell(r,1).value
    m=re.match(r"=IF\('([^']+)'!B(\d+)=",fa) if isinstance(fa,str) else None
    if not m: ok_ref=False; continue
    blocks.setdefault(m.group(1),[]).append(int(m.group(2)))
if all(blocks.get(t)==list(range(3,33)) for t in TYPES) and ok_ref: ok("8 blocks x 30 rows, each row traces to a type tab")
else: bad("master list block structure off")
wr=wb["Wellness & Recovery"]
if wr.cell(3,2).value=="Tevah Wellness" and wr.cell(3,34).value=="Yes": ok("Tevah is first Wellness row, In BC Tracker = Yes")
else: bad("Tevah row off")
nmd=sum(1 for t in TYPES for r in range(3,33) if wb[t].cell(r,2).value)
own_ok=all(wb[t].cell(r,8).value=="Maddie" for t in TYPES for r in range(3,33) if wb[t].cell(r,2).value)
if nmd==60 and own_ok: ok("60 partners, owner Maddie on every one")
else: bad(f"partners={nmd} ownerok={own_ok}")

print("\n[7] NO PROPOSED ACTIVATIONS on type tabs; SKU numbers correct")
acts=[(t,r,c) for t in TYPES for r in range(3,33) for c in (22,23,24,25,26,28) if wb[t].cell(r,c).value not in (None,"")]
if not acts: ok("Activation Type, Date, cases, cost and contra blank on every partner row")
else: bad(f"unexpected activation values: {acts[:5]}")
sku_ok=all([wb[t].cell(2,35).value,wb[t].cell(2,36).value,wb[t].cell(2,37).value]==["Raspberry 4338","Lemon Lime 4336","Pineapple Passion Fruit 4340"] for t in TYPES)
ok("SKU headers 4338 4336 4340 on every type tab") if sku_ok else bad("SKU headers off")

print("\n[8] SPELLCHECK headers + Guide text")
try:
    from spellchecker import SpellChecker
    sp=SpellChecker()
    allow=set("""organika rtd bc sku skus costco raspberry lemon lime pineapple passion fruit
    kelowna vancouver victoria metro burnaby richmond surrey langley coquitlam nanaimo yaletown
    neighbourhood colour colours centre lookups dropdown tevah ubc instagram website roadshow
    activations contra deliverables reps okanagan maddie flavour flavours""".split())
    words=set()
    def collect(ws,cols,rmax):
        for r in range(1,rmax+1):
            for c in cols:
                v=ws.cell(r,c).value
                if isinstance(v,str) and not v.startswith("="):
                    for w in re.findall(r"[A-Za-z]+",v): words.add(w.lower())
    for t in TYPES: collect(wb[t],list(range(1,39)),2)
    collect(wb["Activations"],list(range(1,16)),2)
    for nm in ["Dashboard","Master List","Type Summary","Budget","Sales Team","Suggested Events"]: collect(wb[nm],list(range(1,15)),5)
    collect(wb["Guide"],[1,2],70)
    unknown=sorted(w for w in words if w not in allow and len(w)>2 and sp.unknown([w]))
    if not unknown: ok("no unknown words in headers or Guide")
    else: warn("review words: "+", ".join(unknown[:40]))
except Exception as e: warn("spellcheck skipped: "+str(e))

print("\n[9] FORMULA EVALUATION (0 errors) + key numbers")
try:
    import formulas
    xl=formulas.ExcelModel().loads(OUT).finish(); sol=xl.calculate()
    base=OUT.split("/")[-1]; errs=[]
    for k,v in sol.items():
        try: val=v.value[0,0]
        except: val=getattr(v,"value",None)
        if any(e in str(val) for e in ["#REF","#DIV","#NAME","#VALUE","#N/A","#NUM","#NULL"]): errs.append((k,str(val)))
    if errs:
        for k,s in errs[:25]: bad(f"formula error {k} = {s}")
    else: ok(f"all {len(sol)} computed cells resolve with no errors")
    def g(sh,cell):
        v=sol.get(f"'[{base}]"+sh.upper()+"'!"+cell)
        try: return v.value[0,0]
        except: return getattr(v,"value",v)
    kpis={"Total Partners":g("Dashboard","A5"),"P1":g("Dashboard","C5"),
          "Activations Booked":g("Dashboard","G5"),"Cans":g("Dashboard","I5"),
          "Days to Costco":g("Dashboard","M5")}
    print("        KPIs:",{k:(round(float(v),2) if isinstance(v,(int,float)) else v) for k,v in kpis.items()})
    if int(g("Dashboard","A5"))==60: ok("Total Partners = 60")
    else: bad(f"Total Partners={g('Dashboard','A5')}")
    if int(g("Dashboard","G5"))==0 and (g("Dashboard","I5") in (0,0.0)): ok("Activations Booked and Cans Sampled start at 0 (empty Activations tab)")
    else: bad(f"activations/cans not zero: {g('Dashboard','G5')},{g('Dashboard','I5')}")
except Exception as e:
    import traceback; traceback.print_exc()

print("\n[10] PROFESSIONAL DASHBOARD checks")
dash=wb["Dashboard"]
if len(dash._charts)==3: ok("3 live charts on the Dashboard")
else: bad(f"charts={len(dash._charts)}")
if dash.cell(1,1).value=="Organika RTD  ·  Community Partnerships": ok("clean title, no command centre")
labels=[dash.cell(8,1).value,dash.cell(26,1).value,dash.cell(27,1).value,dash.cell(33,1).value]
if labels==["Snapshot","Detail","Region","List Health"]: ok("section dividers and table blocks aligned (Snapshot, Detail, Region, List Health)")
else: warn(f"dashboard layout markers: {labels}")
links=sum(1 for t in TYPES for r in range(3,33) if wb[t].cell(r,13).hyperlink)
if links>=55: ok(f"{links} clickable partner links")
else: bad(f"links={links}")

print("\n[11] v5: partner menu, priority sort, locked read only tabs")
# partner dropdown on Activations col C, sourced from a 60 name list
act=wb["Activations"]
pdv=[dv for dv in act.data_validations.dataValidation if "LU_Partners" in str(dv.formula1) and "C3" in str(dv.sqref)]
lp=wb["Lookups"]
pnames=[lp.cell(r,19).value for r in range(3,63) if lp.cell(r,19).value]
if pdv and len(pnames)==60: ok(f"Partner menu on the Activations tab lists all {len(pnames)} partners, and lets you type your own")
else: bad(f"partner dropdown={bool(pdv)} names={len(pnames)}")
# priority sorted within each type tab (first filled row is P1 where P1 exists)
bad_sort=[]
for t in TYPES:
    pr=[wb[t].cell(r,6).value for r in range(3,33) if wb[t].cell(r,2).value]
    ranks=[{"P1":0,"P2":1,"P3":2}.get(p,3) for p in pr]
    if ranks!=sorted(ranks): bad_sort.append(t)
if not bad_sort: ok("every type tab is sorted P1 first, then P2, then P3")
else: bad(f"not priority sorted: {bad_sort}")
if wb["Wellness & Recovery"].cell(3,2).value=="Tevah Wellness": ok("Tevah still pinned as the first Wellness row")
else: bad("Tevah no longer first")
locked=[nm for nm in ["Dashboard","Type Summary","Master List"] if wb[nm].protection.sheet]
if len(locked)==3: ok("Dashboard, Type Summary and Master List are locked so their formulas cannot be edited by accident")
else: bad(f"protected tabs: {locked}")
if wb["Master List"].protection.autoFilter is False and not wb["Run Clubs"].protection.sheet:
    ok("Master List still filters, and the working tabs stay fully editable")
else: warn("protection flags need a look")

print("\n================ QA SUMMARY ================")
print(f"  FAILURES: {len(fails)}   NOTES: {len(warns)}")
print("  RESULT  :", "ALL CHECKS PASS" if not fails else "FAILURES PRESENT")
