
# =====================================================================
# START HERE  (the front door: simple, beautiful, anyone can use it)
# =====================================================================
print("building start here...")
from openpyxl.worksheet.hyperlink import Hyperlink
home = wb.create_sheet("Start Here")
home.sheet_view.showGridLines = False
for c in range(1,18): home.column_dimensions[get_column_letter(c)].width = 9.5
home.column_dimensions["A"].width = 3
# hero band
for col in range(2,18):
    home.cell(2,col).fill=fill(C_TITLE); home.cell(3,col).fill=fill(C_TITLE); home.cell(4,col).fill=fill(C_HEADER)
home.merge_cells("B2:Q3")
h=home.cell(2,2,"Community Partnerships"); h.font=Font(name="Arial",size=28,bold=True,color=WHITE)
h.alignment=Alignment(horizontal="left",vertical="center",indent=1)
home.merge_cells("B4:Q4")
sst=home.cell(4,2,"Organika RTD  ·  Find partners, book activations, win the summer 2026 Costco road show.")
sst.font=Font(name="Arial",size=12,bold=False,color=WHITE); sst.alignment=Alignment(horizontal="left",vertical="center",indent=1)
home.row_dimensions[2].height=34; home.row_dimensions[3].height=18; home.row_dimensions[4].height=26
# nav cards
def card(c0, title, desc, target, clr):
    c1=c0+2
    for rr in range(6,10):
        for cc in range(c0,c1+1): home.cell(rr,cc).fill=fill(clr); home.cell(rr,cc).border=BORD
    home.merge_cells(start_row=6,start_column=c0,end_row=7,end_column=c1)
    home.merge_cells(start_row=8,start_column=c0,end_row=9,end_column=c1)
    t=home.cell(6,c0,title); t.font=Font(name="Arial",size=14,bold=True,color=WHITE)
    t.alignment=Alignment(horizontal="left",vertical="center",indent=1,wrap_text=True)
    t.hyperlink=Hyperlink(ref=t.coordinate, location="'%s'!A1"%target, display=title)
    d=home.cell(8,c0,desc); d.font=Font(name="Arial",size=10,bold=False,color=WHITE)
    d.alignment=Alignment(horizontal="left",vertical="top",indent=1,wrap_text=True)
    d.hyperlink=Hyperlink(ref=d.coordinate, location="'%s'!A1"%target, display=desc)
home.row_dimensions[6].height=24; home.row_dimensions[7].height=20; home.row_dimensions[8].height=18; home.row_dimensions[9].height=22
card(2,"Who to contact","Your worklist, most urgent first.","Action List",C_HEADER)
card(6,"Log an activation","Samples, budget and cans by flavour.","Activations","3E7C68")
card(10,"See the picture","Totals, charts and budget.","Dashboard","6FA392")
card(14,"Browse all partners","Every partner in one place.","Master List","8FB3A6")
# how it works
hw=home.cell(12,2,"How it works"); hw.font=Font(name="Arial",size=13,bold=True,color=C_TITLE)
for i,sline in enumerate([
 "1.   Open the Action List to see who to reach out to first.",
 "2.   Work a partner on its green tab. Keep status and next steps current.",
 "3.   When something is booked, log it on the Activations tab. Everything else updates itself.",
]):
    cs=home.cell(13+i,2,sline); cs.font=font(11,False,C_DATA); cs.alignment=A_L; home.row_dimensions[13+i].height=18
home.cell(17,2,"You only ever type into the green partner tabs and the Activations tab. The other tabs read themselves.").font=font(11,False,C_SUB)
home.sheet_properties.tabColor="2E5A4E"
