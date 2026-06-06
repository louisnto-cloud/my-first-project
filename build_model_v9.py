"""
Costco UK D93 Portfolio ROI Model — v9 (enterprise)

Vs v8:
  - Cumulative tab (running totals, break-even week per scenario, line chart).
  - Sensitivity tab (Lift × Demo Cost grid showing Net Profit).
  - Pitch Memo tab (auto-generated exec summary pulling live numbers).
  - Probability weights per scenario + Expected-Value row on Summary.
  - Bar chart on Summary (Net Profit by scenario).
  - Portfolio: SP/Cost stored in GBP (source of truth); display × local FX.
  - Portfolio: 3-year view (Y1 / Y2 / Y3) with per-SKU growth rate.
  - Data validation on key inputs; conditional formatting on Portfolio ROI.
  - README expanded with change log, owner, assumptions, limitations.
"""

import openpyxl
from datetime import date
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList

# Palette
NAVY="1F4E78"; BLUE="2E75B6"; LIGHT_BLUE="DDEBF7"
YELLOW="FFF2CC"; GREY="F2F2F2"; WHITE="FFFFFF"
GREEN="C6EFCE"; RED="FFC7CE"; DARK_GREY="595959"
AMBER="FFE699"

thin = Side(border_style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

f_title       = Font(name="Calibri", size=18, bold=True, color=WHITE)
f_section     = Font(name="Calibri", size=12, bold=True, color=WHITE)
f_subhead     = Font(name="Calibri", size=11, bold=True, color=NAVY)
f_label       = Font(name="Calibri", size=11, bold=True)
f_input       = Font(name="Calibri", size=11, color="0070C0", bold=True)
f_calc        = Font(name="Calibri", size=11)
f_note        = Font(name="Calibri", size=10, italic=True, color=DARK_GREY)
f_kpi_val     = Font(name="Calibri", size=14, bold=True, color=NAVY)
f_white_bold  = Font(name="Calibri", size=11, bold=True, color=WHITE)
f_memo_body   = Font(name="Calibri", size=11)

fill_title   = PatternFill("solid", fgColor=NAVY)
fill_section = PatternFill("solid", fgColor=BLUE)
fill_input   = PatternFill("solid", fgColor=YELLOW)
fill_calc    = PatternFill("solid", fgColor=GREY)
fill_kpi     = PatternFill("solid", fgColor=LIGHT_BLUE)
fill_warn    = PatternFill("solid", fgColor=AMBER)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
right  = Alignment(horizontal="right", vertical="center")
top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

FMT_GBP_HARD   = '_-£* #,##0_-;[Red]-£* #,##0_-;_-£* "-"??_-;_-@_-'
FMT_GBP_HARD_D = '_-£* #,##0.00_-;[Red]-£* #,##0.00_-;_-£* "-"??_-;_-@_-'
FMT_MONEY      = '_-* #,##0_-;[Red]_-* (#,##0)_-;_-* "-"??_-;_-@_-'
FMT_MONEY_D    = '_-* #,##0.00_-;[Red]_-* (#,##0.00)_-;_-* "-"??_-;_-@_-'
FMT_INT  = "#,##0"
FMT_PCT  = "0.0%"
FMT_MULT = '0.0"x"'
FMT_DATE = "dd-mmm"
FMT_DATE_L = "dd-mmm-yyyy"
FMT_RATE = "0.0000"

WEEKS = 26
FIRST_WEEK_NUM = 1
DEFAULT_START_DATE = date(2026, 1, 12)

CURRENCIES = [
    ("GBP", "£",  1.0000),
    ("USD", "$",  1.2700),
    ("CAD", "C$", 1.7100),
    ("EUR", "€",  1.1700),
]

PROMO_TYPES = ["Demo", "End Cap", "TPD", "Markdown"]

SCENARIOS = [
    {"key":"BEST","name":"BEST · Demo + TPD (Wks 5-6)",
     "weight":0.30,
     "promo":{"Demo":[5,6],"End Cap":[],"TPD":[5,6],"Markdown":[]},
     "lift_overrides":{5:3.0,6:3.0},
     "rationale":"Demo + TPD only. ~3x lift in-promo. No end cap, no clearance."},
    {"key":"IDEAL","name":"IDEAL · Demo + End Cap + TPD (Wks 5-6)",
     "weight":0.50,
     "promo":{"Demo":[5,6],"End Cap":[5,6],"TPD":[5,6],"Markdown":[]},
     "lift_overrides":{5:3.6,6:3.6},
     "rationale":"Demo + end cap + TPD stacked. ~3.6x lift."},
    {"key":"WORST","name":"WORST · Launch then markdown clearance",
     "weight":0.20,
     "promo":{"Demo":[5,6],"End Cap":[5,6],"TPD":[5,6],"Markdown":[18,19,20]},
     "lift_overrides":{5:3.0,6:3.0,18:1.0,19:0.3,20:8.7},
     "rationale":"Launch as IDEAL; 50% markdown to clear residual stock Wks 18-20."},
]

PORTFOLIO_SKUS = [
    {"rank":1,"name":"Bee Propolis Spray 30ml x2","sp":15.99,"cost":7.29,"units":17000,"promo":25000,
     "growth":0.10,"rationale":"Lead pitch — strong GP%."},
    {"rank":2,"name":"Mag 8-in-1 90's/120's","sp":None,"cost":None,"units":None,"promo":None,"growth":0.10,"rationale":None},
    {"rank":3,"name":"O1 450g","sp":None,"cost":None,"units":None,"promo":None,"growth":0.10,"rationale":None},
    {"rank":4,"name":"Belli Bliss Raspberry, 450g","sp":None,"cost":None,"units":None,"promo":None,"growth":0.10,"rationale":None},
    {"rank":5,"name":"Daily Boost","sp":None,"cost":None,"units":None,"promo":None,"growth":0.10,"rationale":None},
    {"rank":6,"name":"Berberine 90's/120's","sp":None,"cost":None,"units":None,"promo":None,"growth":0.10,"rationale":None},
    {"rank":7,"name":"L-theanine Capsules 90's/120's","sp":None,"cost":None,"units":None,"promo":None,"growth":0.10,"rationale":None},
]

wb = openpyxl.Workbook()
wb.remove(wb.active)


def wc(ws, r, c, v, *, font=None, fill=None, align=None, border=None, number_format=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font is not None: cell.font = font
    if fill is not None: cell.fill = fill
    if align is not None: cell.alignment = align
    if border is not None: cell.border = border
    if number_format is not None: cell.number_format = number_format
    return cell


def section_header(ws, r, text, span_end_col=4, span_start_col=2):
    ws.merge_cells(start_row=r, start_column=span_start_col, end_row=r, end_column=span_end_col)
    wc(ws, r, span_start_col, text, font=f_section, fill=fill_section, align=left)
    ws.row_dimensions[r].height = 22


def add_currency_widget(ws, row, default="GBP"):
    """Inline currency widget. Returns (fx_ref, sym_ref) absolute refs ($Y/$Z<row>)."""
    wc(ws, row, 2, "Display Currency", font=f_label, align=right)
    wc(ws, row, 3, default, font=f_input, fill=fill_input, align=center, border=box)
    wc(ws, row, 4,
       f'="("&VLOOKUP(C{row},Inputs!$B$13:$D$16,2,FALSE)&"  ·  rate "&TEXT(VLOOKUP(C{row},Inputs!$B$13:$D$16,3,FALSE),"0.0000")&")"',
       font=f_note, align=left)
    wc(ws, row, 25, f"=VLOOKUP(C{row},Inputs!$B$13:$D$16,3,FALSE)", number_format=FMT_RATE)
    wc(ws, row, 26, f"=VLOOKUP(C{row},Inputs!$B$13:$D$16,2,FALSE)")
    dv = DataValidation(type="list", formula1='"GBP,USD,CAD,EUR"', allow_blank=False)
    ws.add_data_validation(dv); dv.add(f"C{row}")
    ws.row_dimensions[row].height = 22
    return (f"$Y${row}", f"$Z${row}")


# =============================================================================
# 1. START HERE (README)
# =============================================================================
rd = wb.create_sheet("Start Here")
rd.sheet_view.showGridLines = False
rd.column_dimensions["A"].width = 4
rd.column_dimensions["B"].width = 105

wc(rd,2,2,"Costco UK — D93 Portfolio ROI Model",
   font=Font(name="Calibri",size=20,bold=True,color=NAVY),align=left)
wc(rd,3,2,"v9 enterprise · 8 tabs · scenario-weighted, multi-year, multi-currency",
   font=Font(name="Calibri",size=12,italic=True,color=DARK_GREY),align=left)

r = 5
def hdr(text):
    global r
    wc(rd, r, 2, text, font=Font(name="Calibri",size=13,bold=True,color=WHITE),
       fill=fill_section, align=left); r += 1

def body(text):
    global r
    wc(rd, r, 2, text, font=Font(name="Calibri",size=11), align=top_wrap)
    rd.row_dimensions[r].height = 18
    r += 1

hdr("WHAT THIS MODEL DOES")
body("")
body("Costs out the Bee Propolis Spray Costco UK market test under three scenarios "
     "(BEST / IDEAL / WORST), rolls those into Y1 numbers, then blends Bee Propolis "
     "with the other six D93 pitch SKUs into a 3-year portfolio view with blended margin and ROI.")
body("")

hdr("HOW TO USE")
body("")
for txt in [
    "1.  Open the Inputs tab. Change any YELLOW cell. Grey cells calculate themselves.",
    "2.  Pick a Display Currency at the top of any output tab (Forecast, Cumulative, Summary, Portfolio, Pitch Memo).",
    "3.  Forecast = 26-week weekly P&L per scenario.",
    "4.  Cumulative = running Net Profit per scenario + break-even week (line chart).",
    "5.  Sensitivity = Lift × Demo Cost grid showing Net Profit (where the model breaks).",
    "6.  Summary = scenario KPI comparison + probability-weighted Expected Value + bar chart.",
    "7.  Portfolio = 3-year roll-up across all seven SKUs with blended margin/ROI.",
    "8.  Pitch Memo = exec-ready 1-pager. Numbers update live.",
]:
    body(txt)
body("")

hdr("COLOUR KEY")
body("")
for label,color,desc in [
    ("Yellow",     YELLOW,     "Editable input"),
    ("Light Blue", LIGHT_BLUE, "Summary KPI"),
    ("Grey",       GREY,       "Calculated"),
    ("Amber",      AMBER,      "Warning / sanity-check"),
]:
    wc(rd,r,2,f"   {label}: {desc}",align=left,
       fill=PatternFill("solid",fgColor=color),font=Font(name="Calibri",size=11))
    r += 1
body("")

hdr("KEY ASSUMPTIONS")
body("")
for txt in [
    "·  Base demand = 276 units / week across 20 warehouses (~13.8 units/wh/week). From the prior 30ml SKU run-rate.",
    "·  Demo lift = 3.0x base. End cap stacked on demo = 3.6x. Industry benchmark for Costco UK supplements.",
    "·  TPD discount = 25% off net (20% off MSRP × 1/0.8 retail margin gross-up).",
    "·  Markdown discount = 50% of net (clearance).",
    "·  Default scenario probabilities: BEST 30%, IDEAL 50%, WORST 20%. Editable on Inputs.",
    "·  Y2 / Y3 growth = 10% per SKU. Editable per SKU on Portfolio.",
    "·  FX rates current as of model build date — refresh on the Inputs FX Rate Table.",
]:
    body(txt)
body("")

hdr("MODEL LIMITATIONS")
body("")
for txt in [
    "·  Demand lift is the same in every promo week — no week-2 fatigue modelled.",
    "·  No seasonality (Q4 retail spike, summer dip).",
    "·  Single FX snapshot — no hedge cost modelled.",
    "·  Portfolio SP / Cost stored in GBP. Switching the display currency converts but does NOT re-quote prices in the new market.",
    "·  Markdown only triggers in the WORST scenario weeks 18-20 (hard-wired in scenario activity grid; edit grid to change).",
]:
    body(txt)
body("")

hdr("CHANGE LOG")
body("")
for txt in [
    "v9  enterprise pass: Cumulative, Sensitivity, Pitch Memo tabs; scenario weighting; 3-year Portfolio; charts; validation.",
    "v8  simplified Inputs / Forecast / Summary / Portfolio.",
    "v7  per-tab currency dropdowns.",
    "v6  currency selection added on Inputs.",
    "v5  removed override column; single yellow column for all inputs.",
    "v4-v2  iterations on layout & style.",
    "v1  original George spreadsheet (rebuilt from scratch).",
]:
    body(txt)


# =============================================================================
# 2. INPUTS
# =============================================================================
inp = wb.create_sheet("Inputs")
inp.sheet_view.showGridLines = False
inp.column_dimensions["A"].width = 2
inp.column_dimensions["B"].width = 38
inp.column_dimensions["C"].width = 16
inp.column_dimensions["D"].width = 78

inp.merge_cells("B2:D2")
wc(inp,2,2,"INPUTS — change any yellow cell",
   font=f_title,fill=fill_title,align=center)
inp.row_dimensions[2].height = 32


def input_row(r, label, value, note, fmt=None, is_formula=False):
    wc(inp,r,2,label,font=f_label,align=left)
    cell_value = f"={value}" if is_formula else value
    cell_font = f_calc if is_formula else f_input
    wc(inp,r,3,cell_value,font=cell_font,fill=fill_input,
       align=right,border=box,number_format=fmt)
    wc(inp,r,4,note,font=f_note,align=top_wrap)
    inp.row_dimensions[r].height = 22


section_header(inp, 4, "1.  Display Currency  (master copy — each tab has its own)")
wc(inp,5,2,"Display Currency",font=f_label,align=left)
wc(inp,5,3,"GBP",font=f_input,fill=fill_input,align=center,border=box)
wc(inp,5,4,"Reference only. Each display tab has its own dropdown.",
   font=f_note,align=top_wrap)
inp.row_dimensions[5].height = 22

dv = DataValidation(type="list", formula1='"GBP,USD,CAD,EUR"', allow_blank=False)
inp.add_data_validation(dv); dv.add("C5")

inp.merge_cells("B7:D7")
wc(inp,7,2,"FX Rate Table — edit rates (rate is per 1 GBP)",
   font=f_subhead,align=left)

wc(inp,12,2,"Code",  font=f_white_bold,fill=fill_section,align=center,border=box)
wc(inp,12,3,"Symbol",font=f_white_bold,fill=fill_section,align=center,border=box)
wc(inp,12,4,"Rate (× GBP)",font=f_white_bold,fill=fill_section,align=left,border=box)

for i,(code,sym,rate) in enumerate(CURRENCIES):
    rr = 13 + i
    wc(inp,rr,2,code,font=f_label,fill=fill_calc,align=center,border=box)
    wc(inp,rr,3,sym, font=f_input,fill=fill_input,align=center,border=box)
    wc(inp,rr,4,rate,font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_RATE)
    inp.row_dimensions[rr].height = 18

section_header(inp, 19, "2.  Product & Distribution")
input_row(20,"SKU","Bee Propolis Spray 30ml x 2","Costco UK market test SKU.")
input_row(21,"Test Start Date (W1)",DEFAULT_START_DATE,
          "Week-commencing date for W1.",fmt=FMT_DATE_L)
input_row(22,"Warehouses in test",20,"Number of Costco UK warehouses.",fmt=FMT_INT)

section_header(inp, 24, "3.  Pricing (GBP)")
input_row(25,"Selling Price (MSRP)",15.99,"Retail per pack.",fmt=FMT_GBP_HARD_D)
input_row(26,"Cost Price (landed)",7.29,"Brand cost delivered.",fmt=FMT_GBP_HARD_D)
input_row(27,"Gross Margin per unit","C25-C26","SP − Cost (auto).",
          fmt=FMT_GBP_HARD_D,is_formula=True)
input_row(28,"Gross Margin %","IFERROR((C25-C26)/C25,0)","GP ÷ SP (auto).",
          fmt=FMT_PCT,is_formula=True)

# Margin validation: warn if < 30% or > 80%
dv_pct = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
inp.add_data_validation(dv_pct)
# (We can't enforce on a calculated cell — instead, conditional format flags low margin)
inp.conditional_formatting.add("C28",
    CellIsRule(operator="lessThan", formula=["0.3"],
               fill=fill_warn, font=Font(bold=True, color="9C5700")))

section_header(inp, 30, "4.  Base Demand")
input_row(31,"Base Units / Week (all warehouses)",276,
          "Organic weekly units across test warehouses, no promo.",fmt=FMT_INT)

section_header(inp, 33, "5.  Promo Economics (GBP)")
input_row(34,"Demo Total £ / week",7403,
          "All-warehouse demo cost in a demo week. ~£199/wh × 20 wh × 1.86 factor.",fmt=FMT_GBP_HARD)
input_row(35,"End Cap Total £ / week",15810,
          "All-warehouse end cap. ~£850/wh × 20 wh × 1.86 ÷ 2-week cycle.",fmt=FMT_GBP_HARD)
input_row(36,"TPD discount % (of net)",0.25,"20% off MSRP ≈ 25% off net.",fmt=FMT_PCT)
input_row(37,"Markdown discount %",0.50,"Clearance pricing.",fmt=FMT_PCT)

section_header(inp, 39, "6.  Scenario Probability Weights  (must sum to 100%)")
for i, sc in enumerate(SCENARIOS):
    rr = 40 + i
    wc(inp,rr,2,f"P({sc['key']})",font=f_label,align=left)
    wc(inp,rr,3,sc["weight"],font=f_input,fill=fill_input,
       align=right,border=box,number_format=FMT_PCT)
    wc(inp,rr,4,sc["rationale"],font=f_note,align=top_wrap)
    inp.row_dimensions[rr].height = 22
# Sum check
wc(inp,43,2,"Sum of weights",font=f_label,align=left)
wc(inp,43,3,"=SUM(C40:C42)",font=f_calc,fill=fill_calc,
   align=right,border=box,number_format=FMT_PCT)
wc(inp,43,4,'="(should be 100%)"',font=f_note,align=top_wrap)
# Warn if sum != 100%
inp.conditional_formatting.add("C43",
    FormulaRule(formula=["ROUND(C43,4)<>1"], fill=fill_warn,
                font=Font(bold=True, color="9C5700")))

# Named ranges
wb.defined_names["StartDate"] = DefinedName("StartDate", attr_text="Inputs!$C$21")
wb.defined_names["WHs"]       = DefinedName("WHs",       attr_text="Inputs!$C$22")
wb.defined_names["Period"]    = DefinedName("Period",    attr_text=f"{WEEKS}")
wb.defined_names["Price"]     = DefinedName("Price",     attr_text="Inputs!$C$25")
wb.defined_names["Cost"]      = DefinedName("Cost",      attr_text="Inputs!$C$26")
wb.defined_names["BaseUnits"] = DefinedName("BaseUnits", attr_text="Inputs!$C$31")
wb.defined_names["DemoCost"]  = DefinedName("DemoCost",  attr_text="Inputs!$C$34")
wb.defined_names["EndCapCost"]= DefinedName("EndCapCost",attr_text="Inputs!$C$35")
wb.defined_names["TPDPct"]    = DefinedName("TPDPct",    attr_text="Inputs!$C$36")
wb.defined_names["MdPct"]     = DefinedName("MdPct",     attr_text="Inputs!$C$37")
wb.defined_names["PBest"]     = DefinedName("PBest",     attr_text="Inputs!$C$40")
wb.defined_names["PIdeal"]    = DefinedName("PIdeal",    attr_text="Inputs!$C$41")
wb.defined_names["PWorst"]    = DefinedName("PWorst",    attr_text="Inputs!$C$42")

section_header(inp, 45, "7.  Scenario activity (1 = on / 0 = off per week) and demand lift")

WEEK_START_COL = 6
WEEK_END_COL = WEEK_START_COL + WEEKS - 1
for col in range(WEEK_START_COL, WEEK_END_COL + 1):
    inp.column_dimensions[get_column_letter(col)].width = 7.5

scenario_grid_start = {}
current = 47
for sc in SCENARIOS:
    inp.merge_cells(start_row=current,start_column=2,end_row=current,end_column=WEEK_END_COL)
    wc(inp,current,2,sc["name"],font=f_white_bold,fill=fill_title,align=left)
    inp.row_dimensions[current].height = 22
    current += 1

    inp.merge_cells(start_row=current,start_column=2,end_row=current,end_column=WEEK_END_COL)
    wc(inp,current,2,f"Rationale: {sc['rationale']}",font=f_note,align=top_wrap)
    inp.row_dimensions[current].height = 24
    current += 1

    wc(inp,current,2,"Week",font=f_label,fill=fill_calc,align=left)
    for i in range(WEEKS):
        wc(inp,current,WEEK_START_COL+i,f"W{FIRST_WEEK_NUM+i}",
           font=f_white_bold,fill=PatternFill("solid",fgColor=BLUE),
           align=center,border=box)
    current += 1

    wc(inp,current,2,"Wk-commencing",font=f_label,fill=fill_calc,align=left)
    for i in range(WEEKS):
        wc(inp,current,WEEK_START_COL+i,f"=StartDate+7*{i}",
           font=Font(name="Calibri",size=9,color=DARK_GREY),
           fill=fill_calc,align=center,border=box,number_format=FMT_DATE)
    current += 1

    scenario_grid_start[sc["key"]] = current

    for promo in PROMO_TYPES:
        wc(inp,current,2,promo,font=f_label,align=left)
        active = set(sc["promo"].get(promo,[]))
        for i in range(WEEKS):
            wknum = FIRST_WEEK_NUM + i
            wc(inp,current,WEEK_START_COL+i,
               1 if wknum in active else 0,
               font=f_input,fill=fill_input,align=center,border=box,
               number_format="0;;;@")
        current += 1

    wc(inp,current,2,"Lift Multiplier (x base)",font=f_label,align=left)
    for i in range(WEEKS):
        wknum = FIRST_WEEK_NUM + i
        wc(inp,current,WEEK_START_COL+i,
           sc["lift_overrides"].get(wknum,1.0),
           font=f_input,fill=fill_input,align=center,border=box,
           number_format=FMT_MULT)
    current += 1
    current += 1

inp.freeze_panes = "F5"


# =============================================================================
# 3. FORECAST
# =============================================================================
fc = wb.create_sheet("Forecast")
fc.sheet_view.showGridLines = False
fc.column_dimensions["A"].width = 2
fc.column_dimensions["B"].width = 28
for col in range(WEEK_START_COL,WEEK_END_COL+1):
    fc.column_dimensions[get_column_letter(col)].width = 11
TOTAL_COL = WEEK_END_COL + 1
fc.column_dimensions[get_column_letter(TOTAL_COL)].width = 14

fc.merge_cells(start_row=2,start_column=2,end_row=2,end_column=TOTAL_COL)
wc(fc,2,2,'="FORECAST — Weekly P&L  (Display: "&$Z$3&" "&$C$3&")"',
   font=f_title,fill=fill_title,align=center)
fc.row_dimensions[2].height = 32

FC_FX, FC_SYM = add_currency_widget(fc, 3, default="GBP")

PNL_ROWS = [
    ("POS Units",                            "units"),
    ('="Revenue ("&$Z$3&")"',                "rev"),
    ('="Raw COGS ("&$Z$3&")"',               "cogs"),
    ('="Gross Profit ("&$Z$3&")"',           "gp"),
    ('="Promo Spend ("&$Z$3&")"',            "promo"),
    ('="Net Profit ("&$Z$3&")"',             "net"),
]
PROMO_ROW_OFFSET = {p:i for i,p in enumerate(PROMO_TYPES)}
LIFT_OFFSET = len(PROMO_TYPES)

fc_row = 5
scenario_summary_rows = {}

for sc in SCENARIOS:
    fc.merge_cells(start_row=fc_row,start_column=2,end_row=fc_row,end_column=TOTAL_COL)
    wc(fc,fc_row,2,sc["name"],font=f_white_bold,fill=fill_title,align=left)
    fc.row_dimensions[fc_row].height = 22
    fc_row += 1

    wc(fc,fc_row,2,"Week",font=f_label,fill=fill_calc,align=left)
    for i in range(WEEKS):
        wc(fc,fc_row,WEEK_START_COL+i,f"W{FIRST_WEEK_NUM+i}",
           font=f_white_bold,fill=PatternFill("solid",fgColor=BLUE),
           align=center,border=box)
    wc(fc,fc_row,TOTAL_COL,"TOTAL",font=f_white_bold,fill=fill_title,align=center,border=box)
    fc_row += 1

    wc(fc,fc_row,2,"Wk-commencing",font=f_label,fill=fill_calc,align=left)
    for i in range(WEEKS):
        wc(fc,fc_row,WEEK_START_COL+i,f"=StartDate+7*{i}",
           font=Font(name="Calibri",size=9,color=DARK_GREY),
           fill=fill_calc,align=center,border=box,number_format=FMT_DATE)
    wc(fc,fc_row,TOTAL_COL,"",fill=fill_calc,border=box)
    fc_row += 1

    scenario_pnl_rows = {}
    grid_top = scenario_grid_start[sc["key"]]
    promo_rows_on_inp = {p:grid_top+PROMO_ROW_OFFSET[p] for p in PROMO_TYPES}
    lift_row_on_inp = grid_top + LIFT_OFFSET

    for label_formula, key in PNL_ROWS:
        scenario_pnl_rows[key] = fc_row
        navy_bold = key in ("gp","net")
        lbl_font = (Font(name="Calibri",size=11,bold=True,color=NAVY)
                    if navy_bold else f_label)
        wc(fc,fc_row,2,label_formula,font=lbl_font,align=left)

        for i in range(WEEKS):
            wcol = get_column_letter(WEEK_START_COL+i)
            if key == "units":
                formula = f"=ROUND(BaseUnits*Inputs!{wcol}{lift_row_on_inp},0)"
            elif key == "rev":
                formula = f"={wcol}{scenario_pnl_rows['units']}*Price*{FC_FX}"
            elif key == "cogs":
                formula = f"={wcol}{scenario_pnl_rows['units']}*Cost*{FC_FX}"
            elif key == "gp":
                formula = f"={wcol}{scenario_pnl_rows['rev']}-{wcol}{scenario_pnl_rows['cogs']}"
            elif key == "promo":
                demo  = f"Inputs!{wcol}{promo_rows_on_inp['Demo']}*DemoCost"
                endcap= f"Inputs!{wcol}{promo_rows_on_inp['End Cap']}*EndCapCost"
                tpd   = (f"Inputs!{wcol}{promo_rows_on_inp['TPD']}"
                         f"*{wcol}{scenario_pnl_rows['units']}*Price*TPDPct")
                md    = (f"Inputs!{wcol}{promo_rows_on_inp['Markdown']}"
                         f"*{wcol}{scenario_pnl_rows['units']}*Price*MdPct")
                formula = f"=({demo}+{endcap}+{tpd}+{md})*{FC_FX}"
            elif key == "net":
                formula = f"={wcol}{scenario_pnl_rows['gp']}-{wcol}{scenario_pnl_rows['promo']}"
            else:
                formula = ""
            num_fmt = FMT_INT if key == "units" else FMT_MONEY
            wc(fc,fc_row,WEEK_START_COL+i,formula,
               font=f_calc,fill=fill_calc,align=right,border=box,number_format=num_fmt)

        first_w = get_column_letter(WEEK_START_COL)
        last_w = get_column_letter(WEEK_END_COL)
        wc(fc,fc_row,TOTAL_COL,f"=SUM({first_w}{fc_row}:{last_w}{fc_row})",
           font=Font(name="Calibri",size=11,bold=True,color=NAVY),
           fill=fill_kpi,align=right,border=box,
           number_format=FMT_INT if key == "units" else FMT_MONEY)
        fc_row += 1

    scenario_summary_rows[sc["key"]] = scenario_pnl_rows
    fc_row += 2

fc.freeze_panes = "C5"


# =============================================================================
# 4. CUMULATIVE  (running totals + break-even)
# =============================================================================
cu = wb.create_sheet("Cumulative")
cu.sheet_view.showGridLines = False
cu.column_dimensions["A"].width = 2
cu.column_dimensions["B"].width = 26
for col in range(WEEK_START_COL, WEEK_END_COL + 1):
    cu.column_dimensions[get_column_letter(col)].width = 11

cu.merge_cells(start_row=2, start_column=2, end_row=2, end_column=WEEK_END_COL)
wc(cu,2,2,'="CUMULATIVE Net Profit & break-even  (Display: "&$Z$3&" "&$C$3&")"',
   font=f_title, fill=fill_title, align=center)
cu.row_dimensions[2].height = 32

CU_FX, CU_SYM = add_currency_widget(cu, 3, default="GBP")
FC_FX_FROM_CU = f"Forecast!{FC_FX}"

# Week header
wc(cu,5,2,"Scenario",font=f_label,fill=fill_calc,align=left,border=box)
for i in range(WEEKS):
    wc(cu,5,WEEK_START_COL+i,f"W{FIRST_WEEK_NUM+i}",
       font=f_white_bold,fill=PatternFill("solid",fgColor=BLUE),
       align=center,border=box)
# Date row
wc(cu,6,2,"Wk-commencing",font=f_label,fill=fill_calc,align=left)
for i in range(WEEKS):
    wc(cu,6,WEEK_START_COL+i,f"=StartDate+7*{i}",
       font=Font(name="Calibri",size=9,color=DARK_GREY),
       fill=fill_calc,align=center,border=box,number_format=FMT_DATE)

cu_row = 7
cumulative_rows_by_scenario = {}
for sc in SCENARIOS:
    wc(cu,cu_row,2,sc["key"],font=f_label,align=left,border=box)
    cumulative_rows_by_scenario[sc["key"]] = cu_row
    forecast_net_row = scenario_summary_rows[sc["key"]]["net"]
    fx_adj = f"*({CU_FX}/{FC_FX_FROM_CU})"
    for i in range(WEEKS):
        wcol = get_column_letter(WEEK_START_COL+i)
        first_w = get_column_letter(WEEK_START_COL)
        # Cumulative net through this week = SUM(Forecast week-cells, W1..Wi) × FX adjust
        formula = (f"=SUM(Forecast!{first_w}{forecast_net_row}:Forecast!{wcol}{forecast_net_row})"
                   f"{fx_adj}")
        wc(cu,cu_row,WEEK_START_COL+i,formula,
           font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_MONEY)
    cu_row += 1

cu_row += 2
# Break-even table
wc(cu,cu_row,2,"Break-Even Analysis",font=f_section,fill=fill_section,align=left,border=box)
cu.merge_cells(start_row=cu_row, start_column=2, end_row=cu_row, end_column=8)
cu_row += 1

wc(cu,cu_row,2,"Scenario",font=f_white_bold,fill=fill_section,align=left,border=box)
wc(cu,cu_row,3,"Break-even Week",font=f_white_bold,fill=fill_section,align=center,border=box)
wc(cu,cu_row,4,"Wk-commencing",font=f_white_bold,fill=fill_section,align=center,border=box)
wc(cu,cu_row,5,"Y1 Net Profit",font=f_white_bold,fill=fill_section,align=right,border=box)
wc(cu,cu_row,6,"Status",font=f_white_bold,fill=fill_section,align=center,border=box)
cu_row += 1

first_w = get_column_letter(WEEK_START_COL)
last_w  = get_column_letter(WEEK_END_COL)

for sc in SCENARIOS:
    crow = cumulative_rows_by_scenario[sc["key"]]
    wc(cu,cu_row,2,sc["key"],font=f_label,align=left,border=box)
    # Break-even week = MATCH(first cumulative value > 0) — using INDEX/MATCH with array eval
    # We use IFERROR(MATCH(TRUE,(range>0),0),0) — array entered. To keep it simple, use SUMPRODUCT trick:
    # First week where cumulative > 0
    be_formula = (f'=IFERROR(MATCH(TRUE,INDEX({first_w}{crow}:{last_w}{crow}>0,0),0),0)')
    wc(cu,cu_row,3,be_formula,
       font=f_calc,fill=fill_calc,align=center,border=box,number_format='"W"0;;""')
    # Date for that week
    wc(cu,cu_row,4,f'=IF(C{cu_row}>0,StartDate+7*(C{cu_row}-1),"never")',
       font=f_calc,fill=fill_calc,align=center,border=box,number_format=FMT_DATE_L)
    # Y1 Net (last cumulative value)
    wc(cu,cu_row,5,f"={last_w}{crow}",
       font=Font(name="Calibri",size=11,bold=True,color=NAVY),
       fill=fill_kpi,align=right,border=box,number_format=FMT_MONEY)
    # Status: positive net → "Pays back" else "Loss-making"
    wc(cu,cu_row,6,f'=IF(E{cu_row}>=0,"Pays back","Loss-making")',
       font=f_calc,fill=fill_calc,align=center,border=box)
    # Conditional fmt on the status text
    cu.conditional_formatting.add(f"F{cu_row}",
        FormulaRule(formula=[f'F{cu_row}="Pays back"'],
                    fill=PatternFill("solid",fgColor=GREEN),
                    font=Font(bold=True,color="006100")))
    cu.conditional_formatting.add(f"F{cu_row}",
        FormulaRule(formula=[f'F{cu_row}="Loss-making"'],
                    fill=PatternFill("solid",fgColor=RED),
                    font=Font(bold=True,color="9C0006")))
    cu_row += 1

# Line chart of cumulative Net Profit by scenario
chart = LineChart()
chart.title = "Cumulative Net Profit by Scenario"
chart.style = 12
chart.y_axis.title = "Net Profit"
chart.x_axis.title = "Week"
chart.height = 10
chart.width = 22
for sc in SCENARIOS:
    crow = cumulative_rows_by_scenario[sc["key"]]
    data = Reference(cu, min_col=WEEK_START_COL, max_col=WEEK_END_COL, min_row=crow, max_row=crow)
    s = Series(data, title=sc["key"])
    chart.series.append(s)
cats = Reference(cu, min_col=WEEK_START_COL, max_col=WEEK_END_COL, min_row=5, max_row=5)
chart.set_categories(cats)
cu.add_chart(chart, f"B{cu_row + 2}")

cu.freeze_panes = "C7"


# =============================================================================
# 5. SENSITIVITY  (Lift × Demo Cost grid → Net Profit, IDEAL scenario)
# =============================================================================
sn = wb.create_sheet("Sensitivity")
sn.sheet_view.showGridLines = False
sn.column_dimensions["A"].width = 2
sn.column_dimensions["B"].width = 28
for c in range(3, 11):
    sn.column_dimensions[get_column_letter(c)].width = 14

sn.merge_cells("B2:J2")
wc(sn,2,2,'="SENSITIVITY — IDEAL scenario Net Profit (GBP)  (×lift vs ×demo cost)"',
   font=f_title,fill=fill_title,align=center)
sn.row_dimensions[2].height = 32

sn.merge_cells("B3:J3")
wc(sn,3,2,
   "Each cell shows Y1 Net Profit if BOTH Demo-week lift multipliers AND the Demo Total £/week "
   "were the values at that row/column intersection. Highlights how robust the IDEAL plan is.",
   font=Font(name="Calibri",size=11,italic=True,color=DARK_GREY),align=top_wrap)
sn.row_dimensions[3].height = 36

# Build sensitivity grid manually (Excel data tables are clunky in openpyxl)
# Rows: Lift multiplier (2.0, 2.5, 3.0, 3.6, 4.0, 4.5, 5.0)
# Cols: Demo Cost £/week (5000, 6000, 7403 [base], 9000, 11000, 13000)
lifts = [2.0, 2.5, 3.0, 3.6, 4.0, 4.5, 5.0]
demo_costs = [5000, 6000, 7403, 9000, 11000, 13000]

# Header
wc(sn,5,2,"Lift ↓  /  Demo Cost →",font=f_white_bold,fill=fill_section,align=center,border=box)
for j, d in enumerate(demo_costs):
    wc(sn,5,3+j,d,font=f_white_bold,fill=fill_section,align=center,border=box,number_format=FMT_GBP_HARD)

# For each (lift, demo cost):
# Net = Σ over 26 weeks of: Units × (Price - Cost) - Promo_in_week
# Where in IDEAL, Demo and EndCap and TPD are on in weeks 5,6 with the given lift,
# other weeks lift = 1.0.
# This is a simplified estimate: 24 base weeks + 2 promo weeks
# - Base weeks: 24 × BaseUnits × (Price-Cost)
# - Promo weeks: 2 × BaseUnits × lift × (Price-Cost) - 2 × (DemoCost + EndCapCost + BaseUnits × lift × Price × TPDPct)

for i, lift in enumerate(lifts):
    rr = 6 + i
    wc(sn,rr,2,lift,font=f_label,fill=fill_section,align=center,border=box,
       number_format=FMT_MULT)
    for j, d in enumerate(demo_costs):
        # Inline formula: 24 weeks base + 2 weeks with promo
        base_24 = "24*BaseUnits*(Price-Cost)"
        promo_units = f"2*BaseUnits*{lift}*(Price-Cost)"
        promo_cost = f"2*({d}+EndCapCost+BaseUnits*{lift}*Price*TPDPct)"
        formula = f"={base_24}+{promo_units}-{promo_cost}"
        cell = wc(sn,rr,3+j,formula,
           font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_GBP_HARD)

# Highlight the base-case row/column (lift 3.6, demo 7403) — corresponds to row index 3, col index 2
# That cell is at row 6+3 = 9, col 3+2 = 5
wc(sn,9,5,"=24*BaseUnits*(Price-Cost)+2*BaseUnits*3.6*(Price-Cost)-2*(7403+EndCapCost+BaseUnits*3.6*Price*TPDPct)",
   font=Font(name="Calibri",size=11,bold=True,color=NAVY),
   fill=fill_kpi,align=right,border=box,number_format=FMT_GBP_HARD)

# Apply 3-color scale on the grid
from openpyxl.formatting.rule import ColorScaleRule
sn.conditional_formatting.add("C6:H12",
    ColorScaleRule(start_type="min", start_color="FFC7CE",
                   mid_type="percentile", mid_value=50, mid_color="FFFFFF",
                   end_type="max", end_color="C6EFCE"))

wc(sn,14,2,"Reading the grid",font=f_subhead,align=left)
sn.merge_cells("B14:J14")
wc(sn,15,2,
   "• Centre cell (lift 3.6x, demo £7,403) ≈ the IDEAL base case. Green = better Net Profit, "
   "red = worse. If your numbers land in the red zone, the plan is exposed.",
   font=Font(name="Calibri",size=11),align=top_wrap)
sn.merge_cells("B15:J15")
sn.row_dimensions[15].height = 36

wc(sn,17,2,
   "Note: this grid is GBP-only (the model's source currency). Switch the display currency "
   "on other tabs for reporting.",
   font=f_note,align=top_wrap)
sn.merge_cells("B17:J17")
sn.row_dimensions[17].height = 24


# =============================================================================
# 6. SUMMARY
# =============================================================================
sm = wb.create_sheet("Summary")
sm.sheet_view.showGridLines = False
sm.column_dimensions["A"].width = 2
sm.column_dimensions["B"].width = 32
for i in range(len(SCENARIOS) + 1):  # +1 for Expected Value column
    sm.column_dimensions[get_column_letter(3+i)].width = 22

end_col = 2 + len(SCENARIOS) + 1
sm.merge_cells(start_row=2,start_column=2,end_row=2,end_column=end_col)
wc(sm,2,2,'="ROI SUMMARY  (Display: "&$Z$3&" "&$C$3&")"',
   font=f_title,fill=fill_title,align=center)
sm.row_dimensions[2].height = 32

SM_FX, SM_SYM = add_currency_widget(sm, 3, default="GBP")
FC_FX_FROM_SM = f"Forecast!{FC_FX}"

sm.merge_cells(start_row=4,start_column=2,end_row=4,end_column=end_col)
wc(sm,4,2,
   '="Test window: W1 ("&TEXT(StartDate,"dd-mmm-yyyy")&")  →  W"&Period&" ("&TEXT(StartDate+7*(Period-1),"dd-mmm-yyyy")&")"',
   font=Font(name="Calibri",size=11,italic=True,color=DARK_GREY),
   fill=fill_calc,align=center)

hr = 6
wc(sm,hr,2,"Metric",font=f_white_bold,fill=fill_section,align=left)
for i,sc in enumerate(SCENARIOS):
    wc(sm,hr,3+i,sc["key"],font=f_white_bold,fill=fill_section,align=center)
wc(sm,hr,3+len(SCENARIOS),"E[Value]",font=f_white_bold,fill=fill_section,align=center)
sm.row_dimensions[hr].height = 24

# Show weights row
wc(sm,hr+1,2,"Probability weight",font=f_note,align=left,border=box)
wc(sm,hr+1,3,"=PBest",font=f_note,fill=fill_calc,align=right,border=box,number_format=FMT_PCT)
wc(sm,hr+1,4,"=PIdeal",font=f_note,fill=fill_calc,align=right,border=box,number_format=FMT_PCT)
wc(sm,hr+1,5,"=PWorst",font=f_note,fill=fill_calc,align=right,border=box,number_format=FMT_PCT)
wc(sm,hr+1,6,"=PBest+PIdeal+PWorst",font=f_note,fill=fill_calc,align=right,border=box,number_format=FMT_PCT)

TC = get_column_letter(TOTAL_COL)
KPIS = [
    ('Units Sold',                             "units",  FMT_INT,    False, False),
    ('="Revenue ("&$Z$3&")"',                  "rev",    FMT_MONEY,  False, True),
    ('="Gross Profit ("&$Z$3&")"',             "gp",     FMT_MONEY,  True,  True),
    ('Gross Margin %',                         "gp_pct", FMT_PCT,    True,  False),
    ('="Promo Spend ("&$Z$3&")"',              "promo",  FMT_MONEY,  False, True),
    ('="Net Profit ("&$Z$3&")"',               "net",    FMT_MONEY,  True,  True),
    ('ROI on Promo Spend',                     "roi",    FMT_PCT,    True,  False),
]

sr = hr + 2
kpi_row_idx = {}
for label,key,fmt,bold,is_money in KPIS:
    kpi_row_idx[key] = sr
    wc(sm,sr,2,label,
       font=Font(name="Calibri",size=11,bold=bold,color=NAVY if bold else "000000"),
       fill=fill_kpi if bold else None,align=left,border=box)
    fx_adj = f"*({SM_FX}/{FC_FX_FROM_SM})"
    for i,sc in enumerate(SCENARIOS):
        rows = scenario_summary_rows[sc["key"]]
        if key in rows:
            base = f"Forecast!{TC}{rows[key]}"
            formula = f"={base}{fx_adj}" if is_money else f"={base}"
        elif key == "gp_pct":
            formula = f"=IFERROR(Forecast!{TC}{rows['gp']}/Forecast!{TC}{rows['rev']},0)"
        elif key == "roi":
            formula = f"=IFERROR(Forecast!{TC}{rows['net']}/Forecast!{TC}{rows['promo']},0)"
        else:
            formula = ""
        wc(sm,sr,3+i,formula,
           font=f_kpi_val if bold else f_calc,
           fill=fill_kpi if bold else None,
           align=right,border=box,number_format=fmt)
    # Expected Value column: SUMPRODUCT of weights and scenario values
    ev_col = 3 + len(SCENARIOS)
    # Cells C{sr}, D{sr}, E{sr} hold BEST/IDEAL/WORST values
    ev_formula = f"=C{sr}*PBest+D{sr}*PIdeal+E{sr}*PWorst"
    wc(sm,sr,ev_col,ev_formula,
       font=f_kpi_val if bold else f_calc,
       fill=fill_kpi if bold else None,
       align=right,border=box,number_format=fmt)
    sm.row_dimensions[sr].height = 26 if bold else 22
    sr += 1

# Conditional formatting on net & roi
last_col_letter = get_column_letter(end_col)
for k in ("net","roi"):
    rr = kpi_row_idx[k]
    rng = f"C{rr}:{last_col_letter}{rr}"
    sm.conditional_formatting.add(rng,
        CellIsRule(operator="lessThan",formula=["0"],
                   fill=PatternFill("solid",fgColor=RED),
                   font=Font(bold=True,color="9C0006")))
    sm.conditional_formatting.add(rng,
        CellIsRule(operator="greaterThanOrEqual",formula=["0"],
                   fill=PatternFill("solid",fgColor=GREEN),
                   font=Font(bold=True,color="006100")))

sr += 2
sm.merge_cells(start_row=sr,start_column=2,end_row=sr,end_column=end_col)
wc(sm,sr,2,"Scenario notes",font=f_section,fill=fill_section,align=left)
sm.row_dimensions[sr].height = 22
sr += 1
for sc in SCENARIOS:
    wc(sm,sr,2,sc["key"],
       font=Font(name="Calibri",size=11,bold=True,color=NAVY),align=top_wrap)
    sm.merge_cells(start_row=sr,start_column=3,end_row=sr,end_column=end_col)
    wc(sm,sr,3,sc["rationale"],font=f_note,align=top_wrap)
    sm.row_dimensions[sr].height = 28
    sr += 1

# Bar chart of Net Profit by scenario
net_row = kpi_row_idx["net"]
chart = BarChart()
chart.type = "col"
chart.style = 12
chart.title = "Net Profit by Scenario (+ Expected Value)"
chart.y_axis.title = "Net Profit"
chart.x_axis.title = "Scenario"
chart.height = 9
chart.width = 18
data = Reference(sm, min_col=3, max_col=end_col, min_row=net_row, max_row=net_row)
cats = Reference(sm, min_col=3, max_col=end_col, min_row=hr, max_row=hr)
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
chart.dataLabels = DataLabelList(showVal=True)
sm.add_chart(chart, f"B{sr + 2}")

sm.freeze_panes = "C7"


# =============================================================================
# 7. PORTFOLIO  (multi-year, FX-aware)
# =============================================================================
pf = wb.create_sheet("Portfolio")
pf.sheet_view.showGridLines = False
pf.column_dimensions["A"].width = 2
widths = {"B":4,"C":32,"D":10,"E":10,"F":8,"G":11,"H":13,"I":12,"J":13,
          "K":7,"L":13,"M":13,"N":13,"O":13,"P":34}
for col,w in widths.items():
    pf.column_dimensions[col].width = w

pf.merge_cells("B2:P2")
wc(pf,2,2,'="PORTFOLIO — D93 Pitch SKUs · 3-Year View  (Display: "&$Z$3&" "&$C$3&")"',
   font=f_title,fill=fill_title,align=center)
pf.row_dimensions[2].height = 32

PF_FX, PF_SYM = add_currency_widget(pf, 3, default="GBP")

pf.merge_cells("B4:P4")
wc(pf,4,2,
   "Enter SP / Cost / Y1 Units / Y1 Promo / Growth % per SKU. All inputs in GBP — display column "
   "converts. Y2 and Y3 use the per-SKU growth rate.",
   font=Font(name="Calibri",size=11,italic=True,color=DARK_GREY),align=top_wrap)
pf.row_dimensions[4].height = 26

HDR_ROW = 6
hdrs = [
    ("B","#"),("C","SKU"),
    ('D','="SP ("&$Z$3&")"'),
    ('E','="Cost ("&$Z$3&")"'),
    ("F","GM %"),("G","Y1 Units"),
    ('H','="Y1 Revenue ("&$Z$3&")"'),
    ('I','="Y1 Promo ("&$Z$3&")"'),
    ('J','="Y1 Net ("&$Z$3&")"'),
    ("K","Growth"),
    ('L','="Y2 Net ("&$Z$3&")"'),
    ('M','="Y3 Net ("&$Z$3&")"'),
    ('N','="3-Yr Net ("&$Z$3&")"'),
    ("O","3-Yr ROI"),
    ("P","Rationale"),
]
for col_letter,txt in hdrs:
    al = left if col_letter in ("C","P") else center
    wc(pf,HDR_ROW,column_index_from_string(col_letter),txt,
       font=f_white_bold,fill=fill_section,align=al,border=box)

sku_start_row = HDR_ROW + 1
for idx,sku in enumerate(PORTFOLIO_SKUS):
    rr = sku_start_row + idx
    wc(pf,rr,2,sku["rank"],font=f_label,fill=fill_calc,align=center,border=box,number_format=FMT_INT)
    wc(pf,rr,3,sku["name"],font=f_label,fill=fill_calc,align=left,border=box)
    # SP & Cost stored in GBP — display × FX
    wc(pf,rr,4,sku["sp"],font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_MONEY_D)
    wc(pf,rr,5,sku["cost"],font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_MONEY_D)
    wc(pf,rr,6,f"=IFERROR((D{rr}-E{rr})/D{rr},0)",
       font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_PCT)
    wc(pf,rr,7,sku["units"],font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_INT)
    # Y1 Revenue = units × SP × FX
    wc(pf,rr,8,f"=IFERROR(G{rr}*D{rr}*{PF_FX},0)",
       font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_MONEY)
    # Y1 Promo (input in GBP) × FX for display
    wc(pf,rr,9,sku["promo"],font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_MONEY)
    # Y1 Net = (Revenue × GM%) − Promo × FX
    wc(pf,rr,10,f"=IFERROR((G{rr}*(D{rr}-E{rr})-I{rr})*{PF_FX},0)",
       font=Font(name="Calibri",size=11,bold=True,color=NAVY),
       fill=fill_kpi,align=right,border=box,number_format=FMT_MONEY)
    # Growth
    wc(pf,rr,11,sku["growth"],font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_PCT)
    # Y2 Net = Y1 Net × (1 + growth)
    wc(pf,rr,12,f"=IFERROR(J{rr}*(1+K{rr}),0)",
       font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_MONEY)
    # Y3 Net = Y2 × (1 + growth)
    wc(pf,rr,13,f"=IFERROR(L{rr}*(1+K{rr}),0)",
       font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_MONEY)
    # 3-Yr Net = Y1 + Y2 + Y3
    wc(pf,rr,14,f"=J{rr}+L{rr}+M{rr}",
       font=Font(name="Calibri",size=11,bold=True,color=NAVY),
       fill=fill_kpi,align=right,border=box,number_format=FMT_MONEY)
    # 3-Yr ROI = (3-Yr Net) ÷ (Y1 Promo × FX) — promo assumed Y1-only
    wc(pf,rr,15,f"=IFERROR(N{rr}/(I{rr}*{PF_FX}),0)",
       font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_PCT)
    wc(pf,rr,16,sku["rationale"],font=f_input,fill=fill_input,align=left,border=box)
    pf.row_dimensions[rr].height = 22

last_sku_row = sku_start_row + len(PORTFOLIO_SKUS) - 1
tr = last_sku_row + 1

for c in range(2,6):
    wc(pf,tr,c,"",fill=fill_title,border=box)
wc(pf,tr,3,"PORTFOLIO TOTAL / BLENDED",
   font=f_white_bold,fill=fill_title,align=left,border=box)
# Blended GM%
wc(pf,tr,6,
   f"=IFERROR(SUMPRODUCT(G{sku_start_row}:G{last_sku_row},D{sku_start_row}:D{last_sku_row}-E{sku_start_row}:E{last_sku_row})/SUM(H{sku_start_row}:H{last_sku_row})*{PF_FX},0)",
   font=f_white_bold,fill=fill_title,align=right,border=box,number_format=FMT_PCT)
for col_letter,fmt in [("G",FMT_INT),("H",FMT_MONEY),("I",FMT_MONEY),
                       ("J",FMT_MONEY),("L",FMT_MONEY),("M",FMT_MONEY),("N",FMT_MONEY)]:
    wc(pf,tr,column_index_from_string(col_letter),
       f"=SUM({col_letter}{sku_start_row}:{col_letter}{last_sku_row})",
       font=f_white_bold,fill=fill_title,align=right,border=box,number_format=fmt)
# K column (growth) — blank for totals row
wc(pf,tr,11,"",fill=fill_title,border=box)
# Blended 3-Yr ROI = total 3-Yr Net ÷ total Promo × FX
wc(pf,tr,15,f"=IFERROR(N{tr}/(I{tr}*{PF_FX}),0)",
   font=f_white_bold,fill=fill_title,align=right,border=box,number_format=FMT_PCT)
wc(pf,tr,16,"Blended across SKUs",font=f_white_bold,fill=fill_title,align=left,border=box)
pf.row_dimensions[tr].height = 26

# Conditional format on ROI column (O) — red if negative, green if positive
pf.conditional_formatting.add(f"O{sku_start_row}:O{last_sku_row}",
    CellIsRule(operator="lessThan",formula=["0"],
               fill=PatternFill("solid",fgColor=RED),
               font=Font(bold=True,color="9C0006")))
pf.conditional_formatting.add(f"O{sku_start_row}:O{last_sku_row}",
    CellIsRule(operator="greaterThanOrEqual",formula=["0.5"],
               fill=PatternFill("solid",fgColor=GREEN),
               font=Font(bold=True,color="006100")))

pf.freeze_panes = "B7"


# =============================================================================
# 8. PITCH MEMO  (exec 1-pager, live)
# =============================================================================
pm = wb.create_sheet("Pitch Memo")
pm.sheet_view.showGridLines = False
pm.column_dimensions["A"].width = 4
pm.column_dimensions["B"].width = 105

pm.merge_cells("B2:B2")
wc(pm,2,2,"PITCH MEMO — D93 Portfolio (live numbers)",
   font=Font(name="Calibri",size=20,bold=True,color=NAVY),align=left)

PM_FX, PM_SYM = add_currency_widget(pm, 3, default="GBP")
PM_FX_FROM_SM = f"Summary!{SM_FX}"
SM_NET_EV = f"Summary!{get_column_letter(3+len(SCENARIOS))}{kpi_row_idx['net']}"
SM_NET_BEST = f"Summary!C{kpi_row_idx['net']}"
SM_NET_IDEAL = f"Summary!D{kpi_row_idx['net']}"
SM_NET_WORST = f"Summary!E{kpi_row_idx['net']}"
SM_ROI_IDEAL = f"Summary!D{kpi_row_idx['roi']}"
SM_GP_PCT_IDEAL = f"Summary!D{kpi_row_idx['gp_pct']}"
PF_3YR = f"Portfolio!N{tr}"
PF_BLENDED_ROI = f"Portfolio!O{tr}"
PF_BLENDED_GM = f"Portfolio!F{tr}"

# FX bridge: Pitch Memo wants its own currency; pull EV from Summary which is in Summary's ccy
pm_fx_bridge = f"*({PM_FX}/{PM_FX_FROM_SM})"
pf_fx_bridge = f"*({PM_FX}/{PF_FX.replace('$Y$3','Portfolio!$Y$3')})"

memo_lines = [
    ("THE ASK", f'="Approve the D93 portfolio pitch for Costco UK — seven SKUs, Bee Propolis Spray as lead, '
                'targeting a 26-week market test starting "&TEXT(StartDate,"dd-mmm-yyyy")&"."'),
    ("HEADLINE",
        '="Expected-value Net Profit on the Bee Propolis test: "&$Z$3&" "&'
        f'TEXT({SM_NET_EV}{pm_fx_bridge},"#,##0")&" across BEST/IDEAL/WORST weighted '
        '"&TEXT(PBest,"0%")&"/"&TEXT(PIdeal,"0%")&"/"&TEXT(PWorst,"0%")&"."'),
    ("BEST CASE",
        f'="If BEST plays out: Net Profit "&$Z$3&" "&TEXT({SM_NET_BEST}{pm_fx_bridge},"#,##0")&"."'),
    ("BASE CASE (IDEAL)",
        f'="IDEAL Net Profit "&$Z$3&" "&TEXT({SM_NET_IDEAL}{pm_fx_bridge},"#,##0")&'
        f'", GM% "&TEXT({SM_GP_PCT_IDEAL},"0.0%")&", ROI on promo "&TEXT({SM_ROI_IDEAL},"0%")&"."'),
    ("WORST CASE",
        f'="WORST Net Profit "&$Z$3&" "&TEXT({SM_NET_WORST}{pm_fx_bridge},"#,##0")&'
        '" (markdown clearance Wks 18-20)."'),
    ("3-YEAR PORTFOLIO",
        f'="Blended 3-year Net Profit across all seven SKUs: "&$Z$3&" "&'
        f'TEXT({PF_3YR}{pf_fx_bridge},"#,##0")&". Blended ROI: "&TEXT({PF_BLENDED_ROI},"0%")&"."'),
    ("WHY THIS MIX",
        f'="Bee Propolis carries the GM% anchor ("&TEXT(Portfolio!F{sku_start_row},"0.0%")&'
        '"); lower-margin SKUs contribute volume. Blended GM% = "&TEXT(' + PF_BLENDED_GM + ',"0.0%")&"."'),
    ("KEY ASSUMPTIONS",
        '="Base demand "&TEXT(BaseUnits,"#,##0")&" units/week across "&TEXT(WHs,"#,##0")&" warehouses. '
        'Demo lift 3.0-3.6x. TPD = "&TEXT(TPDPct,"0%")&" off net."'),
    ("RISKS",
        '="(1) Demo lift below 3x → see Sensitivity tab. (2) Inventory overrun forces markdown → WORST. '
        '(3) FX move on COGS — model is GBP-source, refresh FX Table on Inputs."'),
    ("DECISION",
        '="GO / NO-GO with terms by [date]. Test launches W1 = "&TEXT(StartDate,"dd-mmm-yyyy")&"."'),
]

mr = 5
for label, formula in memo_lines:
    wc(pm,mr,2,label,
       font=Font(name="Calibri",size=11,bold=True,color=WHITE),
       fill=fill_section,align=left)
    pm.row_dimensions[mr].height = 22
    mr += 1
    wc(pm,mr,2,formula,font=f_memo_body,align=top_wrap)
    pm.row_dimensions[mr].height = 36
    mr += 1
    mr += 1  # spacer

# =============================================================================
out = "/home/user/my-first-project/Costco_UK_D93_Portfolio_ROI_v9.xlsx"
wb.save(out)
print(f"Saved: {out}")
