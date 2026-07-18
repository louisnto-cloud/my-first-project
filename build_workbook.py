#!/usr/bin/env python3
"""Build the Organika Sparkling competitor launch analysis workbook (MUV launch)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- palette ----
NAVY   = "1F3A5F"   # headers
TEAL   = "2E7D8A"   # section bands
LIGHT  = "EAF1F4"   # zebra
AMBER  = "FFF2CC"   # flags / unverified
GREEN  = "E2EFDA"   # recommendations
WHITE  = "FFFFFF"
GREY   = "595959"

HEAD_FONT  = Font(name="Calibri", bold=True, color=WHITE, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color=WHITE, size=18)
SUB_FONT   = Font(name="Calibri", bold=True, color=NAVY, size=13)
BODY_FONT  = Font(name="Calibri", size=10, color="000000")
SMALL_ITAL = Font(name="Calibri", size=9, italic=True, color=GREY)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)

def fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)

def style_header(ws, row, ncols, color=NAVY):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD_FONT
        cell.fill = fill(color)
        cell.alignment = CENTER
        cell.border = BORDER

def write_table(ws, start_row, headers, rows, widths=None, zebra=True, header_color=NAVY):
    ncols = len(headers)
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, start_row, ncols, header_color)
    r = start_row + 1
    for i, row in enumerate(rows):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if zebra and i % 2 == 1:
                cell.fill = fill(LIGHT)
        r += 1
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    return r

def banner(ws, row, text, ncols, color=TEAL):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHITE, size=11)
    c.fill = fill(color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return row + 1

wb = Workbook()

# ============================================================== COVER
ws = wb.active
ws.title = "Cover"
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGH", [3, 22, 22, 22, 22, 22, 22, 3]):
    ws.column_dimensions[col].width = w

ws.merge_cells("B2:G2")
t = ws["B2"]; t.value = "Functional Sparkling Beverage — Competitor Launch Analysis"
t.font = TITLE_FONT; t.fill = fill(NAVY); t.alignment = CENTER
ws.row_dimensions[2].height = 46

ws.merge_cells("B3:G3")
s = ws["B3"]; s.value = "Prepared for the Organika Sparkling / “MUV” launch decision  ·  US + Canada"
s.font = Font(bold=True, color=WHITE, size=11); s.fill = fill(TEAL); s.alignment = CENTER
ws.row_dimensions[3].height = 22

meta = [
    ("Scope", "7 competitor brands across 3 categories: modern functional sodas, hydration/electrolyte, wellness/collagen sparkling"),
    ("Angles analyzed", "Channel & retail · Formulation & positioning · Marketing & brand build · Pricing & margin · Outcomes"),
    ("Geography", "United States primary, with explicit read-across to Canada"),
    ("Date prepared", "2026-05-31"),
    ("Method", "Multi-source web research, claims cited per-cell on the Sources tab; low-confidence items flagged amber"),
]
r = 5
for k, v in meta:
    ws.cell(row=r, column=2, value=k).font = SUB_FONT
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    cell = ws.cell(row=r, column=3, value=v); cell.font = BODY_FONT; cell.alignment = WRAP_TOP
    ws.row_dimensions[r].height = 30
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
c = ws.cell(row=r, column=2, value="Tabs in this workbook")
c.font = SUB_FONT
r += 1
tabs = [
    "1. Comparison Matrix — all brands, key metrics side by side",
    "2. Channel & Retail — launch channel and expansion sequence",
    "3. Formulation & Positioning — the claim that won, per brand",
    "4. Marketing & Brand Build — founder story, virality engine",
    "5. Pricing & Margin — per-unit, multipack, margin where known",
    "6. Outcomes — revenue, funding, valuations, exits",
    "7. Canada Read-Across — MUV launch implications & recommendation",
    "8. Sources & Confidence — every citation + flagged items",
]
for line in tabs:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    cell = ws.cell(row=r, column=2, value=line); cell.font = BODY_FONT; cell.alignment = WRAP_TOP
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r+2, end_column=7)
note = ws.cell(row=r, column=2, value=(
    "⚠ Key caveat: A discrete “Organika Sparkling” ready-to-drink can could not be verified in public sources. "
    "Organika's documented fizzy/collagen products are powder (Electrolytes + Enhanced Collagen) and liquid collagen water. "
    "If MUV is a new canned RTD, confirm SKU, format, flavours and price with the brand before finalizing."))
note.font = BODY_FONT; note.fill = fill(AMBER); note.alignment = WRAP_TOP
note.border = BORDER

# ============================================================== 1. MATRIX
ws = wb.create_sheet("1. Comparison Matrix")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:N1")
h = ws["A1"]; h.value = "Competitor Comparison Matrix"; h.font = TITLE_FONT; h.fill = fill(NAVY); h.alignment = CENTER
ws.row_dimensions[1].height = 30

headers = ["Brand","Category","Founded","Founder(s)","First / launch channel","Core claim",
           "Format","Key spec","Unit price","Retail doors / reach","Peak revenue","Outcome / exit"]
rows = [
 ["Poppi","Modern functional soda","2016 (as Mother Beverage)","Allison & Stephen Ellsworth","Dallas farmers market → Whole Foods","Prebiotic soda + apple cider vinegar, low sugar","12oz can","~2g fiber, 2–5g sugar, ~25 cal","~$2.25/can (12pk DTC $26.99)","~36,000+ US stores, 120+ retailers","~$500M (2024 est.)","Acquired by PepsiCo, $1.95B, closed May 2025"],
 ["OLIPOP","Modern functional soda","2018","Ben Goodwin & David Lester","40 NorCal indie grocers (Oct 2019)","\"A new kind of soda\" — 9g fiber, gut health","12oz can","up to 9g fiber, 2–5g sugar, ~35 cal","~$3.00/can (12pk DTC $35.99)","~35,000+ US stores","$400M+ (2024)","Independent; $50M Series C @ $1.85B (Feb 2025)"],
 ["LMNT","Hydration / electrolyte","2018","Robb Wolf + Ketogains team","DTC + podcast sponsorship","Zero-sugar, high-sodium, paleo/keto","Powder sticks (+Sparkling can)","1,000mg sodium, 0g sugar","~$1.00–1.50/stick","DTC-led + Amazon/select retail","$206M sales (2023, SEC filing)","Independent (Reg CF)"],
 ["Liquid I.V.","Hydration / electrolyte","2012","Brandin Cohen (+2)","Amazon → Costco (Jan 2019, ~516 whs)","\"Hydration Multiplier\" — hydrate 2–3x faster","Powder sticks","500mg sodium, 11g carbs (orig)","~$0.67/stick (Costco 30pk)","80,000+ US stores (30,000 at exit)","~$200M at exit; $1B+ now","Acquired by Unilever, Sept 2020 (~$500M reported)"],
 ["Nuun","Hydration / electrolyte","2004","Tim Moxey","Cycling/endurance specialty → REI/grocery","Effervescent low-sugar electrolyte tablet","Effervescent tablets (+powder)","300mg sodium, 1g sugar, 15 cal","~$0.60–0.70/serving","Broad grocery/mass/specialty; #1 run/bike/outdoor","$10M+ (2013)","Acquired by Nestlé Health Science, July 2021"],
 ["Vital Proteins Sparkling","Wellness / collagen sparkling","2025 (sparkling line)","Brand owned by Nestlé (acq. 2020)","Mass retail rollout","Beauty-from-within: VERISOL collagen + Vit C","Can","0g sugar, 15 cal, collagen peptides","$2.50/can ($17.99 12pk club)","Mass (Nestlé distribution)","n/a (line is new)","Active — best-capitalized collagen entrant"],
 ["Recess","Wellness / functional sparkling","2018","Benjamin Witte","DTC → retail","Pivoted CBD → \"relaxation\": magnesium + adaptogens","Can","Magnesium L-threonate, adaptogens","Premium","15,000+ stores (Target, Kroger, Sprouts, WF)","n/a","Independent; $30M Series B (Oct 2025), ~$58M raised"],
 ["Aura Bora","Functional sparkling water","2020","Paul & Maddie Voge","Shark Tank → DTC/retail","Herbal/floral sparkling water, 0 cal/sugar","Can","0 cal, 0 sugar, 0 sodium","~$1.99–2.29/can","11,000 stores","$12M (2024)","Majority stake to Next In Natural (Feb 2025)"],
]
# collagen graveyard mini-note rows
end = write_table(ws, 2, headers, rows,
   widths=[16,16,15,18,24,26,16,18,16,24,18,30])
ws.freeze_panes = "A3"
for rr in range(2, end):
    ws.row_dimensions[rr].height = 58

# graveyard callout
end += 1
ws.merge_cells(start_row=end, start_column=1, end_row=end, end_column=12)
g = ws.cell(row=end, column=1, value=(
    "⚠ Collagen-sparkling graveyard: Goldn Hour (CA, closed 2025), BOONS (acq. then discontinued), Luster & Lum (defunct). "
    "Standalone \"beauty collagen sparkling water\" indie brands have repeatedly flopped — narrow positioning, premium price, "
    "no clinical moat, undercapitalized. Collagen-as-feature inside a broader brand (Vital Proteins, Organika) is the survivable model."))
g.font = BODY_FONT; g.fill = fill(AMBER); g.alignment = WRAP_TOP; g.border = BORDER
ws.row_dimensions[end].height = 56

# ============================================================== 2. CHANNEL
ws = wb.create_sheet("2. Channel & Retail")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:E1")
h = ws["A1"]; h.value = "Channel & Retail Strategy — where they launched and why"; h.font = TITLE_FONT; h.fill = fill(NAVY); h.alignment = CENTER
ws.row_dimensions[1].height = 30
headers = ["Brand","Launch channel (and why)","Expansion sequence","Doors / reach today","Canada presence"]
rows = [
 ["Poppi","Dallas farmers market (2016); a Whole Foods buyer approached the booth within ~3 weeks. Natural channel validated the product first.","Farmers market → Whole Foods national → DTC/Amazon (#1 soda on Amazon 2023) → conventional grocery → mass/club (Target, Kroger, Costco, 7-Eleven, Walmart)","~36,000+ US locations across 120+ retailers","Limited/varying at Walmart Canada, Costco Canada, Amazon.ca; full range via specialty e-comm (Natura Market, Healthy Planet)"],
 ["OLIPOP","40 independent NorCal natural grocers (Oct 2019). Deliberately seeded the natural channel before going wide.","Indie CA grocers → DTC/Amazon (heavy 2020) → natural chains (Sprouts, WF, Wegmans) → conventional (Kroger, Safeway) → mass/club (Target, Walmart, Costco) + Starbucks. 4,500 stores by Nov 2020.","~35,000+ US stores","Thin official presence; mainly online specialty importers (Natura Market 15+ flavours); some Walmart.ca / Sobeys ON"],
 ["LMNT","DTC + podcast sponsorship flywheel — almost no traditional retail at launch. Owned the customer relationship from day one.","DTC/subscription core → Amazon (multiple SKUs) → select retail (Vitamin Shoppe, Target) → added Sparkling can line","DTC-led; Amazon + selective retail","Ships to Canada from drinklmnt.com (free shipping over ~$100 USD)"],
 ["Liquid I.V.","Amazon first, then a national Costco launch (Jan 2019) into ~516 warehouses — Costco was the growth engine.","Amazon → Costco national → broad retail. 30,000+ stores at 2020 exit → 80,000+ today under Unilever.","80,000+ US stores","Costco Canada was the FIRST retailer to bring it to Canada (national); also Amazon.ca / major retail"],
 ["Nuun","Endurance/cycling/triathlon specialty shops (2004). Grassroots within the athlete community.","Run/bike specialty → outdoor & natural (REI, Whole Foods) → broad grocery/mass (Target, Walmart, Sam's, Walgreens)","Broad grocery/mass/specialty; #1 in run, bike, outdoor & natural-foods","Dedicated Canadian Retailers page; distributed in Canada"],
 ["Vital Proteins Sparkling","Mass retail rollout (2025) leveraging Nestlé's distribution muscle — skipped the indie-seeding phase entirely.","Big-brand mass launch from the start","Mass (Nestlé)","Not yet verified at Canadian retail"],
 ["Recess","DTC-first (2018), then retail. Built brand world online before shelf.","DTC → retail chains (Target, Kroger, Sprouts, Whole Foods)","15,000+ stores","Not verified"],
 ["Aura Bora","Shark Tank (S12) launchpad → DTC + retail","DTC/Amazon → grocery (natural + conventional)","11,000 stores","Not verified"],
]
end = write_table(ws, 2, headers, rows, widths=[16,40,44,26,40])
ws.freeze_panes = "A3"
for rr in range(2, end):
    ws.row_dimensions[rr].height = 92

end = banner(ws, end+1, "READ-ACROSS: Costco Canada + Amazon.ca + natural channel are the proven beachhead for new functional beverages in Canada (Liquid I.V., Poppi both entered this way). Organika already holds Canadian shelf relationships imported brands must buy their way into.", 5, color=GREEN)
ws.merge_cells(start_row=end-1, start_column=1, end_row=end-1, end_column=5)
ws.cell(row=end-1, column=1).font = Font(bold=True, color="1E5631", size=10)
ws.cell(row=end-1, column=1).fill = fill(GREEN)
ws.row_dimensions[end-1].height = 44

# ============================================================== 3. FORMULATION
ws = wb.create_sheet("3. Formulation & Positioning")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:E1")
h = ws["A1"]; h.value = "Formulation & Positioning — the claim that won"; h.font = TITLE_FONT; h.fill = fill(NAVY); h.alignment = CENTER
ws.row_dimensions[1].height = 30
headers = ["Brand","Winning claim / positioning","Hero ingredient & key spec","Positioning evolution","Claim risk / notes"]
rows = [
 ["Poppi","\"Prebiotic soda\" — feel-good soda you can drink daily","Prebiotic fiber + ACV + fruit juice; ~2g fiber, 2–5g sugar, ~25 cal","Started as medicinal ACV \"tonic\" (Mother Beverage) → rebranded 2020 to bright, social-first \"prebiotic soda\"","$8.9M class-action settlement (2025) over gut-health claims — 2g fiber called too little to deliver benefit"],
 ["OLIPOP","\"A new kind of soda\" with serious fiber for gut/digestive health","Up to 9g fiber (cassava, chicory, Jerusalem artichoke) + botanicals; ~35 cal","Stable gut-health/fiber positioning since launch — no pivot; ~4x Poppi's fiber is the differentiator","No FTC action or lawsuit found; higher fiber = stronger substantiation"],
 ["LMNT","Zero-sugar, high-sodium electrolytes for paleo/keto & athletes","1,000mg sodium, 200mg K, 60mg Mg, 0g sugar per stick","Salt-first, science-forward; doubled down on \"more sodium than rivals\"","Sodium level is the polarizing hook — loved by keto/athletes, off-putting to mainstream"],
 ["Liquid I.V.","\"Hydration Multiplier\" — hydrate 2–3x faster than water","Cellular Transport Technology; 500mg sodium, 11g carbs (sugar+stevia); +vitamins","Added sugar-free line later to counter the sugar objection","CTT efficacy claim is marketing-led; sugar content a vulnerability vs zero-sugar rivals"],
 ["Nuun","Effervescent low-sugar electrolyte tablet — separates electrolytes from carbs","300mg sodium, 1g sugar, 15 cal, stevia; 10 tablets/tube","Pioneered carb-free electrolytes for endurance; later added powders/RTD","Clean low-sugar story; tablet format is a love-it-or-leave-it differentiator"],
 ["Vital Proteins Sparkling","Beauty-from-within with a clinical collagen claim at mass price","VERISOL collagen peptides + 100% DV Vitamin C; 0g sugar, 15 cal","Brand-extends Nestlé's #1 collagen powder into a can","Strongest collagen claim of the set (clinical ingredient + big-brand trust)"],
 ["Recess","Relaxation / functional mood — sells the benefit, not the molecule","Magnesium L-threonate + adaptogens (Mood line)","Pivoted CBD → relaxation when consumers bought the SOLUTION not the ingredient (CBD now <8% of sales)","Key lesson: lead with the felt benefit, not the buzzy ingredient"],
 ["Aura Bora","Herbal/floral flavor-led sparkling water, zero everything","0 cal / 0 sugar / 0 sodium; herb+fruit+flower flavors","Flavor-forward, no health-claim baggage","Low regulatory risk; competes on taste & brand, not function"],
]
end = write_table(ws, 2, headers, rows, widths=[16,38,34,38,38])
ws.freeze_panes = "A3"
for rr in range(2, end):
    ws.row_dimensions[rr].height = 84
end = banner(ws, end+1, "READ-ACROSS: \"Collagen-as-hero in a can\" has a graveyard. Winners either (a) wrap collagen in a big brand + clinical claim at mass price (Vital Proteins), or (b) sell a broader felt benefit (Recess relaxation). For MUV, lead with a functional benefit + \"by Canada's #1 collagen brand\" halo rather than narrow beauty-from-within.", 5, color=GREEN)
ws.merge_cells(start_row=end-1, start_column=1, end_row=end-1, end_column=5)
ws.cell(row=end-1, column=1).font = Font(bold=True, color="1E5631", size=10)
ws.cell(row=end-1, column=1).fill = fill(GREEN)
ws.row_dimensions[end-1].height = 44

# ============================================================== 4. MARKETING
ws = wb.create_sheet("4. Marketing & Brand Build")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:D1")
h = ws["A1"]; h.value = "Marketing & Brand Build — the virality engine"; h.font = TITLE_FONT; h.fill = fill(NAVY); h.alignment = CENTER
ws.row_dimensions[1].height = 30
headers = ["Brand","Founder story","Growth / virality engine","What specifically drove it"]
rows = [
 ["Poppi","Allison & Stephen Ellsworth; Shark Tank Dec 2018 ($400K for 25% from Rohan Oza)","TikTok-first influencer + two consecutive Super Bowl ads (2024, 2025)","Bright pink aesthetic, creator seeding (32 branded vending machines to creators in 2025 — sparked viral debate), Alix Earle / Jake Shane"],
 ["OLIPOP","Ben Goodwin (formulator, ex-Obi) & David Lester; bootstrapped first cans with $100K from selling Obi","Strong DTC repurchase rates + reactive/competitive social","\"Soda Stories\" campaign; opportunistic social (jabbed at Poppi's vending-machine spend during 2025 Super Bowl) — earned media without paying for the Big Game"],
 ["LMNT","Robb Wolf (biochemist, Paleo author) + Ketogains team (Luis Villaseñor)","Podcast-sponsorship moat — Huberman, Attia, Modern Wisdom, etc. (~10,000 episodes across ~419 podcasts)","Trusted-host endorsement + \"free sample pack with purchase\" funnel. Spent $54.6M on marketing & $7.1M on sampling in 2023 (~26.5% of revenue)"],
 ["Liquid I.V.","Brandin Cohen, college brainchild (+ 2 co-founders)","Influencer + heavy retail sampling + give-back (\"buy one, donate a serving\")","Costco roadshow sampling at scale; cause marketing; later a Unilever-backed Super Bowl ad"],
 ["Nuun","Tim Moxey, British Ironman athlete; idea born running out of drink on a hot ride","Grassroots endurance-community/event marketing","Authentic credibility in run/bike/tri communities; #1 in those specialty channels before going mainstream"],
 ["Vital Proteins Sparkling","Brand founded by Kurt Seidensticker; acquired by Nestlé 2020","Big-brand media + retail muscle + existing collagen loyalty","Leverages Nestlé distribution and the #1-collagen halo rather than scrappy DTC virality"],
 ["Recess","Benjamin Witte","Distinctive pastel brand world + culturally-tuned social","Owned a vibe/aesthetic and the \"relaxation\" cultural moment; design-led brand recall"],
 ["Aura Bora","Paul & Maddie Voge","Shark Tank (Herjavec $200K/15%) + witty, irreverent brand voice","Personality-driven marketing and flavor novelty; punched above its ad budget"],
]
end = write_table(ws, 2, headers, rows, widths=[16,40,40,52])
ws.freeze_panes = "A3"
for rr in range(2, end):
    ws.row_dimensions[rr].height = 86
end = banner(ws, end+1, "READ-ACROSS: The two cheapest, highest-ROI playbooks here are LMNT's trusted-host podcast model and OLIPOP's reactive earned-media social — both winnable for a challenger brand without Super Bowl budgets. Pair with sampling (Costco roadshows) which converted for Liquid I.V.", 4, color=GREEN)
ws.merge_cells(start_row=end-1, start_column=1, end_row=end-1, end_column=4)
ws.cell(row=end-1, column=1).font = Font(bold=True, color="1E5631", size=10)
ws.cell(row=end-1, column=1).fill = fill(GREEN)
ws.row_dimensions[end-1].height = 44

# ============================================================== 5. PRICING
ws = wb.create_sheet("5. Pricing & Margin")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:F1")
h = ws["A1"]; h.value = "Pricing & Margin Structure"; h.font = TITLE_FONT; h.fill = fill(NAVY); h.alignment = CENTER
ws.row_dimensions[1].height = 30
headers = ["Brand","Unit price","Multipack / DTC","Channel mix","Gross margin","Notes"]
rows = [
 ["Poppi","~$2.25/can","12-pack DTC ~$26.99; Costco US 15-pk $24.99 (~$1.67/can)","DTC + retail (split not disclosed; ~250% online growth 2022)","~40–50% (below PepsiCo's typical 55–60%)","Mass-priced to sit near conventional soda"],
 ["OLIPOP","~$3.00/can","12-pack DTC ~$35.99","~30% e-comm / 70% physical retail","Not disclosed","Priced at a premium to Poppi; higher fiber justifies it"],
 ["LMNT","~$1.00–1.50/stick","30-count box; Insider bundle <$1/stick w/ subscribe","DTC-led (subscription core)","~20% NET margin (2023 SEC filing)","Rare hard margin data — and it's net, not gross; healthy for DTC CPG"],
 ["Liquid I.V.","~$0.67/stick (Costco)","30-count variety ~$19.97","Costco/retail heavy","Not disclosed (Unilever-owned)","Club-channel volume model; low per-stick price, high units"],
 ["Nuun","~$0.60–0.70/serving","Tube of 10 tablets ~$6–7","Specialty + grocery/mass","Not disclosed","Tablet format = low COGS per serving"],
 ["Vital Proteins Sparkling","$2.50/can (SRP)","$17.99 / 12-pk club","Mass retail","Not disclosed","Aggressive mass price for a collagen drink — sets the ceiling"],
 ["Recess","Premium","~$20–30 / multipack","DTC + retail","Not disclosed","Premium positioning"],
 ["Aura Bora","~$1.99–2.29/can","~$30–33 / 12-pack","DTC + retail","Not disclosed","Accessible everyday price point"],
]
end = write_table(ws, 2, headers, rows, widths=[16,16,28,26,24,34])
ws.freeze_panes = "A3"
for rr in range(2, end):
    ws.row_dimensions[rr].height = 62

# US vs Canada price block
end += 1
end = banner(ws, end, "US vs CANADA PRICING", 6, color=TEAL)
cad = [
 ["Poppi","~$1.67/can US (Costco 15pk) / $2.00–2.50 regular","~$3.49 CAD/can at Canadian specialty retail","~40–100% higher in Canada (FX, freight, bilingual packaging, importer margin)"],
 ["Vital Proteins","$2.50/can US","No verified Canadian retail price yet","—"],
 ["General read","—","—","Imported functional drinks carry a meaningful Canada premium. A domestic Canadian producer (Organika) has a structural cost/price advantage."],
]
end = write_table(ws, end, ["Brand","US price","Canada price","Spread / note"], cad, widths=None, header_color=TEAL)
for rr in range(end-3, end):
    ws.row_dimensions[rr].height = 40
end = banner(ws, end+1, "READ-ACROSS: Canadian shelf norm for premium functional cans is ~$2.50–$3.49 CAD. As a domestic producer Organika can price competitively in that band AND protect margin where imported US brands cannot. Anchor MUV at/below $3.49 CAD.", 6, color=GREEN)
ws.merge_cells(start_row=end-1, start_column=1, end_row=end-1, end_column=6)
ws.cell(row=end-1, column=1).font = Font(bold=True, color="1E5631", size=10)
ws.cell(row=end-1, column=1).fill = fill(GREEN)
ws.row_dimensions[end-1].height = 44

# ============================================================== 6. OUTCOMES
ws = wb.create_sheet("6. Outcomes")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:F1")
h = ws["A1"]; h.value = "Outcomes — revenue, funding, valuations, exits"; h.font = TITLE_FONT; h.fill = fill(NAVY); h.alignment = CENTER
ws.row_dimensions[1].height = 30
headers = ["Brand","Revenue trajectory","Funding / valuation","Exit","Category share","Profitability"]
rows = [
 ["Poppi","$13M (2020) → $26M (2021) → $65M (2022) → $100M+ (2023) → ~$500M (2024 est.)","VC-backed pre-exit","Acquired by PepsiCo — $1.95B, closed May 19, 2025","~38% of \"modern soda\" (2024)","n/a (acquired)"],
 ["OLIPOP","$852K (2018) → $200M+ (2023) → $400M+ (2024)","$50M Series C @ $1.85B valuation (Feb 2025, JPM Private Capital)","Independent","~32.7% of modern soda (2024)","Cash-profitable early 2024"],
 ["LMNT","$206M sales (2023, SEC Reg CF filing)","Reg CF / Republic investor base","Independent","n/a","~20% net income margin (2023)"],
 ["Liquid I.V.","~$100M within first ~5 yrs → ~$200M at exit → $1B+ now","—","Acquired by Unilever, Sept 2020 (~$500M / ~5x revenue reported; terms undisclosed)","n/a","Unilever later invested $80M+ in a MO plant"],
 ["Nuun","$10M+ (2013)","TSG Consumer Partners growth equity (2017)","Acquired by Nestlé Health Science, July 2021 (terms undisclosed)","#1 in run/bike/outdoor/natural channels","n/a"],
 ["Vital Proteins","(brand) part of Nestlé Health Science","Acquired by Nestlé 2020","Owned by Nestlé","n/a","n/a"],
 ["Recess","—","$30M Series B Oct 2025 (CAVU-led); ~$58M total raised","Independent","n/a","CBD <8% of sales; Mood line ~95%"],
 ["Aura Bora","$12M (2024)","~$22M raised 2019–24","Majority stake to Next In Natural (Feb 2025)","n/a","n/a"],
]
end = write_table(ws, 2, headers, rows, widths=[16,42,36,38,26,28])
ws.freeze_panes = "A3"
for rr in range(2, end):
    ws.row_dimensions[rr].height = 66
end += 1
ws.merge_cells(start_row=end, start_column=1, end_row=end, end_column=6)
m = ws.cell(row=end, column=1, value="Category context: \"Modern soda\" = $1.8B in 2024 (+83% vs $983M in 2023). Collagen sparkling drink market est. $1.2B (2024) → $3.5B by 2034 (11.5% CAGR) — single-source, treat as directional.")
m.font = SMALL_ITAL; m.alignment = WRAP_TOP; m.fill = fill(LIGHT); m.border = BORDER
ws.row_dimensions[end].height = 32

# ============================================================== 7. CANADA
ws = wb.create_sheet("7. Canada Read-Across")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:D1")
h = ws["A1"]; h.value = "Canada Read-Across — MUV Launch Implications"; h.font = TITLE_FONT; h.fill = fill(NAVY); h.alignment = CENTER
ws.row_dimensions[1].height = 30

r = 3
r = banner(ws, r, "ORGANIKA'S STARTING POSITION", 4)
org = [
 ["Company","Canadian, founded 1990, family-run, manufactures in British Columbia; 200+ NHPs"],
 ["Market position","Canada's #1 selling collagen (powder); ~92% CAGR over 3 yrs (brand self-reported)"],
 ["Existing distribution","Costco Canada, Well.ca, Amazon.ca, Vitasave, Natura Market, independent pharmacies & health-food stores"],
 ["Documented fizzy/collagen SKUs","Electrolytes + Enhanced Collagen (fizzy powder); Liquid Marine Collagen / collagen water (liquids). A canned RTD \"Sparkling\" SKU is UNVERIFIED."],
]
r = write_table(ws, r, ["Dimension","Detail"], org, widths=[26,84])
for rr in range(r-len(org), r):
    ws.row_dimensions[rr].height = 34

r += 1
r = banner(ws, r, "REGULATORY — THE KEY GOTCHA (verify before finalizing)", 4, color="B45309")
reg = [
 ["NHP vs Food jurisdiction","Per Health Canada, beverages in food format fall OUTSIDE the scope of Natural Health Products — they're regulated as FOOD (CFIA/Food Directorate). A canned RTD collagen sparkling beverage generally CANNOT carry an NPN the way Organika's powders/liquids do."],
 ["Claims impact","Food health claims are far more restricted than NHP-licensed claims. Organika's NPN-backed beauty/collagen claims do NOT automatically transfer to a can."],
 ["NPN","8-digit Natural Product Number; only licensed NHPs bear one. Likely unavailable for an RTD food-format drink."],
 ["Labelling","Bilingual (English + French) mandatory, including a bilingual Nutrition Facts table."],
]
r = write_table(ws, r, ["Item","Implication"], reg, widths=[26,84], header_color="B45309")
for rr in range(r-len(reg), r):
    ws.row_dimensions[rr].height = 50

r += 1
r = banner(ws, r, "RECOMMENDATIONS FOR THE MUV LAUNCH", 4, color="1E7D32")
rec = [
 ["1. Format & claims","Assume the can is a FOOD (CFIA), not an NHP. Build messaging on permissible food claims + a \"made by Canada's #1 collagen brand\" halo. Get regulatory sign-off on every on-pack claim FIRST — highest-priority de-risking step."],
 ["2. Channel","Lead with Costco Canada + Amazon.ca + natural channel (Whole Foods Canada, Healthy Planet) + Organika's existing pharmacy/health-food base. This mirrors how Liquid I.V. and Poppi entered Canada — and Organika already has the shelf relationships."],
 ["3. Positioning","Avoid the standalone \"beauty collagen sparkling water\" graveyard (Goldn Hour, BOONS, Luster & Lum). Sell a broader felt/functional benefit (à la Recess relaxation) or ride the big-brand + clinical-claim model (à la Vital Proteins). Benefit-forward, not narrow beauty-from-within."],
 ["4. Pricing","Anchor at/below the Canadian premium-functional shelf norm of ~$2.50–$3.49 CAD/can. As a domestic producer, undercut imported US premiums while protecting margin."],
 ["5. Marketing","Run the cheap, high-ROI challenger playbooks: trusted-host podcast sponsorship (LMNT) + reactive earned-media social (OLIPOP) + Costco roadshow sampling (Liquid I.V.). Skip Super Bowl-tier spend."],
 ["6. Moat","Differentiate on a credible, substantiated benefit (OLIPOP's 9g fiber beat Poppi's 2g in court). Don't over-claim — Poppi paid $8.9M for a thin gut-health claim. A real, defensible spec is the moat."],
]
r = write_table(ws, r, ["#","Recommendation"], rec, widths=[20,90], header_color="1E7D32")
for rr in range(r-len(rec), r):
    ws.row_dimensions[rr].height = 62

# ============================================================== 8. SOURCES
ws = wb.create_sheet("8. Sources & Confidence")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:C1")
h = ws["A1"]; h.value = "Sources & Confidence"; h.font = TITLE_FONT; h.fill = fill(NAVY); h.alignment = CENTER
ws.row_dimensions[1].height = 30

r = 3
r = banner(ws, r, "FLAGGED / LOW-CONFIDENCE ITEMS — confirm before publishing", 3, color="B45309")
flags = [
 ["\"Organika Sparkling\" / MUV exact SKU, format, flavours, price","Could not verify a discrete canned RTD; documented products are powder + liquid collagen","Confirm directly with brand"],
 ["Poppi DTC-vs-retail revenue split","Not disclosed","Unverified"],
 ["Poppi 2024 ~$500M revenue","Estimate from analytics aggregator","Directional"],
 ["OLIPOP & most brands' gross margin","Not disclosed (private)","Unverified"],
 ["Liquid I.V. ~$500M / 5x exit value","Press estimate; Unilever says terms undisclosed","Secondary"],
 ["Nuun & Vital Proteins / Nestlé deal values","Terms undisclosed","Unverified"],
 ["LMNT podcast episode count / \"$300M+\" revenue","Social-media estimates","Unverified (use SEC $206M figure instead)"],
 ["Collagen market $1.2B→$3.5B & Organika 92% CAGR","Single-source / self-reported","Directional"],
 ["Canadian per-can prices","From one specialty retailer; banner pricing will be lower","Directional"],
]
r = write_table(ws, r, ["Claim","Why flagged","Status"], flags, widths=[48,42,22], header_color="B45309")
for rr in range(r-len(flags), r):
    ws.row_dimensions[rr].height = 30

r += 1
r = banner(ws, r, "PRIMARY & SECONDARY SOURCES", 3, color=TEAL)
src = [
 ["PepsiCo newsroom — Poppi acquisition ($1.95B, closed May 2025)","pepsico.com/newsroom","Primary"],
 ["Poppi $8.9M gut-health class-action settlement","classaction.org; bevnet.com","Primary/Trade"],
 ["OLIPOP $50M Series C @ $1.85B; revenue & profitability","cnbc.com; bevindustry.com; inc.com","Primary/Trade"],
 ["Modern soda category $1.8B 2024 (+83%); brand shares","bevindustry.com","Trade"],
 ["LMNT FY2023 financials ($206M sales, ~20% net, $54.6M mktg, $7.1M sampling)","SEC Reg CF filing, CIK 1871551","Primary (strongest)"],
 ["Liquid I.V. Costco national launch (Jan 2019, ~516 whs)","liquid-iv.com","Primary"],
 ["Liquid I.V. acquired by Unilever (Sept 2020, terms undisclosed)","unilever.com; bevnet.com","Primary/Trade"],
 ["Unilever $80M+ Jefferson City plant for Liquid I.V.","ded.mo.gov","Primary"],
 ["Nuun acquired by Nestlé Health Science (July 2021)","nestle.com; businesswire.com","Primary"],
 ["Vital Proteins sparkling launch ($2.50/can, VERISOL)","fooddive.com; wwd.com","Trade"],
 ["Recess $30M Series B / relaxation pivot","businesswire.com; beveragedaily.com","Primary/Trade"],
 ["Aura Bora $12M rev / 11,000 stores / Next In Natural stake","inc.com; sharktankblog.com","Trade"],
 ["Liquid I.V. first-to-Canada via Costco Canada","foodincanada.com; costco.ca","Trade/Primary"],
 ["Health Canada NHP labelling & food-format scope","canada.ca/health-canada","Primary (regulatory)"],
 ["CFIA food health claims","inspection.canada.ca","Primary (regulatory)"],
 ["Organika company & distribution","organika.com; costco.ca; well.ca; BC gov export catalogue","Primary"],
 ["Poppi/OLIPOP Canada availability","naturamarket.ca; healthyplanetcanada.com; walmart.ca","Retail"],
]
r = write_table(ws, r, ["Claim / topic","Source(s)","Type"], src, widths=[58,40,18], header_color=TEAL)
for rr in range(r-len(src), r):
    ws.row_dimensions[rr].height = 26
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
n = ws.cell(row=r, column=1, value="Note: Several primary publisher pages (Entrepreneur, CNBC, Wikipedia, some brand sites) block automated fetch (HTTP 403); figures were drawn from search-result extracts of those same outlets plus corroborating trade press. Recommend a manual click-through on the PepsiCo, SEC and BevNET links to lock final numbers before external publication.")
n.font = SMALL_ITAL; n.alignment = WRAP_TOP; n.fill = fill(LIGHT); n.border = BORDER
ws.row_dimensions[r].height = 44

wb.save("/home/user/my-first-project/Organika_Sparkling_Competitor_Launch_Analysis.xlsx")
print("workbook saved; sheets:", wb.sheetnames)
