#!/usr/bin/env python3
"""Loop batch I5+I6+I8:
 I5 Financial Model usability — input/formula legend, cell comments, lock formulas (inputs stay editable).
 I6 Sources — make the first source in each 'Source(s)' cell a clickable link.
 I8 MÜV Peer Set — sodium vs price positioning scatter (labeled per brand)."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.comments import Comment
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; TEAL="2E7D8A"; STEEL="3E5C76"; LIGHT="EAF1F4"; AMBER="FFF2CC"; AMBERHEAD="B45309"; GREEN="E2EFDA"; GREENHEAD="1E7D32"; WHITE="FFFFFF"; GREY="595959"; LINKBLUE="1155CC"; YELHEX="FFF7CC"
def F(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

# ---------- I5: Financial Model usability ----------
fm=wb["Financial Model"]
# legend (top-right area, cols G-H around row 4-6, which are empty)
fm.cell(row=6,column=6,value="LEGEND").font=F(bold=True,size=9,color=NAVY)
lg1=fm.cell(row=7,column=6,value="  edit me  "); lg1.fill=fill(YELHEX); lg1.font=F(bold=True,size=9); lg1.border=BORDER; lg1.alignment=Alignment(horizontal="center")
fm.cell(row=7,column=7,value="= yellow input cell").font=F(size=9,color=GREY)
lg2=fm.cell(row=8,column=6,value="  formula  "); lg2.fill=fill(LIGHT); lg2.font=F(size=9); lg2.border=BORDER; lg2.alignment=Alignment(horizontal="center")
fm.cell(row=8,column=7,value="= calculated (locked)").font=F(size=9,color=GREY)
# comments
def addcomment(cell, text):
    try: fm[cell].comment=Comment(text,"MÜV model")
    except Exception: pass
addcomment("C4","Pick Conservative / Base / Aggressive. The whole P&L, break-even and sensitivity recompute from this one cell.")
# find the price/cogs active-input rows to comment (search col B labels)
for r in range(1,60):
    lbl=fm.cell(row=r,column=2).value
    if lbl=="List price / can (CAD)": addcomment(f"C{r}","Shelf price per can. Edit per scenario in cols C/D/E.")
    if lbl=="COGS / can (CAD)": addcomment(f"C{r}","All-in cost/can (liquid+can+co-pack). Remember freight/duty are extra.")
# lock formulas, keep yellow inputs editable
for row in fm.iter_rows():
    for c in row:
        rgb = getattr(getattr(c.fill,"fgColor",None),"rgb",None)
        is_input = isinstance(rgb,str) and rgb.upper().endswith(YELHEX)
        c.protection = Protection(locked=not is_input)
fm.protection = SheetProtection(sheet=True, selectLockedCells=False, selectUnlockedCells=False,
                                formatCells=False, formatColumns=False, formatRows=False, sort=False, autoFilter=False)

# ---------- I6: Sources clickable ----------
def linkify_sources(ws):
    # find 'Source(s)' or 'Source(s)'/'Source' header column
    hdr_col=hdr_row=None
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and c.value.strip() in ("Source(s)","Source","Source(s) "):
                hdr_col=c.column; hdr_row=c.row; break
        if hdr_col: break
    if not hdr_col: return 0
    n=0
    for r in range(hdr_row+1, ws.max_row+1):
        c=ws.cell(row=r,column=hdr_col)
        if not isinstance(c.value,str) or not c.value.strip(): continue
        first=c.value.replace(",",";").split(";")[0].strip()
        # take a domain-looking token
        tok=first.split()[-1] if " " in first and "." in first.split()[-1] else first
        if "." not in tok: continue
        url = tok if tok.startswith("http") else "https://"+tok.lstrip("/")
        try:
            c.hyperlink=url; c.font=F(size=c.font.size or 10,color=LINKBLUE,underline="single"); n+=1
        except Exception: pass
    return n
linked = linkify_sources(wb["24 Sources"])

# ---------- I8: MÜV Peer Set sodium vs price scatter ----------
ps=wb["MÜV Peer Set"]
base=ps.max_row+3
ps.cell(row=base,column=1,value="Positioning: sodium (mg) vs price/serving — data").font=F(size=9,italic=True,color=GREY)
pts=[("LMNT",1000,1.50),("Liquid I.V.",500,1.56),("Nuun",300,0.75),("Organika sachets",440,1.00)]
# header
ps.cell(row=base+1,column=1,value="Brand"); ps.cell(row=base+1,column=2,value="Sodium"); ps.cell(row=base+1,column=3,value="Price")
for i,(b,na,pr) in enumerate(pts):
    ps.cell(row=base+2+i,column=1,value=b); ps.cell(row=base+2+i,column=2,value=na); ps.cell(row=base+2+i,column=3,value=pr)
ch=ScatterChart(); ch.title="Electrolyte positioning — Sodium (mg) vs Price/serving"; ch.x_axis.title="Sodium mg/serving"; ch.y_axis.title="Price/serving ($)"
ch.height=8.5; ch.width=16; ch.style=2
for i,(b,na,pr) in enumerate(pts):
    xref=Reference(ps,min_col=2,min_row=base+2+i,max_row=base+2+i)
    yref=Reference(ps,min_col=3,min_row=base+2+i,max_row=base+2+i)
    s=Series(yref,xref,title=b); s.marker.symbol="circle"; s.marker.size=9; s.graphicalProperties.line.noFill=True
    ch.series.append(s)
ch.x_axis.delete=False; ch.y_axis.delete=False
anchor="E"+str(base)
ps.add_chart(ch,anchor)
note=ps.cell(row=base+2+len(pts)+1,column=1,value="MÜV sodium/serving TBD (confirm on-can) — plot when known. Aim: daily-wellness zone (moderate sodium, mid price).")
note.font=F(size=8,italic=True,color=GREY)
for rr in range(base,base+2+len(pts)+2): ps.row_dimensions[rr].hidden=True

wb.save(PATH)
print(f"I5 model-lock ok; I6 linked {linked} sources; I8 scatter added.")
