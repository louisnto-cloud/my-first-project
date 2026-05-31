"""
Costco UK D93 Portfolio ROI Model — v8 (simplified)

Same look-and-feel as v7, but trimmed for readability:
  - Inputs: ~10 rows total. Decomposed "rate × factor = total" rolled into
    single editable cells. Fence and Advertising lines removed.
  - Forecast: 6 P&L rows per scenario (Units / Revenue / COGS / GP /
    Promo / Net). Per-promo breakdown is shown on Summary instead.
  - Summary: 7 KPIs (was 16).
  - Portfolio: 11 columns (was 17).
  - Currency widget on every tab compressed to 3 cells.
"""

import openpyxl
from datetime import date
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

NAVY="1F4E78"; BLUE="2E75B6"; LIGHT_BLUE="DDEBF7"
YELLOW="FFF2CC"; GREY="F2F2F2"; WHITE="FFFFFF"
GREEN="C6EFCE"; RED="FFC7CE"; DARK_GREY="595959"

thin = Side(border_style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

f_title       = Font(name="Calibri", size=18, bold=True, color=WHITE)
f_section     = Font(name="Calibri", size=12, bold=True, color=WHITE)
f_subhead     = Font(name="Calibri", size=11, bold=True, color=NAVY)
f_label       = Font(name="Calibri", size=11, bold=True)
f_input       = Font(name="Calibri", size=11, color="0070C0", bold=True)
f_calc        = Font(name="Calibri", size=11)
f_note        = Font(name="Calibri", size=10, italic=True, color=DARK_GREY)
f_kpi_val     = Font(name="Calibri", size=14, bold=True, color=NAVY)
f_white_bold  = Font(name="Calibri", size=11, bold=True, color=WHITE)

fill_title   = PatternFill("solid", fgColor=NAVY)
fill_section = PatternFill("solid", fgColor=BLUE)
fill_input   = PatternFill("solid", fgColor=YELLOW)
fill_calc    = PatternFill("solid", fgColor=GREY)
fill_kpi     = PatternFill("solid", fgColor=LIGHT_BLUE)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
right  = Alignment(horizontal="right", vertical="center")
top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

FMT_GBP_HARD   = '_-£* #,##0_-;[Red]-£* #,##0_-;_-£* "-"??_-;_-@_-'
FMT_GBP_HARD_D = '_-£* #,##0.00_-;[Red]-£* #,##0.00_-;_-£* "-"??_-;_-@_-'
FMT_MONEY      = '_-* #,##0_-;[Red]_-* (#,##0)_-;_-* "-"??_-;_-@_-'
FMT_MONEY_D    = '_-* #,##0.00_-;[Red]_-* (#,##0.00)_-;_-* "-"??_-;_-@_-'
FMT_INT  = "#,##0"
FMT_PCT  = "0.0%"
FMT_MULT = '0.0"x"'
FMT_DATE = "dd-mmm"
FMT_DATE_L = "dd-mmm-yyyy"
FMT_RATE = "0.0000"

WEEKS = 26
FIRST_WEEK_NUM = 1
DEFAULT_START_DATE = date(2026, 1, 12)

CURRENCIES = [
    ("GBP", "£",  1.0000),
    ("USD", "$",  1.2700),
    ("CAD", "C$", 1.7100),
    ("EUR", "€",  1.1700),
]

# Simplified promo set — Fence and Advertising removed
PROMO_TYPES = ["Demo", "End Cap", "TPD", "Markdown"]

SCENARIOS = [
    {"key":"BEST","name":"BEST · Demo + TPD (Wks 5-6)",
     "promo":{"Demo":[5,6],"End Cap":[],"TPD":[5,6],"Markdown":[]},
     "lift_overrides":{5:3.0,6:3.0},
     "rationale":"Two demo weekends with TPD. ~3x lift in-promo weeks."},
    {"key":"IDEAL","name":"IDEAL · Demo + End Cap + TPD (Wks 5-6)",
     "promo":{"Demo":[5,6],"End Cap":[5,6],"TPD":[5,6],"Markdown":[]},
     "lift_overrides":{5:3.6,6:3.6},
     "rationale":"Stacked demo + end cap + TPD. ~3.6x lift."},
    {"key":"WORST","name":"WORST · Launch then markdown clearance",
     "promo":{"Demo":[5,6],"End Cap":[5,6],"TPD":[5,6],"Markdown":[18,19,20]},
     "lift_overrides":{5:3.0,6:3.0,18:1.0,19:0.3,20:8.7},
     "rationale":"Launch as IDEAL, then 50% markdown to clear Wks 18-20."},
]

PORTFOLIO_SKUS = [
    {"rank":1,"name":"Bee Propolis Spray 30ml x2","sp":15.99,"cost":7.29,"units":17000,"promo":25000,
     "rationale":"Lead pitch — strong GP%."},
    {"rank":2,"name":"Mag 8-in-1 90's/120's","sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":3,"name":"O1 450g","sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":4,"name":"Belli Bliss Raspberry, 450g","sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":5,"name":"Daily Boost","sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":6,"name":"Berberine 90's/120's","sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
    {"rank":7,"name":"L-theanine Capsules 90's/120's","sp":None,"cost":None,"units":None,"promo":None,"rationale":None},
]

wb = openpyxl.Workbook()
wb.remove(wb.active)


def wc(ws, r, c, v, *, font=None, fill=None, align=None, border=None, number_format=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font is not None: cell.font = font
    if fill is not None: cell.fill = fill
    if align is not None: cell.alignment = align
    if border is not None: cell.border = border
    if number_format is not None: cell.number_format = number_format
    return cell


def section_header(ws, r, text, span_start=2, span_end=4):
    ws.merge_cells(start_row=r, start_column=span_start, end_row=r, end_column=span_end)
    wc(ws, r, span_start, text, font=f_section, fill=fill_section, align=left)
    ws.row_dimensions[r].height = 22


def add_currency_widget(ws, row, default="GBP"):
    """Compact 3-cell widget: label | dropdown | live "= $ 1.27" summary.
    FX rate and symbol are still pulled via VLOOKUP into hidden helper cells
    at Y/Z so other formulas can reference them."""
    wc(ws, row, 2, "Display Currency", font=f_label, align=right)
    wc(ws, row, 3, default, font=f_input, fill=fill_input, align=center, border=box)
    # Single info cell: "(£ · rate 1.0000)"
    wc(ws, row, 4,
       f'="("&VLOOKUP(C{row},Inputs!$B$13:$D$16,2,FALSE)&"  ·  rate "&TEXT(VLOOKUP(C{row},Inputs!$B$13:$D$16,3,FALSE),"0.0000")&")"',
       font=f_note, align=left)

    # Helper cells (off to the right, narrow column) used by display formulas
    wc(ws, row, 25, f"=VLOOKUP(C{row},Inputs!$B$13:$D$16,3,FALSE)",
       font=f_calc, number_format=FMT_RATE)   # Y{row} = FX rate
    wc(ws, row, 26, f"=VLOOKUP(C{row},Inputs!$B$13:$D$16,2,FALSE)",
       font=f_calc)                            # Z{row} = symbol

    dv = DataValidation(type="list", formula1='"GBP,USD,CAD,EUR"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"C{row}")
    ws.row_dimensions[row].height = 22
    return (f"$Y${row}", f"$Z${row}")


# =============================================================================
# README
# =============================================================================
rd = wb.create_sheet("README")
rd.sheet_view.showGridLines = False
rd.column_dimensions["A"].width = 4
rd.column_dimensions["B"].width = 105

wc(rd,2,2,"Costco UK — D93 Portfolio ROI",
   font=Font(name="Calibri",size=20,bold=True,color=NAVY),align=left)
wc(rd,3,2,"v8 · simplified",
   font=Font(name="Calibri",size=12,italic=True,color=DARK_GREY),align=left)

wc(rd,5,2,"HOW TO USE",font=Font(name="Calibri",size=13,bold=True,color=WHITE),
   fill=fill_section,align=left)

lines = [
    "",
    "1.  Open Inputs. Edit any YELLOW cell. Calculated cells (grey) update automatically.",
    "",
    "2.  Pick a Display Currency at the top of any tab. Each tab is independent — set them",
    "    all the same for a consistent pitch view.",
    "",
    "3.  Forecast: 26-week P&L for the Bee Propolis market test under three scenarios.",
    "    Summary: side-by-side KPIs across the three scenarios.",
    "    Portfolio: annual Y1 roll-up across all seven D93 SKUs with blended margin & ROI.",
    "",
    "4.  Test Start Date (W1) on Inputs drives every week column's date.",
    "",
]
r = 6
for ln in lines:
    wc(rd,r,2,ln,font=Font(name="Calibri",size=11),align=top_wrap); r += 1

wc(rd,r,2,"COLOUR KEY",font=f_subhead); r += 1
for label,color,desc in [
    ("Yellow",     YELLOW,     "Editable input"),
    ("Light Blue", LIGHT_BLUE, "Summary KPI"),
    ("Grey",       GREY,       "Calculated"),
]:
    wc(rd,r,2,f"   {label}: {desc}",align=left,
       fill=PatternFill("solid",fgColor=color),font=Font(name="Calibri",size=11))
    r += 1


# =============================================================================
# INPUTS — slim
# =============================================================================
inp = wb.create_sheet("Inputs")
inp.sheet_view.showGridLines = False
inp.column_dimensions["A"].width = 2
inp.column_dimensions["B"].width = 36
inp.column_dimensions["C"].width = 16
inp.column_dimensions["D"].width = 70

inp.merge_cells("B2:D2")
wc(inp,2,2,"INPUTS — change any yellow cell",
   font=f_title,fill=fill_title,align=center)
inp.row_dimensions[2].height = 32


def input_row(r, label, value, note, fmt=None, is_formula=False):
    wc(inp,r,2,label,font=f_label,align=left)
    cell_value = f"={value}" if is_formula else value
    cell_font = f_calc if is_formula else f_input
    wc(inp,r,3,cell_value,font=cell_font,fill=fill_input,
       align=right,border=box,number_format=fmt)
    wc(inp,r,4,note,font=f_note,align=top_wrap)
    inp.row_dimensions[r].height = 22


# Section 1 — Currency dropdown + shared FX table
section_header(inp, 4, "1.  Display Currency  (master copy — each tab has its own)")
wc(inp,5,2,"Display Currency",font=f_label,align=left)
wc(inp,5,3,"GBP",font=f_input,fill=fill_input,align=center,border=box)
wc(inp,5,4,"Each display tab has its own dropdown; this one is just here for reference.",
   font=f_note,align=top_wrap)
inp.row_dimensions[5].height = 22

dv = DataValidation(type="list", formula1='"GBP,USD,CAD,EUR"', allow_blank=False)
inp.add_data_validation(dv); dv.add("C5")

# FX Table — header at row 7, column titles at row 12, rates at rows 13-16
# (widgets on display tabs reference $B$13:$D$16)
inp.merge_cells("B7:D7")
wc(inp,7,2,"FX Rate Table — edit rates (rate is per 1 GBP)",
   font=f_subhead,align=left)

wc(inp,12,2,"Code",  font=f_white_bold,fill=fill_section,align=center,border=box)
wc(inp,12,3,"Symbol",font=f_white_bold,fill=fill_section,align=center,border=box)
wc(inp,12,4,"Rate (× GBP)",font=f_white_bold,fill=fill_section,align=left,border=box)

for i,(code,sym,rate) in enumerate(CURRENCIES):
    r = 13 + i
    wc(inp,r,2,code,font=f_label,fill=fill_calc,align=center,border=box)
    wc(inp,r,3,sym, font=f_input,fill=fill_input,align=center,border=box)
    wc(inp,r,4,rate,font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_RATE)
    inp.row_dimensions[r].height = 18

# Section 2 — Product & Distribution
section_header(inp, 19, "2.  Product & Distribution")
input_row(20,"SKU","Bee Propolis Spray 30ml x 2","Costco UK market test SKU.")
input_row(21,"Test Start Date (W1)",DEFAULT_START_DATE,
          "Week-commencing date for W1.",fmt=FMT_DATE_L)
input_row(22,"Warehouses in test",20,"Number of Costco UK warehouses.",fmt=FMT_INT)

# Section 3 — Pricing
section_header(inp, 24, "3.  Pricing (GBP)")
input_row(25,"Selling Price (MSRP)",15.99,"Retail per pack.",fmt=FMT_GBP_HARD_D)
input_row(26,"Cost Price (landed)",7.29,"Brand cost delivered.",fmt=FMT_GBP_HARD_D)
input_row(27,"Gross Margin per unit","C25-C26","SP − Cost (auto).",
          fmt=FMT_GBP_HARD_D,is_formula=True)
input_row(28,"Gross Margin %","IFERROR((C25-C26)/C25,0)","GP ÷ SP (auto).",
          fmt=FMT_PCT,is_formula=True)

# Section 4 — Demand
section_header(inp, 30, "4.  Base Demand")
input_row(31,"Base Units / Week (all warehouses)",276,
          "Organic weekly sell-through across all test warehouses, no promo.",fmt=FMT_INT)

# Section 5 — Promo Economics (slim)
section_header(inp, 33, "5.  Promo Economics (GBP)")
input_row(34,"Demo Total £ / week (all warehouses)",7403,
          "All-warehouse demo cost in a demo week. ~£199/wh × 20 wh × 1.86 factor.",fmt=FMT_GBP_HARD)
input_row(35,"End Cap Total £ / week (all warehouses)",15810,
          "All-warehouse end cap cost in an end cap week. ~£850/wh × 20 wh × 1.86 ÷ 2-week cycle.",fmt=FMT_GBP_HARD)
input_row(36,"TPD discount % (of net)",0.25,"20% off MSRP ≈ 25% off net.",fmt=FMT_PCT)
input_row(37,"Markdown discount %",0.50,"Clearance pricing.",fmt=FMT_PCT)

# Named ranges
wb.defined_names["StartDate"] = DefinedName("StartDate", attr_text="Inputs!$C$21")
wb.defined_names["WHs"]       = DefinedName("WHs",       attr_text="Inputs!$C$22")
wb.defined_names["Period"]    = DefinedName("Period",    attr_text=f"{WEEKS}")  # constant
wb.defined_names["Price"]     = DefinedName("Price",     attr_text="Inputs!$C$25")
wb.defined_names["Cost"]      = DefinedName("Cost",      attr_text="Inputs!$C$26")
wb.defined_names["BaseUnits"] = DefinedName("BaseUnits", attr_text="Inputs!$C$31")
wb.defined_names["DemoCost"]  = DefinedName("DemoCost",  attr_text="Inputs!$C$34")
wb.defined_names["EndCapCost"]= DefinedName("EndCapCost",attr_text="Inputs!$C$35")
wb.defined_names["TPDPct"]    = DefinedName("TPDPct",    attr_text="Inputs!$C$36")
wb.defined_names["MdPct"]     = DefinedName("MdPct",     attr_text="Inputs!$C$37")

# Section 6 — Scenario activity & lift
section_header(inp, 39, "6.  Scenario activity (1 = on / 0 = off per week) and demand lift")

WEEK_START_COL = 6
WEEK_END_COL = WEEK_START_COL + WEEKS - 1
for col in range(WEEK_START_COL, WEEK_END_COL + 1):
    inp.column_dimensions[get_column_letter(col)].width = 7.5

scenario_grid_start = {}
current = 41
for sc in SCENARIOS:
    inp.merge_cells(start_row=current,start_column=2,end_row=current,end_column=WEEK_END_COL)
    wc(inp,current,2,sc["name"],font=f_white_bold,fill=fill_title,align=left)
    inp.row_dimensions[current].height = 22
    current += 1

    inp.merge_cells(start_row=current,start_column=2,end_row=current,end_column=WEEK_END_COL)
    wc(inp,current,2,f"Rationale: {sc['rationale']}",font=f_note,align=top_wrap)
    inp.row_dimensions[current].height = 24
    current += 1

    wc(inp,current,2,"Week",font=f_label,fill=fill_calc,align=left)
    for i in range(WEEKS):
        wc(inp,current,WEEK_START_COL+i,f"W{FIRST_WEEK_NUM+i}",
           font=f_white_bold,fill=PatternFill("solid",fgColor=BLUE),
           align=center,border=box)
    current += 1

    wc(inp,current,2,"Wk-commencing",font=f_label,fill=fill_calc,align=left)
    for i in range(WEEKS):
        wc(inp,current,WEEK_START_COL+i,f"=StartDate+7*{i}",
           font=Font(name="Calibri",size=9,color=DARK_GREY),
           fill=fill_calc,align=center,border=box,number_format=FMT_DATE)
    current += 1

    scenario_grid_start[sc["key"]] = current

    for promo in PROMO_TYPES:
        wc(inp,current,2,promo,font=f_label,align=left)
        active = set(sc["promo"].get(promo,[]))
        for i in range(WEEKS):
            wknum = FIRST_WEEK_NUM + i
            wc(inp,current,WEEK_START_COL+i,
               1 if wknum in active else 0,
               font=f_input,fill=fill_input,align=center,border=box,
               number_format="0;;;@")
        current += 1

    wc(inp,current,2,"Lift Multiplier (x base)",font=f_label,align=left)
    for i in range(WEEKS):
        wknum = FIRST_WEEK_NUM + i
        wc(inp,current,WEEK_START_COL+i,
           sc["lift_overrides"].get(wknum,1.0),
           font=f_input,fill=fill_input,align=center,border=box,
           number_format=FMT_MULT)
    current += 1
    current += 1

inp.freeze_panes = "F5"


# =============================================================================
# FORECAST — 6 P&L rows per scenario
# =============================================================================
fc = wb.create_sheet("Forecast")
fc.sheet_view.showGridLines = False
fc.column_dimensions["A"].width = 2
fc.column_dimensions["B"].width = 28
for col in range(WEEK_START_COL,WEEK_END_COL+1):
    fc.column_dimensions[get_column_letter(col)].width = 11
TOTAL_COL = WEEK_END_COL + 1
fc.column_dimensions[get_column_letter(TOTAL_COL)].width = 14

fc.merge_cells(start_row=2,start_column=2,end_row=2,end_column=TOTAL_COL)
wc(fc,2,2,'="FORECAST — Weekly P&L  (Display: "&$Z$3&" "&$C$3&")"',
   font=f_title,fill=fill_title,align=center)
fc.row_dimensions[2].height = 32

FC_FX, FC_SYM = add_currency_widget(fc, 3, default="GBP")

PNL_ROWS = [
    ("POS Units",                                  "units"),
    ('="Revenue ("&$Z$3&")"',                      "rev"),
    ('="Raw COGS ("&$Z$3&")"',                     "cogs"),
    ('="Gross Profit ("&$Z$3&")"',                 "gp"),
    ('="Promo Spend ("&$Z$3&")"',                  "promo"),
    ('="Net Profit ("&$Z$3&")"',                   "net"),
]
PROMO_ROW_OFFSET = {p:i for i,p in enumerate(PROMO_TYPES)}
LIFT_OFFSET = len(PROMO_TYPES)

fc_row = 5
scenario_summary_rows = {}

for sc in SCENARIOS:
    fc.merge_cells(start_row=fc_row,start_column=2,end_row=fc_row,end_column=TOTAL_COL)
    wc(fc,fc_row,2,sc["name"],font=f_white_bold,fill=fill_title,align=left)
    fc.row_dimensions[fc_row].height = 22
    fc_row += 1

    wc(fc,fc_row,2,"Week",font=f_label,fill=fill_calc,align=left)
    for i in range(WEEKS):
        wc(fc,fc_row,WEEK_START_COL+i,f"W{FIRST_WEEK_NUM+i}",
           font=f_white_bold,fill=PatternFill("solid",fgColor=BLUE),
           align=center,border=box)
    wc(fc,fc_row,TOTAL_COL,"TOTAL",
       font=f_white_bold,fill=fill_title,align=center,border=box)
    fc_row += 1

    wc(fc,fc_row,2,"Wk-commencing",font=f_label,fill=fill_calc,align=left)
    for i in range(WEEKS):
        wc(fc,fc_row,WEEK_START_COL+i,f"=StartDate+7*{i}",
           font=Font(name="Calibri",size=9,color=DARK_GREY),
           fill=fill_calc,align=center,border=box,number_format=FMT_DATE)
    wc(fc,fc_row,TOTAL_COL,"",fill=fill_calc,border=box)
    fc_row += 1

    scenario_pnl_rows = {}
    grid_top = scenario_grid_start[sc["key"]]
    promo_rows_on_inp = {p:grid_top+PROMO_ROW_OFFSET[p] for p in PROMO_TYPES}
    lift_row_on_inp = grid_top + LIFT_OFFSET

    for label_formula, key in PNL_ROWS:
        scenario_pnl_rows[key] = fc_row
        bold = key in ("rev","cogs","promo","gp","net","units")
        navy_bold = key in ("gp","net")
        lbl_font = (Font(name="Calibri",size=11,bold=True,color=NAVY)
                    if navy_bold else (f_label if bold else f_calc))
        wc(fc,fc_row,2,label_formula,font=lbl_font,align=left)

        for i in range(WEEKS):
            wcol = get_column_letter(WEEK_START_COL+i)
            if key == "units":
                formula = f"=ROUND(BaseUnits*Inputs!{wcol}{lift_row_on_inp},0)"
            elif key == "rev":
                # rev = units × price × FX
                formula = (f"={wcol}{scenario_pnl_rows['units']}*Price*{FC_FX}")
            elif key == "cogs":
                formula = (f"={wcol}{scenario_pnl_rows['units']}*Cost*{FC_FX}")
            elif key == "gp":
                formula = f"={wcol}{scenario_pnl_rows['rev']}-{wcol}{scenario_pnl_rows['cogs']}"
            elif key == "promo":
                # Promo = Demo + EndCap + TPD + Markdown
                demo  = f"Inputs!{wcol}{promo_rows_on_inp['Demo']}*DemoCost"
                endcap= f"Inputs!{wcol}{promo_rows_on_inp['End Cap']}*EndCapCost"
                tpd   = (f"Inputs!{wcol}{promo_rows_on_inp['TPD']}"
                         f"*{wcol}{scenario_pnl_rows['units']}*Price*TPDPct")
                md    = (f"Inputs!{wcol}{promo_rows_on_inp['Markdown']}"
                         f"*{wcol}{scenario_pnl_rows['units']}*Price*MdPct")
                formula = f"=({demo}+{endcap}+{tpd}+{md})*{FC_FX}"
            elif key == "net":
                formula = f"={wcol}{scenario_pnl_rows['gp']}-{wcol}{scenario_pnl_rows['promo']}"
            else:
                formula = ""
            num_fmt = FMT_INT if key == "units" else FMT_MONEY
            wc(fc,fc_row,WEEK_START_COL+i,formula,
               font=f_calc,fill=fill_calc,align=right,border=box,number_format=num_fmt)

        first_w = get_column_letter(WEEK_START_COL)
        last_w = get_column_letter(WEEK_END_COL)
        tot_formula = f"=SUM({first_w}{fc_row}:{last_w}{fc_row})"
        wc(fc,fc_row,TOTAL_COL,tot_formula,
           font=Font(name="Calibri",size=11,bold=True,color=NAVY),
           fill=fill_kpi,align=right,border=box,
           number_format=FMT_INT if key == "units" else FMT_MONEY)
        fc_row += 1

    scenario_summary_rows[sc["key"]] = scenario_pnl_rows
    fc_row += 2

fc.freeze_panes = "C5"


# =============================================================================
# SUMMARY — 7 KPIs
# =============================================================================
sm = wb.create_sheet("Summary")
sm.sheet_view.showGridLines = False
sm.column_dimensions["A"].width = 2
sm.column_dimensions["B"].width = 32
for i in range(len(SCENARIOS)):
    sm.column_dimensions[get_column_letter(3+i)].width = 22

end_col = 2 + len(SCENARIOS)
sm.merge_cells(start_row=2,start_column=2,end_row=2,end_column=end_col)
wc(sm,2,2,
   '="ROI SUMMARY  (Display: "&$Z$3&" "&$C$3&")"',
   font=f_title,fill=fill_title,align=center)
sm.row_dimensions[2].height = 32

SM_FX, SM_SYM = add_currency_widget(sm, 3, default="GBP")
FC_FX_FROM_SM = f"Forecast!{FC_FX}"

sm.merge_cells(start_row=4,start_column=2,end_row=4,end_column=end_col)
wc(sm,4,2,
   '="Test window: W1 ("&TEXT(StartDate,"dd-mmm-yyyy")&")  →  W"&Period&" ("&TEXT(StartDate+7*(Period-1),"dd-mmm-yyyy")&")"',
   font=Font(name="Calibri",size=11,italic=True,color=DARK_GREY),
   fill=fill_calc,align=center)

hr = 6
wc(sm,hr,2,"Metric",font=f_white_bold,fill=fill_section,align=left)
for i,sc in enumerate(SCENARIOS):
    wc(sm,hr,3+i,sc["key"],font=f_white_bold,fill=fill_section,align=center)
sm.row_dimensions[hr].height = 24

TC = get_column_letter(TOTAL_COL)
KPIS = [
    ('Units Sold',                             "units",  FMT_INT,    False, False),
    ('="Revenue ("&$Z$3&")"',                  "rev",    FMT_MONEY,  False, True),
    ('="Gross Profit ("&$Z$3&")"',             "gp",     FMT_MONEY,  True,  True),
    ('Gross Margin %',                         "gp_pct", FMT_PCT,    True,  False),
    ('="Promo Spend ("&$Z$3&")"',              "promo",  FMT_MONEY,  False, True),
    ('="Net Profit ("&$Z$3&")"',               "net",    FMT_MONEY,  True,  True),
    ('ROI on Promo Spend',                     "roi",    FMT_PCT,    True,  False),
]

sr = hr + 1
kpi_row_idx = {}
for label,key,fmt,bold,is_money in KPIS:
    kpi_row_idx[key] = sr
    wc(sm,sr,2,label,
       font=Font(name="Calibri",size=11,bold=bold,color=NAVY if bold else "000000"),
       fill=fill_kpi if bold else None,align=left,border=box)
    fx_adj = f"*({SM_FX}/{FC_FX_FROM_SM})"
    for i,sc in enumerate(SCENARIOS):
        rows = scenario_summary_rows[sc["key"]]
        if key in rows:
            base = f"Forecast!{TC}{rows[key]}"
            formula = f"={base}{fx_adj}" if is_money else f"={base}"
        elif key == "gp_pct":
            formula = f"=IFERROR(Forecast!{TC}{rows['gp']}/Forecast!{TC}{rows['rev']},0)"
        elif key == "roi":
            formula = f"=IFERROR(Forecast!{TC}{rows['net']}/Forecast!{TC}{rows['promo']},0)"
        else:
            formula = ""
        wc(sm,sr,3+i,formula,
           font=f_kpi_val if bold else f_calc,
           fill=fill_kpi if bold else None,
           align=right,border=box,number_format=fmt)
    sm.row_dimensions[sr].height = 26 if bold else 22
    sr += 1

last_col_letter = get_column_letter(2+len(SCENARIOS))
for k in ("net","roi"):
    r = kpi_row_idx[k]
    rng = f"C{r}:{last_col_letter}{r}"
    sm.conditional_formatting.add(rng,
        CellIsRule(operator="lessThan",formula=["0"],
                   fill=PatternFill("solid",fgColor=RED),
                   font=Font(bold=True,color="9C0006")))
    sm.conditional_formatting.add(rng,
        CellIsRule(operator="greaterThanOrEqual",formula=["0"],
                   fill=PatternFill("solid",fgColor=GREEN),
                   font=Font(bold=True,color="006100")))

sr += 2
sm.merge_cells(start_row=sr,start_column=2,end_row=sr,end_column=end_col)
wc(sm,sr,2,"Why each scenario?",font=f_section,fill=fill_section,align=left)
sm.row_dimensions[sr].height = 22
sr += 1
for sc in SCENARIOS:
    wc(sm,sr,2,sc["key"],
       font=Font(name="Calibri",size=11,bold=True,color=NAVY),align=top_wrap)
    sm.merge_cells(start_row=sr,start_column=3,end_row=sr,end_column=end_col)
    wc(sm,sr,3,sc["rationale"],font=f_note,align=top_wrap)
    sm.row_dimensions[sr].height = 28
    sr += 1

sm.freeze_panes = "C7"


# =============================================================================
# PORTFOLIO — 11 columns
# =============================================================================
pf = wb.create_sheet("Portfolio")
pf.sheet_view.showGridLines = False
pf.column_dimensions["A"].width = 2
widths = {"B":4,"C":34,"D":10,"E":10,"F":8,"G":11,"H":13,"I":12,"J":13,"K":8,"L":34}
for col,w in widths.items():
    pf.column_dimensions[col].width = w

pf.merge_cells("B2:L2")
wc(pf,2,2,
   '="PORTFOLIO — D93 Pitch SKUs · Y1 Roll-Up  (Display: "&$Z$3&" "&$C$3&")"',
   font=f_title,fill=fill_title,align=center)
pf.row_dimensions[2].height = 32

PF_FX, PF_SYM = add_currency_widget(pf, 3, default="GBP")

pf.merge_cells("B4:L4")
wc(pf,4,2,
   "Bee Propolis pre-filled. Fill SP / Cost / Y1 Units / Promo Spend for SKUs 2-7 (yellow cells). "
   "Bottom row blends across SKUs.",
   font=Font(name="Calibri",size=11,italic=True,color=DARK_GREY),align=top_wrap)
pf.row_dimensions[4].height = 24

HDR_ROW = 6
hdrs = [
    ("B","#"),("C","SKU"),
    ('D','="SP ("&$Z$3&")"'),
    ('E','="Cost ("&$Z$3&")"'),
    ("F","GM %"),("G","Y1 Units"),
    ('H','="Revenue ("&$Z$3&")"'),
    ('I','="Promo ("&$Z$3&")"'),
    ('J','="Net Profit ("&$Z$3&")"'),
    ("K","ROI"),
    ("L","Rationale"),
]
for col_letter,txt in hdrs:
    al = left if col_letter in ("C","L") else center
    wc(pf,HDR_ROW,column_index_from_string(col_letter),txt,
       font=f_white_bold,fill=fill_section,align=al,border=box)

sku_start_row = HDR_ROW + 1
for idx,sku in enumerate(PORTFOLIO_SKUS):
    r = sku_start_row + idx
    wc(pf,r,2,sku["rank"],font=f_label,fill=fill_calc,align=center,border=box,number_format=FMT_INT)
    wc(pf,r,3,sku["name"],font=f_label,fill=fill_calc,align=left,border=box)
    wc(pf,r,4,sku["sp"],  font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_MONEY_D)
    wc(pf,r,5,sku["cost"],font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_MONEY_D)
    wc(pf,r,6,f"=IFERROR((D{r}-E{r})/D{r},0)",font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_PCT)
    wc(pf,r,7,sku["units"],font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_INT)
    wc(pf,r,8,f"=IFERROR(G{r}*D{r},0)",font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_MONEY)
    wc(pf,r,9,sku["promo"],font=f_input,fill=fill_input,align=right,border=box,number_format=FMT_MONEY)
    # Net = Revenue × GM% − Promo  (i.e. GP − Promo)
    wc(pf,r,10,f"=IFERROR((G{r}*(D{r}-E{r}))-I{r},0)",
       font=Font(name="Calibri",size=11,bold=True,color=NAVY),
       fill=fill_kpi,align=right,border=box,number_format=FMT_MONEY)
    wc(pf,r,11,f"=IFERROR(J{r}/I{r},0)",
       font=f_calc,fill=fill_calc,align=right,border=box,number_format=FMT_PCT)
    wc(pf,r,12,sku["rationale"],font=f_input,fill=fill_input,align=left,border=box)
    pf.row_dimensions[r].height = 22

last_sku_row = sku_start_row + len(PORTFOLIO_SKUS) - 1
tr = last_sku_row + 1

for c in range(2,6):
    wc(pf,tr,c,"",fill=fill_title,border=box)
wc(pf,tr,3,"PORTFOLIO TOTAL / BLENDED",
   font=f_white_bold,fill=fill_title,align=left,border=box)
wc(pf,tr,6,f"=IFERROR(SUM(H{sku_start_row}:H{last_sku_row})*0+ "
            f"(SUMPRODUCT(G{sku_start_row}:G{last_sku_row},(D{sku_start_row}:D{last_sku_row}-E{sku_start_row}:E{last_sku_row})))"
            f"/SUM(H{sku_start_row}:H{last_sku_row}),0)",
   font=f_white_bold,fill=fill_title,align=right,border=box,number_format=FMT_PCT)
for col_letter,fmt in [("G",FMT_INT),("H",FMT_MONEY),("I",FMT_MONEY),("J",FMT_MONEY)]:
    wc(pf,tr,column_index_from_string(col_letter),
       f"=SUM({col_letter}{sku_start_row}:{col_letter}{last_sku_row})",
       font=f_white_bold,fill=fill_title,align=right,border=box,number_format=fmt)
wc(pf,tr,11,f"=IFERROR(J{tr}/I{tr},0)",
   font=f_white_bold,fill=fill_title,align=right,border=box,number_format=FMT_PCT)
wc(pf,tr,12,"Blended across SKUs",font=f_white_bold,fill=fill_title,align=left,border=box)
pf.row_dimensions[tr].height = 26

pf.freeze_panes = "B7"


out = "/home/user/my-first-project/Costco_UK_D93_Portfolio_ROI_v8.xlsx"
wb.save(out)
print(f"Saved: {out}")
