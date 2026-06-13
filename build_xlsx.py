"""
Builds revenue-forecast.xlsx — a revenue & PROFIT planning workbook driven by
live Excel formulas. Mirrors engine.js / revenue-forecast.html.

Sheets: Model · Scenarios · Projection · Profit Grid · Monthly Plan · Channel View
Yellow cells are inputs; grey/blue cells are formulas. Run: python3 build_xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter as col

# ----- assumptions / placeholders (overwrite in Excel) -----
TARGET, UPC, DAYS, FIXED, GROWTH, PYEARS = 2_500_000, 12, 260, 600_000, 0.15, 3
# name, gross price, mix, cogs, discount%, selling%
CHAN = [
    ("Direct-to-Consumer (DTC)",  300, 20, 90, 0.05, 0.18),
    ("On-Premise / Food Service", 216, 15, 90, 0.10, 0.05),
    ("Off-Premise Retail",        180, 30, 90, 0.12, 0.04),
    ("Wholesale / Distributor",   120, 30, 90, 0.00, 0.03),
    ("Online Marketplace",        150,  5, 90, 0.08, 0.15),
]
SEASON = [0.7, 0.7, 0.9, 1.0, 1.1, 1.2, 1.2, 1.1, 1.0, 1.0, 1.3, 1.8]
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
SCEN = [("Conservative", 2_000_000), ("Base", 2_500_000), ("Aggressive", 3_200_000)]
VOLF = [0.5, 0.75, 1.0, 1.25, 1.5, 1.75]
PRICEF = [0.6, 0.8, 1.0, 1.2, 1.4]

# ----- styles -----
NAVY,BLUE,LIGHT,YELL,GREY,GREEN,WHITE = "16243D","2E5496","DBE4F3","FFF8E1","F2F4F8","E2EFDA","FFFFFF"
CUR,CUR2,NUM,PCT = '"$"#,##0','"$"#,##0.00','#,##0','0.0%'
thin = Side(style="thin", color="C8D0DC"); BD = Border(thin,thin,thin,thin)
F_TITLE=Font(bold=True,size=16,color=WHITE); F_SUB=Font(italic=True,size=9,color="8A93A6")
F_HDR=Font(bold=True,size=10.5,color=WHITE); F_SECT=Font(bold=True,size=12,color=NAVY)
F_LBL=Font(bold=True,color="3A4456"); F_IN=Font(bold=True,color="16243D"); F_CALC=Font(color="3A4456")
F_WHITEB=Font(bold=True,color=WHITE)
fill=lambda c:PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center"); WRAP=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEFTC=Alignment(horizontal="left",vertical="center")

def put(ws,ref,val,font=None,bg=None,fmt=None,al=None,bd=False):
    c=ws[ref]; c.value=val
    if font:c.font=font
    if bg:c.fill=fill(bg)
    if fmt:c.number_format=fmt
    if al:c.alignment=al
    if bd:c.border=BD
    return c
def title(ws,rng,text):
    ws.merge_cells(rng); a=rng.split(":")[0]
    put(ws,a,text,F_TITLE,NAVY,al=LEFTC)
    ws.row_dimensions[int(a[1:])].height=26

wb=Workbook(); M=wb.active; M.title="Model"; M.sheet_view.showGridLines=False
for c,w in {"A":30,"B":12,"C":11,"D":11,"E":11,"F":11,"G":11,"H":12,"I":10,"J":11,"K":13,"L":13,"M":13,"N":13,"O":13}.items():
    M.column_dimensions[c].width=w

title(M,"A1:O1","Revenue & Profit Forecast  —  Cases-Sold Model")
M.merge_cells("A2:O2"); put(M,"A2","Yellow = inputs. Everything else is live formulas. Profit layers in discounts, COGS, cost-to-serve & fixed costs.",F_SUB)

# assumptions
put(M,"A4","ASSUMPTIONS",F_SECT)
asm=[("Annual gross revenue target",TARGET,CUR),("Units per case",UPC,NUM),("Selling days / year",DAYS,NUM),
     ("Annual fixed costs",FIXED,CUR),("YoY growth (projection)",GROWTH,PCT),("Projection years",PYEARS,NUM)]
for i,(lab,val,fmt) in enumerate(asm):
    r=5+i; put(M,f"A{r}",lab,F_LBL); put(M,f"B{r}",val,F_IN,YELL,fmt,bd=True)

# channels
HR=12; F0=13; F1=F0+len(CHAN)-1; TOT=F1+1
put(M,f"A{HR-1}","CHANNELS — pricing & unit economics",F_SECT)
hdr=["Channel","Gross/case","Mix","COGS/case","Disc %","Sell %","Net/case","Contrib/case","Margin","Cases","Gross $","Net $","COGS $","Sell $","Contrib $"]
for j,h in enumerate(hdr):
    put(M,f"{col(j+1)}{HR}",h,F_HDR,BLUE if j>=6 else NAVY,al=WRAP,bd=True)
M.row_dimensions[HR].height=28
for k,(nm,gp,mx,cg,ds,se) in enumerate(CHAN):
    r=F0+k
    put(M,f"A{r}",nm,bd=True)
    put(M,f"B{r}",gp,F_IN,YELL,CUR,bd=True); put(M,f"C{r}",mx,F_IN,YELL,NUM,bd=True)
    put(M,f"D{r}",cg,F_IN,YELL,CUR,bd=True); put(M,f"E{r}",ds,F_IN,YELL,PCT,bd=True); put(M,f"F{r}",se,F_IN,YELL,PCT,bd=True)
    put(M,f"G{r}",f"=B{r}*(1-E{r})",F_CALC,GREY,CUR2,bd=True)
    put(M,f"H{r}",f"=G{r}*(1-F{r})-D{r}",F_CALC,GREY,CUR2,bd=True)
    put(M,f"I{r}",f"=IF(G{r}=0,0,H{r}/G{r})",F_CALC,GREY,PCT,bd=True)
    put(M,f"J{r}",f"=$B$37*(C{r}/SUM($C${F0}:$C${F1}))",F_CALC,GREY,NUM,bd=True)
    put(M,f"K{r}",f"=J{r}*B{r}",F_CALC,GREY,CUR,bd=True)
    put(M,f"L{r}",f"=J{r}*G{r}",F_CALC,GREY,CUR,bd=True)
    put(M,f"M{r}",f"=J{r}*D{r}",F_CALC,GREY,CUR,bd=True)
    put(M,f"N{r}",f"=J{r}*G{r}*F{r}",F_CALC,GREY,CUR,bd=True)
    put(M,f"O{r}",f"=J{r}*H{r}",F_CALC,GREY,CUR,bd=True)
rng=lambda c:f"{c}{F0}:{c}{F1}"
put(M,f"A{TOT}","Blended / Total",F_WHITEB,NAVY,bd=True)
for c,f in {"B":f"=SUMPRODUCT(B{F0}:B{F1},C{F0}:C{F1})/SUM({rng('C')})","C":f"=SUM({rng('C')})",
            "D":f"=SUMPRODUCT(D{F0}:D{F1},C{F0}:C{F1})/SUM({rng('C')})","E":f"=SUMPRODUCT(E{F0}:E{F1},C{F0}:C{F1})/SUM({rng('C')})",
            "F":f"=SUMPRODUCT(F{F0}:F{F1},C{F0}:C{F1})/SUM({rng('C')})","G":f"=SUMPRODUCT(G{F0}:G{F1},C{F0}:C{F1})/SUM({rng('C')})",
            "H":f"=SUMPRODUCT(H{F0}:H{F1},C{F0}:C{F1})/SUM({rng('C')})","I":f"=IF(B{TOT}=0,0,H{TOT}/G{TOT})",
            "J":f"=SUM({rng('J')})","K":f"=SUM({rng('K')})","L":f"=SUM({rng('L')})","M":f"=SUM({rng('M')})",
            "N":f"=SUM({rng('N')})","O":f"=SUM({rng('O')})"}.items():
    fmt=PCT if c in("E","F","I") else CUR2 if c in("B","D","G","H") else NUM if c=="C" or c=="J" else CUR
    put(M,f"{c}{TOT}",f,F_WHITEB,NAVY,fmt,bd=True)

# profit cascade
put(M,"A21","PROFIT CASCADE (at target)",F_SECT)
casc=[("Gross revenue",f"=SUM({rng('K')})",CUR),("Trade discounts",f"=B24-B22",CUR),("Net revenue",f"=SUM({rng('L')})",CUR),
      ("COGS",f"=-SUM({rng('M')})",CUR),("Selling / cost-to-serve",f"=-SUM({rng('N')})",CUR),("Contribution",f"=SUM({rng('O')})",CUR),
      ("Fixed costs",f"=-B8",CUR),("Operating profit",f"=B27+B28",CUR),("Contribution margin",f"=B27/B24",PCT),("Operating margin",f"=B29/B24",PCT)]
for i,(lab,f,fmt) in enumerate(casc):
    r=22+i; put(M,f"A{r}",lab,F_LBL); put(M,f"B{r}",f,F_CALC,GREY,fmt,bd=True)
put(M,"B29",f"=B27+B28",F_WHITEB,GREEN,CUR,bd=True)  # operating profit highlighted

# key metrics (below the cascade)
put(M,"A33","KEY METRICS",F_SECT)
metrics=[("Blended gross / case",f"=SUMPRODUCT(B{F0}:B{F1},C{F0}:C{F1})/SUM({rng('C')})",CUR2),
         ("Blended net / case",f"=SUMPRODUCT(G{F0}:G{F1},C{F0}:C{F1})/SUM({rng('C')})",CUR2),
         ("Blended contribution / case",f"=SUMPRODUCT(H{F0}:H{F1},C{F0}:C{F1})/SUM({rng('C')})",CUR2),
         ("TOTAL cases to hit target","=B5/B34",NUM),
         ("Cases / month","=B37/12",NUM),("Cases / week","=B37/52",NUM),("Cases / selling-day","=B37/B7",NUM),
         ("Price / unit","=B34/B6",CUR2),("Break-even cases","=B8/B36",NUM),("Break-even revenue","=B42*B34",CUR),
         ("Blended discount %",f"=SUMPRODUCT(E{F0}:E{F1},C{F0}:C{F1})/SUM({rng('C')})",PCT),
         ("Blended selling %",f"=SUMPRODUCT(F{F0}:F{F1},C{F0}:C{F1})/SUM({rng('C')})",PCT),
         ("Blended COGS / case",f"=SUMPRODUCT(D{F0}:D{F1},C{F0}:C{F1})/SUM({rng('C')})",CUR2)]
# place metrics in rows 34.. (B col already used by cascade rows 22-31; metrics use B34..B46 which are free)
for i,(lab,f,fmt) in enumerate(metrics):
    r=34+i; put(M,f"A{r}",lab,F_LBL); put(M,f"B{r}",f,F_CALC,LIGHT,fmt,bd=True)
put(M,"B37","=B5/B34",F_WHITEB,BLUE,NUM,bd=True)  # total cases highlighted
M.freeze_panes="A3"

# ---------------- Scenarios ----------------
S=wb.create_sheet("Scenarios"); S.sheet_view.showGridLines=False
for c,w in {"A":18,"B":15,"C":13,"D":15,"E":15,"F":15,"G":12}.items(): S.column_dimensions[c].width=w
title(S,"A1:G1","Scenarios  —  same economics, three goals")
for j,h in enumerate(["Scenario","Revenue target","Cases","Net revenue","Contribution","Operating profit","Op margin"]):
    put(S,f"{col(j+1)}3",h,F_HDR,BLUE,al=WRAP,bd=True)
S.row_dimensions[3].height=26
for i,(lab,tg) in enumerate(SCEN):
    r=4+i; put(S,f"A{r}",lab,F_LBL,bd=True); put(S,f"B{r}",tg,F_IN,YELL,CUR,bd=True)
    put(S,f"C{r}","=B%d/Model!$B$34"%r,F_CALC,GREY,NUM,bd=True)
    put(S,f"D{r}","=C%d*Model!$B$35"%r,F_CALC,GREY,CUR,bd=True)
    put(S,f"E{r}","=C%d*Model!$B$36"%r,F_CALC,GREY,CUR,bd=True)
    put(S,f"F{r}","=E%d-Model!$B$8"%r,F_CALC,GREEN,CUR,bd=True)
    put(S,f"G{r}","=IF(D%d=0,0,F%d/D%d)"%(r,r,r),F_CALC,GREY,PCT,bd=True)

# ---------------- Projection ----------------
P=wb.create_sheet("Projection"); P.sheet_view.showGridLines=False
for c,w in {"A":10,"B":16,"C":13,"D":16,"E":16}.items(): P.column_dimensions[c].width=w
title(P,"A1:E1","Multi-year projection  (constant price & cost structure)")
for j,h in enumerate(["Year","Revenue","Cases","Contribution","Operating profit"]):
    put(P,f"{col(j+1)}3",h,F_HDR,BLUE,al=CEN,bd=True)
for y in range(PYEARS):
    r=4+y; put(P,f"A{r}",f"Y{y+1}",F_LBL,bd=True)
    put(P,f"B{r}",f"=Model!$B$5*(1+Model!$B$9)^{y}",F_CALC,GREY,CUR,bd=True)
    put(P,f"C{r}",f"=B{r}/Model!$B$34",F_CALC,GREY,NUM,bd=True)
    put(P,f"D{r}",f"=C{r}*Model!$B$36",F_CALC,GREY,CUR,bd=True)
    put(P,f"E{r}",f"=D{r}-Model!$B$8",F_CALC,GREEN,CUR,bd=True)

# ---------------- Profit Grid ----------------
G=wb.create_sheet("Profit Grid"); G.sheet_view.showGridLines=False
G.column_dimensions["A"].width=16
for j in range(len(VOLF)): G.column_dimensions[col(2+j)].width=13
title(G,f"A1:{col(1+len(VOLF))}1","Profit grid  —  price/case × cases → operating profit")
put(G,"A2","Green = profit, red = loss. Uses blended discount/selling/COGS.",F_SUB)
put(G,"A4","Price ╲ Cases",F_HDR,NAVY,al=WRAP,bd=True)
for j,vf in enumerate(VOLF):
    put(G,f"{col(2+j)}4",f"=ROUND(Model!$B$37*{vf}/100,0)*100",F_HDR,BLUE,NUM,al=CEN,bd=True)
for i,pf in enumerate(PRICEF):
    r=5+i; put(G,f"A{r}",f"=ROUND(Model!$B$34*{pf}/5,0)*5",F_LBL,LIGHT,CUR,bd=True)
    for j in range(len(VOLF)):
        cc=col(2+j)
        # op = vol*(price*(1-disc)*(1-sell)-cogs) - fixed
        f=(f"={cc}$4*($A{r}*(1-Model!$B$44)*(1-Model!$B$45)-Model!$B$46)-Model!$B$8")
        put(G,f"{cc}{r}",f,F_CALC,fmt=CUR,bd=True)
last=f"{col(1+len(VOLF))}{4+len(PRICEF)}"
G.conditional_formatting.add(f"B5:{last}",
    ColorScaleRule(start_type='min',start_color='E8A0A0',mid_type='num',mid_value=0,mid_color='FFFFFF',end_type='max',end_color='93C47D'))

# ---------------- Monthly Plan ----------------
MO=wb.create_sheet("Monthly Plan"); MO.sheet_view.showGridLines=False
for c,w in {"A":8,"B":13,"C":10,"D":12,"E":15,"F":15,"G":14,"H":15}.items(): MO.column_dimensions[c].width=w
title(MO,"A1:H1","Monthly plan  —  seasonality spread")
put(MO,"A2","Weights are inputs (relative). Edit to model peaks & ramp-ups.",F_SUB)
for j,h in enumerate(["Month","Weight","Mix %","Cases","Gross rev","Contribution","Cum. cases","Cum. rev"]):
    put(MO,f"{col(j+1)}4",h,F_HDR,BLUE,al=WRAP,bd=True)
MO.row_dimensions[4].height=26
for i,mo in enumerate(MONTHS):
    r=5+i; put(MO,f"A{r}",mo,F_LBL,bd=True); put(MO,f"B{r}",SEASON[i],F_IN,YELL,"0.0",bd=True)
    put(MO,f"C{r}",f"=B{r}/SUM($B$5:$B$16)",F_CALC,GREY,PCT,bd=True)
    put(MO,f"D{r}",f"=Model!$B$37*C{r}",F_CALC,GREY,NUM,bd=True)
    put(MO,f"E{r}",f"=Model!$B$22*C{r}",F_CALC,GREY,CUR,bd=True)
    put(MO,f"F{r}",f"=Model!$B$27*C{r}",F_CALC,GREY,CUR,bd=True)
    put(MO,f"G{r}",f"=SUM($D$5:D{r})",F_CALC,GREY,NUM,bd=True)
    put(MO,f"H{r}",f"=SUM($E$5:E{r})",F_CALC,GREY,CUR,bd=True)
put(MO,"A17","TOTAL",F_WHITEB,NAVY,bd=True)
for c,f in {"B":"=SUM(B5:B16)","C":"=SUM(C5:C16)","D":"=SUM(D5:D16)","E":"=SUM(E5:E16)","F":"=SUM(F5:F16)"}.items():
    fmt="0.0" if c=="B" else PCT if c=="C" else NUM if c=="D" else CUR
    put(MO,f"{c}17",f,F_WHITEB,NAVY,fmt,bd=True)

# ---------------- Channel View (pure-play + revenue grid) ----------------
CV=wb.create_sheet("Channel View"); CV.sheet_view.showGridLines=False
CV.column_dimensions["A"].width=28
for c in "BCDEFGHI": CV.column_dimensions[c].width=13
title(CV,"A1:G1","Channel view  —  one channel at a time")
put(CV,"A3","A) Cases to hit target if 100% sold through one channel",F_SECT)
for j,h in enumerate(["Channel","Gross/case","Cases to target","Units","Contribution"]):
    put(CV,f"{col(j+1)}4",h,F_HDR,BLUE,al=WRAP,bd=True)
for k in range(len(CHAN)):
    r=5+k; m=F0+k
    put(CV,f"A{r}",f"=Model!A{m}",bd=True)
    put(CV,f"B{r}",f"=Model!B{m}",F_CALC,GREY,CUR,bd=True)
    put(CV,f"C{r}",f"=Model!$B$5/B{r}",F_CALC,GREY,NUM,bd=True)
    put(CV,f"D{r}",f"=C{r}*Model!$B$6",F_CALC,GREY,NUM,bd=True)
    put(CV,f"E{r}",f"=C{r}*Model!H{m}",F_CALC,GREY,CUR,bd=True)
s2=5+len(CHAN)+2
put(CV,f"A{s2-1}","B) Gross revenue at a given number of cases",F_SECT)
put(CV,f"A{s2}","Channel",F_HDR,BLUE,al=CEN,bd=True); put(CV,f"B{s2}","Gross/case",F_HDR,BLUE,al=WRAP,bd=True)
VP=[5000,10000,15000,20000,25000,30000]
for j,v in enumerate(VP): put(CV,f"{col(3+j)}{s2}",v,F_HDR,BLUE,NUM,al=CEN,bd=True)
for k in range(len(CHAN)):
    r=s2+1+k; m=F0+k
    put(CV,f"A{r}",f"=Model!A{m}",bd=True); put(CV,f"B{r}",f"=Model!B{m}",F_CALC,GREY,CUR,bd=True)
    for j in range(len(VP)):
        cc=col(3+j); put(CV,f"{cc}{r}",f"={cc}${s2}*$B{r}",F_CALC,GREEN if (r+j)%2 else GREY,CUR,bd=True)

wb.calculation.fullCalcOnLoad = True   # force Excel/LibreOffice to recalc on open
wb.save("revenue-forecast.xlsx")
print("wrote revenue-forecast.xlsx — sheets:", wb.sheetnames)
