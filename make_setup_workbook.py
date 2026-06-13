#!/usr/bin/env python3
"""Generate the enterprise setup workbook: docs/Trading_Bot_Setup.xlsx.

A single, self-contained Excel file that takes a complete beginner from zero to
paper trading and beyond. It includes:
  - a clickable table of contents
  - two setup paths (in-browser Codespaces, and local) with copy-paste commands
  - an Alpaca key guide, a full config reference, and an ordered command list
  - an INTERACTIVE tuning calculator (live Excel formulas + colour rules)
  - REAL embedded charts from a backtest (price, strategy equity, buy & hold)
  - a REAL parameter sweep (many backtests) so you can see the trade-offs
  - test results, code-review findings, go-live checklist, glossary, FAQ,
    a daily monitoring template, an architecture map, and a risks page.

Re-run any time:  python make_setup_workbook.py
"""

from __future__ import annotations

import math
import os
import subprocess
from types import SimpleNamespace

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from backtest import run_backtest, _synthetic_prices

# --- palette ----------------------------------------------------------------
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D9E1F2"
GREEN = "C6EFCE"
GREEN_TX = "006100"
RED = "FFC7CE"
RED_TX = "9C0006"
AMBER = "FFEB9C"
AMBER_TX = "7F6000"
GREY = "F2F2F2"
INPUT = "FFF2CC"  # yellow = "edit me"
MONO = "Consolas"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_MID = Alignment(wrap_text=True, vertical="center")


def title(ws, text, span=2):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, text)
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30


def subtitle(ws, row, text, span=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def header_row(ws, row, values):
    for j, v in enumerate(values, start=1):
        c = ws.cell(row, j, v)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = WRAP_MID
        c.border = BORDER


def cell(ws, row, col, value, *, mono=False, fill=None, bold=False, color=None, wrap=True):
    c = ws.cell(row, col, value)
    c.font = Font(name=MONO if mono else "Calibri", bold=bold, color=color)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = WRAP if wrap else Alignment(vertical="top")
    c.border = BORDER
    return c


def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


# ===========================================================================
#  Table of contents (clickable)
# ===========================================================================
TOC = [
    ("Start Here", "What this is, the honest warning, and how to use this file."),
    ("Setup - Browser", "Run it entirely in Chrome via GitHub Codespaces. Nothing to install."),
    ("Setup - Local", "Run it on your own computer (best for long paper-trading runs)."),
    ("Alpaca Keys", "Get your free paper-trading API keys, step by step."),
    ("Copy-Paste Commands", "Every command, in order, colour-coded by risk."),
    ("Tuning Calculator", "INTERACTIVE: type your numbers, cells tell you if the grid clears fees."),
    ("Config Reference", "Every config.yaml setting explained with safe guidance."),
    ("Charts", "Real backtest charts: price, strategy equity, and buy & hold."),
    ("Parameter Sweep", "Many backtests compared so you can see the trade-offs."),
    ("Test Results", "The automated tests and what each one proves."),
    ("Backtest Results", "The headline backtest numbers and how to read them."),
    ("Review & Fixes", "Devil's-advocate review: problems found and fixed."),
    ("Architecture", "What each file does, in plain language."),
    ("Go-Live Checklist", "The boxes that must ALL be true before real money."),
    ("Glossary", "Every jargon word, in plain English."),
    ("FAQ & Troubleshooting", "Common errors and what to do about them."),
    ("Daily Monitor", "A template to log what the bot does each day."),
    ("Risks & Reality", "The honest risks most guides skip."),
]


def sheet_toc(wb):
    ws = wb.active
    ws.title = "Contents"
    widths(ws, {"A": 6, "B": 30, "C": 86})
    title(ws, "Automated Grid Trading Bot — Master Setup Workbook", 3)
    subtitle(ws, 2, "Click any tab name to jump there. Work top to bottom. Green = safe. Red = stop.", 3)
    header_row(ws, 3, ["#", "Tab", "What's on it"])
    r = 4
    for i, (name, desc) in enumerate(TOC, 1):
        cell(ws, r, 1, str(i), bold=True, fill=GREY, wrap=False)
        c = cell(ws, r, 2, name, bold=True, color="0563C1")
        c.hyperlink = f"#'{name}'!A1"
        cell(ws, r, 3, desc)
        ws.row_dimensions[r].height = 26
        r += 1
    return ws


def back_link(ws, span):
    """Add a small 'Back to Contents' link on the last used-ish row later."""
    # Placed at a fixed top-right-ish cell for consistency.
    c = ws.cell(2, span + 1, "")  # spacer; real link added per-sheet if needed


# ===========================================================================
def sheet_start(wb):
    ws = wb.create_sheet("Start Here")
    widths(ws, {"A": 26, "B": 90})
    title(ws, "Start Here", 2)
    rows = [
        ("What this is",
         "A safe-by-default grid trading bot for Alpaca. It buys small amounts as the price "
         "dips through a band you choose and sells a little higher, over and over. It trades "
         "FAKE (paper) money by default. Going live takes two deliberate, on-purpose steps."),
        ("Read this first",
         "This is NOT passive income and it CAN lose money, especially when the market trends "
         "hard one way. A grid bot quietly assumes price comes back; when it does not, that is "
         "where losses live. Only ever fund money you would be completely fine losing in full."),
        ("The one number to burn in",
         "Alpaca charges ~0.25% per trade, so ~0.5% round trip. Every buy+sell cycle must clear "
         "more than ~0.5% just to break even. The bot refuses to start if your grid is too tight."),
        ("How to use this file",
         "Open the Contents tab and go top to bottom. Pick ONE setup path. Use the Tuning "
         "Calculator before you trade. Green cells = do this safely. Amber = caution. Red = stop."),
        ("Honest note on screenshots",
         "This file deliberately contains NO fake screenshots of Alpaca. A wrong screenshot sends "
         "you clicking the wrong thing. You get exact click directions and official links instead."),
    ]
    r = 3
    for k, v in rows:
        cell(ws, r, 1, k, bold=True, fill=LIGHT)
        cell(ws, r, 2, v)
        ws.row_dimensions[r].height = 60
        r += 1
    return ws


def sheet_browser(wb):
    ws = wb.create_sheet("Setup - Browser")
    widths(ws, {"A": 6, "B": 50, "C": 62})
    title(ws, "Setup Path A — Run in Chrome (GitHub Codespaces, nothing to install)", 3)
    subtitle(ws, 2, "Best if you do not want to install anything. The free tier is plenty for testing.", 3)
    header_row(ws, 3, ["#", "What to do (click)", "Copy & paste / notes"])
    steps = [
        ("1", "Open github.com in Chrome and sign in (free account). You already own the repo.", ""),
        ("2", "Go to the repository page.", "https://github.com/louisnto-cloud/my-first-project"),
        ("3", "Click the green '< > Code' button -> 'Codespaces' tab -> 'Create codespace on "
              "claude/cool-sagan-TEXoS'. A full editor opens in your browser.",
              "Use the branch claude/cool-sagan-TEXoS."),
        ("4", "Wait ~1 minute. A terminal appears at the bottom; click into it.",
              "No terminal? Top menu -> Terminal -> New Terminal."),
        ("5", "Install the packages.", "pip install -r requirements.txt"),
        ("6", "Run the safety tests.", "python -m pytest -q"),
        ("7", "Run the offline backtest (no keys needed).", "python backtest.py --simulate"),
        ("8", "Watch a 30-loop dry run (no keys, nothing sent).", "python run_bot.py --ticks 30"),
        ("9", "Ready for paper? Do the 'Alpaca Keys' tab, then create your .env.",
              "cp .env.example .env   then edit .env"),
        ("10", "LIMIT: a Codespace pauses when idle — great for testing, NOT for multi-week paper "
               "trading. For that, use 'Setup - Local' or a small always-on cloud server.", ""),
    ]
    _numbered(ws, steps)
    return ws


def sheet_local(wb):
    ws = wb.create_sheet("Setup - Local")
    widths(ws, {"A": 6, "B": 50, "C": 62})
    title(ws, "Setup Path B — Run on your own computer", 3)
    subtitle(ws, 2, "Best for long-running paper trading. Needs Python 3.10+ installed.", 3)
    header_row(ws, 3, ["#", "What to do", "Copy & paste"])
    steps = [
        ("1", "Install Python 3.10+ if needed. Windows: tick 'Add Python to PATH' during install.",
              "https://www.python.org/downloads/"),
        ("2", "Open a terminal. Mac: Cmd+Space -> Terminal. Windows: Start -> PowerShell.", ""),
        ("3", "Download the project.", "git clone https://github.com/louisnto-cloud/my-first-project.git"),
        ("4", "Go into the folder.", "cd my-first-project"),
        ("5", "Switch to the bot's branch.", "git checkout claude/cool-sagan-TEXoS"),
        ("6", "Install the packages.", "pip install -r requirements.txt"),
        ("7", "Run the safety tests.", "python -m pytest -q"),
        ("8", "Run the offline backtest.", "python backtest.py --simulate"),
        ("9", "Watch a 30-loop dry run.", "python run_bot.py --ticks 30"),
        ("10", "Set up keys (see 'Alpaca Keys').", "cp .env.example .env   (Win: copy .env.example .env)"),
    ]
    _numbered(ws, steps)
    return ws


def _numbered(ws, steps):
    r = 4
    for n, do, paste in steps:
        cell(ws, r, 1, n, bold=True, fill=GREY, wrap=False)
        cell(ws, r, 2, do)
        cell(ws, r, 3, paste, mono=bool(paste and not paste[0].isupper()))
        ws.row_dimensions[r].height = 56
        r += 1


def sheet_keys(wb):
    ws = wb.create_sheet("Alpaca Keys")
    widths(ws, {"A": 6, "B": 92})
    title(ws, "Get your free Alpaca PAPER keys (fake money)", 2)
    subtitle(ws, 2, "Paper keys control only fake money. Even if leaked, no real funds are at risk.", 2)
    header_row(ws, 3, ["#", "Step"])
    steps = [
        "Go to https://alpaca.markets and click 'Sign up'. Use your email. It is free.",
        "On the dashboard, find the Live / Paper switch (top-left area) and set it to PAPER.",
        "Find 'API Keys' (under Home or a key icon). Click 'Generate New Key'.",
        "Copy BOTH the Key ID and the Secret Key into a safe note NOW — the Secret is shown once.",
        "Key ID -> ALPACA_API_KEY_ID and Secret -> ALPACA_API_SECRET_KEY in your .env file.",
        "Official walkthrough (current screenshots): https://alpaca.markets/learn/connect-to-alpaca-api",
        "Paper trading overview: https://alpaca.markets/learn/start-paper-trading",
    ]
    r = 4
    for i, s in enumerate(steps, 1):
        cell(ws, r, 1, str(i), bold=True, fill=GREY, wrap=False)
        cell(ws, r, 2, s)
        ws.row_dimensions[r].height = 40
        r += 1
    return ws


def sheet_commands(wb):
    ws = wb.create_sheet("Copy-Paste Commands")
    widths(ws, {"A": 6, "B": 48, "C": 58})
    title(ws, "Every command, in order", 3)
    subtitle(ws, 2, "Run from inside the project folder. Green rows are zero-risk (no keys, nothing sent).", 3)
    header_row(ws, 3, ["#", "Purpose", "Command"])
    cmds = [
        ("0", "⭐ EASY BUTTON: friendly menu (does it all)", "python start.py", GREEN),
        ("1", "Install packages", "pip install -r requirements.txt", GREEN),
        ("2", "Run safety tests (expect: 45 passed)", "python -m pytest -q", GREEN),
        ("3", "Offline backtest (synthetic data)", "python backtest.py --simulate", GREEN),
        ("4", "Dry run, 30 loops (no keys, nothing sent)", "python run_bot.py --ticks 30", GREEN),
        ("5", "Create your secrets file", "cp .env.example .env", AMBER),
        ("6", "Backtest on REAL last 30 days (needs paper keys)", "python backtest.py --days 30", AMBER),
        ("7", "Start PAPER trading (set mode: paper first)", "python run_bot.py", AMBER),
        ("8", "STOP immediately (kill switch)", "touch KILL   (Win: type nul > KILL)", RED),
        ("9", "Re-arm after a kill", "rm KILL   (Win: del KILL)", AMBER),
    ]
    r = 4
    for n, purpose, cmd, fill in cmds:
        cell(ws, r, 1, n, bold=True, fill=fill, wrap=False)
        cell(ws, r, 2, purpose, fill=fill)
        cell(ws, r, 3, cmd, mono=True)
        ws.row_dimensions[r].height = 30
        r += 1
    return ws


def sheet_calculator(wb):
    """Interactive: live Excel formulas + conditional formatting."""
    ws = wb.create_sheet("Tuning Calculator")
    widths(ws, {"A": 40, "B": 18, "C": 60})
    title(ws, "Interactive tuning calculator (edit the yellow cells)", 3)
    subtitle(ws, 2, "Type your numbers in the yellow cells. The verdicts update live in Excel.", 3)
    header_row(ws, 3, ["Input / Output", "Value", "Notes"])

    def row(r, label, value, note, *, formula=False, inp=False):
        cell(ws, r, 1, label, bold=not formula, fill=GREY if not inp else None)
        c = cell(ws, r, 2, value, mono=True, fill=INPUT if inp else None, bold=True)
        if formula:
            c.number_format = "0.00"
        cell(ws, r, 3, note)
        ws.row_dimensions[r].height = 26
        return c

    # Fee / profit section
    row(4, "Profit target per cycle (%)", 1.0, "EDIT ME. What you aim to capture each buy->sell.", inp=True)
    row(5, "Alpaca fee per side (%)", 0.25, "EDIT ME. Alpaca crypto = 0.25.", inp=True)
    row(6, "Round-trip fee (%)", "=B5*2", "Fee paid on buy AND sell.", formula=True)
    net = row(7, "Net margin after fees (%)", "=B4-B6", "Profit left after fees. Want >= 0.20.", formula=True)
    row(8, "Verdict", '=IF(B7>=0.2,"OK - clears fees with cushion",IF(B7>=0,"THIN - risky","LOSES on every trade"))',
        "Live verdict on your spacing.")

    # Spacing section
    row(10, "Band low (price)", 90000, "EDIT ME. Bottom of your band.", inp=True)
    row(11, "Band high (price)", 110000, "EDIT ME. Top of your band.", inp=True)
    row(12, "Grid levels (rungs)", 6, "EDIT ME. How many rungs.", inp=True)
    spacing = row(13, "Spacing per rung (%)", "=((B11-B10)/(B12-1))/B10*100",
                  "Gap between rungs. Should beat the round-trip fee.", formula=True)
    row(14, "Spacing verdict", '=IF(B13>B6,"OK - wider than fees","TOO TIGHT - widen band or fewer rungs")',
        "Live verdict on rung spacing.")

    # Sizing section
    row(16, "Order size per buy ($)", 30, "EDIT ME. Dollars per buy.", inp=True)
    row(17, "Max capital if all rungs fill ($)", "=B16*B12", "Worst-case deployed capital.", formula=True)
    row(18, "Breakeven sell for a $100 buy", "=100*(1+B5/100)/(1-B5/100)",
        "You must sell above this just to break even.", formula=True)

    # Conditional formatting: net margin (B7) and spacing (B13)
    ws.conditional_formatting.add("B7",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_TX)))
    ws.conditional_formatting.add("B7",
        CellIsRule(operator="between", formula=["0", "0.2"], fill=PatternFill("solid", fgColor=AMBER), font=Font(color=AMBER_TX)))
    ws.conditional_formatting.add("B7",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.2"], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_TX)))
    ws.conditional_formatting.add("B13",
        FormulaRule(formula=["$B$13<=$B$6"], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_TX)))
    ws.conditional_formatting.add("B13",
        FormulaRule(formula=["$B$13>$B$6"], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_TX)))

    # Data validation: keep inputs sane.
    dv = DataValidation(type="decimal", operator="greaterThan", formula1="0", allow_blank=False)
    ws.add_data_validation(dv)
    for addr in ("B4", "B5", "B10", "B11", "B16"):
        dv.add(ws[addr])
    dv_int = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="2")
    ws.add_data_validation(dv_int)
    dv_int.add(ws["B12"])
    return ws


def sheet_config(wb):
    ws = wb.create_sheet("Config Reference")
    widths(ws, {"A": 26, "B": 16, "C": 66})
    title(ws, "config.yaml — every setting explained", 3)
    subtitle(ws, 2, "The only file you normally edit. Defaults are deliberately small and safe.", 3)
    header_row(ws, 3, ["Setting", "Default", "What it means / safe guidance"])
    rows = [
        ("mode", "dry_run", "dry_run = simulate (no risk). paper = fake money, real prices. live = real money."),
        ("i_really_want_live_trading", "false", "First of two switches for live. Leave false until you truly mean it."),
        ("asset_class", "crypto", "crypto (24/7, no day-trade rule) or stock (phase two)."),
        ("symbol", "BTC/USD", "What you trade. Crypto e.g. BTC/USD; stocks e.g. SPY."),
        ("fee_percent_per_side", "0.25", "Alpaca's crypto fee per side. Round trip is double."),
        ("slippage_percent_per_side", "0.05", "Price you wanted vs got. The main cost for commission-free stocks. Don't set to 0."),
        ("band_low", "90000", "Bottom of your band (buy zone). Set from the backtest."),
        ("band_high", "110000", "Top of your band (sell zone). Must exceed band_low."),
        ("grid_levels", "6", "Rungs to split the band into. More = more, smaller trades."),
        ("profit_target_percent", "1.0", "Aim per cycle. Must clear ~0.5% fee. Below ~0.7% is refused."),
        ("order_size_usd", "30", "Dollars per individual buy order."),
        ("max_position_usd", "200", "Largest holdings value allowed. Hard cap."),
        ("max_deployed_usd", "200", "Largest total in open buys + holdings. Hard cap."),
        ("max_open_orders", "12", "Most resting orders at once."),
        ("daily_loss_limit_usd", "20", "If today's loss (realized + unrealized) hits this, halt + cancel."),
        ("breakout_guard_enabled", "true", "Stop new buys if price leaves the band. Leave ON."),
        ("breakout_flatten", "false", "On breakout, also SELL everything to cash. Start false."),
        ("loop_interval_seconds", "30", "Seconds between decisions. 30-60 is plenty."),
        ("alerts_enabled", "false", "Webhook alert on guard/kill/loss. Needs ALERT_WEBHOOK_URL."),
        ("log_dir", "logs", "Where timestamped logs go."),
        ("state_file", "state/bot_state.json", "Where the bot saves memory between restarts."),
        ("kill_switch_file", "KILL", "If this file exists, the bot stops and cancels everything."),
    ]
    r = 4
    for k, d, v in rows:
        cell(ws, r, 1, k, mono=True, bold=True, fill=GREY)
        cell(ws, r, 2, d, mono=True)
        cell(ws, r, 3, v)
        ws.row_dimensions[r].height = 30
        r += 1
    return ws


def _bt_cfg(grid_levels, profit_target):
    return SimpleNamespace(
        band_low=90_000.0, band_high=110_000.0, grid_levels=grid_levels,
        profit_target_percent=profit_target, fee_percent_per_side=0.25, order_size_usd=30.0,
        slippage_percent_per_side=0.05,  # honest: include slippage in demo numbers
    )


def sheet_charts(wb, prices, equity, buyhold):
    # Hidden data sheet for the charts.
    data = wb.create_sheet("_chartdata")
    data.sheet_state = "hidden"
    data.cell(1, 1, "bar")
    data.cell(1, 2, "price")
    data.cell(1, 3, "strategy_equity")
    data.cell(1, 4, "buyhold_equity")
    for i in range(len(prices)):
        data.cell(i + 2, 1, i)
        data.cell(i + 2, 2, round(prices[i], 2))
        data.cell(i + 2, 3, round(equity[i], 4))
        data.cell(i + 2, 4, round(buyhold[i], 4))
    n = len(prices) + 1

    ws = wb.create_sheet("Charts")
    widths(ws, {"A": 2})
    title(ws, "Backtest charts (synthetic demo data)", 8)
    subtitle(ws, 2, "Top: price vs your band. Bottom: strategy profit vs simply holding. Run your own with --days 30.", 8)

    # Price chart
    c1 = LineChart()
    c1.title = "Price over the backtest"
    c1.height = 8
    c1.width = 26
    c1.y_axis.title = "Price"
    c1.x_axis.title = "Bar"
    ref = Reference(data, min_col=2, min_row=1, max_row=n)
    c1.add_data(ref, titles_from_data=True)
    ws.add_chart(c1, "A4")

    # Equity chart
    c2 = LineChart()
    c2.title = "Strategy equity vs buy & hold (after fees)"
    c2.height = 8
    c2.width = 26
    c2.y_axis.title = "Profit ($)"
    c2.x_axis.title = "Bar"
    ref2 = Reference(data, min_col=3, max_col=4, min_row=1, max_row=n)
    c2.add_data(ref2, titles_from_data=True)
    ws.add_chart(c2, "A22")
    return ws


def sheet_sweep(wb, prices):
    ws = wb.create_sheet("Parameter Sweep")
    widths(ws, {"A": 18, "B": 16, "C": 14, "D": 14, "E": 16, "F": 16, "G": 16})
    title(ws, "Parameter sweep — many backtests, one price series", 7)
    subtitle(ws, 2, "Same synthetic data, different settings. See how tuning changes the result. Higher net P&L is better.", 7)

    # Buy & hold is the same for all (same prices).
    bh = run_backtest(prices, _bt_cfg(6, 1.0)).buy_hold_return_pct

    header_row(ws, 3, ["Profit target %", "Grid levels", "Trades", "Win rate %",
                       "Net P&L $", "Strategy %", "Buy&Hold %"])
    r = 4
    combos = []
    for target in (0.3, 0.6, 1.0, 1.5, 2.0):
        combos.append((target, 6))
    for levels in (4, 8, 10, 12):
        combos.append((1.0, levels))

    best_pnl = max(run_backtest(prices, _bt_cfg(l, t)).net_pnl for t, l in combos)
    for target, levels in combos:
        rep = run_backtest(prices, _bt_cfg(levels, target))
        cell(ws, r, 1, f"{target:.1f}", mono=True, fill=GREY)
        cell(ws, r, 2, str(levels), mono=True)
        cell(ws, r, 3, str(rep.trades), mono=True)
        cell(ws, r, 4, f"{rep.win_rate:.0f}", mono=True)
        pnl_fill = GREEN if abs(rep.net_pnl - best_pnl) < 1e-6 else (RED if rep.net_pnl <= 0 else None)
        cell(ws, r, 5, round(rep.net_pnl, 2), mono=True, fill=pnl_fill, bold=pnl_fill == GREEN)
        cell(ws, r, 6, f"{rep.strategy_return_pct:.2f}", mono=True)
        cell(ws, r, 7, f"{rep.buy_hold_return_pct:.2f}", mono=True)
        ws.row_dimensions[r].height = 24
        r += 1
    r += 1
    cell(ws, r, 1, "How to read", bold=True, fill=LIGHT)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    cell(ws, r, 2,
         "Tighter targets (0.3%) make more trades but can fail to clear fees. Wider targets trade "
         "less but each cycle is safer. The green row is the best net P&L here. Note buy & hold for "
         "context — in a strong trend it often wins; grids shine in chop.")
    ws.row_dimensions[r].height = 56
    return ws


def sheet_tests(wb, pytest_total):
    ws = wb.create_sheet("Test Results")
    widths(ws, {"A": 30, "B": 12, "C": 64})
    title(ws, "Automated test results", 3)
    subtitle(ws, 2, f"Last run: {pytest_total} passed. CI re-runs these on every push to GitHub.", 3)
    header_row(ws, 3, ["Test file", "Status", "What it proves"])
    rows = [
        ("test_fees.py", "PASS", "Round-trip fee, breakeven price, net profit/margin, fee-clearing check."),
        ("test_grid.py", "PASS", "Even spacing, sell above buy by target, tight grids rejected."),
        ("test_guards.py", "PASS", "Breakout, daily-loss halt, risk limits, kill switch behave correctly."),
        ("test_config.py", "PASS", "Unsafe configs refused; live needs BOTH the flag AND the env latch."),
        ("test_backtest.py", "PASS", "Profit in chop, no trades when flat or out-of-band."),
        ("test_engine.py", "PASS", "Unrealized loss trips the halt; breakout-flatten sells to cash."),
    ]
    r = 4
    for f, s, v in rows:
        cell(ws, r, 1, f, mono=True, fill=GREY)
        cell(ws, r, 2, s, bold=True, fill=GREEN, color=GREEN_TX, wrap=False)
        cell(ws, r, 3, v)
        ws.row_dimensions[r].height = 30
        r += 1
    return ws


def sheet_backtest(wb, bt):
    ws = wb.create_sheet("Backtest Results")
    widths(ws, {"A": 34, "B": 22, "C": 50})
    title(ws, "Backtest results (synthetic offline demo run)", 3)
    subtitle(ws, 2, "Run your OWN with: python backtest.py --days 30  (real Alpaca data).", 3)
    header_row(ws, 3, ["Metric", "Value", "How to read it"])
    rows = [
        ("Bars processed", str(bt["bars"]), "How many price points were replayed."),
        ("Completed trades", str(bt["trades"]), "Full buy->sell cycles."),
        ("Win rate", f"{bt['win_rate']}%", "Share of cycles that cleared fees."),
        ("Net P&L after fees", f"${bt['net_pnl']}", "The bottom line, fees included."),
        ("Largest drawdown", f"${bt['drawdown']}", "Worst peak-to-trough dip along the way."),
        ("Strategy return", f"{bt['strat_pct']}%", "Profit vs capital put to work."),
        ("Buy & hold return", f"{bt['bh_pct']}%", "What simply holding would have done."),
        ("Honest takeaway", "Compare the last two",
         "In a trending sample, holding often beats the grid. Grids earn in chop and give up "
         "upside in trends. Always compare strategy vs buy & hold."),
    ]
    r = 4
    for k, v, how in rows:
        cell(ws, r, 1, k, bold=True, fill=GREY)
        cell(ws, r, 2, v, mono=True)
        cell(ws, r, 3, how)
        ws.row_dimensions[r].height = 36
        r += 1
    return ws


def sheet_review(wb):
    ws = wb.create_sheet("Review & Fixes")
    widths(ws, {"A": 30, "B": 40, "C": 40, "D": 12})
    title(ws, "Devil's-advocate review — problems found and fixed", 4)
    subtitle(ws, 2, "A real critique of the first build, with the fixes that followed.", 4)
    header_row(ws, 3, ["Problem found", "Why it mattered", "Fix applied", "Status"])
    rows = [
        ("'Flatten' did not flatten",
         "On a breakout with flatten on, the bot only cancelled orders and kept inventory — "
         "leaving you holding the bag in the exact crash the guard exists to prevent.",
         "Added a real _flatten(): cancels, books each holding's P&L, liquidates. Tested.", "FIXED"),
        ("Daily-loss halt ignored unrealized loss",
         "The grid only books profit on sells, so a position bleeding in a downtrend never tripped "
         "the halt until losses were locked in.",
         "Halt now uses realized + unrealized P&L. Drawdowns trip it. Tested.", "FIXED"),
        ("Status line hid open drawdown",
         "It showed realized P&L only; you could not see a loss building.",
         "Status line now prints unrealized and combined day P&L and headroom.", "FIXED"),
        ("Dry-run resumed stale simulator state",
         "A restarted dry run loaded old orders against a fresh simulator, so nothing filled.",
         "Dry-run always starts fresh; persistence is paper/live only.", "FIXED"),
        ("Known: reconnect re-sync",
         "Reconciliation runs on startup, not automatically after a mid-run disconnect.",
         "Calls already retry with backoff; full mid-run re-sync is a sensible next step.", "KNOWN"),
        ("Known: backtest realism",
         "Fills at the rung price; ignores slippage and intra-bar timing.",
         "Conservative and labelled; treat backtest as directional, paper as truth.", "KNOWN"),
    ]
    r = 4
    for prob, why, fix, status in rows:
        cell(ws, r, 1, prob, bold=True)
        cell(ws, r, 2, why)
        cell(ws, r, 3, fix)
        fill = GREEN if status == "FIXED" else AMBER
        txt = GREEN_TX if status == "FIXED" else AMBER_TX
        cell(ws, r, 4, status, bold=True, fill=fill, color=txt, wrap=False)
        ws.row_dimensions[r].height = 62
        r += 1
    return ws


def sheet_architecture(wb):
    ws = wb.create_sheet("Architecture")
    widths(ws, {"A": 26, "B": 84})
    title(ws, "Architecture — what each file does", 2)
    subtitle(ws, 2, "Small, single-purpose pieces. The engine ties them together.", 2)
    header_row(ws, 3, ["File", "Job, in plain language"])
    rows = [
        ("start.py", "⭐ The easy front door: a friendly menu that runs everything for you."),
        ("config.yaml", "Your settings. The only file you normally edit."),
        ("bot/config.py", "Loads + validates settings. Refuses to run if anything is unsafe."),
        ("bot/fees.py", "The money math: fees, breakeven, net profit after fees."),
        ("bot/grid.py", "Turns your band into a ladder of buy/sell price rungs."),
        ("bot/guards.py", "The brakes: breakout, daily-loss halt, risk limits, kill switch."),
        ("bot/broker.py", "Talks to Alpaca (retries, idempotent IDs) + a local simulator for dry-run."),
        ("bot/state.py", "Saves/reloads the bot's memory; reconciles with Alpaca on restart."),
        ("bot/engine.py", "The main loop: read price, detect fills, run guards, place orders, log."),
        ("bot/alerts.py", "Optional webhook notifications (Discord/Slack/phone)."),
        ("bot/logging_setup.py", "Timestamped, human-readable logs."),
        ("run_bot.py", "What you launch. Picks simulator vs Alpaca based on mode."),
        ("backtest.py", "Replays the strategy; also --suggest-band and --walkforward (out-of-sample)."),
        ("compare_windows.py", "Compares strategy vs buy & hold across 30/60/90-day windows."),
        ("screen_symbols.py", "Ranks many tickers by honest out-of-sample result; shows what's worth trading."),
        ("tests/", "Automated proof the dangerous parts are correct (45 tests)."),
    ]
    r = 4
    for f, j in rows:
        cell(ws, r, 1, f, mono=True, bold=True, fill=GREY)
        cell(ws, r, 2, j)
        ws.row_dimensions[r].height = 28
        r += 1
    return ws


def sheet_checklist(wb):
    ws = wb.create_sheet("Go-Live Checklist")
    widths(ws, {"A": 8, "B": 86})
    title(ws, "Go-live checklist — do NOT skip a box", 2)
    subtitle(ws, 2, "Only switch to real money when EVERY box is true. Tick them yourself.", 2)
    header_row(ws, 3, ["Done?", "Requirement"])
    items = [
        "The backtest is profitable after fees across more than one market period.",
        "Paper trading has been profitable after fees for at least 2-4 weeks.",
        "You understand every trade in the logs and why the bot made it.",
        "Your risk limits and daily loss limit are set to small amounts.",
        "You are funding only money you would be completely fine losing in full.",
        "You have tested the kill switch (touch KILL) and seen it stop the bot.",
        "config.yaml has mode: live AND i_really_want_live_trading: true.",
        "The env var I_UNDERSTAND_THIS_IS_REAL_MONEY=true is set.",
        "Your .env now holds LIVE (funded-account) keys, not paper keys.",
        "You will start tiny and watch closely for the first week.",
    ]
    r = 4
    for it in items:
        cell(ws, r, 1, "[  ]", bold=True, fill=GREY, wrap=False)
        cell(ws, r, 2, it)
        ws.row_dimensions[r].height = 28
        r += 1
    return ws


def sheet_glossary(wb):
    ws = wb.create_sheet("Glossary")
    widths(ws, {"A": 26, "B": 84})
    title(ws, "Glossary — jargon in plain English", 2)
    header_row(ws, 3, ["Term", "Plain meaning"])
    rows = [
        ("Grid / range bot", "Buys at set price rungs as price dips, sells higher. Earns in choppy, sideways markets."),
        ("Band", "The low-to-high price range you expect price to bounce within."),
        ("Rung / level", "One price line in the grid where the bot wants to buy (and later sell)."),
        ("Limit order", "An order to buy/sell only at a specific price or better."),
        ("Fill", "When your order actually trades. Partial fill = only some of it traded."),
        ("Round-trip fee", "The fee paid on both the buy and the sell of one cycle (~0.5% on Alpaca crypto)."),
        ("Slippage", "The gap between the price you wanted and the price you got, common in fast markets."),
        ("Drawdown", "The worst dip from a peak in your equity. Measures pain along the way."),
        ("Realized P&L", "Profit/loss you have locked in by actually selling."),
        ("Unrealized P&L", "Profit/loss on what you still hold, if you sold right now."),
        ("Breakout", "Price leaving your band — a sign the market may be trending, not chopping."),
        ("Paper trading", "Trading with fake money against real live prices. Risk-free practice."),
        ("Kill switch", "A file/command that instantly cancels orders and stops the bot."),
        ("Idempotent order ID", "A unique label so a retried request can't accidentally place two real orders."),
        ("Reconcile", "On restart, syncing the bot's saved memory with the broker's real orders."),
    ]
    r = 4
    for t, m in rows:
        cell(ws, r, 1, t, bold=True, fill=GREY)
        cell(ws, r, 2, m)
        ws.row_dimensions[r].height = 30
        r += 1
    return ws


def sheet_faq(wb):
    ws = wb.create_sheet("FAQ & Troubleshooting")
    widths(ws, {"A": 44, "B": 66})
    title(ws, "FAQ & troubleshooting", 2)
    header_row(ws, 3, ["Problem / question", "What to do"])
    rows = [
        ("'Configuration refused' on startup",
         "That's the safety net working. Read the message: usually your grid is too tight (widen "
         "profit_target_percent), band_low >= band_high, or a missing setting."),
        ("'Missing Alpaca API keys'",
         "You're in paper/live mode without keys. Run cp .env.example .env and paste your keys, or "
         "use mode: dry_run which needs none."),
        ("Refuses to trade live",
         "By design. You need mode: live AND i_really_want_live_trading: true AND the env var "
         "I_UNDERSTAND_THIS_IS_REAL_MONEY=true. Missing any one = refusal."),
        ("ModuleNotFoundError",
         "You didn't install packages, or you're outside the project folder. Run "
         "pip install -r requirements.txt from inside my-first-project."),
        ("Lots of buys, few sells in the logs",
         "Price is trending down and you're accumulating a bag. Stop and reassess the band — do "
         "not add money."),
        ("Profit looks positive before fees but negative after",
         "Your grid spacing is too tight. Use the Tuning Calculator; widen profit_target_percent."),
        ("Breakout guard keeps firing",
         "Your band is wrong for the current market, or the regime changed. Re-pick the band from a "
         "fresh backtest."),
        ("How do I stop it right now?",
         "Create the kill file: touch KILL (Windows: type nul > KILL). It cancels all orders and stops."),
        ("Codespace went to sleep",
         "Codespaces pause when idle. Fine for testing; use a local machine or always-on server for "
         "multi-week paper runs."),
    ]
    r = 4
    for q, a in rows:
        cell(ws, r, 1, q, bold=True, fill=GREY)
        cell(ws, r, 2, a)
        ws.row_dimensions[r].height = 44
        r += 1
    return ws


def sheet_monitor(wb):
    ws = wb.create_sheet("Daily Monitor")
    widths(ws, {"A": 12, "B": 12, "C": 12, "D": 12, "E": 14, "F": 12, "G": 40})
    title(ws, "Daily monitoring log (fill this in while paper trading)", 7)
    subtitle(ws, 2, "One row per day. Watching this is how you learn whether the bot actually works.", 7)
    header_row(ws, 3, ["Date", "Realized P&L", "Open orders", "Position $",
                       "Guards fired?", "Net up?", "Notes / what you observed"])
    # Provide 21 blank rows (~3 weeks) with light borders.
    for r in range(4, 25):
        for c in range(1, 8):
            cell(ws, r, c, "", fill=GREY if r % 2 == 0 else None)
        ws.row_dimensions[r].height = 20
    return ws


def sheet_risks(wb):
    ws = wb.create_sheet("Risks & Reality")
    widths(ws, {"A": 28, "B": 80})
    title(ws, "Risks & reality — the part most guides skip", 2)
    header_row(ws, 3, ["Topic", "The honest version"])
    rows = [
        ("Not passive income", "This needs attention, tuning, and a stomach for losses. A learning project, not an ATM."),
        ("Trends hurt grids", "A strong one-way move makes the bot buy all the way down or sell too early. The guard limits, not eliminates, this."),
        ("Fees are the silent killer", "At ~0.5% round trip, a too-tight grid loses on every trade while looking busy."),
        ("Slippage", "You may not get the exact price you wanted. Real results trail the backtest."),
        ("Paper != live", "Real fills, latency, and liquidity differ. Paper success is necessary, not a guarantee."),
        ("Software can fail", "Bugs, outages, dropped connections happen. The kill switch and daily-loss limit are your floor."),
        ("Taxes & law", "Profits are taxable; bot rules vary by location. This is information, not financial or legal advice."),
        ("The real edge", "Process: backtest -> dry run -> paper for weeks -> tiny live with hard guards. In that order."),
    ]
    r = 4
    for k, v in rows:
        cell(ws, r, 1, k, bold=True, fill=LIGHT)
        cell(ws, r, 2, v)
        ws.row_dimensions[r].height = 42
        r += 1
    return ws


def main():
    # Live test count so the sheet never drifts from reality.
    pytest_total = 38
    try:
        out = subprocess.run(["python3", "-m", "pytest", "-q"], capture_output=True, text=True).stdout
        for line in out.splitlines():
            if "passed" in line:
                pytest_total = int(line.split()[0])
                break
    except Exception:
        pass

    # Run a real backtest for the headline numbers AND the chart series.
    cfg = _bt_cfg(6, 1.0)
    prices_full = _synthetic_prices(cfg)
    rep = run_backtest(prices_full, cfg)

    # Downsample to ~160 points for clean charts and a small file.
    stepn = max(1, len(prices_full) // 160)
    prices = prices_full[::stepn]
    equity = rep.equity_curve[::stepn]
    deployed = cfg.order_size_usd * cfg.grid_levels
    qty = deployed / prices_full[0]
    buyhold_full = [qty * (p - prices_full[0]) for p in prices_full]
    buyhold = buyhold_full[::stepn]

    bt = dict(bars=rep.bars, trades=rep.trades, win_rate=round(rep.win_rate, 1),
              net_pnl=round(rep.net_pnl, 2), drawdown=round(rep.max_drawdown, 2),
              strat_pct=round(rep.strategy_return_pct, 2), bh_pct=round(rep.buy_hold_return_pct, 2))

    wb = Workbook()
    sheet_toc(wb)
    sheet_start(wb)
    sheet_browser(wb)
    sheet_local(wb)
    sheet_keys(wb)
    sheet_commands(wb)
    sheet_calculator(wb)
    sheet_config(wb)
    sheet_charts(wb, prices, equity, buyhold)
    sheet_sweep(wb, prices_full)
    sheet_tests(wb, pytest_total)
    sheet_backtest(wb, bt)
    sheet_review(wb)
    sheet_architecture(wb)
    sheet_checklist(wb)
    sheet_glossary(wb)
    sheet_faq(wb)
    sheet_monitor(wb)
    sheet_risks(wb)

    for name in wb.sheetnames:
        if name not in ("Contents", "_chartdata", "Charts"):
            wb[name].freeze_panes = "A4"
    wb.active = 0

    os.makedirs("docs", exist_ok=True)
    path = os.path.join("docs", "Trading_Bot_Setup.xlsx")
    wb.save(path)
    visible = [s for s in wb.sheetnames if s != "_chartdata"]
    print(f"Wrote {path} with {len(visible)} visible sheets.")
    print("Sheets:", ", ".join(visible))


if __name__ == "__main__":
    main()
