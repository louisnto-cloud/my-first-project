#!/usr/bin/env python3
"""
Ingest live-captured Amazon listing data into the competitor-intel workbook.

Reads one or more capture CSV files (schema = capture_template.csv) and writes a
"Verified Listings (Live)" sheet into the workbook, with calculated columns as
live formulas and provenance stamped "Verified live <date>". Idempotent: rebuilds
the sheet from the union of all provided CSV rows, de-duplicated by ASIN.

Usage:
    python3 ingest_captures.py [workbook.xlsx] [capture1.csv capture2.csv ...]
Defaults: newest Amazon_ca_*.xlsx in cwd; all *.csv matching capture*.csv + captures/*.csv
Rows whose asin is blank or starts with EXAMPLE are skipped.
"""
import sys, csv, glob, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import DataBarRule

TODAY = datetime.date.today().isoformat()
SHEET = "Verified Listings (Live)"
C1="1) Sparkling electrolyte / functional sparkling"; C2="2) Electrolyte beverage (RTD)"
C3="3) Functional beverage (still/sparkling)"; C4="4) Caffeine & energy beverage"
C5="5) Sparkling water & soda"
def map_cat(v):
    v=(v or "").strip().lower()
    if v in ("1","c1") or "sparkling electrolyte" in v or "functional sparkling" in v: return C1
    if v in ("2","c2") or "electrolyte" in v: return C2
    if v in ("3","c3") or "functional" in v or "prebiotic" in v or "gut" in v: return C3
    if v in ("4","c4") or "energy" in v or "caffeine" in v: return C4
    if v in ("5","c5") or "sparkling water" in v or "soda" in v: return C5
    return v or "Uncategorized"

HDRS=["ASIN","URL","Brand","Product Title","Category","Format & Size","Unit mL","Pack",
 "Price (CAD)","Price/Unit","Price/100mL","S&S?","S&S %","S&S Price","Coupon (CAD)","Promo",
 "Lowest Effective","Flavours","Caffeine mg","Sodium mg","Potassium mg","Sugar g","Calories",
 "Sweetener","Claims","Star","#Ratings","Badge","BSR","Caffeine/100mL","Sugar/100mL",
 "Health Score","Functional Score","Provenance","Captured Date"]
WIDTHS=[14,28,16,34,24,15,8,7,12,11,11,7,8,12,11,20,14,30,10,9,10,8,9,18,30,7,10,16,18,11,11,11,12,16,13]
MONEY='"$"#,##0.00'; PCT='0%'; NAVY="1F3864"; AMBER="FFF2CC"; GREEN="E2EFDA"
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
def num(x):
    if x is None: return None
    s=str(x).strip().replace("$","").replace(",","").replace("%","")
    if s=="" or s.lower() in ("not listed","n/a","na","none"): return None
    try:
        f=float(s); return int(f) if f==int(f) else f
    except: return None

def fnum(x):  # for ss_pct allow "5" or "0.05" or "5%"
    n=num(x)
    if n is None: return None
    return n/100.0 if n>1 else n

def collect(csvs):
    rows=[]; seen=set()
    for path in csvs:
        try: fh=open(path,newline="",encoding="utf-8-sig")
        except FileNotFoundError: continue
        with fh:
            for d in csv.DictReader(fh):
                a=(d.get("asin") or "").strip()
                if not a or a.upper().startswith("EXAMPLE"): continue
                if a in seen: continue
                seen.add(a); rows.append(d)
    return rows

def build(wbpath, csvs):
    wb=load_workbook(wbpath)
    if SHEET in wb.sheetnames: del wb[SHEET]
    ws=wb.create_sheet(SHEET)
    # place right after the 5 master tabs if possible
    try:
        idx=max(i for i,s in enumerate(wb._sheets) if s.title.startswith("M") and " - " in s.title)+1
        wb._sheets.remove(ws); wb._sheets.insert(idx,ws)
    except ValueError: pass
    ws.sheet_properties.tabColor="00B050"
    for i,w in enumerate(WIDTHS,1): ws.column_dimensions[get_column_letter(i)].width=w
    for i,h in enumerate(HDRS,1):
        c=ws.cell(1,i,h); c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=BORDER
    ws.row_dimensions[1].height=30; ws.freeze_panes="C2"
    rows=collect(csvs)
    r=2
    for d in rows:
        I,H,G,M,O,S,T,U,V,W = (f"I{r}",f"H{r}",f"G{r}",f"M{r}",f"O{r}",f"S{r}",f"T{r}",f"U{r}",f"V{r}",f"W{r}")
        vals=[
          d.get("asin"), d.get("url"), d.get("brand"), d.get("product_title"), map_cat(d.get("category")),
          d.get("format_size"), num(d.get("unit_ml")), num(d.get("pack_count")), num(d.get("price_cad")),
          f'=IF(AND(ISNUMBER({I}),ISNUMBER({H}),{H}<>0),{I}/{H},"")',
          f'=IF(AND(ISNUMBER({I}),ISNUMBER({G}),ISNUMBER({H}),({G}*{H})<>0),{I}/({G}*{H})*100,"")',
          d.get("ss_offered"), fnum(d.get("ss_pct")),
          f'=IF(AND(ISNUMBER({I}),ISNUMBER({M})),{I}*(1-{M}),"")',
          num(d.get("coupon_value_cad")), d.get("promo_text"),
          f'=IF(ISNUMBER({I}),({I}-IF(ISNUMBER({O}),{O},0))*(1-IF(ISNUMBER({M}),{M},0)),"")',
          d.get("flavours"), num(d.get("caffeine_mg")), num(d.get("sodium_mg")), num(d.get("potassium_mg")),
          num(d.get("sugar_g")), num(d.get("calories")), d.get("sweetener"), d.get("claims"),
          num(d.get("star_rating")), num(d.get("num_ratings")), d.get("badge"), d.get("bsr"),
          f'=IF(AND(ISNUMBER({S}),ISNUMBER({G}),{G}<>0),ROUND({S}/{G}*100,1),"")',
          f'=IF(AND(ISNUMBER({V}),ISNUMBER({G}),{G}<>0),ROUND({V}/{G}*100,2),"")',
          f'=IF(ISNUMBER({V}),ROUND(MAX(0,MIN(100,100-{V}*2.5-IF(ISNUMBER({W}),{W},0)*0.12+IF({V}<=1,12,0))),0),"")',
          (f'=IF(ISNUMBER({T}),IF({T}>=100,20,IF({T}>=30,10,0)),0)+IF(ISNUMBER({U}),IF({U}>=100,20,0),0)'
           f'+IF(ISNUMBER({S}),IF({S}>=150,20,IF({S}>=50,12,5)),0)+IF(ISNUMBER({V}),IF({V}<=1,15,0),0)'),
          f"Verified live {TODAY}", TODAY,
        ]
        for ci,v in enumerate(vals,1):
            c=ws.cell(r,ci,v); c.border=BORDER; c.alignment=Alignment(vertical="top",wrap_text=True)
            if ci in (9,10,11,14,15,17): c.number_format=MONEY
            if ci==13: c.number_format=PCT
        r+=1
    last=r-1
    if last>=2:
        t=Table(displayName="tblVerified",ref=f"A1:{get_column_letter(len(HDRS))}{last}")
        t.tableStyleInfo=TableStyleInfo(name="TableStyleMedium7",showRowStripes=True); ws.add_table(t)
        ws.conditional_formatting.add(f"AF2:AF{last}",DataBarRule(start_type="num",start_value=0,end_type="num",end_value=100,color="63BE7B"))
    else:
        n=ws.cell(2,1,"No verified rows yet. Fill capture_template.csv (or export from browser-Claude) and run: python3 ingest_captures.py")
        n.font=Font(italic=True,color="C00000")
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=10)
    try: wb.calculation.fullCalcOnLoad=True
    except Exception: pass
    wb.save(wbpath)
    return last-1

if __name__=="__main__":
    args=sys.argv[1:]
    wbs=[a for a in args if a.endswith(".xlsx")]
    csvs=[a for a in args if a.endswith(".csv")]
    wbpath=wbs[0] if wbs else sorted(glob.glob("Amazon_ca_*.xlsx"))[-1]
    if not csvs:
        csvs=sorted(set(glob.glob("capture*.csv")+glob.glob("captures/*.csv")))
    n=build(wbpath,csvs)
    print(f"Ingested {n} verified rows into '{SHEET}' of {wbpath}")
    print(f"CSV sources: {csvs or '(none)'}")
