"""
Costco UK ROI Model — v4

Two big changes vs v3:

  1.  PORTFOLIO TAB. The Bee Propolis market test is still the deep-dive,
      but there's now a one-year, multi-SKU portfolio roll-up using the
      SKUs from the Asana D93 ranking:
        #1 Bee Propolis Spray 30ml x2  (baseline — pre-filled)
        #2 Mag 8-in-1 90's/120's
        #3 O1 450g
        #4 Belli Bliss Raspberry, 450g
        #5 Daily Boost
        #6 Berberine 90's/120's
        #7 L-theanine Capsules 90's/120's
      Costs/units left blank on rows 2-7 — fill in from Asana.
      Bottom-row blended GP%, Net Margin %, ROI on Promo Spend show how
      lower-margin SKUs are offset by higher-margin SKUs.

  2.  STYLE. Reworked to a finance-analyst look instead of the deck/AI
      style: Tahoma 9pt body, ALL-CAPS section headers with a thin
      underline (no navy fill bars), blue font for inputs, black for
      calcs, green for cross-sheet links, accounting number format.
      Named ranges for the major drivers so formulas read like
      =Price*Units instead of =Inputs!$C$12*F6.
"""

import openpyxl
from datetime import date
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

# -----------------------------------------------------------------------------
# STYLE — finance-analyst, hand-rolled look
# -----------------------------------------------------------------------------
INK = "000000"
INPUT_BLUE = "1F497D"       # dark blue for hardcoded inputs
LINK_GREEN = "00643C"       # dark green for cross-sheet links / pulled values
SUBTLE_YELLOW = "FFFCEB"    # very pale yellow for input cells (almost imperceptible)
SUBTLE_GREY = "F7F7F7"      # very pale grey for header rows / totals
RULE_GREY = "8C8C8C"        # mid-grey for sub-rules

thin_black = Side(border_style="thin", color=INK)
hair_grey = Side(border_style="hair", color=RULE_GREY)
medium_black = Side(border_style="medium", color=INK)

# Fonts (Tahoma 9pt body — financial industry default)
def F(size=9, bold=False, italic=False, color=INK, name="Tahoma"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

f_title       = F(size=14, bold=True)
f_subtitle    = F(size=10, italic=True, color=RULE_GREY)
f_section     = F(size=10, bold=True)            # ALL CAPS section headers
f_label       = F(size=9)
f_label_bold  = F(size=9, bold=True)
f_input       = F(size=9, color=INPUT_BLUE)
f_calc        = F(size=9, color=INK)
f_link        = F(size=9, color=LINK_GREEN)
f_total       = F(size=9, bold=True)
f_note        = F(size=8, italic=True, color=RULE_GREY)
f_colhdr      = F(size=9, bold=True)

fill_input    = PatternFill("solid", fgColor=SUBTLE_YELLOW)
fill_hdr      = PatternFill("solid", fgColor=SUBTLE_GREY)
fill_none     = PatternFill(fill_type=None)

align_left    = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right   = Alignment(horizontal="right", vertical="center")
align_center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_top     = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Accounting / number formats (financial convention: negatives in red parens)
FMT_GBP   = '_(£* #,##0_);[Red]_(£* (#,##0);_(£* "-"_);_(@_)'
FMT_GBP_D = '_(£* #,##0.00_);[Red]_(£* (#,##0.00);_(£* "-"??_);_(@_)'
FMT_NUM   = '_(* #,##0_);[Red]_(* (#,##0);_(* "-"_);_(@_)'
FMT_PCT   = '0.0%;[Red](0.0%)'
FMT_MULT  = '0.00"x"'
FMT_DATE  = 'd-mmm'
FMT_DATE_L = 'd-mmm-yyyy'

# -----------------------------------------------------------------------------
# MODEL CONSTANTS
# -----------------------------------------------------------------------------
WEEKS = 26
FIRST_WEEK_NUM = 1
DEFAULT_START_DATE = date(2026, 1, 12)

SCENARIOS = [
    {
        "key": "BEST",
        "name": "BEST — x2 Demo (Wks 5-6) + TPD",
        "promo": {"Demo": [5, 6], "End Cap": [], "Fence": [],
                  "TPD": [5, 6], "Markdown": [], "Advert": []},
        "lift_overrides": {5: 3.0, 6: 3.0},
        "rationale": ("Two demo weekends with TPD. Demo lifts ~3x in-week. "
                      "No end cap, no markdown."),
    },
    {
        "key": "IDEAL",
        "name": "IDEAL — x2 Demo + End Cap (Wks 5-6) + TPD",
        "promo": {"Demo": [5, 6], "End Cap": [5, 6], "Fence": [],
                  "TPD": [5, 6], "Markdown": [], "Advert": []},
        "lift_overrides": {5: 3.6, 6: 3.6},
        "rationale": ("Demo + end cap + TPD stacked for two weeks. "
                      "Combined velocity ~3.6x base."),
    },
    {
        "key": "WORST",
        "name": "WORST — Demo + EndCap + TPD then Markdown (Wks 18-20)",
        "promo": {"Demo": [5, 6], "End Cap": [5, 6], "Fence": [],
                  "TPD": [5, 6], "Markdown": [18, 19, 20], "Advert": []},
        "lift_overrides": {5: 3.0, 6: 3.0, 18: 1.0, 19: 0.3, 20: 8.7},
        "rationale": ("Promo plan as IDEAL but residual inventory clears at "
                      "50% off over Wks 18-20."),
    },
]

PROMO_TYPES = ["Demo", "End Cap", "Fence", "TPD", "Markdown", "Advert"]

# Portfolio SKUs (D93 ranking from Asana). Bee Propolis is the baseline.
PORTFOLIO_SKUS = [
    {"rank": 1, "code": "UK-D93-#1", "name": "Bee Propolis Spray 30ml x2",
     "pack": "30ml x 2", "sp": 15.99, "cost": 7.29, "units": 17000, "promo": 25000,
     "baseline": True,
     "rationale": "Lead pitch — strong GP%, supported by demo + TPD."},
    {"rank": 2, "code": "UK-D93-#2", "name": "Mag 8-in-1 90's/120's",
     "pack": "", "sp": None, "cost": None, "units": None, "promo": None,
     "baseline": False, "rationale": ""},
    {"rank": 3, "code": "UK-D93-#3", "name": "O1 450g",
     "pack": "", "sp": None, "cost": None, "units": None, "promo": None,
     "baseline": False, "rationale": ""},
    {"rank": 4, "code": "UK-D93-#4", "name": "Belli Bliss Raspberry, 450g",
     "pack": "", "sp": None, "cost": None, "units": None, "promo": None,
     "baseline": False, "rationale": ""},
    {"rank": 5, "code": "UK-D93-#5", "name": "Daily Boost",
     "pack": "", "sp": None, "cost": None, "units": None, "promo": None,
     "baseline": False, "rationale": ""},
    {"rank": 6, "code": "UK-D93-#6", "name": "Berberine 90's/120's",
     "pack": "", "sp": None, "cost": None, "units": None, "promo": None,
     "baseline": False, "rationale": ""},
    {"rank": 7, "code": "UK-D93-#7", "name": "L-theanine Capsules 90's/120's",
     "pack": "", "sp": None, "cost": None, "units": None, "promo": None,
     "baseline": False, "rationale": ""},
]

# -----------------------------------------------------------------------------
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Default font for the whole workbook
def set_default_font(ws):
    """Apply Tahoma 9 to default cell style (rendered Excel will still
    follow per-cell fonts, but this prevents Calibri showing on empty cells)."""
    pass  # per-cell fonts handle this; openpyxl can't easily set workbook default


def write_cell(ws, row, col, value, *, font=None, fill=None, align=None,
               border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font is not None: c.font = font
    if fill is not None: c.fill = fill
    if align is not None: c.alignment = align
    if border is not None: c.border = border
    if number_format is not None: c.number_format = number_format
    return c


def section_header(ws, row, text, span_cols, underline=True):
    start, end = span_cols
    if isinstance(start, str): start = openpyxl.utils.column_index_from_string(start)
    if isinstance(end, str):   end   = openpyxl.utils.column_index_from_string(end)
    ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
    c = ws.cell(row=row, column=start, value=text.upper())
    c.font = f_section
    c.alignment = align_left
    if underline:
        for col_i in range(start, end + 1):
            cc = ws.cell(row=row, column=col_i)
            cc.border = Border(bottom=thin_black)
    ws.row_dimensions[row].height = 18


# =============================================================================
# README
# =============================================================================
rd = wb.create_sheet("README")
rd.sheet_view.showGridLines = False
rd.column_dimensions["A"].width = 3
rd.column_dimensions["B"].width = 110

write_cell(rd, 2, 2, "Costco UK — D93 Portfolio ROI Model",
           font=F(size=16, bold=True), align=align_left)
write_cell(rd, 3, 2, "v4.0  ·  Bee Propolis market test (baseline) + 1-year portfolio roll-up",
           font=f_subtitle, align=align_left)

section_header(rd, 5, "How to use", ("B", "B"))

readme_lines = [
    "",
    "1.  This workbook does two things side-by-side:",
    "      (a) the deep-dive market-test ROI for Bee Propolis (the #1-ranked pitch SKU),",
    "      (b) a one-year portfolio view across all seven D93 SKUs, with blended margin and ROI.",
    "",
    "2.  Change only the cells with BLUE text. Black text is calculated. Green text is a cross-sheet link.",
    "",
    "3.  Bee Propolis detail lives on the Inputs, Forecast, and Summary sheets. The detail model uses",
    "    a 26-week test horizon, 20 warehouses, and three scenarios (BEST / IDEAL / WORST).",
    "",
    "4.  Annual portfolio numbers live on the Portfolio sheet. Each SKU is one row. Bee Propolis row 1",
    "    is pre-filled at indicative annual numbers; rows 2-7 are blank — paste in COGS, selling price,",
    "    and forecast units from Asana / commercial. The blended row at the bottom does the math.",
    "",
    "5.  Override cells. On the Inputs sheet, every calculated input has an OVERRIDE column next to it.",
    "    Leave it blank to use the formula; type a value to force the model to use yours.",
    "",
    "6.  Dates. Change the 'Test Start Date (W1)' on Inputs and every week header on every sheet shifts.",
    "",
]
r = 6
for line in readme_lines:
    write_cell(rd, r, 2, line, font=f_label, align=align_top)
    r += 1

section_header(rd, r, "Key concepts", ("B", "B")); r += 1
concepts = [
    ("Base Demand — organic Costco velocity with no promo. From the prior 30ml SKU run-rate: "
     "~500 lifetime units across a 29-week reference cycle, scaled to 20 warehouses. ≈ 13.8 units / wh / week."),
    ("Lift Multiplier — sales multiplier on a given week. 1.0x = base, 3.0x = demo, 3.6x = demo + end cap, "
     "8.0x+ = clearance markdown."),
    ("TPD — Temporary Price Discount. 20% off MSRP, which equals a 25% deduction on net (Costco's 20% "
     "retail margin grosses it up: 0.20 / 0.80 = 0.25)."),
    ("Demo cost — £199 / warehouse / weekend, grossed up by 1.86x for staffing & POS."),
    ("End cap — £850 / warehouse / 2-week cycle, same 1.86x factor. Per-week is half."),
    ("ROI on Promo Spend — Net Profit ÷ Promo Spend. Tells you what £1 of promo investment returned."),
    ("Blended Margin / ROI (Portfolio) — totals are summed across SKUs; blended % = total GP ÷ total Revenue. "
     "This is the number that matters for the overall pitch — a low-margin SKU is fine if higher-margin "
     "SKUs in the portfolio more than offset it."),
]
for c in concepts:
    write_cell(rd, r, 2, "•  " + c, font=f_label, align=align_top)
    rd.row_dimensions[r].height = 36
    r += 1

r += 1
section_header(rd, r, "Colour key", ("B", "B")); r += 1
for label, font, fill, desc in [
    ("Blue text",  f_input, fill_input, "Editable input. Type your own number."),
    ("Black text", f_calc,  None,       "Calculated. Don't edit — formula will rebuild."),
    ("Green text", f_link,  None,       "Cross-sheet link. Pulled from another tab."),
]:
    c = write_cell(rd, r, 2, f"  {label} — {desc}", font=font, fill=fill, align=align_left)
    r += 1


# =============================================================================
# INPUTS — Bee Propolis baseline
# =============================================================================
inp = wb.create_sheet("Inputs")
inp.sheet_view.showGridLines = False

inp.column_dimensions["A"].width = 3
inp.column_dimensions["B"].width = 38
inp.column_dimensions["C"].width = 14
inp.column_dimensions["D"].width = 14
inp.column_dimensions["E"].width = 78

write_cell(inp, 2, 2, "Inputs & Assumptions — Bee Propolis 30ml x 2",
           font=f_title, align=align_left)
write_cell(inp, 3, 2, "Edit blue cells only. Optional overrides in column D.",
           font=f_subtitle, align=align_left)

# Column header strip
for col, txt, align in [(2, "Field", align_left), (3, "Value", align_center),
                        (4, "Override", align_center), (5, "Notes", align_left)]:
    write_cell(inp, 5, col, txt, font=f_colhdr, fill=fill_hdr, align=align,
               border=Border(bottom=thin_black))


def row_input(ws, row, label, value, note, fmt=None):
    write_cell(ws, row, 2, label, font=f_label, align=align_left)
    c = write_cell(ws, row, 3, value, font=f_input, fill=fill_input,
                   align=align_right, border=Border(bottom=hair_grey),
                   number_format=fmt)
    write_cell(ws, row, 5, note, font=f_note, align=align_top)
    ws.row_dimensions[row].height = 18
    return c


def row_calc(ws, row, label, formula, note, fmt=None, allow_override=True):
    write_cell(ws, row, 2, label, font=f_label, align=align_left)
    d_letter = "D"
    if allow_override:
        wrapped = f'=IF({d_letter}{row}="",{formula},{d_letter}{row})'
    else:
        wrapped = f"={formula}"
    c = write_cell(ws, row, 3, wrapped, font=f_calc,
                   align=align_right, border=Border(bottom=hair_grey),
                   number_format=fmt)
    if allow_override:
        write_cell(ws, row, 4, None, font=f_input, fill=fill_input,
                   align=align_right, border=Border(bottom=hair_grey),
                   number_format=fmt)
    write_cell(ws, row, 5, note, font=f_note, align=align_top)
    ws.row_dimensions[row].height = 18
    return c


# Section 1 — Product & Distribution
section_header(inp, 7, "1.  Product & Distribution", ("B", "E"))
row_input(inp, 8, "SKU", "Bee Propolis Spray 30ml x 2", "Costco UK market test SKU.")
row_input(inp, 9, "Test Start Date (W1)", DEFAULT_START_DATE,
          "Week-commencing date for W1. Drives every date header.", fmt=FMT_DATE_L)
row_input(inp, 10, "Test Warehouses", 20, "Number of Costco UK warehouses in test.", fmt=FMT_NUM)
row_input(inp, 11, "Test Period (Weeks)", WEEKS,
          f"Forecast horizon ({WEEKS} weeks, W1 to W{WEEKS}).", fmt=FMT_NUM)

# Section 2 — Pricing & Cost
section_header(inp, 13, "2.  Pricing & Cost (per unit, £)", ("B", "E"))
row_input(inp, 14, "Selling Price (MSRP)", 15.99, "Retail per pack.", fmt=FMT_GBP_D)
row_input(inp, 15, "Cost Price (landed)", 7.29, "Brand cost delivered.", fmt=FMT_GBP_D)
row_calc (inp, 16, "Gross Margin / unit", "C14-C15",
          "SP less landed cost. Override is display-only.", fmt=FMT_GBP_D)
row_calc (inp, 17, "Gross Margin %", "IFERROR((C14-C15)/C14,0)",
          "GP ÷ SP. Override is display-only.", fmt=FMT_PCT)

# Section 3 — Base Demand
section_header(inp, 19, "3.  Base Demand (no promo)", ("B", "E"))
row_input(inp, 20, "Units / Warehouse / Week", 13.79,
          "From prior 30ml SKU run-rate (500 × 0.8 / 29).", fmt="#,##0.00")
row_calc (inp, 21, "Base Units / Week (all warehouses)", "C20*C10",
          "Per-wh × warehouses. Override flows into Forecast.", fmt=FMT_NUM)

# Section 4 — Promo Economics
section_header(inp, 23, "4.  Promo Economics", ("B", "E"))
row_input(inp, 24, "Demo £ / warehouse / week", 199, "Costco standard demo rate.", fmt=FMT_GBP)
row_input(inp, 25, "Demo cost factor", 1.86, "Staffing + POS overhead multiplier.", fmt=FMT_MULT)
row_calc (inp, 26, "Demo Total £ / week", "C24*C10*C25",
          "Rate × warehouses × factor. Override sets the all-wh demo cost directly.", fmt=FMT_GBP)
row_input(inp, 27, "End Cap £ / warehouse / 2-week cycle", 850, "Costco standard end cap rate.", fmt=FMT_GBP)
row_input(inp, 28, "End Cap cost factor", 1.86, "Same as demo.", fmt=FMT_MULT)
row_calc (inp, 29, "End Cap Total £ / week", "C27*C10*C28/2",
          "Rate × warehouses × factor ÷ 2. Override sets it directly.", fmt=FMT_GBP)
row_input(inp, 30, "Fence £ / week (when active)", 0, "Not budgeted.", fmt=FMT_GBP)
row_input(inp, 31, "Advertising £ / week (when active)", 0, "Not budgeted.", fmt=FMT_GBP)
row_input(inp, 32, "TPD discount % (of net)", 0.25,
          "20% off MSRP = 25% deduction on net (Costco's 20% retail margin).", fmt=FMT_PCT)
row_input(inp, 33, "Markdown discount %", 0.50, "Clearance pricing.", fmt=FMT_PCT)

# Named ranges (real Excel-modeller convention)
wb.defined_names["StartDate"]   = DefinedName("StartDate",   attr_text="Inputs!$C$9")
wb.defined_names["WHs"]         = DefinedName("WHs",         attr_text="Inputs!$C$10")
wb.defined_names["Period"]      = DefinedName("Period",      attr_text="Inputs!$C$11")
wb.defined_names["Price"]       = DefinedName("Price",       attr_text="Inputs!$C$14")
wb.defined_names["Cost"]        = DefinedName("Cost",        attr_text="Inputs!$C$15")
wb.defined_names["BaseUnits"]   = DefinedName("BaseUnits",   attr_text="Inputs!$C$21")
wb.defined_names["DemoCost"]    = DefinedName("DemoCost",    attr_text="Inputs!$C$26")
wb.defined_names["EndCapCost"]  = DefinedName("EndCapCost",  attr_text="Inputs!$C$29")
wb.defined_names["FenceCost"]   = DefinedName("FenceCost",   attr_text="Inputs!$C$30")
wb.defined_names["AdvertCost"]  = DefinedName("AdvertCost",  attr_text="Inputs!$C$31")
wb.defined_names["TPDPct"]      = DefinedName("TPDPct",      attr_text="Inputs!$C$32")
wb.defined_names["MdPct"]       = DefinedName("MdPct",       attr_text="Inputs!$C$33")

# --- Scenario grids ----------------------------------------------------------
section_header(inp, 35, "5.  Scenario activity & demand lift", ("B", "E"))
write_cell(inp, 36, 2,
           "Toggle 1 = active, 0 = off per week. Lift is the demand multiplier vs base. "
           "Dates auto-fill from Test Start Date.",
           font=f_note, align=align_top)

WEEK_START_COL = 6
WEEK_END_COL = WEEK_START_COL + WEEKS - 1
for col in range(WEEK_START_COL, WEEK_END_COL + 1):
    inp.column_dimensions[get_column_letter(col)].width = 7.2

scenario_grid_start = {}
current = 38
for sc in SCENARIOS:
    write_cell(inp, current, 2, sc["name"], font=f_label_bold, align=align_left)
    inp.row_dimensions[current].height = 16
    current += 1

    write_cell(inp, current, 2, f"Rationale: {sc['rationale']}", font=f_note, align=align_top)
    inp.merge_cells(start_row=current, start_column=2, end_row=current, end_column=WEEK_END_COL)
    inp.row_dimensions[current].height = 24
    current += 1

    # Week header
    write_cell(inp, current, 2, "Week", font=f_colhdr, fill=fill_hdr, align=align_left,
               border=Border(top=thin_black, bottom=hair_grey))
    for i in range(WEEKS):
        write_cell(inp, current, WEEK_START_COL + i, f"W{FIRST_WEEK_NUM + i}",
                   font=f_colhdr, fill=fill_hdr, align=align_center,
                   border=Border(top=thin_black, bottom=hair_grey))
    current += 1

    # Date sub-header
    write_cell(inp, current, 2, "W/c", font=f_note, fill=fill_hdr, align=align_left,
               border=Border(bottom=thin_black))
    for i in range(WEEKS):
        write_cell(inp, current, WEEK_START_COL + i, f"=StartDate+7*{i}",
                   font=f_note, fill=fill_hdr, align=align_center,
                   border=Border(bottom=thin_black), number_format=FMT_DATE)
    current += 1

    scenario_grid_start[sc["key"]] = current

    for promo in PROMO_TYPES:
        write_cell(inp, current, 2, promo, font=f_label, align=align_left)
        active_weeks = set(sc["promo"].get(promo, []))
        for i in range(WEEKS):
            wknum = FIRST_WEEK_NUM + i
            val = 1 if wknum in active_weeks else 0
            write_cell(inp, current, WEEK_START_COL + i, val,
                       font=f_input, fill=fill_input, align=align_center,
                       border=Border(bottom=hair_grey),
                       number_format="0;;;@")
        current += 1

    write_cell(inp, current, 2, "Lift multiplier (x base)", font=f_label_bold, align=align_left)
    for i in range(WEEKS):
        wknum = FIRST_WEEK_NUM + i
        val = sc["lift_overrides"].get(wknum, 1.0)
        write_cell(inp, current, WEEK_START_COL + i, val,
                   font=f_input, fill=fill_input, align=align_center,
                   border=Border(top=thin_black, bottom=thin_black),
                   number_format=FMT_MULT)
    current += 1
    current += 1

inp.freeze_panes = "F7"


# =============================================================================
# FORECAST
# =============================================================================
fc = wb.create_sheet("Forecast")
fc.sheet_view.showGridLines = False
fc.column_dimensions["A"].width = 3
fc.column_dimensions["B"].width = 34
for col in range(WEEK_START_COL, WEEK_END_COL + 1):
    fc.column_dimensions[get_column_letter(col)].width = 9.5
TOTAL_COL = WEEK_END_COL + 1
fc.column_dimensions[get_column_letter(TOTAL_COL)].width = 13

write_cell(fc, 2, 2, "Forecast — Weekly P&L by Scenario", font=f_title, align=align_left)
write_cell(fc, 3, 2, "All values calculated from Inputs. Don't edit.",
           font=f_subtitle, align=align_left)

PNL_ROWS = [
    ("POS Units",                 "units"),
    ("  Selling Price",           "price"),
    ("Revenue",                   "rev"),
    ("  Cost Price",              "cost"),
    ("Raw COGS",                  "cogs"),
    ("  Demo Spend",              "demo"),
    ("  End Cap Spend",           "endcap"),
    ("  Fence Spend",             "fence"),
    ("  TPD Spend",               "tpd"),
    ("  Markdown Spend",          "markdown"),
    ("  Advertising Spend",       "advert"),
    ("Total Promo Spend",         "promo_total"),
    ("Gross Profit",              "gp"),
    ("Net Profit (after promo)",  "net"),
]
PROMO_ROW_OFFSET = {p: i for i, p in enumerate(PROMO_TYPES)}
LIFT_OFFSET = len(PROMO_TYPES)

fc_row = 5
scenario_summary_rows = {}

for sc in SCENARIOS:
    write_cell(fc, fc_row, 2, sc["name"], font=f_label_bold, align=align_left,
               border=Border(bottom=thin_black))
    fc.merge_cells(start_row=fc_row, start_column=2, end_row=fc_row, end_column=TOTAL_COL)
    for col_i in range(2, TOTAL_COL + 1):
        fc.cell(row=fc_row, column=col_i).border = Border(bottom=thin_black)
    fc_row += 1

    # Week header
    write_cell(fc, fc_row, 2, "Week", font=f_colhdr, fill=fill_hdr,
               align=align_left, border=Border(bottom=hair_grey))
    for i in range(WEEKS):
        write_cell(fc, fc_row, WEEK_START_COL + i, f"W{FIRST_WEEK_NUM + i}",
                   font=f_colhdr, fill=fill_hdr, align=align_center,
                   border=Border(bottom=hair_grey))
    write_cell(fc, fc_row, TOTAL_COL, "TOTAL", font=f_colhdr, fill=fill_hdr,
               align=align_center, border=Border(bottom=hair_grey))
    fc_row += 1

    # Date sub-header
    write_cell(fc, fc_row, 2, "W/c", font=f_note, fill=fill_hdr, align=align_left,
               border=Border(bottom=thin_black))
    for i in range(WEEKS):
        write_cell(fc, fc_row, WEEK_START_COL + i, f"=StartDate+7*{i}",
                   font=f_note, fill=fill_hdr, align=align_center,
                   border=Border(bottom=thin_black), number_format=FMT_DATE)
    write_cell(fc, fc_row, TOTAL_COL, "", font=f_note, fill=fill_hdr,
               border=Border(bottom=thin_black))
    fc_row += 1

    scenario_pnl_rows = {}
    grid_top = scenario_grid_start[sc["key"]]
    promo_rows_on_inp = {p: grid_top + PROMO_ROW_OFFSET[p] for p in PROMO_TYPES}
    lift_row_on_inp = grid_top + LIFT_OFFSET

    for label, key in PNL_ROWS:
        scenario_pnl_rows[key] = fc_row
        bold = key in ("rev", "cogs", "promo_total", "gp", "net")
        lbl_font = f_label_bold if bold else f_label
        write_cell(fc, fc_row, 2, label, font=lbl_font, align=align_left)

        # Top/bottom borders for total rows
        border_total = Border(top=thin_black, bottom=thin_black) if key == "net" else \
                       Border(top=thin_black) if key in ("promo_total", "gp") else None

        for i in range(WEEKS):
            week_col = WEEK_START_COL + i
            wcol = get_column_letter(week_col)
            if key == "units":
                formula = f"=ROUND(BaseUnits*Inputs!{wcol}{lift_row_on_inp},0)"
            elif key == "price":
                formula = f"=Price"
            elif key == "rev":
                formula = f"={wcol}{scenario_pnl_rows['units']}*{wcol}{scenario_pnl_rows['price']}"
            elif key == "cost":
                formula = f"=Cost"
            elif key == "cogs":
                formula = f"={wcol}{scenario_pnl_rows['units']}*{wcol}{scenario_pnl_rows['cost']}"
            elif key == "demo":
                formula = f"=Inputs!{wcol}{promo_rows_on_inp['Demo']}*DemoCost"
            elif key == "endcap":
                formula = f"=Inputs!{wcol}{promo_rows_on_inp['End Cap']}*EndCapCost"
            elif key == "fence":
                formula = f"=Inputs!{wcol}{promo_rows_on_inp['Fence']}*FenceCost"
            elif key == "tpd":
                formula = (f"=Inputs!{wcol}{promo_rows_on_inp['TPD']}"
                           f"*{wcol}{scenario_pnl_rows['units']}"
                           f"*{wcol}{scenario_pnl_rows['price']}*TPDPct")
            elif key == "markdown":
                formula = (f"=Inputs!{wcol}{promo_rows_on_inp['Markdown']}"
                           f"*{wcol}{scenario_pnl_rows['units']}"
                           f"*{wcol}{scenario_pnl_rows['price']}*MdPct")
            elif key == "advert":
                formula = f"=Inputs!{wcol}{promo_rows_on_inp['Advert']}*AdvertCost"
            elif key == "promo_total":
                formula = (f"=SUM({wcol}{scenario_pnl_rows['demo']}:"
                           f"{wcol}{scenario_pnl_rows['advert']})")
            elif key == "gp":
                formula = f"={wcol}{scenario_pnl_rows['rev']}-{wcol}{scenario_pnl_rows['cogs']}"
            elif key == "net":
                formula = f"={wcol}{scenario_pnl_rows['gp']}-{wcol}{scenario_pnl_rows['promo_total']}"
            else:
                formula = ""
            num_fmt = (FMT_NUM if key == "units"
                       else FMT_GBP_D if key in ("price", "cost")
                       else FMT_GBP)
            write_cell(fc, fc_row, week_col, formula,
                       font=f_total if bold else f_calc,
                       align=align_right,
                       border=border_total,
                       number_format=num_fmt)

        first_w = get_column_letter(WEEK_START_COL)
        last_w = get_column_letter(WEEK_END_COL)
        if key in ("price", "cost"):
            total_formula = "=Price" if key == "price" else "=Cost"
        else:
            total_formula = f"=SUM({first_w}{fc_row}:{last_w}{fc_row})"
        num_fmt = (FMT_NUM if key == "units"
                   else FMT_GBP_D if key in ("price", "cost")
                   else FMT_GBP)
        write_cell(fc, fc_row, TOTAL_COL, total_formula,
                   font=f_total, align=align_right,
                   border=border_total or Border(left=thin_black),
                   number_format=num_fmt)
        fc_row += 1

    scenario_summary_rows[sc["key"]] = scenario_pnl_rows
    fc_row += 2

fc.freeze_panes = "C5"


# =============================================================================
# SUMMARY — Bee Propolis scenario comparison
# =============================================================================
sm = wb.create_sheet("Summary")
sm.sheet_view.showGridLines = False
sm.column_dimensions["A"].width = 3
sm.column_dimensions["B"].width = 32
for i in range(len(SCENARIOS)):
    sm.column_dimensions[get_column_letter(3 + i)].width = 18

write_cell(sm, 2, 2, "Summary — Bee Propolis Scenario Comparison", font=f_title, align=align_left)
write_cell(sm, 3, 2,
           f'="Test window: W1 ("&TEXT(StartDate,"d-mmm-yyyy")&") through W"&Period&" ("&TEXT(StartDate+7*(Period-1),"d-mmm-yyyy")&")"',
           font=f_subtitle, align=align_left)

# Header
hr = 5
write_cell(sm, hr, 2, "Metric", font=f_colhdr, fill=fill_hdr, align=align_left,
           border=Border(bottom=thin_black))
for i, sc in enumerate(SCENARIOS):
    write_cell(sm, hr, 3 + i, sc["key"], font=f_colhdr, fill=fill_hdr,
               align=align_center, border=Border(bottom=thin_black))

TC = get_column_letter(TOTAL_COL)
KPIS = [
    ("Units Sold",          "units",       FMT_NUM, False),
    ("Revenue",             "rev",         FMT_GBP, False),
    ("Raw COGS",            "cogs",        FMT_GBP, False),
    ("Gross Profit",        "gp",          FMT_GBP, True),
    ("Gross Margin %",      "gp_pct",      FMT_PCT, True),
    ("Demo Spend",          "demo",        FMT_GBP, False),
    ("End Cap Spend",       "endcap",      FMT_GBP, False),
    ("Fence Spend",         "fence",       FMT_GBP, False),
    ("TPD Spend",           "tpd",         FMT_GBP, False),
    ("Markdown Spend",      "markdown",    FMT_GBP, False),
    ("Advertising Spend",   "advert",      FMT_GBP, False),
    ("Total Promo Spend",   "promo_total", FMT_GBP, True),
    ("Net Profit",          "net",         FMT_GBP, True),
    ("Net Margin %",        "net_pct",     FMT_PCT, True),
    ("ROI on Promo Spend",  "roi",         FMT_PCT, True),
    ("Avg Units / WH / Wk", "uphpw",       FMT_NUM, False),
]

sr = hr + 1
sum_metric_rows = {}  # for portfolio cross-reference
for label, key, fmt, bold in KPIS:
    sum_metric_rows[key] = sr
    write_cell(sm, sr, 2, label,
               font=f_total if bold else f_label, align=align_left,
               border=Border(bottom=hair_grey) if not bold else Border(top=thin_black, bottom=thin_black) if key == "roi" else Border(top=thin_black))
    for i, sc in enumerate(SCENARIOS):
        rows = scenario_summary_rows[sc["key"]]
        if key in rows:
            formula = f"=Forecast!{TC}{rows[key]}"
        elif key == "gp_pct":
            formula = f"=IFERROR(Forecast!{TC}{rows['gp']}/Forecast!{TC}{rows['rev']},0)"
        elif key == "net_pct":
            formula = f"=IFERROR(Forecast!{TC}{rows['net']}/Forecast!{TC}{rows['rev']},0)"
        elif key == "roi":
            formula = f"=IFERROR(Forecast!{TC}{rows['net']}/Forecast!{TC}{rows['promo_total']},0)"
        elif key == "uphpw":
            formula = f"=IFERROR(Forecast!{TC}{rows['units']}/WHs/Period,0)"
        else:
            formula = ""
        write_cell(sm, sr, 3 + i, formula,
                   font=f_total if bold else f_link,
                   align=align_right, number_format=fmt,
                   border=Border(bottom=hair_grey) if not bold else Border(top=thin_black, bottom=thin_black) if key == "roi" else Border(top=thin_black))
    sr += 1

# Conditional formatting on Net Profit + ROI rows
last_col = get_column_letter(2 + len(SCENARIOS))
for k in ("net", "roi"):
    r = sum_metric_rows[k]
    sm.conditional_formatting.add(
        f"C{r}:{last_col}{r}",
        CellIsRule(operator="lessThan", formula=["0"], font=F(size=9, bold=True, color="9C0006")))

# Scenario rationale block
sr += 2
section_header(sm, sr, "Scenario rationale", ("B", get_column_letter(2 + len(SCENARIOS))))
sr += 1
for sc in SCENARIOS:
    write_cell(sm, sr, 2, sc["key"], font=f_label_bold, align=align_top)
    sm.merge_cells(start_row=sr, start_column=3, end_row=sr, end_column=2 + len(SCENARIOS))
    write_cell(sm, sr, 3, sc["rationale"], font=f_label, align=align_top)
    sm.row_dimensions[sr].height = 30
    sr += 1

sm.freeze_panes = "C6"


# =============================================================================
# PORTFOLIO — annual roll-up across SKUs
# =============================================================================
pf = wb.create_sheet("Portfolio")
pf.sheet_view.showGridLines = False

# Layout: 1 row per SKU + 1 total row
# Columns: A spacer, B Rank, C Code, D SKU Name, E Pack, F SP, G Cost,
#          H GM/u (calc), I GM% (calc), J Annual Units, K Revenue (calc),
#          L COGS (calc), M GP (calc), N Promo Spend, O Net Profit (calc),
#          P Net % (calc), Q ROI on Promo (calc), R Notes/Rationale
pf.column_dimensions["A"].width = 3
widths = {"B": 5, "C": 12, "D": 32, "E": 12, "F": 9, "G": 9, "H": 9, "I": 8,
          "J": 11, "K": 13, "L": 13, "M": 13, "N": 12, "O": 13, "P": 8, "Q": 8, "R": 40}
for col, w in widths.items():
    pf.column_dimensions[col].width = w

write_cell(pf, 2, 2, "Portfolio — D93 Pitch SKUs (Annual Y1 Roll-Up)",
           font=f_title, align=align_left)
write_cell(pf, 3, 2,
           "Bee Propolis pre-filled at indicative annual numbers. Other rows: paste in SP, "
           "Cost, Y1 Units, and Promo Spend from Asana. Bottom row blends across SKUs.",
           font=f_subtitle, align=align_left)
pf.merge_cells("B3:R3")

# Column headers
hdrs = [
    ("B", "#"), ("C", "Code"), ("D", "SKU"), ("E", "Pack"),
    ("F", "SP £"), ("G", "Cost £"), ("H", "GM/u £"), ("I", "GM %"),
    ("J", "Y1 Units"), ("K", "Revenue £"), ("L", "COGS £"), ("M", "GP £"),
    ("N", "Promo £"), ("O", "Net Profit £"), ("P", "Net %"),
    ("Q", "ROI %"), ("R", "Rationale / status"),
]
HDR_ROW = 5
for col_letter, txt in hdrs:
    align = align_left if col_letter in ("C", "D", "E", "R") else align_center
    write_cell(pf, HDR_ROW, openpyxl.utils.column_index_from_string(col_letter), txt,
               font=f_colhdr, fill=fill_hdr, align=align,
               border=Border(top=thin_black, bottom=thin_black))

# SKU rows
sku_start_row = HDR_ROW + 1
last_sku_row = sku_start_row + len(PORTFOLIO_SKUS) - 1

for idx, sku in enumerate(PORTFOLIO_SKUS):
    r = sku_start_row + idx

    # Direct inputs
    write_cell(pf, r, 2, sku["rank"], font=f_label, align=align_center, number_format=FMT_NUM)
    write_cell(pf, r, 3, sku["code"], font=f_label, align=align_left)
    write_cell(pf, r, 4, sku["name"], font=f_label_bold, align=align_left)
    write_cell(pf, r, 5, sku["pack"] or None, font=f_input, fill=fill_input, align=align_left)
    write_cell(pf, r, 6, sku["sp"],   font=f_input, fill=fill_input, align=align_right, number_format=FMT_GBP_D)
    write_cell(pf, r, 7, sku["cost"], font=f_input, fill=fill_input, align=align_right, number_format=FMT_GBP_D)

    # Calc cells
    write_cell(pf, r, 8, f"=IFERROR(F{r}-G{r},0)",  font=f_calc, align=align_right, number_format=FMT_GBP_D)
    write_cell(pf, r, 9, f"=IFERROR((F{r}-G{r})/F{r},0)", font=f_calc, align=align_right, number_format=FMT_PCT)

    # More inputs
    write_cell(pf, r, 10, sku["units"], font=f_input, fill=fill_input, align=align_right, number_format=FMT_NUM)

    # Calcs
    write_cell(pf, r, 11, f"=IFERROR(J{r}*F{r},0)", font=f_calc, align=align_right, number_format=FMT_GBP)
    write_cell(pf, r, 12, f"=IFERROR(J{r}*G{r},0)", font=f_calc, align=align_right, number_format=FMT_GBP)
    write_cell(pf, r, 13, f"=IFERROR(K{r}-L{r},0)", font=f_total, align=align_right, number_format=FMT_GBP)

    # Promo input
    write_cell(pf, r, 14, sku["promo"], font=f_input, fill=fill_input, align=align_right, number_format=FMT_GBP)

    # Net & ratios
    write_cell(pf, r, 15, f"=IFERROR(M{r}-N{r},0)", font=f_total, align=align_right, number_format=FMT_GBP)
    write_cell(pf, r, 16, f"=IFERROR(O{r}/K{r},0)", font=f_calc, align=align_right, number_format=FMT_PCT)
    write_cell(pf, r, 17, f"=IFERROR(O{r}/N{r},0)", font=f_calc, align=align_right, number_format=FMT_PCT)

    write_cell(pf, r, 18, sku["rationale"], font=f_input if not sku["baseline"] else f_label,
               fill=fill_input if not sku["baseline"] else None, align=align_left)

    pf.row_dimensions[r].height = 18

# Total / blended row
tr = last_sku_row + 1
write_cell(pf, tr, 2, "", border=Border(top=thin_black, bottom=thin_black))
write_cell(pf, tr, 3, "", border=Border(top=thin_black, bottom=thin_black))
write_cell(pf, tr, 4, "PORTFOLIO TOTAL / BLENDED", font=f_total, align=align_left,
           border=Border(top=thin_black, bottom=thin_black))
for col_i in range(5, 19):
    pf.cell(row=tr, column=col_i).border = Border(top=thin_black, bottom=thin_black)

# Blank for SP/Cost/GM/u (can't average meaningfully)
write_cell(pf, tr, 5, "", border=Border(top=thin_black, bottom=thin_black))
write_cell(pf, tr, 6, "", border=Border(top=thin_black, bottom=thin_black))
write_cell(pf, tr, 7, "", border=Border(top=thin_black, bottom=thin_black))
write_cell(pf, tr, 8, "", border=Border(top=thin_black, bottom=thin_black))

# Blended GM% = SUM(GP) / SUM(Revenue)
write_cell(pf, tr, 9, f"=IFERROR(SUM(M{sku_start_row}:M{last_sku_row})/SUM(K{sku_start_row}:K{last_sku_row}),0)",
           font=f_total, align=align_right, number_format=FMT_PCT,
           border=Border(top=thin_black, bottom=thin_black))

# Sums
for col_letter in ("J", "K", "L", "M", "N", "O"):
    write_cell(pf, tr, openpyxl.utils.column_index_from_string(col_letter),
               f"=SUM({col_letter}{sku_start_row}:{col_letter}{last_sku_row})",
               font=f_total, align=align_right,
               number_format=FMT_NUM if col_letter == "J" else FMT_GBP,
               border=Border(top=thin_black, bottom=thin_black))

# Blended Net% and ROI%
write_cell(pf, tr, 16, f"=IFERROR(O{tr}/K{tr},0)", font=f_total, align=align_right,
           number_format=FMT_PCT, border=Border(top=thin_black, bottom=thin_black))
write_cell(pf, tr, 17, f"=IFERROR(O{tr}/N{tr},0)", font=f_total, align=align_right,
           number_format=FMT_PCT, border=Border(top=thin_black, bottom=thin_black))
write_cell(pf, tr, 18, "Blended across all SKUs",
           font=f_note, align=align_left,
           border=Border(top=thin_black, bottom=thin_black))

pf.row_dimensions[tr].height = 22

# Note below
nr = tr + 2
write_cell(pf, nr, 2,
           "Note: GM% and Net% on the total row are weighted by revenue (Σ GP / Σ Rev). "
           "ROI is Σ Net Profit / Σ Promo Spend.  Blue cells are inputs — fill in from Asana for rows 2-7.",
           font=f_note, align=align_top)
pf.merge_cells(start_row=nr, start_column=2, end_row=nr, end_column=18)
pf.row_dimensions[nr].height = 28

pf.freeze_panes = "B6"


# =============================================================================
out = "/home/user/my-first-project/Costco_UK_D93_Portfolio_ROI_v4.xlsx"
wb.save(out)
print(f"Saved: {out}")
