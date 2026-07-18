#!/usr/bin/env python3
"""Iteration 1 content: 'Scenario Comparison' tab — all three scenarios side-by-side
with computed per-can economics, 3-yr P&L rollup, EBITDA and break-even, plus a chart."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.properties import PageSetupProperties
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; TEAL="2E7D8A"; STEEL="3E5C76"; LIGHT="EAF1F4"; GREEN="E2EFDA"; GREENHEAD="1E7D32"; AMBER="FFF2CC"; AMBERHEAD="B45309"; GOLD="C9A227"; WHITE="FFFFFF"; GREY="595959"; LINKBLUE="1155CC"
def F(**k): return Font(name="Calibri",**k)
HEAD=F(bold=True,color=WHITE,size=10); TITLE=F(bold=True,color=WHITE,size=18); BODY=F(size=10); BODYB=F(size=10,bold=True); SMALL=F(size=9,italic=True,color=GREY)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top"); CTR=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)
def fill(c): return PatternFill("solid",fgColor=c)

if "Scenario Comparison" in wb.sheetnames: del wb["Scenario Comparison"]

# --- scenario assumptions (mirror of Financial Model) ---
S={"Conservative":dict(price=3.29,cogs=1.20,ret=0.38,dist=0.18,trade=0.25,case=12,
      doors=[300,1200,3500],vel=[2.0,3.0,3.5],mkt=300000/2,fixed=250000,slot=50000),
   "Base":dict(price=3.49,cogs=1.05,ret=0.35,dist=0.15,trade=0.20,case=12,
      doors=[500,2000,6000],vel=[3.0,4.0,5.0],mkt=300000,fixed=350000,slot=100000),
   "Aggressive":dict(price=3.79,cogs=0.95,ret=0.33,dist=0.12,trade=0.15,case=12,
      doors=[900,3500,11000],vel=[4.5,6.0,7.5],mkt=600000,fixed=500000,slot=200000)}
# note: mkt for Conservative is 150000 (fix inline)
S["Conservative"]["mkt"]=150000

def econ(s):
    net=s["price"]*(1-s["ret"])*(1-s["dist"]); gp=net-s["cogs"]; trade=net*s["trade"]; cont=gp-trade
    yrs=[]
    for i in range(3):
        units=s["doors"][i]*s["vel"][i]*52
        nrev=units*net; cogs=units*s["cogs"]; gprof=nrev-cogs; trd=units*trade
        slot=s["slot"] if i==0 else 0
        ebitda=gprof-trd-s["mkt"]-s["fixed"]-slot
        yrs.append(dict(units=units,nrev=nrev,gprof=gprof,ebitda=ebitda))
    be_units=(s["mkt"]+s["fixed"]+s["slot"])/cont if cont>0 else float('inf')
    return dict(net=net,gp=gp,gm=gp/net,cont=cont,yrs=yrs,be=be_units)

E={k:econ(v) for k,v in S.items()}

ws=wb.create_sheet("Scenario Comparison")
ws.sheet_view.showGridLines=False
ws.merge_cells("A1:E1"); t=ws["A1"]; t.value="Scenario Comparison — Conservative · Base · Aggressive"; t.font=TITLE; t.fill=fill(NAVY); t.alignment=CTR; ws.row_dimensions[1].height=30
ws.merge_cells("A2:E2"); s=ws["A2"]; s.value="Full envelope of the Financial Model, computed for all three scenarios (edit source assumptions on ‘Financial Model’)"; s.font=F(bold=True,color=WHITE,size=10); s.fill=fill(TEAL); s.alignment=CTR; ws.row_dimensions[2].height=18

def row(r,label,vals,fmt,bold=False,bg=None):
    c=ws.cell(row=r,column=1,value=label); c.font=BODYB if bold else BODY; c.alignment=LEFT; c.border=BORDER
    if bg: c.fill=fill(bg)
    for j,v in enumerate(vals):
        cc=ws.cell(row=r,column=2+j,value=v); cc.font=BODYB if bold else BODY; cc.alignment=CTR; cc.border=BORDER; cc.number_format=fmt
        if bg: cc.fill=fill(bg)

# header
for j,h in enumerate(["Metric","Conservative","Base","Aggressive"]):
    c=ws.cell(row=4,column=1+j,value=h); c.font=HEAD; c.fill=fill(STEEL); c.alignment=CTR; c.border=BORDER
order=["Conservative","Base","Aggressive"]
r=5
row(r,"List price / can",[S[k]["price"] for k in order],'"$"#,##0.00'); r+=1
row(r,"COGS / can",[S[k]["cogs"] for k in order],'"$"#,##0.00'); r+=1
row(r,"Net revenue / can",[E[k]["net"] for k in order],'"$"#,##0.00'); r+=1
row(r,"Gross margin %",[E[k]["gm"] for k in order],'0.0%'); r+=1
row(r,"Contribution / can",[E[k]["cont"] for k in order],'"$"#,##0.00',bold=True,bg=LIGHT); r+=1
row(r,"Break-even units (Yr-1 load)",[E[k]["be"] for k in order],'#,##0'); r+=1
row(r,"Year-1 units",[E[k]["yrs"][0]["units"] for k in order],'#,##0'); r+=1
row(r,"Year-1 EBITDA",[E[k]["yrs"][0]["ebitda"] for k in order],'"$"#,##0',bold=True); r+=1
row(r,"Year-3 net revenue",[E[k]["yrs"][2]["nrev"] for k in order],'"$"#,##0'); r+=1
row(r,"Year-3 EBITDA",[E[k]["yrs"][2]["ebitda"] for k in order],'"$"#,##0',bold=True,bg=GREEN); r+=1
last=r-1
ws.column_dimensions["A"].width=28
for col in "BCDE": ws.column_dimensions[col].width=16

# chart: Year-3 net revenue & EBITDA by scenario — build a small data block then chart
dr=r+2
ws.cell(row=dr,column=1,value="(chart data)").font=SMALL
ws.cell(row=dr+1,column=1,value="Scenario"); ws.cell(row=dr+1,column=2,value="Y3 Net Rev"); ws.cell(row=dr+1,column=3,value="Y3 EBITDA")
for i,k in enumerate(order):
    ws.cell(row=dr+2+i,column=1,value=k)
    ws.cell(row=dr+2+i,column=2,value=round(E[k]["yrs"][2]["nrev"]))
    ws.cell(row=dr+2+i,column=3,value=round(E[k]["yrs"][2]["ebitda"]))
ch=BarChart(); ch.type="col"; ch.title="Year-3 Net Revenue vs EBITDA by Scenario (CAD)"; ch.height=8; ch.width=16
ch.add_data(Reference(ws,min_col=2,max_col=3,min_row=dr+1,max_row=dr+1+len(order)),titles_from_data=True)
ch.set_categories(Reference(ws,min_col=1,min_row=dr+2,max_row=dr+1+len(order)))
ch.dataLabels=DataLabelList(); ch.dataLabels.showVal=True
ws.add_chart(ch,"A"+str(last+2))
for rr in range(dr,dr+2+len(order)): ws.row_dimensions[rr].hidden=True

nc=last+20
ws.merge_cells(start_row=nc,start_column=1,end_row=nc,end_column=5)
c=ws.cell(row=nc,column=1,value=("Read: the Base case is the planning anchor; Conservative shows the downside if velocity/doors lag and costs run high; "
 "Aggressive shows the ceiling if the sparkling wedge + Costco beachhead over-deliver. All figures are illustrative — edit the yellow inputs on "
 "‘Financial Model’ (the active scenario there drives the interactive P&L; this tab always shows all three)."))
c.font=F(bold=True,color=GREENHEAD,size=10); c.fill=fill(GREEN); c.alignment=WRAP; c.border=BORDER; ws.row_dimensions[nc].height=52
link=ws.cell(row=1,column=14,value="↩ Contents"); link.hyperlink="#'01 Contents'!A1"; link.font=F(bold=True,color=LINKBLUE,underline="single",size=8); link.alignment=Alignment(horizontal="right")
ws.sheet_properties.tabColor=GOLD
ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0; ws.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)

# contents entry + place after Financial Model
toc=wb["01 Contents"]
r=toc.max_row+1
a=toc.cell(row=r,column=1,value="Scenario Comparison"); a.hyperlink="#'Scenario Comparison'!A1"; a.font=F(bold=True,color=LINKBLUE,underline="single",size=10); a.fill=fill(AMBER); a.border=BORDER; a.alignment=LEFT
b=toc.cell(row=r,column=2,value="All 3 scenarios side-by-side (per-can, EBITDA, break-even) + chart"); b.font=BODY; b.border=BORDER; b.alignment=WRAP; toc.row_dimensions[r].height=18
sc=wb["Scenario Comparison"]; wb._sheets.remove(sc)
idx=wb.sheetnames.index("Financial Model")+1
wb._sheets.insert(idx,sc)

wb.save(PATH)
print("Scenario Comparison added. sheets:",len(wb.sheetnames))
