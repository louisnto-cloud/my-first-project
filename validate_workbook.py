#!/usr/bin/env python3
"""Integrity validator for the enterprise workbook. Exits non-zero on failure.
Run after the build pipeline to catch corruption, missing tabs, or stale content."""
import sys
from openpyxl import load_workbook

PATH="Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
errors=[]; warns=[]

def check(cond, msg):
    if not cond: errors.append(msg)

try:
    wb=load_workbook(PATH)
except Exception as e:
    print(f"FATAL: workbook will not open: {e}"); sys.exit(2)

names=wb.sheetnames
# 1. Required tabs present
required=["00 Cover","01 Contents","02 Executive Summary","Dashboard","Financial Model",
          "07 Comparison Matrix","MÜV Peer Set","19 Canada Regulatory","Canada Regulatory v2",
          "Electrolyte BFY Deep-Dive","Go-NoGo Decision","Battlecards","Verification Log",
          "22 Risk Register","24 Sources","25 Glossary"]
for r in required:
    check(r in names, f"missing required tab: {r}")

# 2. Dashboard has charts
try: check(len(wb["Dashboard"]._charts)>=5, "Dashboard should have >=5 charts")
except Exception as e: errors.append(f"Dashboard chart check failed: {e}")

# 3. Financial Model has formulas + scenario dropdown
fm=wb["Financial Model"]
formulas=[c.value for row in fm.iter_rows() for c in row if isinstance(c.value,str) and c.value.startswith("=")]
check(len(formulas)>=25, f"Financial Model should have >=25 formulas (found {len(formulas)})")
check(any(dv.type=="list" for dv in fm.data_validations.dataValidation), "Financial Model missing scenario dropdown")

# 4. Contents nav links present
toc=wb["01 Contents"]
links=sum(1 for r in range(1,toc.max_row+1) if toc.cell(row=r,column=1).hyperlink)
check(links>=25, f"Contents should have >=25 nav links (found {links})")

# 5. Stale-content guard: MÜV must NOT be described as a powder OUTSIDE a correction/retraction context
STALE_TERMS=("powder drink mix","sparkling/effervescent POWDER","IS a sparkling POWDER","MÜV is a POWDER")
CORRECTION_CUES=("retract","corrected","superseded","not a powder","was WRONG","earlier","interim","downgrad")
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if not isinstance(c.value,str): continue
            v=c.value
            if any(t.lower() in v.lower() for t in STALE_TERMS):
                if not any(cue.lower() in v.lower() for cue in CORRECTION_CUES):
                    warns.append(f"possible stale 'MÜV powder' claim in {ws.title}!{c.coordinate}: {v[:70]}")

# 5b. Consistency (I7): front-door + brief present; Home link on every content tab
for r in ["🏠 Start Here","Executive Brief 1-page"]:
    check(r in names, f"missing navigation/brief tab: {r}")
missing_home=[]
for ws in wb.worksheets:
    if ws.title=="🏠 Start Here": continue
    has_home=any((isinstance(c.value,str) and "Home" in c.value and c.hyperlink) for row in ws.iter_rows(min_row=1,max_row=2) for c in row)
    if not has_home: missing_home.append(ws.title)
check(not missing_home, f"tabs missing a Home link: {missing_home[:6]}")

# 5c. Contrast/accessibility (I18): flag solid-fill cells whose text is low-contrast on the fill
def _lum(hex6):
    r,g,b=(int(hex6[i:i+2],16)/255 for i in (0,2,4))
    f=lambda c:(c/12.92 if c<=0.03928 else ((c+0.055)/1.055)**2.4)
    return 0.2126*f(r)+0.7152*f(g)+0.0722*f(b)
def _ratio(a,b):
    la,lb=_lum(a),_lum(b); hi,lo=max(la,lb),min(la,lb); return (hi+0.05)/(lo+0.05)
low_contrast=[]
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if not isinstance(c.value,str) or not c.value.strip(): continue
            fillc=getattr(getattr(c.fill,"fgColor",None),"rgb",None)
            if not (isinstance(fillc,str) and len(fillc)==8 and c.fill.patternType=="solid"): continue
            if fillc[-6:].upper() in ("FFFFFF","000000"): continue  # skip pure/no-fill sentinels
            fontc=getattr(getattr(c.font,"color",None),"rgb",None)
            fhex = fontc[-6:] if (isinstance(fontc,str) and len(fontc)==8) else "000000"  # default black
            try:
                if _ratio(fhex, fillc[-6:]) < 2.0: low_contrast.append(f"{ws.title}!{c.coordinate}")
            except Exception: pass
if low_contrast: warns.append(f"low-contrast text cells (<2.0): {low_contrast[:8]}")

# 6. Go/NoGo verdict formula present
g=wb["Go-NoGo Decision"]
has_verdict=any(isinstance(c.value,str) and "GO" in str(c.value) and c.value.startswith("=IF") for row in g.iter_rows() for c in row)
check(has_verdict, "Go-NoGo verdict IF() formula missing")

print(f"tabs={len(names)} charts_ok formulas={len(formulas)} navlinks={links}")
for w in warns: print("WARN:", w)
if errors:
    print(f"\nFAILED with {len(errors)} error(s):")
    for e in errors: print("  ERROR:", e)
    sys.exit(1)
print(f"OK — {len(names)} tabs, {len(warns)} warning(s), 0 errors.")
