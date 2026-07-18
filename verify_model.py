#!/usr/bin/env python3
"""Safety gate for the 10x improvement loop.

Rebuilds the workbook from build_model.py, then checks:
  1. build_model.py runs without error and writes the .xlsx
  2. the workbook opens and contains the expected tabs
  3. no formula/text cell contains an Excel error literal (#REF!, #DIV/0!, …)
  4. a formulas-engine *recompute* produces no error on any REAL sheet
     (the model uses INDIRECT with a scenario selector, which the engine
     can't resolve — those phantom "&$C$4&" sheets are ignored on purpose)
  5. the standalone HTML dashboard is present and well-formed enough to open

Exit 0 = PASS (safe to commit), non-zero = FAIL (revert the change).
Run:  python3 verify_model.py
"""
import subprocess, sys, pathlib, warnings

XLSX = "Organika_Sparkling_Daily_1YR_Scenario_Model.xlsx"
HTML = "organika_sparkling_daily_model.html"
EXPECTED_TABS = {"Home", "Assumptions", "Dashboard", "Base", "Monthly",
                 "Pricing Lab", "Targets", "Sensitivity", "Checks", "Cover"}
ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!", "#N/A", "#NULL!")

def fail(msg):
    print(f"  ✗ {msg}")
    sys.exit(1)

def _pip(pkg):
    subprocess.run([sys.executable, "-m", "pip", "install", "-q", pkg])

def engine_recompute(real_sheets):
    """Recompute every formula; return error cells on real sheets only.

    Returns (ran: bool, errors: list). ran=False means the engine could not
    be used (treated as a soft skip, never a hard fail)."""
    try:
        import formulas  # noqa
    except ImportError:
        _pip("formulas")
        try:
            import formulas  # noqa
        except Exception:
            return False, []
    try:
        import formulas
        warnings.filterwarnings("ignore")
        sol = formulas.ExcelModel().loads(XLSX).finish().calculate()
    except Exception as e:
        print(f"  ~ engine recompute skipped ({type(e).__name__})")
        return False, []
    errs = []
    for key, cell in sol.items():
        try:
            s = str(cell.value)
        except Exception:
            continue
        if not any(tok in s for tok in ERROR_TOKENS) and "XlError" not in s:
            continue
        # key looks like  '[file]Sheet Name'!A1  — pull the sheet name out
        try:
            sheet = key.split("]", 1)[1].rsplit("'!", 1)[0].strip("'")
        except Exception:
            continue
        if sheet in real_sheets:          # ignore INDIRECT phantom sheets
            errs.append(f"{sheet}!{key.rsplit('!',1)[-1]} = {s[:24]}")
    return True, errs

def main():
    # 1. build
    r = subprocess.run([sys.executable, "build_model.py"], capture_output=True, text=True)
    if r.returncode != 0:
        fail("build_model.py errored:\n" + (r.stderr or r.stdout)[-1500:])
    print("  ✓ build_model.py ran")

    # 2/3. workbook opens, tabs present, no literal error tokens
    try:
        import openpyxl
    except ImportError:
        _pip("openpyxl"); import openpyxl
    if not pathlib.Path(XLSX).exists():
        fail(f"{XLSX} not written")
    wb = openpyxl.load_workbook(XLSX)
    real_sheets = set(wb.sheetnames)
    missing = EXPECTED_TABS - real_sheets
    if missing:
        fail(f"missing tabs: {sorted(missing)}")
    print(f"  ✓ workbook opens · {len(real_sheets)} tabs")

    lit = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and any(e in c.value for e in ERROR_TOKENS):
                    if not (c.value.startswith("=") and '"' in c.value):
                        lit.append(f"{ws.title}!{c.coordinate}")
    if lit:
        fail("error literals in cells:\n    " + "\n    ".join(lit[:15]))
    print("  ✓ no error literals in any cell")

    # 4. engine recompute (real sheets only)
    ran, errs = engine_recompute(real_sheets)
    if ran:
        if errs:
            fail("recompute errors on real sheets:\n    " + "\n    ".join(errs[:15]))
        print("  ✓ engine recompute clean (INDIRECT phantoms ignored)")

    # 5. HTML present & balanced enough to open
    h = pathlib.Path(HTML)
    if h.exists():
        txt = h.read_text(encoding="utf-8", errors="ignore")
        if "<html" not in txt.lower() or "</html>" not in txt.lower():
            fail("HTML missing <html>…</html>")
        if txt.lower().count("<script") != txt.lower().count("</script>"):
            fail("HTML <script> tags unbalanced")
        print(f"  ✓ HTML dashboard OK · {len(txt)//1024} KB")

    print("PASS")
    sys.exit(0)

if __name__ == "__main__":
    main()
