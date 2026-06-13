#!/usr/bin/env python3
"""
Integrity test suite for the Amazon.ca Beverage Competitor-Intelligence workbook.

Turns the README's honesty promises into machine-checked guarantees and guards
against regressions in the generator. Loads the newest Amazon_ca_*.xlsx (formulas
read as strings) and asserts:

  1. Structure      — every expected tab is present.
  2. Integrity      — NO commercial cell (price/S&S/coupon/promo/rating/#ratings/
                      badge/BSR/URL/coupon-value) is fabricated; all are blank.
  3. Formulas       — every computed column actually holds a formula, not a literal.
  4. Nutrition sane — no negatives / impossible magnitudes in the stable nutrition.
  5. Cross-tab      — master SKU rows reconcile with the build's reported total.
  6. Navigation     — every README TOC hyperlink resolves to a real sheet.
  7. Provenance     — every Enrichment Log row carries a date and a source.

Exit code 0 = all pass; 1 = at least one hard failure. Warnings never fail the run.
Usage:  python3 test_competitor_intel.py [workbook.xlsx]
"""
import sys, glob, re
from openpyxl import load_workbook

EXPECTED_SHEETS=["README & Methodology","Executive Dashboard","Strategic Insights",
 "Data Dictionary","M1 - Sparkling Electrolyte","M2 - Electrolyte RTD",
 "M3 - Functional Beverage","M4 - Caffeine & Energy","M5 - Sparkling Water & Soda",
 "Brand Roll-up","Category Benchmarks","Nutrition Scoreboard","Pricing & Promo Analysis",
 "Subscription Strategy","Why They Win","Velocity Estimate","Flavour Map",
 "Live-Capture Protocol","QA & Integrity","Enrichment Log","Sources"]
MASTERS=[s for s in EXPECTED_SHEETS if s.startswith("M") and " - " in s]

# field-number prefixes (token before first ".") grouped by expected behaviour
COMMERCIAL_BLANK={"6","9","10","12","13","23","24","25","26","30","H2"}
FORMULA_COLS={"7","8","11","14","C1","C2","C3","C4","C5","C6"}
PROV_COLS={"P1","P2","P3"}
NUTRITION={"16":("caffeine",0,500),"17":("sodium",0,2000),
           "18":("potassium",0,2000),"19":("sugar",0,100),"20":("calories",0,500)}

FAIL=[]; WARN=[]; npass=0
def ok(cond,msg):
    global npass
    if cond: npass+=1
    else: FAIL.append(msg)
def warn(cond,msg):
    if not cond: WARN.append(msg)
def prefix(h): return str(h).split(".")[0].strip() if h else ""

def header_map(ws):
    return {prefix(c.value):c.column for c in ws[1] if c.value}

def main():
    all_wbs=sorted(glob.glob("Amazon_ca_*.xlsx"))
    path=sys.argv[1] if len(sys.argv)>1 else all_wbs[-1]
    wb=load_workbook(path)  # formulas stay as strings
    print(f"Testing: {path}\n")
    warn(len(all_wbs)<=1,
         f"[housekeeping] {len(all_wbs)} dated workbooks present ({', '.join(all_wbs)}); keep only the current build tracked")

    # 1. structure
    for s in EXPECTED_SHEETS:
        ok(s in wb.sheetnames, f"[structure] missing tab: {s}")

    total_rows=0
    for mname in MASTERS:
        if mname not in wb.sheetnames: continue
        ws=wb[mname]; hm=header_map(ws)
        last=ws.max_row
        # find true last data row (col 1 brand populated)
        data_rows=[r for r in range(2,last+1) if ws.cell(r,1).value not in (None,"")]
        total_rows+=len(data_rows)

        # 2. commercial integrity — these columns must be blank everywhere
        for pre in COMMERCIAL_BLANK:
            col=hm.get(pre)
            if not col: continue
            bad=[r for r in data_rows if ws.cell(r,col).value not in (None,"")]
            ok(not bad, f"[integrity] {mname} col {pre} has {len(bad)} fabricated commercial value(s) (rows {bad[:5]})")

        # 3. formulas present in computed columns
        for pre in FORMULA_COLS:
            col=hm.get(pre)
            if not col or not data_rows: continue
            sample=ws.cell(data_rows[0],col).value
            ok(isinstance(sample,str) and sample.startswith("="),
               f"[formula] {mname} col {pre} is not a formula (got {sample!r})")

        # 4. nutrition sanity
        for pre,(field,lo,hi) in NUTRITION.items():
            col=hm.get(pre)
            if not col: continue
            for r in data_rows:
                v=ws.cell(r,col).value
                if isinstance(v,(int,float)):
                    ok(lo<=v<=hi, f"[nutrition] {mname} {field} out of range at row {r}: {v}")

        # 4b. provenance columns must be stamped on every data row (honesty trail)
        for pre in PROV_COLS:
            col=hm.get(pre)
            if not col: continue
            blank=[r for r in data_rows if str(ws.cell(r,col).value or "").strip()==""]
            ok(not blank, f"[provenance] {mname} col {pre} missing on {len(blank)} row(s)")

    # 5. cross-tab reconciliation: master SKU rows must equal Brand Roll-up SKU sum
    ok(total_rows>0, "[cross-tab] no SKU data rows found in master tabs")
    warn(total_rows>=2000, f"[cross-tab] master SKU rows ({total_rows}) lower than expected (~2900)")
    print(f"   master SKU rows counted: {total_rows}")
    if "Brand Roll-up" in wb.sheetnames:
        br=wb["Brand Roll-up"]; hdrs={(c.value or "").lower():c.column for c in br[1] if c.value}
        sku_col=next((v for k,v in hdrs.items() if "sku" in k),2)
        s=sum(br.cell(r,sku_col).value for r in range(2,br.max_row+1)
              if isinstance(br.cell(r,sku_col).value,(int,float)))
        ok(s==total_rows,
           f"[cross-tab] Brand Roll-up SKU sum ({s}) != master rows ({total_rows})")

    # 6. navigation — README TOC hyperlinks resolve
    rd=wb["README & Methodology"]
    links=[c for row in rd.iter_rows(min_col=2,max_col=2) for c in row if c.hyperlink]
    ok(len(links)>=len(EXPECTED_SHEETS),
       f"[nav] TOC has {len(links)} links, expected >= {len(EXPECTED_SHEETS)}")
    for c in links:
        tgt=c.hyperlink.target or ""
        m=re.search(r"#'?([^'!]+)'?!", tgt)
        ok(bool(m) and m.group(1) in wb.sheetnames,
           f"[nav] TOC link does not resolve: {tgt!r}")

    # 7. provenance — Enrichment Log rows carry a date + a source
    if "Enrichment Log" in wb.sheetnames:
        el=wb["Enrichment Log"]; hm=header_map(el)
        # columns by position: 1 Date ... last Source (find by header text)
        hdrs={(c.value or "").lower():c.column for c in el[1] if c.value}
        date_col=next((v for k,v in hdrs.items() if "date" in k),1)
        src_col=next((v for k,v in hdrs.items() if "source" in k),el.max_column)
        # a true data row has a date (YYYY-MM-DD) in col 1 — excludes the trailing note row
        datelike=re.compile(r"^\d{4}-\d{2}-\d{2}")
        rows=[r for r in range(2,el.max_row+1)
              if datelike.match(str(el.cell(r,1).value or ""))]
        ok(len(rows)>0, "[provenance] Enrichment Log has no entries")
        missing=[r for r in rows if not str(el.cell(r,src_col).value or "").strip()]
        ok(not missing, f"[provenance] {len(missing)} Enrichment Log rows lack a source (rows {missing[:5]})")

    # report
    print(f"\n{'='*60}")
    print(f"PASS: {npass}   WARN: {len(WARN)}   FAIL: {len(FAIL)}")
    for w in WARN: print(f"  ⚠ {w}")
    for f in FAIL: print(f"  ✗ {f}")
    if FAIL:
        print("\nRESULT: FAILED"); sys.exit(1)
    print("\nRESULT: PASSED — all integrity guarantees hold."); sys.exit(0)

if __name__=="__main__":
    main()
