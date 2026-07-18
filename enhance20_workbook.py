#!/usr/bin/env python3
"""Hardening-mode high-value add: 'FAQ & Objections' tab — anticipated leadership pushback + crisp answers."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; STEEL="3E5C76"; LIGHT="EAF1F4"; GREEN="E2EFDA"; GREENHEAD="1E7D32"; GREY="595959"; GOLD="C9A227"; WHITE="FFFFFF"; LINKBLUE="1155CC"
def F(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top")

if "FAQ & Objections" in wb.sheetnames: del wb["FAQ & Objections"]
ws=wb.create_sheet("FAQ & Objections")
ws.sheet_view.showGridLines=False
ws.merge_cells("A1:C1"); t=ws["A1"]; t.value="FAQ & Objections — answers for the leadership room"; t.font=F(bold=True,color=WHITE,size=15); t.fill=fill(NAVY); t.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=26
ws.merge_cells("A2:C2"); s=ws["A2"]; s.value="Anticipated pushback and the one-line answer. Back each with the named tab."; s.font=F(italic=True,color=GREY,size=10); s.alignment=Alignment(horizontal="center")
for j,h in enumerate(["Objection","Answer","Proof tab"]):
    x=ws.cell(row=4,column=1+j,value=h); x.font=F(bold=True,color=WHITE,size=10); x.fill=fill(STEEL); x.alignment=Alignment(horizontal="center"); x.border=BORDER
faq=[
 ("“Isn't this just another me-too electrolyte?”","A flavored SPARKLING electrolyte is genuine white space; we own the daily-wellness sparkling occasion, made in Canada.","MÜV Peer Set / Deep-Dive"),
 ("“Nestlé owns Vital Proteins, PepsiCo owns Poppi — can we even compete?”","We don't fight them on their shelf. We win the sparkling daily-wellness occasion in Canada where we already hold distribution + a collagen-brand halo.","15 Outcomes & M&A / GTM"),
 ("“Collagen efficacy is contested — what's the legal exposure?”","Exactly why the CAN sells hydration/electrolyte (food claims). Collagen beauty claims stay on the NHP powder line, not the can.","Canada Regulatory v2"),
 ("“Why a can, not a powder — powders ship cheaper?”","Sparkling is the wedge and the sensory differentiator; BC manufacturing protects margin. Regulatorily both are food now, so it's a commercial choice.","Electrolyte BFY Deep-Dive"),
 ("“What's the regulatory risk in Canada?”","It's a Supplemented Food — design to that framework from day one (SFFt, cautions). ORS carve-out doesn't apply. Confirm the claim set with counsel.","Canada Regulatory v2"),
 ("“Is the financial model real?”","It's illustrative scaffolding — every input is flagged. Replace with co-packer/distributor quotes. Even the Conservative case has a defined break-even.","Financial Model / Assumptions Audit"),
 ("“Why not just extend the existing powder line?”","The powder line stays (and keeps its collagen claims). MÜV captures the sparkling RTD occasion the powder can't.","MÜV Peer Set"),
 ("“What exactly do you need from us?”","Approve Phase 0: confirm MÜV specs, regulatory sign-off, co-packer quote, and the Costco Canada beachhead.","Next Steps & Gates"),
]
r=5
for i,(q,a,p) in enumerate(faq):
    ws.cell(row=r,column=1,value=q).font=F(bold=True,size=10); ws.cell(row=r,column=1).alignment=WRAP; ws.cell(row=r,column=1).border=BORDER
    ws.cell(row=r,column=2,value=a).font=F(size=10); ws.cell(row=r,column=2).alignment=WRAP; ws.cell(row=r,column=2).border=BORDER
    ws.cell(row=r,column=3,value=p).font=F(size=9,italic=True,color=GREENHEAD); ws.cell(row=r,column=3).alignment=WRAP; ws.cell(row=r,column=3).border=BORDER
    if i%2:
        for col in (1,2,3): ws.cell(row=r,column=col).fill=fill(LIGHT)
    ws.row_dimensions[r].height=42; r+=1
ws.column_dimensions["A"].width=34; ws.column_dimensions["B"].width=64; ws.column_dimensions["C"].width=24
h=ws.cell(row=1,column=13,value="🏠 Home"); h.hyperlink="#'🏠 Start Here'!A1"; h.font=F(bold=True,color=LINKBLUE,underline="single",size=8); h.alignment=Alignment(horizontal="right")
ws.sheet_properties.tabColor=GOLD
ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0; ws.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
# place after How to Present; contents
s2=wb["FAQ & Objections"]; wb._sheets.remove(s2); wb._sheets.insert(wb.sheetnames.index("How to Present")+1,s2)
toc=wb["01 Contents"]; row=toc.max_row+1
a=toc.cell(row=row,column=1,value="FAQ & Objections"); a.hyperlink="#'FAQ & Objections'!A1"; a.font=F(bold=True,color=LINKBLUE,underline="single",size=10); a.fill=fill(GOLD); a.border=BORDER; a.alignment=Alignment(vertical="center")
b=toc.cell(row=row,column=2,value="Anticipated leadership pushback + crisp answers, with proof tabs"); b.font=F(size=10); b.border=BORDER; b.alignment=Alignment(vertical="center")
wb.save(PATH)
print("FAQ & Objections tab added. sheets:",len(wb.sheetnames))
