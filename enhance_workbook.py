#!/usr/bin/env python3
"""10x enhancement pass: adds a live financial model, a visual dashboard with charts,
conditional formatting, auto-filters, tab-color grouping, print setup and metadata
to the enterprise workbook."""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)

NAVY="14304F"; TEAL="2E7D8A"; STEEL="3E5C76"; LIGHT="EAF1F4"; LIGHT2="F4F8FA"
AMBER="FFF2CC"; AMBERHEAD="B45309"; GREEN="E2EFDA"; GREENHEAD="1E7D32"
GOLD="C9A227"; WHITE="FFFFFF"; GREY="595959"; REDH="9C2A2A"
def F(**k): return Font(name="Calibri", **k)
HEAD=F(bold=True,color=WHITE,size=10); TITLE=F(bold=True,color=WHITE,size=18)
BODY=F(size=10); BODYB=F(size=10,bold=True); SMALL=F(size=9,italic=True,color=GREY)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top"); CTR=Alignment(horizontal="center",vertical="center",wrap_text=True)
def fill(c): return PatternFill("solid",fgColor=c)
YELLOW=fill("FFF7CC")  # editable input cells

def sheet_title(ws,text,n,sub=None):
    ws.sheet_view.showGridLines=False
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    c=ws.cell(row=1,column=1,value=text); c.font=TITLE; c.fill=fill(NAVY); c.alignment=CTR
    ws.row_dimensions[1].height=30
    if sub:
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=n)
        s=ws.cell(row=2,column=1,value=sub); s.font=F(bold=True,color=WHITE,size=10); s.fill=fill(TEAL); s.alignment=CTR
        ws.row_dimensions[2].height=20
        return 4
    return 3
def banner(ws,row,text,n,color=STEEL):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=n)
    c=ws.cell(row=row,column=1,value=text); c.font=F(bold=True,color=WHITE,size=10); c.fill=fill(color)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1); return row+1

# =====================================================================
# 1) FINANCIAL MODEL (live formulas)
# =====================================================================
fm=wb.create_sheet("Financial Model")
nr=sheet_title(fm,"MUV Financial Model — Unit Economics & 3-Year Build",6,
   "Edit the yellow input cells; all white cells are live formulas. Illustrative defaults — replace with your numbers")
fm.column_dimensions["A"].width=3
fm.column_dimensions["B"].width=34
for col in "CDEF": fm.column_dimensions[col].width=15

def lbl(r,t,bold=False):
    c=fm.cell(row=r,column=2,value=t); c.font=BODYB if bold else BODY; c.alignment=Alignment(vertical="center")
def inp(r,col,val,fmt=None):
    c=fm.cell(row=r,column=col,value=val); c.font=BODYB; c.fill=YELLOW; c.border=BORDER; c.alignment=CTR
    if fmt: c.number_format=fmt
    return c
def calc(r,col,formula,fmt=None,bold=False):
    c=fm.cell(row=r,column=col,value=formula); c.font=BODYB if bold else BODY; c.border=BORDER; c.alignment=CTR
    c.fill=fill(LIGHT2)
    if fmt: c.number_format=fmt
    return c

nr=banner(fm,nr,"PER-UNIT ASSUMPTIONS  (edit yellow)",6)
a=nr
lbl(a,"List / shelf price per can (CAD)"); inp(a,3,3.49,'"$"#,##0.00')
lbl(a+1,"COGS per can (liquid+can+co-pack, CAD)"); inp(a+1,3,1.05,'"$"#,##0.00')
lbl(a+2,"Retailer margin %"); inp(a+2,3,0.35,'0%')
lbl(a+3,"Distributor margin %"); inp(a+3,3,0.15,'0%')
lbl(a+4,"Trade spend % of net revenue"); inp(a+4,3,0.20,'0%')
lbl(a+5,"Cans per case"); inp(a+5,3,12,'0')
PRICE=f"$C${a}"; COGS=f"$C${a+1}"; RETM=f"$C${a+2}"; DISTM=f"$C${a+3}"; TRADE=f"$C${a+4}"
nr=a+7
nr=banner(fm,nr,"DERIVED PER-CAN ECONOMICS",6)
d=nr
lbl(d,"Brand net revenue per can",True);      calc(d,3,f"={PRICE}*(1-{RETM})*(1-{DISTM})",'"$"#,##0.00',True)
lbl(d+1,"Gross profit per can",True);          calc(d+1,3,f"=C{d}-{COGS}",'"$"#,##0.00',True)
lbl(d+2,"Gross margin %",True);                calc(d+2,3,f"=C{d+1}/C{d}",'0.0%',True)
lbl(d+3,"Trade spend per can");                calc(d+3,3,f"=C{d}*{TRADE}",'"$"#,##0.00')
lbl(d+4,"Contribution per can (after trade)",True); calc(d+4,3,f"=C{d+1}-C{d+3}",'"$"#,##0.00',True)
NETCAN=f"$C${d}"; GPCAN=f"$C${d+1}"; TRADECAN=f"$C${d+3}"; CONTCAN=f"$C${d+4}"
nr=d+6

nr=banner(fm,nr,"3-YEAR VOLUME BUILD  (edit yellow rows)",6)
h=nr
for j,t in enumerate(["Metric","Year 1","Year 2","Year 3"]):
    c=fm.cell(row=h,column=2+j,value=t); c.font=HEAD; c.fill=fill(NAVY); c.alignment=CTR; c.border=BORDER
rows_spec=[
 ("Distribution (doors)",[500,2000,6000],'#,##0',"input"),
 ("Velocity (units/store/week)",[3,4,5],'0.0',"input"),
 ("Annual units (= doors×velocity×52)",None,'#,##0',"units"),
 ("Net revenue (CAD)",None,'"$"#,##0',"net"),
 ("Gross profit (CAD)",None,'"$"#,##0',"gp"),
 ("Trade spend (CAD)",None,'"$"#,##0',"trade"),
 ("Contribution (CAD)",None,'"$"#,##0',"cont"),
]
r=h+1
rowref={}
for name,vals,fmt,kind in rows_spec:
    lbl(r,name, kind in ("units","net","gp","trade","cont","input") and name.endswith(")") or kind in("net","cont"))
    fm.cell(row=r,column=2).font=BODYB if kind in("net","cont") else BODY
    for j,col in enumerate([3,4,5]):
        if kind=="input":
            inp(r,col,vals[j],fmt)
        else:
            L=get_column_letter(col)
            if kind=="units":    f=f"={L}{rowref['Distribution (doors)']}*{L}{rowref['Velocity (units/store/week)']}*52"
            elif kind=="net":    f=f"={L}{rowref['units']}*{NETCAN}"
            elif kind=="gp":     f=f"={L}{rowref['units']}*{GPCAN}"
            elif kind=="trade":  f=f"={L}{rowref['units']}*{TRADECAN}"
            elif kind=="cont":   f=f"={L}{rowref['units']}*{CONTCAN}"
            calc(r,col,f,fmt, kind in("net","cont"))
    rowref[name]=r
    if kind=="units": rowref["units"]=r
    fm.row_dimensions[r].height=18
    r+=1
YEAR_HDR=h; NET_ROW=rowref["Net revenue (CAD)"]; CONT_ROW=rowref["Contribution (CAD)"]
nr=r+1
fm.merge_cells(start_row=nr,start_column=2,end_row=nr,end_column=6)
note=fm.cell(row=nr,column=2,value=("Model logic: net/can = list × (1−retailer %) × (1−distributor %). Contribution excludes fixed overhead, slotting, "
 "marketing and freight/FX — add those below for a full P&L. Defaults illustrate a disciplined Canadian beachhead → scale path; "
 "they are NOT a forecast. Calibrate every yellow cell with real quotes (co-packer COGS, distributor terms, buyer velocity)."))
note.font=SMALL; note.alignment=WRAP; note.fill=fill(AMBER); note.border=BORDER; fm.row_dimensions[nr].height=46

# =====================================================================
# 2) DASHBOARD with charts
# =====================================================================
dash=wb.create_sheet("Dashboard")
nr=sheet_title(dash,"Executive Dashboard",12,"Visual snapshot — category sizing, leader trajectories, share, competitive scores & MUV model")
for col in "ABCDEFGHIJKL": dash.column_dimensions[col].width=11

# KPI strip
kpis=[("Modern soda (US, 2024)","$1.8B","+83% YoY"),
      ("Poppi exit (PepsiCo)","$1.95B","May 2025"),
      ("OLIPOP valuation","$1.85B","Feb 2025"),
      ("Collagen-sparkling indie survival","Low","graveyard"),
      ("Canada beachhead","Costco+Amazon","proven path")]
r0=nr
for i,(t,v,s) in enumerate(kpis):
    c0=1+i*2+ (1 if i else 1)
# simpler: place 5 KPI cards across 2-col blocks
col=1
for t,v,s in kpis:
    dash.merge_cells(start_row=r0,start_column=col,end_row=r0,end_column=col+1)
    a1=dash.cell(row=r0,column=col,value=t); a1.font=F(bold=True,color=WHITE,size=8); a1.fill=fill(STEEL); a1.alignment=CTR
    dash.merge_cells(start_row=r0+1,start_column=col,end_row=r0+1,end_column=col+1)
    a2=dash.cell(row=r0+1,column=col,value=v); a2.font=F(bold=True,color=NAVY,size=14); a2.fill=fill(LIGHT); a2.alignment=CTR
    dash.merge_cells(start_row=r0+2,start_column=col,end_row=r0+2,end_column=col+1)
    a3=dash.cell(row=r0+2,column=col,value=s); a3.font=F(italic=True,color=GREY,size=8); a3.fill=fill(LIGHT); a3.alignment=CTR
    for rr in (r0,r0+1,r0+2):
        for cc in (col,col+1): dash.cell(row=rr,column=cc).border=BORDER
    col+=2
dash.row_dimensions[r0].height=22; dash.row_dimensions[r0+1].height=22; dash.row_dimensions[r0+2].height=16

# ---- helper data tables (placed low, columns further right won't collide with charts) ----
DATA_ROW=60
def put_table(start_row,start_col,title,headers,rows):
    dash.cell(row=start_row,column=start_col,value=title).font=BODYB
    for j,hh in enumerate(headers):
        dash.cell(row=start_row+1,column=start_col+j,value=hh).font=SMALL
    for i,row in enumerate(rows):
        for j,val in enumerate(row):
            dash.cell(row=start_row+2+i,column=start_col+j,value=val)
    return start_row+2+len(rows)

# Category sizes (US selected, $B)
cat=[("Modern soda",1.8),("Hydration powder",1.5),("Electrolyte powder",2.8),
     ("Sparkling water",6.9),("Collagen bev (US)",0.05),("Beauty-from-within (global)",3.24)]
put_table(DATA_ROW,1,"Category sizes ($B)",["Category","$B"],cat)
# Revenue trajectory
yrs=[2020,2021,2022,2023,2024]
poppi=[13,26,65,100,500]; olipop=[20,40,73,200,400]
put_table(DATA_ROW,5,"Revenue ($M)",["Year","Poppi","OLIPOP"],list(zip(yrs,poppi,olipop)))
# Share pie
share=[("Poppi",38),("OLIPOP",32.7),("Zevia",12.1),("Jarritos",10.3),("Fever-Tree",4),("Other",2.9)]
put_table(DATA_ROW,9,"Modern soda share %",["Brand","%"],share)
# Competitive scores (computed)
weights=[0.15,0.15,0.15,0.10,0.10,0.10,0.15,0.10]
sc={"Poppi":[5,5,2,3,3,5,5,2],"OLIPOP":[5,5,4,3,4,4,4,2],"LMNT":[4,3,3,4,3,5,4,2],
    "Liquid I.V.":[4,5,3,3,3,4,5,3],"Vital Proteins":[4,5,4,3,3,3,5,2],"Recess":[3,3,3,3,4,4,3,1],
    "MUV (Organika)":[3,3,3,4,3,2,2,5]}
scores=[(k,round(sum(s*w for s,w in zip(v,weights)),2)) for k,v in sc.items()]
put_table(DATA_ROW,13,"Weighted score",["Brand","Score"],scores)

# ---- Charts ----
# Bar: category sizes
c1=BarChart(); c1.type="col"; c1.title="Selected Category Sizes (US, $B)"; c1.height=7.5; c1.width=15
d=Reference(dash,min_col=2,min_row=DATA_ROW+1,max_row=DATA_ROW+1+len(cat))
cats=Reference(dash,min_col=1,min_row=DATA_ROW+2,max_row=DATA_ROW+1+len(cat))
c1.add_data(d,titles_from_data=True); c1.set_categories(cats); c1.legend=None
c1.y_axis.title="$B"; dash.add_chart(c1,"A8")

# Line: revenue trajectory
c2=LineChart(); c2.title="Revenue Trajectory ($M) — Poppi vs OLIPOP"; c2.height=7.5; c2.width=15
d2=Reference(dash,min_col=6,min_row=DATA_ROW+1,max_col=7,max_row=DATA_ROW+1+len(yrs))
cats2=Reference(dash,min_col=5,min_row=DATA_ROW+2,max_row=DATA_ROW+1+len(yrs))
c2.add_data(d2,titles_from_data=True); c2.set_categories(cats2); c2.y_axis.title="$M"
for s in c2.series: s.smooth=False
dash.add_chart(c2,"G8")

# Pie: share
c3=PieChart(); c3.title="Modern Soda Share 2024 (%)"; c3.height=7.5; c3.width=11
d3=Reference(dash,min_col=10,min_row=DATA_ROW+1,max_row=DATA_ROW+1+len(share))
cats3=Reference(dash,min_col=9,min_row=DATA_ROW+2,max_row=DATA_ROW+1+len(share))
c3.add_data(d3,titles_from_data=True); c3.set_categories(cats3)
c3.dataLabels=DataLabelList(); c3.dataLabels.showPercent=False; c3.dataLabels.showVal=True
dash.add_chart(c3,"A24")

# Bar: competitive scores
c4=BarChart(); c4.type="bar"; c4.title="Weighted Competitive Score (1–5)"; c4.height=7.5; c4.width=15
d4=Reference(dash,min_col=14,min_row=DATA_ROW+1,max_row=DATA_ROW+1+len(scores))
cats4=Reference(dash,min_col=13,min_row=DATA_ROW+2,max_row=DATA_ROW+1+len(scores))
c4.add_data(d4,titles_from_data=True); c4.set_categories(cats4); c4.legend=None
c4.dataLabels=DataLabelList(); c4.dataLabels.showVal=True
dash.add_chart(c4,"E24")

# Bar: financial model output (cross-sheet reference)
c5=BarChart(); c5.type="col"; c5.title="MUV Model — Net Revenue vs Contribution (CAD)"; c5.height=7.5; c5.width=15
dnet=Reference(fm,min_col=2,min_row=NET_ROW,max_col=5,max_row=NET_ROW)
dcont=Reference(fm,min_col=2,min_row=CONT_ROW,max_col=5,max_row=CONT_ROW)
c5.add_data(dnet,titles_from_data=True,from_rows=True)
c5.add_data(dcont,titles_from_data=True,from_rows=True)
yrcat=Reference(fm,min_col=3,min_row=YEAR_HDR,max_col=5,max_row=YEAR_HDR)
c5.set_categories(yrcat); c5.y_axis.title="CAD"
dash.add_chart(c5,"J24")

# hide helper data rows visually (group/outline)
for rr in range(DATA_ROW, DATA_ROW+2+len(cat)+2):
    dash.row_dimensions[rr].outlineLevel=1; dash.row_dimensions[rr].hidden=True
dash.merge_cells(start_row=40,start_column=1,end_row=40,end_column=12)
src=dash.cell(row=40,column=1,value="Charts are illustrative and built from the cited research (see ‘24 Sources’). Revenue/share figures rounded; OLIPOP 2020–22 interpolated. Financial-model bars are driven live by the ‘Financial Model’ tab inputs.")
src.font=SMALL; src.alignment=WRAP; dash.row_dimensions[40].height=28

# =====================================================================
# 3) CONDITIONAL FORMATTING on scoring sheet
# =====================================================================
sm=wb["17 Scoring Model"]
# find weights row
wrow=None
for r in range(1,40):
    if sm.cell(row=r,column=1).value and str(sm.cell(row=r,column=1).value).startswith("WEIGHTS"):
        wrow=r; break
if wrow:
    hdr=wrow+1; first=hdr+1; last=first+6  # 7 brands
    # color scale on 8 criteria cols B..I
    rng=f"B{first}:I{last}"
    sm.conditional_formatting.add(rng,ColorScaleRule(
        start_type='num',start_value=1,start_color='F8696B',
        mid_type='num',mid_value=3,mid_color='FFEB84',
        end_type='num',end_value=5,end_color='63BE7B'))
    # data bar on weighted score col J
    sm.conditional_formatting.add(f"J{first}:J{last}",
        DataBarRule(start_type='num',start_value=0,end_type='num',end_value=5,color="2E7D8A"))

# Financial model: data bars on net revenue & contribution
fm.conditional_formatting.add(f"C{NET_ROW}:E{NET_ROW}",
    DataBarRule(start_type='min',end_type='max',color="3E5C76"))
fm.conditional_formatting.add(f"C{CONT_ROW}:E{CONT_ROW}",
    DataBarRule(start_type='min',end_type='max',color="1E7D32"))

# =====================================================================
# 4) AUTO-FILTER on comparison matrix
# =====================================================================
cm=wb["07 Comparison Matrix"]
cm.auto_filter.ref="A3:L11"

# =====================================================================
# 5) TAB COLORS (section grouping) + reorder new sheets to front
# =====================================================================
group={
 "00 Cover":NAVY,"01 Contents":NAVY,"02 Executive Summary":NAVY,"03 Methodology":NAVY,
 "Dashboard":GOLD,"Financial Model":GOLD,
 "04 Market Sizing":TEAL,"05 Trends & Macro":TEAL,"06 Consumer":TEAL,
 "07 Comparison Matrix":STEEL,"08 Extended Landscape":STEEL,"09 Channel & Retail":STEEL,
 "10 Channel Economics":STEEL,"11 Formulation":STEEL,"12 Marketing":STEEL,
 "13 Launch Playbooks":STEEL,"14 Patterns & Failures":STEEL,"15 Outcomes & M&A":STEEL,
 "16 SWOT":STEEL,"17 Scoring Model":STEEL,
 "18 Canada Market":GREENHEAD,"19 Canada Regulatory":GREENHEAD,
 "20 GTM Recommendation":AMBERHEAD,"21 Launch Plan":AMBERHEAD,"22 Risk Register":AMBERHEAD,"23 KPI Dashboard":AMBERHEAD,
 "24 Sources":GREY,"25 Glossary":GREY,
}
for name,color in group.items():
    if name in wb.sheetnames: wb[name].sheet_properties.tabColor=color

# move Dashboard + Financial Model to just after Executive Summary
dash_ws=wb["Dashboard"]; fmws=wb["Financial Model"]
wb._sheets.remove(dash_ws); wb._sheets.remove(fmws)
idx=wb.sheetnames.index("02 Executive Summary")+1
wb._sheets.insert(idx,dash_ws); wb._sheets.insert(idx+1,fmws)

# =====================================================================
# 6) PRINT SETUP + METADATA
# =====================================================================
from openpyxl.worksheet.properties import PageSetupProperties
for ws in wb.worksheets:
    ws.page_setup.orientation="landscape"
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
    ws.oddFooter.right.text="&P of &N"
    ws.oddFooter.left.text="Organika MUV — Competitor Intelligence (confidential draft)"
wb.properties.title="Functional & Wellness Sparkling — Competitor Intelligence (MUV)"
wb.properties.creator="Organika MUV launch analysis"
wb.properties.subject="Competitive intelligence & Canada market-entry dossier"
wb.properties.keywords="functional beverage, prebiotic soda, collagen, hydration, Canada, MUV, Organika"

wb.save(PATH)
print("enhanced & saved:",PATH)
print("total sheets:",len(wb.sheetnames))
print("order:")
for s in wb.sheetnames: print("  -",s)
