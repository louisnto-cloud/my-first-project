#!/usr/bin/env python3
"""Reconcile competitive framing to MÜV's actual peer set: adds a 'MÜV Peer Set' tab
and a 'how to read this' framing banner so direct peers vs category-adjacent are explicit."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; TEAL="2E7D8A"; STEEL="3E5C76"; LIGHT="EAF1F4"; LIGHT2="F4F8FA"
AMBER="FFF2CC"; AMBERHEAD="B45309"; GREEN="E2EFDA"; GREENHEAD="1E7D32"; GOLD="C9A227"; WHITE="FFFFFF"; GREY="595959"; LINKBLUE="1155CC"
def F(**k): return Font(name="Calibri",**k)
HEAD=F(bold=True,color=WHITE,size=10); TITLE=F(bold=True,color=WHITE,size=18); BODY=F(size=10); BODYB=F(size=10,bold=True); SMALL=F(size=9,italic=True,color=GREY)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top"); CTR=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)
def fill(c): return PatternFill("solid",fgColor=c)
def title(ws,text,n,sub=None):
    ws.sheet_view.showGridLines=False
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    c=ws.cell(row=1,column=1,value=text); c.font=TITLE; c.fill=fill(NAVY); c.alignment=CTR; ws.row_dimensions[1].height=30
    if sub:
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=n)
        s=ws.cell(row=2,column=1,value=sub); s.font=F(bold=True,color=WHITE,size=10); s.fill=fill(TEAL); s.alignment=CTR; ws.row_dimensions[2].height=20
        return 4
    return 3
def banner(ws,row,text,n,color=STEEL):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=n)
    c=ws.cell(row=row,column=1,value=text); c.font=F(bold=True,color=WHITE,size=10); c.fill=fill(color); c.alignment=Alignment(horizontal="left",vertical="center",indent=1); return row+1
def tbl(ws,start,headers,rows,widths=None,hcolor=NAVY,rowh=None):
    for j,h in enumerate(headers,1):
        c=ws.cell(row=start,column=j,value=h); c.font=HEAD; c.fill=fill(hcolor); c.alignment=CTR; c.border=BORDER
    r=start+1
    for i,row in enumerate(rows):
        for j,val in enumerate(row,1):
            c=ws.cell(row=r,column=j,value=val); c.font=BODY; c.alignment=WRAP; c.border=BORDER
            if i%2: c.fill=fill(LIGHT)
        if rowh: ws.row_dimensions[r].height=rowh
        r+=1
    if widths:
        for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
    return r
def callout(ws,row,text,n,fillc=GREEN,fontc=GREENHEAD,h=44):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=n)
    c=ws.cell(row=row,column=1,value=text); c.font=F(bold=True,color=fontc,size=10); c.fill=fill(fillc); c.alignment=WRAP; c.border=BORDER; ws.row_dimensions[row].height=h; return row+1

# ============================================================ MÜV PEER SET
ps=wb.create_sheet("MÜV Peer Set")
nr=title(ps,"MÜV Peer Set — Direct Competitors vs Category-Adjacent",7,
   "Reconciles the workbook to MÜV's actual lane: a daily-wellness SPARKLING ELECTROLYTE (food-regulated)")
nr=callout(ps,nr,("HOW TO READ THIS WORKBOOK: Tabs 07–17 profile the BROAD functional-beverage market (Poppi/OLIPOP/Vital Proteins). Those are "
 "CATEGORY-ADJACENT — they prove the better-for-you thesis and supply the LAUNCH PLAYBOOK, but they are NOT MÜV's direct shelf rivals. MÜV's "
 "DIRECT peer set is electrolyte/hydration + sparkling (this tab + ‘Electrolyte BFY Deep-Dive’). Read the two together: borrow the soda brands' "
 "go-to-market, but benchmark price/format/claims against the electrolyte brands."),7,fillc=AMBER,fontc=AMBERHEAD,h=64)

nr=banner(ps,nr,"DIRECT PEERS — same occasion / same shelf (benchmark MÜV here)",7,color=GREENHEAD)
nr=tbl(ps,nr,["Brand","Format","Sodium/serv","Price/serv","Lane","Canada channel","Why a direct peer"],[
 ["LMNT","Powder stick (+Sparkling can)","1,000 mg","~$1.50","High-sodium performance","Amazon.ca + DTC","Defines the premium electrolyte ceiling; sparkling-can overlap"],
 ["Liquid I.V.","Powder stick","500 mg","~$1.56 (~$1.00 Costco)","Daily wellness","Costco Canada (2023) → mass","Closest positioning twin (daily wellness); Costco beachhead model"],
 ["Nuun","Effervescent tablet","300 mg","~$0.75","Everyday/endurance","Natural + Amazon.ca","Fizz/effervescence overlap; value anchor"],
 ["BioSteel","Powder / RTD","moderate","mid","Athletic/clean sport","Broad omnichannel (CA)","Canadian rival; the cautionary tale (see Deep-Dive)"],
 ["Prime","RTD + powder","low","hype","Youth/viral","Walmart/convenience","Sparkling-adjacent RTD; hype-cycle warning"],
 ["Gatorlyte / Powerade","RTD","high","mass","Sport/mass","Full distribution","Incumbent sport hydration on the same cooler shelf"],
 ["Organika electrolyte sachets","3.5g powder sachet","~440mg (sibling)","~$29.99/30 (Costco)","Daily, Canadian-made","Costco/Well.ca/Amazon.ca","Organika's own line — internal reference & cannibalization watch"],
 ["MÜV (this product)","Sparkling RTD CAN","TBD ⚠","$14.99/pack","Daily-wellness sparkling","Costco/Amazon/Well.ca","— the subject —"],
],widths=[20,22,12,16,20,22,30],rowh=40)

nr=banner(ps,nr,"CATEGORY-ADJACENT — different occasion; context & playbook only",7,color=STEEL)
nr=tbl(ps,nr,["Brand","Category","What MÜV borrows","What does NOT transfer"],[
 ["Poppi","Prebiotic soda (can)","TikTok creator seeding; mass-priced premium; PepsiCo exit proof","Soda occasion; 2g-fiber claim model (and its $8.9M lawsuit)"],
 ["OLIPOP","Prebiotic soda (can)","Natural-channel seeding; sell-in on repurchase data","Dessert-soda positioning; 9g fiber formulation"],
 ["Vital Proteins Sparkling","Collagen sparkling (can)","Beauty-from-within demand proof; clinical-claim model","Collagen-beauty claims (NHP-only in Canada — keep off MÜV can)"],
 ["Recess","Relaxation sparkling","“Sell the felt benefit, not the molecule” lesson","Adaptogen/CBD relaxation occasion"],
 ["Aura Bora","Sparkling water","Personality-led cheap marketing; flavour novelty","No functional/electrolyte claim"],
],widths=[20,22,40,34],rowh=42)

nr=callout(ps,nr,("RECONCILED THESIS: MÜV should be PRICED & FORMULATED against the DIRECT peers (≤ LMNT/Liquid I.V. per-serving economics, sparkling as the wedge, "
 "Supplemented-Food claims) while BORROWING the GTM from the adjacent soda winners (one obsessive trial tactic, sell-in on velocity/repeat, sequenced "
 "channel ladder). Win the daily-wellness sparkling occasion, made in Canada — don't fight Poppi for the soda shelf or Vital Proteins for beauty claims."),7,h=58)
c=ps.cell(row=1,column=14,value="↩ Contents"); c.hyperlink="#'01 Contents'!A1"; c.font=F(bold=True,color=LINKBLUE,underline="single",size=8); c.alignment=Alignment(horizontal="right")
ps.sheet_properties.tabColor=GOLD
ps.page_setup.orientation="landscape"; ps.page_setup.fitToWidth=1; ps.page_setup.fitToHeight=0; ps.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)

# ============================================================ FRAMING NOTE on 07 Comparison Matrix
# (no row insert — openpyxl shifts merged cells poorly; write into the safe column-N gutter instead)
cm=wb["07 Comparison Matrix"]
b=cm.cell(row=2,column=1,value="LENS: broad functional-beverage set (context + playbook). MÜV's DIRECT electrolyte/sparkling rivals → ‘MÜV Peer Set’ tab.")
b.font=F(bold=True,color=AMBERHEAD,size=9); b.alignment=Alignment(horizontal="left",vertical="center")

# ============================================================ CONTENTS + ORDER
toc=wb["01 Contents"]
r=toc.max_row+1
a=toc.cell(row=r,column=1,value="MÜV Peer Set"); a.hyperlink="#'MÜV Peer Set'!A1"; a.font=F(bold=True,color=LINKBLUE,underline="single",size=10); a.fill=fill(AMBER); a.border=BORDER; a.alignment=LEFT
bb=toc.cell(row=r,column=2,value="MÜV's direct electrolyte/sparkling rivals vs category-adjacent context (reconciliation)"); bb.font=BODY; bb.border=BORDER; bb.alignment=WRAP; toc.row_dimensions[r].height=18

s=wb["MÜV Peer Set"]; wb._sheets.remove(s)
idx=wb.sheetnames.index("07 Comparison Matrix")+1
wb._sheets.insert(idx,s)

wb.properties.description="34-tab dossier: interactive scenario P&L, dashboard, Go/No-Go, battlecards, verification log, Canada-first electrolyte/BFY deep-dive, and a reconciled MÜV peer set. MÜV verified as an RTD sparkling can in a food-regulated lane."
wb.save(PATH)
print("peer-set + framing added. sheets:",len(wb.sheetnames))
PY=None
