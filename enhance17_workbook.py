#!/usr/bin/env python3
"""Loop batch I27+I26+I25:
 I27 add MÜV as a highlighted SUBJECT row in the Comparison Matrix (row 13, below merged callout; no inserts).
 I26 Methodology: expand 'what would strengthen this' (append rows).
 I25 Go-NoGo: append a one-line 'how to use' hint."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
NAVY="14304F"; STEEL="3E5C76"; LIGHT="EAF1F4"; AMBER="FFF2CC"; AMBERHEAD="B45309"; GREEN="E2EFDA"; GREENHEAD="1E7D32"; GOLD="C9A227"; GREY="595959"; WHITE="FFFFFF"
def F(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top")

# ---------- I27: MÜV subject row in Comparison Matrix ----------
cm=wb["07 Comparison Matrix"]
row=13  # below merged graveyard callout at row 12
muv=["MÜV (Organika) — SUBJECT","Sparkling electrolyte (RTD)","Organika est. 1990","Organika (BC, Canada)",
     "Costco CA / Amazon.ca / Well.ca (planned)","Daily-wellness sparkling hydration + prebiotic fibre",
     "RTD sparkling CAN (SKU 4338)","0 sugar, caffeine-free, magnesium glycinate, Fibersol; sodium TBD",
     "$14.99 / pack","Canadian beachhead (planned)","n/a (pre-launch)","Food-regulated (Supplemented Food)"]
for j,v in enumerate(muv,1):
    c=cm.cell(row=row,column=j,value=v); c.font=F(bold=(j==1),size=9,color=(NAVY if j==1 else "000000"))
    c.fill=fill(GOLD if j==1 else GREEN); c.alignment=WRAP; c.border=BORDER
cm.row_dimensions[row].height=54
cm.cell(row=row+1,column=1,value="↑ MÜV shown for benchmarking; its direct rivals (LMNT/Liquid I.V./Nuun) are on the ‘MÜV Peer Set’ tab.").font=F(italic=True,size=8,color=GREY)

# ---------- I26: Methodology 'what would strengthen this' expansion ----------
me=wb["03 Methodology"]
r=me.max_row+2
me.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
h=me.cell(row=r,column=1,value="WHAT WOULD STRENGTHEN THIS FURTHER (priority order)"); h.font=F(bold=True,color=WHITE,size=10); h.fill=fill(STEEL); h.alignment=Alignment(horizontal="left",vertical="center",indent=1)
me.row_dimensions[r].height=18; r+=1
items=[
 ("1. Confirm MÜV on-can specs","Sodium mg, pack size, and whether a collagen variant exists — the last open input on the model and claims."),
 ("2. Licensed Canadian scanner data","NielsenIQ/Circana Canada pull for the sparkling-electrolyte sub-segment to replace US-proxy sizing."),
 ("3. Regulatory-counsel opinion","Written classification + permissible claim set under the Supplemented Foods framework."),
 ("4. Real cost inputs","Co-packer COGS quote + distributor terms to replace the illustrative model assumptions."),
 ("5. Distributor interviews","UNFI Canada / Tree of Life / Purity Life on listing terms and velocity expectations."),
]
for i,(a,b) in enumerate(items):
    x=me.cell(row=r,column=1,value=a); x.font=F(bold=True,size=10); x.border=BORDER; x.alignment=WRAP
    y=me.cell(row=r,column=2,value=b); y.font=F(size=10); y.border=BORDER; y.alignment=WRAP
    if i%2: x.fill=fill(LIGHT); y.fill=fill(LIGHT)
    me.row_dimensions[r].height=30; r+=1

# ---------- I25: Go-NoGo how-to hint ----------
g=wb["Go-NoGo Decision"]
r=g.max_row+2
g.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
c=g.cell(row=r,column=1,value="HOW TO USE: edit the yellow Weight and Score (1–5) cells; the weighted total and the GO / GO-conditional / NO-GO verdict recompute automatically.")
c.font=F(italic=True,bold=True,size=9,color=GREENHEAD); c.fill=fill(GREEN); c.alignment=Alignment(wrap_text=True,vertical="center"); c.border=BORDER
g.row_dimensions[r].height=26

wb.save(PATH)
print("I27 MÜV row + I26 methodology expansion + I25 Go-NoGo hint added.")
