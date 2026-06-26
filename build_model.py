"""
Organika Sparkling Daily - Enterprise One-Year Operating Model
==============================================================
Single source of truth (Assumptions) -> formula-driven scenario P&Ls
(Low/Base/High/Stretch) -> Monthly phasing -> Dashboard -> Sensitivity ->
Checks, with Cover / Guide / Review Log governance tabs.

3 flavours (Pineapple, Passionfruit, Lime Lemon) x (Singles 12-ct + 4-Pack) = 6 SKUs.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName

OUT = "Organika_Sparkling_Daily_1YR_Scenario_Model.xlsx"
FY = "FY2026"
LINE = "Sparkling Daily"
VERSION = "v2.0"
BUILT = "2026-06-05"

FLAVOURS = ["Pineapple", "Passionfruit", "Lime Lemon"]
FORMATS = ["Singles", "4-Pack"]
SKU_COLS = ["C", "D", "E", "F", "G", "H"]          # 6 SKU columns
TOT = "I"                                           # total column
SKUS = [(fmt, fl) for fmt in FORMATS for fl in FLAVOURS]
SCEN = ["Low", "Base", "High", "Stretch"]

# ---- per-SKU base economics (Base scenario, CDN $, per case) -------------
# direct COGS = ingredients + packaging + manufacturing + freight_in
SKU_DATA = [
    # fmt, flavour, units, cp, disc%, ingr, pack, mfg, frt, wh, apc, apt, aps, vol
    # Singles = case of 12 single cans;  4-Pack = case of 6 four-packs (24 cans)
    ("Singles", "Pineapple",    12, 15.00, 0.07, 2.16, 2.64, 1.44, 0.29, 0.72, 0.36, 0.60, 0.24, 15000),
    ("Singles", "Passionfruit", 12, 15.00, 0.07, 2.16, 2.64, 1.44, 0.29, 0.72, 0.36, 0.60, 0.24, 11000),
    ("Singles", "Lime Lemon",   12, 14.00, 0.07, 2.06, 2.54, 1.40, 0.28, 0.70, 0.36, 0.60, 0.24, 9000),
    ("4-Pack",  "Pineapple",     6, 33.00, 0.08, 4.32, 7.08, 2.88, 0.58, 1.44, 0.72, 1.20, 0.48, 6000),
    ("4-Pack",  "Passionfruit",  6, 33.00, 0.08, 4.32, 7.08, 2.88, 0.58, 1.44, 0.72, 1.20, 0.48, 5000),
    ("4-Pack",  "Lime Lemon",    6, 31.00, 0.08, 4.12, 6.88, 2.80, 0.56, 1.40, 0.72, 1.20, 0.48, 4000),
]
SKU_KEYS = ["units", "cp", "disc", "ingr", "pack", "mfg", "frt", "wh", "apc", "apt", "aps", "vol"]

# ---- company operating assumptions (annual, Base) ------------------------
COMPANY = {
    "salaries": 180000, "ga": 60000, "mktg": 40000, "slotting": 20000,
    "log_pct": 0.04, "da": 25000, "interest": 10000, "tax": 0.27,
    "elast": 0.0,   # price elasticity of demand (advanced; 0 = off, e.g. -1.3 = elastic)
}
# ---- scenario levers: volume x, price x, trade-disc delta (decimal) -------
LEVERS = {
    "Low":     (0.70, 0.95,  0.01),
    "Base":    (1.00, 1.00,  0.00),
    "High":    (1.30, 1.00,  0.00),
    "Stretch": (1.60, 1.05, -0.01),
}
# ---- seasonality (% of annual volume), Jan..Dec, sums to 1.0 -------------
SEASON = [0.06, 0.06, 0.07, 0.08, 0.10, 0.11, 0.12, 0.11, 0.09, 0.08, 0.06, 0.06]
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# =====================================================================
#  styling
# =====================================================================
INK = "1F2A44"; BAND = "2E4374"; SECT = "D7E3F4"; INP = "FFF7CC"
TOTF = "EAF0FB"; WHITE = "FFFFFF"; GOOD = "1E7B34"; BAD = "B00020"; SUBF = "F2F6FC"

f_title = Font(name="Calibri", size=16, bold=True, color=INK)
f_h2 = Font(name="Calibri", size=12, bold=True, color=INK)
f_sub = Font(name="Calibri", size=10, italic=True, color="555555")
f_hdr = Font(name="Calibri", size=10, bold=True, color=WHITE)
f_sec = Font(name="Calibri", size=10, bold=True, color=INK)
f_lbl = Font(name="Calibri", size=10, color="222222")
f_lblb = Font(name="Calibri", size=10, bold=True, color=INK)
f_in = Font(name="Calibri", size=10, color="0033AA")
f_calc = Font(name="Calibri", size=10, color="222222")
f_good = Font(name="Calibri", size=10, bold=True, color=GOOD)
f_bad = Font(name="Calibri", size=10, bold=True, color=BAD)

fill_band = PatternFill("solid", fgColor=BAND)
fill_sec = PatternFill("solid", fgColor=SECT)
fill_in = PatternFill("solid", fgColor=INP)
fill_tot = PatternFill("solid", fgColor=TOTF)
fill_sub = PatternFill("solid", fgColor=SUBF)

thin = Side(style="thin", color="C9D2E0")
b_all = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center")
rght = Alignment(horizontal="right", vertical="center")

CASES = "#,##0"; MON = "$#,##0;[Red]($#,##0)"; MON2 = "$#,##0.00;[Red]($#,##0.00)"
PCT = "0.0%"; PCT2 = "0.00%"; MULT = "0.00\\x"

# tab colours (grouping)
TAB = {"Home": "0A84FF", "Cover": "808A99", "Guide": "808A99", "Review Log": "808A99",
       "Assumptions": "E8A33D", "Dashboard": "2E7D32",
       "Low": "5B6FA6", "Base": "2E4374", "High": "5B6FA6", "Stretch": "5B6FA6",
       "Monthly": "00838F", "Pricing Lab": "6A1B9A", "Targets": "6A1B9A",
       "Sensitivity": "EF6C00", "Checks": "B00020"}


def polish(ws, fit_w=1, landscape=True, title_rows=None):
    """Print setup + footer for a professional PDF/print output."""
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = fit_w
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_margins = openpyxl.worksheet.page.PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6)
    ws.oddFooter.left.text = f"{LINE} • {FY}"
    ws.oddFooter.center.text = "&A"
    ws.oddFooter.right.text = f"{VERSION} • Page &P of &N"
    if title_rows:
        ws.print_title_rows = title_rows

wb = openpyxl.Workbook()


def setcell(ws, coord, val, font=f_lbl, fmt=None, fill=None, align=None, border=False):
    c = ws[coord]
    c.value = val
    c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if align: c.alignment = align
    if border: c.border = b_all
    return c


def section(ws, row, text, span_to=TOT):
    ws.merge_cells(f"B{row}:{span_to}{row}")
    c = ws[f"B{row}"]; c.value = text; c.font = f_sec; c.alignment = left
    last = openpyxl.utils.column_index_from_string(span_to)
    for cc in range(2, last + 1):
        ws.cell(row=row, column=cc).fill = fill_sec


# =====================================================================
#  ASSUMPTIONS  (single source of truth) -> returns row map A
# =====================================================================
def build_assumptions(ws):
    ws.sheet_view.showGridLines = False
    A = {}
    setcell(ws, "B1", f"{LINE} - Assumptions (single source of truth)", f_title)
    setcell(ws, "B2", "Edit YELLOW cells only. Every scenario, the monthly view, sensitivity and checks read from this sheet.", f_sub)

    # SKU table header
    ws.merge_cells("C4:E4"); setcell(ws, "C4", "SINGLES", f_hdr, fill=fill_band, align=center)
    ws.merge_cells("F4:H4"); setcell(ws, "F4", "4-PACK", f_hdr, fill=fill_band, align=center)
    setcell(ws, "B5", "Per case (CDN $)", f_hdr, fill=fill_band, align=left, border=True)
    for i, (fmt, fl) in enumerate(SKUS):
        setcell(ws, f"{SKU_COLS[i]}5", fl, f_hdr, fill=fill_band, align=center, border=True)

    r = 6
    def line(label, key=None, fmt=MON2, calc=None, bold=False, pct=False):
        nonlocal r
        A[key or label] = r
        setcell(ws, f"B{r}", "   " + label, f_lblb if bold else f_lbl, align=left)
        for i, col in enumerate(SKU_COLS):
            d = dict(zip(SKU_KEYS, SKU_DATA[i][2:]))
            cell = ws[f"{col}{r}"]
            if calc:
                cell.value = calc(col)
                cell.font = f_calc; cell.fill = fill_tot
            else:
                cell.value = d[key]
                cell.font = f_in; cell.fill = fill_in
            cell.number_format = fmt; cell.alignment = rght; cell.border = b_all
        r += 1

    section(ws, r, "SKU DEFINITION"); r += 1
    # text rows: format & flavour
    A["fmt"] = r
    setcell(ws, f"B{r}", "   Format", f_lbl, align=left)
    for i, col in enumerate(SKU_COLS):
        setcell(ws, f"{col}{r}", SKUS[i][0], f_calc, align=center, border=True)
    r += 1
    A["flav"] = r
    setcell(ws, f"B{r}", "   Flavour", f_lbl, align=left)
    for i, col in enumerate(SKU_COLS):
        setcell(ws, f"{col}{r}", SKUS[i][1], f_calc, align=center, border=True)
    r += 1
    line("Units per case", "units", CASES)

    section(ws, r, "PRICE (per case)"); r += 1
    line("CP - List price / case", "cp", MON2, bold=True)
    line("Trade discount (% of gross)", "disc", PCT)
    A["ppu"] = r
    setcell(ws, f"B{r}", "   Price per unit (info)", f_lbl, align=left)
    for col in SKU_COLS:
        setcell(ws, f"{col}{r}", f"=IFERROR({col}{A['cp']}/{col}{A['units']},0)", f_calc, MON2, fill_tot, rght, True)
    r += 1

    section(ws, r, "COGS (per case)"); r += 1
    line("Ingredients", "ingr")
    line("Packaging", "pack")
    line("Manufacturing / Co-pack", "mfg")
    line("Inbound Freight", "frt")
    A["cogs_d"] = r
    setcell(ws, f"B{r}", "   COGS Direct / case", f_lblb, align=left)
    for col in SKU_COLS:
        setcell(ws, f"{col}{r}", f"=SUM({col}{A['ingr']}:{col}{A['frt']})", f_lblb, MON2, fill_tot, rght, True)
    r += 1
    line("Warehousing / Indirect", "wh")
    A["cos"] = r
    setcell(ws, f"B{r}", "   Cost of Sales / case", f_lblb, align=left)
    for col in SKU_COLS:
        setcell(ws, f"{col}{r}", f"={col}{A['cogs_d']}+{col}{A['wh']}", f_lblb, MON2, fill_tot, rght, True)
    r += 1

    section(ws, r, "A&P (per case)"); r += 1
    line("Consumer", "apc")
    line("Trade", "apt")
    line("Sales / Broker", "aps")
    A["ap"] = r
    setcell(ws, f"B{r}", "   A&P / case", f_lblb, align=left)
    for col in SKU_COLS:
        setcell(ws, f"{col}{r}", f"=SUM({col}{A['apc']}:{col}{A['aps']})", f_lblb, MON2, fill_tot, rght, True)
    r += 1

    section(ws, r, "VOLUME"); r += 1
    line("Base annual volume (cases)", "vol", CASES, bold=True)

    # ---- company operating assumptions (single column C) ----
    r += 1
    section(ws, r, "COMPANY OPERATING ASSUMPTIONS (annual)"); r += 1
    comp_rows = [
        ("Salaries & wages", "salaries", MON), ("G&A / overhead", "ga", MON),
        ("Marketing overhead (brand)", "mktg", MON), ("Slotting / listing fees", "slotting", MON),
        ("Outbound logistics (% of net sales)", "log_pct", PCT), ("Depreciation & amortisation", "da", MON),
        ("Interest expense", "interest", MON), ("Tax rate", "tax", PCT),
        ("Price elasticity (advanced; 0 = off)", "elast", "0.0"),
    ]
    for label, key, fmt in comp_rows:
        A["co_" + key] = r
        setcell(ws, f"B{r}", "   " + label, f_lbl, align=left)
        setcell(ws, f"C{r}", COMPANY[key], f_in, fmt, fill_in, rght, True)
        r += 1

    # ---- scenario levers ----
    r += 1
    section(ws, r, "SCENARIO LEVERS"); r += 1
    setcell(ws, f"B{r}", "   Scenario", f_hdr, fill=fill_band, align=left, border=True)
    setcell(ws, f"C{r}", "Volume x", f_hdr, fill=fill_band, align=center, border=True)
    setcell(ws, f"D{r}", "Price x", f_hdr, fill=fill_band, align=center, border=True)
    setcell(ws, f"E{r}", "Trade disc Δ", f_hdr, fill=fill_band, align=center, border=True)
    r += 1
    A["lever0"] = r
    for s in SCEN:
        v, p, d = LEVERS[s]
        setcell(ws, f"B{r}", "   " + s, f_lblb, align=left, border=True)
        setcell(ws, f"C{r}", v, f_in, MULT, fill_in, rght, True)
        setcell(ws, f"D{r}", p, f_in, MULT, fill_in, rght, True)
        setcell(ws, f"E{r}", d, f_in, PCT, fill_in, rght, True)
        r += 1

    # ---- seasonality ----
    r += 1
    section(ws, r, "SEASONALITY (% of annual volume)", span_to="N"); r += 1
    for j, m in enumerate(MONTHS):
        setcell(ws, f"{get_column_letter(3+j)}{r}", m, f_hdr, fill=fill_band, align=center, border=True)
    setcell(ws, f"O{r}", "Check", f_hdr, fill=fill_band, align=center, border=True)
    r += 1
    A["season"] = r
    for j in range(12):
        setcell(ws, f"{get_column_letter(3+j)}{r}", SEASON[j], f_in, PCT, fill_in, rght, True)
    setcell(ws, f"O{r}", f"=SUM(C{r}:N{r})", f_lblb, PCT, fill_tot, rght, True)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    for col in SKU_COLS + ["I", "J", "K", "L", "M", "N", "O"]:
        ws.column_dimensions[col].width = 12
    ws.freeze_panes = "C6"
    return A


# =====================================================================
#  SCENARIO P&L
# =====================================================================
SR = {  # scenario row map
    "title": 1, "sub": 2, "note": 3,
    "grp": 5, "sku": 6,
    "vh": 7, "cases": 8,
    "ph": 9, "cp": 10, "disc": 11, "cosd": 12, "cosi": 13, "ap_c": 14,
    "plh": 15, "grev": 16, "tdisc": 17, "nsales": 18, "cogsd": 19, "cogsi": 20,
    "cos": 21, "gp": 22, "gppct": 23, "ap": 24, "cm": 25, "cmpct": 26,
    "oph": 27, "sal": 28, "ga": 29, "mkt": 30, "slot": 31, "log": 32, "topex": 33,
    "ebitda": 34, "ebitdapct": 35, "da": 36, "ebit": 37, "int": 38, "ebt": 39,
    "tax": 40, "ni": 41, "nipct": 42,
    "kh": 43, "nspc": 44, "cmpc": 45,
}


def build_scenario(ws, name, A):
    ws.sheet_view.showGridLines = False
    v, p, d = LEVERS[name]
    idx = SCEN.index(name)
    lr = A["lever0"] + idx  # this scenario's lever row in Assumptions
    LV = f"'Assumptions'!$C${lr}"; LP = f"'Assumptions'!$D${lr}"; LD = f"'Assumptions'!$E${lr}"

    setcell(ws, "B1", f"{LINE} - {name.upper()} Scenario  ({FY})", f_title)
    setcell(ws, "B2", f"Levers vs Base:  Volume {v:.0%}  •  Price {p:.0%}  •  Trade-disc Δ {d:+.0%}   (all driven from Assumptions)", f_sub)
    setcell(ws, "B3", "CDN $.  All figures calculate from the Assumptions tab - nothing here is hand-keyed.", f_sub)

    ws.merge_cells(f"C{SR['grp']}:E{SR['grp']}"); setcell(ws, f"C{SR['grp']}", "SINGLES", f_hdr, fill=fill_band, align=center)
    ws.merge_cells(f"F{SR['grp']}:H{SR['grp']}"); setcell(ws, f"F{SR['grp']}", "4-PACK", f_hdr, fill=fill_band, align=center)
    ws.merge_cells(f"{TOT}{SR['grp']}:{TOT}{SR['sku']}"); setcell(ws, f"{TOT}{SR['grp']}", "TOTAL\nDAILY", f_hdr, fill=fill_band, align=center)
    setcell(ws, f"B{SR['sku']}", "CDN $", f_hdr, fill=fill_band, align=left, border=True)
    for i, (fmt, fl) in enumerate(SKUS):
        setcell(ws, f"{SKU_COLS[i]}{SR['sku']}", fl, f_hdr, fill=fill_band, align=center, border=True)

    def lab(row, text, bold=False):
        setcell(ws, f"B{row}", "   " + text, f_lblb if bold else f_lbl, align=left)

    section(ws, SR["vh"], "VOLUME"); lab(SR["cases"], "Physical Cases", True)
    section(ws, SR["ph"], "PRICE & COST - PER CASE (from Assumptions x levers)")
    lab(SR["cp"], "CP - Gross Rev / case", True); lab(SR["disc"], "Trade discount %")
    lab(SR["cosd"], "COGS Direct / case"); lab(SR["cosi"], "COGS Indirect / case"); lab(SR["ap_c"], "A&P / case")
    section(ws, SR["plh"], "P&L  ($)")
    for rk, t, bold in [("grev", "Gross Revenue", False), ("tdisc", "Trade Discounts", False),
                        ("nsales", "Net Sales", True), ("cogsd", "COGS Direct", False),
                        ("cogsi", "COGS Indirect", False), ("cos", "Cost of Sales", False),
                        ("gp", "Gross Profit", True), ("gppct", "GP %", False),
                        ("ap", "A&P", False), ("cm", "Contribution Margin", True), ("cmpct", "CM %", False)]:
        lab(SR[rk], t, bold)
    section(ws, SR["oph"], "OPERATING EXPENSES & PROFIT  (company total)")
    for rk, t, bold in [("sal", "Salaries & wages", False), ("ga", "G&A / overhead", False),
                        ("mkt", "Marketing overhead", False), ("slot", "Slotting / listing", False),
                        ("log", "Outbound logistics", False), ("topex", "Total Operating Expenses", True),
                        ("ebitda", "EBITDA", True), ("ebitdapct", "EBITDA %", False),
                        ("da", "Depreciation & amort.", False), ("ebit", "EBIT", True),
                        ("int", "Interest", False), ("ebt", "Pre-tax Profit (EBT)", True),
                        ("tax", "Tax", False), ("ni", "Net Income", True), ("nipct", "Net Income %", False)]:
        lab(SR[rk], t, bold)
    section(ws, SR["kh"], "PER-CASE KPIs"); lab(SR["nspc"], "Net Sales / case"); lab(SR["cmpc"], "Contribution / case")

    money_rows = ["grev", "tdisc", "nsales", "cogsd", "cogsi", "cos", "gp", "ap", "cm"]
    pct_rows = ["gppct", "cmpct"]

    # per-SKU columns
    for i, col in enumerate(SKU_COLS):
        ac = col  # Assumptions uses same column letters
        setcell(ws, f"{col}{SR['cases']}", f"='Assumptions'!{ac}{A['vol']}*{LV}*({LP}^'Assumptions'!$C${A['co_elast']})", f_calc, CASES, fill_tot, rght, True)
        setcell(ws, f"{col}{SR['cp']}", f"='Assumptions'!{ac}{A['cp']}*{LP}", f_calc, MON2, fill_tot, rght, True)
        setcell(ws, f"{col}{SR['disc']}", f"='Assumptions'!{ac}{A['disc']}+{LD}", f_calc, PCT, fill_tot, rght, True)
        setcell(ws, f"{col}{SR['cosd']}", f"='Assumptions'!{ac}{A['cogs_d']}", f_calc, MON2, fill_tot, rght, True)
        setcell(ws, f"{col}{SR['cosi']}", f"='Assumptions'!{ac}{A['wh']}", f_calc, MON2, fill_tot, rght, True)
        setcell(ws, f"{col}{SR['ap_c']}", f"='Assumptions'!{ac}{A['ap']}", f_calc, MON2, fill_tot, rght, True)
        f = {
            "grev": f"={col}{SR['cases']}*{col}{SR['cp']}",
            "tdisc": f"={col}{SR['grev']}*{col}{SR['disc']}",
            "nsales": f"={col}{SR['grev']}-{col}{SR['tdisc']}",
            "cogsd": f"={col}{SR['cases']}*{col}{SR['cosd']}",
            "cogsi": f"={col}{SR['cases']}*{col}{SR['cosi']}",
            "cos": f"={col}{SR['cogsd']}+{col}{SR['cogsi']}",
            "gp": f"={col}{SR['nsales']}-{col}{SR['cos']}",
            "gppct": f"=IFERROR({col}{SR['gp']}/{col}{SR['nsales']},0)",
            "ap": f"={col}{SR['cases']}*{col}{SR['ap_c']}",
            "cm": f"={col}{SR['gp']}-{col}{SR['ap']}",
            "cmpct": f"=IFERROR({col}{SR['cm']}/{col}{SR['nsales']},0)",
            "nspc": f"=IFERROR({col}{SR['nsales']}/{col}{SR['cases']},0)",
            "cmpc": f"=IFERROR({col}{SR['cm']}/{col}{SR['cases']},0)",
        }
        for rk, formula in f.items():
            fmt = PCT if rk in pct_rows else (MON2 if rk in ("nspc", "cmpc") else MON)
            setcell(ws, f"{col}{SR[rk]}", formula, f_calc, fmt, None, rght, True)

    # total column
    T = TOT
    setcell(ws, f"{T}{SR['cases']}", f"=SUM(C{SR['cases']}:H{SR['cases']})", f_lblb, CASES, fill_tot, rght, True)
    for rk, src in [("cp", "grev"), ("disc", "tdisc"), ("cosd", "cogsd"), ("cosi", "cogsi"), ("ap_c", "ap")]:
        base = "grev" if rk == "disc" else "cases"
        setcell(ws, f"{T}{SR[rk]}", f"=IFERROR({T}{SR[src]}/{T}{SR[base]},0)", f_lblb,
                PCT if rk == "disc" else MON2, fill_tot, rght, True)
    for rk in ["grev", "tdisc", "cogsd", "cogsi", "ap"]:
        setcell(ws, f"{T}{SR[rk]}", f"=SUM(C{SR[rk]}:H{SR[rk]})", f_lblb, MON, fill_tot, rght, True)
    tf = {
        "nsales": f"={T}{SR['grev']}-{T}{SR['tdisc']}",
        "cos": f"={T}{SR['cogsd']}+{T}{SR['cogsi']}",
        "gp": f"={T}{SR['nsales']}-{T}{SR['cos']}",
        "gppct": f"=IFERROR({T}{SR['gp']}/{T}{SR['nsales']},0)",
        "cm": f"={T}{SR['gp']}-{T}{SR['ap']}",
        "cmpct": f"=IFERROR({T}{SR['cm']}/{T}{SR['nsales']},0)",
        # opex (company total)
        "sal": f"='Assumptions'!$C${A['co_salaries']}",
        "ga": f"='Assumptions'!$C${A['co_ga']}",
        "mkt": f"='Assumptions'!$C${A['co_mktg']}",
        "slot": f"='Assumptions'!$C${A['co_slotting']}",
        "log": f"={T}{SR['nsales']}*'Assumptions'!$C${A['co_log_pct']}",
        "topex": f"=SUM({T}{SR['sal']}:{T}{SR['log']})",
        "ebitda": f"={T}{SR['cm']}-{T}{SR['topex']}",
        "ebitdapct": f"=IFERROR({T}{SR['ebitda']}/{T}{SR['nsales']},0)",
        "da": f"='Assumptions'!$C${A['co_da']}",
        "ebit": f"={T}{SR['ebitda']}-{T}{SR['da']}",
        "int": f"='Assumptions'!$C${A['co_interest']}",
        "ebt": f"={T}{SR['ebit']}-{T}{SR['int']}",
        "tax": f"=MAX(0,{T}{SR['ebt']}*'Assumptions'!$C${A['co_tax']})",
        "ni": f"={T}{SR['ebt']}-{T}{SR['tax']}",
        "nipct": f"=IFERROR({T}{SR['ni']}/{T}{SR['nsales']},0)",
        "nspc": f"=IFERROR({T}{SR['nsales']}/{T}{SR['cases']},0)",
        "cmpc": f"=IFERROR({T}{SR['cm']}/{T}{SR['cases']},0)",
    }
    for rk, formula in tf.items():
        fmt = PCT if rk in ("gppct", "cmpct", "ebitdapct", "nipct") else (MON2 if rk in ("nspc", "cmpc") else MON)
        setcell(ws, f"{T}{SR[rk]}", formula, f_lblb, fmt, fill_tot, rght, True)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    for col in SKU_COLS: ws.column_dimensions[col].width = 12
    ws.column_dimensions[TOT].width = 14
    ws.freeze_panes = "C7"
    ws.row_dimensions[SR["grp"]].height = 16
    ws.row_dimensions[SR["sku"]].height = 26


# =====================================================================
#  MONTHLY PHASING  (scenario selectable)
# =====================================================================
def build_monthly(ws, A):
    ws.sheet_view.showGridLines = False
    setcell(ws, "B1", f"{LINE} - Monthly Phasing ({FY})", f_title)
    setcell(ws, "B2", "Pick a scenario; volume-driven lines follow the seasonality curve, fixed costs spread evenly.", f_sub)
    setcell(ws, "B4", "Scenario:", f_lblb, align=left)
    sel = setcell(ws, "C4", "Base", f_in, fill=fill_in, align=center, border=True)
    dv = DataValidation(type="list", formula1='"Low,Base,High,Stretch"', allow_blank=False)
    ws.add_data_validation(dv); dv.add(sel)
    setcell(ws, "D4", "← change me", f_sub, align=left)

    hdr = 6
    setcell(ws, f"B{hdr}", "CDN $", f_hdr, fill=fill_band, align=left, border=True)
    for j, m in enumerate(MONTHS):
        setcell(ws, f"{get_column_letter(3+j)}{hdr}", m, f_hdr, fill=fill_band, align=center, border=True)
    setcell(ws, f"O{hdr}", "FY", f_hdr, fill=fill_band, align=center, border=True)

    S = "INDIRECT(\"'\"&$C$4&\"'!" + TOT  # prefix for selected scenario total cells
    def scen(rk):  # annual value of selected scenario's total
        return f"{S}{SR[rk]}\")"

    rows = []
    r = hdr + 1
    def add(label, kind, expr_builder, fmt=MON, bold=False):
        nonlocal r
        setcell(ws, f"B{r}", "   " + label, f_lblb if bold else f_lbl, align=left)
        for j in range(12):
            mc = get_column_letter(3 + j)
            setcell(ws, f"{mc}{r}", expr_builder(mc, j), f_calc, fmt, None, rght, True)
        setcell(ws, f"O{r}", f"=SUM(C{r}:N{r})", f_lblb, fmt, fill_tot, rght, True)
        rows.append((label, r)); r += 1
        return r - 1

    seasref = lambda mc: f"'Assumptions'!{mc}${A['season']}"
    r_cases = add("Physical Cases", "v", lambda mc, j: f"={scen('cases')}*{seasref(mc)}", CASES, True)
    r_net = add("Net Sales", "v", lambda mc, j: f"={scen('nsales')}*{seasref(mc)}", MON, True)
    r_gp = add("Gross Profit", "v", lambda mc, j: f"={scen('gp')}*{seasref(mc)}", MON)
    r_ap = add("A&P", "v", lambda mc, j: f"={scen('ap')}*{seasref(mc)}", MON)
    r_cm = add("Contribution Margin", "v", lambda mc, j: f"={scen('cm')}*{seasref(mc)}", MON, True)
    r_log = add("Outbound logistics", "v", lambda mc, j: f"={mc}{r_net}*'Assumptions'!$C${A['co_log_pct']}", MON)
    fixed = f"('Assumptions'!$C${A['co_salaries']}+'Assumptions'!$C${A['co_ga']}+'Assumptions'!$C${A['co_mktg']}+'Assumptions'!$C${A['co_slotting']})/12"
    r_fix = add("Fixed opex (salaries, G&A, mktg, slotting)", "f", lambda mc, j: f"={fixed}", MON)
    r_ebitda = add("EBITDA", "c", lambda mc, j: f"={mc}{r_cm}-{mc}{r_log}-{mc}{r_fix}", MON, True)
    r_da = add("Depreciation & amort.", "f", lambda mc, j: f"='Assumptions'!$C${A['co_da']}/12", MON)
    r_ebit = add("EBIT", "c", lambda mc, j: f"={mc}{r_ebitda}-{mc}{r_da}", MON, True)
    r_int = add("Interest", "f", lambda mc, j: f"='Assumptions'!$C${A['co_interest']}/12", MON)
    r_ebt = add("Pre-tax Profit", "c", lambda mc, j: f"={mc}{r_ebit}-{mc}{r_int}", MON, True)
    r_tax = add("Tax", "c", lambda mc, j: f"=MAX(0,{mc}{r_ebt}*'Assumptions'!$C${A['co_tax']})", MON)
    r_ni = add("Net Income", "c", lambda mc, j: f"={mc}{r_ebt}-{mc}{r_tax}", MON, True)
    # cumulative net income
    r_cum = r_ni + 1
    setcell(ws, f"B{r_cum}", "   Cumulative Net Income", f_lblb, align=left)
    setcell(ws, f"C{r_cum}", f"=C{r_ni}", f_calc, MON, None, rght, True)
    for j in range(1, 12):
        mc = get_column_letter(3 + j); pc = get_column_letter(2 + j)
        setcell(ws, f"{mc}{r_cum}", f"={pc}{r_cum}+{mc}{r_ni}", f_calc, MON, None, rght, True)
    setcell(ws, f"O{r_cum}", f"=N{r_cum}", f_lblb, MON, fill_tot, rght, True)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    for j in range(12): ws.column_dimensions[get_column_letter(3+j)].width = 10
    ws.column_dimensions["O"].width = 13
    ws.freeze_panes = "C7"


# =====================================================================
#  DASHBOARD  (clean comparison table)
# =====================================================================
def build_dashboard(ws):
    ws.sheet_view.showGridLines = False
    setcell(ws, "B1", f"{LINE} - Scenario Dashboard ({FY})", f_title)
    setcell(ws, "B2", "Company totals across the four scenarios. Driven live from each scenario tab.", f_sub)
    hdr = 4
    setcell(ws, f"B{hdr}", "Metric", f_hdr, fill=fill_band, align=left, border=True)
    for j, s in enumerate(SCEN):
        setcell(ws, f"{get_column_letter(3+j)}{hdr}", s, f_hdr, fill=fill_band, align=center, border=True)
    metrics = [
        ("Physical Cases", "cases", CASES), ("Gross Revenue", "grev", MON),
        ("Trade Discounts", "tdisc", MON), ("Net Sales", "nsales", MON),
        ("Gross Profit", "gp", MON), ("GP %", "gppct", PCT),
        ("A&P", "ap", MON), ("Contribution Margin", "cm", MON), ("CM %", "cmpct", PCT),
        ("Total Opex", "topex", MON), ("EBITDA", "ebitda", MON), ("EBITDA %", "ebitdapct", PCT),
        ("Net Income", "ni", MON), ("Net Income %", "nipct", PCT),
        ("Blended CP / case", "cp", MON2), ("Net Sales / case", "nspc", MON2),
    ]
    first = hdr + 1
    for i, (name, rk, fmt) in enumerate(metrics):
        row = first + i
        bold = name in ("Net Sales", "Gross Profit", "Contribution Margin", "EBITDA", "Net Income")
        setcell(ws, f"B{row}", name, f_lblb if bold else f_lbl, align=left, border=True)
        for j, s in enumerate(SCEN):
            c = setcell(ws, f"{get_column_letter(3+j)}{row}", f"='{s}'!{TOT}{SR[rk]}", f_calc, fmt, None, rght, True)
            if s == "Base": c.fill = fill_tot
    ws._dash_rows = {rk: first + i for i, (_, rk, _) in enumerate(metrics)}

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 24
    for j in range(len(SCEN)): ws.column_dimensions[get_column_letter(3+j)].width = 14
    ws.freeze_panes = "C5"


# =====================================================================
#  SENSITIVITY  (two-way EBITDA + break-even)
# =====================================================================
def build_sensitivity(ws, A):
    ws.sheet_view.showGridLines = False
    setcell(ws, "B1", f"{LINE} - Sensitivity & Break-even ({FY})", f_title)
    setcell(ws, "B2", "EBITDA as price and volume flex around Base. Exact closed-form from Assumptions.", f_sub)

    # helper constants
    volr = f"'Assumptions'!C{A['vol']}:H{A['vol']}"
    cpr = f"'Assumptions'!C{A['cp']}:H{A['cp']}"
    dr = f"'Assumptions'!C{A['disc']}:H{A['disc']}"
    cosr = f"'Assumptions'!C{A['cos']}:H{A['cos']}"
    apr = f"'Assumptions'!C{A['ap']}:H{A['ap']}"
    setcell(ws, "B4", "K1  Base gross-less-disc (Σ vol·cp·(1-d))", f_lbl, align=left)
    setcell(ws, "C4", f"=SUMPRODUCT({volr},{cpr},(1-{dr}))", f_calc, MON, fill_tot, rght, True)
    setcell(ws, "B5", "K2  Base variable cost (Σ vol·(COS+A&P))", f_lbl, align=left)
    setcell(ws, "C5", f"=SUMPRODUCT({volr},({cosr}+{apr}))", f_calc, MON, fill_tot, rght, True)
    setcell(ws, "B6", "Fixed opex (sal+G&A+mktg+slot)", f_lbl, align=left)
    setcell(ws, "C6", f"='Assumptions'!C{A['co_salaries']}+'Assumptions'!C{A['co_ga']}+'Assumptions'!C{A['co_mktg']}+'Assumptions'!C{A['co_slotting']}", f_calc, MON, fill_tot, rght, True)
    setcell(ws, "B7", "Logistics % of net", f_lbl, align=left)
    setcell(ws, "C7", f"='Assumptions'!C{A['co_log_pct']}", f_calc, PCT, fill_tot, rght, True)
    setcell(ws, "B8", "Base total cases", f_lbl, align=left)
    setcell(ws, "C8", f"=SUM({volr})", f_calc, CASES, fill_tot, rght, True)
    setcell(ws, "B9", "Price elasticity (from Assumptions)", f_lbl, align=left)
    setcell(ws, "C9", f"='Assumptions'!C{A['co_elast']}", f_calc, "0.0", fill_tot, rght, True)
    K1, K2, FX, LOG, E = "$C$4", "$C$5", "$C$6", "$C$7", "$C$9"

    # two-way table
    vol_mults = [0.70, 0.90, 1.00, 1.10, 1.30, 1.50, 1.70]
    price_mults = [0.90, 0.95, 1.00, 1.05, 1.10, 1.15]
    top = 11
    setcell(ws, f"B{top}", "EBITDA  ($)", f_h2, align=left)
    setcell(ws, f"B{top+1}", "Price x  \\  Volume x", f_hdr, fill=fill_band, align=center, border=True)
    for j, mv in enumerate(vol_mults):
        c = setcell(ws, f"{get_column_letter(3+j)}{top+1}", mv, f_hdr, MULT, fill_band, center, True)
    for i, mp in enumerate(price_mults):
        rr = top + 2 + i
        setcell(ws, f"B{rr}", mp, f_hdr, MULT, fill_band, center, True)
        for j, mv in enumerate(vol_mults):
            mcv = f"{get_column_letter(3+j)}${top+1}"   # volume header (row fixed)
            mpr = f"$B{rr}"                                # price header (col fixed)
            # EBITDA = mv·mp^E·(mp·K1−K2) − Fixed − Log%·mv·mp^(E+1)·K1   (E=0 ⇒ linear)
            formula = (f"={mcv}*({mpr}^{E})*({mpr}*{K1}-{K2})-{FX}-{LOG}*{mcv}*({mpr}^({E}+1))*{K1}")
            c = setcell(ws, f"{get_column_letter(3+j)}{rr}", formula, f_calc, MON, None, rght, True)
            if abs(mv - 1.0) < 1e-9 and abs(mp - 1.0) < 1e-9:
                c.fill = fill_tot

    # break-even
    be = top + 2 + len(price_mults) + 2
    setcell(ws, f"B{be}", "BREAK-EVEN (at Base price)", f_h2, align=left)
    setcell(ws, f"B{be+1}", "Volume x for EBITDA = 0", f_lbl, align=left)
    setcell(ws, f"C{be+1}", f"=IFERROR({FX}/(({K1}-{K2})-{LOG}*{K1}),0)", f_calc, MULT, fill_tot, rght, True)
    setcell(ws, f"B{be+2}", "Break-even cases", f_lbl, align=left)
    setcell(ws, f"C{be+2}", f"=C{be+1}*{'$C$8'}", f_calc, CASES, fill_tot, rght, True)
    setcell(ws, f"B{be+3}", "Base case volume", f_lbl, align=left)
    setcell(ws, f"C{be+3}", "=$C$8", f_calc, CASES, fill_tot, rght, True)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    for j in range(len(vol_mults)): ws.column_dimensions[get_column_letter(3+j)].width = 13


# =====================================================================
#  CHECKS
# =====================================================================
def build_checks(ws, A):
    ws.sheet_view.showGridLines = False
    setcell(ws, "B1", f"{LINE} - Model Checks", f_title)
    setcell(ws, "B2", "Automated validation. Everything should read PASS.", f_sub)
    hdr = 4
    for col, t in [("B", "Check"), ("C", "Result"), ("D", "Status")]:
        setcell(ws, f"{col}{hdr}", t, f_hdr, fill=fill_band, align=left if col == "B" else center, border=True)
    checks = [
        ("Seasonality sums to 100%", f"=SUM('Assumptions'!C{A['season']}:N{A['season']})",
         f"=IF(ABS(SUM('Assumptions'!C{A['season']}:N{A['season']})-1)<0.0001,\"PASS\",\"FAIL\")", PCT),
        ("All CP / case > 0", f"=MIN('Assumptions'!C{A['cp']}:H{A['cp']})",
         f"=IF(MIN('Assumptions'!C{A['cp']}:H{A['cp']})>0,\"PASS\",\"FAIL\")", MON2),
        ("All volumes >= 0", f"=MIN('Assumptions'!C{A['vol']}:H{A['vol']})",
         f"=IF(MIN('Assumptions'!C{A['vol']}:H{A['vol']})>=0,\"PASS\",\"FAIL\")", CASES),
        ("Trade discount within 0-50%", f"=MAX('Assumptions'!C{A['disc']}:H{A['disc']})",
         f"=IF(AND(MIN('Assumptions'!C{A['disc']}:H{A['disc']})>=0,MAX('Assumptions'!C{A['disc']}:H{A['disc']})<=0.5),\"PASS\",\"FAIL\")", PCT),
        ("Base GP% between 0 and 100%", f"='Base'!{TOT}{SR['gppct']}",
         f"=IF(AND('Base'!{TOT}{SR['gppct']}>0,'Base'!{TOT}{SR['gppct']}<1),\"PASS\",\"FAIL\")", PCT),
        ("COGS Direct < CP (each SKU)", "=SUMPRODUCT(--('Assumptions'!C%d:H%d>='Assumptions'!C%d:H%d))" % (A['cogs_d'], A['cogs_d'], A['cp'], A['cp']),
         "=IF(SUMPRODUCT(--('Assumptions'!C%d:H%d>='Assumptions'!C%d:H%d))=0,\"PASS\",\"FAIL\")" % (A['cogs_d'], A['cogs_d'], A['cp'], A['cp']), CASES),
        ("Tax rate within 0-50%", f"='Assumptions'!C{A['co_tax']}",
         f"=IF(AND('Assumptions'!C{A['co_tax']}>=0,'Assumptions'!C{A['co_tax']}<=0.5),\"PASS\",\"FAIL\")", PCT),
        ("All scenario volume levers > 0", f"=MIN('Assumptions'!C{A['lever0']}:C{A['lever0']+3})",
         f"=IF(MIN('Assumptions'!C{A['lever0']}:C{A['lever0']+3})>0,\"PASS\",\"FAIL\")", MULT),
        ("Price elasticity <= 0 (demand sane)", f"='Assumptions'!C{A['co_elast']}",
         f"=IF('Assumptions'!C{A['co_elast']}<=0,\"PASS\",\"FAIL\")", "0.0"),
        ("Dashboard Net Sales ties to Base tab", f"='Dashboard'!D{4+4}-'Base'!{TOT}{SR['nsales']}",
         f"=IF(ABS('Dashboard'!D{4+4}-'Base'!{TOT}{SR['nsales']})<1,\"PASS\",\"FAIL\")", MON),
    ]
    for i, (name, res, status, fmt) in enumerate(checks):
        rr = hdr + 1 + i
        setcell(ws, f"B{rr}", name, f_lbl, align=left, border=True)
        setcell(ws, f"C{rr}", res, f_calc, fmt, None, rght, True)
        sc = setcell(ws, f"D{rr}", status, f_good, align=center, border=True)
    # overall
    last = hdr + len(checks)
    setcell(ws, f"B{last+2}", "OVERALL", f_lblb, align=left)
    setcell(ws, f"D{last+2}", f'=IF(COUNTIF(D{hdr+1}:D{last},"FAIL")=0,"ALL PASS","REVIEW")', f_good, align=center, border=True)
    # green PASS / red FAIL|REVIEW
    rng = f"D{hdr+1}:D{last+2}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("PASS",D{hdr+1}))'],
        fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100", bold=True)))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("FAIL",D{hdr+1})),ISNUMBER(SEARCH("REVIEW",D{hdr+1})))'],
        fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006", bold=True)))
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12


# =====================================================================
#  COVER / GUIDE / REVIEW LOG
# =====================================================================
def build_cover(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 60
    setcell(ws, "B2", "ORGANIKA", f_title)
    setcell(ws, "B3", f"{LINE} — One-Year Operating Model", f_h2)

    # ---- live KPI banner (Base scenario) ----
    setcell(ws, "B5", f"BASE-CASE SNAPSHOT ({FY})", f_sec, fill=fill_sec, align=left)
    for cc in "CDEFG":
        ws[f"{cc}5"].fill = fill_sec
    kpis = [
        ("Net Sales", f"='Base'!{TOT}{SR['nsales']}", MON),
        ("Gross Profit %", f"='Base'!{TOT}{SR['gppct']}", PCT),
        ("EBITDA", f"='Base'!{TOT}{SR['ebitda']}", MON),
        ("Net Income", f"='Base'!{TOT}{SR['ni']}", MON),
        ("Total Cases", f"='Base'!{TOT}{SR['cases']}", CASES),
        ("Break-even Cases", "='Sensitivity'!$C$23", CASES),
    ]
    f_kpi = Font(name="Calibri", size=14, bold=True, color=INK)
    for j, (label, formula, fmt) in enumerate(kpis):
        col = get_column_letter(2 + j)
        setcell(ws, f"{col}6", label, f_hdr, fill=fill_band, align=center, border=True)
        setcell(ws, f"{col}7", formula, f_kpi, fmt, fill_tot, center, True)
    ws.row_dimensions[7].height = 26
    for cc in "BCDEFG":
        ws.column_dimensions[cc].width = 16
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 60

    meta = [
        ("Purpose", "Model FY net sales, margin and EBITDA for Sparkling Daily across pricing & volume scenarios."),
        ("Scope", "3 flavours (Pineapple, Passionfruit, Lime Lemon) × Singles (12-ct) + 4-Pack = 6 SKUs; one fiscal year."),
        ("Fiscal year", FY), ("Version", VERSION), ("Built", BUILT),
        ("Currency", "CDN $"), ("Owner", "Commercial Finance"),
        ("Status", "DRAFT — figures are placeholders pending real inputs"),
    ]
    r = 9
    for k, v in meta:
        setcell(ws, f"B{r}", k, f_lblb, align=left); setcell(ws, f"C{r}", v, f_lbl, align=left); r += 1
    r += 1
    setcell(ws, f"B{r}", "TAB INDEX", f_sec, fill=fill_sec, align=left)
    setcell(ws, f"C{r}", "", f_lbl, fill=fill_sec); r += 1
    index = [
        ("Guide", "How to use + glossary of terms"),
        ("Review Log", "Devil's-advocate critique and how each issue was resolved"),
        ("Assumptions", "SINGLE SOURCE OF TRUTH — edit yellow cells here"),
        ("Dashboard", "Scenario comparison (table)"),
        ("Low / Base / High / Stretch", "Full P&L per scenario (6 SKUs + total → Net Income)"),
        ("Monthly", "Seasonalised monthly P&L for a selected scenario"),
        ("Sensitivity", "Two-way EBITDA table + break-even volume"),
        ("Checks", "Automated validation (should read ALL PASS)"),
    ]
    for k, v in index:
        setcell(ws, f"B{r}", k, f_lblb, align=left); setcell(ws, f"C{r}", v, f_lbl, align=left); r += 1
    r += 1
    setcell(ws, f"B{r}", "LEGEND", f_sec, fill=fill_sec, align=left); setcell(ws, f"C{r}", "", fill=fill_sec); r += 1
    setcell(ws, f"B{r}", "Yellow cell", f_in, fill=fill_in, align=center, border=True)
    setcell(ws, f"C{r}", "Input — safe to edit", f_lbl, align=left); r += 1
    setcell(ws, f"B{r}", "Blue tint", f_lblb, fill=fill_tot, align=center, border=True)
    setcell(ws, f"C{r}", "Calculated total — do not overwrite", f_lbl, align=left); r += 1


def build_guide(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 110
    lines = [
        (f"{LINE} — Model Guide", f_title), ("", None),
        ("START HERE", f_sec),
        ("Open the Home tab. Pick a scenario from the dropdown and the KPI cards update instantly.", f_lbl),
        ("To change the numbers, go to Assumptions and edit the yellow cells. That's the only place you type.", f_lbl),
        ("Advanced (optional): set Price elasticity on Assumptions to a negative number (e.g. -1.3) to make", f_lbl),
        ("volume respond to price; the Pricing Lab then highlights the wholesale price that earns the most.", f_lbl),
        ("", None),
        ("HOW IT FITS TOGETHER", f_sec),
        ("Assumptions is the only place you key numbers. Scenario tabs, Monthly, Dashboard, Sensitivity and Checks all read from it.", f_lbl),
        ("Change a price, cost or volume once on Assumptions and the entire workbook updates.", f_lbl),
        ("", None),
        ("STEP BY STEP", f_sec),
        ("1. On Assumptions, set each SKU's CP, trade discount, COGS build-up, A&P and base volume (yellow cells).", f_lbl),
        ("2. Set company operating assumptions (salaries, G&A, logistics %, D&A, interest, tax).", f_lbl),
        ("3. Set the four scenario levers (volume ×, price ×, trade-disc Δ) and the seasonality curve.", f_lbl),
        ("4. Read results on Dashboard; drill into Low/Base/High/Stretch; phase a scenario on Monthly.", f_lbl),
        ("5. Stress-test on Sensitivity; confirm Checks reads ALL PASS.", f_lbl),
        ("", None),
        ("P&L LOGIC (per SKU, then company)", f_sec),
        ("Gross Revenue = Cases × CP/case", f_lbl),
        ("Net Sales = Gross Revenue − Trade Discounts   (discount = % of gross)", f_lbl),
        ("Gross Profit = Net Sales − Cost of Sales        Cost of Sales = COGS Direct + Indirect", f_lbl),
        ("Contribution Margin = Gross Profit − A&P", f_lbl),
        ("EBITDA = Σ Contribution − Operating Expenses (fixed + variable logistics)", f_lbl),
        ("EBIT = EBITDA − D&A;  Pre-tax = EBIT − Interest;  Net Income = Pre-tax − Tax (tax floored at 0).", f_lbl),
        ("", None),
        ("GLOSSARY", f_sec),
        ("CP — 'Cost Price' / list price per case charged to the customer (your wholesale price).", f_lbl),
        ("Case — a shipping case; Singles (12 cans) and 4-Pack (6 four-packs) hold different unit counts.", f_lbl),
        ("Trade discount — off-invoice allowances/promo funding, expressed as % of gross revenue.", f_lbl),
        ("A&P — Advertising & Promotion (consumer + trade + sales/broker).", f_lbl),
        ("Contribution Margin — profit a SKU contributes before company fixed costs.", f_lbl),
        ("EBITDA — earnings before interest, tax, depreciation & amortisation.", f_lbl),
        ("", None),
        ("CAVEATS", f_sec),
        ("All loaded numbers are illustrative placeholders. Replace with real figures before relying on output.", f_lbl),
        ("Single-currency (CDN $), one fiscal year. No balance sheet / cash flow — this is an operating P&L.", f_lbl),
    ]
    r = 1
    for text, font in lines:
        c = ws[f"B{r}"]; c.value = text
        if font: c.font = font
        c.alignment = left; r += 1


def build_review_log(ws):
    ws.sheet_view.showGridLines = False
    setcell(ws, "B1", f"{LINE} — Review Log (devil's advocate)", f_title)
    setcell(ws, "B2", "Issues identified in the first-pass model and how each was resolved in this build.", f_sub)
    hdr = 4
    for col, t, w in [("B", "#", 4), ("C", "Issue / risk", 52), ("D", "Resolution", 60), ("E", "Status", 12)]:
        setcell(ws, f"{col}{hdr}", t, f_hdr, fill=fill_band, align=left, border=True)
        ws.column_dimensions[col].width = w
    items = [
        ("No single source of truth — costs hand-keyed in ~24 places across tabs.",
         "Added Assumptions tab; all scenarios/monthly/sensitivity read from it (DRY)."),
        ("Trade discount modelled as flat $/case — unrealistic.",
         "Now % of gross revenue, per SKU, flexed by scenario."),
        ("COGS was a single lump — no cost visibility.",
         "Built up from ingredients, packaging, co-pack, inbound freight + warehousing."),
        ("P&L stopped at contribution — not a real P&L.",
         "Extended to Opex → EBITDA → EBIT → Pre-tax → Net Income."),
        ("No operating leverage — all costs implicitly variable.",
         "Split fixed (salaries/G&A/mktg/slotting) vs variable (logistics % of net)."),
        ("Annual only — no phasing.",
         "Added Monthly tab driven by a seasonality curve."),
        ("No way to switch scenario in one view.",
         "Monthly has a dropdown selector (INDIRECT-driven)."),
        ("No sensitivity or break-even.",
         "Two-way EBITDA table (price × volume) + break-even volume, closed-form."),
        ("No validation of inputs/outputs.",
         "Checks tab with PASS/FAIL incl. a tie-out of Dashboard to Base."),
        ("No governance / context.",
         "Cover (purpose, version, owner, legend), Guide + glossary, this Review Log."),
        ("Tax could show a fictitious credit on losses.",
         "Tax floored at 0 via MAX(0, EBT×rate)."),
        ("Inputs vs calcs indistinguishable.",
         "Colour coding: yellow = input, blue tint = calculated total; freeze panes throughout."),
        ("Flavour names were placeholders.",
         "Set to Pineapple, Passionfruit, Lime Lemon (editable on Assumptions)."),
        ("Price and volume were independent — raising CP with no volume loss is unrealistic.",
         "Added optional price elasticity; volume now responds to price. Pricing Lab finds the best CP."),
        ("Workbook opened on a dense tab — not approachable.",
         "Added a clean Home screen (scenario picker + live KPI cards + guided navigation); simple-first tab order."),
    ]
    for i, (issue, res) in enumerate(items):
        rr = hdr + 1 + i
        setcell(ws, f"B{rr}", i + 1, f_lbl, align=center, border=True)
        setcell(ws, f"C{rr}", issue, f_lbl, align=left, border=True).alignment = Alignment(wrap_text=True, vertical="top")
        setcell(ws, f"D{rr}", res, f_lbl, align=left, border=True).alignment = Alignment(wrap_text=True, vertical="top")
        setcell(ws, f"E{rr}", "Resolved", f_good, align=center, border=True)
    ws.column_dimensions["A"].width = 2


# =====================================================================
#  PRICING LAB  (per-SKU CP ladder)
# =====================================================================
def build_pricing(ws, A):
    ws.sheet_view.showGridLines = False
    setcell(ws, "B1", f"{LINE} - Pricing Lab ({FY})", f_title)
    setcell(ws, "B2", "How wholesale price (CP) drives margin. Each column flexes every SKU's CP vs its Assumptions base; volume & costs held at base.", f_sub)

    ladders = [0.85, 0.90, 0.95, 1.00, 1.05, 1.10, 1.15]
    base_col_idx = ladders.index(1.00)

    def table(top, title, kind, fmt):
        setcell(ws, f"B{top}", title, f_h2, align=left)
        hr = top + 1
        setcell(ws, f"B{hr}", "SKU  \\  CP vs base", f_hdr, fill=fill_band, align=left, border=True)
        for j, m in enumerate(ladders):
            setcell(ws, f"{get_column_letter(3+j)}{hr}", m, f_hdr, MULT, fill_band, center, True)
        for i, col in enumerate(SKU_COLS):
            rr = hr + 1 + i
            nm = f"{SKUS[i][1]} {SKUS[i][0]}"
            setcell(ws, f"B{rr}", "   " + nm, f_lbl, align=left, border=True)
            cp = f"'Assumptions'!${col}${A['cp']}"
            disc = f"'Assumptions'!${col}${A['disc']}"
            cos = f"'Assumptions'!${col}${A['cos']}"
            ap = f"'Assumptions'!${col}${A['ap']}"
            vol = f"'Assumptions'!${col}${A['vol']}"
            for j, m in enumerate(ladders):
                mc = f"{get_column_letter(3+j)}${hr}"
                netpc = f"({cp}*{mc}*(1-{disc}))"   # net price per case
                if kind == "gppct":
                    f = f"=IFERROR(1-{cos}/{netpc},0)"
                elif kind == "contrib":
                    f = f"={vol}*({mc}^'Assumptions'!$C${A['co_elast']})*({netpc}-{cos}-{ap})"
                else:  # gp per case
                    f = f"={netpc}-{cos}"
                c = setcell(ws, f"{get_column_letter(3+j)}{rr}", f, f_calc, fmt, None, rght, True)
                if j == base_col_idx:
                    c.fill = fill_tot
        return hr + 1, hr + len(SKU_COLS)  # data first/last rows

    f1, l1 = table(4, "Gross Profit %  (by SKU, by CP)", "gppct", PCT)
    t2 = l1 + 3
    f2, l2 = table(t2, "Annual Contribution $  (by SKU, by CP — demand-adjusted)", "contrib", MON)
    # "Best CP in range" insight column (price that maximises contribution within the ladder)
    hr2 = f2 - 1
    setcell(ws, f"K{hr2}", "Best CP ×", f_hdr, fill=fill_band, align=center, border=True)
    for i in range(len(SKU_COLS)):
        rr = f2 + i
        setcell(ws, f"K{rr}", f"=INDEX($C${hr2}:$I${hr2},MATCH(MAX(C{rr}:I{rr}),C{rr}:I{rr},0))",
                f_lblb, MULT, fill_tot, center, True)
    ws.column_dimensions["K"].width = 11

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    for j in range(len(ladders)):
        ws.column_dimensions[get_column_letter(3+j)].width = 11
    ws.freeze_panes = "C5"


# =====================================================================
#  TARGETS  (closed-form goal-seek)
# =====================================================================
def build_targets(ws, A):
    ws.sheet_view.showGridLines = False
    setcell(ws, "B1", f"{LINE} - Targets / Goal-Seek ({FY})", f_title)
    setcell(ws, "B2", "Set a goal; the model solves the price or volume needed - exact closed form, no iteration.", f_sub)

    volr = f"'Assumptions'!C{A['vol']}:H{A['vol']}"
    cpr = f"'Assumptions'!C{A['cp']}:H{A['cp']}"
    dr = f"'Assumptions'!C{A['disc']}:H{A['disc']}"
    cosr = f"'Assumptions'!C{A['cos']}:H{A['cos']}"
    apr = f"'Assumptions'!C{A['ap']}:H{A['ap']}"
    # helper constants
    setcell(ws, "B4", "Model constants (from Assumptions)", f_h2, align=left)
    rows = [
        ("K1  Net @ base (Σ vol·cp·(1-d))", f"=SUMPRODUCT({volr},{cpr},(1-{dr}))", MON, "K1"),
        ("Kcos  Σ vol·Cost of Sales", f"=SUMPRODUCT({volr},{cosr})", MON, "KCOS"),
        ("Kap  Σ vol·A&P", f"=SUMPRODUCT({volr},{apr})", MON, "KAP"),
        ("Fixed opex", f"='Assumptions'!C{A['co_salaries']}+'Assumptions'!C{A['co_ga']}+'Assumptions'!C{A['co_mktg']}+'Assumptions'!C{A['co_slotting']}", MON, "FX"),
        ("Logistics % of net", f"='Assumptions'!C{A['co_log_pct']}", PCT, "LOG"),
        ("Base total cases", f"=SUM({volr})", CASES, "CASES"),
    ]
    cellref = {}
    for i, (label, f, fmt, key) in enumerate(rows):
        rr = 5 + i
        setcell(ws, f"B{rr}", "   " + label, f_lbl, align=left)
        setcell(ws, f"C{rr}", f, f_calc, fmt, fill_tot, rght, True)
        cellref[key] = f"$C${rr}"
    K1, KCOS, KAP, FX, LOG, CASES_C = (cellref[k] for k in ["K1", "KCOS", "KAP", "FX", "LOG", "CASES"])

    # editable targets
    t = 13
    setcell(ws, f"B{t}", "Your targets (edit yellow)", f_h2, align=left)
    setcell(ws, f"B{t+1}", "   Target company GP %", f_lbl, align=left)
    g_gp = setcell(ws, f"C{t+1}", 0.45, f_in, PCT, fill_in, rght, True)
    setcell(ws, f"B{t+2}", "   Target EBITDA ($)", f_lbl, align=left)
    g_eb = setcell(ws, f"C{t+2}", 150000, f_in, MON, fill_in, rght, True)
    GP_T = f"$C${t+1}"; EB_T = f"$C${t+2}"

    # answers
    a = t + 4
    setcell(ws, f"B{a}", "Solved answers", f_h2, align=left)
    ans = [
        ("Price × to hit target GP % (base volume)",
         f"=IFERROR({KCOS}/({K1}*(1-{GP_T})),\"n/a\")", MULT),
        ("→ implied blended CP change",
         f"=IFERROR({KCOS}/({K1}*(1-{GP_T}))-1,\"n/a\")", PCT),
        ("Volume × to hit target EBITDA (base price)",
         f"=IFERROR(({EB_T}+{FX})/(({K1}-{KCOS}-{KAP})-{LOG}*{K1}),\"n/a\")", MULT),
        ("→ implied total cases",
         f"=IFERROR(({EB_T}+{FX})/(({K1}-{KCOS}-{KAP})-{LOG}*{K1})*{CASES_C},\"n/a\")", CASES),
        ("Price × to hit target EBITDA (base vol; ε=0 approx)",
         f"=IFERROR(({EB_T}+{FX}+{KCOS}+{KAP})/({K1}*(1-{LOG})),\"n/a\")", MULT),
        ("Break-even volume × (EBITDA = 0, base price)",
         f"=IFERROR({FX}/(({K1}-{KCOS}-{KAP})-{LOG}*{K1}),\"n/a\")", MULT),
        ("→ break-even cases",
         f"=IFERROR({FX}/(({K1}-{KCOS}-{KAP})-{LOG}*{K1})*{CASES_C},\"n/a\")", CASES),
    ]
    for i, (label, f, fmt) in enumerate(ans):
        rr = a + 1 + i
        bold = not label.startswith("→")
        setcell(ws, f"B{rr}", ("   " if not bold else "   ") + label, f_lblb if bold else f_lbl, align=left)
        setcell(ws, f"C{rr}", f, f_calc, fmt, fill_tot, rght, True)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16


# =====================================================================
#  HOME  (Apple-style front screen — simple, beautiful, live)
# =====================================================================
ACCENT = "0A84FF"; CARD = "F4F7FB"; INKSOFT = "5B6472"


def build_home(ws, ws_dash, A):
    ws.sheet_view.showGridLines = False
    f_eyebrow = Font(name="Calibri", size=11, bold=True, color=ACCENT)
    f_hero = Font(name="Calibri", size=30, bold=True, color=INK)
    f_heros = Font(name="Calibri", size=12, color=INKSOFT)
    f_kval = Font(name="Calibri", size=18, bold=True, color=INK)
    f_klbl = Font(name="Calibri", size=9, bold=True, color=INKSOFT)
    f_nav = Font(name="Calibri", size=11, bold=True, color=INK)
    f_navd = Font(name="Calibri", size=10, color=INKSOFT)
    card = PatternFill("solid", fgColor=CARD)
    accent_fill = PatternFill("solid", fgColor=ACCENT)

    ws.column_dimensions["A"].width = 3
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 17
    ws.column_dimensions["B"].width = 19

    setcell(ws, "B2", "ORGANIKA", f_eyebrow, align=left)
    setcell(ws, "B3", "Sparkling Daily", f_hero, align=left)
    ws.merge_cells("B3:E3"); ws.row_dimensions[3].height = 40
    setcell(ws, "B4", f"One-Year Operating Model  ·  {FY}", f_heros, align=left)
    ws.merge_cells("B4:E4")

    # scenario selector
    setcell(ws, "B6", "Scenario", f_klbl, align=left)
    sel = setcell(ws, "C6", "Base", Font(name="Calibri", size=12, bold=True, color="0033AA"),
                  fill=fill_in, align=center, border=True)
    dv = DataValidation(type="list", formula1='"Low,Base,High,Stretch"', allow_blank=False)
    ws.add_data_validation(dv); dv.add(sel)
    setcell(ws, "D6", "▼ pick a scenario — everything below updates", f_navd, align=left)
    ws.merge_cells("D6:G6")

    # KPI cards (selected scenario, via INDIRECT)
    S = "INDIRECT(\"'\"&$C$6&\"'!" + TOT
    def scn(rk): return f"{S}{SR[rk]}\")"
    kpis = [
        ("NET SALES", scn("nsales"), MON), ("GROSS PROFIT %", scn("gppct"), PCT),
        ("EBITDA", scn("ebitda"), MON), ("EBITDA %", scn("ebitdapct"), PCT),
        ("NET INCOME", scn("ni"), MON), ("PHYSICAL CASES", scn("cases"), CASES),
    ]
    top = 8
    for j, (label, formula, fmt) in enumerate(kpis):
        col = get_column_letter(2 + j)
        lc = setcell(ws, f"{col}{top}", label, f_klbl, fill=card, align=center, border=False)
        vc = setcell(ws, f"{col}{top+1}", f"={formula}", f_kval, fmt, card, center, False)
    ws.row_dimensions[top].height = 18
    ws.row_dimensions[top + 1].height = 30

    # secondary line: break-even + blended CP
    setcell(ws, f"B{top+3}", "Break-even volume", f_klbl, align=left)
    setcell(ws, f"C{top+3}", "='Sensitivity'!$C$23", f_kval, CASES, None, left, False)
    setcell(ws, f"E{top+3}", "Blended CP / case", f_klbl, align=left)
    setcell(ws, f"F{top+3}", f"={scn('cp')}", f_kval, MON2, None, left, False)

    # navigation / table of contents
    nav_top = top + 5
    setcell(ws, f"B{nav_top}", "WHERE TO GO", f_eyebrow, align=left)
    nav = [
        ("1.  Assumptions", "The ONE place you type. Change a yellow cell, everything updates."),
        ("2.  Dashboard", "All four scenarios side by side."),
        ("Base · Low · High · Stretch", "The full profit story for each scenario."),
        ("Monthly", "The chosen scenario spread across the 12 months."),
        ("Pricing Lab", "See the wholesale price that makes the most money."),
        ("Targets", "Set a goal — it tells you the price or volume to get there."),
        ("Sensitivity", "How profit moves with price and volume."),
        ("Checks", "Built-in health check (should say ALL PASS)."),
    ]
    r = nav_top + 1
    for name, desc in nav:
        setcell(ws, f"B{r}", name, f_nav, align=left)
        setcell(ws, f"C{r}", desc, f_navd, align=left)
        ws.merge_cells(f"C{r}:D{r}")
        r += 1
    setcell(ws, f"B{r+1}", f"{VERSION}  ·  built {BUILT}  ·  all $ CDN  ·  figures are placeholders until you add real numbers", f_navd, align=left)
    ws.merge_cells(f"B{r+1}:G{r+1}")


# =====================================================================
#  ASSEMBLE
# =====================================================================
ws_home = wb.active; ws_home.title = "Home"
ws_asmp = wb.create_sheet("Assumptions")
ws_dash = wb.create_sheet("Dashboard")
scen_ws = {s: wb.create_sheet(s) for s in SCEN}
ws_month = wb.create_sheet("Monthly")
ws_price = wb.create_sheet("Pricing Lab")
ws_target = wb.create_sheet("Targets")
ws_sens = wb.create_sheet("Sensitivity")
ws_checks = wb.create_sheet("Checks")
ws_guide = wb.create_sheet("Guide")
ws_review = wb.create_sheet("Review Log")
ws_cover = wb.create_sheet("Cover")

A = build_assumptions(ws_asmp)
for s in SCEN:
    build_scenario(scen_ws[s], s, A)
build_monthly(ws_month, A)
build_sensitivity(ws_sens, A)          # before dashboard/cover/home (they reference C23)
build_pricing(ws_price, A)
build_targets(ws_target, A)
build_dashboard(ws_dash)
build_checks(ws_checks, A)
build_cover(ws_cover)
build_guide(ws_guide)
build_review_log(ws_review)
build_home(ws_home, ws_dash, A)

# ---- simple-first tab order (Apple-style) ----
order = ["Home", "Assumptions", "Dashboard", "Base", "Low", "High", "Stretch",
         "Monthly", "Pricing Lab", "Targets", "Sensitivity", "Checks",
         "Guide", "Review Log", "Cover"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

# ---- named ranges for key inputs (usability) ----
def dn(name, sheet, cell):
    ref = f"'{sheet}'!{cell}"
    wb.defined_names.add(DefinedName(name, attr_text=ref))
dn("CP_Pineapple_Singles", "Assumptions", f"$C${A['cp']}")
dn("BaseVolume_Total", "Assumptions", f"$C${A['vol']}:$H${A['vol']}")
dn("TaxRate", "Assumptions", f"$C${A['co_tax']}")
dn("LogisticsPct", "Assumptions", f"$C${A['co_log_pct']}")

# ---- tab colours + print setup across the workbook ----
for ws in wb.worksheets:
    if ws.title in TAB:
        ws.sheet_properties.tabColor = TAB[ws.title]
    landscape = ws.title not in ("Cover", "Guide", "Review Log", "Targets")
    polish(ws, landscape=landscape)
# repeat header rows when printing the long tabs
for s in SCEN:
    scen_ws[s].print_title_rows = "5:6"
ws_asmp.print_title_rows = "4:5"
ws_dash.print_title_rows = "4:4"
ws_month.print_title_rows = "6:6"

for ws in wb.worksheets:
    ws.sheet_view.tabSelected = False
ws_home.sheet_view.tabSelected = True
wb.active = wb._sheets.index(ws_home)

wb.save(OUT)
print("Saved", OUT)
print("Tabs:", wb.sheetnames)
