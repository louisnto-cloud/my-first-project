# -*- coding: utf-8 -*-
"""
Enterprise CPG Commercial Decision System  -  Excel workbook generator.

Builds: CPG_Commercial_Decision_System.xlsx

Design philosophy
-----------------
* One source of truth for assumptions (Inputs_and_Control_Center).
* A single scenario selector (selScenario) drives the ENTIRE workbook.
* Deduction / cost elements are entered on one of three bases
  (per-unit $, % of gross, fixed $) and rolled into named "buckets".
  Both the detailed Calc engine and the compact Scenario table consume the
  same buckets, so Base scenario == Calc engine == Dashboard, exactly.
* Formulas use NAMED RANGES so they read like English and survive row edits.
* Inputs are unlocked / yellow; everything else is locked grey/white.

Author: Commercial Finance Systems
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- palette ----
NAVY      = "1F3864"   # headers
BLUE      = "2E5496"
STEEL     = "4472C4"
LTBLUE    = "D9E1F2"   # section band
INPUTYEL  = "FFF2CC"   # input cell
INPUTBRD  = "BF9000"
GREY      = "F2F2F2"   # calc band
GREEN     = "C6EFCE"; GREENT = "006100"
YELLOW    = "FFEB9C"; YELLOWT = "9C6500"
RED       = "FFC7CE"; REDT   = "9C0006"
WHITE     = "FFFFFF"
DKTXT     = "203864"

thin  = Side(style="thin",   color="BFBFBF")
med   = Side(style="medium", color="808080")
box   = Border(left=thin, right=thin, top=thin, bottom=thin)
botln = Border(bottom=Side(style="medium", color=NAVY))

def font(sz=10, b=False, color="000000", italic=False, name="Calibri"):
    return Font(name=name, size=sz, bold=b, color=color, italic=italic)

def fill(hexc):
    return PatternFill("solid", fgColor=hexc)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
RIGHT  = Alignment(horizontal="right",  vertical="center")
LEFTW  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

FMT_USD0 = '#,##0;[Red](#,##0)'
FMT_USD2 = '#,##0.00;[Red](#,##0.00)'
FMT_PCT1 = '0.0%;[Red](0.0%)'
FMT_PCT0 = '0%'
FMT_NUM0 = '#,##0'
FMT_X2   = '0.00"x"'

wb = Workbook()
wb.remove(wb.active)

# names: key -> "'Sheet'!$C$row"  (used for defined names + readable formulas)
NAME = {}

def ref(sheet, col, row):
    return "'%s'!$%s$%d" % (sheet.title, col, row)

def define(name, sheet, col, row):
    r = ref(sheet, col, row)
    NAME[name] = r
    wb.defined_names[name] = DefinedName(name, attr_text=r)
    return r

def define_range(name, sheet, c1, r1, c2, r2):
    r = "'%s'!$%s$%d:$%s$%d" % (sheet.title, c1, r1, c2, r2)
    NAME[name] = r
    wb.defined_names[name] = DefinedName(name, attr_text=r)
    return r

# ---------------------------------------------------------- cell helpers ----
def W(ws, coord, val=None, *, f=None, fillc=None, fmt=None, align=None,
      border=None, locked=True, wrap=False):
    c = ws[coord]
    if val is not None:
        c.value = val
    if f:      c.font = f
    if fillc:  c.fill = fill(fillc)
    if fmt:    c.number_format = fmt
    if align:  c.alignment = align
    elif wrap: c.alignment = LEFTW
    if border: c.border = border
    c.protection = Protection(locked=locked)
    return c

def title_block(ws, title, subtitle):
    ws.merge_cells("B2:N2")
    W(ws, "B2", title, f=font(16, True, WHITE), fillc=NAVY, align=LEFT)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(2, 15):
        ws.cell(row=2, column=col).fill = fill(NAVY)
    ws.row_dimensions[2].height = 26
    ws.merge_cells("B3:N3")
    W(ws, "B3", subtitle, f=font(9, False, DKTXT, italic=True), fillc=LTBLUE, align=LEFT)
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(2, 15):
        ws.cell(row=3, column=col).fill = fill(LTBLUE)
    ws.row_dimensions[3].height = 16

def section(ws, row, label, span=12, col0=2):
    ws.merge_cells(start_row=row, start_column=col0, end_row=row, end_column=col0+span-1)
    c = ws.cell(row=row, column=col0, value=label)
    c.font = font(11, True, WHITE)
    c.fill = fill(BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for cc in range(col0, col0+span):
        ws.cell(row=row, column=cc).fill = fill(BLUE)
    ws.row_dimensions[row].height = 20

def colhdr(ws, row, pairs):
    """pairs: list of (coord, text)"""
    for coord, text in pairs:
        W(ws, coord, text, f=font(9, True, WHITE), fillc=STEEL, align=CENTER, border=box)

# scenario names used everywhere
SCN = ["Base Case", "Promo Case", "High Volume Case", "Low Volume Case",
       "Cost Inflation Case", "Margin Recovery Case", "Custom Scenario"]
BIZMODELS = ["Direct Retailer", "Distributor", "Club", "FDM", "Hybrid"]
CCY = ["USD", "CAD", "JPY", "KRW"]

# =====================================================================
#  TAB 9 (built early so lists exist): LOOKUPS_AND_ASSUMPTIONS
# =====================================================================
look = wb.create_sheet("Lookups_and_Assumptions")
look.sheet_properties.tabColor = "808080"
title_block(look, "LOOKUPS & ASSUMPTIONS", "Pick-lists, FX seed values and reference tables. Edit lists here; the rest of the workbook references them.")

section(look, 5, "Scenario list (selScenario)")
for i, s in enumerate(SCN):
    W(look, "B%d" % (6+i), s, f=font(10), fillc=WHITE, border=box, align=LEFT)
define_range("scnList", look, "B", 6, "B", 6+len(SCN)-1)

section(look, 14, "Business model list (selBusinessModel)")
for i, s in enumerate(BIZMODELS):
    W(look, "B%d" % (15+i), s, f=font(10), fillc=WHITE, border=box, align=LEFT)
define_range("bizList", look, "B", 15, "B", 15+len(BIZMODELS)-1)

section(look, 21, "Currency list (selBaseCurrency / selReportingCurrency)")
for i, s in enumerate(CCY):
    W(look, "B%d" % (22+i), s, f=font(10), fillc=WHITE, border=box, align=LEFT)
define_range("ccyList", look, "B", 22, "B", 22+len(CCY)-1)

section(look, 28, "Decision status list")
for i, s in enumerate(["GO", "REVIEW", "STOP"]):
    W(look, "B%d" % (29+i), s, f=font(10), fillc=WHITE, border=box, align=LEFT)

look.column_dimensions["A"].width = 2
look.column_dimensions["B"].width = 28
look.sheet_view.showGridLines = False

# =====================================================================
#  TAB 2: INPUTS_AND_CONTROL_CENTER   (built early; holds most named cells)
# =====================================================================
inp = wb.create_sheet("Inputs_and_Control_Center")
inp.sheet_properties.tabColor = INPUTBRD
title_block(inp, "INPUTS & CONTROL CENTER",
            "Single source of truth. ONLY yellow cells are user inputs. Change an assumption and every tab recalculates instantly.")

# ---- control bar ----
section(inp, 5, "CONTROL BAR  —  set the active scenario & targets here", span=8)
ctrl = [
    ("B6", "Model Name",            "C6", "CPG Commercial Decision System",  "text"),
    ("B7", "Active Scenario",       "C7", "Base Case",                       "scn"),
    ("B8", "Business Model",        "C8", "Distributor",                     "biz"),
    ("B9", "Base Currency",         "C9", "USD",                             "ccy"),
    ("B10","Reporting Currency",    "C10","CAD",                             "ccy"),
]
for lab_c, lab, val_c, val, kind in ctrl:
    W(inp, lab_c, lab, f=font(10, True, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
    W(inp, val_c, val, f=font(10, True), fillc=INPUTYEL, border=box, align=LEFT, locked=False)
define("selScenario",         inp, "C", 7)
define("selBusinessModel",    inp, "C", 8)
define("selBaseCurrency",     inp, "C", 9)
define("selReportingCurrency",inp, "C", 10)

targets = [
    ("E6","Target Gross Margin %",  "F6", 0.42, FMT_PCT1, "TargetGM"),
    ("E7","Target Net Margin %",    "F7", 0.12, FMT_PCT1, "TargetNM"),
    ("E8","Target Promo ROI (x)",   "F8", 1.00, FMT_X2,  "TargetROI"),
    ("E9","Max Trade Spend % Gross","F9", 0.22, FMT_PCT1, "MaxTradePct"),
    ("E10","Min Contribution / Unit","F10",0.15, FMT_USD2, "MinContribUnit"),
]
for lab_c, lab, val_c, val, fmt, nm in targets:
    W(inp, lab_c, lab, f=font(10, True, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
    W(inp, val_c, val, f=font(10, True), fillc=INPUTYEL, border=box, fmt=fmt, align=CENTER, locked=False)
    define(nm, inp, "F", int(val_c[1:]))

# decision + integrity status (formulas filled after calc refs exist)
W(inp, "H6", "DECISION STATUS", f=font(10, True, WHITE), fillc=NAVY, border=box, align=CENTER)
W(inp, "H7", "MODEL INTEGRITY", f=font(10, True, WHITE), fillc=NAVY, border=box, align=CENTER)
# placeholders, set later
inp.merge_cells("I6:J6"); inp.merge_cells("I7:J7")

r = 13
def sec(label, span=8):
    global r
    section(inp, r, label, span=span); r += 1

def inrow(label, val, fmt=FMT_USD2, name=None, note="", unit=""):
    """label B, value C (input,yellow), note/unit D."""
    global r
    W(inp, "B%d" % r, label, f=font(10, DKTXT and "000000"), fillc=WHITE, border=box, align=LEFT)
    W(inp, "C%d" % r, val, f=font(10, True), fillc=INPUTYEL, border=box, fmt=fmt, align=RIGHT, locked=False)
    W(inp, "D%d" % r, unit, f=font(9, False, "808080"), fillc=WHITE, border=box, align=LEFT)
    if note:
        W(inp, "E%d" % r, note, f=font(9, False, "808080"), fillc=WHITE, align=LEFT)
    if name:
        define(name, inp, "C", r)
    r += 1

# ---- A. Business context ----
sec("A.  BUSINESS CONTEXT")
inrow("SKU / Product Name", "Zephyr Sparkling 355ml 12pk", '@', "SKU", unit="text")
inrow("Brand", "Zephyr", '@', unit="text")
inrow("Customer / Retailer / Distributor", "National Grocery via UDF Distributor", '@', unit="text")
inrow("Channel", "Grocery", '@', unit="text")
inrow("Market / Region", "Canada", '@', unit="text")
inrow("Launch vs Existing flag", "Existing", '@', unit="text")
inrow("Units per Case", 12, FMT_NUM0, "UnitsPerCase", unit="units/case")
inrow("Servings per Unit", 1, FMT_NUM0, unit="svgs/unit")
inrow("Number of Stores", 1200, FMT_NUM0, "NumStores", unit="stores")

# ---- B. Pricing ----
sec("B.  PRICING INPUTS")
inrow("List Price / unit",            1.80, FMT_USD2, "ListPrice", unit="$/unit")
inrow("Invoice Price / unit",         1.65, FMT_USD2, "InvoicePrice", unit="$/unit")
inrow("Wholesale Price / unit (company sells at)", 1.50, FMT_USD2, "WholesalePrice", unit="$/unit",
      note="Net invoice price the company realizes before off-invoice deductions")
inrow("Suggested Retail Price / unit", 2.49, FMT_USD2, "SRP", unit="$/unit")
inrow("Promo Retail Price / unit",     1.99, FMT_USD2, "PromoRetail", unit="$/unit")
inrow("Promo Depth %",                 0.20, FMT_PCT1, "PromoDepth", unit="% off SRP")
inrow("Retailer Target Margin %",      0.30, FMT_PCT1, "RetailerMargin", unit="% of retail")
inrow("Distributor Target Margin %",   0.12, FMT_PCT1, "DistributorMargin", unit="% of dist sell")

# ---- C. Cost stack (per unit) ----
sec("C.  COST STACK  (per unit, except broker/distributor %)")
colhdr(inp, r, [("B%d"%r,"Cost element"),("C%d"%r,"$ / unit"),("D%d"%r,"Note")]); r += 1
cost_start = r
cost_items = [
    ("COGS / Product Cost", 0.55),
    ("Packaging", 0.09),
    ("Co-Manufacturing", 0.04),
    ("Inbound Freight", 0.03),
    ("Outbound Freight", 0.05),
    ("Duty / Tariffs / Import", 0.01),
    ("Warehousing / Pick-Pack", 0.02),
    ("Compliance / Quality / Reg", 0.01),
    ("Returns & Spoilage allowance", 0.02),
    ("Other Variable Cost", 0.01),
]
for lab, v in cost_items:
    W(inp, "B%d"%r, lab, f=font(10), fillc=WHITE, border=box, align=LEFT)
    W(inp, "C%d"%r, v, f=font(10, True), fillc=INPUTYEL, border=box, fmt=FMT_USD2, align=RIGHT, locked=False)
    W(inp, "D%d"%r, "$/unit", f=font(9,False,"808080"), fillc=WHITE, border=box, align=LEFT)
    r += 1
cost_end = r-1
# subtotal vc_pu
W(inp, "B%d"%r, "Σ Variable Cost per Unit (sum)", f=font(10, True, DKTXT), fillc=GREY, border=box, align=LEFT)
W(inp, "C%d"%r, "=SUM(C%d:C%d)"%(cost_start,cost_end), f=font(10, True), fillc=GREY, border=box, fmt=FMT_USD2, align=RIGHT)
define("vc_pu", inp, "C", r); r += 1
# broker / distributor % of gross
W(inp, "B%d"%r, "Broker Fee % of Gross", f=font(10), fillc=WHITE, border=box, align=LEFT)
W(inp, "C%d"%r, 0.03, f=font(10, True), fillc=INPUTYEL, border=box, fmt=FMT_PCT1, align=RIGHT, locked=False)
define("BrokerPct", inp, "C", r); r += 1
W(inp, "B%d"%r, "Distributor Fee % of Gross", f=font(10), fillc=WHITE, border=box, align=LEFT)
W(inp, "C%d"%r, 0.00, f=font(10, True), fillc=INPUTYEL, border=box, fmt=FMT_PCT1, align=RIGHT, locked=False)
W(inp, "E%d"%r, "Use only if company PAYS the distributor a fee on top of wholesale; default 0 = distributor margin sits in the price chain (avoids double-count).",
  f=font(8, False, "808080"), fillc=WHITE, align=LEFT)
define("DistributorPct", inp, "C", r); r += 1
W(inp, "B%d"%r, "Σ Variable Cost % of Gross (broker+distributor)", f=font(10, True, DKTXT), fillc=GREY, border=box, align=LEFT)
W(inp, "C%d"%r, "=BrokerPct+DistributorPct", f=font(10, True), fillc=GREY, border=box, fmt=FMT_PCT1, align=RIGHT)
define("vc_pct", inp, "C", r); r += 1

# ---- D. Trade spend & deductions (3-basis entry) ----
sec("D.  TRADE INVESTMENT  (enter on ONE basis per line)")
hdr = r
colhdr(inp, hdr, [("B%d"%hdr,"Trade element"),("C%d"%hdr,"$ / unit"),
                  ("D%d"%hdr,"% of Gross"),("E%d"%hdr,"Fixed $"),("F%d"%hdr,"Basis flag")])
r += 1
trade_start = r
trade_items = [
    ("TPD (Temporary Price Discount)", None, 0.04, None),
    ("End Cap", None, None, 10000),
    ("Fence / Feature Promotion", None, 0.01, None),
    ("Display Allowance", None, None, 8000),
    ("Advertising Support", None, None, 6000),
    ("Flyer Support", None, None, 4000),
    ("Digital Media Support", None, 0.004, None),
    ("Demo Spend", None, None, 2500),
    ("Listing / Slotting Fees (amortized)", None, None, 5000),
    ("Scan Backs", 0.02, None, None),
    ("Bill Backs", 0.015, None, None),
    ("Off-Invoice Discounts", None, 0.01, None),
    ("Markdown Support", None, 0.005, None),
    ("Retailer Funding", None, None, 3000),
    ("Lump-Sum Trade Spend", None, None, 6000),
    ("Co-Op Marketing", None, 0.003, None),
    ("Free Fill / Free Goods", 0.005, None, None),
]
for lab, pu, pc, fx in trade_items:
    W(inp, "B%d"%r, lab, f=font(10), fillc=WHITE, border=box, align=LEFT)
    W(inp, "C%d"%r, pu if pu is not None else 0, f=font(10), fillc=INPUTYEL, border=box, fmt=FMT_USD2, align=RIGHT, locked=False)
    W(inp, "D%d"%r, pc if pc is not None else 0, f=font(10), fillc=INPUTYEL, border=box, fmt=FMT_PCT1, align=RIGHT, locked=False)
    W(inp, "E%d"%r, fx if fx is not None else 0, f=font(10), fillc=INPUTYEL, border=box, fmt=FMT_USD0, align=RIGHT, locked=False)
    # basis flag (text) for auditability
    basis = "per-unit" if pu is not None else ("%gross" if pc is not None else "fixed $")
    W(inp, "F%d"%r, basis, f=font(9,False,"808080"), fillc=WHITE, border=box, align=CENTER)
    r += 1
trade_end = r-1
W(inp, "B%d"%r, "Σ Trade buckets (per-unit / %gross / fixed)", f=font(10, True, DKTXT), fillc=GREY, border=box, align=LEFT)
W(inp, "C%d"%r, "=SUM(C%d:C%d)"%(trade_start,trade_end), f=font(10, True), fillc=GREY, border=box, fmt=FMT_USD2, align=RIGHT)
W(inp, "D%d"%r, "=SUM(D%d:D%d)"%(trade_start,trade_end), f=font(10, True), fillc=GREY, border=box, fmt=FMT_PCT1, align=RIGHT)
W(inp, "E%d"%r, "=SUM(E%d:E%d)"%(trade_start,trade_end), f=font(10, True), fillc=GREY, border=box, fmt=FMT_USD0, align=RIGHT)
define("ded_trade_pu",  inp, "C", r)
define("ded_trade_pct", inp, "D", r)
define("ded_trade_fix", inp, "E", r)
r += 1

# ---- D2. Other gross-to-net deductions (NOT counted as trade investment) ----
sec("D2.  OTHER GROSS-TO-NET DEDUCTIONS  (financial / non-trade)")
hdr = r
colhdr(inp, hdr, [("B%d"%hdr,"Deduction element"),("C%d"%hdr,"$ / unit"),
                  ("D%d"%hdr,"% of Gross"),("E%d"%hdr,"Fixed $"),("F%d"%hdr,"Basis flag")])
r += 1
other_start = r
other_items = [
    ("Trade Accrual", None, 0.010, None),
    ("Early Payment Discount", None, 0.010, None),
    ("Customer-Specific Deductions", None, 0.004, None),
    ("Chargebacks", None, 0.002, None),
    ("Shortage / Damage / Claims", None, 0.002, None),
]
for lab, pu, pc, fx in other_items:
    W(inp, "B%d"%r, lab, f=font(10), fillc=WHITE, border=box, align=LEFT)
    W(inp, "C%d"%r, pu if pu is not None else 0, f=font(10), fillc=INPUTYEL, border=box, fmt=FMT_USD2, align=RIGHT, locked=False)
    W(inp, "D%d"%r, pc if pc is not None else 0, f=font(10), fillc=INPUTYEL, border=box, fmt=FMT_PCT1, align=RIGHT, locked=False)
    W(inp, "E%d"%r, fx if fx is not None else 0, f=font(10), fillc=INPUTYEL, border=box, fmt=FMT_USD0, align=RIGHT, locked=False)
    basis = "per-unit" if pu is not None else ("%gross" if pc is not None else "fixed $")
    W(inp, "F%d"%r, basis, f=font(9,False,"808080"), fillc=WHITE, border=box, align=CENTER)
    r += 1
other_end = r-1
W(inp, "B%d"%r, "Σ Other deduction buckets", f=font(10, True, DKTXT), fillc=GREY, border=box, align=LEFT)
W(inp, "C%d"%r, "=SUM(C%d:C%d)"%(other_start,other_end), f=font(10, True), fillc=GREY, border=box, fmt=FMT_USD2, align=RIGHT)
W(inp, "D%d"%r, "=SUM(D%d:D%d)"%(other_start,other_end), f=font(10, True), fillc=GREY, border=box, fmt=FMT_PCT1, align=RIGHT)
W(inp, "E%d"%r, "=SUM(E%d:E%d)"%(other_start,other_end), f=font(10, True), fillc=GREY, border=box, fmt=FMT_USD0, align=RIGHT)
define("ded_other_pu",  inp, "C", r)
define("ded_other_pct", inp, "D", r)
define("ded_other_fix", inp, "E", r)
r += 1

# ---- E. Volume & distribution ----
sec("E.  VOLUME & DISTRIBUTION ASSUMPTIONS")
inrow("Forecast Unit Sales (annual)", 500000, FMT_NUM0, "ForecastUnits", unit="units")
inrow("Base (non-promo) Volume", 380000, FMT_NUM0, "BaseUnits", unit="units")
inrow("Expected Promo Lift %", 0.35, FMT_PCT1, "PromoLiftPct", unit="% of base")
inrow("Cannibalization %", 0.10, FMT_PCT1, "CannibPct", unit="% of incr.")
inrow("Forward Buy %", 0.08, FMT_PCT1, "ForwardBuyPct", unit="% of incr.")
inrow("Pantry Loading %", 0.05, FMT_PCT1, "PantryPct", unit="% of incr.")
inrow("Post-Promo Dip %", 0.07, FMT_PCT1, "PostDipPct", unit="% of incr.")
inrow("Velocity per Store per Week", 6.5, FMT_USD2, "VelocityPSPW", unit="units/store/wk")
inrow("Number of Promo Weeks", 4, FMT_NUM0, "PromoWeeks", unit="weeks")
inrow("Fill Rate %", 0.97, FMT_PCT1, "FillRate", unit="service level")
inrow("ACV % (distribution)", 0.85, FMT_PCT1, "ACV", unit="% ACV")

# ---- F. Fixed & program costs ----
sec("F.  FIXED & PROGRAM COSTS  (annual $)")
fx_start = r
fixed_items = [
    ("Fixed Monthly Support (annualized)", 24000),
    ("Customer-Specific Support Spend", 15000),
    ("Broker Support Cost", 9000),
    ("Account Management Cost", 12000),
    ("Merchandising Cost", 8000),
    ("Annual Listing Fee Amortization", 6000),
    ("Promotional Development Cost", 5000),
    ("Other Fixed Commercial Cost", 3000),
]
for lab, v in fixed_items:
    W(inp, "B%d"%r, lab, f=font(10), fillc=WHITE, border=box, align=LEFT)
    W(inp, "C%d"%r, v, f=font(10, True), fillc=INPUTYEL, border=box, fmt=FMT_USD0, align=RIGHT, locked=False)
    W(inp, "D%d"%r, "$/yr", f=font(9,False,"808080"), fillc=WHITE, border=box, align=LEFT)
    r += 1
fx_end = r-1
W(inp, "B%d"%r, "Σ Total Fixed Commercial Cost", f=font(10, True, DKTXT), fillc=GREY, border=box, align=LEFT)
W(inp, "C%d"%r, "=SUM(C%d:C%d)"%(fx_start,fx_end), f=font(10, True), fillc=GREY, border=box, fmt=FMT_USD0, align=RIGHT)
define("FixedCommercialTotal", inp, "C", r); r += 1

# ---- G. Governance ----
sec("G.  GOVERNANCE & DECISION THRESHOLDS")
inrow("Minimum Gross Margin %", 0.38, FMT_PCT1, "MinGM", unit="hurdle")
inrow("Minimum Net Margin %", 0.08, FMT_PCT1, "MinNM", unit="hurdle")
inrow("Price-Increase Approval Threshold %", 0.05, FMT_PCT1, "PriceApprovalThr", unit="needs sign-off above")
W(inp, "B%d"%r, "Management Commentary", f=font(10), fillc=WHITE, border=box, align=LEFT)
W(inp, "C%d"%r, "Distributor path; watch trade % vs 22% ceiling.", f=font(10, italic=True), fillc=INPUTYEL, border=box, align=LEFT, locked=False)
inp.merge_cells("C%d:F%d"%(r,r)); r += 2

inp.column_dimensions["A"].width = 2
inp.column_dimensions["B"].width = 38
for c in "CDE":
    inp.column_dimensions[c].width = 13
inp.column_dimensions["F"].width = 12
for c in "GHIJ":
    inp.column_dimensions[c].width = 14
inp.freeze_panes = "B6"
inp.sheet_view.showGridLines = False

KPI_SNAP_ROW = r  # remember where to place KPI snapshot after Calc built

# =====================================================================
#  TAB 4: SCENARIOS  (driver table + compact P&L per scenario)
# =====================================================================
scn = wb.create_sheet("Scenarios")
scn.sheet_properties.tabColor = STEEL
title_block(scn, "SCENARIO ENGINE",
            "Each scenario is a set of DRIVER MULTIPLIERS on the base inputs. The active scenario (selScenario) is read by the Calc engine via INDEX/MATCH.")

# scenario header row
SCN_HDR = 6
W(scn, "B%d"%SCN_HDR, "Driver  (multiplier on base; 1.00 = base)", f=font(10, True, WHITE), fillc=NAVY, border=box, align=LEFT)
cols = ["C","D","E","F","G","H","I"]
for i, s in enumerate(SCN):
    W(scn, "%s%d"%(cols[i],SCN_HDR), s, f=font(9, True, WHITE), fillc=NAVY, border=box, align=CENTER)
define_range("scnHeaders", scn, "C", SCN_HDR, "I", SCN_HDR)

# driver rows
drivers = [
    ("Volume multiplier",        [1.00,1.30,1.20,0.80,1.00,0.95,1.00], "volMult"),
    ("Wholesale price mult",     [1.00,0.95,1.00,1.00,1.00,1.06,1.00], "priceMult"),
    ("Variable cost mult",       [1.00,1.00,0.98,1.02,1.12,1.00,1.00], "vcMult"),
    ("Trade spend mult",         [1.00,1.35,1.05,1.00,1.00,0.85,1.00], "tradeMult"),
    ("Fill-rate mult",           [1.00,1.00,1.00,0.97,1.00,1.00,1.00], "fillMult"),
    ("Cannibalization mult",     [1.00,1.30,1.10,0.90,1.00,1.00,1.00], "cannibMult"),
    ("Promo active flag (1/0)",  [0,1,0,0,0,0,0],                       "promoFlag"),
    ("FX stress mult",           [1.00,1.00,1.00,1.00,1.00,1.00,1.00], "fxMult"),
]
DRV0 = SCN_HDR+1
for di,(lab, vals, nm) in enumerate(drivers):
    rr = DRV0+di
    W(scn, "B%d"%rr, lab, f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
    for i,v in enumerate(vals):
        is_flag = (nm=="promoFlag")
        W(scn, "%s%d"%(cols[i],rr), v, f=font(10, True), fillc=INPUTYEL, border=box,
          fmt=(FMT_NUM0 if is_flag else '0.00'), align=CENTER, locked=False)
    define_range(nm, scn, "C", rr, "I", rr)   # whole driver row
DRVend = DRV0+len(drivers)-1

# helper to get the column letter for a given scenario index inside the compact P&L
def scell(metric_row, col):  # not used heavily
    return "%s%d"%(col, metric_row)

# ----- compact P&L per scenario column -----
section_row = DRVend+2
section(scn, section_row, "SCENARIO RESULTS  (full P&L computed per scenario, identical bucket math to the Calc engine)", span=8)
PNL0 = section_row+1
# header
W(scn, "B%d"%PNL0, "Metric", f=font(9, True, WHITE), fillc=STEEL, border=box, align=LEFT)
for i,s in enumerate(SCN):
    W(scn, "%s%d"%(cols[i],PNL0), s, f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)
W(scn, "K%d"%PNL0, "Hurdle test", f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)

# rows: build formulas per column referencing driver cells in same column.
# driver cell rows:
dvol, dprice, dvc, dtrade, dfill, dcannib, dpromo, dfx = [DRV0+i for i in range(8)]

metric_defs = []  # (label, fmt, func(col)->formula)
def F(col):  # dict of per-column cell addresses for drivers
    return dict(vol="%s%d"%(col,dvol), price="%s%d"%(col,dprice), vc="%s%d"%(col,dvc),
                trade="%s%d"%(col,dtrade), fill="%s%d"%(col,dfill), cannib="%s%d"%(col,dcannib),
                promo="%s%d"%(col,dpromo), fx="%s%d"%(col,dfx))

# We will assign output row numbers as we append:
row_index = {}
def add_metric(key, label, fmt, formula_fn, band=False):
    metric_defs.append((key, label, fmt, formula_fn, band))

# define formulas (col is letter)
add_metric("units","Effective Units", FMT_NUM0,
    lambda c,R: "=ForecastUnits*%s*MIN(1,FillRate*%s)"%(F(c)["vol"],F(c)["fill"]))
add_metric("gross","Gross Sales", FMT_USD0,
    lambda c,R: "=%s*WholesalePrice*%s"%(R["units"]%c, F(c)["price"]))
add_metric("trade","Trade Spend $", FMT_USD0,
    lambda c,R: "=(ded_trade_pct*%s+ded_trade_pu*%s+ded_trade_fix)*%s"%(R["gross"]%c,R["units"]%c,F(c)["trade"]))
add_metric("otherded","Other GtN Deductions $", FMT_USD0,
    lambda c,R: "=(ded_other_pct*%s+ded_other_pu*%s+ded_other_fix)*%s"%(R["gross"]%c,R["units"]%c,F(c)["trade"]))
add_metric("gtn","Total Gross-to-Net Deductions", FMT_USD0,
    lambda c,R: "=%s+%s"%(R["trade"]%c,R["otherded"]%c))
add_metric("net","Net Sales", FMT_USD0,
    lambda c,R: "=%s-%s"%(R["gross"]%c,R["gtn"]%c))
add_metric("vcost","Total Variable Cost", FMT_USD0,
    lambda c,R: "=(vc_pu*%s)*%s+vc_pct*%s"%(F(c)["vc"],R["units"]%c,R["gross"]%c))
add_metric("gp","Gross Profit $", FMT_USD0,
    lambda c,R: "=%s-%s"%(R["gross"]%c,R["vcost"]%c))
add_metric("gmpct","Gross Margin %", FMT_PCT1,
    lambda c,R: "=IF(%s=0,0,%s/%s)"%(R["gross"]%c,R["gp"]%c,R["gross"]%c))
add_metric("contrib","Contribution $", FMT_USD0,
    lambda c,R: "=%s-%s"%(R["net"]%c,R["vcost"]%c))
add_metric("cmpct","Contribution Margin %", FMT_PCT1,
    lambda c,R: "=IF(%s=0,0,%s/%s)"%(R["net"]%c,R["contrib"]%c,R["net"]%c))
add_metric("np","Net Profit $", FMT_USD0,
    lambda c,R: "=%s-FixedCommercialTotal"%(R["contrib"]%c))
add_metric("nmpct","Net Margin %", FMT_PCT1,
    lambda c,R: "=IF(%s=0,0,%s/%s)"%(R["net"]%c,R["np"]%c,R["net"]%c))
add_metric("tradepct","Trade Spend % Gross", FMT_PCT1,
    lambda c,R: "=IF(%s=0,0,%s/%s)"%(R["gross"]%c,R["trade"]%c,R["gross"]%c))
add_metric("ppu","Profit / Unit", FMT_USD2,
    lambda c,R: "=IF(%s=0,0,%s/%s)"%(R["units"]%c,R["np"]%c,R["units"]%c))
add_metric("cpu","Contribution / Unit", FMT_USD2,
    lambda c,R: "=IF(%s=0,0,%s/%s)"%(R["units"]%c,R["contrib"]%c,R["units"]%c))
add_metric("beunits","Break-Even Units", FMT_NUM0,
    lambda c,R: '=IF(%s<=0,"n/a",FixedCommercialTotal/%s)'%(R["cpu"]%c,R["cpu"]%c))
add_metric("promospend","Promo Spend (vs Base) $", FMT_USD0,
    lambda c,R: "=%s-$C$%d"%(R["trade"]%c, row_index["trade"]))   # base col C
add_metric("incrnp","Incremental Net Profit (vs Base) $", FMT_USD0,
    lambda c,R: "=%s-$C$%d"%(R["np"]%c, row_index["np"]))
add_metric("roi","Promo ROI (x)", FMT_X2,
    lambda c,R: "=IF(%s<=0,0,%s/%s)"%(R["promospend"]%c,R["incrnp"]%c,R["promospend"]%c))

# assign rows
cur = PNL0+1
for key, label, fmt, fn, band in metric_defs:
    row_index[key] = cur
    cur += 1
# Now Rmap: key -> "%s"+row  -> we need format strings like "C12" => "%s12"
Rmap = {k: "%s"+str(v) for k,v in row_index.items()}

# write metric rows
for key, label, fmt, fn, band in metric_defs:
    rr = row_index[key]
    W(scn, "B%d"%rr, label, f=font(10, True if key in ("net","np","roi") else False, DKTXT),
      fillc=(GREY if key in ("gtn","net","gp","contrib","np") else WHITE), border=box, align=LEFT)
    for i,c in enumerate(cols):
        W(scn, "%s%d"%(c,rr), fn(c, Rmap), f=font(10), fillc=WHITE, border=box, fmt=fmt, align=RIGHT)

# Decision row
DEC = cur
W(scn, "B%d"%DEC, "DECISION STATUS", f=font(10, True, WHITE), fillc=NAVY, border=box, align=LEFT)
for c in cols:
    gm = Rmap["gmpct"]%c; nm = Rmap["nmpct"]%c; cpu=Rmap["cpu"]%c; tp=Rmap["tradepct"]%c
    roi=Rmap["roi"]%c; promo="%s%d"%(c,dpromo)
    # GO if all hurdles pass; STOP if value destroyed; else REVIEW
    formula = ('=IF(OR({np}<0,{cpu}<=0),"STOP",'
               'IF(AND({gm}>=MinGM,{nm}>=MinNM,{cpu}>=MinContribUnit,{tp}<=MaxTradePct,'
               'OR({promo}=0,{roi}>=TargetROI)),"GO","REVIEW"))').format(
                   np=Rmap["np"]%c, cpu=cpu, gm=gm, nm=nm, tp=tp, promo=promo, roi=roi)
    W(scn, "%s%d"%(c,DEC), formula, f=font(10, True), fillc=WHITE, border=box, align=CENTER)
# conditional format decision row
rng = "C%d:I%d"%(DEC,DEC)
scn.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"GO"'], fill=fill(GREEN), font=font(10,True,GREENT)))
scn.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"REVIEW"'], fill=fill(YELLOW), font=font(10,True,YELLOWT)))
scn.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"STOP"'], fill=fill(RED), font=font(10,True,REDT)))

# variance vs base & best/worst
VAR = DEC+2
section(scn, VAR, "VARIANCE & RANKING", span=8)
W(scn, "B%d"%(VAR+1), "Net Profit variance vs Base", f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
for c in cols:
    W(scn, "%s%d"%(c,VAR+1), "=%s-$C$%d"%(Rmap["np"]%c, row_index["np"]), f=font(10), fillc=WHITE, border=box, fmt=FMT_USD0, align=RIGHT)
W(scn, "B%d"%(VAR+2), "Best-case Net Profit", f=font(10, True, DKTXT), fillc=GREY, border=box, align=LEFT)
W(scn, "C%d"%(VAR+2), "=MAX(C%d:I%d)"%(row_index["np"],row_index["np"]), f=font(10, True), fillc=GREY, border=box, fmt=FMT_USD0, align=RIGHT)
W(scn, "B%d"%(VAR+3), "Worst-case Net Profit", f=font(10, True, DKTXT), fillc=GREY, border=box, align=LEFT)
W(scn, "C%d"%(VAR+3), "=MIN(C%d:I%d)"%(row_index["np"],row_index["np"]), f=font(10, True), fillc=GREY, border=box, fmt=FMT_USD0, align=RIGHT)

# remember key scenario output rows for cross-tab INDEX
SCN_ROWS = row_index
SCN_DECROW = DEC

scn.column_dimensions["A"].width = 2
scn.column_dimensions["B"].width = 34
for c in cols:
    scn.column_dimensions[c].width = 13
scn.column_dimensions["K"].width = 14
scn.freeze_panes = "C%d"%(PNL0+1)
scn.sheet_view.showGridLines = False

# convenience: a formula that returns active-scenario value of a given metric row
def active(metric_key):
    rr = SCN_ROWS[metric_key]
    return "INDEX(Scenarios!$C$%d:$I$%d,MATCH(selScenario,scnHeaders,0))"%(rr,rr)
def active_decision():
    return "INDEX(Scenarios!$C$%d:$I$%d,MATCH(selScenario,scnHeaders,0))"%(SCN_DECROW,SCN_DECROW)

# =====================================================================
#  TAB 3: CALCULATIONS_GROSS_TO_NET  (detailed ACTIVE-scenario engine)
# =====================================================================
cal = wb.create_sheet("Calculations_Gross_to_Net")
cal.sheet_properties.tabColor = "548235"
title_block(cal, "CALCULATIONS — GROSS TO NET ENGINE",
            "Detailed line-item P&L for the ACTIVE scenario. Active multipliers are pulled from the Scenario tab via INDEX/MATCH. Formulas only — no hardcodes.")

cr = 5
def cput(label, formula, fmt=FMT_USD0, name=None, bold=False, band=None, note=""):
    global cr
    fillc = band if band else WHITE
    W(cal, "B%d"%cr, label, f=font(10, bold, DKTXT), fillc=fillc, border=box, align=LEFT)
    W(cal, "C%d"%cr, formula, f=font(10, bold), fillc=fillc, border=box, fmt=fmt, align=RIGHT)
    if note:
        W(cal, "D%d"%cr, note, f=font(9, False, "808080"), fillc=WHITE, align=LEFT)
    if name:
        define(name, cal, "C", cr)
    cr += 1

# active multiplier resolution block
section(cal, cr, "0.  ACTIVE ASSUMPTION RESOLUTION  (driven by selScenario)", span=3); cr += 1
def act_mult(label, drvname, name):
    cput(label, "=INDEX(%s,MATCH(selScenario,scnHeaders,0))"%drvname, '0.00', name=name, band=GREY)
act_mult("Active Volume mult",     "volMult",    "actVolM")
act_mult("Active Price mult",      "priceMult",  "actPriceM")
act_mult("Active Var-cost mult",   "vcMult",     "actVcM")
act_mult("Active Trade mult",      "tradeMult",  "actTradeM")
act_mult("Active Fill mult",       "fillMult",   "actFillM")
act_mult("Active Cannib mult",     "cannibMult", "actCannibM")
act_mult("Active Promo flag",      "promoFlag",  "actPromoF")
act_mult("Active FX stress mult",  "fxMult",     "actFxM")

# 1. Volume engine
section(cal, cr, "1.  VOLUME ENGINE", span=3); cr += 1
cput("Effective Fill Rate", "=MIN(1,FillRate*actFillM)", FMT_PCT1, name="actFill", band=GREY)
cput("Effective Units Sold", "=ForecastUnits*actVolM*actFill", FMT_NUM0, name="actUnits", bold=True)
cput("Case Sales", "=actUnits/UnitsPerCase", FMT_NUM0, name="actCases")
cput("Base (non-promo) Volume", "=BaseUnits*actVolM", FMT_NUM0, name="actBase")
cput("Gross Promo Lift Units", "=actBase*PromoLiftPct*actPromoF", FMT_NUM0, name="actLift")
cput("Cannibalized Volume", "=actLift*CannibPct*actCannibM", FMT_NUM0, name="actCannibVol")
cput("Forward-Buy Volume", "=actLift*ForwardBuyPct", FMT_NUM0, name="actFwd")
cput("Post-Promo Dip Volume", "=actLift*PostDipPct", FMT_NUM0, name="actDip")
cput("TRUE Net Incremental Volume", "=actLift-actCannibVol-actDip", FMT_NUM0, name="actTrueIncr", bold=True, band=GREY)

# 2. Revenue engine
section(cal, cr, "2.  REVENUE ENGINE", span=3); cr += 1
cput("Active Wholesale Price /unit", "=WholesalePrice*actPriceM", FMT_USD2, name="actPrice", band=GREY)
cput("Gross Sales", "=actUnits*actPrice", FMT_USD0, name="actGross", bold=True)
cput("On-Invoice Deductions", "=(ded_trade_pu*actUnits)*actTradeM", FMT_USD0, name="actOnInv",
     note="per-unit trade applied on invoice")
cput("Off-Invoice Deductions", "=(ded_trade_pct*actGross+ded_trade_fix)*actTradeM", FMT_USD0, name="actOffInv",
     note="% and fixed trade")
cput("Trade Spend $", "=(ded_trade_pct*actGross+ded_trade_pu*actUnits+ded_trade_fix)*actTradeM",
     FMT_USD0, name="actTradeSpend", bold=True)
cput("Other GtN Deductions $", "=(ded_other_pct*actGross+ded_other_pu*actUnits+ded_other_fix)*actTradeM",
     FMT_USD0, name="actOtherDed")
cput("Total Gross-to-Net Deductions", "=actTradeSpend+actOtherDed", FMT_USD0, name="actGtN", bold=True, band=GREY)
cput("Net Sales", "=actGross-actGtN", FMT_USD0, name="actNet", bold=True, band=GREY)
cput("Net Realized Price /unit", "=IF(actUnits=0,0,actNet/actUnits)", FMT_USD2, name="actNRPunit")
cput("Net Realized Price /case", "=IF(actCases=0,0,actNet/actCases)", FMT_USD2, name="actNRPcase")

# 3. Cost engine
section(cal, cr, "3.  COST ENGINE", span=3); cr += 1
cput("Variable Cost per Unit (scaled)", "=vc_pu*actVcM", FMT_USD2, name="actVCpu", band=GREY)
cput("Total Per-Unit Variable Cost", "=actVCpu*actUnits", FMT_USD0)
cput("Broker Fee $", "=BrokerPct*actGross", FMT_USD0, name="actBroker")
cput("Distributor Fee $", "=DistributorPct*actGross", FMT_USD0, name="actDist")
cput("Total Variable Cost", "=actVCpu*actUnits+vc_pct*actGross", FMT_USD0, name="actVCost", bold=True, band=GREY)
cput("Fixed Commercial Cost", "=FixedCommercialTotal", FMT_USD0, name="actFixed")
cput("Total Direct Cost", "=actVCost+actFixed", FMT_USD0, bold=True)

# 4. Profitability engine
section(cal, cr, "4.  PROFITABILITY ENGINE", span=3); cr += 1
cput("Gross Profit $", "=actGross-actVCost", FMT_USD0, name="actGP", bold=True)
cput("Gross Margin %", "=IF(actGross=0,0,actGP/actGross)", FMT_PCT1, name="actGM", bold=True, band=GREY)
cput("Contribution $", "=actNet-actVCost", FMT_USD0, name="actContrib", bold=True)
cput("Contribution Margin %", "=IF(actNet=0,0,actContrib/actNet)", FMT_PCT1, name="actCM")
cput("Net Profit $", "=actContrib-actFixed", FMT_USD0, name="actNP", bold=True, band=GREY)
cput("Net Margin %", "=IF(actNet=0,0,actNP/actNet)", FMT_PCT1, name="actNM", bold=True, band=GREY)
cput("Profit per Unit", "=IF(actUnits=0,0,actNP/actUnits)", FMT_USD2, name="actPPU")
cput("Profit per Case", "=IF(actCases=0,0,actNP/actCases)", FMT_USD2, name="actPPC")
cput("Contribution per Unit", "=IF(actUnits=0,0,actContrib/actUnits)", FMT_USD2, name="actCPU")
cput("Trade Spend % of Gross", "=IF(actGross=0,0,actTradeSpend/actGross)", FMT_PCT1, name="actTradePctG")
cput("Trade Spend % of Net", "=IF(actNet=0,0,actTradeSpend/actNet)", FMT_PCT1, name="actTradePctN")
cput("Trade Efficiency ($NP / $trade)", "=IF(actTradeSpend=0,0,actNP/actTradeSpend)", FMT_X2, name="actTradeEff")

# 5. Customer path economics
section(cal, cr, "5.  CUSTOMER PATH ECONOMICS", span=3); cr += 1
cput("Distributor Buy Price /unit", "=actPrice", FMT_USD2, name="actDistBuy", note="company wholesale price (no markup)")
cput("Distributor Sell Price /unit", "=IF(selBusinessModel=\"Distributor\",actPrice/(1-DistributorMargin),actPrice)",
     FMT_USD2, name="actDistSell")
cput("Distributor $ Margin /unit", "=actDistSell-actDistBuy", FMT_USD2, name="actDistMgnU")
cput("Distributor Margin %", "=IF(actDistSell=0,0,actDistMgnU/actDistSell)", FMT_PCT1, name="actDistMgnPct")
cput("Retail Shelf Price /unit (req'd)", "=actDistSell/(1-RetailerMargin)", FMT_USD2, name="actRetailReq")
cput("Retailer $ Margin /unit", "=actRetailReq-actDistSell", FMT_USD2, name="actRetMgnU")
cput("Retailer Margin %", "=IF(actRetailReq=0,0,actRetMgnU/actRetailReq)", FMT_PCT1, name="actRetMgnPct")
cput("Company Net Realization /unit", "=actNRPunit", FMT_USD2)
cput("Company Net Realization /case", "=actNRPcase", FMT_USD2)

# 6. Margin bridge
section(cal, cr, "6.  MARGIN BRIDGE  (Gross Sales → Net Profit)", span=3); cr += 1
cput("Gross Sales", "=actGross", FMT_USD0, band=GREY)
cput("  less On-Invoice Deductions", "=-actOnInv", FMT_USD0)
cput("  less Off-Invoice Deductions", "=-actOffInv", FMT_USD0)
cput("→ Net Sales", "=actGross-actOnInv-actOffInv-actOtherDed", FMT_USD0, bold=True)
cput("  less Product & Landed Cost", "=-(actVCpu*actUnits)", FMT_USD0)
cput("  less Broker & Distributor Fees", "=-(actBroker+actDist)", FMT_USD0)
cput("→ Contribution", "=actContrib", FMT_USD0, bold=True, band=GREY)
cput("  less Fixed Program Costs", "=-actFixed", FMT_USD0)
cput("→ NET PROFIT", "=actNP", FMT_USD0, bold=True, band=GREY)

# 7. Incremental economics (active vs Base scenario)
section(cal, cr, "7.  INCREMENTAL ECONOMICS  (active vs Base scenario)", span=3); cr += 1
base_net = "Scenarios!$C$%d"%SCN_ROWS["net"]
base_np  = "Scenarios!$C$%d"%SCN_ROWS["np"]
base_trade = "Scenarios!$C$%d"%SCN_ROWS["trade"]
base_vc  = "Scenarios!$C$%d"%SCN_ROWS["vcost"]
cput("Incremental Net Sales", "=actNet-%s"%base_net, FMT_USD0, name="incrNet")
cput("Incremental Trade Spend", "=actTradeSpend-%s"%base_trade, FMT_USD0, name="incrTrade")
cput("Incremental Variable Cost", "=actVCost-%s"%base_vc, FMT_USD0, name="incrVC")
cput("Incremental Contribution", "=incrNet-incrVC", FMT_USD0, name="incrContrib", bold=True)
cput("Incremental Net Profit", "=actNP-%s"%base_np, FMT_USD0, name="incrNP", bold=True, band=GREY)
cput("Promo Spend (incremental trade)", "=MAX(0,incrTrade)", FMT_USD0, name="promoSpend")
cput("Promotion ROI ($NP per $spend)", "=IF(promoSpend=0,0,incrNP/promoSpend)", FMT_X2, name="actROI", bold=True, band=GREY)
cput("Payback Multiple", "=IF(promoSpend=0,0,(incrNP+promoSpend)/promoSpend)", FMT_X2, name="actPayback")

# 8. Break-even engine
section(cal, cr, "8.  BREAK-EVEN ENGINE", span=3); cr += 1
cput("Break-Even Units", '=IF(actCPU<=0,"contrib<=0",actFixed/actCPU)', FMT_NUM0, name="beUnits", bold=True)
cput("Break-Even Cases", '=IF(actCPU<=0,"n/a",beUnits/UnitsPerCase)', FMT_NUM0, name="beCases")
# Break-even wholesale price to hit Target NM  (algebraic solution; see Read_Me logic notes)
# a = trade%+other% (scaled), b = (trade_pu+other_pu)*tradeM, c = (trade_fix+other_fix)*tradeM
# vc = vc_pu*vcM ; f = vc_pct ; U = actUnits ; Fixed = actFixed ; t=TargetNM
cput("  helper a (ded % of gross)", "=(ded_trade_pct+ded_other_pct)*actTradeM", FMT_PCT1, name="beA")
cput("  helper b (ded $/unit)", "=(ded_trade_pu+ded_other_pu)*actTradeM", FMT_USD2, name="beB")
cput("  helper c (ded fixed $)", "=(ded_trade_fix+ded_other_fix)*actTradeM", FMT_USD0, name="beC")
cput("Break-Even Wholesale Price (Target NM)",
     "=IF((1-beA-vc_pct-TargetNM*(1-beA))=0,0,"
     "((beB+vc_pu*actVcM)*actUnits+beC+actFixed-TargetNM*(beB*actUnits+beC))"
     "/((1-beA-vc_pct-TargetNM*(1-beA))*actUnits))",
     FMT_USD2, name="beWholesale", bold=True, band=GREY)
cput("Break-Even Retail Price", "=beWholesale/(1-RetailerMargin)/IF(selBusinessModel=\"Distributor\",(1-DistributorMargin),1)",
     FMT_USD2, name="beRetail")
cput("Break-Even Promo Lift %",
     '=IF(actCPU<=0,"n/a",promoSpend/(actCPU*actBase))', FMT_PCT1, name="beLift")
cput("Break-Even Velocity /store/wk",
     '=IF(OR(NumStores=0,PromoWeeks=0),"n/a",beUnits/NumStores/PromoWeeks)', FMT_USD2, name="beVel")
cput("Break-Even Store Count",
     '=IF(OR(VelocityPSPW=0,PromoWeeks=0),"n/a",beUnits/VelocityPSPW/PromoWeeks)', FMT_NUM0, name="beStores")
cput("Break-Even Promo Weeks",
     '=IF(OR(VelocityPSPW=0,NumStores=0),"n/a",beUnits/VelocityPSPW/NumStores)', FMT_NUM0, name="beWeeks")

# 9. Sensitivity engine (data-table style: deltas applied to active)
section(cal, cr, "9.  SENSITIVITY ENGINE  (impact on Net Profit / margin)", span=6); cr += 1
SENS_HDR = cr
W(cal, "B%d"%cr, "Driver shock", f=font(9, True, WHITE), fillc=STEEL, border=box, align=LEFT)
for i,h in enumerate(["-10%","-5%","Active","+5%","+10%"]):
    W(cal, "%s%d"%(["C","D","E","F","G"][i],cr), h, f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)
cr += 1
# Price -> Net Profit  (approx: NP changes ~ dPrice*Units*(1-a) feeding through)
def sens_row(label, deltas, base_formula_fn):
    global cr
    W(cal, "B%d"%cr, label, f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
    for i,d in enumerate(deltas):
        col=["C","D","E","F","G"][i]
        W(cal, "%s%d"%(col,cr), base_formula_fn(d), f=font(10), fillc=WHITE, border=box, fmt=FMT_USD0, align=RIGHT)
    cr += 1
# Net profit sensitivity to price: NP + dP*Units*(1-a-f effect approx via net & margin)
sens_row("Price → Net Profit", [-0.10,-0.05,0,0.05,0.10],
         lambda d: "=actNP+(%g)*actPrice*actUnits*(1-beA-vc_pct)"%d)
sens_row("Volume → Net Profit", [-0.10,-0.05,0,0.05,0.10],
         lambda d: "=actNP+(%g)*actContrib"%d)
sens_row("Cost inflation → Net Profit", [0.10,0.05,0,-0.05,-0.10],
         lambda d: "=actNP-(%g)*actVCpu*actUnits"%d)   # +cost reduces NP; col order high->low
sens_row("Trade % → Net Profit", [0.10,0.05,0,-0.05,-0.10],
         lambda d: "=actNP-(%g)*actTradeSpend"%d)
sens_row("FX shock → Net Profit", [-0.10,-0.05,0,0.05,0.10],
         lambda d: "=actNP*(1+(%g))"%d)

# Target price / target cost logic
section(cal, cr, "10.  TARGET PRICE / TARGET COST LOGIC", span=4); cr += 1
cput("Wholesale Price to hit Target GM",
     "=(vc_pu*actVcM)/(1-TargetGM-vc_pct)", FMT_USD2, name="tgtPriceGM",
     note="solves GM%=target holding cost")
cput("Wholesale Price to hit Target NM", "=beWholesale", FMT_USD2)
cput("Variable Cost /unit to hit Target GM (price fixed)",
     "=actPrice*(1-TargetGM)-vc_pct*actPrice", FMT_USD2, name="tgtCostGM")
cput("Required cost reduction /unit", "=MAX(0,actVCpu-tgtCostGM)", FMT_USD2, name="tgtCostGap")

cal.column_dimensions["A"].width = 2
cal.column_dimensions["B"].width = 40
cal.column_dimensions["C"].width = 16
cal.column_dimensions["D"].width = 34
for c in "EFG":
    cal.column_dimensions[c].width = 13
cal.freeze_panes = "B5"
cal.sheet_view.showGridLines = False

# Active decision string (computed here for reuse)
define("actDecisionRow", scn, "C", SCN_DECROW)  # not a single cell; we use active_decision()

# =====================================================================
#  Fill Inputs KPI snapshot + control-bar status (needs Calc names)
# =====================================================================
section(inp, KPI_SNAP_ROW, "H.  LIVE KPI SNAPSHOT  (active scenario)", span=8)
ks = KPI_SNAP_ROW+1
snap = [
    ("Gross Sales","=actGross",FMT_USD0), ("Net Sales","=actNet",FMT_USD0),
    ("Gross Margin %","=actGM",FMT_PCT1), ("Net Margin %","=actNM",FMT_PCT1),
    ("Trade Spend $","=actTradeSpend",FMT_USD0), ("Trade Spend % Gross","=actTradePctG",FMT_PCT1),
    ("Contribution $","=actContrib",FMT_USD0), ("Net Profit $","=actNP",FMT_USD0),
    ("Profit / Unit","=actPPU",FMT_USD2), ("Break-Even Units","=beUnits",FMT_NUM0),
    ("Promotion ROI (x)","=actROI",FMT_X2),
]
half = (len(snap)+1)//2
for i,(lab,frm,fmt) in enumerate(snap):
    cc = "B" if i < half else "E"
    vc = "C" if i < half else "F"
    rr = ks + (i if i < half else i-half)
    W(inp, "%s%d"%(cc,rr), lab, f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
    W(inp, "%s%d"%(vc,rr), frm, f=font(10, True), fillc=WHITE, border=box, fmt=fmt, align=RIGHT)
# decision cell
decr = ks + half + 1
W(inp, "B%d"%decr, "DECISION STATUS", f=font(11, True, WHITE), fillc=NAVY, border=box, align=LEFT)
W(inp, "C%d"%decr, "="+active_decision(), f=font(12, True), fillc=WHITE, border=box, align=CENTER)
inp.conditional_formatting.add("C%d"%decr, CellIsRule(operator="equal", formula=['"GO"'], fill=fill(GREEN), font=font(12,True,GREENT)))
inp.conditional_formatting.add("C%d"%decr, CellIsRule(operator="equal", formula=['"REVIEW"'], fill=fill(YELLOW), font=font(12,True,YELLOWT)))
inp.conditional_formatting.add("C%d"%decr, CellIsRule(operator="equal", formula=['"STOP"'], fill=fill(RED), font=font(12,True,REDT)))

# control bar decision + integrity (set placeholders now -> reference)
W(inp, "I6", "="+active_decision(), f=font(11, True), fillc=WHITE, border=box, align=CENTER)
inp.conditional_formatting.add("I6:J6", CellIsRule(operator="equal", formula=['"GO"'], fill=fill(GREEN), font=font(11,True,GREENT)))
inp.conditional_formatting.add("I6:J6", CellIsRule(operator="equal", formula=['"REVIEW"'], fill=fill(YELLOW), font=font(11,True,YELLOWT)))
inp.conditional_formatting.add("I6:J6", CellIsRule(operator="equal", formula=['"STOP"'], fill=fill(RED), font=font(11,True,REDT)))
# integrity references the Checks tab overall flag (defined later as chkOverall)

# =====================================================================
#  TAB 5: PROMOTIONS_AND_ROI
# =====================================================================
pro = wb.create_sheet("Promotions_and_ROI")
pro.sheet_properties.tabColor = "C55A11"
title_block(pro, "PROMOTIONS & ROI",
            "Tactic-by-tactic incremental economics. Each promo is judged on TRUE incremental profit, not gross sales lift. Flags expose value-destroying programs.")

section(pro, 5, "PER-TACTIC PROMOTION ANALYSIS  (evaluated at base volume & active cost)", span=13)
PH = 6
promo_cols = [
    ("B","Tactic"),("C","Spend $"),("D","Base Vol"),("E","Lift %"),
    ("F","Incr Vol (gross)"),("G","Cannib"),("H","True Incr Vol"),
    ("I","Contrib/unit"),("J","Incr Contrib $"),("K","Incr Net Profit $"),
    ("L","ROI (x)"),("M","BE Lift %"),("N","Verdict"),
]
for cidx,(col,txt) in enumerate(promo_cols):
    W(pro, "%s%d"%(col,PH), txt, f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)

# tactic list with spend ($) and assumed lift % of base; cannib% as fraction of incr
tactics = [
    ("TPD",            "=ded_trade_pct*actGross*0.30+0", 0.18, 0.10),
    ("End Cap",        18000, 0.12, 0.15),
    ("Fence / Feature",12000, 0.10, 0.12),
    ("Display Allowance",12000, 0.14, 0.18),
    ("Advertising",    9000,  0.08, 0.05),
    ("Flyer Support",  7500,  0.07, 0.06),
    ("Digital Support",6000,  0.06, 0.04),
    ("Demo Support",   4000,  0.05, 0.03),
]
PR0 = PH+1
prow = PR0
for name_t, spend, lift, cannib in tactics:
    W(pro, "B%d"%prow, name_t, f=font(10), fillc=WHITE, border=box, align=LEFT)
    # spend
    if isinstance(spend, str):
        W(pro, "C%d"%prow, "=18000", f=font(10), fillc=INPUTYEL, border=box, fmt=FMT_USD0, align=RIGHT, locked=False)
    else:
        W(pro, "C%d"%prow, spend, f=font(10), fillc=INPUTYEL, border=box, fmt=FMT_USD0, align=RIGHT, locked=False)
    W(pro, "D%d"%prow, "=actBase", f=font(10), fillc=WHITE, border=box, fmt=FMT_NUM0, align=RIGHT)
    W(pro, "E%d"%prow, lift, f=font(10), fillc=INPUTYEL, border=box, fmt=FMT_PCT1, align=RIGHT, locked=False)
    W(pro, "F%d"%prow, "=D%d*E%d"%(prow,prow), f=font(10), fillc=WHITE, border=box, fmt=FMT_NUM0, align=RIGHT)
    W(pro, "G%d"%prow, cannib, f=font(10), fillc=INPUTYEL, border=box, fmt=FMT_PCT1, align=RIGHT, locked=False)
    W(pro, "H%d"%prow, "=F%d*(1-G%d)"%(prow,prow), f=font(10), fillc=WHITE, border=box, fmt=FMT_NUM0, align=RIGHT)
    W(pro, "I%d"%prow, "=actCPU", f=font(10), fillc=WHITE, border=box, fmt=FMT_USD2, align=RIGHT)
    W(pro, "J%d"%prow, "=H%d*I%d"%(prow,prow), f=font(10), fillc=WHITE, border=box, fmt=FMT_USD0, align=RIGHT)
    W(pro, "K%d"%prow, "=J%d-C%d"%(prow,prow), f=font(10, True), fillc=WHITE, border=box, fmt=FMT_USD0, align=RIGHT)
    W(pro, "L%d"%prow, "=IF(C%d=0,0,K%d/C%d)"%(prow,prow,prow), f=font(10, True), fillc=WHITE, border=box, fmt=FMT_X2, align=RIGHT)
    W(pro, "M%d"%prow, "=IF(OR(actCPU<=0,D%d=0),\"n/a\",C%d/(actCPU*D%d))"%(prow,prow,prow), f=font(10), fillc=WHITE, border=box, fmt=FMT_PCT1, align=RIGHT)
    W(pro, "N%d"%prow, '=IF(K%d<0,"STOP",IF(L%d>=TargetROI,"GO","REVIEW"))'%(prow,prow),
      f=font(10, True), fillc=WHITE, border=box, align=CENTER)
    prow += 1
PRend = prow-1
# combined stack row
W(pro, "B%d"%prow, "COMBINED STACK", f=font(10, True, DKTXT), fillc=GREY, border=box, align=LEFT)
for col in "CFHJK":
    W(pro, "%s%d"%(col,prow), "=SUM(%s%d:%s%d)"%(col,PR0,col,PRend), f=font(10, True), fillc=GREY, border=box,
      fmt=(FMT_USD0 if col in "CJK" else FMT_NUM0), align=RIGHT)
W(pro, "L%d"%prow, "=IF(C%d=0,0,K%d/C%d)"%(prow,prow,prow), f=font(10, True), fillc=GREY, border=box, fmt=FMT_X2, align=RIGHT)
W(pro, "N%d"%prow, '=IF(K%d<0,"STOP",IF(L%d>=TargetROI,"GO","REVIEW"))'%(prow,prow),
  f=font(10, True), fillc=GREY, border=box, align=CENTER)
STACKROW = prow

# conditional format verdict column
vrng = "N%d:N%d"%(PR0,STACKROW)
pro.conditional_formatting.add(vrng, CellIsRule(operator="equal", formula=['"GO"'], fill=fill(GREEN), font=font(10,True,GREENT)))
pro.conditional_formatting.add(vrng, CellIsRule(operator="equal", formula=['"REVIEW"'], fill=fill(YELLOW), font=font(10,True,YELLOWT)))
pro.conditional_formatting.add(vrng, CellIsRule(operator="equal", formula=['"STOP"'], fill=fill(RED), font=font(10,True,REDT)))
# ROI below threshold -> red text
pro.conditional_formatting.add("L%d:L%d"%(PR0,STACKROW),
    CellIsRule(operator="lessThan", formula=["TargetROI"], font=font(10,True,REDT)))

# flags section
FL = STACKROW+2
section(pro, FL, "AUTOMATIC FLAGS", span=8)
flags = [
    ("Any tactic ROI below hurdle", '=IF(COUNTIF(L%d:L%d,"<"&TargetROI)>0,"⚠ YES","OK")'%(PR0,PRend)),
    ("Any negative incremental profit", '=IF(COUNTIF(K%d:K%d,"<0")>0,"⚠ YES","OK")'%(PR0,PRend)),
    ("Combined stack ROI below hurdle", '=IF(L%d<TargetROI,"⚠ YES","OK")'%STACKROW),
    ("Combined trade > max trade %", '=IF((C%d)/actGross>MaxTradePct,"⚠ YES","OK")'%STACKROW),
    ("Profit per $1 promo spend (stack)", "=IF(C%d=0,0,K%d/C%d)"%(STACKROW,STACKROW,STACKROW)),
]
fr = FL+1
for lab, frm in flags:
    W(pro, "B%d"%fr, lab, f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
    W(pro, "C%d"%fr, frm, f=font(10, True), fillc=WHITE, border=box, align=CENTER,
      fmt=(FMT_X2 if "per $1" in lab else None))
    pro.conditional_formatting.add("C%d"%fr, CellIsRule(operator="equal", formula=['"⚠ YES"'], fill=fill(RED), font=font(10,True,REDT)))
    pro.conditional_formatting.add("C%d"%fr, CellIsRule(operator="equal", formula=['"OK"'], fill=fill(GREEN), font=font(10,True,GREENT)))
    fr += 1

pro.column_dimensions["A"].width = 2
pro.column_dimensions["B"].width = 22
for col,txt in promo_cols[1:]:
    pro.column_dimensions[col].width = 12
pro.freeze_panes = "B%d"%(PH+1)
pro.sheet_view.showGridLines = False

# =====================================================================
#  TAB 6: CURRENCY_AND_FX
# =====================================================================
fxs = wb.create_sheet("Currency_and_FX")
fxs.sheet_properties.tabColor = "7030A0"
title_block(fxs, "CURRENCY & FX",
            "Edit the rate table (USD value of 1 unit of each currency). Switch base/reporting currency on the Inputs tab. KPIs translate automatically.")

section(fxs, 5, "FX RATE TABLE  (USD value of 1 unit of currency)", span=4)
W(fxs, "B6", "Currency", f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)
W(fxs, "C6", "USD per 1 unit", f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)
W(fxs, "D6", "Note", f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)
rates = [("USD",1.0,"reference"),("CAD",0.73,"editable"),("JPY",0.0064,"editable"),("KRW",0.00072,"editable")]
FXTOP = 7
for i,(c,v,note) in enumerate(rates):
    rr=FXTOP+i
    W(fxs, "B%d"%rr, c, f=font(10, True), fillc=WHITE, border=box, align=CENTER)
    W(fxs, "C%d"%rr, v, f=font(10, True), fillc=INPUTYEL, border=box, fmt='0.00000', align=RIGHT, locked=False)
    W(fxs, "D%d"%rr, note, f=font(9,False,"808080"), fillc=WHITE, border=box, align=LEFT)
FXBOT=FXTOP+len(rates)-1
define_range("fxCcy", fxs, "B", FXTOP, "B", FXBOT)
define_range("fxRate", fxs, "C", FXTOP, "C", FXBOT)

# rate lookups for base & reporting
section(fxs, 12, "CONVERSION FACTORS", span=4)
W(fxs, "B13", "Base currency (USD value of 1 unit)", f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
W(fxs, "C13", "=INDEX(fxRate,MATCH(selBaseCurrency,fxCcy,0))", f=font(10, True), fillc=WHITE, border=box, fmt='0.00000', align=RIGHT)
define("fxBaseRate", fxs, "C", 13)
W(fxs, "B14", "Reporting currency (USD value of 1 unit)", f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
W(fxs, "C14", "=INDEX(fxRate,MATCH(selReportingCurrency,fxCcy,0))", f=font(10, True), fillc=WHITE, border=box, fmt='0.00000', align=RIGHT)
define("fxRptRate", fxs, "C", 14)
W(fxs, "B15", "Base → Reporting factor", f=font(10, True, DKTXT), fillc=GREY, border=box, align=LEFT)
W(fxs, "C15", "=fxBaseRate/fxRptRate", f=font(10, True), fillc=GREY, border=box, fmt='0.00000', align=RIGHT)
define("fxFactor", fxs, "C", 15)
W(fxs, "D15", "value_in_reporting = value_base × this", f=font(9,False,"808080"), align=LEFT)

# translated KPI table across all 4 currencies
section(fxs, 18, "TRANSLATED KPI TABLE  (active scenario, all currencies)", span=6)
TKH=19
W(fxs, "B%d"%TKH, "KPI (in base currency)", f=font(9, True, WHITE), fillc=STEEL, border=box, align=LEFT)
W(fxs, "C%d"%TKH, "Base value", f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)
for i,c in enumerate(CCY):
    W(fxs, "%s%d"%(["D","E","F","G"][i],TKH), c, f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)
kpis_fx = [("Gross Sales","actGross"),("Net Sales","actNet"),("Trade Spend $","actTradeSpend"),
           ("Contribution $","actContrib"),("Net Profit $","actNP")]
tr=TKH+1
for lab,nm in kpis_fx:
    W(fxs, "B%d"%tr, lab, f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
    W(fxs, "C%d"%tr, "=%s"%nm, f=font(10, True), fillc=WHITE, border=box, fmt=FMT_USD0, align=RIGHT)
    for i,c in enumerate(CCY):
        col=["D","E","F","G"][i]
        # value in currency c = base_value * fxBaseRate / rate(c)
        W(fxs, "%s%d"%(col,tr),
          "=%s*fxBaseRate/INDEX(fxRate,MATCH(\"%s\",fxCcy,0))"%(nm,c),
          f=font(10), fillc=WHITE, border=box, fmt=FMT_USD0, align=RIGHT)
    tr += 1

fxs.column_dimensions["A"].width = 2
fxs.column_dimensions["B"].width = 32
for c in "CDEFG":
    fxs.column_dimensions[c].width = 14
fxs.sheet_view.showGridLines = False

# =====================================================================
#  TAB 8: CHECKS_CONTROLS_AND_AUDIT
# =====================================================================
chk = wb.create_sheet("Checks_Controls_and_Audit")
chk.sheet_properties.tabColor = "C00000"
title_block(chk, "CHECKS, CONTROLS & AUDIT",
            "Automated integrity tests. If any CRITICAL check fails, the model is flagged UNSAFE and the banner turns red. Review before any business decision.")

# banner (filled after overall computed)
chk.merge_cells("B5:N5")
W(chk, "B5", "MODEL INTEGRITY BANNER", f=font(12, True, WHITE), fillc=NAVY, align=CENTER)
for cc in range(2,15): chk.cell(row=5,column=cc).fill=fill(NAVY)
chk.row_dimensions[5].height=22

section(chk, 7, "CONTROL CHECKS", span=8)
CH=8
W(chk, "B%d"%CH, "#", f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)
W(chk, "C%d"%CH, "Check", f=font(9, True, WHITE), fillc=STEEL, border=box, align=LEFT)
W(chk, "D%d"%CH, "Expected", f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)
W(chk, "E%d"%CH, "Actual / delta", f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)
W(chk, "F%d"%CH, "Status", f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)
W(chk, "G%d"%CH, "Severity", f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)

# reconciliation reference: active vs scenario-selected column
act_net_scn = active("net")
checks = [
    ("Net Sales ≤ Gross Sales", '=IF(actNet<=actGross,"PASS","FAIL")', "actNet", "TRUE", "Critical"),
    ("Calc Net Sales = Scenario-table Net Sales",
     '=IF(ABS(actNet-(%s))<1,"PASS","FAIL")'%act_net_scn, "ABS<$1", "=actNet-(%s)"%act_net_scn, "Critical"),
    ("Calc Net Profit = Scenario-table Net Profit",
     '=IF(ABS(actNP-(%s))<1,"PASS","FAIL")'%active("np"), "ABS<$1", "=actNP-(%s)"%active("np"), "Critical"),
    ("Trade Spend reconciles (Calc vs Scenario)",
     '=IF(ABS(actTradeSpend-(%s))<1,"PASS","FAIL")'%active("trade"), "ABS<$1", "=actTradeSpend-(%s)"%active("trade"), "Critical"),
    ("Gross Margin %% ties to $ (GP/Gross)",
     '=IF(ABS(actGM-IF(actGross=0,0,actGP/actGross))<0.0001,"PASS","FAIL")', "≈0", "", "High"),
    ("Net Margin %% ties to $ (NP/Net)",
     '=IF(ABS(actNM-IF(actNet=0,0,actNP/actNet))<0.0001,"PASS","FAIL")', "≈0", "", "High"),
    ("Per-unit × units ties to total NP",
     '=IF(ABS(actPPU*actUnits-actNP)<1,"PASS","FAIL")', "ABS<$1", "=actPPU*actUnits-actNP", "High"),
    ("Scenario selector is valid",
     '=IF(COUNTIF(scnHeaders,selScenario)=1,"PASS","FAIL")', "found", "", "Critical"),
    ("Break-even contribution is positive",
     '=IF(actCPU>0,"PASS","WARN")', ">0", "=actCPU", "High"),
    ("Trade Spend %% within ceiling",
     '=IF(actTradePctG<=MaxTradePct,"PASS","WARN")', "≤max", "=actTradePctG", "Medium"),
    ("Gross Margin %% meets minimum hurdle",
     '=IF(actGM>=MinGM,"PASS","WARN")', "≥min", "=actGM", "Medium"),
    ("Net Margin %% meets minimum hurdle",
     '=IF(actNM>=MinNM,"PASS","WARN")', "≥min", "=actNM", "Medium"),
    ("FX factor reconciles (base/report)",
     '=IF(ABS(fxFactor-fxBaseRate/fxRptRate)<0.0000001,"PASS","FAIL")', "≈0", "", "High"),
    ("FX translation non-zero when rate set",
     '=IF(AND(fxBaseRate>0,fxRptRate>0),"PASS","FAIL")', ">0", "", "High"),
    ("No negative Effective Units",
     '=IF(actUnits>=0,"PASS","FAIL")', "≥0", "=actUnits", "Critical"),
    ("Required assumptions present (price, units)",
     '=IF(AND(WholesalePrice>0,ForecastUnits>0,UnitsPerCase>0),"PASS","FAIL")', "all>0", "", "Critical"),
    ("Fixed cost subtotal non-negative",
     '=IF(FixedCommercialTotal>=0,"PASS","FAIL")', "≥0", "=FixedCommercialTotal", "Medium"),
    ("Net Realized Price ≤ Wholesale Price",
     '=IF(actNRPunit<=actPrice+0.0001,"PASS","WARN")', "≤", "=actPrice-actNRPunit", "Medium"),
]
cr2=CH+1
chk_first=cr2
for i,(lab, statf, exp, act, sev) in enumerate(checks):
    W(chk, "B%d"%cr2, i+1, f=font(9), fillc=WHITE, border=box, align=CENTER)
    W(chk, "C%d"%cr2, lab, f=font(10), fillc=WHITE, border=box, align=LEFT)
    W(chk, "D%d"%cr2, exp, f=font(9,False,"808080"), fillc=WHITE, border=box, align=CENTER)
    if act:
        W(chk, "E%d"%cr2, act, f=font(9), fillc=WHITE, border=box, fmt=FMT_USD2, align=RIGHT)
    else:
        W(chk, "E%d"%cr2, "—", f=font(9,False,"808080"), fillc=WHITE, border=box, align=CENTER)
    W(chk, "F%d"%cr2, statf, f=font(10, True), fillc=WHITE, border=box, align=CENTER)
    W(chk, "G%d"%cr2, sev, f=font(9), fillc=WHITE, border=box, align=CENTER)
    cr2 += 1
chk_last=cr2-1
# conditional format status
srng="F%d:F%d"%(chk_first,chk_last)
chk.conditional_formatting.add(srng, CellIsRule(operator="equal", formula=['"PASS"'], fill=fill(GREEN), font=font(10,True,GREENT)))
chk.conditional_formatting.add(srng, CellIsRule(operator="equal", formula=['"WARN"'], fill=fill(YELLOW), font=font(10,True,YELLOWT)))
chk.conditional_formatting.add(srng, CellIsRule(operator="equal", formula=['"FAIL"'], fill=fill(RED), font=font(10,True,REDT)))

# summary
SU=chk_last+2
section(chk, SU, "MODEL HEALTH SUMMARY", span=6)
W(chk, "B%d"%(SU+1), "Total checks", f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
W(chk, "C%d"%(SU+1), "=COUNTA(C%d:C%d)"%(chk_first,chk_last), f=font(10, True), fillc=WHITE, border=box, align=CENTER)
W(chk, "B%d"%(SU+2), "Passes", f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
W(chk, "C%d"%(SU+2), '=COUNTIF(F%d:F%d,"PASS")'%(chk_first,chk_last), f=font(10, True), fillc=WHITE, border=box, align=CENTER)
W(chk, "B%d"%(SU+3), "Warnings", f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
W(chk, "C%d"%(SU+3), '=COUNTIF(F%d:F%d,"WARN")'%(chk_first,chk_last), f=font(10, True), fillc=WHITE, border=box, align=CENTER)
W(chk, "B%d"%(SU+4), "Critical failures", f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
# count FAIL among Critical severity rows
W(chk, "C%d"%(SU+4), '=COUNTIFS(F%d:F%d,"FAIL",G%d:G%d,"Critical")'%(chk_first,chk_last,chk_first,chk_last),
  f=font(10, True), fillc=WHITE, border=box, align=CENTER)
W(chk, "B%d"%(SU+5), "OVERALL MODEL STATUS", f=font(11, True, WHITE), fillc=NAVY, border=box, align=LEFT)
W(chk, "C%d"%(SU+5),
  '=IF(C%d>0,"UNSAFE — CRITICAL FAIL",IF(COUNTIF(F%d:F%d,"FAIL")>0,"REVIEW — CHECK FAILS",IF(C%d>0,"CAUTION — WARNINGS","HEALTHY")))'
  %(SU+4,chk_first,chk_last,SU+3),
  f=font(11, True), fillc=WHITE, border=box, align=CENTER)
define("chkOverall", chk, "C", SU+5)
chk.conditional_formatting.add("C%d"%(SU+5), CellIsRule(operator="equal", formula=['"HEALTHY"'], fill=fill(GREEN), font=font(11,True,GREENT)))
chk.conditional_formatting.add("C%d"%(SU+5), FormulaRule(formula=['ISNUMBER(SEARCH("UNSAFE",C%d))'%(SU+5)], fill=fill(RED), font=font(11,True,REDT)))
chk.conditional_formatting.add("C%d"%(SU+5), FormulaRule(formula=['ISNUMBER(SEARCH("REVIEW",C%d))'%(SU+5)], fill=fill(YELLOW), font=font(11,True,YELLOWT)))
chk.conditional_formatting.add("C%d"%(SU+5), FormulaRule(formula=['ISNUMBER(SEARCH("CAUTION",C%d))'%(SU+5)], fill=fill(YELLOW), font=font(11,True,YELLOWT)))

# fill banner now
W(chk, "B5", '=IF(chkOverall="HEALTHY","✔ MODEL HEALTHY — SAFE TO USE","⚠ "&chkOverall)', f=font(12, True, WHITE), align=CENTER)
chk.conditional_formatting.add("B5", FormulaRule(formula=['ISNUMBER(SEARCH("HEALTHY",B5))'], fill=fill(GREENT), font=font(12,True,WHITE)))
chk.conditional_formatting.add("B5", FormulaRule(formula=['ISNUMBER(SEARCH("UNSAFE",B5))'], fill=fill(REDT), font=font(12,True,WHITE)))

chk.column_dimensions["A"].width = 2
chk.column_dimensions["B"].width = 6
chk.column_dimensions["C"].width = 42
chk.column_dimensions["D"].width = 12
chk.column_dimensions["E"].width = 16
chk.column_dimensions["F"].width = 12
chk.column_dimensions["G"].width = 11
chk.freeze_panes = "B%d"%(CH+1)
chk.sheet_view.showGridLines = False

# now wire control-bar integrity on Inputs
W(inp, "I7", "=chkOverall", f=font(11, True), fillc=WHITE, border=box, align=CENTER)
inp.conditional_formatting.add("I7:J7", FormulaRule(formula=['ISNUMBER(SEARCH("HEALTHY",I7))'], fill=fill(GREEN), font=font(11,True,GREENT)))
inp.conditional_formatting.add("I7:J7", FormulaRule(formula=['ISNUMBER(SEARCH("UNSAFE",I7))'], fill=fill(RED), font=font(11,True,REDT)))

# =====================================================================
#  TAB 7: DASHBOARD_EXECUTIVE_VIEW
# =====================================================================
dsh = wb.create_sheet("Dashboard_Executive_View")
dsh.sheet_properties.tabColor = NAVY
title_block(dsh, "EXECUTIVE DASHBOARD",
            "One-screen readout for CFO / GM / SVP Sales / RGM / Commercial Finance. Color = green above hurdle, yellow caution, red below hurdle / value destruction.")

def kpi_card(coord_lab, coord_val, label, formula, fmt, big=False):
    W(dsh, coord_lab, label, f=font(9, True, DKTXT), fillc=LTBLUE, border=box, align=CENTER)
    W(dsh, coord_val, formula, f=font(14 if big else 12, True), fillc=WHITE, border=box, fmt=fmt, align=CENTER)

# A. header summary
section(dsh, 5, "A.  HEADER", span=12)
hdr_items = [
    ("B6","Product","C6","='Inputs_and_Control_Center'!C14"),
    ("E6","Customer","F6","='Inputs_and_Control_Center'!C16"),
    ("H6","Channel","I6","='Inputs_and_Control_Center'!C17"),
    ("B7","Scenario","C7","=selScenario"),
    ("E7","Base Ccy","F7","=selBaseCurrency"),
    ("H7","Reporting Ccy","I7","=selReportingCurrency"),
]
for lc,lab,vc,frm in hdr_items:
    W(dsh, lc, lab, f=font(9, True, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
    W(dsh, vc, frm, f=font(10, True), fillc=WHITE, border=box, align=LEFT)
# decision + integrity big
W(dsh, "K6", "DECISION", f=font(10, True, WHITE), fillc=NAVY, border=box, align=CENTER)
W(dsh, "L6", "="+active_decision(), f=font(14, True), fillc=WHITE, border=box, align=CENTER)
W(dsh, "K7", "INTEGRITY", f=font(10, True, WHITE), fillc=NAVY, border=box, align=CENTER)
W(dsh, "L7", "=chkOverall", f=font(10, True), fillc=WHITE, border=box, align=CENTER)
dsh.conditional_formatting.add("L6", CellIsRule(operator="equal", formula=['"GO"'], fill=fill(GREEN), font=font(14,True,GREENT)))
dsh.conditional_formatting.add("L6", CellIsRule(operator="equal", formula=['"REVIEW"'], fill=fill(YELLOW), font=font(14,True,YELLOWT)))
dsh.conditional_formatting.add("L6", CellIsRule(operator="equal", formula=['"STOP"'], fill=fill(RED), font=font(14,True,REDT)))
dsh.conditional_formatting.add("L7", FormulaRule(formula=['ISNUMBER(SEARCH("HEALTHY",L7))'], fill=fill(GREEN), font=font(10,True,GREENT)))
dsh.conditional_formatting.add("L7", FormulaRule(formula=['ISNUMBER(SEARCH("UNSAFE",L7))'], fill=fill(RED), font=font(10,True,REDT)))

# B. core financial KPIs
section(dsh, 9, "B.  CORE FINANCIAL KPIs (active scenario, base currency)", span=12)
core = [
    ("Gross Sales","=actGross",FMT_USD0),("Net Sales","=actNet",FMT_USD0),
    ("Gross Profit $","=actGP",FMT_USD0),("Gross Margin %","=actGM",FMT_PCT1),
    ("Contribution $","=actContrib",FMT_USD0),("Contribution Margin %","=actCM",FMT_PCT1),
    ("Net Profit $","=actNP",FMT_USD0),("Net Margin %","=actNM",FMT_PCT1),
]
positions = ["B","D","F","H","J","L"]
rr=10
ci=0
for lab,frm,fmt in core:
    col = ["B","D","F","H"][ci%4]
    if ci%4==0 and ci>0: rr+=3
    lc="%s%d"%(col,rr); vc="%s%d"%(col,rr+1)
    kpi_card(lc,vc,lab,frm,fmt)
    ci+=1
# margin cells get hurdle coloring
for addr,hur in [("D11","actGM"),]:
    pass
# color GM/NM vs hurdle
dsh.conditional_formatting.add("H11", CellIsRule(operator="greaterThanOrEqual", formula=["MinGM"], fill=fill(GREEN)))
dsh.conditional_formatting.add("H11", CellIsRule(operator="lessThan", formula=["MinGM"], fill=fill(RED)))

# C. trade & deductions
TR=16
section(dsh, TR, "C.  TRADE & DEDUCTION KPIs", span=12)
trade_kpis=[("Total Trade Spend $","=actTradeSpend",FMT_USD0),
            ("Trade % of Gross","=actTradePctG",FMT_PCT1),
            ("Trade % of Net","=actTradePctN",FMT_PCT1),
            ("Total GtN Deductions","=actGtN",FMT_USD0),
            ("Trade Efficiency (x)","=actTradeEff",FMT_X2),
            ("GtN Leakage % of Gross","=IF(actGross=0,0,actGtN/actGross)",FMT_PCT1)]
ci=0; rr=TR+1
for lab,frm,fmt in trade_kpis:
    col=["B","D","F","H","J","L"][ci]
    kpi_card("%s%d"%(col,rr),"%s%d"%(col,rr+1),lab,frm,fmt)
    ci+=1
dsh.conditional_formatting.add("D%d"%(rr+1), CellIsRule(operator="greaterThan", formula=["MaxTradePct"], fill=fill(RED), font=font(12,True,REDT)))
dsh.conditional_formatting.add("D%d"%(rr+1), CellIsRule(operator="lessThanOrEqual", formula=["MaxTradePct"], fill=fill(GREEN)))

# D. unit economics + E. break-even
UE=21
section(dsh, UE, "D.  UNIT ECONOMICS    |    E.  BREAK-EVEN", span=12)
ue=[("Net Realized Price /unit","=actNRPunit",FMT_USD2),
    ("Net Realized Price /case","=actNRPcase",FMT_USD2),
    ("Profit / Unit","=actPPU",FMT_USD2),
    ("Profit / Case","=actPPC",FMT_USD2),
    ("Break-Even Units","=beUnits",FMT_NUM0),
    ("BE Wholesale Price","=beWholesale",FMT_USD2)]
ci=0; rr=UE+1
for lab,frm,fmt in ue:
    col=["B","D","F","H","J","L"][ci]
    kpi_card("%s%d"%(col,rr),"%s%d"%(col,rr+1),lab,frm,fmt)
    ci+=1

# F. customer path economics
CP=25
section(dsh, CP, "F.  CUSTOMER PATH ECONOMICS", span=12)
cp=[("Retailer Margin %","=actRetMgnPct",FMT_PCT1),
    ("Retailer $ /unit","=actRetMgnU",FMT_USD2),
    ("Distributor Margin %","=actDistMgnPct",FMT_PCT1),
    ("Distributor $ /unit","=actDistMgnU",FMT_USD2),
    ("Company Net Real. /unit","=actNRPunit",FMT_USD2),
    ("Retail Shelf Price","=actRetailReq",FMT_USD2)]
ci=0; rr=CP+1
for lab,frm,fmt in cp:
    col=["B","D","F","H","J","L"][ci]
    kpi_card("%s%d"%(col,rr),"%s%d"%(col,rr+1),lab,frm,fmt)
    ci+=1

# G. scenario comparison (mini table)
SCMP=29
section(dsh, SCMP, "G.  SCENARIO COMPARISON", span=12)
W(dsh, "B%d"%(SCMP+1), "Metric", f=font(9, True, WHITE), fillc=STEEL, border=box, align=LEFT)
for i,s in enumerate(["Base","Active","Best NP","Worst NP","Δ vs Base"]):
    W(dsh, "%s%d"%(["D","E","F","G","H"][i],SCMP+1), s, f=font(9, True, WHITE), fillc=STEEL, border=box, align=CENTER)
cmp_rows=[("Net Profit $","np",FMT_USD0),("Net Margin %","nmpct",FMT_PCT1),
          ("Gross Margin %","gmpct",FMT_PCT1),("Trade % Gross","tradepct",FMT_PCT1)]
rr=SCMP+2
for lab,key,fmt in cmp_rows:
    W(dsh, "B%d"%rr, lab, f=font(10, DKTXT), fillc=LTBLUE, border=box, align=LEFT)
    W(dsh, "D%d"%rr, "=Scenarios!C%d"%SCN_ROWS[key], f=font(10), fillc=WHITE, border=box, fmt=fmt, align=RIGHT)
    W(dsh, "E%d"%rr, "="+active(key), f=font(10, True), fillc=WHITE, border=box, fmt=fmt, align=RIGHT)
    W(dsh, "F%d"%rr, "=MAX(Scenarios!C%d:I%d)"%(SCN_ROWS[key],SCN_ROWS[key]), f=font(10), fillc=WHITE, border=box, fmt=fmt, align=RIGHT)
    W(dsh, "G%d"%rr, "=MIN(Scenarios!C%d:I%d)"%(SCN_ROWS[key],SCN_ROWS[key]), f=font(10), fillc=WHITE, border=box, fmt=fmt, align=RIGHT)
    W(dsh, "H%d"%rr, "=E%d-D%d"%(rr,rr), f=font(10), fillc=WHITE, border=box, fmt=fmt, align=RIGHT)
    rr+=1

# H. gross-to-net waterfall (data for chart)
WF=rr+1
section(dsh, WF, "H.  GROSS-TO-NET WATERFALL  &  TRADE MIX", span=12)
wf_labels=[("Gross Sales","=actGross"),("On-Invoice","=-actOnInv"),("Off-Invoice","=-actOffInv"),
           ("Other GtN","=-actOtherDed"),("Net Sales","=actNet"),("Variable Cost","=-actVCost"),
           ("Contribution","=actContrib"),("Fixed","=-actFixed"),("Net Profit","=actNP")]
wstart=WF+1
for i,(lab,frm) in enumerate(wf_labels):
    W(dsh, "B%d"%(wstart+i), lab, f=font(9, DKTXT), fillc=WHITE, border=box, align=LEFT)
    W(dsh, "C%d"%(wstart+i), frm, f=font(9, True), fillc=WHITE, border=box, fmt=FMT_USD0, align=RIGHT)
wend=wstart+len(wf_labels)-1
# bar chart for waterfall steps
from openpyxl.chart import BarChart, Reference
bc = BarChart(); bc.type="col"; bc.title="Gross-to-Net Bridge"; bc.height=7; bc.width=14
data = Reference(dsh, min_col=3, min_row=wstart, max_row=wend)
cats = Reference(dsh, min_col=2, min_row=wstart, max_row=wend)
bc.add_data(data, titles_from_data=False); bc.set_categories(cats)
bc.legend=None
dsh.add_chart(bc, "E%d"%WF)

# scenario NP chart
bc2=BarChart(); bc2.type="col"; bc2.title="Net Profit by Scenario"; bc2.height=7; bc2.width=14
d2=Reference(scn, min_col=3, max_col=9, min_row=SCN_ROWS["np"], max_row=SCN_ROWS["np"])
c2=Reference(scn, min_col=3, max_col=9, min_row=SCN_HDR, max_row=SCN_HDR)
bc2.add_data(d2, from_rows=True, titles_from_data=False); bc2.set_categories(c2); bc2.legend=None
dsh.add_chart(bc2, "I%d"%WF)

dsh.column_dimensions["A"].width = 2
for c in "BCDEFGHIJKL":
    dsh.column_dimensions[c].width = 14
dsh.sheet_view.showGridLines = False
dsh.freeze_panes = "B5"

# =====================================================================
#  TAB 1: READ_ME_AND_DEFINITIONS  (created last, moved to front)
# =====================================================================
rm = wb.create_sheet("Read_Me_and_Definitions")
rm.sheet_properties.tabColor = "404040"
title_block(rm, "READ ME & DEFINITIONS",
            "Enterprise CPG Commercial Decision System — purpose, version control, metric definitions, and logic notes.")

rr=5
def block(label):
    global rr
    section(rm, rr, label, span=10); rr+=1
def line(a,b, boldlab=True):
    global rr
    W(rm, "B%d"%rr, a, f=font(10, boldlab, DKTXT), fillc=WHITE, border=box, align=LEFT)
    W(rm, "C%d"%rr, b, f=font(10), fillc=WHITE, border=box, align=LEFT, wrap=True)
    rm.merge_cells("C%d:J%d"%(rr,rr))
    rr+=1

block("A.  MODEL PURPOSE")
line("Purpose","Enterprise commercial operating tool to evaluate pricing, promotion, trade investment, distribution structure and customer profitability — and produce a GO / REVIEW / STOP decision.")
line("Decisions supported","Should we take this deal? Run this promo? What wholesale price hits target margin? What is true net realized price? How much incremental volume to break even? What if cost / fill / FX move?")
line("Primary users","CFO, GM, SVP Sales, RGM lead, Commercial Finance, Category & Customer teams.")
line("Owner","Commercial Finance Systems (model owner). Business owner: RGM lead.")
line("Model grain","One product × one customer/retailer/distributor path × one scenario × one market × one currency. Template is reusable across SKUs/customers.")

block("B.  VERSION CONTROL")
line("Model name","CPG Commercial Decision System")
line("Version","1.0")
line("Owner","Commercial Finance Systems")
line("Created","2026-06-26")
line("Last update","2026-06-26")
W(rm,"B%d"%rr,"Change log",f=font(10,True,DKTXT),fillc=WHITE,border=box,align=LEFT)
W(rm,"C%d"%rr,"v1.0 — initial build: GtN engine, scenario engine, promo ROI, FX, dashboard, audit.",f=font(10),fillc=WHITE,border=box,align=LEFT,wrap=True)
rm.merge_cells("C%d:J%d"%(rr,rr)); rr+=2

block("C.  METRIC DEFINITIONS")
W(rm,"B%d"%rr,"Metric",f=font(9,True,WHITE),fillc=STEEL,border=box,align=CENTER)
W(rm,"C%d"%rr,"Definition / formula",f=font(9,True,WHITE),fillc=STEEL,border=box,align=CENTER)
rm.merge_cells("C%d:J%d"%(rr,rr)); rr+=1
defs=[
 ("Gross Sales","Effective Units × Active Wholesale Price."),
 ("On-Invoice Deductions","Trade deductions taken on the invoice (per-unit trade elements)."),
 ("Off-Invoice Deductions","Trade taken after invoice: scan/bill backs, markdown, display, % & fixed trade."),
 ("Gross-to-Net Deductions","On-invoice + off-invoice trade + other financial deductions (accruals, EPD, customer deductions, chargebacks, claims)."),
 ("Net Sales","Gross Sales − Total Gross-to-Net Deductions."),
 ("Gross Profit / Margin %","Gross Sales − Total Variable Cost ; ÷ Gross Sales."),
 ("Contribution / Margin %","Net Sales − Total Variable Cost ; ÷ Net Sales."),
 ("Net Profit / Margin %","Contribution − Fixed Commercial Cost ; ÷ Net Sales."),
 ("Trade Spend $ / % Gross / % Net","Sum of trade-investment elements ; ÷ Gross ; ÷ Net."),
 ("Net Realized Price /unit /case","Net Sales ÷ Effective Units (or ÷ cases)."),
 ("Profit per Unit / Case","Net Profit ÷ Effective Units (or ÷ cases)."),
 ("Incremental Volume / Rev / Contrib / NP","Active scenario minus Base scenario for each line."),
 ("Promotion ROI","Incremental Net Profit ÷ Promo Spend (incremental trade)."),
 ("Break-Even Units / Cases","Fixed Spend ÷ Contribution per Unit (÷ units/case)."),
 ("Break-Even Wholesale / Retail Price","Price that drives Net Margin to target (algebraic solve); retail grossed up for distributor & retailer margin."),
 ("Break-Even Promo Lift %","Promo Spend ÷ (Contribution/unit × Base Volume)."),
 ("Retailer / Distributor Margin %","Channel mark-up margins implied by the selected business model."),
 ("Trade Efficiency","Net Profit ÷ Trade Spend ($ profit per $ trade)."),
]
for a,b in defs:
    W(rm,"B%d"%rr,a,f=font(10,True,DKTXT),fillc=WHITE,border=box,align=LEFT)
    W(rm,"C%d"%rr,b,f=font(10),fillc=WHITE,border=box,align=LEFT,wrap=True)
    rm.merge_cells("C%d:J%d"%(rr,rr)); rr+=1

block("D.  LOGIC NOTES")
notes=[
 ("Variable cost","Per-unit: COGS, packaging, co-man, inbound/outbound freight, duty, warehousing, compliance, returns/spoilage, other. Plus broker & distributor fees as % of gross."),
 ("Fixed commercial cost","Annualized account support, broker support, account mgmt, merchandising, listing amortization, promo development, other."),
 ("Distributor fees","Modeled as % of gross (variable) and, where relevant, in distributor margin grossing-up of shelf price."),
 ("Retailer margin","Shelf price = distributor sell ÷ (1 − retailer margin); distributor sell = wholesale ÷ (1 − distributor margin) when business model = Distributor."),
 ("Scenario overrides","Each scenario stores driver MULTIPLIERS (1.00 = base) for volume, price, var-cost, trade, fill, cannibalization, promo flag, FX. The Calc engine pulls the active column via INDEX/MATCH(selScenario)."),
 ("FX translation","Rates are USD value of 1 currency unit. value_in_reporting = value_base × fxBaseRate ÷ fxReportingRate. Edit rates on Currency_and_FX."),
 ("Break-even logic","Units: Fixed ÷ contribution/unit. Price: closed-form solution setting Net Margin = target (see Calc §8 helpers a,b,c)."),
 ("Recommendation logic","GO when GM≥min, NM≥min, contribution/unit≥min, trade%≤max, and (no promo OR ROI≥hurdle). STOP when net profit<0 or contribution/unit≤0. Otherwise REVIEW."),
 ("User vs calculated","User inputs = YELLOW cells (Inputs tab, scenario drivers, FX rates, promo spend/lift). Everything else is a locked formula."),
]
for a,b in notes:
    W(rm,"B%d"%rr,a,f=font(10,True,DKTXT),fillc=WHITE,border=box,align=LEFT)
    W(rm,"C%d"%rr,b,f=font(10),fillc=WHITE,border=box,align=LEFT,wrap=True)
    rm.merge_cells("C%d:J%d"%(rr,rr)); rr+=1

block("E.  HOW TO USE  (60-second guide)")
steps=[
 "1. Open Inputs_and_Control_Center. Set the Active Scenario, business model and currencies in the control bar.",
 "2. Enter pricing, cost, trade, volume and fixed-cost assumptions (yellow cells only).",
 "3. Read the Live KPI Snapshot and DECISION STATUS at the bottom of Inputs.",
 "4. Open Dashboard_Executive_View for the one-screen readout and waterfall.",
 "5. Use Scenarios to compare all 7 cases; use Promotions_and_ROI to test tactics.",
 "6. ALWAYS confirm Checks_Controls_and_Audit shows MODEL HEALTHY before deciding.",
]
for s in steps:
    W(rm,"B%d"%rr,s,f=font(10),fillc=WHITE,border=box,align=LEFT,wrap=True)
    rm.merge_cells("B%d:J%d"%(rr,rr)); rr+=1

rm.column_dimensions["A"].width=2
rm.column_dimensions["B"].width=30
for c in "CDEFGHIJ":
    rm.column_dimensions[c].width=13
rm.sheet_view.showGridLines=False
rm.freeze_panes="B5"

# =====================================================================
#  ORDER TABS, PROTECTION, VALIDATION
# =====================================================================
order = ["Read_Me_and_Definitions","Inputs_and_Control_Center","Calculations_Gross_to_Net",
         "Scenarios","Promotions_and_ROI","Currency_and_FX","Dashboard_Executive_View",
         "Checks_Controls_and_Audit","Lookups_and_Assumptions"]
wb._sheets.sort(key=lambda ws: order.index(ws.title))

# data validations
def add_dv(ws, cells, listref):
    dv = DataValidation(type="list", formula1="="+listref, allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    for cell in cells:
        dv.add(ws[cell])

add_dv(inp, ["C7"], "scnList")
add_dv(inp, ["C8"], "bizList")
add_dv(inp, ["C9","C10"], "ccyList")

# sheet protection: lock formulas, keep yellow inputs editable (locked=False already set)
for ws in [rm, inp, cal, scn, pro, fxs, dsh, chk, look]:
    ws.protection.sheet = True
    ws.protection.password = "cpg"
    ws.protection.formatCells = False
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True
    ws.protection.enableFormatColumns = True
    ws.protection.enableFormatRows = True

wb.properties.title = "CPG Commercial Decision System"
wb.properties.creator = "Commercial Finance Systems"
wb.properties.company = "Enterprise CPG"

import os
out = os.path.join(os.path.dirname(__file__), "..", "CPG_Commercial_Decision_System.xlsx")
out = os.path.abspath(out)
wb.save(out)
print("SAVED:", out)
print("Defined names:", len(NAME))
