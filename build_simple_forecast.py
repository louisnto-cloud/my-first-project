"""Organika RTD FY26-27 Forecast - Simple Edition (v5).

Same calculation engine as the enterprise build, but only 10 tabs are
visible. Engine/governance internals are hidden but still working.

Visible tabs (in order):
  1.  Home           - the one-pager
  2.  Controls       - the 5 knobs
  3.  Doors          - 12 monthly inputs per row -> 52 weekly outputs
  4.  Pricing        - net price + landed cost per SKU
  5.  Marketing      - sampling, events, rep samples, buffer
  6.  Trade Spend    - channel listing + slotting + scan/promo
  7.  OPEX           - reps, A&P, freight, G&A
  8.  P&L            - Revenue -> Trade -> COGS -> A&P -> SG&A -> EBITDA
  9.  Scenarios      - Conservative / Base / Stretch
 10.  Dashboard      - monthly + quarterly roll-ups by SKU/channel/province
 11.  Risks          - top risks scored by likelihood x impact

Hidden (still functional):
  Forecast Weekly, Production Plan, Revenue (engine)
  Assumptions, Class Mix, Velocity (advanced inputs)
  Validation Lists, Change Log (governance internals)
"""

from datetime import date, timedelta

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

# ---------- Domain ----------
SKUS = ["MUV Raspberry Lemon", "MUV Lime", "MUV PFP", "LCA Energy"]
SKU_BRAND = {s: ("LCA" if "LCA" in s else "MUV") for s in SKUS}
PROVINCES = ["BC", "AB", "SK", "MB", "ON", "QC", "NB", "NS", "PE", "NL"]
CHANNELS = ["Natural", "Specialty", "FDM", "Costco",
            "On-Premise", "Convenience", "Gym & Fitness", "Private Liquor", "RAS"]
NEW_CHANNELS = {"On-Premise", "Convenience", "Gym & Fitness", "Private Liquor", "RAS"}

START_DATE = date(2026, 9, 1)
N_WEEKS = 52
WEEK_DATES = [START_DATE + timedelta(days=7 * i) for i in range(N_WEEKS)]
WEEK_LABELS = [f"W{i + 1:02d} ({d.strftime('%b %d')})" for i, d in enumerate(WEEK_DATES)]
MONTH_ORDER = ["Sep 2026", "Oct 2026", "Nov 2026", "Dec 2026", "Jan 2027",
               "Feb 2027", "Mar 2027", "Apr 2027", "May 2027", "Jun 2027",
               "Jul 2027", "Aug 2027"]
WEEK_TO_MONTH = [date(d.year, d.month, 1).strftime("%b %Y") for d in WEEK_DATES]
QUARTERS = ["Q1 (Sep-Nov)", "Q2 (Dec-Feb)", "Q3 (Mar-May)", "Q4 (Jun-Aug)"]
MONTH_Q = {m: q for m, q in zip(MONTH_ORDER,
            sum([[q]*3 for q in QUARTERS], []))}

FW_FIRST = 5; FW_LAST = FW_FIRST + N_WEEKS - 1
def fw_col(i): return get_column_letter(FW_FIRST + i - 1)

DOORS_M_FIRST = 5; DOORS_M_LAST = 16  # E..P
DOORS_W_FIRST = 17                     # Q
DOORS_W_LAST = DOORS_W_FIRST + N_WEEKS - 1
MONTH_COL = {m: get_column_letter(DOORS_M_FIRST + i) for i, m in enumerate(MONTH_ORDER)}
def doors_w_col(i): return get_column_letter(DOORS_W_FIRST + i - 1)

REP_M_FIRST = 2; REP_W_FIRST = 14
REP_MONTH_COL = {m: get_column_letter(REP_M_FIRST + i) for i, m in enumerate(MONTH_ORDER)}

# ---------- Style ----------
INK = "1D1D1F"; GREY = "6E6E73"; SUBTLE = "F5F5F7"
BLUE = "0071E3"; GREEN = "34C759"; ORANGE = "FF9500"; RED = "FF3B30"

H_FONT = Font(name="Helvetica", bold=True, color="FFFFFF", size=11)
H_FILL = PatternFill("solid", fgColor=INK)
TITLE = Font(name="Helvetica", bold=True, size=22, color=INK)
HERO_LABEL = Font(name="Helvetica", size=10, color=GREY)
HERO = Font(name="Helvetica", bold=True, size=32, color=INK)
SECTION = Font(name="Helvetica", bold=True, size=13, color=INK)
BODY = Font(name="Helvetica", size=11, color=INK)
MUTED = Font(name="Helvetica", size=10, color=GREY)
SMALL_LABEL = Font(name="Helvetica", size=10, color=GREY)
BIG_NUM = Font(name="Helvetica", bold=True, size=18, color=INK)
INP_FONT = Font(name="Helvetica", color=BLUE, size=11)
LINK_FONT = Font(name="Helvetica", color=GREEN, size=11)
BOLD = Font(name="Helvetica", bold=True, size=11, color=INK)
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREEN_F = PatternFill("solid", fgColor="D1F2D1")
RED_F = PatternFill("solid", fgColor="FFD6D6")
SUBT_FILL = PatternFill("solid", fgColor=SUBTLE)
MO_FILL = PatternFill("solid", fgColor="FFF8E5")

def head_row(ws, row, cols):
    for j, val in enumerate(cols, start=1):
        c = ws.cell(row=row, column=j, value=val)
        c.font = H_FONT; c.fill = H_FILL
        c.alignment = Alignment(horizontal="center")

def setup_print(ws, landscape=True, header_rows=0):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

# ---------- README hidden (info baked into Home) ----------

def build_assumptions(ws):
    ws.title = "Assumptions"
    ws["A1"] = "Assumptions"; ws["A1"].font = TITLE
    ws["A3"] = "Units per case"; ws["B3"] = 24; ws["B3"].font = INP_FONT
    ws["A4"] = "Production buffer"; ws["B4"] = 0.10; ws["B4"].font = INP_FONT; ws["B4"].number_format = "0%"
    ws["A5"] = "Scenario multiplier"; ws["B5"] = 1.0; ws["B5"].font = INP_FONT; ws["B5"].fill = YELLOW
    # Backwards compatibility: B4/B5/B6 layout used by engine formulas
    # Engine refs: B4=units, B5=buffer, B6=multiplier
    ws["A6"] = "(engine alias)"; ws["B6"] = "=B5"
    for c, w in zip("AB", (24, 14)):
        ws.column_dimensions[c].width = w

def build_class_mix(ws):
    ws.title = "Class Mix"
    ws["A1"] = "Class Mix"; ws["A1"].font = TITLE
    head_row(ws, 3, ["Channel", "Class A %", "Class B %", "Class C %", "Sum"])
    rows = [("Natural", 0.20, 0.50, 0.30), ("Specialty", 0.30, 0.50, 0.20),
            ("FDM", 0.50, 0.35, 0.15), ("Costco", 0.80, 0.20, 0.00),
            ("On-Premise", 0.25, 0.50, 0.25), ("Convenience", 0.15, 0.50, 0.35),
            ("Gym & Fitness", 0.30, 0.50, 0.20), ("Private Liquor", 0.25, 0.55, 0.20),
            ("RAS", 0.10, 0.40, 0.50)]
    for i, (ch, a, b, c) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=ch).font = BODY
        for col, v in zip((2, 3, 4), (a, b, c)):
            cell = ws.cell(row=i, column=col, value=v)
            cell.font = INP_FONT; cell.number_format = "0%"
            if ch in NEW_CHANNELS: cell.fill = YELLOW
        ws.cell(row=i, column=5, value=f"=B{i}+C{i}+D{i}").number_format = "0%"
    ws.conditional_formatting.add(f"E4:E{3+len(rows)}",
        FormulaRule(formula=[f"E4=1"], fill=GREEN_F))
    ws.conditional_formatting.add(f"E4:E{3+len(rows)}",
        FormulaRule(formula=[f"E4<>1"], fill=RED_F))
    for c, w in zip("ABCDE", (20, 12, 12, 12, 10)):
        ws.column_dimensions[c].width = w

def build_velocity(ws):
    ws.title = "Velocity"
    ws["A1"] = "Velocity (units/door/week)"; ws["A1"].font = TITLE
    head_row(ws, 3, ["SKU", "Channel", "Vel A", "Vel B", "Vel C",
                     "A %", "B %", "C %", "Blended"])
    muv = {"Natural": (12, 8, 5), "Specialty": (10, 7, 4), "FDM": (20, 12, 6),
           "Costco": (200, 120, 60), "On-Premise": (6, 4, 2), "Convenience": (8, 5, 3),
           "Gym & Fitness": (10, 7, 4), "Private Liquor": (5, 3, 2), "RAS": (4, 3, 2)}
    lime_o = {"Natural": (10, 7, 4), "Specialty": (9, 6, 4), "FDM": (18, 11, 5), "Costco": (180, 110, 55)}
    pfp_o = {"Natural": (14, 9, 5), "Specialty": (11, 7, 4), "FDM": (22, 13, 6), "Costco": (220, 130, 65)}
    lca = {"Natural": (8, 5, 3), "Specialty": (8, 5, 3), "FDM": (15, 10, 5), "Costco": (180, 100, 50),
           "On-Premise": (8, 5, 3), "Convenience": (12, 8, 4), "Gym & Fitness": (12, 8, 5),
           "Private Liquor": (3, 2, 1), "RAS": (3, 2, 1)}
    cm = "'Class Mix'!$A$4:$D$200"
    row = 4
    for sku in SKUS:
        for ch in CHANNELS:
            if sku == "MUV Raspberry Lemon": a, b, c = muv[ch]
            elif sku == "MUV Lime": a, b, c = lime_o.get(ch, muv[ch])
            elif sku == "MUV PFP": a, b, c = pfp_o.get(ch, muv[ch])
            else: a, b, c = lca[ch]
            ws.cell(row=row, column=1, value=sku).font = BODY
            ws.cell(row=row, column=2, value=ch).font = BODY
            for col, v in zip((3, 4, 5), (a, b, c)):
                cell = ws.cell(row=row, column=col, value=v); cell.font = INP_FONT
                if ch in NEW_CHANNELS: cell.fill = YELLOW
            for col, mi in zip((6, 7, 8), (2, 3, 4)):
                cc = ws.cell(row=row, column=col, value=f"=IFERROR(VLOOKUP(B{row},{cm},{mi},FALSE),0)")
                cc.font = LINK_FONT; cc.number_format = "0%"
            ws.cell(row=row, column=9, value=f"=C{row}*F{row}+D{row}*G{row}+E{row}*H{row}").number_format = "0.00"
            row += 1
    for c, w in zip("ABCDEFGHI", (22, 16, 8, 8, 8, 10, 10, 10, 10)):
        ws.column_dimensions[c].width = w

def build_pricing(ws):
    ws.title = "Pricing"
    ws["A1"] = "Pricing"; ws["A1"].font = TITLE
    ws["A2"] = "Net price = invoice. Landed cost = ingredients + co-man + freight in."
    ws["A2"].font = MUTED
    head_row(ws, 4, ["SKU", "Net price/case", "Landed cost/case", "GP/case", "GP %", "Notes"])
    rows = [("MUV Raspberry Lemon", 28, 14.5, "Confirm with Rijo."),
            ("MUV Lime", 28, 14.5, "Same as Raspberry Lemon."),
            ("MUV PFP", 28, 14.5, "Same as other MUV SKUs."),
            ("LCA Energy", 32, 16.0, "Higher price point. Cost is placeholder.")]
    for i, (sku, p, c, n) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=sku).font = BODY
        ws.cell(row=i, column=2, value=p).font = INP_FONT
        ws.cell(row=i, column=3, value=c).font = INP_FONT
        ws.cell(row=i, column=4, value=f"=B{i}-C{i}")
        gp = ws.cell(row=i, column=5, value=f"=D{i}/B{i}"); gp.number_format = "0%"
        ws.cell(row=i, column=6, value=n).font = MUTED
        for k in (2, 3, 4):
            ws.cell(row=i, column=k).number_format = "$#,##0.00"
    ws.conditional_formatting.add("E5:E8",
        CellIsRule(operator="lessThan", formula=["0.40"], fill=RED_F))
    for c, w in zip("ABCDEF", (22, 16, 18, 12, 8, 50)):
        ws.column_dimensions[c].width = w
    setup_print(ws, landscape=False)

def build_doors(ws):
    ws.title = "Doors"
    ws["A1"] = "Active doors"; ws["A1"].font = TITLE
    ws["A2"] = "Edit the 12 monthly columns (yellow tint). The 52 weekly columns auto-fill."
    ws["A2"].font = MUTED

    for j, h in enumerate(["Brand", "Province", "Channel", "FY total"], start=1):
        c = ws.cell(row=4, column=j, value=h); c.font = H_FONT; c.fill = H_FILL
    for j, m in enumerate(MONTH_ORDER):
        c = ws.cell(row=4, column=DOORS_M_FIRST + j, value=m); c.font = H_FONT; c.fill = H_FILL
    for j, lb in enumerate(WEEK_LABELS):
        c = ws.cell(row=4, column=DOORS_W_FIRST + j, value=lb); c.font = H_FONT; c.fill = H_FILL

    # Baseline door plan (from prior versions, faithfully)
    plan = {}
    def sm(b, p, ch, vs): plan[(b, p, ch)] = dict(zip(MONTH_ORDER, vs))

    # MUV existing
    sm("MUV","BC","Natural",[80,95,110,125,140,155,170,180,190,195,200,200])
    sm("MUV","BC","Specialty",[10,15,20,25,30,35,42,48,52,56,58,60])
    sm("MUV","BC","FDM",[0,0,8,15,20,25,28,30,33,36,38,40])
    sm("MUV","BC","Costco",[0,0,0,0,0,0,60,0,0,25,25,25])
    sm("MUV","AB","Natural",[5,10,18,28,38,48,58,68,75,82,87,90])
    sm("MUV","AB","Specialty",[3,5,8,12,16,20,24,27,30,32,33,35])
    sm("MUV","AB","FDM",[0,0,0,5,10,15,18,22,25,27,28,30])
    sm("MUV","AB","Costco",[0,0,0,0,0,0,40,0,0,20,22,22])
    sm("MUV","SK","Natural",[0,0,2,5,8,12,15,18,20,22,24,25])
    sm("MUV","SK","Specialty",[0,0,1,2,4,6,8,10,12,13,14,15])
    sm("MUV","SK","FDM",[0,0,0,0,2,5,7,10,12,13,14,15])
    sm("MUV","SK","Costco",[0,0,0,0,0,0,0,0,0,5,6,7])
    sm("MUV","MB","Natural",[0,0,2,4,7,10,13,15,17,19,21,22])
    sm("MUV","MB","Specialty",[0,0,1,2,3,5,7,8,10,11,12,13])
    sm("MUV","MB","FDM",[0,0,0,0,2,4,6,8,10,12,13,14])
    sm("MUV","MB","Costco",[0,0,0,0,0,0,0,0,0,4,5,6])
    sm("MUV","ON","Natural",[10,18,28,40,55,75,95,110,125,135,145,150])
    sm("MUV","ON","Specialty",[5,8,12,17,23,28,33,38,42,45,48,50])
    sm("MUV","ON","FDM",[0,0,5,12,20,28,35,42,48,53,57,60])
    sm("MUV","ON","Costco",[0,0,0,0,0,0,0,0,0,30,32,35])
    sm("MUV","QC","Natural",[20,25,30,35,40,45,50,55,60,65,70,75])
    sm("MUV","QC","Specialty",[5,8,11,14,17,20,23,25,28,30,32,32])
    sm("MUV","QC","FDM",[0,0,0,5,8,12,15,18,20,22,23,25])
    sm("MUV","QC","Costco",[0,0,0,0,0,0,0,0,0,20,22,22])
    sm("MUV","NB","Natural",[0,0,1,3,5,7,9,11,13,14,15,16])
    sm("MUV","NB","Specialty",[0,0,0,1,2,4,5,6,7,8,9,10])
    sm("MUV","NB","FDM",[0,0,0,0,0,2,3,5,7,8,9,10])
    sm("MUV","NB","Costco",[0,0,0,0,0,0,0,0,0,3,4,5])
    sm("MUV","NS","Natural",[0,0,1,3,5,7,9,11,13,14,15,16])
    sm("MUV","NS","Specialty",[0,0,0,1,2,4,5,6,7,8,9,10])
    sm("MUV","NS","FDM",[0,0,0,0,0,2,3,5,7,8,9,10])
    sm("MUV","NS","Costco",[0,0,0,0,0,0,0,0,0,3,4,5])
    sm("MUV","PE","Natural",[0,0,0,1,2,3,4,5,6,6,7,7])
    sm("MUV","PE","Specialty",[0,0,0,0,1,2,2,3,3,4,4,4])
    sm("MUV","PE","FDM",[0,0,0,0,0,0,1,2,2,3,3,3])
    sm("MUV","PE","Costco",[0]*12)
    sm("MUV","NL","Natural",[0,0,0,1,2,3,4,5,6,6,7,7])
    sm("MUV","NL","Specialty",[0,0,0,0,1,2,2,3,3,4,4,4])
    sm("MUV","NL","FDM",[0,0,0,0,0,0,1,2,2,3,3,3])
    sm("MUV","NL","Costco",[0]*12)

    # New channels
    new = {
      "On-Premise": {"BC":[30,40,55,70,80,90,100,105,110,115,118,120],"AB":[20,28,38,48,56,62,68,72,75,78,80,82],
                     "SK":[5,8,12,15,18,21,24,26,28,29,30,30],"MB":[5,8,12,15,18,21,24,26,28,29,30,30],
                     "ON":[40,55,75,90,105,120,130,140,145,150,152,155],"QC":[30,45,60,75,85,95,105,110,115,118,120,122],
                     "NB":[3,5,7,9,10,11,12,13,14,14,15,15],"NS":[3,5,7,9,10,11,12,13,14,14,15,15],
                     "PE":[1,1,2,2,3,3,4,4,4,5,5,5],"NL":[1,1,2,2,3,3,4,4,4,5,5,5]},
      "Convenience": {"BC":[20,30,45,60,70,80,88,92,95,98,100,100],"AB":[15,22,32,45,55,62,68,72,75,78,80,80],
                      "SK":[5,8,12,16,20,23,25,27,28,29,30,30],"MB":[5,8,12,16,20,23,25,27,28,29,30,30],
                      "ON":[50,75,105,130,150,170,180,190,195,198,200,200],"QC":[30,50,70,90,105,115,122,126,128,130,130,130],
                      "NB":[3,5,7,9,11,12,13,14,14,15,15,15],"NS":[3,5,7,9,11,12,13,14,14,15,15,15],
                      "PE":[1,2,2,3,3,4,4,5,5,5,5,5],"NL":[1,2,2,3,3,4,4,5,5,5,5,5]},
      "Gym & Fitness": {"BC":[15,22,30,38,44,50,54,56,58,59,60,60],"AB":[10,16,24,32,38,42,45,47,49,50,50,50],
                        "SK":[3,5,7,9,11,12,13,14,14,15,15,15],"MB":[3,5,7,9,11,12,13,14,14,15,15,15],
                        "ON":[25,38,55,70,82,90,95,98,99,100,100,100],"QC":[15,22,32,42,50,55,58,59,60,60,60,60],
                        "NB":[2,3,4,5,6,7,7,8,8,8,8,8],"NS":[2,3,4,5,6,7,7,8,8,8,8,8],
                        "PE":[0,0,1,1,2,2,2,3,3,3,3,3],"NL":[0,0,1,1,2,2,2,3,3,3,3,3]},
      "Private Liquor": {"BC":[10,15,22,28,34,40,44,47,49,50,50,50],"AB":[8,12,18,24,30,34,37,39,40,40,40,40],
                         "SK":[0,0,2,4,6,8,10,12,14,15,15,15],"MB":[0,0,2,4,6,8,10,11,12,13,14,15],
                         "ON":[0]*12,"QC":[0]*12,"NB":[0]*12,"NS":[0]*12,"PE":[0]*12,"NL":[0]*12},
      "RAS": {"BC":[0]*12,"AB":[0]*12,"SK":[0]*12,"MB":[0]*12,
              "ON":[5,8,12,16,20,24,26,28,29,30,30,30],"QC":[0]*12,
              "NB":[2,3,5,6,7,8,9,10,10,10,10,10],"NS":[2,3,5,6,7,8,9,10,10,10,10,10],
              "PE":[0]*12,"NL":[0]*12},
    }
    for ch, pm in new.items():
        for p, vs in pm.items():
            sm("MUV", p, ch, vs)

    # LCA (BC ramps; others zero)
    sm("LCA","BC","Natural",[0,0,0,0,5,10,15,20,25,28,30,32])
    sm("LCA","BC","Specialty",[0,0,0,0,2,5,8,12,15,17,18,20])
    sm("LCA","BC","On-Premise",[0,0,0,0,3,6,10,14,18,20,22,24])
    sm("LCA","BC","Convenience",[0,0,0,0,5,10,15,20,25,28,30,32])
    sm("LCA","BC","Gym & Fitness",[0,0,0,0,3,6,10,13,16,18,20,22])
    for p in PROVINCES:
        for ch in CHANNELS:
            if ("LCA", p, ch) not in plan:
                sm("LCA", p, ch, [0]*12)

    row = 5
    for b in ("MUV", "LCA"):
        for p in PROVINCES:
            for ch in CHANNELS:
                vs = plan[(b, p, ch)]
                ws.cell(row=row, column=1, value=b).font = BODY
                ws.cell(row=row, column=2, value=p).font = BODY
                ws.cell(row=row, column=3, value=ch).font = BODY
                first_w = doors_w_col(1); last_w = doors_w_col(N_WEEKS)
                ws.cell(row=row, column=4, value=f"=SUM({first_w}{row}:{last_w}{row})")
                for j, m in enumerate(MONTH_ORDER):
                    cell = ws.cell(row=row, column=DOORS_M_FIRST + j, value=vs[m])
                    cell.font = INP_FONT
                    cell.fill = YELLOW if ch in NEW_CHANNELS else MO_FILL
                for wi in range(1, N_WEEKS + 1):
                    mc = MONTH_COL[WEEK_TO_MONTH[wi - 1]]
                    cell = ws.cell(row=row, column=DOORS_W_FIRST + wi - 1, value=f"=${mc}{row}")
                    cell.font = LINK_FONT
                row += 1
    total_row = row
    ws.cell(row=total_row, column=1, value="Total").font = BOLD
    ws.cell(row=total_row, column=4, value=f"=SUM(D5:D{total_row-1})").font = BOLD
    for j in range(12):
        cl = get_column_letter(DOORS_M_FIRST + j)
        cell = ws.cell(row=total_row, column=DOORS_M_FIRST + j,
                       value=f"=SUM({cl}5:{cl}{total_row-1})")
        cell.font = BOLD; cell.fill = SUBT_FILL
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(DOORS_W_FIRST + wi - 1)
        cell = ws.cell(row=total_row, column=DOORS_W_FIRST + wi - 1,
                       value=f"=SUM({cl}5:{cl}{total_row-1})")
        cell.font = BOLD; cell.fill = SUBT_FILL

    for c, w in zip("ABCD", (8, 8, 16, 10)):
        ws.column_dimensions[c].width = w
    for j in range(12):
        ws.column_dimensions[get_column_letter(DOORS_M_FIRST + j)].width = 11
    for wi in range(1, N_WEEKS + 1):
        ws.column_dimensions[get_column_letter(DOORS_W_FIRST + wi - 1)].width = 11
    ws.freeze_panes = "E5"
    return total_row

def build_marketing(ws):
    ws.title = "Marketing"
    ws["A1"] = "Marketing & sampling (cases/week)"; ws["A1"].font = TITLE
    ws["A2"] = "Account sampling rate per channel + events + rep samples + buffer."
    ws["A2"].font = MUTED

    ws.cell(row=4, column=1, value="Summary").font = H_FONT
    ws.cell(row=4, column=1).fill = H_FILL
    for j, lb in enumerate(WEEK_LABELS):
        c = ws.cell(row=4, column=FW_FIRST + j, value=lb); c.font = H_FONT; c.fill = H_FILL
    for i, lb in enumerate(["Account sampling", "Events", "Rep samples", "Buffer", "Total"], start=5):
        ws.cell(row=i, column=1, value=lb).font = BOLD

    # Section 1
    sec1 = 12
    ws.cell(row=sec1, column=1, value="Account sampling").font = SECTION
    ws.cell(row=sec1 + 1, column=1, value="Channel").font = H_FONT
    ws.cell(row=sec1 + 1, column=1).fill = H_FILL
    ws.cell(row=sec1 + 1, column=2, value="Cases/door/week").font = H_FONT
    ws.cell(row=sec1 + 1, column=2).fill = H_FILL
    for j, lb in enumerate(WEEK_LABELS):
        c = ws.cell(row=sec1 + 1, column=FW_FIRST + j, value=lb); c.font = H_FONT; c.fill = H_FILL
    rates = {"Natural": 0.025, "Specialty": 0.020, "FDM": 0.0125, "Costco": 0.050,
             "On-Premise": 0.040, "Convenience": 0.005, "Gym & Fitness": 0.050,
             "Private Liquor": 0.010, "RAS": 0.010}
    sec1_first = sec1 + 2
    for i, ch in enumerate(CHANNELS):
        r = sec1_first + i
        ws.cell(row=r, column=1, value=ch).font = BODY
        rate = ws.cell(row=r, column=2, value=rates[ch])
        rate.font = INP_FONT; rate.number_format = "0.0000"
        if ch in NEW_CHANNELS: rate.fill = YELLOW
        for wi in range(1, N_WEEKS + 1):
            col = FW_FIRST + wi - 1
            dwc = doors_w_col(wi)
            ws.cell(row=r, column=col,
                value=f"=$B{r}*SUMIFS(Doors!{dwc}:{dwc},Doors!$C:$C,$A{r})").font = LINK_FONT
    sec1_last = sec1_first + len(CHANNELS) - 1
    sec1_total = sec1_last + 1
    ws.cell(row=sec1_total, column=1, value="Sampling total").font = BOLD
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(FW_FIRST + wi - 1)
        c = ws.cell(row=sec1_total, column=FW_FIRST + wi - 1,
                    value=f"=SUM({cl}{sec1_first}:{cl}{sec1_last})")
        c.font = BOLD; c.fill = SUBT_FILL

    # Section 2 events
    sec2 = sec1_total + 2
    ws.cell(row=sec2, column=1, value="Events").font = SECTION
    sub = sec2 + 1
    for col, lb in zip("ABCD", ["Event", "Cases", "Start week", "End week"]):
        c = ws.cell(row=sub, column="ABCD".index(col) + 1, value=lb)
        c.font = H_FONT; c.fill = H_FILL
    for j, lb in enumerate(WEEK_LABELS):
        c = ws.cell(row=sub, column=FW_FIRST + j, value=lb); c.font = H_FONT; c.fill = H_FILL

    def wfirst(m):
        for i, mm in enumerate(WEEK_TO_MONTH, start=1):
            if mm == m: return i
        return 1
    def wlast(m):
        last = 1
        for i, mm in enumerate(WEEK_TO_MONTH, start=1):
            if mm == m: last = i
        return last

    events = [
        ("CHFA West (Vancouver)", 25, wfirst("Sep 2026"), wfirst("Sep 2026")),
        ("Indie Alley demo (recurring)", 36, 1, N_WEEKS),
        ("Marche Tao launch (QC)", 15, wfirst("Sep 2026"), wfirst("Sep 2026")),
        ("Costco BC roadshow", 60, wfirst("Mar 2027"), wfirst("Mar 2027")),
        ("Costco AB roadshow", 40, wfirst("Mar 2027"), wfirst("Mar 2027")),
        ("CHFA East (Toronto)", 25, wfirst("Apr 2027"), wfirst("Apr 2027")),
        ("BCAA Wellness Tour", 30, wfirst("Apr 2027"), wlast("May 2027")),
        ("Toronto Vegfest", 20, wfirst("Jun 2027") + 2, wfirst("Jun 2027") + 2),
        ("Calgary Stampede", 25, wfirst("Jul 2027") + 1, wfirst("Jul 2027") + 1),
        ("Yoga/fitness partners (recurring)", 36, 1, N_WEEKS),
        ("Influencer seeding (recurring)", 48, 1, N_WEEKS),
        ("Restaurant launch dinners", 20, wfirst("Oct 2026"), wlast("May 2027")),
        ("Convenience sales blitz (recurring)", 24, 1, N_WEEKS),
    ]
    sec2_first = sub + 1
    for i, (lb, total, s, e) in enumerate(events):
        r = sec2_first + i
        ws.cell(row=r, column=1, value=lb).font = BODY
        for col, v in zip((2, 3, 4), (total, s, e)):
            ws.cell(row=r, column=col, value=v).font = INP_FONT
        for wi in range(1, N_WEEKS + 1):
            col = FW_FIRST + wi - 1
            ws.cell(row=r, column=col,
                value=f"=IF(AND({wi}>=$C{r},{wi}<=$D{r}),$B{r}/($D{r}-$C{r}+1),0)").font = LINK_FONT
    sec2_last = sec2_first + len(events) - 1
    sec2_total = sec2_last + 1
    ws.cell(row=sec2_total, column=1, value="Events total").font = BOLD
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(FW_FIRST + wi - 1)
        c = ws.cell(row=sec2_total, column=FW_FIRST + wi - 1,
                    value=f"=SUM({cl}{sec2_first}:{cl}{sec2_last})")
        c.font = BOLD; c.fill = SUBT_FILL

    # Section 3 rep samples
    sec3 = sec2_total + 2
    ws.cell(row=sec3, column=1, value="Rep samples").font = SECTION
    rate_row = sec3 + 1
    ws.cell(row=rate_row, column=1, value="Cases/rep/week").font = BODY
    ws.cell(row=rate_row, column=2, value=1.0).font = INP_FONT
    sub3 = sec3 + 2
    ws.cell(row=sub3, column=1, value="Province").font = H_FONT
    ws.cell(row=sub3, column=1).fill = H_FILL
    for j, m in enumerate(MONTH_ORDER):
        c = ws.cell(row=sub3, column=REP_M_FIRST + j, value=m); c.font = H_FONT; c.fill = H_FILL
    for j, lb in enumerate(WEEK_LABELS):
        c = ws.cell(row=sub3, column=REP_W_FIRST + j, value=lb); c.font = H_FONT; c.fill = H_FILL
    rep_plan = {"BC":[2,2,2,2,2,3,3,3,3,3,3,3], "AB":[1,1,1,1,1,1,2,2,2,2,2,2],
                "SK":[0,0,0,0,0,1,1,1,1,1,1,1], "MB":[0,0,0,0,0,1,1,1,1,1,1,1],
                "ON":[1,1,1,1,1,1,2,2,2,2,2,2], "QC":[1,1,1,1,1,1,1,1,2,2,2,2],
                "NB":[0,0,0,0,0,0,1,1,1,1,1,1], "NS":[0,0,0,0,0,0,1,1,1,1,1,1],
                "PE":[0]*12, "NL":[0]*12}
    sec3_first = sub3 + 1
    for i, p in enumerate(PROVINCES):
        r = sec3_first + i
        ws.cell(row=r, column=1, value=p).font = BODY
        for j, m in enumerate(MONTH_ORDER):
            cell = ws.cell(row=r, column=REP_M_FIRST + j, value=rep_plan[p][j])
            cell.font = INP_FONT; cell.fill = MO_FILL
        for wi in range(1, N_WEEKS + 1):
            mc = REP_MONTH_COL[WEEK_TO_MONTH[wi - 1]]
            ws.cell(row=r, column=REP_W_FIRST + wi - 1, value=f"=${mc}{r}").font = LINK_FONT
    total_reps = sec3_first + len(PROVINCES)
    sample_row = total_reps + 1
    ws.cell(row=total_reps, column=1, value="Total reps").font = BOLD
    ws.cell(row=sample_row, column=1, value="Rep samples total").font = BOLD
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(REP_W_FIRST + wi - 1)
        ws.cell(row=total_reps, column=REP_W_FIRST + wi - 1,
                value=f"=SUM({cl}{sec3_first}:{cl}{sec3_first + len(PROVINCES) - 1})").font = BOLD
        c = ws.cell(row=sample_row, column=REP_W_FIRST + wi - 1,
                    value=f"=$B${rate_row}*{cl}{total_reps}")
        c.font = BOLD; c.fill = SUBT_FILL
    # Project rep samples to standard FW columns
    link_row = sample_row + 1
    ws.cell(row=link_row, column=1, value="(linked)").font = MUTED
    for wi in range(1, N_WEEKS + 1):
        src = get_column_letter(REP_W_FIRST + wi - 1)
        ws.cell(row=link_row, column=FW_FIRST + wi - 1, value=f"={src}{sample_row}").font = LINK_FONT

    # Section 4 buffer
    sec4 = link_row + 2
    ws.cell(row=sec4, column=1, value="Buffer").font = SECTION
    sub4 = sec4 + 1
    for col, lb in zip("ABC", ["Item", "Cases/week", "Notes"]):
        c = ws.cell(row=sub4, column="ABC".index(col) + 1, value=lb); c.font = H_FONT; c.fill = H_FILL
    for j, lb in enumerate(WEEK_LABELS):
        c = ws.cell(row=sub4, column=FW_FIRST + j, value=lb); c.font = H_FONT; c.fill = H_FILL
    buf = [("Gifting", 0.5, "Aaron + Teresa allocation"),
           ("R&D samples", 0.75, "Formulation tests"),
           ("Photo & content", 0.5, "Brand shoots"),
           ("Freight returns", 0.25, "Damaged or returned"),
           ("Unallocated", 0.75, "Reserve")]
    sec4_first = sub4 + 1
    for i, (lb, v, n) in enumerate(buf):
        r = sec4_first + i
        ws.cell(row=r, column=1, value=lb).font = BODY
        ws.cell(row=r, column=2, value=v).font = INP_FONT
        ws.cell(row=r, column=3, value=n).font = MUTED
        for wi in range(1, N_WEEKS + 1):
            ws.cell(row=r, column=FW_FIRST + wi - 1, value=f"=$B{r}").font = LINK_FONT
    sec4_last = sec4_first + len(buf) - 1
    sec4_total = sec4_last + 1
    ws.cell(row=sec4_total, column=1, value="Buffer total").font = BOLD
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(FW_FIRST + wi - 1)
        c = ws.cell(row=sec4_total, column=FW_FIRST + wi - 1,
                    value=f"=SUM({cl}{sec4_first}:{cl}{sec4_last})")
        c.font = BOLD; c.fill = SUBT_FILL

    # Wire summary
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(FW_FIRST + wi - 1)
        ws.cell(row=5, column=FW_FIRST + wi - 1, value=f"={cl}{sec1_total}").font = LINK_FONT
        ws.cell(row=6, column=FW_FIRST + wi - 1, value=f"={cl}{sec2_total}").font = LINK_FONT
        ws.cell(row=7, column=FW_FIRST + wi - 1, value=f"={cl}{link_row}").font = LINK_FONT
        ws.cell(row=8, column=FW_FIRST + wi - 1, value=f"={cl}{sec4_total}").font = LINK_FONT
        c = ws.cell(row=9, column=FW_FIRST + wi - 1, value=f"=SUM({cl}5:{cl}8)")
        c.font = BOLD; c.fill = SUBT_FILL

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    for wi in range(1, N_WEEKS + 1):
        ws.column_dimensions[get_column_letter(FW_FIRST + wi - 1)].width = 11
    for j in range(12):
        ws.column_dimensions[get_column_letter(REP_M_FIRST + j)].width = 11
    for wi in range(1, N_WEEKS + 1):
        ws.column_dimensions[get_column_letter(REP_W_FIRST + wi - 1)].width = 11
    ws.freeze_panes = "E5"
    return {"acct": 5, "events": 6, "reps": 7, "buf": 8, "total": 9}

def build_fw(ws):
    ws.title = "Forecast Weekly"
    ws["A1"] = "Forecast weekly cases (engine)"; ws["A1"].font = TITLE
    headers = ["SKU", "Province", "Channel", "Brand"] + WEEK_LABELS + ["FY cases", "FY revenue"]
    for j, val in enumerate(headers, start=1):
        c = ws.cell(row=4, column=j, value=val); c.font = H_FONT; c.fill = H_FILL
    fy_cases_c = get_column_letter(FW_FIRST + N_WEEKS)
    fy_rev_c = get_column_letter(FW_FIRST + N_WEEKS + 1)
    row = 5
    for sku in SKUS:
        brand = SKU_BRAND[sku]
        for p in PROVINCES:
            for ch in CHANNELS:
                ws.cell(row=row, column=1, value=sku).font = BODY
                ws.cell(row=row, column=2, value=p).font = BODY
                ws.cell(row=row, column=3, value=ch).font = BODY
                ws.cell(row=row, column=4, value=brand).font = BODY
                for wi in range(1, N_WEEKS + 1):
                    col = FW_FIRST + wi - 1
                    dwc = doors_w_col(wi)
                    f = (f"=SUMIFS(Doors!{dwc}:{dwc},Doors!$A:$A,$D{row},"
                         f"Doors!$B:$B,$B{row},Doors!$C:$C,$C{row})"
                         f"*SUMIFS(Velocity!$I:$I,Velocity!$A:$A,$A{row},"
                         f"Velocity!$B:$B,$C{row})/Assumptions!$B$3*Assumptions!$B$5")
                    ws.cell(row=row, column=col, value=f).font = LINK_FONT
                ws.cell(row=row, column=FW_FIRST + N_WEEKS,
                        value=f"=SUM({fw_col(1)}{row}:{fw_col(N_WEEKS)}{row})")
                ws.cell(row=row, column=FW_FIRST + N_WEEKS + 1,
                        value=f"={fy_cases_c}{row}*IFERROR(VLOOKUP($A{row},Pricing!$A$5:$B$200,2,FALSE),0)").number_format = "$#,##0"
                row += 1
    total = row
    ws.cell(row=total, column=1, value="Total").font = BOLD
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(FW_FIRST + wi - 1)
        ws.cell(row=total, column=FW_FIRST + wi - 1, value=f"=SUM({cl}5:{cl}{total-1})").font = BOLD
    for cl in (fy_cases_c, fy_rev_c):
        ws.cell(row=total, column=ws[f"{cl}4"].column,
                value=f"=SUM({cl}5:{cl}{total-1})").font = BOLD
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["C"].width = 16
    for wi in range(1, N_WEEKS + 1):
        ws.column_dimensions[get_column_letter(FW_FIRST + wi - 1)].width = 11
    ws.freeze_panes = "E5"
    return total

def build_production(ws, fw_total, mk):
    ws.title = "Production Plan"
    ws["A1"] = "Production plan"; ws["A1"].font = TITLE
    fy = FW_FIRST + N_WEEKS
    fy_col = get_column_letter(fy)
    head_row(ws, 4, ["SKU"] + [""]*3 + WEEK_LABELS + ["FY total"])
    for i, sku in enumerate(SKUS):
        r = 5 + i
        ws.cell(row=r, column=1, value=sku).font = BODY
        for wi in range(1, N_WEEKS + 1):
            cl = get_column_letter(FW_FIRST + wi - 1)
            ws.cell(row=r, column=FW_FIRST + wi - 1,
                    value=f"=SUMIFS('Forecast Weekly'!{cl}:{cl},'Forecast Weekly'!$A:$A,$A{r})").font = LINK_FONT
        ws.cell(row=r, column=fy, value=f"=SUM({fw_col(1)}{r}:{fw_col(N_WEEKS)}{r})")
    tsr = 5 + len(SKUS)  # total sales row
    ws.cell(row=tsr, column=1, value="Total sales cases").font = BOLD
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(FW_FIRST + wi - 1)
        c = ws.cell(row=tsr, column=FW_FIRST + wi - 1, value=f"=SUM({cl}5:{cl}{tsr-1})")
        c.font = BOLD; c.fill = SUBT_FILL
    ws.cell(row=tsr, column=fy, value=f"=SUM({fw_col(1)}{tsr}:{fw_col(N_WEEKS)}{tsr})").font = BOLD

    mh = tsr + 2
    ws.cell(row=mh, column=1, value="Marketing breakdown").font = SECTION
    rows = [("Sampling", mk["acct"]), ("Events", mk["events"]),
            ("Rep samples", mk["reps"]), ("Buffer", mk["buf"])]
    mfirst = mh + 1
    for i, (lb, mr) in enumerate(rows):
        r = mfirst + i
        ws.cell(row=r, column=1, value=lb).font = BODY
        for wi in range(1, N_WEEKS + 1):
            cl = get_column_letter(FW_FIRST + wi - 1)
            ws.cell(row=r, column=FW_FIRST + wi - 1, value=f"=Marketing!{cl}{mr}").font = LINK_FONT
        ws.cell(row=r, column=fy, value=f"=SUM({fw_col(1)}{r}:{fw_col(N_WEEKS)}{r})")
    mtr = mfirst + len(rows)
    ws.cell(row=mtr, column=1, value="Total marketing").font = BOLD
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(FW_FIRST + wi - 1)
        c = ws.cell(row=mtr, column=FW_FIRST + wi - 1, value=f"=SUM({cl}{mfirst}:{cl}{mtr-1})")
        c.font = BOLD; c.fill = SUBT_FILL
    ws.cell(row=mtr, column=fy, value=f"=SUM({fw_col(1)}{mtr}:{fw_col(N_WEEKS)}{mtr})").font = BOLD

    tpr = mtr + 2
    ws.cell(row=tpr, column=1, value="Total cases to produce").font = BOLD
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(FW_FIRST + wi - 1)
        ws.cell(row=tpr, column=FW_FIRST + wi - 1, value=f"={cl}{tsr}+{cl}{mtr}").font = BOLD
    ws.cell(row=tpr, column=fy, value=f"=SUM({fw_col(1)}{tpr}:{fw_col(N_WEEKS)}{tpr})").font = BOLD

    br = tpr + 1
    ws.cell(row=br, column=1, value="Production with buffer").font = BOLD
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(FW_FIRST + wi - 1)
        c = ws.cell(row=br, column=FW_FIRST + wi - 1, value=f"={cl}{tpr}*(1+Assumptions!$B$4)")
        c.font = BOLD; c.fill = SUBT_FILL
    ws.cell(row=br, column=fy, value=f"=SUM({fw_col(1)}{br}:{fw_col(N_WEEKS)}{br})").font = BOLD

    ws.column_dimensions["A"].width = 30
    for wi in range(1, N_WEEKS + 1):
        ws.column_dimensions[get_column_letter(FW_FIRST + wi - 1)].width = 11
    ws.column_dimensions[fy_col].width = 14
    ws.freeze_panes = "B5"
    return {"sales": tsr, "mk": mtr, "prod": tpr, "buf": br, "fy_col": fy}

def build_revenue(ws):
    ws.title = "Revenue"
    ws["A1"] = "Revenue and gross profit"; ws["A1"].font = TITLE
    fy = FW_FIRST + N_WEEKS
    pr = "Pricing!$A$5:$D$200"
    head_row(ws, 4, ["SKU"] + [""]*3 + WEEK_LABELS + ["FY total"])
    for i, sku in enumerate(SKUS):
        r = 5 + i
        ws.cell(row=r, column=1, value=sku).font = BODY
        for wi in range(1, N_WEEKS + 1):
            cl = get_column_letter(FW_FIRST + wi - 1)
            ws.cell(row=r, column=FW_FIRST + wi - 1,
                    value=f"=SUMIFS('Forecast Weekly'!{cl}:{cl},'Forecast Weekly'!$A:$A,$A{r})*"
                          f"VLOOKUP($A{r},{pr},2,FALSE)").number_format = "$#,##0"
        ws.cell(row=r, column=fy, value=f"=SUM({fw_col(1)}{r}:{fw_col(N_WEEKS)}{r})").number_format = "$#,##0"
    rtr = 5 + len(SKUS)
    ws.cell(row=rtr, column=1, value="Total revenue").font = BOLD
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(FW_FIRST + wi - 1)
        c = ws.cell(row=rtr, column=FW_FIRST + wi - 1, value=f"=SUM({cl}5:{cl}{rtr-1})")
        c.font = BOLD; c.fill = SUBT_FILL; c.number_format = "$#,##0"
    ws.cell(row=rtr, column=fy, value=f"=SUM({fw_col(1)}{rtr}:{fw_col(N_WEEKS)}{rtr})").number_format = "$#,##0"

    gh = rtr + 2
    ws.cell(row=gh, column=1, value="Gross profit").font = SECTION
    head_row(ws, gh + 1, ["SKU"] + [""]*3 + WEEK_LABELS + ["FY total"])
    for i, sku in enumerate(SKUS):
        r = gh + 2 + i
        ws.cell(row=r, column=1, value=sku).font = BODY
        for wi in range(1, N_WEEKS + 1):
            cl = get_column_letter(FW_FIRST + wi - 1)
            ws.cell(row=r, column=FW_FIRST + wi - 1,
                    value=f"=SUMIFS('Forecast Weekly'!{cl}:{cl},'Forecast Weekly'!$A:$A,$A{r})*"
                          f"VLOOKUP($A{r},{pr},4,FALSE)").number_format = "$#,##0"
        ws.cell(row=r, column=fy, value=f"=SUM({fw_col(1)}{r}:{fw_col(N_WEEKS)}{r})").number_format = "$#,##0"
    gtr = gh + 2 + len(SKUS)
    ws.cell(row=gtr, column=1, value="Total GP").font = BOLD
    for wi in range(1, N_WEEKS + 1):
        cl = get_column_letter(FW_FIRST + wi - 1)
        c = ws.cell(row=gtr, column=FW_FIRST + wi - 1, value=f"=SUM({cl}{gh+2}:{cl}{gtr-1})")
        c.font = BOLD; c.fill = SUBT_FILL; c.number_format = "$#,##0"
    ws.cell(row=gtr, column=fy, value=f"=SUM({fw_col(1)}{gtr}:{fw_col(N_WEEKS)}{gtr})").number_format = "$#,##0"

    ws.column_dimensions["A"].width = 22
    for wi in range(1, N_WEEKS + 1):
        ws.column_dimensions[get_column_letter(FW_FIRST + wi - 1)].width = 11
    ws.column_dimensions[get_column_letter(fy)].width = 14
    ws.freeze_panes = "B5"
    return {"rev": rtr, "gp": gtr, "fy_col": fy}

def build_trade(ws, doors_total):
    ws.title = "Trade Spend"
    ws["A1"] = "Trade spend"; ws["A1"].font = TITLE
    ws["A2"] = "Listing fees (one-time) + slotting (per door) + scan/promo (% of revenue)."
    ws["A2"].font = MUTED
    head_row(ws, 4, ["Channel", "Listing fee", "Slotting/door", "Scan/promo %",
                     "Listing $", "Slotting $", "Scan/promo $", "Total"])
    rows = [
        ("Natural", 5000, 0, 0.02, False),
        ("Specialty", 3000, 0, 0.02, False),
        ("FDM", 25000, 500, 0.05, False),
        ("Costco", 50000, 0, 0.08, False),
        ("On-Premise", 5000, 50, 0.03, True),
        ("Convenience", 15000, 250, 0.04, True),
        ("Gym & Fitness", 2000, 0, 0.02, True),
        ("Private Liquor", 4000, 100, 0.03, True),
        ("RAS", 1000, 0, 0.02, True),
    ]
    for i, (ch, lf, sl, sp, new) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=ch).font = BODY
        for col, v in zip((2, 3, 4), (lf, sl, sp)):
            cell = ws.cell(row=i, column=col, value=v); cell.font = INP_FONT
            if new: cell.fill = YELLOW
        ws.cell(row=i, column=5, value=f"=B{i}")
        ws.cell(row=i, column=6,
                value=f"=$C{i}*SUMIFS(Doors!{doors_w_col(N_WEEKS)}:{doors_w_col(N_WEEKS)},Doors!$C:$C,$A{i})")
        ws.cell(row=i, column=7,
                value=f"=$D{i}*SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,$A{i})")
        ws.cell(row=i, column=8, value=f"=E{i}+F{i}+G{i}")
        for k in (2, 3, 5, 6, 7, 8): ws.cell(row=i, column=k).number_format = "$#,##0"
        ws.cell(row=i, column=4).number_format = "0.0%"
    tr = 5 + len(rows)
    ws.cell(row=tr, column=1, value="Total").font = BOLD
    for k in (5, 6, 7, 8):
        cl = get_column_letter(k)
        cell = ws.cell(row=tr, column=k, value=f"=SUM({cl}5:{cl}{tr-1})")
        cell.font = BOLD; cell.fill = SUBT_FILL; cell.number_format = "$#,##0"
    for c, w in zip("ABCDEFGH", (16, 14, 14, 14, 14, 14, 14, 14)):
        ws.column_dimensions[c].width = w
    setup_print(ws)
    return {"total_row": tr}

def build_opex(ws):
    ws.title = "OPEX"
    ws["A1"] = "OPEX (FY26-27)"; ws["A1"].font = TITLE
    ws["A2"] = "Reps, A&P, freight, G&A. Type column drives the P&L."
    ws["A2"].font = MUTED
    head_row(ws, 4, ["Line item", "FY amount", "Type", "Notes"])
    rows = [
        ("Sales reps (5 FTE loaded)", 360000, "SG&A", "5 reps blended ~$72K each"),
        ("Sales lead (Louis)", 130000, "SG&A", ""),
        ("Brokers (commission)", 140000, "SG&A", "~3% of revenue"),
        ("T&E", 60000, "SG&A", "Trade shows, field"),
        ("Freight to DCs", 180000, "Logistics", "Outbound"),
        ("Warehousing 3PL", 50000, "Logistics", ""),
        ("Paid digital", 120000, "A&P", "Meta, TikTok, Google"),
        ("Influencer & content", 90000, "A&P", "Creators + content"),
        ("Events & sampling cash", 180000, "A&P", "CHFA, Stampede, etc."),
        ("PR / agency", 60000, "A&P", "Retainer"),
        ("Brand design / packaging", 30000, "A&P", "Refresh + new SKU"),
        ("Insurance & compliance", 25000, "G&A", "Product liability"),
        ("Software & data", 18000, "G&A", "Spins/Nielsen"),
        ("Slotting (in Trade Spend tab)", 90000, "Trade", "Avoid double-count"),
        ("Contingency", 60000, "G&A", ""),
    ]
    for i, (lb, amt, t, n) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=lb).font = BODY
        ws.cell(row=i, column=2, value=amt).font = INP_FONT
        ws.cell(row=i, column=2).number_format = "$#,##0"
        ws.cell(row=i, column=3, value=t).font = BODY
        ws.cell(row=i, column=4, value=n).font = MUTED
    last = 4 + len(rows)
    tot = last + 1
    ws.cell(row=tot, column=1, value="Total OPEX").font = BOLD
    ws.cell(row=tot, column=2, value=f"=SUM(B5:B{last})").number_format = "$#,##0"
    ws.cell(row=tot, column=2).font = BOLD; ws.cell(row=tot, column=2).fill = SUBT_FILL
    sub = tot + 2
    ws.cell(row=sub, column=1, value="By type").font = SECTION
    for i, t in enumerate(["SG&A", "A&P", "Logistics", "G&A", "Trade"], start=sub + 1):
        ws.cell(row=i, column=1, value=t).font = BODY
        ws.cell(row=i, column=2,
                value=f'=SUMIFS(B5:B{last},C5:C{last},"{t}")').number_format = "$#,##0"
    for c, w in zip("ABCD", (40, 16, 12, 50)):
        ws.column_dimensions[c].width = w
    setup_print(ws, landscape=False)
    return {"first": 5, "last": last, "total_row": tot}

def build_pl(ws, rev, trade, opex):
    ws.title = "P&L"
    ws["A1"] = "P&L (FY26-27, CAD)"; ws["A1"].font = TITLE
    head_row(ws, 4, ["Line", "Amount", "% of revenue", "Notes"])
    fy_rev = get_column_letter(rev["fy_col"])
    revc = f"Revenue!{fy_rev}{rev['rev']}"
    gpc = f"Revenue!{fy_rev}{rev['gp']}"
    trc = f"'Trade Spend'!H{trade['total_row']}"
    opx = lambda t: (f"SUMIFS(OPEX!B{opex['first']}:B{opex['last']},"
                     f"OPEX!C{opex['first']}:C{opex['last']},\"{t}\")")
    rows = [
        ("Gross revenue", f"={revc}", "Top line"),
        ("Trade spend", f"=-{trc}", "Channel cash"),
        ("Net revenue", "=B5+B6", "Revenue net of trade"),
        ("COGS (landed)", f"={gpc}-{revc}", "Computed"),
        ("Gross profit", f"={gpc}+B6", "After trade"),
        ("A&P", f"=-{opx('A&P')}", "Marketing cash"),
        ("SG&A", f"=-{opx('SG&A')}", "Reps + brokers"),
        ("Logistics", f"=-{opx('Logistics')}", "Freight + 3PL"),
        ("G&A", f"=-{opx('G&A')}", "Admin"),
        ("EBITDA", "=B9+B10+B11+B12+B13", "Pre-tax operating income"),
        ("EBITDA margin", "=B14/B7", "% of net revenue"),
    ]
    for i, (lb, fml, n) in enumerate(rows, start=5):
        bold = lb in ("Net revenue", "Gross profit", "EBITDA", "EBITDA margin")
        ws.cell(row=i, column=1, value=lb).font = Font(name="Helvetica", bold=bold, size=11, color=INK)
        c = ws.cell(row=i, column=2, value=fml)
        c.number_format = "$#,##0"
        if bold: c.fill = SUBT_FILL; c.font = Font(name="Helvetica", bold=True, size=11, color=INK)
        if lb == "EBITDA margin": c.number_format = "0.0%"
        if lb != "EBITDA margin":
            pct = ws.cell(row=i, column=3, value=f"=B{i}/$B$5"); pct.number_format = "0.0%"
        ws.cell(row=i, column=4, value=n).font = MUTED
    ws.conditional_formatting.add("B14",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_F))
    ws.conditional_formatting.add("B14",
        CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_F))
    for c, w in zip("ABCD", (36, 18, 14, 50)):
        ws.column_dimensions[c].width = w
    setup_print(ws, landscape=False)

def build_scenarios(ws, rev, trade, opex, prod):
    ws.title = "Scenarios"
    ws["A1"] = "Scenarios"; ws["A1"].font = TITLE
    ws["A2"] = "Conservative / Base / Stretch side by side. Edit column B."
    ws["A2"].font = MUTED
    head_row(ws, 4, ["Scenario", "Multiplier", "Sales cases", "Revenue",
                     "Gross profit", "Trade spend", "Net revenue", "OPEX cash",
                     "EBITDA (est.)", "Margin"])
    fy_rev = get_column_letter(rev["fy_col"])
    fy_prod = get_column_letter(prod["fy_col"])
    base_rev = f"Revenue!{fy_rev}{rev['rev']}"
    base_gp = f"Revenue!{fy_rev}{rev['gp']}"
    base_cases = f"'Production Plan'!{fy_prod}{prod['sales']}"
    trade_t = f"'Trade Spend'!H{trade['total_row']}"
    opx = (f"SUMIFS(OPEX!B{opex['first']}:B{opex['last']},"
           f"OPEX!C{opex['first']}:C{opex['last']},\"A&P\")+"
           f"SUMIFS(OPEX!B{opex['first']}:B{opex['last']},"
           f"OPEX!C{opex['first']}:C{opex['last']},\"SG&A\")+"
           f"SUMIFS(OPEX!B{opex['first']}:B{opex['last']},"
           f"OPEX!C{opex['first']}:C{opex['last']},\"Logistics\")+"
           f"SUMIFS(OPEX!B{opex['first']}:B{opex['last']},"
           f"OPEX!C{opex['first']}:C{opex['last']},\"G&A\")")
    # IFERROR wrap to guard against divide-by-zero
    for i, (name, mult) in enumerate([("Conservative", 0.85), ("Base", 1.00), ("Stretch", 1.15)], start=5):
        ws.cell(row=i, column=1, value=name).font = BODY
        c = ws.cell(row=i, column=2, value=mult); c.font = INP_FONT; c.fill = YELLOW
        ws.cell(row=i, column=3, value=f"=IFERROR({base_cases}/Assumptions!$B$5*$B{i},0)").number_format = "#,##0"
        ws.cell(row=i, column=4, value=f"=IFERROR({base_rev}/Assumptions!$B$5*$B{i},0)").number_format = "$#,##0"
        ws.cell(row=i, column=5, value=f"=IFERROR({base_gp}/Assumptions!$B$5*$B{i},0)").number_format = "$#,##0"
        ws.cell(row=i, column=6, value=f"=IFERROR({trade_t}/Assumptions!$B$5*$B{i},0)").number_format = "$#,##0"
        ws.cell(row=i, column=7, value=f"=D{i}-F{i}").number_format = "$#,##0"
        ws.cell(row=i, column=8, value=f"={opx}").number_format = "$#,##0"
        ws.cell(row=i, column=9, value=f"=E{i}-F{i}-H{i}").number_format = "$#,##0"
        ws.cell(row=i, column=10, value=f"=IFERROR(I{i}/G{i},0)").number_format = "0.0%"
    ws.conditional_formatting.add("I5:I7", ColorScaleRule(
        start_type="num", start_value=-500000, start_color="FFB3B3",
        mid_type="num", mid_value=0, mid_color="FFFFCC",
        end_type="num", end_value=500000, end_color="B3F2B3"))
    for i, w in enumerate((14, 12, 14, 14, 14, 14, 14, 14, 14, 12), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    setup_print(ws)

def build_dashboard(ws, prod, rev, doors_total, fw_total):
    ws.title = "Dashboard"
    ws["A1"] = "Dashboard"; ws["A1"].font = TITLE
    fy_rev = get_column_letter(rev["fy_col"])
    fy_prod = get_column_letter(prod["fy_col"])
    month_cols = {m: [] for m in MONTH_ORDER}
    for wi, m in enumerate(WEEK_TO_MONTH, start=1):
        month_cols[m].append(fw_col(wi))

    ws.cell(row=3, column=1, value="By SKU").font = SECTION
    head_row(ws, 4, ["SKU"] + MONTH_ORDER + ["FY total"])
    for i, sku in enumerate(SKUS):
        r = 5 + i
        ws.cell(row=r, column=1, value=sku).font = BODY
        for j, m in enumerate(MONTH_ORDER):
            terms = [f"SUMIFS('Forecast Weekly'!{c}:{c},'Forecast Weekly'!$A:$A,$A{r})"
                     for c in month_cols[m]]
            cell = ws.cell(row=r, column=2 + j, value="=" + "+".join(terms))
            cell.font = LINK_FONT; cell.number_format = "#,##0"
        ws.cell(row=r, column=2 + len(MONTH_ORDER),
                value=f"=SUM(B{r}:M{r})").number_format = "#,##0"

    cb = 5 + len(SKUS) + 2
    ws.cell(row=cb, column=1, value="By channel").font = SECTION
    head_row(ws, cb + 1, ["Channel"] + MONTH_ORDER + ["FY total"])
    for i, ch in enumerate(CHANNELS):
        r = cb + 2 + i
        ws.cell(row=r, column=1, value=ch).font = BODY
        for j, m in enumerate(MONTH_ORDER):
            terms = [f"SUMIFS('Forecast Weekly'!{c}:{c},'Forecast Weekly'!$C:$C,$A{r})"
                     for c in month_cols[m]]
            cell = ws.cell(row=r, column=2 + j, value="=" + "+".join(terms))
            cell.font = LINK_FONT; cell.number_format = "#,##0"
        ws.cell(row=r, column=2 + len(MONTH_ORDER),
                value=f"=SUM(B{r}:M{r})").number_format = "#,##0"

    qb = cb + 2 + len(CHANNELS) + 2
    ws.cell(row=qb, column=1, value="By quarter (channel)").font = SECTION
    head_row(ws, qb + 1, ["Channel"] + QUARTERS)
    q_cols = {q: [] for q in QUARTERS}
    for wi, m in enumerate(WEEK_TO_MONTH, start=1):
        q_cols[MONTH_Q[m]].append(fw_col(wi))
    for i, ch in enumerate(CHANNELS):
        r = qb + 2 + i
        ws.cell(row=r, column=1, value=ch).font = BODY
        for j, q in enumerate(QUARTERS):
            terms = [f"SUMIFS('Forecast Weekly'!{c}:{c},'Forecast Weekly'!$C:$C,$A{r})"
                     for c in q_cols[q]]
            cell = ws.cell(row=r, column=2 + j, value="=" + "+".join(terms))
            cell.font = LINK_FONT; cell.number_format = "#,##0"

    db = qb + 2 + len(CHANNELS) + 2
    ws.cell(row=db, column=1, value="Doors by channel (end of month)").font = SECTION
    head_row(ws, db + 1, ["Channel"] + MONTH_ORDER)
    for i, ch in enumerate(CHANNELS):
        r = db + 2 + i
        ws.cell(row=r, column=1, value=ch).font = BODY
        for j, m in enumerate(MONTH_ORDER):
            cl = MONTH_COL[m]
            cell = ws.cell(row=r, column=2 + j,
                value=f"=SUMIFS(Doors!{cl}:{cl},Doors!$C:$C,$A{r})")
            cell.font = LINK_FONT; cell.number_format = "#,##0"

    ws.column_dimensions["A"].width = 26
    for c in range(2, 2 + max(len(MONTH_ORDER), len(QUARTERS)) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    setup_print(ws)

def build_risks(ws):
    ws.title = "Risks"
    ws["A1"] = "Risks"; ws["A1"].font = TITLE
    head_row(ws, 4, ["ID", "Risk", "Likelihood (1-5)", "Impact (1-5)", "Score",
                     "Mitigation", "Owner", "Status"])
    rows = [
        ("R01", "Co-man capacity slips", 3, 5, "Lock Q4 PO 8 wks ahead.", "Rijo", "Amber"),
        ("R02", "Costco roadshow doesn't convert", 3, 4, "Brief on KPIs, pre-align with buyer.", "Aaron", "Amber"),
        ("R03", "New-channel velocity wrong (overstated)", 4, 4, "Validate at 50 doors per channel.", "Louis", "Amber"),
        ("R04", "LCA Energy GP under 45%", 3, 4, "Renegotiate co-man.", "Rijo", "Red"),
        ("R05", "Sales rep hiring delayed", 3, 3, "Pipeline 2 candidates per role.", "Louis", "Amber"),
        ("R06", "Trade spend exceeds plan", 3, 3, "Monthly scorecard.", "Aaron", "Green"),
        ("R07", "On-premise broker gap", 3, 3, "Hire specialist or use food distributor.", "Louis", "Amber"),
        ("R08", "Convenience slotting overruns", 3, 3, "Negotiate in-kind (free fills).", "Aaron", "Green"),
        ("R09", "Private Liquor regulation change", 2, 3, "Monitor BC/AB boards.", "Louis", "Green"),
        ("R10", "RAS rural freight", 3, 2, "Group via LCBO hub.", "Rijo", "Green"),
        ("R11", "Event spend doesn't convert", 3, 3, "Tag with unique URL/code.", "Teresa", "Amber"),
        ("R12", "MUV/LCA cannibalisation BC Natural", 2, 2, "Differentiate POS.", "Louis", "Green"),
    ]
    for i, (id_, r, l, im, mit, own, st) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=id_).font = BODY
        ws.cell(row=i, column=2, value=r).font = BODY
        ws.cell(row=i, column=3, value=l).font = INP_FONT
        ws.cell(row=i, column=4, value=im).font = INP_FONT
        ws.cell(row=i, column=5, value=f"=C{i}*D{i}")
        ws.cell(row=i, column=6, value=mit).font = MUTED
        ws.cell(row=i, column=7, value=own).font = BODY
        ws.cell(row=i, column=8, value=st).font = BODY
    last = 4 + len(rows)
    ws.conditional_formatting.add(f"E5:E{last}", ColorScaleRule(
        start_type="num", start_value=1, start_color="B3F2B3",
        mid_type="num", mid_value=9, mid_color="FFFFCC",
        end_type="num", end_value=25, end_color="FFB3B3"))
    dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5,
                        allow_blank=False, showErrorMessage=True)
    ws.add_data_validation(dv); dv.add(f"C5:D{last}")
    dvs = DataValidation(type="list", formula1='"Green,Amber,Red"', allow_blank=True)
    ws.add_data_validation(dvs); dvs.add(f"H5:H{last}")
    for c, w in zip("ABCDEFGH", (6, 38, 14, 12, 8, 40, 10, 10)):
        ws.column_dimensions[c].width = w
    setup_print(ws)

def build_home(ws, prod, rev, trade, opex, doors_total):
    ws.title = "Home"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for c in "BCDE": ws.column_dimensions[c].width = 22
    ws.column_dimensions["F"].width = 3

    ws.cell(row=2, column=2, value="ORGANIKA RTD").font = Font(name="Helvetica", bold=True, size=11, color=GREY)
    ws.cell(row=3, column=2, value="FY26-27 Forecast").font = TITLE
    ws.cell(row=4, column=2,
            value="September 1, 2026 - August 31, 2027 . Base scenario . CAD").font = MUTED

    fy_rev = get_column_letter(rev["fy_col"])
    fy_prod = get_column_letter(prod["fy_col"])
    rev_c = f"Revenue!{fy_rev}{rev['rev']}"
    gp_c = f"Revenue!{fy_rev}{rev['gp']}"
    trade_c = f"'Trade Spend'!H{trade['total_row']}"
    opx_c = (f"(SUMIFS(OPEX!B{opex['first']}:B{opex['last']},"
             f"OPEX!C{opex['first']}:C{opex['last']},\"A&P\")"
             f"+SUMIFS(OPEX!B{opex['first']}:B{opex['last']},"
             f"OPEX!C{opex['first']}:C{opex['last']},\"SG&A\")"
             f"+SUMIFS(OPEX!B{opex['first']}:B{opex['last']},"
             f"OPEX!C{opex['first']}:C{opex['last']},\"Logistics\")"
             f"+SUMIFS(OPEX!B{opex['first']}:B{opex['last']},"
             f"OPEX!C{opex['first']}:C{opex['last']},\"G&A\"))")

    # Hero strip
    heroes = [
        ("REVENUE", f"={rev_c}", "$#,##0", "Annual sales"),
        ("GROSS PROFIT", f"={gp_c}", "$#,##0", "After landed cost"),
        ("EBITDA", f"={gp_c}-{trade_c}-{opx_c}", "$#,##0", "Operating income"),
        ("CASES TO PRODUCE", f"='Production Plan'!{fy_prod}{prod['buf']}", "#,##0", "With 10% buffer"),
    ]
    for i, (lb, fml, fmt, sub) in enumerate(heroes):
        col = 2 + i
        ws.cell(row=6, column=col, value=lb).font = HERO_LABEL
        v = ws.cell(row=7, column=col, value=fml); v.font = HERO; v.number_format = fmt
        v.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=8, column=col, value=sub).font = MUTED
    ws.row_dimensions[7].height = 44

    # Story
    ws.cell(row=11, column=2, value="THE STORY").font = HERO_LABEL
    ws.cell(row=12, column=2,
            value="Year 1 of a 9-channel national footprint. Doors ramp ~360 -> ~3,000. "
                  "EBITDA negative in Y1 from one-time listing/slotting + full OPEX. "
                  "Path to positive in FY27-28 by amortising slotting and scaling revenue.").font = BODY
    ws.merge_cells(start_row=12, start_column=2, end_row=12, end_column=5)
    ws.cell(row=12, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[12].height = 60

    # At a glance
    ws.cell(row=15, column=2, value="AT A GLANCE").font = HERO_LABEL
    quick = [
        ("GP margin", f"={gp_c}/{rev_c}", "0.0%"),
        ("Sales cases", f"='Production Plan'!{fy_prod}{prod['sales']}", "#,##0"),
        ("Doors at year-end", f"=Doors!{doors_w_col(N_WEEKS)}{doors_total}", "#,##0"),
        ("Revenue per door", f"={rev_c}/Doors!{doors_w_col(N_WEEKS)}{doors_total}", "$#,##0"),
        ("New-channel rev share",
         f"=(SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"On-Premise\")+"
         f"SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"Convenience\")+"
         f"SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"Gym & Fitness\")+"
         f"SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"Private Liquor\")+"
         f"SUMIFS('Forecast Weekly'!$BF:$BF,'Forecast Weekly'!$C:$C,\"RAS\"))/{rev_c}", "0.0%"),
        ("Trade spend %", f"={trade_c}/{rev_c}", "0.0%"),
        ("Scenario multiplier", "=Assumptions!B5", "0.00"),
        ("Production buffer", "=Assumptions!B4", "0%"),
    ]
    for i, (lb, fml, fmt) in enumerate(quick):
        r = 17 + (i // 2) * 2; c = 2 + (i % 2) * 2
        ws.cell(row=r, column=c, value=lb).font = SMALL_LABEL
        v = ws.cell(row=r + 1, column=c, value=fml); v.font = BIG_NUM; v.number_format = fmt

    # Navigation
    nav = 26
    ws.cell(row=nav, column=2, value="WHERE TO GO").font = HERO_LABEL
    pointers = [
        ("Adjust the model", "Controls - the 5 knobs"),
        ("Door plan", "Doors - 12 monthly columns drive 52 weekly"),
        ("Pricing", "Pricing - net price + landed cost per SKU"),
        ("Marketing", "Marketing - sampling, events, rep samples, buffer"),
        ("Trade spend", "Trade Spend - channel listing, slotting, scan/promo"),
        ("OPEX", "OPEX - reps, A&P, freight, G&A"),
        ("Full P&L", "P&L - revenue -> EBITDA"),
        ("Compare scenarios", "Scenarios"),
        ("Drill down", "Dashboard - monthly + quarterly by SKU/channel"),
        ("Risks", "Risks - top 12 scored"),
    ]
    for i, (act, where) in enumerate(pointers, start=nav + 2):
        ws.cell(row=i, column=2, value=act).font = BOLD
        ws.cell(row=i, column=3, value=where).font = MUTED
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
    ws.cell(row=nav + 2 + len(pointers) + 2, column=2,
            value="Owner: Louis . Source of truth. Press F9 to recalc.").font = MUTED

def build_controls(ws):
    ws.title = "Controls"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 60

    ws.cell(row=2, column=2, value="CONTROLS").font = Font(name="Helvetica", bold=True, size=11, color=GREY)
    ws.cell(row=3, column=2, value="Five knobs").font = TITLE
    ws.cell(row=4, column=2, value="Edit these five cells. Everything downstream recalculates.").font = MUTED
    ws.cell(row=6, column=2, value="INPUT").font = HERO_LABEL
    ws.cell(row=6, column=3, value="VALUE").font = HERO_LABEL
    ws.cell(row=6, column=4, value="WHAT IT DOES").font = HERO_LABEL
    knobs = [
        ("Scenario multiplier", "=Assumptions!B5", "0.00",
         "0.85 conservative, 1.00 base, 1.15 stretch. Scales sales."),
        ("Production buffer", "=Assumptions!B4", "0%",
         "Cushion on production. Industry norm 10-15%."),
        ("Units per case", "=Assumptions!B3", "#,##0",
         "Fixed at 24 cans."),
        ("MUV price per case", "=Pricing!B5", "$#,##0.00",
         "Net price for all three MUV SKUs."),
        ("LCA price per case", "=Pricing!B8", "$#,##0.00",
         "Net price for LCA Energy."),
    ]
    for i, (lb, fml, fmt, n) in enumerate(knobs, start=8):
        ws.cell(row=i, column=2, value=lb).font = Font(name="Helvetica", bold=True, size=12, color=INK)
        c = ws.cell(row=i, column=3, value=fml)
        c.font = Font(name="Helvetica", bold=True, size=14, color=BLUE); c.number_format = fmt
        ws.cell(row=i, column=4, value=n).font = MUTED
        ws.row_dimensions[i].height = 26

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Create in dependency order, then reorder for display
    ass_ws = wb.create_sheet("Assumptions"); build_assumptions(ass_ws)
    cm_ws = wb.create_sheet("Class Mix"); build_class_mix(cm_ws)
    vel_ws = wb.create_sheet("Velocity"); build_velocity(vel_ws)
    pr_ws = wb.create_sheet("Pricing"); build_pricing(pr_ws)
    do_ws = wb.create_sheet("Doors"); doors_total = build_doors(do_ws)
    mk_ws = wb.create_sheet("Marketing"); mk_refs = build_marketing(mk_ws)
    fw_ws = wb.create_sheet("Forecast Weekly"); fw_total = build_fw(fw_ws)
    pp_ws = wb.create_sheet("Production Plan"); prod_refs = build_production(pp_ws, fw_total, mk_refs)
    re_ws = wb.create_sheet("Revenue"); rev_refs = build_revenue(re_ws)
    tr_ws = wb.create_sheet("Trade Spend"); trade_refs = build_trade(tr_ws, doors_total)
    op_ws = wb.create_sheet("OPEX"); opex_refs = build_opex(op_ws)
    pl_ws = wb.create_sheet("P&L"); build_pl(pl_ws, rev_refs, trade_refs, opex_refs)
    sc_ws = wb.create_sheet("Scenarios"); build_scenarios(sc_ws, rev_refs, trade_refs, opex_refs, prod_refs)
    db_ws = wb.create_sheet("Dashboard"); build_dashboard(db_ws, prod_refs, rev_refs, doors_total, fw_total)
    ri_ws = wb.create_sheet("Risks"); build_risks(ri_ws)
    co_ws = wb.create_sheet("Controls"); build_controls(co_ws)
    ho_ws = wb.create_sheet("Home"); build_home(ho_ws, prod_refs, rev_refs, trade_refs, opex_refs, doors_total)

    # Hide engine + advanced tabs
    for name in ("Forecast Weekly", "Production Plan", "Revenue",
                 "Assumptions", "Class Mix", "Velocity"):
        wb[name].sheet_state = "hidden"

    # Tab colors
    for name in ("Home", "Controls"):
        wb[name].sheet_properties.tabColor = BLUE
    for name in ("P&L", "Scenarios", "Dashboard"):
        wb[name].sheet_properties.tabColor = INK
    for name in ("Doors", "Pricing", "Marketing", "Trade Spend", "OPEX"):
        wb[name].sheet_properties.tabColor = ORANGE
    wb["Risks"].sheet_properties.tabColor = GREY

    # Reorder: Home, Controls, Doors, Pricing, Marketing, Trade Spend, OPEX,
    #          P&L, Scenarios, Dashboard, Risks, (hidden engine)
    desired = ["Home", "Controls", "Doors", "Pricing", "Marketing", "Trade Spend",
               "OPEX", "P&L", "Scenarios", "Dashboard", "Risks",
               "Forecast Weekly", "Production Plan", "Revenue",
               "Assumptions", "Class Mix", "Velocity"]
    wb._sheets = [wb[n] for n in desired]
    wb.active = wb.sheetnames.index("Home")

    out = "/home/user/my-first-project/Organika_RTD_Forecast.xlsx"
    wb.save(out)
    print(f"Wrote: {out}")


if __name__ == "__main__":
    main()
