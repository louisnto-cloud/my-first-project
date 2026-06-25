#!/usr/bin/env python3
"""
Build `organika-cpg-pnl.xlsx` — an enterprise CPG P&L model for
Organika MÜV Sparkling Electrolytes.

Everything is live formulas: a bill-of-materials COGS build-up feeds a
channel engine (DTC / Amazon / Retail / Club), which rolls up to a blended
gross-sales-to-EBITDA P&L, plus a price×COGS sensitivity heatmap and a
Base/Bull/Bear scenario block.  Run:  python3 build_pnl_workbook.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment

# palette ------------------------------------------------------------------
NAVY="14213F"; INK="0E1320"; TEAL="2BB3A3"; BLUE="4F8CFF"; TEALBG="E8F5F3"
AMBER="FFF3C4"; AMBERB="E6C84B"; GREY="93A0BA"; LINE="C9D2E0"
RED="F8696B"; YEL="FFEB84"; GRN="63BE7B"; POSBG="E4F7EE"
MONEY='"$"#,##0'; MONEY2='"$"#,##0.00'; CENTS='"$"#,##0.000'; PCT='0.0%'; PCT0='0%'; NUM='#,##0'
thin=Side(style="thin",color=LINE); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def font(sz=11,b=False,color="222222",it=False): return Font(name="Calibri",size=sz,bold=b,color=color,italic=it)
def fill(h): return PatternFill("solid",fgColor=h)
def sty(c,*,f=None,fl=None,nf=None,al=None,bd=False,wrap=False):
    if f:c.font=f
    if fl:c.fill=fl
    if nf:c.number_format=nf
    if al:c.alignment=Alignment(horizontal=al,vertical="center",wrap_text=wrap)
    elif wrap:c.alignment=Alignment(vertical="center",wrap_text=True)
    if bd:c.border=BORDER
def title(ws,text,span,sub=None):
    ws.merge_cells(f"A1:{get_column_letter(span)}1"); c=ws["A1"]; c.value=text
    sty(c,f=font(16,True,"FFFFFF"),fl=fill(NAVY),al="left"); ws.row_dimensions[1].height=28
    if sub:
        ws.merge_cells(f"A2:{get_column_letter(span)}2"); s=ws["A2"]; s.value=sub
        sty(s,f=font(9.5,False,"FFFFFF"),fl=fill(NAVY),al="left"); ws.row_dimensions[2].height=16
def section(ws,row,text,span):
    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}"); c=ws[f"A{row}"]; c.value=text.upper()
    sty(c,f=font(10,True,TEAL),al="left")
def lbl(ws,ref,text,b=False,color="333333"):
    c=ws[ref]; c.value=text; sty(c,f=font(11,b,color),al="left"); return c
def inp(ws,ref,val,nf=MONEY,cm=None):
    c=ws[ref]; c.value=val; sty(c,f=font(11,True,"1A1A1A"),fl=fill(AMBER),nf=nf,al="right")
    c.border=Border(left=Side(style="thin",color=AMBERB),right=Side(style="thin",color=AMBERB),
                    top=Side(style="thin",color=AMBERB),bottom=Side(style="thin",color=AMBERB))
    if cm: c.comment=Comment(cm,"Model")
    return c
def out(ws,ref,formula,nf=MONEY,b=False,hl=False):
    c=ws[ref]; c.value=formula
    sty(c,f=font(11,b,"11302C" if hl else "1A1A1A"),fl=fill(TEALBG) if hl else None,nf=nf,al="right",bd=True)
    return c

wb=Workbook()

# ============================================================ 1. P&L (assumptions + statement)
pl=wb.active; pl.title="P&L"; pl.sheet_properties.tabColor=BLUE; pl.sheet_view.showGridLines=False
for col,w in {"A":2.5,"B":34,"C":16,"D":15,"E":11}.items(): pl.column_dimensions[col].width=w
title(pl,"Organika MÜV Sparkling Electrolytes — Annual P&L",5,
      "Yellow = inputs. COGS is built on COGS_BOM; revenue & channel costs on Channels. This is the blended roll-up.")
section(pl,4,"Operating assumptions",5)
inp(pl,"C5",250000,NUM,"Total annual cases across all channels."); lbl(pl,"B5","Annual volume (cases)")
inp(pl,"C6",12,NUM); lbl(pl,"B6","Units / case")
lbl(pl,"B7","Total units (cans)"); out(pl,"C7","=C5*C6",NUM)
inp(pl,"C8",0.08,PCT,"Corporate brand marketing as % of net sales (beyond channel ad)."); lbl(pl,"B8","Brand marketing (% net sales)")
inp(pl,"C9",600000,MONEY); lbl(pl,"B9","Sales & team salaries")
inp(pl,"C10",850000,MONEY); lbl(pl,"B10","G&A / overhead")
inp(pl,"C11",180000,MONEY); lbl(pl,"B11","R&D / innovation")
inp(pl,"C12",150000,MONEY); lbl(pl,"B12","Depreciation & amortization")
lbl(pl,"B13","Fixed costs (sales + G&A + R&D)",b=True); out(pl,"C13","=C9+C10+C11",MONEY,b=True)
# statement references (filled after Channels totals row is known)
CH="Channels"; TR=9   # Channels totals row (set when building Channels)
section(pl,16,"P&L statement — gross sales to EBIT",5)
hdr=["Line","Total","Per case","% net"]
for i,h in enumerate(hdr):
    c=pl.cell(row=17,column=2+i,value=h); sty(c,f=font(10,True,"FFFFFF"),fl=fill(NAVY),al="left" if i==0 else "right",bd=True)
# (label, formula_total, kind)  kind: 'pos'|'ded'|'tot'
def chref(cell): return f"{CH}!${cell[0]}${TR}"
NET=f"{CH}!$O${TR}"
lines=[
 ("Gross sales",            f"={chref('L')}","pos"),
 ("  Less: trade & promo",  f"=-{chref('M')}","ded"),
 ("  Less: returns & spoilage", f"=-{chref('N')}","ded"),
 ("Net sales",              f"={chref('O')}","tot"),
 ("  Less: COGS",           f"=-{chref('P')}","ded"),
 ("Gross profit",           f"={chref('Q')}","tot"),
 ("  Less: fulfillment & freight", f"=-{chref('T')}","ded"),
 ("  Less: platform & selling fees", f"=-{chref('R')}","ded"),
 ("  Less: channel marketing (PPC/CAC)", f"=-{chref('S')}","ded"),
 ("  Less: slotting & listing", f"=-{chref('U')}","ded"),
 ("Contribution margin",    f"={chref('V')}","tot"),
 ("  Less: brand marketing", f"=-C8*{NET}","ded"),
 ("  Less: sales & team",   "=-C9","ded"),
 ("  Less: G&A / overhead", "=-C10","ded"),
 ("  Less: R&D",            "=-C11","ded"),
 ("EBITDA",                 f"={chref('V')}-C8*{NET}-C13","tot"),
 ("  Less: D&A",            "=-C12","ded"),
 ("EBIT",                   f"={chref('V')}-C8*{NET}-C13-C12","tot"),
]
r=18; PL_FIRST=r
for label,formula,kind in lines:
    lc=pl.cell(row=r,column=2,value=label)
    sty(lc,f=font(11,kind=='tot',"333333" if kind!='ded' else GREY),al="left",bd=True)
    tc=pl.cell(row=r,column=3,value=formula); sty(tc,f=font(11,kind=='tot'),nf=MONEY,al="right",bd=True)
    if kind=='tot': tc.fill=fill(TEALBG)
    pc=pl.cell(row=r,column=4,value=f"=C{r}/$C$5"); sty(pc,f=font(10,False,GREY),nf=MONEY2,al="right",bd=True)
    nc=pl.cell(row=r,column=5,value=f"=C{r}/{NET}"); sty(nc,f=font(10,False,GREY),nf=PCT,al="right",bd=True)
    r+=1
# margin callouts
pl.conditional_formatting.add(f"C{PL_FIRST+15}", CellIsRule(operator="lessThan",formula=["0"],fill=fill(RED)))
pl.conditional_formatting.add(f"C{PL_FIRST+15}", CellIsRule(operator="greaterThanOrEqual",formula=["0"],fill=fill(GRN)))
lbl(pl,f"B{r+1}","Break-even cases (EBITDA = 0)",b=True)
# break-even = fixed / (contribution_per_case - mktg% * net_per_case)
out(pl,f"C{r+1}", f"=C13/(({chref('V')}/C5)-C8*({NET}/C5))", NUM, b=True, hl=True)

# ============================================================ 2. COGS_BOM
cb=wb.create_sheet("COGS_BOM"); cb.sheet_properties.tabColor=TEAL; cb.sheet_view.showGridLines=False
for col,w in {"A":2.5,"B":34,"C":14}.items(): cb.column_dimensions[col].width=w
title(cb,"Bill of materials — landed cost per 355 ml can",3,
      "Edit any line (yellow). Groups roll up, then a yield-loss allowance grosses up the survivors.")
groups=[
 ("Liquid & ingredients",[("Carbonated water (RO + CO₂)",0.014),("Sodium / potassium citrate blend",0.055),
   ("Magnesium bisglycinate",0.048),("Fibersol® prebiotic fiber",0.030),("Natural flavour",0.040),
   ("Acids (citric / malic / phosphoric)",0.010),("Stevia extract",0.022)]),
 ("Primary packaging",[("Aluminium can 355 ml (printed)",0.150),("Can end / lid",0.022),("Tab",0.004)]),
 ("Secondary packaging",[("Tray (per can)",0.016),("Shrink film (per can)",0.006)]),
 ("Tertiary / logistics",[("Case / carton (per can)",0.006),("Pallet & stretch wrap (per can)",0.003)]),
 ("Co-pack conversion",[("Co-pack fill & seam fee",0.090),("Blending / batching",0.018),("QA / lab / coding",0.008)]),
]
r=4; subtotal_cells=[]
for gname,items in groups:
    section(cb,r,gname,3); r+=1
    first=r
    for nm,cost in items:
        lbl(cb,f"B{r}",nm); inp(cb,f"C{r}",cost,CENTS); r+=1
    sc=cb.cell(row=r,column=2,value="  subtotal"); sty(sc,f=font(10,True,GREY),al="left")
    st=cb.cell(row=r,column=3,value=f"=SUM(C{first}:C{r-1})"); sty(st,f=font(10,True,"333333"),nf=CENTS,al="right",bd=True)
    subtotal_cells.append(f"C{r}"); r+=2
section(cb,r,"Roll-up to COGS / unit",3); r+=1
lbl(cb,f"B{r}","Sum of components"); out(cb,f"C{r}","="+"+".join(subtotal_cells),CENTS); SUMrow=r; r+=1
lbl(cb,f"B{r}","Inbound freight / unit"); inp(cb,f"C{r}",0.016,CENTS,"Freight of ingredients & packaging to the co-packer."); FRT=r; r+=1
lbl(cb,f"B{r}","Landed cost before loss"); out(cb,f"C{r}",f"=C{SUMrow}+C{FRT}",CENTS); BEF=r; r+=1
lbl(cb,f"B{r}","Yield / scrap loss"); inp(cb,f"C{r}",0.04,PCT,"Fill loss & spoilage. Survivors carry it: cost ÷ (1 − loss)."); LOSS=r; r+=1
lbl(cb,f"B{r}","COGS / unit (landed)",b=True); out(cb,f"C{r}",f"=C{BEF}/(1-C{LOSS})",CENTS,b=True,hl=True); COGS_UNIT=f"COGS_BOM!$C${r}"; r+=1
lbl(cb,f"B{r}","COGS / case",b=True); out(cb,f"C{r}",f"=C{r-1}*'P&L'!C6",MONEY2,b=True)

# ============================================================ 3. Channels (the engine)
ch=wb.create_sheet("Channels"); ch.sheet_properties.tabColor="23E0A0"; ch.sheet_view.showGridLines=False
heads=["Channel","Mix %","SRP","Sell price","Trade %","Spoil %","Fee %","Ad %","Ful $","Slot $",
       "Units","Gross","Trade","Spoil","Net","COGS","Gross profit","Fees","Ad","Ful","Slot","Contribution","CM %"]
ws_w=[20,8,8,9,8,8,8,8,8,8,11,12,10,9,12,12,12,10,10,9,9,13,8]
for i,w in enumerate(ws_w): ch.column_dimensions[get_column_letter(i+1)].width=w
title(ch,"Channel economics — route-to-market P&L",len(heads),
      "Each channel carries its own price realization and cost stack. Columns K–W are live formulas; the totals row feeds the P&L.")
HR=4
for i,h in enumerate(heads):
    c=ch.cell(row=HR,column=i+1,value=h); sty(c,f=font(9,True,"FFFFFF"),fl=fill(NAVY),al="left" if i==0 else "right",bd=True)
ch.freeze_panes="B5"
# channel seed: name, mix, srp, sell, trade, spoil, fee, ad, ful, slot
chan=[
 ("DTC (own site)",0.12,2.99,2.99,0.12,0.02,0.03,0.22,0.85,0.00),
 ("Amazon",0.18,2.99,2.79,0.08,0.03,0.15,0.13,0.48,0.00),
 ("Retail / Distributor",0.52,2.99,1.55,0.15,0.02,0.05,0.02,0.11,0.07),
 ("Club / Mass",0.18,2.79,1.40,0.08,0.01,0.03,0.01,0.09,0.10),
]
first=HR+1
for k,row in enumerate(chan):
    rr=first+k
    name,mix,srp,sell,trade,spoil,fee,ad,ful,slot=row
    vals=[(1,name,None,"left"),(2,mix,PCT,None),(3,srp,MONEY2,None),(4,sell,MONEY2,None),(5,trade,PCT,None),
          (6,spoil,PCT,None),(7,fee,PCT,None),(8,ad,PCT,None),(9,ful,CENTS,None),(10,slot,CENTS,None)]
    for col,v,nf,al in vals:
        c=ch.cell(row=rr,column=col,value=v); sty(c,f=font(10,col==1),fl=fill(AMBER),nf=nf,al=al or ("left" if col==1 else "right"),bd=True)
    F={11:f"='P&L'!$C$7*B{rr}", 12:f"=D{rr}*K{rr}", 13:f"=L{rr}*E{rr}", 14:f"=L{rr}*F{rr}",
       15:f"=L{rr}-M{rr}-N{rr}", 16:f"={COGS_UNIT}*K{rr}", 17:f"=O{rr}-P{rr}", 18:f"=L{rr}*G{rr}",
       19:f"=L{rr}*H{rr}", 20:f"=I{rr}*K{rr}", 21:f"=J{rr}*K{rr}", 22:f"=Q{rr}-R{rr}-S{rr}-T{rr}-U{rr}",
       23:f"=IF(O{rr}<>0,V{rr}/O{rr},\"\")"}
    nfmt={11:NUM,12:MONEY,13:MONEY,14:MONEY,15:MONEY,16:MONEY,17:MONEY,18:MONEY,19:MONEY,20:MONEY,21:MONEY,22:MONEY,23:PCT}
    for col,formula in F.items():
        c=ch.cell(row=rr,column=col,value=formula); sty(c,f=font(10),nf=nfmt[col],al="right",bd=True)
TR=first+len(chan)   # totals row  (= 9 with 4 channels)
tc=ch.cell(row=TR,column=1,value="Company total"); sty(tc,f=font(10,True,"FFFFFF"),fl=fill(NAVY),al="left",bd=True)
for col in range(2,24):
    c=ch.cell(row=TR,column=col); sty(c,fl=fill(NAVY),bd=True)
# totals
ch.cell(row=TR,column=2,value=f"=SUM(B{first}:B{TR-1})").number_format=PCT
for col in [11,12,13,14,15,16,17,18,19,20,21,22]:
    L=get_column_letter(col)
    c=ch.cell(row=TR,column=col,value=f"=SUM({L}{first}:{L}{TR-1})")
    sty(c,f=font(10,True,"FFFFFF"),fl=fill(NAVY),nf=NUM if col==11 else MONEY,al="right",bd=True)
ch.cell(row=TR,column=2).font=font(10,True,"FFFFFF"); ch.cell(row=TR,column=2).fill=fill(NAVY); ch.cell(row=TR,column=2).alignment=Alignment(horizontal="right")
cmw=ch.cell(row=TR,column=23,value=f"=V{TR}/O{TR}"); sty(cmw,f=font(10,True,"FFFFFF"),fl=fill(NAVY),nf=PCT,al="right",bd=True)
# heat CM% per channel
ch.conditional_formatting.add(f"W{first}:W{TR-1}",
    ColorScaleRule(start_type="num",start_value=0,start_color=RED,mid_type="num",mid_value=0.25,mid_color=YEL,end_type="num",end_value=0.4,end_color=GRN))

# fix P&L break-even + statement already reference TR via chref/NET (TR known=9). Re-point if needed:
assert TR==9, f"channels totals row changed to {TR}; update P&L refs"

# ============================================================ 4. Sensitivity
sn=wb.create_sheet("Sensitivity"); sn.sheet_properties.tabColor="FFC24B"; sn.sheet_view.showGridLines=False
for col in range(1,9): sn.column_dimensions[get_column_letter(col)].width=12
title(sn,"Sensitivity — EBITDA % across price × COGS",8,
      "Every cell re-runs the blended model for a scaled selling-price and COGS. Centre = today. Edit the range.")
lbl(sn,"B3","Range ±"); inp(sn,"C3",0.20,PCT,"How far to sweep price and COGS around today.")
O=f"Channels!$O${TR}"; P=f"Channels!$P${TR}"; R=f"Channels!$R${TR}"; Sd=f"Channels!$S${TR}"
Tt=f"Channels!$T${TR}"; U=f"Channels!$U${TR}"; MK="'P&L'!$C$8"; FX="'P&L'!$C$13"
GT=6  # header row for COGS multipliers
cor=sn.cell(row=GT,column=2,value="Price ↓ / COGS →"); sty(cor,f=font(9,True,"FFFFFF"),fl=fill(NAVY),al="center",bd=True)
for j in range(7):  # cogs multipliers across C..I  (low cost left)
    col=3+j; c=sn.cell(row=GT,column=col,value=f"=1+$C$3*(({j}/3)-1)")
    sty(c,f=font(9,True,"FFFFFF"),fl=fill(NAVY),nf="0%",al="center",bd=True)
for i in range(7):  # price multipliers down  (high price top)
    rr=GT+1+i; pc=sn.cell(row=rr,column=2,value=f"=1+$C$3*(1-({i}/3))")
    sty(pc,f=font(9,True,"FFFFFF"),fl=fill(NAVY),nf="0%",al="center",bd=True)
    for j in range(7):
        col=3+j; cl=get_column_letter(col)
        pf=f"$B{rr}"; cf=f"{cl}${GT}"
        ebitda=f"({pf}*{O}-{cf}*{P}-{pf}*({R}+{Sd})-({Tt}+{U})-{MK}*{pf}*{O}-{FX})"
        formula=f"={ebitda}/({pf}*{O})"
        c=sn.cell(row=rr,column=col,value=formula); sty(c,f=font(10,False,"1A1A1A"),nf="0%",al="center",bd=True)
sn.conditional_formatting.add(f"C{GT+1}:I{GT+7}",
    ColorScaleRule(start_type="min",start_color=RED,mid_type="percentile",mid_value=50,mid_color=YEL,end_type="max",end_color=GRN))
lbl(sn,"B15","Green = higher EBITDA margin. The steeper the colour shift along a row vs a column, the bigger that lever.",color=GREY)
sn["B15"].font=font(10,False,GREY,it=True)

# ============================================================ 5. Scenario Lab (side by side)
sc=wb.create_sheet("Scenario Lab"); sc.sheet_properties.tabColor="7C7BFF"; sc.sheet_view.showGridLines=False
NCOL=6; LAST=get_column_letter(2+NCOL)
sc.column_dimensions["A"].width=2.5; sc.column_dimensions["B"].width=20
for i in range(NCOL): sc.column_dimensions[get_column_letter(3+i)].width=13
title(sc,"Scenario Lab — compare full plans side by side",2+NCOL,
      "Each column is a plan. Edit the yellow dials; results recompute live and the best in each row turns green.")
section(sc,4,"Model basis (live, from your plan)",2+NCOL)
lbl(sc,"B5","Base price / can");    out(sc,"C5",f"=Channels!$L${TR}/Channels!$K${TR}",MONEY2)
lbl(sc,"B6","Base cost / can");     out(sc,"C6",f"={COGS_UNIT}",CENTS)
lbl(sc,"B7","Base volume (cases)"); out(sc,"C7","='P&L'!$C$5",NUM)
lbl(sc,"B8","Fixed costs");         out(sc,"C8","='P&L'!$C$13",MONEY)
# seeds: name, price, cost, cases, trade, marketing
seeds=[("Base",1.919,0.58125,250000,1.0,0.08),("Bull",2.015,0.52313,350000,0.85,0.08),
       ("Bear",1.861,0.63938,187500,1.15,0.08),("Scenario 4",1.919,0.58125,250000,1.0,0.08),
       ("Scenario 5",1.919,0.58125,250000,1.0,0.08),("Scenario 6",1.919,0.58125,250000,1.0,0.08)]
HR=11
hc=sc.cell(row=HR,column=2,value="Scenario"); sty(hc,f=font(10,True,"FFFFFF"),fl=fill(NAVY),al="left",bd=True)
for i,seed in enumerate(seeds):
    c=sc.cell(row=HR,column=3+i,value=seed[0]); sty(c,f=font(11,True,"1A1A1A"),fl=fill(AMBER),al="center",bd=True)
dials=[("Price / can",1,MONEY2),("Cost / can",2,CENTS),("Cases / year",3,NUM),("Trade level (×)",4,"0.00"),("Marketing (% net)",5,PCT)]
DR0=12
for di,(dlabel,si,nf) in enumerate(dials):
    rr=DR0+di; lbl(sc,f"B{rr}",dlabel)
    for i,seed in enumerate(seeds): inp(sc,f"{get_column_letter(3+i)}{rr}",seed[si],nf)
PRC,CST,CAS,TRD,MKT=DR0,DR0+1,DR0+2,DR0+3,DR0+4
NET,GMr,CMr,BE,EB,EBP=18,19,20,21,22,23
MP,MC,MV,CMD=25,26,27,28
Lc=f"Channels!$L${TR}";Mtc=f"Channels!$M${TR}";Nc=f"Channels!$N${TR}";Pcc=f"Channels!$P${TR}"
Rcc=f"Channels!$R${TR}";Scc=f"Channels!$S${TR}";Tcc=f"Channels!$T${TR}";Ucc=f"Channels!$U${TR}"
res=[("Net sales",NET),("Gross margin %",GMr),("Contribution %",CMr),("Break-even (cases)",BE),("EBITDA",EB),("EBITDA %",EBP)]
for dlabel,rr in res: lbl(sc,f"B{rr}",dlabel,b=(rr in (EB,EBP)))
for i in range(NCOL):
    X=get_column_letter(3+i)
    sc.cell(row=MP,column=3+i,value=f"={X}{PRC}/$C$5")          # price multiplier
    sc.cell(row=MC,column=3+i,value=f"={X}{CST}/$C$6")          # cost multiplier
    sc.cell(row=MV,column=3+i,value=f"={X}{CAS}/$C$7")          # volume multiplier
    sc.cell(row=CMD,column=3+i,value=f"=({X}{NET}-{X}{MC}*{X}{MV}*{Pcc})-{X}{MP}*{X}{MV}*{Rcc}-{X}{MP}*{X}{MV}*{Scc}-{X}{MV}*{Tcc}-{X}{MV}*{Ucc}")  # contribution $
    cells={
      NET:(f"={X}{MP}*{X}{MV}*({Lc}-{X}{TRD}*{Mtc}-{Nc})",MONEY),
      GMr:(f"=({X}{NET}-{X}{MC}*{X}{MV}*{Pcc})/{X}{NET}",PCT),
      CMr:(f"={X}{CMD}/{X}{NET}",PCT),
      EB :(f"={X}{CMD}-{X}{MKT}*{X}{NET}-$C$8",MONEY),
      EBP:(f"={X}{EB}/{X}{NET}",PCT),
      BE :(f"=IF(({X}{EB}+$C$8)>0,$C$8*{X}{CAS}/({X}{EB}+$C$8),\"n/a\")",NUM),
    }
    for rr,(formula,nf) in cells.items():
        c=sc.cell(row=rr,column=3+i,value=formula); sty(c,f=font(12 if rr==EB else 11,rr in (EB,EBP)),nf=nf,al="right",bd=True)
for r in (MP,MC,MV,CMD): sc.row_dimensions[r].hidden=True
# conditional formatting: best in each row → green; EBITDA red/green font
for rr in (NET,GMr,CMr,EB,EBP):
    sc.conditional_formatting.add(f"C{rr}:{LAST}{rr}", FormulaRule(formula=[f"C{rr}=MAX($C{rr}:${LAST}{rr})"], fill=fill(POSBG)))
sc.conditional_formatting.add(f"C{BE}:{LAST}{BE}", FormulaRule(formula=[f"C{BE}=MIN($C{BE}:${LAST}{BE})"], fill=fill(POSBG)))
for rr in (EB,EBP):
    sc.conditional_formatting.add(f"C{rr}:{LAST}{rr}", CellIsRule(operator="lessThan",formula=["0"],font=Font(name="Calibri",bold=True,color="C00000")))
    sc.conditional_formatting.add(f"C{rr}:{LAST}{rr}", CellIsRule(operator="greaterThanOrEqual",formula=["0"],font=Font(name="Calibri",bold=True,color="1F9E6E")))
chart=BarChart(); chart.type="col"; chart.title="EBITDA by scenario"; chart.legend=None; chart.height=7.5; chart.width=18
chart.add_data(Reference(sc,min_col=3,max_col=2+NCOL,min_row=EB,max_row=EB),from_rows=True,titles_from_data=False)
chart.set_categories(Reference(sc,min_col=3,max_col=2+NCOL,min_row=HR,max_row=HR))
sc.add_chart(chart,"B30")
lbl(sc,"B29","Edit any yellow dial and the whole column recomputes. To add a scenario, copy column H and edit its dials.",color=GREY)
sc["B29"].font=font(9,False,GREY,it=True)

# ============================================================ 6. Guide
gd=wb.create_sheet("Guide"); gd.sheet_properties.tabColor=NAVY; gd.sheet_view.showGridLines=False
gd.column_dimensions["A"].width=2.5; gd.column_dimensions["B"].width=112
title(gd,"Guide — reading a CPG P&L",2)
G=[("","" ),
 ("The waterfall","h"),
 ("Gross Sales − Trade − Returns = NET SALES","m"),
 ("Net Sales − COGS = GROSS PROFIT   (Gross Margin %)","m"),
 ("Gross Profit − fulfillment − fees − channel ad − slotting = CONTRIBUTION MARGIN  (CM %)","m"),
 ("Contribution − brand marketing − sales − G&A − R&D = EBITDA   (EBITDA %)","m"),
 ("",""),
 ("Why contribution margin is the number to watch","h"),
 ("Gross margin flatters DTC and Amazon because their shelf prices are high — but those channels","" ),
 ("bleed it back through referral fees, fulfillment and customer acquisition. Contribution margin is","" ),
 ("what each channel actually leaves to cover corporate overhead. Manage the mix on contribution.","" ),
 ("",""),
 ("Where the levers are (see the Sensitivity & tornado)","h"),
 ("• COGS — every cent off a can flows straight to EBITDA at full volume.","" ),
 ("• Price & trade — highest-leverage revenue dials, but trade is sticky once retailers expect it.","" ),
 ("• Mix — shifting volume to higher-contribution channels lifts EBITDA without touching a cost.","" ),
 ("• Volume vs fixed cost — overhead is fixed, so scale is its own lever (see break-even).","" ),
 ("",""),
 ("Sheet map","h"),
 ("P&L          – assumptions (yellow) + the blended statement, per-case and % of net sales.","" ),
 ("COGS_BOM     – bill of materials → landed cost per can; feeds every channel.","" ),
 ("Channels     – the revenue engine; one column block per route to market, totals feed the P&L.","" ),
 ("Sensitivity  – EBITDA % across a price × COGS grid, coloured as a heatmap.","" ),
 ("Scenario Lab – compare many full plans side by side; edit each column's dials, best in row turns green.","" ),
 ("",""),
 ("Defaults model Organika MÜV Sparkling Electrolytes (355 ml can, zero-sugar; magnesium bisglycinate","i"),
 ("+ Fibersol prebiotic fiber) as a representative premium functional beverage. Replace with your actuals.","i"),
]
r=4
for text,kind in G:
    c=gd.cell(row=r,column=2,value=text)
    if kind=="h": sty(c,f=font(12,True,TEAL))
    elif kind=="m": sty(c,f=Font(name="Consolas",size=10,color="11302C"),fl=fill(TEALBG))
    elif kind=="i": sty(c,f=font(10,False,GREY,it=True))
    else: sty(c,f=font(11,False,"333333"))
    r+=1

wb.active=0
OUT="organika-cpg-pnl.xlsx"; wb.save(OUT)
print(f"Wrote {OUT} with sheets: {', '.join(s.title for s in wb.worksheets)}")
