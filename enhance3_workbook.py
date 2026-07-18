#!/usr/bin/env python3
"""Final polish: clickable nav hyperlinks on Contents, interactivity guide on Cover,
'back to contents' links, metadata refresh."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
PATH="/home/user/my-first-project/Organika_Sparkling_Competitor_Intelligence_ENTERPRISE.xlsx"
wb=load_workbook(PATH)
LINKBLUE="1155CC"; AMBER="FFF2CC"; AMBERHEAD="B45309"; GOLD="C9A227"
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top")

# --- Contents: make tab cells clickable internal links ---
toc=wb["01 Contents"]
for r in range(1,toc.max_row+1):
    cell=toc.cell(row=r,column=1)
    name=cell.value
    if name and name in wb.sheetnames and name!="Tab":
        cell.hyperlink=f"#'{name}'!A1"
        cell.font=Font(name="Calibri",bold=True,color=LINKBLUE,underline="single",size=10)

# --- 'Back to Contents' link in a fixed corner of each content sheet (top-right cell) ---
for name in wb.sheetnames:
    if name in ("00 Cover","01 Contents"): continue
    ws=wb[name]
    # place a small link in a far cell that won't collide (column N, row 1)
    c=ws.cell(row=1,column=14,value="↩ Contents")
    c.hyperlink="#'01 Contents'!A1"
    c.font=Font(name="Calibri",bold=True,color=LINKBLUE,underline="single",size=8)
    c.alignment=Alignment(horizontal="right")

# --- Cover: add an interactivity guide + jump link ---
cov=wb["00 Cover"]
# find first empty-ish row near bottom
r=cov.max_row+2
cov.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
c=cov.cell(row=r,column=2,value="INTERACTIVE TABS (gold) — built to be used, not just read:")
c.font=Font(name="Calibri",bold=True,color=AMBERHEAD,size=10); c.fill=fill(AMBER); c.alignment=WRAP; c.border=BORDER
cov.row_dimensions[r].height=18
guide=[
 ("Dashboard","5 live charts + KPI strip — the one-screen snapshot"),
 ("Financial Model","Pick a scenario (Conservative/Base/Aggressive) and the whole P&L, break-even & sensitivity heatmap recompute"),
 ("Go-NoGo Decision","Edit weights/scores; the verdict (GO / GO-conditional / NO-GO) auto-computes"),
 ("Battlecards","One quick-reference block per key competitor"),
 ("01 Contents","Click any tab name to jump straight to it"),
]
r+=1
for tab,desc in guide:
    link=cov.cell(row=r,column=2,value="▸ "+tab)
    link.hyperlink=f"#'{tab}'!A1"; link.font=Font(name="Calibri",bold=True,color=LINKBLUE,underline="single",size=10)
    cov.merge_cells(start_row=r,start_column=3,end_row=r,end_column=8)
    d=cov.cell(row=r,column=3,value=desc); d.font=Font(name="Calibri",size=10); d.alignment=WRAP
    cov.row_dimensions[r].height=16; r+=1

wb.properties.title="Functional & Wellness Sparkling — Competitor Intelligence & Market-Entry Dossier (MUV)"
wb.properties.description="30-tab analyst dossier with interactive scenario P&L, dashboard, Go/No-Go and battlecards. Confidence-tagged, fully cited."
wb.save(PATH)
print("polished & saved. sheets:",len(wb.sheetnames))
PY=None
