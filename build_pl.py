#!/usr/bin/env python3
"""Build the Organika RTD P&L workbook: 3 SKUs x 7 channels, single Aug 1 - Nov 1 2026 period."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.shapes import GraphicalProperties

# ---- palette / styles (matched to the original template) ----
NAVY   = "FF1F3864"
MIDBLU = "FF2E75B6"
LTBLU  = "FFEAF2FB"
BLUE   = "FF0000FF"   # input-cell font
BLACK  = "FF000000"   # formula-cell font
GREY   = "FF595959"
WHITE  = "FFFFFFFF"
INPUT_FILL = "FFFFF2CC"   # light amber highlight on every editable input
ACCENT = "FF0071E3"       # Apple-blue accent for steps / emphasis
INK    = "FF1D1D1F"       # near-black text
MIST   = "FFF5F5F7"       # soft light-grey panel
SOFT   = "FFF3F3F6"       # inherited / auto cell tint
BANDBG = "FFDCE6F4"       # soft-blue section divider (unified look)

CUR2 = '\\$#,##0.00_);"($"#,##0.00\\);\\-'   # per-case currency
CUR0 = '\\$#,##0_);"($"#,##0\\);\\-'          # total currency
INT  = '#,##0_);(#,##0\\);\\-'                 # case counts
PCT  = '0.0%'

def font(sz=10, bold=False, color=BLACK):
    return Font(name="Arial", size=sz, bold=bold, color=color)
def fill(c):
    return PatternFill("solid", fgColor=c)

CHANNELS = ["Shopify", "Amazon", "On-premise", "Retail", "Distributor", "Corporate Office",
            "Purity Life", "Natural & Specialty Direct", "FDM Direct", "Convenience", "Fitness Centre"]
CH_COLS  = [get_column_letter(2 + i) for i in range(len(CHANNELS))]   # B..H
TOT_COL  = get_column_letter(2 + len(CHANNELS))                        # I
LAST_COL = TOT_COL
FIRST, LAST = CH_COLS[0], CH_COLS[-1]                                  # B, H

# row -> (label, kind) ; kind drives number format
ROWS = {
    9:  ("Physical Cases",            "int"),
    10: ("Gross Revenue",             "cur0"),
    11: ("Discounts",                 "cur0"),
    12: ("Net Sales",                 "cur0"),
    13: ("    COGS Direct",           "cur0"),
    14: ("    COGS Indirect",         "cur0"),
    15: ("Cost of Sales",             "cur0"),
    16: ("Gross Profit",              "cur0"),
    17: ("GP %",                      "pct"),
    19: ("    A&P Consumer",          "cur0"),
    20: ("    A&P Trade",             "cur0"),
    21: ("    A&P Sales",             "cur0"),
    22: ("A&P",                       "cur0"),
    24: ("Contrib After A&P (CAAP)",  "cur0"),
    26: ("Gross Rev - per case",      "cur2"),
    27: ("Discounts - per case",      "cur2"),
    28: ("Net Sales - per case",      "cur2"),
    29: ("    COGS Direct - per case","cur2"),
    30: ("    COGS Indirect - per case","cur2"),
    31: ("Cost of Sales - per case",  "cur2"),
    32: ("Gross Profit - per case",   "cur2"),
    33: ("GP %",                      "pct"),
    34: ("A&P - per case",            "cur2"),
    35: ("CAAP - per case",           "cur2"),
}
FMT = {"int": INT, "cur0": CUR0, "cur2": CUR2, "pct": PCT}
SPACERS = {3: 7.5, 5: 6.0, 7: 7.5, 18: 6.0, 23: 6.0, 25: 6.0}
PCT_ROWS = (17, 33)
# input rows (blue) on SKU tabs, for channel columns B..H
# (COGS rows 29/30 are NOT inputs — they pull from the Cost Build-up tab)
INPUT_ROWS = (9, 19, 20, 21, 26, 27)

# Cost Build-up tab — output cells the SKU P&Ls reference for COGS per case
COST_SHEET = "Cost Build-up"
R_DIRECT, R_INDIRECT = 44, 45
COST_DIRECT_REF   = f"='{COST_SHEET}'!$D${R_DIRECT}"
COST_INDIRECT_REF = f"='{COST_SHEET}'!$D${R_INDIRECT}"
COST_TOTAL_REF    = f"'{COST_SHEET}'!$D${R_INDIRECT+1}"   # COGS total / case (no leading =)

# Assumptions tab — single control panel; SKU P&Ls reference these cells
ASSUMP = "Assumptions"
A_CANS, A_4PK = "$C$7", "$C$8"            # cans/case, 4-packs/case
A_MINGM, A_TGTGM = "$C$9", "$C$10"        # min / target GM%
A_SPLIT = ("$C$11", "$C$12", "$C$13")     # A&P split: consumer / trade / sales
A_TOTROW = {"Lime Lemon": 16, "Passion Fruit Pineapple": 17, "Raspberry": 18}
A_CH_ROW0 = 23                            # first channel row in the pricing table
TTL_SHEET = "TTL ORGANIKA RTD"


def sku_channel_formula(row, col, ci, sku_row):
    """Driver formulas for a channel column on a SKU sheet — sourced from Assumptions."""
    ch = A_CH_ROW0 + ci
    if row == 9:  return f"={ASSUMP}!$C${sku_row}*{ASSUMP}!$F${ch}"          # cases = total × mix%
    if row == 26: return f"={ASSUMP}!$B${ch}"                                # gross/case = CP
    if row == 27: return f"={ASSUMP}!$B${ch}*{ASSUMP}!$D${ch}"              # disc/case = CP × disc%
    if row == 22: return f"={col}12*{ASSUMP}!$E${ch}"                        # A&P = net × A&P%
    if row == 19: return f"={col}22*{ASSUMP}!{A_SPLIT[0]}"                   # A&P consumer
    if row == 20: return f"={col}22*{ASSUMP}!{A_SPLIT[1]}"                   # A&P trade
    if row == 21: return f"={col}22*{ASSUMP}!{A_SPLIT[2]}"                   # A&P sales
    return formula(row, col, False, None)


def build_sheet(ws, title, section, is_total, sku_sheets=None, sku_row=None):
    # column widths
    ws.column_dimensions["A"].width = 30
    for col in CH_COLS + [TOT_COL]:
        ws.column_dimensions[col].width = 13

    # row 1 title
    ws.merge_cells(f"A1:{LAST_COL}1")
    c = ws["A1"]; c.value = title
    c.font = font(13, True, WHITE); c.fill = fill(NAVY); c.alignment = Alignment("left", "center")
    ws.row_dimensions[1].height = 27.75

    # row 2 subtitle
    ws.merge_cells(f"A2:{LAST_COL}2")
    c = ws["A2"]; c.value = "CDN $ | Per case of 24 cans (6×4×355 ml) | Period: Aug 1 – Nov 1, 2026"
    c.font = font(9, False, GREY); c.alignment = Alignment("left", "center")
    ws.row_dimensions[2].height = 15

    # row 4 column headers
    a = ws["A4"]; a.value = "CDN $"; a.font = font(10, True, WHITE); a.fill = fill(MIDBLU)
    a.alignment = Alignment("left", "center", wrap_text=True)
    for col, name in zip(CH_COLS, CHANNELS):
        c = ws[f"{col}4"]; c.value = name; c.font = font(10, True, WHITE); c.fill = fill(MIDBLU)
        c.alignment = Alignment("center", "center", wrap_text=True)
    c = ws[f"{TOT_COL}4"]; c.value = "Total"; c.font = font(10, True, WHITE); c.fill = fill(NAVY)
    c.alignment = Alignment("center", "center", wrap_text=True)
    ws.row_dimensions[4].height = 45

    # row 6 legend
    ws.merge_cells(f"A6:{LAST_COL}6")
    legend = ("⚫ All values are formulas — edit inputs on the \U0001F7E1 Assumptions tab (pricing, mix, A&P)"
              if not is_total else "⚫ Roll-up of the three SKU tabs — all formulas")
    c = ws["A6"]; c.value = legend
    c.font = font(8, False, GREY); c.alignment = Alignment("left", "center")
    ws.row_dimensions[6].height = 13.5

    # row 8 section header
    ws.merge_cells(f"A8:{LAST_COL}8")
    c = ws["A8"]; c.value = section; c.font = font(11, True, INK); c.fill = fill(BANDBG)
    c.alignment = Alignment("left", "center")
    ws.row_dimensions[8].height = 15

    # spacers
    for r, h in SPACERS.items():
        ws.row_dimensions[r].height = h

    # labels + values
    for row, (label, kind) in ROWS.items():
        ws.row_dimensions[row].height = 15
        la = ws[f"A{row}"]; la.value = label; la.font = font(10, False, BLACK)
        la.alignment = Alignment("left", "center")
        numfmt = FMT[kind]
        for col in CH_COLS + [TOT_COL]:
            cell = ws[f"{col}{row}"]
            cell.number_format = numfmt
            cell.alignment = Alignment("right", "center")
            cell.font = font(10, False, BLACK)
            if is_total or col == TOT_COL:
                cell.value = formula(row, col, is_total, sku_sheets)
            else:
                cell.value = sku_channel_formula(row, col, CH_COLS.index(col), sku_row)
        # pct band fill across the whole row
        if row in PCT_ROWS:
            for col in ["A"] + CH_COLS + [TOT_COL]:
                ws[f"{col}{row}"].fill = fill(LTBLU)

    # note on the Total cell
    tgt = ("Sum of the 3 SKU tabs" if is_total
           else "= Assumptions total cases × channel mix% (edit on Assumptions)")
    ws[f"{TOT_COL}9"].comment = Comment(tgt, "Model")

    ws.freeze_panes = "B9"
    ws.sheet_view.showGridLines = False


def formula(row, col, is_total, sku_sheets):
    """Return the formula/value for a computed cell (col is a channel col B..H or the Total col)."""
    is_tot_col = (col == TOT_COL)

    if is_total:
        # ----- TTL roll-up sheet -----
        sumsheets = "+".join(f"'{s}'!{col}{row}" for s in sku_sheets)
        if is_tot_col:
            # Total column: $ rows = sum across channels; per-case = blended; pct = ratio
            if row in (17, 33):
                num, den = (16, 12) if row == 17 else (32, 28)
                return f'=IFERROR({TOT_COL}{num}/{TOT_COL}{den},"-")'
            if row in PC_BLENDED:                       # per-case blended from totals
                return PC_BLENDED[row](TOT_COL)
            return f"=SUM({FIRST}{row}:{LAST}{row})"
        else:
            # channel column on TTL: $/case rows = cross-sheet sum; pct & per-case = derived here
            if row in (17, 33):
                num, den = (16, 12) if row == 17 else (32, 28)
                return f'=IFERROR({col}{num}/{col}{den},"-")'
            if row in PC_BLENDED:
                return PC_BLENDED[row](col)
            return f"={sumsheets}"

    # ----- SKU sheet -----
    if is_tot_col:
        if row == 9:
            return f"=SUM({FIRST}9:{LAST}9)"
        if row in (10, 11, 13, 14, 19, 20, 21):
            return f"=SUM({FIRST}{row}:{LAST}{row})"
        if row == 12: return f"={TOT_COL}10-{TOT_COL}11"
        if row == 15: return f"={TOT_COL}13+{TOT_COL}14"
        if row == 16: return f"={TOT_COL}12-{TOT_COL}15"
        if row == 17: return f'=IFERROR({TOT_COL}16/{TOT_COL}12,"-")'
        if row == 22: return f"={TOT_COL}19+{TOT_COL}20+{TOT_COL}21"
        if row == 24: return f"={TOT_COL}16-{TOT_COL}22"
        if row in PC_BLENDED: return PC_BLENDED[row](TOT_COL)
        if row == 33: return f'=IFERROR({TOT_COL}32/{TOT_COL}28,"-")'
        if row == 35: return f"={TOT_COL}32-{TOT_COL}34"
    else:
        # channel columns on a SKU sheet
        if row == 29: return COST_DIRECT_REF     # COGS direct / case  ← Cost Build-up
        if row == 30: return COST_INDIRECT_REF   # COGS indirect / case ← Cost Build-up
        if row == 10: return f"={col}26*{col}9"
        if row == 11: return f"={col}27*{col}9"
        if row == 12: return f"={col}10-{col}11"
        if row == 13: return f"={col}29*{col}9"
        if row == 14: return f"={col}30*{col}9"
        if row == 15: return f"={col}13+{col}14"
        if row == 16: return f"={col}12-{col}15"
        if row == 17: return f'=IFERROR({col}16/{col}12,"-")'
        if row == 22: return f"={col}19+{col}20+{col}21"
        if row == 24: return f"={col}16-{col}22"
        if row == 28: return f"={col}26-{col}27"
        if row == 31: return f"={col}29+{col}30"
        if row == 32: return f"={col}28-{col}31"
        if row == 33: return f'=IFERROR({col}32/{col}28,"-")'
        if row == 34: return f'=IFERROR({col}22/{col}9,"-")'
        if row == 35: return f"={col}32-{col}34"
    return None


# per-case rows that are blended weighted-averages (total$/cases) on Total cols & TTL sheet
PC_BLENDED = {
    26: lambda c: f'=IFERROR({c}10/{c}9,"-")',
    27: lambda c: f'=IFERROR({c}11/{c}9,"-")',
    28: lambda c: f'=IFERROR({c}12/{c}9,"-")',
    29: lambda c: f'=IFERROR({c}13/{c}9,"-")',
    30: lambda c: f'=IFERROR({c}14/{c}9,"-")',
    31: lambda c: f'=IFERROR({c}15/{c}9,"-")',
    32: lambda c: f'=IFERROR({c}16/{c}9,"-")',
    34: lambda c: f'=IFERROR({c}22/{c}9,"-")',
    35: lambda c: f'=IFERROR({c}32-{c}34,0)',
}

def build_dashboard(ws, sku_sheets):
    """Front-page dashboard: total P&L, blended per-case, by-SKU and by-channel matrices."""
    ttl = "TTL ORGANIKA RTD"
    widths = {"A": 26, "B": 15, "C": 15, "D": 16, "E": 13, "F": 14, "G": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    def band(coord_range, text):
        first = coord_range.split(":")[0]
        ws.merge_cells(coord_range)
        c = ws[first]; c.value = text; c.font = font(10, True, INK); c.fill = fill(BANDBG)
        c.alignment = Alignment("left", "center")

    def kpi(r, lab_col, lab, val_col, ref, kind):
        ws[f"{lab_col}{r}"].value = lab
        ws[f"{lab_col}{r}"].font = font(10, False, BLACK)
        ws[f"{lab_col}{r}"].alignment = Alignment("left", "center")
        v = ws[f"{val_col}{r}"]; v.value = ref; v.font = font(10, False, BLACK)
        v.number_format = FMT[kind]; v.alignment = Alignment("right", "center")

    # title + subtitle
    ws.merge_cells("A1:G1")
    c = ws["A1"]; c.value = "ORGANIKA RTD — DASHBOARD"
    c.font = font(13, True, WHITE); c.fill = fill(NAVY); c.alignment = Alignment("left", "center")
    ws.row_dimensions[1].height = 27.75
    ws.merge_cells("A2:G2")
    c = ws["A2"]; c.value = ("CDN $ | 6×4×355 ml (24 cans/case) | Aug 1 – Nov 1, 2026 | "
                             f"{len(sku_sheets)} SKUs × {len(CHANNELS)} channels")
    c.font = font(9, False, GREY); c.alignment = Alignment("left", "center")
    ws.row_dimensions[3].height = 6

    # ---- block headers ----
    band("A4:B4", "TOTAL P&L — ALL SKUs × ALL CHANNELS")
    band("D4:E4", "BLENDED — PER CASE")

    left = [("Physical Cases", 9, "int"), ("Gross Revenue", 10, "cur0"), ("Discounts", 11, "cur0"),
            ("Net Sales", 12, "cur0"), ("Cost of Sales", 15, "cur0"), ("Gross Profit", 16, "cur0"),
            ("GP %", 17, "pct"), ("A&P", 22, "cur0"), ("CAAP", 24, "cur0")]
    right = [("Gross Rev / case", 26, "cur2"), ("Discounts / case", 27, "cur2"),
             ("Net Sales / case", 28, "cur2"), ("COGS / case", 31, "cur2"),
             ("Gross Profit / case", 32, "cur2"), ("GP %", 33, "pct"),
             ("A&P / case", 34, "cur2"), ("CAAP / case", 35, "cur2")]
    for i, (lab, r, kind) in enumerate(left):
        kpi(5 + i, "A", lab, "B", f"='{ttl}'!{TOT_COL}{r}", kind)
    for i, (lab, r, kind) in enumerate(right):
        kpi(5 + i, "D", lab, "E", f"='{ttl}'!{TOT_COL}{r}", kind)
    for r in (11, 17):  # GP% bands
        for col in ("A", "B"):
            if r == 11: ws[f"{col}{r}"].fill = fill(LTBLU)
    ws["A11"].fill = fill(LTBLU); ws["B11"].fill = fill(LTBLU)   # left GP%
    ws["D10"].fill = fill(LTBLU); ws["E10"].fill = fill(LTBLU)   # right GP%

    # ---- BY SKU matrix ----
    hdr = ["", "Cases", "Net Sales", "Gross Profit", "GP %", "A&P", "CAAP"]
    metrics = [("int", 9), ("cur0", 12), ("cur0", 16), ("pct", 17), ("cur0", 22), ("cur0", 24)]
    band("A15:G15", "BY SKU")
    ws["A16"].value = "SKU"
    for j, h in enumerate(hdr[1:], start=1):
        cc = ws.cell(row=16, column=1 + j, value=h)
        cc.font = font(10, True, WHITE); cc.fill = fill(MIDBLU)
        cc.alignment = Alignment("center", "center")
    ws["A16"].font = font(10, True, WHITE); ws["A16"].fill = fill(MIDBLU)
    ws["A16"].alignment = Alignment("left", "center")
    for i, s in enumerate(sku_sheets):
        r = 17 + i
        ws[f"A{r}"].value = s; ws[f"A{r}"].font = font(10, False, BLACK)
        for j, (kind, src) in enumerate(metrics, start=1):
            cell = ws.cell(row=r, column=1 + j, value=f"='{s}'!{TOT_COL}{src}")
            cell.number_format = FMT[kind]; cell.font = font(10, False, BLACK)
            cell.alignment = Alignment("right", "center")
    rt = 17 + len(sku_sheets)
    ws[f"A{rt}"].value = "TOTAL"; ws[f"A{rt}"].font = font(10, True, BLACK)
    for j, (kind, src) in enumerate(metrics, start=1):
        cell = ws.cell(row=rt, column=1 + j, value=f"='{ttl}'!{TOT_COL}{src}")
        cell.number_format = FMT[kind]; cell.font = font(10, True, BLACK)
        cell.alignment = Alignment("right", "center")

    # ---- BY CHANNEL matrix ----
    base = rt + 2
    band(f"A{base}:G{base}", "BY CHANNEL")
    hrow = base + 1
    ws[f"A{hrow}"].value = "Channel"; ws[f"A{hrow}"].font = font(10, True, WHITE)
    ws[f"A{hrow}"].fill = fill(MIDBLU); ws[f"A{hrow}"].alignment = Alignment("left", "center")
    for j, h in enumerate(hdr[1:], start=1):
        cc = ws.cell(row=hrow, column=1 + j, value=h)
        cc.font = font(10, True, WHITE); cc.fill = fill(MIDBLU)
        cc.alignment = Alignment("center", "center")
    for i, (name, col) in enumerate(zip(CHANNELS, CH_COLS)):
        r = hrow + 1 + i
        ws[f"A{r}"].value = name; ws[f"A{r}"].font = font(10, False, BLACK)
        ws[f"A{r}"].alignment = Alignment("left", "center")
        for j, (kind, src) in enumerate(metrics, start=1):
            cell = ws.cell(row=r, column=1 + j, value=f"='{ttl}'!{col}{src}")
            cell.number_format = FMT[kind]; cell.font = font(10, False, BLACK)
            cell.alignment = Alignment("right", "center")
    rc = hrow + 1 + len(CHANNELS)
    ws[f"A{rc}"].value = "TOTAL"; ws[f"A{rc}"].font = font(10, True, BLACK)
    for j, (kind, src) in enumerate(metrics, start=1):
        cell = ws.cell(row=rc, column=1 + j, value=f"='{ttl}'!{TOT_COL}{src}")
        cell.number_format = FMT[kind]; cell.font = font(10, True, BLACK)
        cell.alignment = Alignment("right", "center")

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


def build_cost_buildup(ws):
    """Bottom-up COGS from the Bevmax co-packing quote. Outputs feed the SKU P&Ls."""
    for col, w in {"A": 34, "B": 12, "C": 12, "D": 13, "E": 11}.items():
        ws.column_dimensions[col].width = w
    BL, BK = font(10, False, BLUE), font(10, False, BLACK)

    def band(rng, text):
        first = rng.split(":")[0]; ws.merge_cells(rng)
        c = ws[first]; c.value = text; c.font = font(10, True, INK); c.fill = fill(BANDBG)
        c.alignment = Alignment("left", "center")

    def put(cell, val, *, inp=False, fmt=None, bold=False, align="right"):
        c = ws[cell]; c.value = val
        c.font = Font(name="Arial", size=10, bold=bold, color=(BLUE if inp else BLACK))
        if inp: c.fill = fill(INPUT_FILL)
        if fmt: c.number_format = fmt
        c.alignment = Alignment(align, "center")

    def label(row, text, bold=False, indent=False):
        c = ws[f"A{row}"]; c.value = ("    " + text) if indent else text
        c.font = font(10, bold, BLACK); c.alignment = Alignment("left", "center")

    ws.merge_cells("A1:E1")
    t = ws["A1"]; t.value = "ORGANIKA RTD — BOTTOM-UP COST (Bevmax quote)"
    t.font = font(13, True, WHITE); t.fill = fill(NAVY); t.alignment = Alignment("left", "center")
    ws.row_dimensions[1].height = 27.75
    ws.merge_cells("A2:E2")
    s = ws["A2"]; s.value = ("CDN $ | Quote OG2026-05-21 v15 (4-pk w/ sleeves) | Pack Summer 2026 | "
                             "FOB Airdrie AB | excl. GST")
    s.font = font(9, False, GREY); s.alignment = Alignment("left", "center")

    # run parameters
    band("A4:E4", "RUN PARAMETERS")
    label(6, "Total run size");      put("D6", 40500, inp=True, fmt=INT);  ws["B6"]="cans"
    label(7, "Number of SKUs");      put("D7", 3,     inp=True, fmt=INT);  ws["B7"]="SKUs"
    label(8, "Cans per SKU");        put("D8", "=D6/D7", fmt=INT);          ws["B8"]="cans"
    label(9, "Cans per case");       put("D9", 24,    inp=True, fmt=INT);  ws["B9"]="cans"
    label(10, "Cans per 4-pack");    put("D10", 4,    inp=True, fmt=INT);  ws["B10"]="cans"
    label(11, "Cases per run");      put("D11", "=D6/D9", fmt='#,##0.0');   ws["B11"]="cases"
    for r in (6,7,8,9,10,11):
        ws[f"B{r}"].font = font(9, False, GREY)

    # cost lines (Total $ per quote; $/can = Total / run cans)
    band("A13:E13", "BEVMAX QUOTE — COST LINES")
    for col, h in zip("ABCDE", ["Cost line","UofM","Rate $","Total $","$/can"]):
        c = ws[f"{col}14"]; c.value = h; c.font = font(9, True, WHITE); c.fill = fill(MIDBLU)
        c.alignment = Alignment("left" if col=="A" else "center", "center")
    lines = [  # label, uom, rate, total (from quote)
        ("Tolling (blend/fill/pasteurize/carbonate)", "per can", 0.300, 12150.00),
        ("4-packing", "per can", 0.050, 2025.00),
        ("Skid packing", "per run", None, 439.45),
        ("Ingredients", "per can", 0.160, 6480.00),
        ("Brite can + 202 LOE end", "per can", 0.320, 12960.00),
        ("Sleeves + application", "per label", 0.163, 6601.50),
        ("Sleeve-perforation setup (one-time)", "per setup", None, 489.00),
        ("4-pk printed carton", "per carton", 0.411, 6165.00),
        ("24-pk white tray", "per tray", 0.860, 1451.25),
        ("Grip sheets", "per tray", 1.250, 175.78),
        ("Tray labels", "per tray", 0.050, 168.75),
    ]
    r = 15
    for lab, uom, rate, total in lines:
        label(r, lab, indent=True)
        ws[f"B{r}"].value = uom; ws[f"B{r}"].font = font(9, False, GREY)
        ws[f"B{r}"].alignment = Alignment("center", "center")
        if rate is not None: put(f"C{r}", rate, inp=True, fmt=CUR2)
        else: put(f"C{r}", "—", fmt='General', align="center")
        put(f"D{r}", total, inp=True, fmt=CUR2)
        put(f"E{r}", f"=D{r}/$D$6", fmt=CUR2)
        r += 1
    sub = r
    label(sub, "Subtotal — Bevmax (FOB Airdrie)", bold=True)
    put(f"D{sub}", f"=SUM(D15:D{sub-1})", fmt=CUR2, bold=True)
    put(f"E{sub}", f"=D{sub}/$D$6", fmt=CUR2, bold=True)

    # per-unit from Bevmax
    label(sub+2, "Bevmax cost / can");      put(f"D{sub+2}", f"=D{sub}/$D$6", fmt=CUR2)
    label(sub+3, "Bevmax cost / 4-pack");   put(f"D{sub+3}", f"=D{sub+2}*D10", fmt=CUR2)
    label(sub+4, "Bevmax cost / case (24)");put(f"D{sub+4}", f"=D{sub+2}*D9", fmt=CUR2, bold=True)
    bev_case = f"D{sub+4}"

    # adders -> landed direct
    band(f"A{sub+6}:E{sub+6}", "ADDERS TO LANDED COST (per case)")
    label(sub+7, "Allergen handling (amortized)")
    put(f"D{sub+7}", f"=167.5/D11", fmt=CUR2)
    label(sub+8, "Inbound freight to DC");  put(f"D{sub+8}", 0, inp=True, fmt=CUR2)
    label(sub+9, "Testing / misc");         put(f"D{sub+9}", 0, inp=True, fmt=CUR2)
    rdir = R_DIRECT  # 44
    # overhead
    band(f"A{sub+11}:E{sub+11}", "OVERHEAD (% of landed direct)")
    label(sub+12, "OH cost %");     put(f"C{sub+12}", 0.05, inp=True, fmt=PCT)
    label(sub+13, "Other COGS %");  put(f"C{sub+13}", 0.05, inp=True, fmt=PCT)

    # outputs to P&L  (D44 direct, D45 indirect)
    band(f"A{R_DIRECT-1}:E{R_DIRECT-1}", "→ COGS TO P&L (per case)")
    label(R_DIRECT, "COGS Direct / case", bold=True)
    put(f"D{R_DIRECT}", f"={bev_case}+D{sub+7}+D{sub+8}+D{sub+9}", fmt=CUR2, bold=True)
    label(R_INDIRECT, "COGS Indirect / case", bold=True)
    put(f"D{R_INDIRECT}", f"=(C{sub+12}+C{sub+13})*D{R_DIRECT}", fmt=CUR2, bold=True)
    label(R_INDIRECT+1, "COGS Total / case", bold=True)
    put(f"D{R_INDIRECT+1}", f"=D{R_DIRECT}+D{R_INDIRECT}", fmt=CUR2, bold=True)
    label(R_INDIRECT+2, "COGS / 4-pack");  put(f"D{R_INDIRECT+2}", f"=D{R_INDIRECT+1}/D9*D10", fmt=CUR2)
    label(R_INDIRECT+3, "COGS / can");     put(f"D{R_INDIRECT+3}", f"=D{R_INDIRECT+1}/D9", fmt=CUR2)
    for rr in (R_DIRECT, R_INDIRECT, R_INDIRECT+1):
        for cc in "ABCDE": ws[f"{cc}{rr}"].fill = fill(LTBLU)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


def build_blended(ws, sku_sheets):
    """Volume-weighted blended P&L across all channels — per SKU and total."""
    ttl = "TTL ORGANIKA RTD"
    ws.column_dimensions["A"].width = 30
    for col in "BCDE": ws.column_dimensions[col].width = 16

    ws.merge_cells("A1:E1")
    t = ws["A1"]; t.value = "ORGANIKA RTD — BLENDED (all channels)"
    t.font = font(13, True, WHITE); t.fill = fill(NAVY); t.alignment = Alignment("left", "center")
    ws.row_dimensions[1].height = 27.75
    ws.merge_cells("A2:E2")
    s = ws["A2"]; s.value = ("CDN $ | Volume-weighted average across all 11 channels | "
                             "per case = 24 cans | Aug 1 – Nov 1, 2026")
    s.font = font(9, False, GREY); s.alignment = Alignment("left", "center")

    # header
    cols = ["B","C","D","E"]
    names = sku_sheets + ["TOTAL"]
    for col, nm in zip(cols, names):
        c = ws[f"{col}4"]; c.value = nm; c.font = font(10, True, WHITE)
        c.fill = fill(NAVY if nm == "TOTAL" else MIDBLU)
        c.alignment = Alignment("center", "center", wrap_text=True)
    ws["A4"].value = "CDN $"; ws["A4"].font = font(10, True, WHITE); ws["A4"].fill = fill(MIDBLU)
    ws["A4"].alignment = Alignment("left", "center"); ws.row_dimensions[4].height = 28

    def section(row, text):
        ws.merge_cells(f"A{row}:E{row}")
        c = ws[f"A{row}"]; c.value = text; c.font = font(11, True, INK); c.fill = fill(BANDBG)
        c.alignment = Alignment("left", "center")

    def line(row, lab, src, kind, indent=False, band=False):
        a = ws[f"A{row}"]; a.value = ("    " + lab) if indent else lab
        a.font = font(10, False, BLACK); a.alignment = Alignment("left", "center")
        refs = [f"='{s}'!{TOT_COL}{src}" for s in sku_sheets] + [f"='{ttl}'!{TOT_COL}{src}"]
        for col, ref in zip(cols, refs):
            c = ws[f"{col}{row}"]; c.value = ref; c.font = font(10, False, BLACK)
            c.number_format = FMT[kind]; c.alignment = Alignment("right", "center")
        if band:
            for col in ["A"]+cols: ws[f"{col}{row}"].fill = fill(LTBLU)

    section(6, "PERIOD TOTALS (Aug 1 – Nov 1)")
    rows_tot = [("Physical Cases",9,"int"),("Gross Revenue",10,"cur0"),("Discounts",11,"cur0"),
                ("Net Sales",12,"cur0"),("COGS Direct",13,"cur0",True),("COGS Indirect",14,"cur0",True),
                ("Cost of Sales",15,"cur0"),("Gross Profit",16,"cur0"),("GP %",17,"pct",False,True),
                ("A&P",22,"cur0"),("Contrib After A&P (CAAP)",24,"cur0")]
    r = 7
    for spec in rows_tot:
        lab,src,kind = spec[0],spec[1],spec[2]
        indent = len(spec)>3 and spec[3]; band = len(spec)>4 and spec[4]
        line(r, lab, src, kind, indent=indent, band=band); r += 1

    r += 1
    section(r, "BLENDED PER CASE (avg across channels)"); r += 1
    rows_pc = [("Gross Rev / case",26,"cur2"),("Discounts / case",27,"cur2"),
               ("Net Sales / case",28,"cur2"),("COGS / case",31,"cur2"),
               ("Gross Profit / case",32,"cur2"),("GP %",33,"pct",False,True),
               ("A&P / case",34,"cur2"),("CAAP / case",35,"cur2")]
    for spec in rows_pc:
        lab,src,kind = spec[0],spec[1],spec[2]
        band = len(spec)>4 and spec[4]
        line(r, lab, src, kind, band=band); r += 1

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"


def build_scenarios(ws):
    """What-if sandbox: trial CP (wholesale) & MSRP points; compare scenarios; sensitivity grids."""
    cogs = COST_TOTAL_REF   # COGS total / case, from Cost Build-up
    ws.column_dimensions["A"].width = 30
    for col in "BCDEFGH": ws.column_dimensions[col].width = 13
    SCN = ["B","C","D","E","F"]

    def band(rng, text):
        first = rng.split(":")[0]; ws.merge_cells(rng)
        c = ws[first]; c.value = text; c.font = font(11, True, INK); c.fill = fill(BANDBG)
        c.alignment = Alignment("left", "center")
    def lab(r, text, bold=False):
        c = ws[f"A{r}"]; c.value = text; c.font = font(10, bold, BLACK)
        c.alignment = Alignment("left", "center")
    def put(cell, val, *, inp=False, fmt=None, bold=False, band_fill=False):
        c = ws[cell]; c.value = val
        c.font = Font(name="Arial", size=10, bold=bold, color=(BLUE if inp else BLACK))
        if inp: c.fill = fill(INPUT_FILL)
        elif band_fill: c.fill = fill(LTBLU)
        if fmt: c.number_format = fmt
        c.alignment = Alignment("right", "center")

    # title
    ws.merge_cells("A1:H1")
    t = ws["A1"]; t.value = "ORGANIKA RTD — SCENARIOS & PRICE SENSITIVITY"
    t.font = font(13, True, WHITE); t.fill = fill(NAVY); t.alignment = Alignment("left", "center")
    ws.row_dimensions[1].height = 27.75
    ws.merge_cells("A2:H2")
    s = ws["A2"]; s.value = ("CDN $ per case (24 cans = 6 × 4-pack) | 🟡 yellow = edit | "
                             f"COGS pulled live from Cost Build-up ({cogs.split('!')[1]})")
    s.font = font(9, False, GREY); s.alignment = Alignment("left", "center")

    # ---------- Section 1: scenario comparison ----------
    band("A4:F4", "SCENARIO COMPARISON  (CP = wholesale price, MSRP = shelf price)")
    names = ["Conservative", "Base", "Stretch", "Scenario 4", "Scenario 5"]
    lab(5, "Scenario")
    for col, nm in zip(SCN, names):
        c = ws[f"{col}5"]; c.value = nm; c.font = font(10, True, BLUE); c.fill = fill(INPUT_FILL)
        c.alignment = Alignment("center", "center")
    seed = {  # col -> (CP, MSRP, promo, volume, A&P/case)
        "B": (54, 84, 0, 500, 0), "C": (60, 96, 0, 500, 0), "D": (66, 108, 0, 500, 0)}
    inputs = [(6,"CP — wholesale ($/case)",CUR2,0),(7,"MSRP — retail ($/case)",CUR2,1),
              (8,"Promo / discount (% off CP)",PCT,2),(9,"Volume (cases)",INT,3),
              (10,"A&P ($/case)",CUR2,4)]
    for r, text, fmt, idx in inputs:
        lab(r, text)
        for col in SCN:
            v = seed.get(col, (None,)*5)[idx]
            put(f"{col}{r}", v, inp=True, fmt=fmt)
    calcs = [
        (11, "Net price ($/case)",   lambda c: f"={c}6*(1-{c}8)", CUR2, False),
        (12, "COGS ($/case)",        lambda c: f"={cogs}",        CUR2, False),
        (13, "Organika GP ($/case)", lambda c: f"={c}11-{c}12",   CUR2, False),
        (14, "Organika GP %",        lambda c: f'=IFERROR({c}13/{c}11,"-")', PCT, True),
        (15, "Customer margin %",    lambda c: f'=IFERROR(({c}7-{c}6)/{c}7,"-")', PCT, True),
        (16, "CAAP ($/case)",        lambda c: f"={c}13-{c}10",   CUR2, False),
        (18, "CP / 4-pack",          lambda c: f"={c}6/6",        CUR2, False),
        (19, "MSRP / 4-pack",        lambda c: f"={c}7/6",        CUR2, False),
        (20, "Net sales — period ($)",lambda c: f"={c}11*{c}9",   CUR0, False),
        (21, "Gross profit — period ($)",lambda c: f"={c}13*{c}9",CUR0, False),
        (22, "CAAP — period ($)",    lambda c: f"={c}16*{c}9",    CUR0, False),
    ]
    for r, text, fn, fmt, bandf in calcs:
        lab(r, text, bold=(r in (14,15)))
        if bandf: ws[f"A{r}"].fill = fill(LTBLU)
        for col in SCN:
            put(f"{col}{r}", fn(col), fmt=fmt, band_fill=bandf)
    ws["A17"].value = "— per-pack & period —"; ws["A17"].font = font(8, False, GREY)
    # heatmap on the two margin rows
    for rng in ("B14:F14", "B15:F15"):
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',
            mid_type='num', mid_value=0.35, mid_color='FFEB84',
            end_type='num', end_value=0.7, end_color='63BE7B'))

    # ---------- Section 2: CP sensitivity ----------
    band("A24:G24", "PRICE SENSITIVITY — Organika margin as CP varies")
    for a, t1, b, val, fmt in [("A25","CP start ($/case)","B25",48,CUR2),
                               ("C25","step ($)","D25",4,CUR2),
                               ("E25","ref MSRP ($/case)","F25",96,CUR2),
                               ("G25","promo %","H25",0,PCT)]:
        ws[a].value = t1; ws[a].font = font(9, False, GREY); ws[a].alignment = Alignment("right","center")
        put(b, val, inp=True, fmt=fmt)
    hdr2 = ["CP /case","CP /4-pack","Net /case","COGS /case","GP /case","GP %","Cust margin %"]
    for col, h in zip("ABCDEFG", hdr2):
        c = ws[f"{col}26"]; c.value = h; c.font = font(9, True, WHITE); c.fill = fill(MIDBLU)
        c.alignment = Alignment("center", "center", wrap_text=True)
    for i in range(10):
        r = 27 + i
        put(f"A{r}", f"=$B$25+{i}*$D$25", fmt=CUR2)
        put(f"B{r}", f"=A{r}/6", fmt=CUR2)
        put(f"C{r}", f"=A{r}*(1-$H$25)", fmt=CUR2)
        put(f"D{r}", f"={cogs}", fmt=CUR2)
        put(f"E{r}", f"=C{r}-D{r}", fmt=CUR2)
        put(f"F{r}", f'=IFERROR(E{r}/C{r},"-")', fmt=PCT)
        put(f"G{r}", f'=IFERROR(($F$25-A{r})/$F$25,"-")', fmt=PCT)
    for rng in ("F27:F36", "G27:G36"):
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',
            mid_type='num', mid_value=0.35, mid_color='FFEB84',
            end_type='num', end_value=0.7, end_color='63BE7B'))

    # ---------- Section 3: CP x MSRP customer-margin grid ----------
    band("A39:H39", "CUSTOMER (RETAILER) MARGIN %  —  CP (down) × MSRP (across)")
    ws["A40"].value = "CP ↓  /  MSRP →"; ws["A40"].font = font(9, True, BLACK)
    ws["A40"].alignment = Alignment("left", "center")
    msrp_cols = "BCDEFGH"
    for j, col in enumerate(msrp_cols):
        put(f"{col}40", 84 + j*6, inp=True, fmt=CUR2)   # MSRP header points (editable)
    for i in range(8):
        r = 41 + i
        put(f"A{r}", f"=$B$25+{i}*$D$25", fmt=CUR2)      # CP ladder (synced to sensitivity)
        for col in msrp_cols:
            put(f"{col}{r}", f'=IFERROR(({col}$40-$A{r})/{col}$40,"-")', fmt=PCT)
    ws.conditional_formatting.add("B41:H48", ColorScaleRule(
        start_type='num', start_value=0, start_color='F8696B',
        mid_type='num', mid_value=0.3, mid_color='FFEB84',
        end_type='num', end_value=0.5, end_color='63BE7B'))

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"


def build_assumptions(ws, sku_sheets):
    """Single control panel: global drivers, per-flavour volume, per-channel pricing/mix, checks."""
    for col, w in {"A":30,"B":13,"C":13,"D":11,"E":11,"F":10,"G":12,"H":13,"I":14}.items():
        ws.column_dimensions[col].width = w
    GREEN = PatternFill("solid", fgColor="C6EFCE"); RED = PatternFill("solid", fgColor="FFC7CE")

    def band(rng, text):
        first = rng.split(":")[0]; ws.merge_cells(rng)
        c = ws[first]; c.value = text; c.font = font(11, True, INK); c.fill = fill(BANDBG)
        c.alignment = Alignment("left", "center")
    def lab(r, text, col="A", bold=False):
        c = ws[f"{col}{r}"]; c.value = text; c.font = font(10, bold, BLACK)
        c.alignment = Alignment("left", "center")
    def inp(cell, val, fmt):
        c = ws[cell]; c.value = val; c.font = font(10, False, BLUE); c.fill = fill(INPUT_FILL)
        c.number_format = fmt; c.alignment = Alignment("right", "center")
    def calc(cell, val, fmt, bold=False):
        c = ws[cell]; c.value = val; c.font = font(10, bold, BLACK)
        c.number_format = fmt; c.alignment = Alignment("right", "center")

    ws.merge_cells("A1:I1")
    t = ws["A1"]; t.value = "ORGANIKA RTD — ASSUMPTIONS (control panel)"
    t.font = font(13, True, WHITE); t.fill = fill(NAVY); t.alignment = Alignment("left","center")
    ws.row_dimensions[1].height = 27.75
    ws.merge_cells("A2:I2")
    s = ws["A2"]; s.value = "🟡 Type in yellow cells. Start with the Master price (top-right) — that's all you need. CDN $ per case (24 cans)."
    s.font = font(9, False, GREY); s.alignment = Alignment("left","center")

    # QUICK START — one price drives every channel
    band("E4:I4", "▶  QUICK START — one price for every channel")
    for rr,(lt,val,fmt) in [(5,("Master CP  ($/case)",72,CUR2)),(6,("Master MSRP  ($/case)",108,CUR2)),
                            (7,("Master discount  (%)",0,PCT)),(8,("Master A&P  (% of net)",0.08,PCT))]:
        ws.merge_cells(f"E{rr}:F{rr}")
        c=ws[f"E{rr}"]; c.value=lt; c.font=font(10,False,INK); c.alignment=Alignment("left","center")
        inp(f"G{rr}",val,fmt)
    ws.merge_cells("E9:I9")
    nn=ws["E9"]; nn.value="↳ Channels inherit these. Type over any single channel below to override it."
    nn.font=font(8,False,GREY); nn.alignment=Alignment("left","center")

    band("A4:C4", "GLOBAL")
    lab(5,"Period start");        ws["C5"]="Aug 1, 2026"; ws["C5"].font=font(10,False,BLUE); ws["C5"].fill=fill(INPUT_FILL); ws["C5"].alignment=Alignment("right","center")
    lab(6,"Period end");          ws["C6"]="Nov 1, 2026"; ws["C6"].font=font(10,False,BLUE); ws["C6"].fill=fill(INPUT_FILL); ws["C6"].alignment=Alignment("right","center")
    lab(7,"Cans per case");       inp("C7",24,INT)
    lab(8,"Four-packs per case"); inp("C8",6,INT)
    lab(9,"Minimum GM% (gate)");  inp("C9",0.50,PCT)
    lab(10,"Target GM%");         inp("C10",0.60,PCT)
    lab(11,"A&P split — Consumer %"); inp("C11",0.50,PCT)
    lab(12,"A&P split — Trade %");    inp("C12",0.30,PCT)
    lab(13,"A&P split — Sales %");    inp("C13",0.20,PCT)
    calc("E11","=C11+C12+C13",PCT); ws["D11"].value="sum →"; ws["D11"].font=font(8,False,GREY); ws["D11"].alignment=Alignment("right","center")

    band("A15:I15", "VOLUME — total cases per flavour (whole period)")
    for i,s2 in enumerate(sku_sheets):
        lab(16+i, s2); inp(f"C{16+i}", 500, INT)
    lab(19,"TOTAL cases", bold=True); calc("C19","=SUM(C16:C18)",INT,bold=True)

    band("A21:I21", "CHANNEL MIX  &  (optional) per-channel price overrides")
    heads = ["Channel","CP $/case","MSRP $/case","Disc %","A&P % net","Mix %  🟡","CP /4-pack","Cust margin %","Cases (all SKUs)"]
    for col,h in zip("ABCDEFGHI", heads):
        c = ws[f"{col}22"]; c.value=h; c.font=font(9,True,WHITE); c.fill=fill(MIDBLU)
        c.alignment=Alignment("left" if col=="A" else "center","center",wrap_text=True)
    ws.row_dimensions[22].height = 28
    mix = [0.12,0.10,0.06,0.18,0.20,0.02,0.08,0.10,0.08,0.04,0.02]   # sums to 1.00
    def softref(cell, ref, fmt):   # inherits the Master value; user can type over
        c=ws[cell]; c.value=ref; c.font=font(10,False,INK); c.fill=fill(SOFT)
        c.number_format=fmt; c.alignment=Alignment("right","center")
    for i,name in enumerate(CHANNELS):
        r = A_CH_ROW0 + i
        lab(r, name)
        softref(f"B{r}","=$G$5",CUR2); softref(f"C{r}","=$G$6",CUR2)
        softref(f"D{r}","=$G$7",PCT); softref(f"E{r}","=$G$8",PCT)
        inp(f"F{r}",mix[i],PCT)
        calc(f"G{r}",f"=B{r}/$C$8",CUR2)
        calc(f"H{r}",f'=IFERROR((C{r}-B{r})/C{r},"-")',PCT)
        calc(f"I{r}",f"=F{r}*$C$19",INT)
    rt = A_CH_ROW0 + len(CHANNELS)
    lab(rt,"TOTAL", bold=True)
    calc(f"F{rt}",f"=SUM(F{A_CH_ROW0}:F{rt-1})",PCT,bold=True)
    calc(f"I{rt}",f"=SUM(I{A_CH_ROW0}:I{rt-1})",INT,bold=True)
    ws.conditional_formatting.add(f"H{A_CH_ROW0}:H{rt-1}", ColorScaleRule(
        start_type='num',start_value=0,start_color='F8696B',
        mid_type='num',mid_value=0.3,mid_color='FFEB84',
        end_type='num',end_value=0.5,end_color='63BE7B'))

    cr = rt + 2
    band(f"A{cr}:I{cr}", "CHECKS")
    blended_gp = f"='{TTL_SHEET}'!M17"
    lab(cr+1,"Channel mix sums to 100%")
    ws[f"C{cr+1}"] = f'=IF(ABS(F{rt}-1)<0.005,"✓ 100%","✗ "&TEXT(F{rt},"0.0%"))'
    lab(cr+2,"Blended GP% (live)"); calc(f"C{cr+2}", blended_gp, PCT)
    lab(cr+3,"GP% ≥ minimum?")
    ws[f"C{cr+3}"] = f'=IF(C{cr+2}>=C9,"✓ ≥ "&TEXT(C9,"0%"),"✗ below min")'
    lab(cr+4,"GP% ≥ target?")
    ws[f"C{cr+4}"] = f'=IF(C{cr+2}>=C10,"✓ ≥ "&TEXT(C10,"0%"),"– below target")'
    lab(cr+5,"Total CAAP > 0?")
    ws[f"C{cr+5}"] = f'=IF(\'{TTL_SHEET}\'!M24>0,"✓","✗")'
    for r in range(cr+1, cr+6):
        ws[f"C{r}"].alignment = Alignment("right","center"); ws[f"C{r}"].font = font(10,True,BLACK)
    ws.conditional_formatting.add(f"C{cr+1}:C{cr+5}", FormulaRule(formula=[f'ISNUMBER(SEARCH("✓",C{cr+1}))'], fill=GREEN))
    ws.conditional_formatting.add(f"C{cr+1}:C{cr+5}", FormulaRule(formula=[f'ISNUMBER(SEARCH("✗",C{cr+1}))'], fill=RED))

    # data validation
    pct_dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    pos_dv = DataValidation(type="decimal", operator="greaterThan", formula1="0", allow_blank=True)
    ws.add_data_validation(pct_dv); ws.add_data_validation(pos_dv)
    for rng in ("C9","C10","C11:C13","G7","G8",f"D{A_CH_ROW0}:D{rt-1}",f"E{A_CH_ROW0}:E{rt-1}",f"F{A_CH_ROW0}:F{rt-1}"):
        pct_dv.add(rng)
    for rng in ("G5","G6",f"B{A_CH_ROW0}:B{rt-1}",f"C{A_CH_ROW0}:C{rt-1}","C16:C18"):
        pos_dv.add(rng)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


def build_cover(ws):
    """Start Here — a friendly, guided home screen. Three steps, what's inside, the numbers."""
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    for col in "CDEFGH": ws.column_dimensions[col].width = 13
    ws.sheet_view.showGridLines = False

    # hero
    ws.merge_cells("A1:H3")
    h = ws["A1"]; h.value = "Organika RTD — Sparkling Electrolytes"
    h.font = Font(name="Arial", size=22, bold=True, color=WHITE); h.fill = fill("FF0A2540")
    h.alignment = Alignment("left", "center", indent=1)
    for rr in (1,2,3): ws.row_dimensions[rr].height = 24
    ws.merge_cells("A4:H4")
    s = ws["A4"]; s.value = "   Launch business case  ·  Aug 1 – Nov 1, 2026  ·  CDN $  ·  v1.1"
    s.font = font(10, False, GREY); s.alignment = Alignment("left", "center")
    ws.merge_cells("A6:H6")
    p = ws["A6"]; p.value = "Everything updates automatically. You only do three things 👇"
    p.font = font(12, True, INK); p.alignment = Alignment("left", "center"); ws.row_dimensions[6].height = 22

    # three step cards
    steps = [
        ("1", "Set your price", "On the Assumptions tab, type one Master price (top-right) — every channel uses it. Want different prices per channel? Just type over any channel cell.", "FF0071E3"),
        ("2", "Set your channel mix", "On Assumptions, enter the % of cases each channel sells. The green check confirms it totals 100%.", "FF34C759"),
        ("3", "Read the answer", "Open Executive Summary: your margin, a clear go / no-go verdict, the price you'd need to hit target, and best / base / worst.", "FFFF9500"),
    ]
    rr = 8
    for num, title, desc, color in steps:
        ws.merge_cells(f"A{rr}:A{rr+1}")
        n = ws[f"A{rr}"]; n.value = num; n.font = Font(name="Arial", size=26, bold=True, color=WHITE)
        n.fill = fill(color); n.alignment = Alignment("center", "center")
        tt = ws[f"B{rr}"]; tt.value = title; tt.font = font(13, True, INK); tt.alignment = Alignment("left","bottom")
        ws.merge_cells(f"B{rr+1}:H{rr+1}")
        d = ws[f"B{rr+1}"]; d.value = desc; d.font = font(10, False, GREY); d.alignment = Alignment("left","top", wrap_text=True)
        ws.merge_cells(f"C{rr}:H{rr}")
        ws.row_dimensions[rr].height = 20; ws.row_dimensions[rr+1].height = 30
        rr += 3

    def band(row, text):
        ws.merge_cells(f"A{row}:H{row}")
        c = ws[f"A{row}"]; c.value = "  " + text; c.font = font(10, True, INK); c.fill = fill(MIST)
        c.alignment = Alignment("left","center"); ws.row_dimensions[row].height = 18
    def line(row, text, bold=False):
        ws.merge_cells(f"B{row}:H{row}")
        c = ws[f"B{row}"]; c.value = text; c.font = font(10, bold, BLACK)
        c.alignment = Alignment("left","center", wrap_text=True)

    band(rr, "WHAT'S INSIDE"); rr += 1
    for nm, desc in [("📋 Executive Summary","the one-screen answer — margin, verdict, scenarios"),
                     ("💰 Business Case","break-even, ROI, payback, inventory, cannibalization, waterfall"),
                     ("🎛️  Assumptions","the only place you type — price, mix, targets"),
                     ("📊 Dashboard","P&L with by-channel / by-SKU tables and charts"),
                     ("🧪 Scenarios","trial any CP × MSRP; sensitivity heatmaps"),
                     ("⚖️  Blended","volume-weighted P&L across all channels"),
                     ("📑 TTL + 3 flavour tabs","the detailed per-SKU P&Ls"),
                     ("🏭 Cost Build-up","bottom-up cost from the Bevmax quote")]:
        line(rr, f"{nm}  —  {desc}"); rr += 1
    rr += 1
    band(rr, "THE NUMBERS BEHIND IT"); rr += 1
    for txt in ["1 case = 24 cans = 6 × 4-pack (355 ml).  Production run 40,500 cans (13,500 / flavour).",
                "Bevmax landed cost $1.21 / can = $29.10 / case.  Fully-loaded COGS $32.12 / case.",
                "Run = 1,687.5 cases vs sell-in plan 1,500 → ~188 cases surplus inventory.",
                "Source: Bevmax quote OG2026-05-21 v15 · FOB Airdrie AB, excl. GST · elliott.zhong@organika.com"]:
        line(rr, txt); rr += 1
    ws.freeze_panes = "A5"


def build_exec(ws):
    """Executive summary: headline P&L, margin-gate verdict, blended per-case, scenario snapshot."""
    AMBER = PatternFill("solid", fgColor="FFEB9C"); GREEN = PatternFill("solid", fgColor="C6EFCE")
    RED = PatternFill("solid", fgColor="FFC7CE")
    ws.column_dimensions["A"].width=26
    for col in "BCDEFG": ws.column_dimensions[col].width=14
    def tr(cell): return f"='{TTL_SHEET}'!{cell}"
    def band(rng,text):
        first=rng.split(":")[0]; ws.merge_cells(rng)
        c=ws[first]; c.value=text; c.font=font(11,True,INK); c.fill=fill(BANDBG); c.alignment=Alignment("left","center")
    def kpi(r,lc,lt,vc,ref,fmt,bandf=False):
        a=ws[f"{lc}{r}"]; a.value=lt; a.font=font(10,False,BLACK); a.alignment=Alignment("left","center")
        v=ws[f"{vc}{r}"]; v.value=ref; v.font=font(10, bandf, BLACK); v.number_format=fmt; v.alignment=Alignment("right","center")
        if bandf:
            ws[f"{lc}{r}"].fill=fill(LTBLU); ws[f"{vc}{r}"].fill=fill(LTBLU)

    ws.merge_cells("A1:G1")
    t=ws["A1"]; t.value="ORGANIKA RTD — EXECUTIVE SUMMARY"
    t.font=font(13,True,WHITE); t.fill=fill(NAVY); t.alignment=Alignment("left","center"); ws.row_dimensions[1].height=27.75
    ws.merge_cells("A2:G2")
    s=ws["A2"]; s.value="CDN $ | Aug 1 – Nov 1, 2026 | 3 SKUs × 11 channels | blended = volume-weighted"
    s.font=font(9,False,GREY); s.alignment=Alignment("left","center")

    band("A4:G4","RECOMMENDATION")
    ws.merge_cells("A5:G5")
    verdict = ("=IF('{t}'!M17>=Assumptions!C10,\"✅ PASS — blended GP% \"&TEXT('{t}'!M17,\"0.0%\")&"
               "\" meets the \"&TEXT(Assumptions!C10,\"0%\")&\" target\","
               "IF('{t}'!M17>=Assumptions!C9,\"🟡 OK — GP% \"&TEXT('{t}'!M17,\"0.0%\")&"
               "\" clears the \"&TEXT(Assumptions!C9,\"0%\")&\" minimum (under target)\","
               "\"❌ BELOW MIN — GP% \"&TEXT('{t}'!M17,\"0.0%\")&\" under \"&TEXT(Assumptions!C9,\"0%\")))").format(t=TTL_SHEET)
    v=ws["A5"]; v.value=verdict; v.font=font(12,True,INK); v.alignment=Alignment("left","center"); ws.row_dimensions[5].height=30
    ws.conditional_formatting.add("A5", FormulaRule(formula=['ISNUMBER(SEARCH("PASS",A5))'], fill=GREEN))
    ws.conditional_formatting.add("A5", FormulaRule(formula=['ISNUMBER(SEARCH("OK",A5))'], fill=AMBER))
    ws.conditional_formatting.add("A5", FormulaRule(formula=['ISNUMBER(SEARCH("BELOW",A5))'], fill=RED))

    band("A7:B7","PERIOD TOTALS"); band("D7:E7","BLENDED — PER CASE")
    left=[("Physical cases","M9","int"),("Gross revenue","M10","cur0"),("Net sales","M12","cur0"),
          ("Cost of sales","M15","cur0"),("Gross profit","M16","cur0"),("Blended GP %","M17","pct"),
          ("A&P","M22","cur0"),("CAAP","M24","cur0")]
    right=[("Net sales / case","M28","cur2"),("COGS / case","M31","cur2"),("Gross profit / case","M32","cur2"),
           ("Blended GP %","M33","pct"),("A&P / case","M34","cur2"),("CAAP / case","M35","cur2")]
    for i,(lt,c,fmt) in enumerate(left):
        kpi(8+i,"A",lt,"B",tr(c),FMT[fmt],bandf=(c=="M17"))
    for i,(lt,c,fmt) in enumerate(right):
        kpi(8+i,"D",lt,"E",tr(c),FMT[fmt],bandf=(c=="M33"))

    band("A17:G17","MARGIN GATE  &  PRICE-TO-HIT-MARGIN")
    kpi(18,"A","Minimum GM%","B","=Assumptions!C9",PCT)
    kpi(19,"A","Target GM%","B","=Assumptions!C10",PCT)
    kpi(20,"A","Blended GP% (live)","B",tr("M17"),PCT,bandf=True)
    # solve the CP (wholesale $/case) needed to hit each margin, at current COGS
    kpi(18,"D","CP /case to hit min GM%","E",f"={COST_TOTAL_REF}/(1-Assumptions!C9)",CUR2)
    kpi(19,"D","CP /case to hit target GM%","E",f"={COST_TOTAL_REF}/(1-Assumptions!C10)",CUR2)
    kpi(20,"D","COGS /case (fully loaded)","E",f"={COST_TOTAL_REF}",CUR2,bandf=True)

    band("A22:G22","PRICE SCENARIOS  (from Scenarios tab)")
    for col,nm in zip(["B","C","D"],["Conservative","Base","Stretch"]):
        c=ws[f"{col}23"]; c.value=nm; c.font=font(10,True,WHITE); c.fill=fill(MIDBLU); c.alignment=Alignment("center","center")
    ws["A23"].value="Scenario"; ws["A23"].font=font(10,True,WHITE); ws["A23"].fill=fill(MIDBLU)
    for r,(lt,srow,fmt) in [(24,("Organika GP %",14,PCT)),(25,("GP $/case",13,CUR2)),(26,("Retailer margin %",15,PCT))]:
        ws[f"A{r}"].value=lt; ws[f"A{r}"].font=font(10,False,BLACK); ws[f"A{r}"].alignment=Alignment("left","center")
        for col,scol in zip(["B","C","D"],["B","C","D"]):
            c=ws[f"{col}{r}"]; c.value=f"=Scenarios!{scol}{srow}"; c.font=font(10,False,BLACK)
            c.number_format=fmt; c.alignment=Alignment("right","center")

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


def build_business_case(ws):
    """Investment-decision layer: break-even, ROI/payback, inventory, category, cannibalization, scenarios + waterfall."""
    CASES1 = '#,##0.0'
    GREEN = PatternFill("solid", fgColor="C6EFCE"); RED = PatternFill("solid", fgColor="FFC7CE")
    for col, w in {"A":36,"B":14,"C":14,"D":14,"E":19,"F":10,"G":10,"H":13,"I":9,"J":9,"K":9}.items():
        ws.column_dimensions[col].width = w
    t = lambda c: f"='{TTL_SHEET}'!{c}"
    cost = lambda c: f"='{COST_SHEET}'!{c}"
    apb = f"'{TTL_SHEET}'!M22/'{TTL_SHEET}'!M12"   # blended A&P % of net

    def band(rng, text):
        first = rng.split(":")[0]; ws.merge_cells(rng)
        c = ws[first]; c.value=text; c.font=font(11,True,INK); c.fill=fill(BANDBG); c.alignment=Alignment("left","center")
    def lab(r, text, col="A", bold=False, sz=10):
        c=ws[f"{col}{r}"]; c.value=text; c.font=font(sz,bold,BLACK); c.alignment=Alignment("left","center")
    def inp(cell, val, fmt):
        c=ws[cell]; c.value=val; c.font=font(10,False,BLUE); c.fill=fill(INPUT_FILL)
        c.number_format=fmt; c.alignment=Alignment("right","center")
    def calc(cell, val, fmt, bold=False, band=False):
        c=ws[cell]; c.value=val; c.font=font(10,bold,BLACK); c.number_format=fmt; c.alignment=Alignment("right","center")
        if band: c.fill=fill(LTBLU)
    def note(cell, text):
        c=ws[cell]; c.value=text; c.font=font(8,False,GREY); c.alignment=Alignment("left","center")

    ws.merge_cells("A1:G1")
    h=ws["A1"]; h.value="ORGANIKA RTD — BUSINESS CASE (break-even, ROI, risk)"
    h.font=font(13,True,WHITE); h.fill=fill(NAVY); h.alignment=Alignment("left","center"); ws.row_dimensions[1].height=27.75
    ws.merge_cells("A2:G2")
    s=ws["A2"]; s.value="CDN $ | Aug 1 – Nov 1, 2026 | pulls from TTL + Cost Build-up | 🟡 edit yellow"
    s.font=font(9,False,GREY); s.alignment=Alignment("left","center")

    # A — discretionary launch spend (NOT in COGS, to avoid double-count)
    band("A4:G4","DISCRETIONARY LAUNCH SPEND  (not already in COGS)")
    lab(5,"Slotting / listing fees");                 inp("B5",10000,CUR0)
    lab(6,"Fixed launch marketing (beyond A&P %)");   inp("B6",15000,CUR0)
    lab(7,"Other one-time (not in COGS)");            inp("B7",0,CUR0)
    lab(8,"Total launch spend", bold=True);           calc("B8","=SUM(B5:B7)",CUR0,bold=True)
    note("C9","Bevmax setup & allergen are already inside COGS — not added here.")

    # B — break-even, ROI, payback
    band("A11:G11","BREAK-EVEN · ROI · PAYBACK")
    lab(12,"Blended CAAP / case (contribution)");     calc("B12",t("M35"),CUR2)
    lab(13,"Break-even cases (cover launch spend)");  calc("B13",'=IFERROR(B8/B12,"-")',INT,bold=True)
    lab(14,"Plan cases (period)");                    calc("B14",t("M9"),INT)
    lab(15,"Margin of safety (cases)");               calc("B15","=B14-B13",INT)
    lab(16,"Margin of safety (%)");                   calc("B16",'=IFERROR(B15/B14,"-")',PCT,band=True)
    lab(17,"Total CAAP (period)");                    calc("B17",t("M24"),CUR0)
    lab(18,"Launch profit (CAAP − launch spend)", bold=True); calc("B18","=B17-B8",CUR0,bold=True)
    lab(19,"ROI on launch spend");                    calc("B19",'=IFERROR(B18/B8,"-")',PCT,band=True)

    # C — production run & inventory
    band("A21:G21","PRODUCTION RUN & INVENTORY (sell-through)")
    lab(22,"Run cost — cash to produce");             calc("B22",cost("D26"),CUR0); note("C22","Bevmax run subtotal")
    lab(23,"Deposit required (50%, 4 wks pre-prod)"); inp("B23",24552.87,CUR0)
    lab(24,"Cans produced");                          calc("B24",cost("D6"),INT)
    lab(25,"Cases produced");                         calc("B25",cost("D11"),CASES1)
    lab(26,"Cases sold (plan)");                      calc("B26",t("M9"),INT)
    lab(27,"Sell-through %");                         calc("B27",'=IFERROR(B26/B25,"-")',PCT,band=True)
    lab(28,"Ending inventory (cases)");               calc("B28","=B25-B26",CASES1)
    lab(29,"Ending inventory value @ COGS");          calc("B29",f'=B28*{COST_TOTAL_REF}',CUR0)
    lab(30,"Est. surplus warehousing cost");          calc("B30","=MAX(0,(F28-F29))*ROUNDUP(B28/F30,0)*F31",CUR0)
    lab(28,"Months stored","E"); inp("F28",6,INT)
    lab(29,"Free months","E");   inp("F29",3,INT)
    lab(30,"Cases / pallet","E");inp("F30",60,INT)
    lab(31,"Warehousing $/pallet/mo","E"); inp("F31",100,CUR0)

    # D — category benchmark & cannibalization
    band("A33:G33","CATEGORY BENCHMARK & CANNIBALIZATION")
    lab(34,"Category average GM% (benchmark)");       inp("B34",0.55,PCT)
    lab(35,"Our blended GP%");                        calc("B35",t("M17"),PCT)
    lab(36,"vs category (pts)");                      calc("B36","=B35-B34",PCT,band=True)
    lab(37,"Cannibalization rate (% of our volume)"); inp("B37",0.20,PCT)
    lab(38,"Existing-SKU CAAP / case (lost)");        inp("B38",20,CUR2)
    lab(39,"Cannibalized cases");                     calc("B39","=B14*B37",INT)
    lab(40,"Lost CAAP (cannibalized)");               calc("B40","=B39*B38",CUR0)
    lab(41,"Net incremental CAAP (after cannib.)", bold=True); calc("B41","=B18-B40",CUR0,bold=True)

    # E — best / base / worst
    band("A43:G43","BEST / BASE / WORST CASE")
    for col,nm in zip(["B","C","D"],["Worst","Base","Best"]):
        c=ws[f"{col}44"]; c.value=nm; c.font=font(10,True,WHITE); c.fill=fill(MIDBLU); c.alignment=Alignment("center","center")
    ws["A44"].value="Driver"; ws["A44"].font=font(10,True,WHITE); ws["A44"].fill=fill(MIDBLU)
    lab(45,"CP / case (wholesale)")
    inp("B45",60,CUR2); inp("C45",72,CUR2); inp("D45",84,CUR2)
    lab(46,"COGS / case");      calc("B46",f"={COST_TOTAL_REF}*(1+G45)",CUR2); calc("C46",f"={COST_TOTAL_REF}",CUR2); calc("D46",f"={COST_TOTAL_REF}*(1-G46)",CUR2)
    lab(47,"A&P % of net");     calc("B47",f"={apb}",PCT); calc("C47",f"={apb}",PCT); calc("D47",f"={apb}",PCT)
    lab(48,"Volume (cases)");   calc("B48","=B14*(1-G47)",INT); calc("C48","=B14",INT); calc("D48","=B14*(1+G48)",INT)
    lab(49,"CAAP / case");      calc("B49","=B45-B46-B45*B47",CUR2); calc("C49","=C45-C46-C45*C47",CUR2); calc("D49","=D45-D46-D45*D47",CUR2)
    lab(50,"Total CAAP",bold=True); calc("B50","=B49*B48",CUR0,bold=True); calc("C50","=C49*C48",CUR0,bold=True); calc("D50","=D49*D48",CUR0,bold=True)
    lab(51,"GP %");             calc("B51",'=IFERROR((B45-B46)/B45,"-")',PCT,band=True); calc("C51",'=IFERROR((C45-C46)/C45,"-")',PCT,band=True); calc("D51",'=IFERROR((D45-D46)/D45,"-")',PCT,band=True)
    # confidence factors (editable)
    lab(45,"Worst COGS +%","E"); inp("G45",0.15,PCT)
    lab(46,"Best COGS −%","E");  inp("G46",0.10,PCT)
    lab(47,"Worst vol −%","E");  inp("G47",0.25,PCT)
    lab(48,"Best vol +%","E");   inp("G48",0.20,PCT)

    # checks
    ws.conditional_formatting.add("B36", FormulaRule(formula=["B36>=0"], fill=GREEN))
    ws.conditional_formatting.add("B36", FormulaRule(formula=["B36<0"], fill=RED))

    # ---- P&L waterfall (per case) ----
    hv = {"H4":"P&L bridge / case","H5":"Gross Rev","H6":"Discounts","H7":"COGS","H8":"A&P","H9":"CAAP"}
    for k,v in hv.items():
        ws[k]=v; ws[k].font=font(8,False,GREY)
    # base (invisible) / decrease / increase  (cols I, J, K)
    ws["I5"]=0;            ws["K5"]=t("M26"); ws["J5"]=0
    ws["I6"]=t("M28");     ws["J6"]=t("M27"); ws["K6"]=0
    ws["I7"]=t("M32");     ws["J7"]=t("M31"); ws["K7"]=0
    ws["I8"]=t("M35");     ws["J8"]=t("M34"); ws["K8"]=0
    ws["I9"]=0;            ws["K9"]=t("M35"); ws["J9"]=0
    for r in range(5,10):
        for c in "IJK": ws[f"{c}{r}"].number_format=CUR2; ws[f"{c}{r}"].font=font(8,False,GREY)
    wf = BarChart(); wf.type="col"; wf.grouping="stacked"; wf.overlap=100
    wf.title="P&L bridge — per case"; wf.legend=None; wf.height=8; wf.width=15
    wf.add_data(Reference(ws,min_col=9,min_row=5,max_row=9), titles_from_data=False)   # base
    wf.add_data(Reference(ws,min_col=10,min_row=5,max_row=9), titles_from_data=False)  # decrease
    wf.add_data(Reference(ws,min_col=11,min_row=5,max_row=9), titles_from_data=False)  # increase
    wf.set_categories(Reference(ws,min_col=8,min_row=5,max_row=9))
    wf.series[0].graphicalProperties = GraphicalProperties(noFill=True)
    wf.series[1].graphicalProperties = GraphicalProperties(solidFill="F4B183")  # decreases (orange)
    wf.series[2].graphicalProperties = GraphicalProperties(solidFill="4472C4")  # totals (blue)
    wf.y_axis.numFmt = '$#,##0'
    ws.add_chart(wf, "A53")

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


# ---- build workbook ----
SKUS = ["Lime Lemon", "Passion Fruit Pineapple", "Raspberry"]
TAB_COLORS = {"Lime Lemon": "FF92D050", "Passion Fruit Pineapple": "FFFFC000", "Raspberry": "FFC00000"}

wb = openpyxl.Workbook()
ttl = wb.active
ttl.title = TTL_SHEET
ttl.sheet_properties.tabColor = NAVY
build_sheet(ttl, "ORGANIKA RTD — TOTAL (3 SKUs)", "Organika RTD — All SKUs (Lime Lemon + Passion Fruit Pineapple + Raspberry)",
            is_total=True, sku_sheets=SKUS)

for sku in SKUS:
    ws = wb.create_sheet(sku)
    ws.sheet_properties.tabColor = TAB_COLORS[sku]
    build_sheet(ws, f"ORGANIKA RTD — {sku.upper()}", f"Organika RTD — {sku} (6×4×355 ml)",
                is_total=False, sku_row=A_TOTROW[sku])

cost = wb.create_sheet(COST_SHEET); cost.sheet_properties.tabColor = "FF548235"
build_cost_buildup(cost)
blend = wb.create_sheet("Blended"); blend.sheet_properties.tabColor = "FF7030A0"
build_blended(blend, SKUS)
scen = wb.create_sheet("Scenarios"); scen.sheet_properties.tabColor = "FFED7D31"
build_scenarios(scen)
dash = wb.create_sheet("Dashboard"); dash.sheet_properties.tabColor = "FF4472C4"
build_dashboard(dash, SKUS)
assum = wb.create_sheet(ASSUMP); assum.sheet_properties.tabColor = "FFFFC000"
build_assumptions(assum, SKUS)
cover = wb.create_sheet("Start Here"); cover.sheet_properties.tabColor = "FF0A2540"
build_cover(cover)
exec_ = wb.create_sheet("Executive Summary"); exec_.sheet_properties.tabColor = "FFC00000"
build_exec(exec_)
bcase = wb.create_sheet("Business Case"); bcase.sheet_properties.tabColor = "FFED7D31"
build_business_case(bcase)

# charts: GP% and volume by channel (Dashboard), CAAP by SKU (Exec)
def bar(anchor_ws, title, data_col, cat_ws, r0, r1, numfmt=None):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.legend = None
    ch.height = 7.5; ch.width = 15
    ch.add_data(Reference(cat_ws, min_col=data_col, min_row=r0, max_row=r1), titles_from_data=False)
    ch.set_categories(Reference(cat_ws, min_col=1, min_row=r0, max_row=r1))
    if numfmt: ch.y_axis.numFmt = numfmt
    return ch
dash.add_chart(bar(dash, "GP % by channel", 5, dash, 24, 34, "0%"), "A37")
dash.add_chart(bar(dash, "Volume (cases) by channel", 2, dash, 24, 34), "A54")
exec_.add_chart(bar(exec_, "CAAP by SKU", 7, dash, 17, 19), "A28")

# negative values flash red (losses) on the summary views
REDFILL = PatternFill("solid", fgColor="FFC7CE")
neg = CellIsRule(operator="lessThan", formula=["0"], fill=REDFILL)
blend.conditional_formatting.add("B7:E27", neg)
dash.conditional_formatting.add("B8:G35", neg)
exec_.conditional_formatting.add("B8:E26", neg)

# print / PDF setup: landscape, fit to one page wide, footer with file + tab + page
for ws in wb.worksheets:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.oddFooter.center.text = "&F  |  &A  |  &P of &N"
    ws.oddFooter.center.size = 8

# final tab order
order = ["Start Here", "Executive Summary", "Business Case", ASSUMP, "Dashboard", "Scenarios",
         "Blended", TTL_SHEET] + SKUS + [COST_SHEET]
wb._sheets = [wb[name] for name in order]

# force recalculation when opened in Excel / Google Sheets / LibreOffice
wb.calculation = CalcProperties(fullCalcOnLoad=True)

out = "/home/user/my-first-project/Organika_RTD_PL_Aug-Nov2026.xlsx"
wb.save(out)
print("Saved:", out)
print("Sheets:", wb.sheetnames)
